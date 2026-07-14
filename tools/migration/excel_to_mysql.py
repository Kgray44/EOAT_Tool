from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

MISSING_TOKENS = {"", "n/a", "na", "none", "unknown", "unknown / not checked", "not checked"}
MACHINE_PATTERN = re.compile(r"^\d+$")


@dataclass(frozen=True)
class MigrationIssue:
    severity: str
    code: str
    sheet: str
    row: int | None
    field: str | None
    source_value: str | None
    description: str
    suggested_resolution: str


@dataclass
class MigrationReport:
    batch_uuid: str
    batch_name: str
    source_workbook: str
    source_checksum_sha256: str
    started_at: str
    completed_at: str = ""
    dry_run: bool = True
    schema_revision: str = "20260713_0001"
    workbook_schema_version: str = ""
    source_rows: dict[str, int] = field(default_factory=dict)
    staged_counts: dict[str, int] = field(default_factory=dict)
    rejected_rows: int = 0
    warnings: int = 0
    errors: int = 0
    duplicate_counts: dict[str, int] = field(default_factory=dict)
    unresolved_relationships: int = 0
    document_paths_checked: int = 0
    document_paths_missing: int = 0
    issues: list[MigrationIssue] = field(default_factory=list)
    source_unchanged: bool = False

    def add_issue(self, issue: MigrationIssue) -> None:
        self.issues.append(issue)
        if issue.severity == "ERROR":
            self.errors += 1
        else:
            self.warnings += 1

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["issues"] = [asdict(issue) for issue in self.issues]
        return payload


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _missing(value: Any) -> bool:
    return _text(value).casefold() in MISSING_TOKENS


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime | date):
        return value.isoformat()
    return value


def _rows(sheet) -> list[tuple[int, dict[str, Any]]]:
    iterator = sheet.iter_rows(values_only=True)
    headers = [_text(value) for value in next(iterator)]
    result: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(iterator, start=2):
        if not any(value not in (None, "") for value in values):
            continue
        result.append(
            (row_number, {header: _json_value(value) for header, value in zip(headers, values, strict=False)})
        )
    return result


def _checksum(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _canonical_area(value: Any) -> str:
    text = _text(value)
    return "CL" if text.casefold() in {"cleanroom", "cl"} else "P4"


def _issue(
    severity: str,
    code: str,
    sheet: str,
    row: int | None,
    field: str | None,
    value: Any,
    description: str,
    resolution: str,
) -> MigrationIssue:
    return MigrationIssue(severity, code, sheet, row, field, _text(value) or None, description, resolution)


def analyze_workbook(source_workbook: str | Path, *, batch_name: str | None = None) -> MigrationReport:
    source = Path(source_workbook).resolve()
    before = source.stat()
    before_checksum = _checksum(source)
    started_at = datetime.now(timezone.utc)
    report = MigrationReport(
        batch_uuid=str(uuid4()),
        batch_name=batch_name or f"legacy-excel-dry-run-{started_at:%Y%m%d-%H%M%S}",
        source_workbook=str(source),
        source_checksum_sha256=before_checksum,
        started_at=started_at.isoformat(),
    )

    workbook = load_workbook(source, read_only=True, data_only=True)
    required_sheets = {"EOAT Inventory", "Photo Index", "_EOAT_App_Metadata"}
    missing_sheets = required_sheets.difference(workbook.sheetnames)
    for name in sorted(missing_sheets):
        report.add_issue(
            _issue(
                "ERROR",
                "MISSING_SHEET",
                name,
                None,
                None,
                None,
                "Required source sheet is absent.",
                "Restore the required sheet before import.",
            )
        )
    if missing_sheets:
        report.rejected_rows += 1
        report.completed_at = datetime.now(timezone.utc).isoformat()
        return report

    inventory_rows = _rows(workbook["EOAT Inventory"])
    photo_rows = _rows(workbook["Photo Index"])
    report.source_rows = {"EOAT Inventory": len(inventory_rows), "Photo Index": len(photo_rows)}
    metadata = {
        str(key): _text(value)
        for key, value in workbook["_EOAT_App_Metadata"].iter_rows(min_row=2, values_only=True)
        if key
    }
    report.workbook_schema_version = metadata.get("schema_version", "")

    plants: set[str] = set()
    areas: set[tuple[str, str]] = set()
    eoats: dict[str, dict[str, Any]] = {}
    machines: set[tuple[str, str]] = set()
    tools: dict[str, dict[str, Any]] = {}
    eoat_machine: set[tuple[str, str, str]] = set()
    eoat_tool: set[tuple[str, str]] = set()
    tool_machine: set[tuple[str, str, str]] = set()
    audit_ids: list[str] = []
    audited_locations: dict[str, set[str]] = defaultdict(set)
    rejected_source_rows: set[int] = set()
    part_candidates: set[tuple[str, str]] = set()

    for row_number, row in inventory_rows:
        plant_code = _canonical_area(row.get("Plant/Area"))
        area_name = _text(row.get("Plant/Area")) or "Unknown"
        plants.add("P4")
        areas.add(("P4", area_name))
        eoat_id = _text(row.get("EOAT Assembly ID"))
        audit_id = _text(row.get("Audit ID"))
        machine = _text(row.get("Press/Machine #"))
        tool_number = _text(row.get("Tool #"))
        entry_type = _text(row.get("Entry Type"))

        if audit_id:
            audit_ids.append(audit_id)
        else:
            report.add_issue(
                _issue(
                    "ERROR",
                    "MISSING_AUDIT_ID",
                    "EOAT Inventory",
                    row_number,
                    "Audit ID",
                    None,
                    "The row has no traceable audit identifier.",
                    "Assign a unique Audit ID.",
                )
            )
            rejected_source_rows.add(row_number)
        if not eoat_id:
            report.add_issue(
                _issue(
                    "ERROR",
                    "MISSING_EOAT_IDENTIFIER",
                    "EOAT Inventory",
                    row_number,
                    "EOAT Assembly ID",
                    None,
                    "The row cannot be associated with an EOAT.",
                    "Assign or recover the EOAT business identifier.",
                )
            )
            rejected_source_rows.add(row_number)
            continue

        eoat_candidate = {
            "business_identifier": eoat_id,
            "eoat_type": _text(row.get("EOAT Type")),
            "connection_type": _text(row.get("Connection Type")),
            "cleanroom_classification": _text(row.get("Cleanroom/Non-Cleanroom")),
            "number_of_parts_picked": _text(row.get("Number of Parts Picked")),
            "number_of_vacuum_cups": _text(row.get("# of Cups")),
            "number_of_grippers": _text(row.get("# of Grippers")),
            "notes": _text(row.get("Notes")),
            "source_row": row_number,
        }
        prior = eoats.get(eoat_id)
        if prior:
            conflicting = [
                key
                for key in ("eoat_type", "connection_type", "cleanroom_classification")
                if prior.get(key) and eoat_candidate.get(key) and prior[key] != eoat_candidate[key]
            ]
            if conflicting:
                report.add_issue(
                    _issue(
                        "WARNING",
                        "CONFLICTING_EOAT_ATTRIBUTES",
                        "EOAT Inventory",
                        row_number,
                        ", ".join(conflicting),
                        eoat_id,
                        "Repeated EOAT rows disagree on stable profile attributes.",
                        "Review the named fields and choose authoritative values before import.",
                    )
                )
        else:
            eoats[eoat_id] = eoat_candidate

        if machine and not _missing(machine):
            if MACHINE_PATTERN.fullmatch(machine):
                machines.add((plant_code, machine))
                eoat_machine.add((eoat_id, plant_code, machine))
                if tool_number and not _missing(tool_number):
                    tool_machine.add((tool_number, plant_code, machine))
                if entry_type.casefold() == "audited":
                    audited_locations[eoat_id].add(f"{plant_code}:{machine}")
            else:
                report.add_issue(
                    _issue(
                        "WARNING",
                        "AMBIGUOUS_MACHINE_NUMBER",
                        "EOAT Inventory",
                        row_number,
                        "Press/Machine #",
                        machine,
                        "Machine value contains qualifiers or multiple meanings and cannot be normalized safely.",
                        "Replace with one canonical machine number and retain the qualifier in notes.",
                    )
                )
        else:
            report.add_issue(
                _issue(
                    "WARNING",
                    "MISSING_MACHINE",
                    "EOAT Inventory",
                    row_number,
                    "Press/Machine #",
                    machine,
                    "No machine relationship can be created for this row.",
                    "Confirm whether the EOAT is in storage or the machine value is missing.",
                )
            )

        if tool_number and not _missing(tool_number):
            tools.setdefault(
                tool_number, {"business_identifier": tool_number, "tool_number": tool_number, "source_row": row_number}
            )
            eoat_tool.add((eoat_id, tool_number))
        else:
            report.add_issue(
                _issue(
                    "WARNING",
                    "MISSING_TOOL",
                    "EOAT Inventory",
                    row_number,
                    "Tool #",
                    tool_number,
                    "No tool relationship can be created for this row.",
                    "Confirm the tool number or explicitly mark the row as not tool-specific.",
                )
            )

        part_name = _text(row.get("Part Name/Description"))
        if part_name:
            part_candidates.add((tool_number, part_name))

        for field_name in (
            "Sensors Present?",
            "Vacuum Confirmation Present?",
            "Part-Present Detection Present?",
            "Quick Disconnects Present?",
        ):
            value = _text(row.get(field_name))
            if value and value.casefold() not in {
                "yes",
                "no",
                "y",
                "n",
                "true",
                "false",
                "present",
                "not present",
                *MISSING_TOKENS,
            }:
                report.add_issue(
                    _issue(
                        "WARNING",
                        "INVALID_BOOLEAN",
                        "EOAT Inventory",
                        row_number,
                        field_name,
                        value,
                        "Boolean-like value is outside the accepted mapping vocabulary.",
                        "Map it explicitly to true, false, or unknown.",
                    )
                )

    duplicate_audits = sum(count - 1 for count in Counter(audit_ids).values() if count > 1)
    report.duplicate_counts["audit_identifiers"] = duplicate_audits
    if duplicate_audits:
        report.add_issue(
            _issue(
                "ERROR",
                "DUPLICATE_AUDIT_IDENTIFIER",
                "EOAT Inventory",
                None,
                "Audit ID",
                duplicate_audits,
                "Audit IDs are not unique.",
                "Assign unique audit identifiers while preserving the original value in import provenance.",
            )
        )

    for eoat_id, locations in sorted(audited_locations.items()):
        if len(locations) > 1:
            report.add_issue(
                _issue(
                    "WARNING",
                    "CONFLICTING_CURRENT_ASSIGNMENT",
                    "EOAT Inventory",
                    None,
                    "Press/Machine #",
                    ", ".join(sorted(locations)),
                    f"{eoat_id} has multiple audited machine assignments and no explicit removal timeline.",
                    "Review audit dates and create approved installation-history intervals; do not infer a current location automatically.",
                )
            )

    if part_candidates:
        report.add_issue(
            _issue(
                "WARNING",
                "PART_IDENTIFIER_AMBIGUITY",
                "EOAT Inventory",
                None,
                "Part Name/Description",
                len(part_candidates),
                "The workbook contains part names but no independent Part Number field; Tool # cannot safely be assumed to be the part number.",
                "Provide a part-number crosswalk before importing parts and tool_parts.",
            )
        )

    photo_ids: list[str] = []
    project_root = source.parents[2]
    documents: set[str] = set()
    photos: set[str] = set()
    for row_number, row in photo_rows:
        photo_id = _text(row.get("Photo ID"))
        if photo_id:
            photo_ids.append(photo_id)
        relative = _text(row.get("Stored Relative Path"))
        fallback_folder = _text(row.get("Folder Path"))
        filename = _text(row.get("Stored Filename")) or _text(row.get("Photo Filename"))
        candidate = project_root / relative if relative else project_root / fallback_folder / filename
        if relative or (fallback_folder and filename):
            report.document_paths_checked += 1
            documents.add(str(candidate))
            photos.add(photo_id or str(candidate))
            if not candidate.exists():
                report.document_paths_missing += 1
                report.add_issue(
                    _issue(
                        "WARNING",
                        "BROKEN_PHOTO_PATH",
                        "Photo Index",
                        row_number,
                        "Stored Relative Path",
                        relative or str(candidate),
                        "Referenced photo file is unavailable at dry-run time.",
                        "Repair the stored relative path or restore the file; retain the metadata record.",
                    )
                )
        else:
            report.add_issue(
                _issue(
                    "WARNING",
                    "MISSING_PHOTO_PATH",
                    "Photo Index",
                    row_number,
                    "Stored Relative Path",
                    None,
                    "Photo metadata has no resolvable file path.",
                    "Supply a controlled relative path; do not discard the metadata row.",
                )
            )

    duplicate_photos = sum(count - 1 for count in Counter(photo_ids).values() if count > 1)
    report.duplicate_counts["photo_identifiers"] = duplicate_photos
    if duplicate_photos:
        report.add_issue(
            _issue(
                "ERROR",
                "DUPLICATE_PHOTO_IDENTIFIER",
                "Photo Index",
                None,
                "Photo ID",
                duplicate_photos,
                "Photo IDs are not unique.",
                "Assign unique photo IDs while preserving original IDs in import provenance.",
            )
        )

    report.staged_counts = {
        "plants": len(plants),
        "areas": len(areas),
        "eoats": len(eoats),
        "machines": len(machines),
        "tools": len(tools),
        "parts": 0,
        "part_candidates_requiring_crosswalk": len(part_candidates),
        "eoat_machine_compatibility": len(eoat_machine),
        "eoat_tool_compatibility": len(eoat_tool),
        "tool_machine_compatibility": len(tool_machine),
        "audit_records": len(inventory_rows) - len(rejected_source_rows),
        "installation_records": 0,
        "documents": len(documents),
        "photos": len(photos),
        "import_rows": len(inventory_rows) + len(photo_rows),
    }
    report.rejected_rows = len(rejected_source_rows)
    report.unresolved_relationships = sum(
        1
        for issue in report.issues
        if issue.code
        in {
            "AMBIGUOUS_MACHINE_NUMBER",
            "MISSING_MACHINE",
            "MISSING_TOOL",
            "CONFLICTING_CURRENT_ASSIGNMENT",
            "PART_IDENTIFIER_AMBIGUITY",
        }
    )
    workbook.close()
    report.completed_at = datetime.now(timezone.utc).isoformat()
    after = source.stat()
    report.source_unchanged = (
        before.st_size == after.st_size
        and before.st_mtime_ns == after.st_mtime_ns
        and before_checksum == _checksum(source)
    )
    return report


def write_report(report: MigrationReport, output: str | Path) -> tuple[Path, Path]:
    target = Path(output)
    if target.suffix.lower() in {".json", ".md"}:
        target = target.with_suffix("")
    target.parent.mkdir(parents=True, exist_ok=True)
    json_path = target.with_suffix(".json")
    md_path = target.with_suffix(".md")
    json_path.write_text(json.dumps(report.to_dict(), indent=2, ensure_ascii=False), encoding="utf-8")
    counts = "\n".join(f"- `{name}`: {count}" for name, count in report.staged_counts.items())
    issues = (
        "\n".join(
            f"| {item.severity} | {item.code} | {item.sheet} | {item.row or ''} | {item.field or ''} | {item.description} |"
            for item in report.issues
        )
        or "| - | - | - | - | - | No issues detected |"
    )
    md_path.write_text(
        f"""# EOAT Atlas Excel-to-MySQL Dry-Run Report

- Batch: `{report.batch_uuid}` ({report.batch_name})
- Source: `{report.source_workbook}`
- SHA-256: `{report.source_checksum_sha256}`
- Workbook schema: `{report.workbook_schema_version or 'unknown'}`
- Target Alembic revision: `{report.schema_revision}`
- Source unchanged: `{report.source_unchanged}`
- Rejected source rows: `{report.rejected_rows}`
- Warnings: `{report.warnings}`
- Errors: `{report.errors}`
- Unresolved relationships: `{report.unresolved_relationships}`
- Photo paths checked/missing: `{report.document_paths_checked}` / `{report.document_paths_missing}`

## Staged counts

{counts}

`installation_records` remains zero by design: the workbook has repeated audited/current-looking rows but no reliable removal timeline, so the dry run refuses to invent installation history.

## Validation findings

| Severity | Code | Sheet | Row | Field | Description |
|---|---|---|---:|---|---|
{issues}
""",
        encoding="utf-8",
    )
    return json_path, md_path


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Inspect and stage the EOAT Atlas legacy workbook for MySQL migration."
    )
    parser.add_argument("--source-workbook", required=True)
    parser.add_argument("--report-output", required=True)
    parser.add_argument("--import-batch-name")
    parser.add_argument("--database-profile", default="development")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--validate-only", action="store_true")
    parser.add_argument("--execute", action="store_true")
    parser.add_argument("--reset-imported-data", action="store_true")
    args = parser.parse_args()
    selected = sum(bool(value) for value in (args.dry_run, args.validate_only, args.execute))
    if selected != 1:
        parser.error("Choose exactly one of --dry-run, --validate-only, or --execute.")
    if args.execute:
        if args.database_profile != "development":
            parser.error("Executable import is restricted to --database-profile development in this phase.")
        from tools.migration.import_pipeline import execute_import

        report = execute_import(
            args.source_workbook,
            batch_name=args.import_batch_name or "controlled-development-import",
            reset_imported_data=args.reset_imported_data,
            report_output=args.report_output,
        )
        print(json.dumps(report.to_dict(), ensure_ascii=False))
        return 0 if report.transaction_result in {"COMMITTED", "SAFE_STOP_ALREADY_IMPORTED"} else 1
    report = analyze_workbook(args.source_workbook, batch_name=args.import_batch_name)
    json_path, md_path = write_report(report, args.report_output)
    print(
        json.dumps(
            {"json_report": str(json_path), "markdown_report": str(md_path), "summary": report.to_dict()},
            ensure_ascii=False,
        )
    )
    return 1 if report.errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
