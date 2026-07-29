"""Controlled, fail-closed import of press capacity into existing machines.

The Press Capacity workbook is not authority for a tool relationship: its NGW
part number is deliberately retained as provenance until an approved crosswalk
exists.  This module therefore updates only ``machines.press_capacity_tons``.
It never creates, deletes, or changes assignment or compatibility rows.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections.abc import Iterable, Mapping
from dataclasses import asdict, dataclass, field, replace
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import uuid4

if __package__ in {None, ""}:
    # Keep the governed CLI usable as ``python tools/migration/...py`` as
    # well as ``python -m``; production runbooks use an explicit script path.
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from server.eoat_api.data_state import record_import_completion
from server.eoat_api.database import models as db
from server.eoat_api.database.session import create_session_factory
from server.eoat_api.release_provenance import ensure_application_release

_MACHINE_HEADERS = {"machine no.", "machine no", "machine #", "machine number", "press", "press #", "press/machine #"}
_TONNAGE_HEADERS = {"press tonnage", "tonnage", "capacity", "u.s. tons", "us tons"}
_MACHINE_TOKEN = re.compile(r"^(?:machine|press)?\s*#?\s*(\d+)$", re.IGNORECASE)
_PRESS_SECTION = re.compile(
    r"^\s*press\s*#?\s*(\d+)\s*(?:[-–—]\s*(\d+(?:\.\d+)?)\s*(?:t|ton|tons)\b)?",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class CapacitySourceRow:
    sheet: str
    row_number: int
    machine_numbers: tuple[str, ...]
    tonnage: Decimal | None
    raw_values: dict[str, Any]
    issue: str | None = None
    capacity_source: str = "press_capacity_workbook"


@dataclass(frozen=True)
class CapacityUpdate:
    machine_number: str
    source_rows: tuple[int, ...]
    source_tonnage: Decimal
    existing_tonnage: Decimal | None
    action: str


@dataclass
class PressCapacityReport:
    source_file_name: str
    source_sha256: str
    source_rows: int
    source_machine_count: int
    matched_machines: int
    supplementary_sources: dict[str, str] = field(default_factory=dict)
    unmatched_machines: list[str] = field(default_factory=list)
    conflicting_source_values: dict[str, list[str]] = field(default_factory=dict)
    conflicting_existing_values: dict[str, dict[str, str]] = field(default_factory=dict)
    invalid_rows: list[dict[str, Any]] = field(default_factory=list)
    updates: list[CapacityUpdate] = field(default_factory=list)
    status: str = "DRY_RUN_COMPLETE"
    batch_uuid: str | None = None
    data_changed: bool = False

    @property
    def safe_to_execute(self) -> bool:
        return not (
            self.unmatched_machines
            or self.conflicting_source_values
            or self.conflicting_existing_values
            or self.invalid_rows
        )

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        for update in payload["updates"]:
            update["source_tonnage"] = str(update["source_tonnage"])
            update["existing_tonnage"] = str(update["existing_tonnage"]) if update["existing_tonnage"] is not None else None
        return payload


def _digest(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _normalise_header(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").strip().casefold().split())


def _normalise_machine(value: Any) -> str | None:
    match = _MACHINE_TOKEN.fullmatch(str(value or "").strip())
    return str(int(match.group(1))) if match else None


def _machine_numbers(value: Any) -> tuple[str, ...] | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    parts = [part.strip() for part in raw.split(",")]
    if not parts or any(not part for part in parts):
        return None
    result = tuple(_normalise_machine(part) for part in parts)
    if any(number is None for number in result) or len(set(result)) != len(result):
        return None
    return tuple(number for number in result if number is not None)


def _tonnage(value: Any) -> Decimal | None:
    if value is None or not str(value).strip():
        return None
    try:
        parsed = Decimal(str(value).replace(",", "").strip())
    except InvalidOperation:
        return None
    return parsed if parsed > 0 else None


def _header_indexes(headers: Iterable[Any]) -> tuple[int, int] | None:
    normalized = [_normalise_header(header) for header in headers]
    machine = next((index for index, header in enumerate(normalized) if header in _MACHINE_HEADERS), None)
    tonnage = next((index for index, header in enumerate(normalized) if header in _TONNAGE_HEADERS), None)
    return (machine, tonnage) if machine is not None and tonnage is not None else None


def _section_machine_and_tonnage(value: Any) -> tuple[tuple[str, ...], Decimal | None] | None:
    """Parse grouped P4 headers such as ``Press 27 - 165T - 45mm Screw``."""
    match = _PRESS_SECTION.match(str(value or ""))
    if match is None:
        return None
    return (str(int(match.group(1))),), _tonnage(match.group(2))


def _read_master_press_capacities(source_workbook: str | Path) -> dict[str, Decimal]:
    """Read the approved master press list used only to fill blank P4 group labels."""
    source = Path(source_workbook).resolve()
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            iterator = worksheet.iter_rows(values_only=True)
            headers = next(iterator, None)
            if not headers:
                continue
            normalized = [_normalise_header(header) for header in headers]
            machine_index = next((index for index, header in enumerate(normalized) if header in _MACHINE_HEADERS), None)
            tonnage_index = next((index for index, header in enumerate(normalized) if header in _TONNAGE_HEADERS), None)
            if machine_index is None or tonnage_index is None:
                continue
            result: dict[str, Decimal] = {}
            for values in iterator:
                numbers = _machine_numbers(values[machine_index] if machine_index < len(values) else None)
                tons = _tonnage(values[tonnage_index] if tonnage_index < len(values) else None)
                if numbers is None or tons is None:
                    continue
                for number in numbers:
                    existing = result.get(number)
                    if existing is not None and existing != tons:
                        raise ValueError(f"Conflicting master press capacities for machine {number}.")
                    result[number] = tons
            if result:
                return result
    finally:
        workbook.close()
    raise ValueError("No worksheet has recognized machine and tonnage headers in the master press list.")


def read_press_capacity_workbook(
    source_workbook: str | Path,
    *,
    master_press_list: str | Path | None = None,
) -> list[CapacitySourceRow]:
    """Read every worksheet with recognized capacity headers, without writing it."""
    source = Path(source_workbook).resolve()
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        rows: list[CapacitySourceRow] = []
        for worksheet in workbook.worksheets:
            iterator = worksheet.iter_rows(values_only=True)
            headers = next(iterator, None)
            if not headers:
                continue
            indexes = _header_indexes(headers)
            labels = [str(header or "").strip() for header in headers]
            if indexes is None:
                machine_index = next(
                    (index for index, header in enumerate(_normalise_header(item) for item in headers) if header in _MACHINE_HEADERS),
                    None,
                )
                if machine_index is None:
                    continue
                for row_number, values in enumerate(iterator, start=2):
                    section = _section_machine_and_tonnage(values[machine_index] if machine_index < len(values) else None)
                    if section is None:
                        continue
                    machine_numbers, tons = section
                    raw = {labels[index]: value for index, value in enumerate(values) if labels[index]}
                    rows.append(
                        CapacitySourceRow(
                            sheet=worksheet.title,
                            row_number=row_number,
                            machine_numbers=machine_numbers,
                            tonnage=tons,
                            raw_values=raw,
                            issue=None if tons is not None else "INVALID_PRESS_TONNAGE",
                        )
                    )
                continue
            machine_index, tonnage_index = indexes
            for row_number, values in enumerate(iterator, start=2):
                raw = {labels[index]: value for index, value in enumerate(values) if labels[index]}
                if not any(value is not None and str(value).strip() for value in values):
                    continue
                machine_numbers = _machine_numbers(values[machine_index] if machine_index < len(values) else None)
                tons = _tonnage(values[tonnage_index] if tonnage_index < len(values) else None)
                issue = None
                if machine_numbers is None:
                    issue = "INVALID_MACHINE_NUMBER"
                elif tons is None:
                    issue = "INVALID_PRESS_TONNAGE"
                rows.append(
                    CapacitySourceRow(
                        sheet=worksheet.title,
                        row_number=row_number,
                        machine_numbers=machine_numbers or (),
                        tonnage=tons,
                        raw_values=raw,
                        issue=issue,
                    )
                )
        if master_press_list:
            master_capacities = _read_master_press_capacities(master_press_list)
            rows = [
                replace(
                    row,
                    tonnage=master_capacities[row.machine_numbers[0]],
                    issue=None,
                    capacity_source="master_press_list",
                )
                if row.issue == "INVALID_PRESS_TONNAGE"
                and len(row.machine_numbers) == 1
                and row.machine_numbers[0] in master_capacities
                else row
                for row in rows
            ]
        if not rows:
            raise ValueError("No worksheet has a supported press-capacity layout.")
        return rows
    finally:
        workbook.close()


def plan_press_capacity_import(
    source_workbook: str | Path,
    existing_capacities: Mapping[str, Decimal | None],
    *,
    master_press_list: str | Path | None = None,
) -> PressCapacityReport:
    """Build a deterministic, non-mutating plan against one plant's machines."""
    source = Path(source_workbook).resolve()
    rows = read_press_capacity_workbook(source, master_press_list=master_press_list)
    grouped: dict[str, list[CapacitySourceRow]] = {}
    invalid_rows: list[dict[str, Any]] = []
    for row in rows:
        if row.issue:
            invalid_rows.append({"sheet": row.sheet, "row_number": row.row_number, "issue": row.issue})
            continue
        for number in row.machine_numbers:
            grouped.setdefault(number, []).append(row)
    updates: list[CapacityUpdate] = []
    source_conflicts: dict[str, list[str]] = {}
    existing_conflicts: dict[str, dict[str, str]] = {}
    unmatched: list[str] = []
    matched = 0
    for number, rows_for_machine in sorted(grouped.items(), key=lambda item: int(item[0])):
        values = {row.tonnage for row in rows_for_machine if row.tonnage is not None}
        if len(values) != 1:
            source_conflicts[number] = sorted(str(value) for value in values)
            continue
        source_tonnage = next(iter(values))
        existing = existing_capacities.get(number)
        if number not in existing_capacities:
            unmatched.append(number)
            continue
        matched += 1
        if existing is not None and Decimal(existing) != source_tonnage:
            existing_conflicts[number] = {"existing": str(existing), "source": str(source_tonnage)}
            continue
        updates.append(
            CapacityUpdate(
                machine_number=number,
                source_rows=tuple(sorted(row.row_number for row in rows_for_machine)),
                source_tonnage=source_tonnage,
                existing_tonnage=existing,
                action="UNCHANGED" if existing is not None else "SET_PRESS_CAPACITY",
            )
        )
    return PressCapacityReport(
        source_file_name=source.name,
        source_sha256=_digest(source),
        supplementary_sources={Path(master_press_list).name: _digest(Path(master_press_list))} if master_press_list else {},
        source_rows=len(rows),
        source_machine_count=len(grouped),
        matched_machines=matched,
        unmatched_machines=unmatched,
        conflicting_source_values=source_conflicts,
        conflicting_existing_values=existing_conflicts,
        invalid_rows=invalid_rows,
        updates=updates,
    )


def _plant_capacities(session: Session, plant_code: str) -> dict[str, tuple[int, Decimal | None]]:
    records = session.execute(
        select(db.Machine.id, db.Machine.machine_number, db.Machine.press_capacity_tons)
        .join(db.Plant, db.Machine.plant_id == db.Plant.id)
        .where(db.Plant.plant_code == plant_code)
    ).all()
    return {number: (machine_id, capacity) for machine_id, number, capacity in records}


def run_press_capacity_import(
    source_workbook: str | Path,
    *,
    plant_code: str,
    execute: bool = False,
    master_press_list: str | Path | None = None,
) -> PressCapacityReport:
    """Run the plan; writes require an explicit flag and an entirely clean plan."""
    factory = create_session_factory(migration=True)
    with factory() as session:
        database_rows = _plant_capacities(session, plant_code)
        report = plan_press_capacity_import(
            source_workbook,
            {key: value[1] for key, value in database_rows.items()},
            master_press_list=master_press_list,
        )
        if not execute:
            return report
        if not report.safe_to_execute:
            report.status = "SAFE_STOP_REVIEW_REQUIRED"
            return report
        prior = session.scalar(
            select(db.ImportBatch).where(
                db.ImportBatch.source_type == "press_capacity_workbook",
                db.ImportBatch.source_file_checksum == report.source_sha256,
                db.ImportBatch.status == "COMPLETED",
                db.ImportBatch.dry_run.is_(False),
            )
        )
        if prior:
            report.status = "SAFE_STOP_ALREADY_IMPORTED"
            report.batch_uuid = prior.batch_uuid
            return report
        # The read-only plan and duplicate check start SQLAlchemy's implicit
        # transaction.  End it before beginning the single write transaction.
        session.rollback()
        with session.begin():
            batch = db.ImportBatch(
                batch_uuid=str(uuid4()),
                batch_name="Controlled press-capacity import",
                source_type="press_capacity_workbook",
                source_file_name=report.source_file_name,
                source_file_checksum=report.source_sha256,
                started_at=datetime.now(timezone.utc),
                status="RUNNING",
                dry_run=False,
                application_release_id=ensure_application_release(session).id,
                records_discovered=report.source_rows,
                notes=f"Plant {plant_code}; compatibility and assignments intentionally untouched.",
            )
            session.add(batch)
            session.flush()
            source_rows = read_press_capacity_workbook(source_workbook, master_press_list=master_press_list)
            changed = 0
            for update in report.updates:
                machine_id, existing = database_rows[update.machine_number]
                if existing is None:
                    machine = session.get(db.Machine, machine_id)
                    if machine is None:
                        raise RuntimeError(f"Machine disappeared during import: {update.machine_number}")
                    machine.press_capacity_tons = update.source_tonnage
                    changed += 1
            # ImportRow is deliberately unique by physical source row.  A
            # comma-separated source row can map to several machines, so it
            # receives one provenance row with every normalized target rather
            # than one duplicate row per target.
            updates_by_machine = {update.machine_number: update for update in report.updates}
            for row in source_rows:
                if row.issue or row.tonnage is None:
                    continue
                row_updates = [updates_by_machine[number] for number in row.machine_numbers]
                imported = any(update.existing_tonnage is None for update in row_updates)
                session.add(
                    db.ImportRow(
                        import_batch_id=batch.id,
                        source_sheet=row.sheet,
                        source_row_number=row.row_number,
                        source_identifier=", ".join(row.machine_numbers),
                        target_entity_type="machine" if len(row.machine_numbers) == 1 else "machine_capacity_row",
                        target_entity_id=database_rows[row.machine_numbers[0]][0] if len(row.machine_numbers) == 1 else None,
                        status="IMPORTED" if imported else "UNCHANGED",
                        raw_values_json=row.raw_values,
                        normalized_values_json={
                            "machine_numbers": list(row.machine_numbers),
                            "press_capacity_tons": str(row.tonnage),
                        },
                    )
                )
            batch.records_imported = changed
            batch.records_rejected = 0
            batch.warnings_count = 0
            batch.status = "COMPLETED"
            batch.completed_at = datetime.now(timezone.utc)
            record_import_completion(
                session,
                source=f"press_capacity_workbook:{report.source_sha256[:12]}",
                changed_data=bool(changed),
            )
            report.batch_uuid = batch.batch_uuid
            report.data_changed = bool(changed)
            report.status = "COMPLETED"
    return report


def write_immutable_receipt(report: PressCapacityReport, directory: str | Path) -> Path:
    """Write a redacted receipt once; a prior receipt is never overwritten."""
    output = Path(directory)
    output.mkdir(parents=True, exist_ok=True)
    target = output / f"press-capacity-{report.source_sha256[:16]}.json"
    payload = json.dumps(report.to_dict(), indent=2, sort_keys=True, default=str) + "\n"
    with target.open("x", encoding="utf-8") as stream:
        stream.write(payload)
    return target


def main() -> int:
    parser = argparse.ArgumentParser(description="Dry-run-first controlled press-capacity importer.")
    parser.add_argument("source_workbook")
    parser.add_argument("--plant-code", required=True)
    parser.add_argument("--execute", action="store_true", help="Apply only a conflict-free plan.")
    parser.add_argument("--master-press-list", help="Approved master press list used only for missing P4 group tonnage.")
    parser.add_argument("--receipt-directory", required=True)
    args = parser.parse_args()
    report = run_press_capacity_import(
        args.source_workbook,
        plant_code=args.plant_code,
        execute=args.execute,
        master_press_list=args.master_press_list,
    )
    receipt = write_immutable_receipt(report, args.receipt_directory)
    print(json.dumps({**report.to_dict(), "receipt": str(receipt)}, indent=2, default=str))
    return 0 if report.status in {"DRY_RUN_COMPLETE", "COMPLETED", "SAFE_STOP_ALREADY_IMPORTED"} and report.safe_to_execute else 2


if __name__ == "__main__":
    raise SystemExit(main())
