from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import subprocess
import sys
from collections import defaultdict
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from sqlalchemy import URL, create_engine, text

try:
    from tools.eoat_location_state import STATE_INSTALLED, STATE_STORED, classify_eoat_locations
except ModuleNotFoundError:  # Direct script execution places tools/ on sys.path.
    from eoat_location_state import STATE_INSTALLED, STATE_STORED, classify_eoat_locations

TOOL_VERSION = "1.1.0"
ACTIVE_DATA_SHEETS = {"EOAT Inventory", "Photo Index"}
DERIVED_SHEETS = {"Audit by Press"}
METADATA_SHEETS = {"_EOAT_App_Metadata"}
TEMPLATE_SHEETS = {"Issue Log", "KPI Baseline", "Interview Notes", "Pilot Candidates", "FMEA Draft", "Action Items"}
MISSING_TOKENS = {"", "n/a", "na", "none", "null", "not available", "unknown", "tbd", "-"}
FAILURE_RESULTS = {
    "missing_database_entity",
    "missing_database_relationship",
    "missing_database_value",
    "conflicting_database_value",
    "truncated_database_value",
    "ambiguous_match",
    "unmapped_source_field",
    "invalid_source_value_lost",
}

OUTPUT_SCHEMAS: dict[str, list[str]] = {
    "workbook_sheet_summary.csv": [
        "sheet_name", "visibility", "used_range", "header_rows", "candidate_data_rows", "fully_blank_rows",
        "formula_cells", "constant_cells", "meaningful_nonempty_cells", "merged_ranges", "tables", "hyperlinks",
        "comments", "hidden_rows", "hidden_columns", "data_validations", "named_ranges", "formula_cache_missing",
        "formatting_business_meaning_candidates",
    ],
    "field_mapping_matrix.csv": [
        "workbook_sheet", "source_header", "normalized_header", "source_data_type", "example_nonempty_values",
        "source_row_count", "intended_entity", "destination_mysql_table", "destination_mysql_column",
        "relationship_or_transformation", "normalization_rules", "null_blank_handling", "multi_value_parsing_behavior",
        "comparison_method", "status", "justification",
    ],
    "source_row_reconciliation.csv": [
        "sheet", "source_row", "business_key", "source_fingerprint", "destination_table", "destination_primary_key",
        "result", "details",
    ],
    "source_field_reconciliation.csv": [
        "sheet", "source_row", "business_key", "source_header", "source_value", "destination_table",
        "destination_column", "destination_primary_key", "database_value", "comparison_result", "history_evidence",
    ],
    "relationship_reconciliation.csv": [
        "sheet", "source_row", "source_key", "relationship_type", "source_member", "destination_primary_key", "result", "details",
    ],
    "missing_database_entities.csv": ["sheet", "source_row", "entity_type", "business_key", "details"],
    "missing_database_relationships.csv": ["sheet", "source_row", "relationship_type", "left_key", "right_key", "details"],
    "field_mismatches.csv": [
        "sheet", "source_row", "business_key", "source_header", "source_value", "database_value", "classification", "details",
    ],
    "unmapped_source_fields.csv": ["sheet", "source_header", "populated_count", "examples", "reason"],
    "ambiguous_matches.csv": ["sheet", "source_row", "business_key", "match_count", "details"],
    "duplicate_source_keys.csv": ["sheet", "key_type", "key", "count", "source_rows"],
    "duplicate_database_keys.csv": ["table", "column", "key", "count", "primary_keys"],
    "import_issue_reconciliation.csv": [
        "workbook_sheet", "workbook_row", "source_field", "source_value", "issue_category", "importer_action",
        "database_result", "original_value_preserved", "resolution_status", "manual_review_required",
    ],
    "document_photo_link_reconciliation.csv": [
        "sheet", "source_row", "source_identifier", "source_path", "database_path", "document_id", "photo_id",
        "linked_entities", "file_exists", "result", "details",
    ],
    "database_only_records.csv": ["entity_type", "primary_key", "business_key", "reason"],
}


DIRECT_MAPPINGS: dict[tuple[str, str], tuple[str, str, str, str]] = {
    ("EOAT Inventory", "Audit ID"): ("audit_record", "audit_records", "audit_identifier", "direct"),
    ("EOAT Inventory", "Audit Date"): ("audit_record", "audit_records", "audit_date", "direct"),
    ("EOAT Inventory", "EOAT Assembly ID"): ("eoat", "eoats", "business_identifier", "direct"),
    ("EOAT Inventory", "Press/Machine #"): ("machine", "machines", "machine_number", "normalized_relationship"),
    ("EOAT Inventory", "Tool #"): ("tool", "tools", "tool_number", "normalized_relationship"),
    ("EOAT Inventory", "Plant/Area"): ("area", "areas", "area_name", "normalized_relationship"),
    ("EOAT Inventory", "EOAT Type"): ("eoat", "eoat_types", "display_name", "normalized_relationship"),
    ("EOAT Inventory", "Connection Type"): ("eoat", "connection_types", "display_name", "normalized_relationship"),
    ("EOAT Inventory", "Cleanroom/Non-Cleanroom"): (
        "eoat", "cleanroom_classifications", "display_name", "normalized_relationship"
    ),
    ("EOAT Inventory", "Number of Parts Picked"): ("eoat", "eoats", "number_of_parts_picked", "direct"),
    ("EOAT Inventory", "# of Cups"): ("eoat", "eoats", "number_of_vacuum_cups", "direct"),
    ("EOAT Inventory", "# of Grippers"): ("eoat", "eoats", "number_of_grippers", "direct"),
    ("EOAT Inventory", "Sensors Present?"): ("eoat", "eoats", "sensors_present", "direct"),
    ("EOAT Inventory", "Part-Present Detection Present?"): (
        "eoat", "eoats", "part_present_sensor_present", "direct"
    ),
    ("EOAT Inventory", "Vacuum Confirmation Present?"): (
        "eoat", "eoats", "vacuum_confirmation_sensor_present", "direct"
    ),
    ("EOAT Inventory", "Quick Disconnects Present?"): (
        "eoat", "eoats", "quick_disconnect_present", "direct"
    ),
    ("EOAT Inventory", "Cup Type/Material"): ("eoat", "eoats", "cup_material", "direct"),
    ("EOAT Inventory", "Notes"): ("audit_record", "audit_records", "notes", "direct"),
    ("Photo Index", "Photo ID"): ("photo", "documents", "document_number", "direct"),
    ("Photo Index", "Date Taken"): ("photo", "photos", "captured_at", "direct"),
    ("Photo Index", "Photo Filename"): ("document", "documents", "file_name", "direct"),
    ("Photo Index", "Stored Filename"): ("document", "documents", "file_name", "direct"),
    ("Photo Index", "Stored Relative Path"): ("document", "documents", "storage_path", "derived_equivalent"),
    ("Photo Index", "Folder Path"): ("document", "documents", "storage_path", "derived_equivalent"),
    ("Photo Index", "Description"): ("photo", "photos", "caption", "direct"),
    ("Photo Index", "Notes"): ("document", "documents", "description", "direct"),
    ("Photo Index", "EOAT Assembly ID"): ("document_link", "document_links", "entity_id", "normalized_relationship"),
    ("Photo Index", "Tool #"): ("document_link", "document_links", "entity_id", "normalized_relationship"),
    ("Photo Index", "Press/Machine #"): ("document_link", "document_links", "entity_id", "normalized_relationship"),
}


@dataclass(frozen=True)
class DatabaseConfig:
    host: str
    port: int
    database: str
    username: str
    password: str
    source: str

    @property
    def account_category(self) -> str:
        lowered = self.username.casefold()
        if "migrat" in lowered:
            return "migration_account"
        if "app" in lowered or "runtime" in lowered:
            return "runtime_application_account"
        if "read" in lowered or "audit" in lowered:
            return "read_only_or_audit_account"
        return "configured_database_account"


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def normalized_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", clean_text(value).casefold()).strip("_")


def normalized_value(value: Any, *, header: str = "") -> str:
    result = clean_text(value)
    result = re.sub(r"[ \t]+", " ", result).strip()
    lowered = result.casefold()
    if lowered in MISSING_TOKENS:
        return ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2} 00:00:00(?:\.0+)?", result):
        return result[:10]
    if lowered in {"yes", "y", "true", "present", "1"}:
        return "true"
    if lowered in {"no", "n", "false", "not present", "0"}:
        return "false"
    if "path" in header.casefold() or "file" in header.casefold():
        result = result.replace("/", "\\")
        result = re.sub(r"\\+", r"\\", result)
    return result.casefold()


def comparison(source: Any, destination: Any, *, header: str = "") -> str:
    if clean_text(source) == clean_text(destination):
        return "exact_match"
    if normalized_value(source, header=header) == normalized_value(destination, header=header):
        return "normalized_match"
    source_normalized = normalized_value(source, header=header)
    destination_normalized = normalized_value(destination, header=header)
    if destination_normalized and len(destination_normalized) < len(source_normalized) and source_normalized.startswith(destination_normalized):
        return "truncated_database_value"
    return "conflicting_database_value"


def split_multi_value(value: Any) -> list[str]:
    raw = clean_text(value)
    if not raw:
        return []
    return [part.strip() for part in re.split(r"[,;|\n]+", raw) if part.strip()]


def source_fingerprint(sheet: str, row_number: int, values: dict[str, Any]) -> str:
    payload = json.dumps(
        {key: clean_text(value) for key, value in sorted(values.items())},
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(f"{sheet}|{row_number}|{payload}".encode()).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def json_value(value: Any) -> Any:
    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return value
    return value


def is_presentation_note(values: Iterable[Any]) -> bool:
    populated = [clean_text(value) for value in values if clean_text(value)]
    return bool(populated) and all(value.casefold().startswith("last updated:") for value in populated)


def find_header_row(sheet) -> int:
    if sheet.title == "Audit by Press":
        return 3
    if sheet.tables:
        first = next(iter(sheet.tables.values()))
        _, min_row, _, _ = range_boundaries(first.ref)
        return min_row
    best_row, best_score = 1, -1
    keywords = {"id", "date", "notes", "tool", "machine", "eoat", "key", "value"}
    for row in range(1, min(sheet.max_row, 10) + 1):
        values = [clean_text(sheet.cell(row, col).value) for col in range(1, sheet.max_column + 1)]
        score = sum(bool(value) for value in values) + sum(
            any(word in value.casefold() for word in keywords) for value in values if value
        )
        if score > best_score:
            best_row, best_score = row, score
    return best_row


def candidate_rows(sheet, header_row: int) -> list[int]:
    rows = []
    for row in range(header_row + 1, sheet.max_row + 1):
        values = [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]
        if any(clean_text(value) for value in values):
            rows.append(row)
    return rows


def workbook_inventory(path: Path) -> tuple[dict[str, Any], dict[str, dict[int, dict[str, Any]]], dict[str, list[str]]]:
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=True)
    cached = load_workbook(path, read_only=False, data_only=True, keep_links=True)
    defined_names = []
    for item in workbook.defined_names.values():
        defined_names.append({"name": item.name, "value": item.attr_text, "hidden": bool(item.hidden)})
    sheets: list[dict[str, Any]] = []
    row_data: dict[str, dict[int, dict[str, Any]]] = {}
    headers_by_sheet: dict[str, list[str]] = {}
    for sheet in workbook.worksheets:
        cached_sheet = cached[sheet.title]
        header_row = find_header_row(sheet)
        headers = [clean_text(sheet.cell(header_row, col).value) or f"Column {get_column_letter(col)}" for col in range(1, sheet.max_column + 1)]
        headers_by_sheet[sheet.title] = headers
        candidates = candidate_rows(sheet, header_row)
        row_data[sheet.title] = {
            row: {headers[col - 1]: sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)}
            for row in candidates
        }
        formula_cells = []
        constant_count = 0
        hyperlink_cells = []
        comment_cells = []
        formula_cache_missing = []
        formatting_candidates = []
        nonempty_rows = set()
        meaningful_nonempty = 0
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if value not in (None, ""):
                    nonempty_rows.add(cell.row)
                    if isinstance(value, str) and value.startswith("="):
                        formula_cells.append(cell.coordinate)
                        if cached_sheet[cell.coordinate].value is None:
                            formula_cache_missing.append(cell.coordinate)
                    else:
                        constant_count += 1
                    if cell.row > header_row and sheet.title in ACTIVE_DATA_SHEETS | METADATA_SHEETS:
                        meaningful_nonempty += 1
                if cell.hyperlink:
                    hyperlink_cells.append({"cell": cell.coordinate, "target": cell.hyperlink.target})
                if cell.comment:
                    comment_cells.append({"cell": cell.coordinate, "author": cell.comment.author, "text": cell.comment.text})
                has_fill = cell.fill and cell.fill.fill_type and cell.fill.fgColor.type != "indexed"
                if (has_fill or cell.font.strike or cell.font.color is not None) and value in (None, ""):
                    formatting_candidates.append({
                        "cell": cell.coordinate, "style_id": cell.style_id, "fill": cell.fill.fgColor.rgb,
                        "font_color": getattr(cell.font.color, "rgb", None), "strikethrough": bool(cell.font.strike),
                    })
        tables = [{"name": table.name, "display_name": table.displayName, "ref": table.ref} for table in sheet.tables.values()]
        validations = []
        if sheet.data_validations:
            for validation in sheet.data_validations.dataValidation:
                validations.append({
                    "type": validation.type, "ranges": str(validation.sqref), "formula1": validation.formula1,
                    "formula2": validation.formula2,
                })
        named_for_sheet = [
            item for item in defined_names if item["value"] and sheet.title.casefold() in item["value"].casefold()
        ]
        hidden_rows = [index for index, dimension in sheet.row_dimensions.items() if dimension.hidden]
        hidden_columns = [key for key, dimension in sheet.column_dimensions.items() if dimension.hidden]
        blank_rows = sum(row not in nonempty_rows for row in range(1, sheet.max_row + 1))
        sheets.append({
            "sheet_name": sheet.title,
            "visibility": sheet.sheet_state,
            "used_range": sheet.calculate_dimension(),
            "header_rows": [header_row],
            "candidate_data_rows": len(candidates),
            "candidate_row_numbers": candidates,
            "fully_blank_rows": blank_rows,
            "formula_cells": len(formula_cells),
            "formula_cell_addresses": formula_cells,
            "constant_cells": constant_count,
            "meaningful_nonempty_cells": meaningful_nonempty,
            "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
            "tables": tables,
            "hyperlinks": hyperlink_cells,
            "comments": comment_cells,
            "hidden_rows": hidden_rows,
            "hidden_columns": hidden_columns,
            "data_validations": validations,
            "named_ranges": named_for_sheet,
            "formula_cache_missing": formula_cache_missing,
            "formatting_business_meaning_candidates": formatting_candidates[:250],
            "conditional_formatting_rule_count": sum(len(item.rules) for item in sheet.conditional_formatting),
        })
    inventory = {
        "path": str(path),
        "file_size": path.stat().st_size,
        "last_modified": datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat(),
        "sha256": sha256_file(path),
        "calculation_mode": getattr(workbook.calculation, "calcMode", None),
        "defined_names": defined_names,
        "sheet_count": len(sheets),
        "sheet_names": workbook.sheetnames,
        "sheets": sheets,
    }
    workbook.close()
    cached.close()
    return inventory, row_data, headers_by_sheet


def source_type(values: list[Any]) -> str:
    names = {type(value).__name__ for value in values if value is not None}
    return ",".join(sorted(names)) or "blank"


def mapping_matrix(
    rows: dict[str, dict[int, dict[str, Any]]], headers: dict[str, list[str]]
) -> list[dict[str, Any]]:
    result = []
    for sheet, sheet_headers in headers.items():
        for header in sheet_headers:
            values = [record.get(header) for record in rows.get(sheet, {}).values() if clean_text(record.get(header))]
            if sheet in DERIVED_SHEETS:
                entity, table, column, status = "report_view", "audit_records", "details_json", "derived_equivalent"
                transform = "Generated presentation view derived from EOAT Inventory"
                justification = "The sheet is a refreshed grouped report; underlying facts originate in EOAT Inventory."
            elif sheet in TEMPLATE_SHEETS:
                entity, table, column, status = "template", "", "", "presentation_only"
                transform = "Empty workbook template column"
                justification = "No candidate business-data values exist; Last Updated markers are presentation metadata."
            elif sheet in METADATA_SHEETS:
                entity, table, column, status = "workbook_metadata", "", "", "unmapped_failure"
                transform = "No importer mapping detected"
                justification = "Populated workbook application metadata is not represented by the matching import batch."
            elif (sheet, header) in DIRECT_MAPPINGS:
                entity, table, column, status = DIRECT_MAPPINGS[(sheet, header)]
                transform = "Importer direct/normalized mapping plus import_rows.raw_values_json provenance"
                justification = "Importer code supplies a structured destination and preserves the complete source row."
            elif sheet in ACTIVE_DATA_SHEETS:
                entity, table, column, status = "source_row", "import_rows", "raw_values_json", "preserved_in_provenance"
                transform = "Exact source header/value retained in JSON provenance"
                justification = "No structured field mapping exists, but the complete populated source value is retained."
            else:
                entity, table, column, status = "unknown", "", "", "unmapped_failure"
                transform = "No mapping"
                justification = "No defensible database representation was found."
            result.append({
                "workbook_sheet": sheet,
                "source_header": header,
                "normalized_header": normalized_header(header),
                "source_data_type": source_type(values),
                "example_nonempty_values": json.dumps([clean_text(value) for value in values[:3]], ensure_ascii=False),
                "source_row_count": len(values),
                "intended_entity": entity,
                "destination_mysql_table": table,
                "destination_mysql_column": column,
                "relationship_or_transformation": transform,
                "normalization_rules": "trim; normalize line endings; integer-like floats; casefold only for case-insensitive comparison",
                "null_blank_handling": "Blank remains blank/null; blank is never coerced to false, zero, No, or Unknown",
                "multi_value_parsing_behavior": "Comma/semicolon/pipe/newline members compared individually when relationship semantics apply",
                "comparison_method": "exact then permitted normalized comparison; structured fields and relationships checked separately",
                "status": status,
                "justification": justification,
            })
    return result


def load_database_config(value: str | None) -> DatabaseConfig:
    if value in (None, "", "development"):
        base = Path(os.environ.get("LOCALAPPDATA", Path.home() / "AppData" / "Local"))
        path = base / "EOAT Atlas Development" / "database.env"
    else:
        path = Path(value).expanduser()
    if not path.is_file():
        raise FileNotFoundError(f"Database environment file is unavailable: {path}")
    settings: dict[str, str] = {}
    for raw in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw.strip()
        if line and not line.startswith("#") and "=" in line:
            key, item = line.split("=", 1)
            if key.strip().startswith("EOAT_"):
                settings[key.strip()] = item.strip()
    return DatabaseConfig(
        host=settings.get("EOAT_DB_HOST", "127.0.0.1"),
        port=int(settings.get("EOAT_DB_PORT", "3306")),
        database=settings.get("EOAT_DB_NAME", "eoat_atlas_dev"),
        username=settings.get("EOAT_DB_USER", "eoat_atlas_app"),
        password=settings.get("EOAT_DB_PASSWORD", ""),
        source=str(path),
    )


def rows_as_dicts(connection, query: str, parameters: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    return [dict(row._mapping) for row in connection.execute(text(query), parameters or {})]


def scalar(connection, query: str, parameters: dict[str, Any] | None = None) -> Any:
    return connection.execute(text(query), parameters or {}).scalar()


def table_exists(connection, table: str) -> bool:
    return bool(scalar(
        connection,
        "SELECT COUNT(*) FROM information_schema.tables WHERE table_schema=DATABASE() AND table_name=:table",
        {"table": table},
    ))


def database_snapshot(config: DatabaseConfig, workbook_hash: str, *, read_only: bool = True) -> dict[str, Any]:
    engine = create_engine(URL.create(
        "mysql+pymysql", username=config.username, password=config.password, host=config.host, port=config.port,
        database=config.database, query={"charset": "utf8mb4"}
    ), pool_pre_ping=True)
    with engine.connect() as connection:
        if read_only:
            connection.exec_driver_sql("SET SESSION TRANSACTION READ ONLY")
        info = {
            "host": config.host,
            "port": config.port,
            "schema": config.database,
            "account_category": config.account_category,
            "server_version": scalar(connection, "SELECT VERSION()"),
            "authenticated_account": scalar(connection, "SELECT CURRENT_USER()"),
            "alembic_revision": scalar(connection, "SELECT version_num FROM alembic_version"),
            "table_count": int(scalar(
                connection,
                "SELECT COUNT(*) FROM information_schema.tables WHERE table_schema=DATABASE() AND table_type='BASE TABLE'",
            ) or 0),
            "environment": "development" if config.database.casefold().endswith(("_dev", "_development")) else (
                "test" if "test" in config.database.casefold() else "production_or_unclassified"
            ),
            "read_only_session_requested": read_only,
        }
        batches = rows_as_dicts(connection, """
            SELECT * FROM import_batches
            WHERE LOWER(source_file_checksum)=LOWER(:checksum) AND status='COMPLETED' AND dry_run=0
            ORDER BY id DESC
        """, {"checksum": workbook_hash})
        batch = batches[0] if batches else None
        batch_id = batch["id"] if batch else -1
        import_rows = rows_as_dicts(connection, """
            SELECT id,import_batch_id,source_sheet,source_row_number,source_identifier,target_entity_type,
                   target_entity_id,status,raw_values_json,normalized_values_json,error_summary
            FROM import_rows WHERE import_batch_id=:batch_id ORDER BY source_sheet,source_row_number
        """, {"batch_id": batch_id})
        for record in import_rows:
            record["raw_values_json"] = json_value(record.get("raw_values_json")) or {}
            record["normalized_values_json"] = json_value(record.get("normalized_values_json")) or {}
        issues = rows_as_dicts(connection, """
            SELECT ii.*,ir.source_sheet,ir.source_row_number,ir.source_identifier,ir.status AS import_row_status,
                   ir.raw_values_json
            FROM import_issues ii LEFT JOIN import_rows ir ON ir.id=ii.import_row_id
            WHERE ii.import_batch_id=:batch_id ORDER BY ii.id
        """, {"batch_id": batch_id})
        for record in issues:
            record["resolution_notes"] = json_value(record.get("resolution_notes")) or {}
            record["raw_values_json"] = json_value(record.get("raw_values_json")) or {}
        eoats = rows_as_dicts(connection, """
            SELECT e.*,et.display_name AS eoat_type,ct.display_name AS connection_type,
                   cc.display_name AS cleanroom_classification,ast.display_name AS status
            FROM eoats e
            LEFT JOIN eoat_types et ON et.id=e.eoat_type_id
            LEFT JOIN connection_types ct ON ct.id=e.connection_type_id
            LEFT JOIN cleanroom_classifications cc ON cc.id=e.cleanroom_classification_id
            LEFT JOIN asset_statuses ast ON ast.id=e.status_id
        """)
        audits = rows_as_dicts(connection, """
            SELECT a.*,e.business_identifier AS eoat_identifier,m.machine_number,t.tool_number,
                   ast.display_name AS status
            FROM audit_records a
            LEFT JOIN eoats e ON e.id=a.eoat_id
            LEFT JOIN machines m ON m.id=a.machine_id
            LEFT JOIN tools t ON t.id=a.tool_id
            LEFT JOIN asset_statuses ast ON ast.id=a.status_id
        """)
        for record in audits:
            record["details_json"] = json_value(record.get("details_json")) or {}
        machines = rows_as_dicts(connection, "SELECT m.*,a.area_name FROM machines m LEFT JOIN areas a ON a.id=m.area_id")
        tools = rows_as_dicts(connection, "SELECT * FROM tools")
        relationships = {
            "eoat_machine": rows_as_dicts(connection, """
                SELECT c.id,e.business_identifier AS eoat_identifier,m.machine_number,c.is_active,c.reason
                FROM eoat_machine_compatibility c JOIN eoats e ON e.id=c.eoat_id JOIN machines m ON m.id=c.machine_id
            """),
            "eoat_tool": rows_as_dicts(connection, """
                SELECT c.id,e.business_identifier AS eoat_identifier,t.tool_number,c.is_active,c.reason
                FROM eoat_tool_compatibility c JOIN eoats e ON e.id=c.eoat_id JOIN tools t ON t.id=c.tool_id
            """),
            "tool_machine": rows_as_dicts(connection, """
                SELECT c.id,t.tool_number,m.machine_number,c.is_active,c.reason
                FROM tool_machine_compatibility c JOIN tools t ON t.id=c.tool_id JOIN machines m ON m.id=c.machine_id
            """),
            "installations": rows_as_dicts(connection, """
                SELECT i.id,e.business_identifier AS eoat_identifier,m.machine_number,
                       (i.removed_at IS NULL) AS is_current,i.installed_at,i.removed_at
                FROM eoat_installations i JOIN eoats e ON e.id=i.eoat_id JOIN machines m ON m.id=i.machine_id
            """),
            "storage_assignments": rows_as_dicts(connection, """
                SELECT a.id,e.business_identifier AS eoat_identifier,l.location_code,
                       (a.removed_from_storage_at IS NULL) AS is_current,a.stored_at,a.removed_from_storage_at
                FROM eoat_storage_assignments a JOIN eoats e ON e.id=a.eoat_id
                JOIN storage_locations l ON l.id=a.storage_location_id
            """),
        }
        photos = rows_as_dicts(connection, """
            SELECT ir.source_row_number,ir.source_identifier,ir.status AS import_status,
                   p.id AS photo_id,p.captured_at,p.caption,p.photo_view_type,
                   d.id AS document_id,d.document_number,d.file_name,d.storage_path,d.description
            FROM import_rows ir
            LEFT JOIN photos p ON ir.target_entity_type='photo' AND p.id=ir.target_entity_id
            LEFT JOIN documents d ON d.id=p.document_id
            WHERE ir.import_batch_id=:batch_id AND ir.source_sheet='Photo Index'
        """, {"batch_id": batch_id})
        document_links = rows_as_dicts(connection, """
            SELECT dl.id,dl.document_id,dl.entity_type,dl.entity_id,
                   CASE dl.entity_type
                       WHEN 'eoat' THEN (SELECT business_identifier FROM eoats WHERE id=dl.entity_id)
                       WHEN 'tool' THEN (SELECT tool_number FROM tools WHERE id=dl.entity_id)
                       WHEN 'machine' THEN (SELECT machine_number FROM machines WHERE id=dl.entity_id)
                       ELSE CAST(dl.entity_id AS CHAR)
                   END AS entity_identifier
            FROM document_links dl
        """)
        history = rows_as_dicts(connection, """
            SELECT id,entity_type,entity_id,occurred_at,summary,reason,previous_values_json,new_values_json,metadata_json
            FROM entity_history_events ORDER BY id
        """)
        audit_log = rows_as_dicts(connection, """
            SELECT id,entity_type,entity_id,occurred_at,action,previous_values_json,new_values_json,changed_fields_json,
                   reason,source,success FROM change_audit_log ORDER BY id
        """)
        database_only_extras = []
        for entity_type, query in (
            ("tag", "SELECT id,tag_code AS business_key FROM tags"),
            ("annotation", "SELECT id,annotation_uuid AS business_key FROM annotations"),
            ("annotation_target", "SELECT id,target_uuid AS business_key FROM annotation_targets"),
            ("entity_tag", "SELECT id,COALESCE(source_record_identifier,CONCAT(entity_type,':',entity_id,':',tag_id)) AS business_key FROM entity_tags"),
            ("maintenance_event", "SELECT id,event_uuid AS business_key FROM maintenance_events"),
        ):
            if table_exists(connection, query.split(" FROM ", 1)[1].split()[0]):
                for record in rows_as_dicts(connection, query):
                    database_only_extras.append({"entity_type": entity_type, **record})
        counts = {}
        for item in rows_as_dicts(connection, """
            SELECT table_name AS name FROM information_schema.tables
            WHERE table_schema=DATABASE() AND table_type='BASE TABLE' ORDER BY table_name
        """):
            name = item["name"]
            counts[name] = int(scalar(connection, f"SELECT COUNT(*) FROM `{name}`") or 0)
        duplicate_database_keys = []
        for table, column, pk in (
            ("eoats", "business_identifier", "id"), ("machines", "machine_number", "id"),
            ("tools", "tool_number", "id"), ("audit_records", "audit_identifier", "id"),
            ("documents", "document_number", "id"),
        ):
            duplicate_database_keys.extend(rows_as_dicts(connection, f"""
                SELECT '{table}' AS `table`,'{column}' AS `column`,CAST(`{column}` AS CHAR) AS `key`,
                       COUNT(*) AS `count`,GROUP_CONCAT(`{pk}` ORDER BY `{pk}`) AS primary_keys
                FROM `{table}` WHERE `{column}` IS NOT NULL
                GROUP BY `{column}` HAVING COUNT(*)>1
            """))
        foreign_keys = rows_as_dicts(connection, """
            SELECT k.table_name AS table_name,k.column_name AS column_name,
                   k.referenced_table_name AS referenced_table_name,
                   k.referenced_column_name AS referenced_column_name,k.constraint_name AS constraint_name
            FROM information_schema.key_column_usage k
            WHERE k.table_schema=DATABASE() AND k.referenced_table_name IS NOT NULL
            ORDER BY k.table_name,k.constraint_name,k.ordinal_position
        """)
        orphan_checks = []
        for fk in foreign_keys:
            count = int(scalar(connection, f"""
                SELECT COUNT(*) FROM `{fk['table_name']}` c
                LEFT JOIN `{fk['referenced_table_name']}` p
                  ON c.`{fk['column_name']}`=p.`{fk['referenced_column_name']}`
                WHERE c.`{fk['column_name']}` IS NOT NULL AND p.`{fk['referenced_column_name']}` IS NULL
            """) or 0)
            orphan_checks.append(fk | {"orphan_count": count, "status": "PASS" if count == 0 else "FAIL"})
    engine.dispose()
    return {
        "info": info, "batch": batch, "import_rows": import_rows, "issues": issues, "eoats": eoats,
        "audits": audits, "machines": machines, "tools": tools, "relationships": relationships, "photos": photos,
        "document_links": document_links, "history": history, "audit_log": audit_log, "counts": counts,
        "duplicate_database_keys": duplicate_database_keys, "orphan_checks": orphan_checks,
        "database_only_extras": database_only_extras,
    }


def duplicate_source_keys(rows: dict[str, dict[int, dict[str, Any]]]) -> list[dict[str, Any]]:
    output = []
    for sheet, field in (("EOAT Inventory", "Audit ID"), ("Photo Index", "Photo ID")):
        grouped: dict[str, list[int]] = defaultdict(list)
        for row_number, record in rows.get(sheet, {}).items():
            key = clean_text(record.get(field))
            if key:
                grouped[normalized_value(key)].append(row_number)
        for key, source_rows in grouped.items():
            if len(source_rows) > 1:
                output.append({
                    "sheet": sheet, "key_type": field, "key": key, "count": len(source_rows),
                    "source_rows": json.dumps(source_rows),
                })
    return output


def _indexed(records: list[dict[str, Any]], key: str) -> dict[str, list[dict[str, Any]]]:
    result: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        value = clean_text(record.get(key))
        if value:
            result[normalized_value(value)].append(record)
    return result


def _direct_database_value(
    sheet: str,
    header: str,
    source: dict[str, Any],
    indexes: dict[str, dict[str, list[dict[str, Any]]]],
) -> tuple[str, Any, int | None, int]:
    if sheet == "EOAT Inventory":
        audit_key = normalized_value(source.get("Audit ID"))
        eoat_key = normalized_value(source.get("EOAT Assembly ID"))
        audits = indexes["audits"].get(audit_key, [])
        eoats = indexes["eoats"].get(eoat_key, [])
        audit_fields = {
            "Audit ID": "audit_identifier", "Audit Date": "audit_date", "Notes": "notes",
        }
        eoat_fields = {
            "EOAT Assembly ID": "business_identifier", "Number of Parts Picked": "number_of_parts_picked",
            "# of Cups": "number_of_vacuum_cups", "# of Grippers": "number_of_grippers",
            "Sensors Present?": "sensors_present", "Part-Present Detection Present?": "part_present_sensor_present",
            "Vacuum Confirmation Present?": "vacuum_confirmation_sensor_present",
            "Quick Disconnects Present?": "quick_disconnect_present", "Cup Type/Material": "cup_material",
        }
        if header in audit_fields:
            return "audit_records", audits[0].get(audit_fields[header]) if len(audits) == 1 else None, (
                audits[0]["id"] if len(audits) == 1 else None
            ), len(audits)
        if header in eoat_fields:
            return "eoats", eoats[0].get(eoat_fields[header]) if len(eoats) == 1 else None, (
                eoats[0]["id"] if len(eoats) == 1 else None
            ), len(eoats)
    if sheet == "Photo Index":
        photo_key = normalized_value(source.get("Photo ID"))
        photos = indexes["photos"].get(photo_key, [])
        fields = {
            "Photo ID": "document_number", "Date Taken": "captured_at", "Photo Filename": "file_name",
            "Stored Filename": "file_name", "Description": "caption", "Notes": "description",
        }
        if header in fields:
            return "documents/photos", photos[0].get(fields[header]) if len(photos) == 1 else None, (
                photos[0].get("photo_id") if len(photos) == 1 else None
            ), len(photos)
    return "", None, None, 0


def find_change_history(
    database: dict[str, Any], entity_type: str, entity_id: int | None, field_name: str, source_value: Any
) -> str:
    if entity_id is None:
        return ""
    aliases = {entity_type, entity_type.removesuffix("_record"), entity_type.removesuffix("s")}
    for record in [*database.get("audit_log", []), *database.get("history", [])]:
        if clean_text(record.get("entity_type")) not in aliases or int(record.get("entity_id") or -1) != int(entity_id):
            continue
        previous = json_value(record.get("previous_values_json")) or {}
        changed = json_value(record.get("changed_fields_json")) or []
        if isinstance(previous, dict) and field_name in previous and comparison(source_value, previous[field_name], header=field_name) in {
            "exact_match", "normalized_match"
        }:
            timestamp = clean_text(record.get("occurred_at"))
            reason = clean_text(record.get("reason"))
            return f"event_id={record.get('id')}; timestamp={timestamp}; reason={reason}; field={field_name}"
        if field_name in changed and isinstance(previous, dict):
            for old_value in previous.values():
                if comparison(source_value, old_value, header=field_name) in {"exact_match", "normalized_match"}:
                    return f"event_id={record.get('id')}; timestamp={clean_text(record.get('occurred_at'))}; field={field_name}"
    return ""


def issue_original_preserved(issue: dict[str, Any]) -> bool:
    source_field = clean_text(issue.get("field_name"))
    raw = json_value(issue.get("raw_values_json")) or {}
    resolution = json_value(issue.get("resolution_notes")) or {}
    return bool(
        issue.get("source_value") is not None
        or (source_field and isinstance(raw, dict) and source_field in raw)
        or (issue.get("issue_code") and source_field and isinstance(resolution, dict))
    )


def reconcile(
    workbook_path: Path,
    inventory: dict[str, Any],
    source_rows: dict[str, dict[int, dict[str, Any]]],
    mappings: list[dict[str, Any]],
    database: dict[str, Any],
    location_audit: dict[str, Any],
    *,
    include_file_existence_check: bool,
) -> dict[str, Any]:
    mapping_index = {(item["workbook_sheet"], item["source_header"]): item for item in mappings}
    import_row_index = {
        (item["source_sheet"], int(item["source_row_number"])): item for item in database["import_rows"]
    }
    indexes = {
        "audits": _indexed(database["audits"], "audit_identifier"),
        "eoats": _indexed(database["eoats"], "business_identifier"),
        "machines": _indexed(database["machines"], "machine_number"),
        "tools": _indexed(database["tools"], "tool_number"),
        "photos": _indexed(database["photos"], "source_identifier"),
    }
    relation_sets = {
        "eoat_machine": {(normalized_value(item["eoat_identifier"]), normalized_value(item["machine_number"])): item
                         for item in database["relationships"]["eoat_machine"]},
        "eoat_tool": {(normalized_value(item["eoat_identifier"]), normalized_value(item["tool_number"])): item
                      for item in database["relationships"]["eoat_tool"]},
        "tool_machine": {(normalized_value(item["tool_number"]), normalized_value(item["machine_number"])): item
                         for item in database["relationships"]["tool_machine"]},
        "installations": {(normalized_value(item["eoat_identifier"]), normalized_value(item["machine_number"])): item
                          for item in database["relationships"]["installations"] if item.get("is_current")},
    }
    document_links: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for link in database["document_links"]:
        document_links[int(link["document_id"])].append(link)

    row_results: list[dict[str, Any]] = []
    field_results: list[dict[str, Any]] = []
    relationship_results: list[dict[str, Any]] = []
    missing_entities: list[dict[str, Any]] = []
    missing_relationships: list[dict[str, Any]] = []
    mismatches: list[dict[str, Any]] = []
    ambiguous: list[dict[str, Any]] = []
    document_results: list[dict[str, Any]] = []

    for sheet, rows in source_rows.items():
        for row_number, source in rows.items():
            if sheet == "EOAT Inventory":
                business_key = clean_text(source.get("Audit ID"))
            elif sheet == "Photo Index":
                business_key = clean_text(source.get("Photo ID")) or f"row-{row_number}"
            elif sheet in METADATA_SHEETS:
                business_key = clean_text(source.get("key")) or f"row-{row_number}"
            else:
                business_key = f"row-{row_number}"
            fingerprint = source_fingerprint(sheet, row_number, source)
            if sheet in DERIVED_SHEETS:
                row_results.append({
                    "sheet": sheet, "source_row": row_number, "business_key": business_key,
                    "source_fingerprint": fingerprint, "destination_table": "audit_records/details_json",
                    "destination_primary_key": "", "result": "equivalent_normalized_relationship",
                    "details": "Derived grouped report; source facts are reconciled through EOAT Inventory.",
                })
                continue
            if sheet in TEMPLATE_SHEETS and is_presentation_note(source.values()):
                row_results.append({
                    "sheet": sheet, "source_row": row_number, "business_key": business_key,
                    "source_fingerprint": fingerprint, "destination_table": "", "destination_primary_key": "",
                    "result": "intentionally_excluded", "details": "Presentation-only Last Updated marker.",
                })
                continue
            if sheet in METADATA_SHEETS:
                row_results.append({
                    "sheet": sheet, "source_row": row_number, "business_key": business_key,
                    "source_fingerprint": fingerprint, "destination_table": "", "destination_primary_key": "",
                    "result": "unmapped_source_field", "details": "Workbook application metadata is not preserved in MySQL.",
                })
                for header, value in source.items():
                    if clean_text(value):
                        field_results.append({
                            "sheet": sheet, "source_row": row_number, "business_key": business_key,
                            "source_header": header, "source_value": clean_text(value), "destination_table": "",
                            "destination_column": "", "destination_primary_key": "", "database_value": "",
                            "comparison_result": "unmapped_source_field", "history_evidence": "",
                        })
                continue
            if sheet not in ACTIVE_DATA_SHEETS:
                row_results.append({
                    "sheet": sheet, "source_row": row_number, "business_key": business_key,
                    "source_fingerprint": fingerprint, "destination_table": "", "destination_primary_key": "",
                    "result": "intentionally_excluded", "details": "Non-data workbook content.",
                })
                continue
            imported = import_row_index.get((sheet, row_number))
            if imported is None:
                row_result = "missing_database_entity"
                destination_pk = ""
                missing_entities.append({
                    "sheet": sheet, "source_row": row_number, "entity_type": "import_row",
                    "business_key": business_key, "details": "No source provenance record exists for the matching import batch.",
                })
            else:
                row_result = "exact_match" if imported["status"] == "IMPORTED" else "invalid_source_value_preserved"
                destination_pk = imported.get("target_entity_id") or ""
            row_results.append({
                "sheet": sheet, "source_row": row_number, "business_key": business_key,
                "source_fingerprint": fingerprint, "destination_table": imported.get("target_entity_type", "") if imported else "",
                "destination_primary_key": destination_pk, "result": row_result,
                "details": imported.get("error_summary", "") if imported else "Missing import provenance row",
            })
            for header, value in source.items():
                if not clean_text(value):
                    continue
                mapping = mapping_index[(sheet, header)]
                raw = imported.get("raw_values_json", {}) if imported else {}
                database_value = raw.get(header) if isinstance(raw, dict) else None
                result = comparison(value, database_value, header=header) if imported and header in raw else "missing_database_value"
                destination_table = mapping["destination_mysql_table"]
                destination_column = mapping["destination_mysql_column"]
                destination_primary_key = imported.get("id") if imported else ""
                history_evidence = "Source value retained in import_rows.raw_values_json."
                placeholder_photo = sheet == "Photo Index" and not clean_text(source.get("Photo ID")) and not (
                    clean_text(source.get("Stored Relative Path")) or clean_text(source.get("Folder Path"))
                )
                if mapping["status"] == "direct" and not placeholder_photo:
                    current_table, current_value, current_pk, match_count = _direct_database_value(
                        sheet, header, source, indexes
                    )
                    if match_count > 1:
                        result = "ambiguous_match"
                        ambiguous.append({
                            "sheet": sheet, "source_row": row_number, "business_key": business_key,
                            "match_count": match_count, "details": f"Multiple current records matched {current_table}.",
                        })
                    elif match_count == 1:
                        current_result = comparison(value, current_value, header=header)
                        destination_table, destination_primary_key = current_table, current_pk
                        database_value = current_value
                        if current_result == "conflicting_database_value":
                            evidence = find_change_history(
                                database, mapping["intended_entity"], current_pk, mapping["destination_mysql_column"], value
                            )
                            if evidence:
                                result = "current_value_changed_after_import_with_history"
                                history_evidence = evidence
                            else:
                                result = current_result
                                history_evidence = "Original source is preserved in provenance, but no field-specific change proof was established."
                    elif header not in {"Stored Filename"}:
                        result = "missing_database_entity"
                if result in FAILURE_RESULTS:
                    mismatches.append({
                        "sheet": sheet, "source_row": row_number, "business_key": business_key,
                        "source_header": header, "source_value": clean_text(value), "database_value": clean_text(database_value),
                        "classification": result, "details": history_evidence,
                    })
                field_results.append({
                    "sheet": sheet, "source_row": row_number, "business_key": business_key,
                    "source_header": header, "source_value": clean_text(value), "destination_table": destination_table,
                    "destination_column": destination_column, "destination_primary_key": destination_primary_key,
                    "database_value": clean_text(database_value), "comparison_result": result,
                    "history_evidence": history_evidence,
                })

    def add_relationship(sheet: str, row: int, source_key: str, kind: str, member: str, found: dict[str, Any] | None,
                         left: str, right: str, details: str = "") -> None:
        result = "equivalent_normalized_relationship" if found else "missing_database_relationship"
        relationship_results.append({
            "sheet": sheet, "source_row": row, "source_key": source_key, "relationship_type": kind,
            "source_member": member, "destination_primary_key": found.get("id", "") if found else "",
            "result": result, "details": details or (found.get("reason", "") if found else "Expected relationship not found"),
        })
        if not found:
            missing_relationships.append({
                "sheet": sheet, "source_row": row, "relationship_type": kind, "left_key": left, "right_key": right,
                "details": details or "Expected normalized relationship not found.",
            })

    for row_number, source in source_rows.get("EOAT Inventory", {}).items():
        audit = clean_text(source.get("Audit ID"))
        eoat = clean_text(source.get("EOAT Assembly ID"))
        machine = clean_text(source.get("Press/Machine #"))
        tool = clean_text(source.get("Tool #"))
        if eoat and not indexes["eoats"].get(normalized_value(eoat)):
            missing_entities.append({"sheet": "EOAT Inventory", "source_row": row_number, "entity_type": "eoat", "business_key": eoat,
                                     "details": "EOAT business identifier not found."})
        if machine and machine.casefold() not in MISSING_TOKENS and machine.isdigit():
            if not indexes["machines"].get(normalized_value(machine)):
                missing_entities.append({"sheet": "EOAT Inventory", "source_row": row_number, "entity_type": "machine",
                                         "business_key": machine, "details": "Machine number not found."})
            if eoat:
                add_relationship("EOAT Inventory", row_number, audit, "eoat_to_machine_compatibility", machine,
                                 relation_sets["eoat_machine"].get((normalized_value(eoat), normalized_value(machine))), eoat, machine)
        if tool and tool.casefold() not in MISSING_TOKENS:
            if not indexes["tools"].get(normalized_value(tool)):
                missing_entities.append({"sheet": "EOAT Inventory", "source_row": row_number, "entity_type": "tool",
                                         "business_key": tool, "details": "Tool number not found."})
            if eoat:
                add_relationship("EOAT Inventory", row_number, audit, "eoat_to_tool_compatibility", tool,
                                 relation_sets["eoat_tool"].get((normalized_value(eoat), normalized_value(tool))), eoat, tool)
        if tool and tool.casefold() not in MISSING_TOKENS and machine.isdigit():
            add_relationship("EOAT Inventory", row_number, audit, "tool_to_machine_compatibility", machine,
                             relation_sets["tool_machine"].get((normalized_value(tool), normalized_value(machine))), tool, machine)

    storage_by_eoat = {
        normalized_value(item.get("eoat_identifier")): item
        for item in database["relationships"].get("storage_assignments", []) if item.get("is_current")
    }
    for state in location_audit["records"]:
        eoat = clean_text(state["eoat_identifier"])
        row_number = int(clean_text(state.get("rows")).split(",", 1)[0]) if clean_text(state.get("rows")) else 0
        physical_state = state["determined_physical_state"]
        if physical_state == STATE_INSTALLED:
            machine = clean_text(state["machine_number"])
            add_relationship(
                "EOAT Inventory", row_number, eoat, "current_eoat_installation", machine,
                relation_sets["installations"].get((normalized_value(eoat), normalized_value(machine))), eoat, machine,
                "Physically verified audit evidence establishes an observed current installation; schema cannot safely represent unknown original installed_at.",
            )
        elif physical_state == STATE_STORED:
            add_relationship(
                "EOAT Inventory", row_number, eoat, "current_eoat_storage", clean_text(state["storage_location"]),
                storage_by_eoat.get(normalized_value(eoat)), eoat, clean_text(state["storage_location"]),
                "Explicit audit note establishes cabinet storage, but the cabinet identifier and original stored_at are not recorded.",
            )

    project_root = workbook_path.parents[2]
    for row_number, source in source_rows.get("Photo Index", {}).items():
        source_id = clean_text(source.get("Photo ID")) or f"row-{row_number}"
        photo_matches = indexes["photos"].get(normalized_value(source.get("Photo ID")), [])
        relative = clean_text(source.get("Stored Relative Path"))
        folder = clean_text(source.get("Folder Path"))
        filename = clean_text(source.get("Stored Filename")) or clean_text(source.get("Photo Filename"))
        candidate = project_root / relative if relative else project_root / folder / filename
        if not relative and not (folder and filename):
            document_results.append({
                "sheet": "Photo Index", "source_row": row_number, "source_identifier": source_id, "source_path": "",
                "database_path": "", "document_id": "", "photo_id": "", "linked_entities": "[]",
                "file_exists": "not_checked", "result": "invalid_source_value_preserved",
                "details": "Placeholder photo row is preserved as DEFERRED provenance.",
            })
            continue
        if len(photo_matches) != 1:
            result = "reference_missing_from_database" if not photo_matches else "ambiguous_match"
            document_results.append({
                "sheet": "Photo Index", "source_row": row_number, "source_identifier": source_id,
                "source_path": str(candidate), "database_path": "", "document_id": "", "photo_id": "",
                "linked_entities": "[]", "file_exists": candidate.exists() if include_file_existence_check else "not_checked",
                "result": result, "details": f"Matching photo records: {len(photo_matches)}",
            })
            continue
        photo = photo_matches[0]
        links = document_links.get(int(photo["document_id"]), [])
        expected_links = {
            ("eoat", clean_text(source.get("EOAT Assembly ID"))), ("tool", clean_text(source.get("Tool #"))),
            ("machine", clean_text(source.get("Press/Machine #"))),
        }
        expected_links = {(kind, value) for kind, value in expected_links if value and value.casefold() not in MISSING_TOKENS}
        actual_links = {(item["entity_type"], clean_text(item["entity_identifier"])) for item in links}
        missing_links = expected_links - actual_links
        path_result = comparison(str(candidate), photo.get("storage_path"), header="path")
        if path_result in {"conflicting_database_value", "truncated_database_value"}:
            result = "reference_truncated" if path_result == "truncated_database_value" or clean_text(photo.get("storage_path")) in str(candidate) else "reference_missing_from_database"
        elif missing_links:
            result = "reference_attached_to_wrong_entity"
        elif include_file_existence_check and not candidate.exists():
            result = "reference_present_file_missing"
        else:
            result = "reference_present_file_exists" if include_file_existence_check else "exact_match"
        document_results.append({
            "sheet": "Photo Index", "source_row": row_number, "source_identifier": source_id,
            "source_path": str(candidate), "database_path": photo.get("storage_path") or "",
            "document_id": photo.get("document_id") or "", "photo_id": photo.get("photo_id") or "",
            "linked_entities": json.dumps(sorted(actual_links)),
            "file_exists": candidate.exists() if include_file_existence_check else "not_checked", "result": result,
            "details": f"Missing expected links: {sorted(missing_links)}" if missing_links else "",
        })

    issue_results = []
    for issue in database["issues"]:
        resolution = issue.get("resolution_notes") if isinstance(issue.get("resolution_notes"), dict) else {}
        status = clean_text(resolution.get("status")) or "UNRESOLVED"
        source_field = clean_text(issue.get("field_name"))
        raw = issue.get("raw_values_json") if isinstance(issue.get("raw_values_json"), dict) else {}
        preserved = issue_original_preserved(issue)
        issue_results.append({
            "workbook_sheet": issue.get("source_sheet") or "EOAT Inventory",
            "workbook_row": issue.get("source_row_number") or "",
            "source_field": source_field,
            "source_value": clean_text(issue.get("source_value")),
            "issue_category": issue.get("issue_code"),
            "importer_action": issue.get("description") or "",
            "database_result": issue.get("import_row_status") or "ISSUE_PRESERVED",
            "original_value_preserved": preserved,
            "resolution_status": status,
            "manual_review_required": status in {"UNRESOLVED", "DEFERRED"},
        })

    unmapped = []
    for item in mappings:
        if item["status"] == "unmapped_failure" and int(item["source_row_count"]):
            unmapped.append({
                "sheet": item["workbook_sheet"], "source_header": item["source_header"],
                "populated_count": item["source_row_count"], "examples": item["example_nonempty_values"],
                "reason": item["justification"],
            })

    source_keys = {
        "eoat": {normalized_value(row.get("EOAT Assembly ID")) for row in source_rows.get("EOAT Inventory", {}).values()
                 if clean_text(row.get("EOAT Assembly ID"))},
        "machine": {normalized_value(row.get("Press/Machine #")) for row in source_rows.get("EOAT Inventory", {}).values()
                    if clean_text(row.get("Press/Machine #")).isdigit()},
        "tool": {normalized_value(row.get("Tool #")) for row in source_rows.get("EOAT Inventory", {}).values()
                 if clean_text(row.get("Tool #")) and clean_text(row.get("Tool #")).casefold() not in MISSING_TOKENS},
        "audit": {normalized_value(row.get("Audit ID")) for row in source_rows.get("EOAT Inventory", {}).values()
                  if clean_text(row.get("Audit ID"))},
        "photo": {normalized_value(row.get("Photo ID")) for row in source_rows.get("Photo Index", {}).values()
                  if clean_text(row.get("Photo ID"))},
    }
    database_only = []
    for entity, records, key, pk in (
        ("eoat", database["eoats"], "business_identifier", "id"), ("machine", database["machines"], "machine_number", "id"),
        ("tool", database["tools"], "tool_number", "id"), ("audit", database["audits"], "audit_identifier", "id"),
        ("photo", database["photos"], "source_identifier", "photo_id"),
    ):
        for record in records:
            business_key = clean_text(record.get(key))
            if business_key and normalized_value(business_key) not in source_keys[entity]:
                database_only.append({
                    "entity_type": entity, "primary_key": record.get(pk), "business_key": business_key,
                    "reason": "Valid database record not present in the authoritative workbook snapshot.",
                })
    for record in database.get("database_only_extras", []):
        database_only.append({
            "entity_type": record["entity_type"], "primary_key": record.get("id"),
            "business_key": clean_text(record.get("business_key")),
            "reason": "Additional valid database information from a later or separate source migration.",
        })

    return {
        "source_row_reconciliation.csv": row_results,
        "source_field_reconciliation.csv": field_results,
        "relationship_reconciliation.csv": relationship_results,
        "missing_database_entities.csv": missing_entities,
        "missing_database_relationships.csv": missing_relationships,
        "field_mismatches.csv": mismatches,
        "unmapped_source_fields.csv": unmapped,
        "ambiguous_matches.csv": ambiguous,
        "duplicate_source_keys.csv": duplicate_source_keys(source_rows),
        "duplicate_database_keys.csv": database["duplicate_database_keys"],
        "import_issue_reconciliation.csv": issue_results,
        "document_photo_link_reconciliation.csv": document_results,
        "database_only_records.csv": database_only,
    }


def _csv_write(path: Path, fields: list[str], records: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for record in records:
            writer.writerow({key: json.dumps(value, ensure_ascii=False) if isinstance(value, list | dict) else value
                             for key, value in record.items()})


def repository_info(root: Path) -> dict[str, Any]:
    def git(*args: str) -> str:
        result = subprocess.run(["git", "-C", str(root), *args], capture_output=True, text=True, check=True)
        return result.stdout.strip()
    status = git("status", "--short", "--branch")
    return {
        "path": str(root.resolve()),
        "branch": git("branch", "--show-current"),
        "commit": git("rev-parse", "HEAD"),
        "status": status,
        "dirty": len(status.splitlines()) > 1,
    }


def percentage(numerator: int, denominator: int) -> float:
    return round((100.0 * numerator / denominator), 4) if denominator else 100.0


def summarize(
    inventory: dict[str, Any], mappings: list[dict[str, Any]], database: dict[str, Any], results: dict[str, Any]
) -> tuple[dict[str, Any], str]:
    row_records = results["source_row_reconciliation.csv"]
    field_records = results["source_field_reconciliation.csv"]
    relationship_records = results["relationship_reconciliation.csv"]
    row_failures = sum(item["result"] in FAILURE_RESULTS or item["result"] == "unmapped_source_field" for item in row_records)
    field_failures = sum(item["comparison_result"] in FAILURE_RESULTS for item in field_records)
    relationship_failures = sum(item["result"] == "missing_database_relationship" for item in relationship_records)
    unresolved_issues = sum(item["manual_review_required"] for item in results["import_issue_reconciliation.csv"])
    lost_issues = sum(not item["original_value_preserved"] for item in results["import_issue_reconciliation.csv"])
    orphan_count = sum(int(item["orphan_count"]) for item in database["orphan_checks"])
    truncations = sum(item["classification"] == "truncated_database_value" for item in results["field_mismatches.csv"])
    truncations += sum(item["result"] == "reference_truncated" for item in results["document_photo_link_reconciliation.csv"])
    blockers = {
        "missing_entities": len(results["missing_database_entities.csv"]),
        "missing_relationships": len(results["missing_database_relationships.csv"]),
        "field_mismatches": len(results["field_mismatches.csv"]),
        "unmapped_fields": len(results["unmapped_source_fields.csv"]),
        "ambiguous_matches": len(results["ambiguous_matches.csv"]),
        "duplicate_source_keys": len(results["duplicate_source_keys.csv"]),
        "duplicate_database_keys": len(results["duplicate_database_keys.csv"]),
        "truncations": truncations,
        "unaccounted_import_issues": lost_issues,
        "unresolved_import_issues": unresolved_issues,
        "orphaned_foreign_keys": orphan_count,
    }
    material_failure = any(value for value in blockers.values())
    verdict = "FAIL — SOURCE INFORMATION MISSING OR UNRESOLVED" if material_failure else "PASS — COMPLETE INFORMATION SUPERSET VERIFIED"
    summary = {
        "verdict": verdict,
        "source_rows_total": len(row_records),
        "source_rows_reconciled": len(row_records) - row_failures,
        "source_row_coverage_percent": percentage(len(row_records) - row_failures, len(row_records)),
        "meaningful_fields_total": len(field_records),
        "meaningful_fields_reconciled": len(field_records) - field_failures,
        "source_field_coverage_percent": percentage(len(field_records) - field_failures, len(field_records)),
        "relationships_total": len(relationship_records),
        "relationships_reconciled": len(relationship_records) - relationship_failures,
        "relationship_coverage_percent": percentage(len(relationship_records) - relationship_failures, len(relationship_records)),
        "workbook_sheets": inventory["sheet_count"],
        "meaningful_nonempty_cells": sum(item["meaningful_nonempty_cells"] for item in inventory["sheets"]),
        "mapping_columns": len(mappings),
        "blockers": blockers,
        "database_only_records": len(results["database_only_records.csv"]),
        "external_file_missing": sum(
            item["result"] == "reference_present_file_missing" for item in results["document_photo_link_reconciliation.csv"]
        ),
    }
    return summary, verdict


def executive_summary_text(summary: dict[str, Any], inventory: dict[str, Any], database: dict[str, Any]) -> str:
    blockers = summary["blockers"]
    batch = database.get("batch") or {}
    return f"""# Excel-to-MySQL Information Superset Verification

## Verdict

**{summary['verdict']}**

MySQL has not been proven to contain every meaningful workbook fact because the strict audit found material blockers shown below. The matching import batch does preserve complete source-row JSON for the primary data sheets, but provenance alone does not resolve missing current assignments, unmapped workbook metadata, or unresolved import decisions.

## Coverage

| Measure | Reconciled | Total | Coverage |
|---|---:|---:|---:|
| Source rows | {summary['source_rows_reconciled']} | {summary['source_rows_total']} | {summary['source_row_coverage_percent']:.4f}% |
| Meaningful populated fields | {summary['meaningful_fields_reconciled']} | {summary['meaningful_fields_total']} | {summary['source_field_coverage_percent']:.4f}% |
| Normalized relationships | {summary['relationships_reconciled']} | {summary['relationships_total']} | {summary['relationship_coverage_percent']:.4f}% |

## Required questions

1. **Does MySQL contain at least all meaningful workbook information?** {"Yes" if not any(blockers.values()) else "No; strict parity is not established."}
2. **Were source entities missing?** {blockers['missing_entities']} missing entity findings.
3. **Were source relationships missing?** {blockers['missing_relationships']} missing relationship findings.
4. **Were values truncated or altered?** {blockers['truncations']} truncation findings and {blockers['field_mismatches']} total field mismatch findings.
5. **Were workbook columns unmapped?** {blockers['unmapped_fields']} populated unmapped column findings.
6. **Were later changes preserved with evidence?** Field-specific history was required for current-value conflicts; unexplained conflicts remain failures.
7. **Were import warnings/rejections accounted for?** {len(database['issues'])} database issue records were inspected; {blockers['unaccounted_import_issues']} lack preserved source evidence and {blockers['unresolved_import_issues']} still require review.
8. **Are there blockers?** Yes if any nonzero item appears in the blocker table.

## Blockers

| Category | Count |
|---|---:|
""" + "\n".join(f"| {key.replace('_', ' ').title()} | {value} |" for key, value in blockers.items()) + f"""

## Source and database identity

- Workbook: `{inventory['path']}`
- Workbook SHA-256: `{inventory['sha256']}`
- Sheets inventoried: {inventory['sheet_count']} of {inventory['sheet_count']}
- Matching import batch: `{batch.get('batch_uuid', 'none')}`
- Database: `{database['info']['host']}:{database['info']['port']}/{database['info']['schema']}`
- MySQL / Alembic: `{database['info']['server_version']}` / `{database['info']['alembic_revision']}`
- Database-only records are informational: {summary['database_only_records']}
- Missing external files are reported separately and are not database-loss failures: {summary['external_file_missing']}
"""


def run(args: argparse.Namespace) -> tuple[int, Path, dict[str, Any]]:
    workbook_path = Path(args.workbook)
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Authoritative workbook is unavailable: {workbook_path}")
    repository_root = Path(__file__).resolve().parents[1]
    repo = repository_info(repository_root)
    inventory, source_rows, headers = workbook_inventory(workbook_path)
    mappings = mapping_matrix(source_rows, headers)
    config = load_database_config(args.database_environment)
    database = database_snapshot(config, inventory["sha256"], read_only=args.read_only)
    location_audit = classify_eoat_locations(source_rows.get("EOAT Inventory", {}), database)
    results = reconcile(
        workbook_path, inventory, source_rows, mappings, database, location_audit,
        include_file_existence_check=args.include_file_existence_check,
    )
    summary, verdict = summarize(inventory, mappings, database, results)

    if args.output_directory:
        output = Path(args.output_directory)
    else:
        stamp = datetime.now().astimezone().strftime("%Y%m%d_%H%M%S_%z")
        output = repository_root / "reports" / "mysql_excel_parity" / stamp
    output.mkdir(parents=True, exist_ok=False)

    (output / "workbook_inventory.json").write_text(json.dumps(inventory, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8")
    sheet_rows = []
    for item in inventory["sheets"]:
        sheet_rows.append({key: (len(item[key]) if isinstance(item.get(key), list) else item.get(key, ""))
                           for key in OUTPUT_SCHEMAS["workbook_sheet_summary.csv"]})
    _csv_write(output / "workbook_sheet_summary.csv", OUTPUT_SCHEMAS["workbook_sheet_summary.csv"], sheet_rows)
    _csv_write(output / "field_mapping_matrix.csv", OUTPUT_SCHEMAS["field_mapping_matrix.csv"], mappings)
    for name, schema in OUTPUT_SCHEMAS.items():
        if name in {"workbook_sheet_summary.csv", "field_mapping_matrix.csv"}:
            continue
        _csv_write(output / name, schema, results.get(name, []))
    normalization_rules = {
        "tool_version": TOOL_VERSION,
        "rules": [
            "trim leading/trailing whitespace", "normalize CRLF/CR to LF", "collapse repeated horizontal whitespace",
            "compare case-insensitively for business identifiers", "normalize integer-like Excel floats without losing string leading zeros",
            "normalize Yes/No, TRUE/FALSE, present/not present, and 1/0 only for boolean fields",
            "normalize slash direction and repeated separators for path comparison",
            "never equate blank with false, No, zero, Unknown, or Not Installed",
        ],
    }
    (output / "normalization_rules.json").write_text(json.dumps(normalization_rules, indent=2) + "\n", encoding="utf-8")
    location_fields = [
        "eoat_identifier", "workbook_source", "sheet", "rows", "workbook_location_fields",
        "current_database_state", "determined_physical_state", "machine_number", "storage_location",
        "confidence", "evidence", "required_database_correction", "unresolved_ambiguity",
        "normalized_location_parity",
    ]
    _csv_write(output / "eoat_location_state.csv", location_fields, location_audit["records"])
    (output / "state_aware_location_parity.json").write_text(
        json.dumps(location_audit["metrics"], indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8"
    )
    integrity = {
        "foreign_key_checks": database["orphan_checks"],
        "duplicate_database_keys": database["duplicate_database_keys"],
        "table_counts": database["counts"],
        "status": "PASS" if not any(int(item["orphan_count"]) for item in database["orphan_checks"]) else "FAIL",
    }
    (output / "sql_integrity_results.json").write_text(json.dumps(integrity, indent=2, default=str) + "\n", encoding="utf-8")
    (output / "executive_summary.md").write_text(executive_summary_text(summary, inventory, database), encoding="utf-8")
    log_lines = [
        f"tool_version={TOOL_VERSION}", f"workbook={workbook_path}", f"workbook_sha256={inventory['sha256']}",
        f"database={config.host}:{config.port}/{config.database}", f"read_only={args.read_only}",
        f"matching_import_batch={(database.get('batch') or {}).get('batch_uuid', '')}", f"verdict={verdict}",
        f"totals={json.dumps(summary, sort_keys=True)}",
    ]
    (output / "verification.log").write_text("\n".join(log_lines) + "\n", encoding="utf-8")

    release_file = repository_root / "release_metadata.json"
    release = json.loads(release_file.read_text(encoding="utf-8")) if release_file.is_file() else {}
    hashes = {
        path.name: sha256_file(path) for path in sorted(output.iterdir())
        if path.is_file() and path.name != "verification_manifest.json"
    }
    manifest = {
        "run_id": output.name,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "repository": repo,
        "application_version": release.get("app_version"),
        "api_version": "1.3.0",
        "workbook_path": str(workbook_path),
        "workbook_hash": inventory["sha256"],
        "database_host": config.host,
        "database_port": config.port,
        "database_schema": config.database,
        "database_account_category": config.account_category,
        "alembic_revision": database["info"]["alembic_revision"],
        "tool_version": TOOL_VERSION,
        "verification_status": verdict,
        "totals": summary,
        "output_file_hashes": hashes,
        "exact_command": subprocess.list2cmdline([sys.executable, *sys.argv]),
        "credentials_redacted": True,
    }
    (output / "verification_manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8")
    exit_code = 1 if verdict.startswith("FAIL") else 0
    if args.fail_on_warning and summary["external_file_missing"]:
        exit_code = 2
    if args.json:
        print(json.dumps({"output_directory": str(output), "exit_code": exit_code, **summary}, indent=2))
    else:
        print(verdict)
        print(f"Output: {output}")
    return exit_code, output, summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Prove Excel Master Tracker information coverage in EOAT Atlas MySQL.")
    parser.add_argument("--workbook", required=True, help="Authoritative EOAT_Master_Tracker.xlsx path")
    parser.add_argument("--database-environment", default="development", help="database.env path or 'development'")
    parser.add_argument("--output-directory", help="New directory for the complete verification evidence set")
    parser.add_argument("--strict", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--read-only", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--include-file-existence-check", action="store_true")
    parser.add_argument("--fail-on-warning", action="store_true")
    parser.add_argument("--json", action="store_true", help="Print machine-readable result summary")
    parser.add_argument("--verbose", action="store_true")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    try:
        exit_code, _, _ = run(args)
        return exit_code
    except Exception as exc:
        if getattr(args, "verbose", False):
            raise
        print(f"BLOCKED — VERIFICATION COULD NOT BE COMPLETED: {exc}", file=sys.stderr)
        return 3


if __name__ == "__main__":
    raise SystemExit(main())
