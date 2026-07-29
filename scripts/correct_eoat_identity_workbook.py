"""Safely materialize the governed physical-identity mapping in a tracker."""

# ruff: noqa: E402

from __future__ import annotations

import argparse
import hashlib
import os
import sys
import tempfile
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

REPOSITORY_ROOT = Path(__file__).resolve().parents[1]
if str(REPOSITORY_ROOT) not in sys.path:
    sys.path.insert(0, str(REPOSITORY_ROOT))

from scripts.generate_eoat_identity_crosswalk import read_inventory_rows
from tools.eoat_location_normalization import (
    identity_resolution,
    load_policy,
    physical_eoat_identifier,
    physical_eoat_uuid,
)
from tools.eoat_physical_identity import build_crosswalk

WORKSHEET = "EOAT Inventory"
IDENTITY_HEADERS = (
    "Original EOAT Assembly ID",
    "Canonical Physical EOAT ID",
    "Physical EOAT UUID",
    "Identity Resolution",
    "Owner Decision Reference",
)


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _header_row(sheet) -> int:
    return next(
        number for number, values in enumerate(sheet.iter_rows(values_only=True), 1)
        if "EOAT Assembly ID" in values and "Audit Date" in values
    )


def correct_workbook(workbook_path: Path, *, expected_sha256: str, backup_path: Path) -> dict[str, object]:
    """Apply only the approved identity columns and save atomically.

    The pre-change hash and an independent byte-for-byte backup are mandatory;
    this function never fabricates an event, time, user, location, or asset.
    """
    if _sha256(workbook_path) != expected_sha256:
        raise RuntimeError("Workbook SHA-256 changed before correction; refusing to overwrite it")
    if not backup_path.is_file() or _sha256(backup_path) != expected_sha256:
        raise RuntimeError("Verified byte-for-byte workbook backup is required before correction")
    source_crosswalk = build_crosswalk(read_inventory_rows(workbook_path), source_workbook_sha256=expected_sha256)
    policy = load_policy()
    owner_reference = str(policy["identity_correction"]["owner_decision_reference"])
    keep_vba = workbook_path.suffix.casefold() == ".xlsm"
    workbook = load_workbook(workbook_path, read_only=False, data_only=False, keep_vba=keep_vba)
    try:
        sheet = workbook[WORKSHEET]
        header_row = _header_row(sheet)
        headers = {_text(cell.value): cell.column for cell in sheet[header_row] if _text(cell.value)}
        source_column = headers["EOAT Assembly ID"]
        template_column = sheet.cell(header_row, source_column)
        added = []
        for header in IDENTITY_HEADERS:
            if header in headers:
                continue
            column = sheet.max_column + 1
            cell = sheet.cell(header_row, column, header)
            cell._style = copy(template_column._style)
            cell.font = copy(template_column.font)
            cell.fill = copy(template_column.fill)
            cell.border = copy(template_column.border)
            cell.alignment = copy(template_column.alignment)
            cell.number_format = template_column.number_format
            cell.protection = copy(template_column.protection)
            sheet.column_dimensions[get_column_letter(column)].width = sheet.column_dimensions[get_column_letter(source_column)].width
            headers[header] = column
            added.append(header)
        for number in range(header_row + 1, sheet.max_row + 1):
            entry_type = _text(sheet.cell(number, headers["Entry Type"]).value)
            source_identifier = _text(sheet.cell(number, source_column).value)
            if not entry_type and not source_identifier:
                continue
            source_cell = sheet.cell(number, source_column)
            for header in IDENTITY_HEADERS:
                target_cell = sheet.cell(number, headers[header])
                if header in added:
                    target_cell._style = copy(source_cell._style)
                    target_cell.font = copy(source_cell.font)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.border = copy(source_cell.border)
                    target_cell.alignment = copy(source_cell.alignment)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = copy(source_cell.protection)
            canonical = physical_eoat_identifier(source_identifier, number, entry_type)
            if canonical:
                sheet.cell(number, headers["Original EOAT Assembly ID"], source_identifier)
                sheet.cell(number, headers["Canonical Physical EOAT ID"], canonical)
                sheet.cell(number, headers["Physical EOAT UUID"], physical_eoat_uuid(canonical))
                sheet.cell(number, headers["Identity Resolution"], identity_resolution(source_identifier, number, entry_type))
                sheet.cell(number, headers["Owner Decision Reference"], owner_reference)
                # The legacy workbook's visible identifier is now explicitly the
                # canonical physical identifier for audited rows.  The original
                # value above remains immutable provenance.
                sheet.cell(number, source_column, canonical)
            else:
                sheet.cell(number, headers["Original EOAT Assembly ID"], source_identifier or None)
                sheet.cell(number, headers["Canonical Physical EOAT ID"], None)
                sheet.cell(number, headers["Physical EOAT UUID"], None)
                sheet.cell(number, headers["Identity Resolution"], identity_resolution(source_identifier, number, entry_type))
                sheet.cell(number, headers["Owner Decision Reference"], owner_reference)
        for table in sheet.tables.values():
            start, end = table.ref.split(":")
            end_column = get_column_letter(sheet.max_column)
            end_row = "".join(filter(str.isdigit, end))
            table.ref = f"{start}:{end_column}{end_row}"
        with tempfile.NamedTemporaryFile(prefix=workbook_path.stem + ".", suffix=workbook_path.suffix, dir=workbook_path.parent, delete=False) as handle:
            temp_path = Path(handle.name)
        workbook.save(temp_path)
    finally:
        workbook.close()
    try:
        # Reopen prior to replacement so corruption cannot replace the source.
        reopened = load_workbook(temp_path, read_only=True, data_only=False, keep_vba=keep_vba)
        try:
            if WORKSHEET not in reopened.sheetnames:
                raise RuntimeError("Corrected workbook is missing EOAT Inventory")
            corrected_rows = read_inventory_rows(temp_path)
            audited = [row for _, row in corrected_rows if _text(row.get("Entry Type")).casefold() == "audited"]
            if len(audited) != 67:
                raise RuntimeError(f"Corrected workbook audit count is {len(audited)}, expected 67")
            canonical = [_text(row.get("Canonical Physical EOAT ID")) for row in audited]
            if len(set(canonical)) != 66 or any(not value for value in canonical):
                raise RuntimeError("Corrected workbook does not contain 66 nonblank canonical physical identifiers")
        finally:
            reopened.close()
        os.replace(temp_path, workbook_path)
    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
    return {
        "original_sha256": expected_sha256,
        "backup_sha256": _sha256(backup_path),
        "corrected_sha256": _sha256(workbook_path),
        "added_identity_headers": added,
        "source_crosswalk_rows": len(source_crosswalk),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--expected-sha256", required=True)
    parser.add_argument("--backup", required=True, type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    if not args.apply:
        raise SystemExit("Refusing to modify a workbook without --apply")
    result = correct_workbook(args.workbook.resolve(), expected_sha256=args.expected_sha256, backup_path=args.backup.resolve())
    print(result)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
