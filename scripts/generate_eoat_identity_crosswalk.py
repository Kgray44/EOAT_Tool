"""Generate the owner-governed EOAT physical-identity crosswalk artifacts."""

# ruff: noqa: E402

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

REPOSITORY_ROOT = Path(__file__).resolve().parents[1]
if str(REPOSITORY_ROOT) not in sys.path:
    sys.path.insert(0, str(REPOSITORY_ROOT))

from tools.eoat_physical_identity import build_crosswalk, validate_crosswalk

WORKSHEET = "EOAT Inventory"


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def read_inventory_rows(workbook_path: Path) -> list[tuple[int, dict[str, object]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        sheet = workbook[WORKSHEET]
        header_number = next(
            number for number, values in enumerate(sheet.iter_rows(values_only=True), 1)
            if "EOAT Assembly ID" in values and "Audit Date" in values
        )
        rows = list(sheet.iter_rows(min_row=header_number, values_only=True))
        headers = [_text(value) for value in rows[0]]
        return [
            (
                header_number + offset,
                {headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]},
            )
            for offset, values in enumerate(rows[1:], 1)
            if any(value is not None and _text(value) for value in values)
        ]
    finally:
        workbook.close()


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def write_artifacts(workbook: Path, output_dir: Path) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=False)
    source_hash = _sha256(workbook)
    crosswalk = build_crosswalk(read_inventory_rows(workbook), source_workbook_sha256=source_hash)
    validation = validate_crosswalk(crosswalk)
    json_path = output_dir / "eoat-physical-identity-crosswalk.json"
    csv_path = output_dir / "eoat-physical-identity-crosswalk.csv"
    report_path = output_dir / "eoat-physical-identity-crosswalk-review.md"
    json_path.write_text(json.dumps(crosswalk, indent=2, sort_keys=True, default=str) + "\n", encoding="utf-8")
    fieldnames = list(crosswalk[0])
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in crosswalk:
            writer.writerow({key: json.dumps(value, sort_keys=True, default=str) if isinstance(value, dict | list) else value for key, value in row.items()})
    report_path.write_text(
        "# EOAT physical-identity crosswalk review\n\n"
        f"- Source workbook SHA-256: `{source_hash}`\n"
        f"- Audited rows: {validation.audited_rows}\n"
        f"- Physical EOAT UUIDs: {validation.physical_units}\n"
        f"- Canonical identifiers: {validation.canonical_identifiers}\n"
        f"- Only shared identity: {validation.duplicate_physical_identifier} ({validation.duplicate_audit_rows} audit rows)\n\n"
        "This report is private correction evidence. The JSON and CSV retain the full source-preserving row mapping.\n",
        encoding="utf-8",
    )
    artifacts = {path.name: _sha256(path) for path in (json_path, csv_path, report_path)}
    (output_dir / "SHA256SUMS.json").write_text(json.dumps(artifacts, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return {**artifacts, "source_workbook_sha256": source_hash}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    arguments = parser.parse_args()
    result = write_artifacts(arguments.workbook.resolve(), arguments.output_dir.resolve())
    print(json.dumps(result, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
