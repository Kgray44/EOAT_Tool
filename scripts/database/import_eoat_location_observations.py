"""Import workbook evidence as observations, never as fabricated movements."""

# ruff: noqa: E402, I001

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from datetime import date, datetime
from pathlib import Path
from uuid import NAMESPACE_URL, uuid5

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[2]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from scripts.database.production_data_migration import REQUIRED_REVISION, connect, database_identity, read_env  # noqa: E402
from tools.eoat_location_state import (  # noqa: E402
    STATE_CONFLICT, STATE_INACTIVE, STATE_INSTALLED, STATE_STORED, classify_eoat_locations,
)
from tools.eoat_location_normalization import (  # noqa: E402
    is_stored_machine_reference,
    load_policy,
    normalize_machine_reference,
    normalized_source_rows,
)

TOOL_VERSION = "1.0.0"
WORKSHEET = "EOAT Inventory"


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(_text(value)[:10])
    except ValueError:
        return None


def read_rows(path: Path) -> dict[int, dict]:
    book = load_workbook(path, read_only=True, data_only=True)
    sheet = book[WORKSHEET]
    header_row = next(
        number for number, values in enumerate(sheet.iter_rows(values_only=True), 1)
        if "EOAT Assembly ID" in values and "Audit Date" in values
    )
    values = list(sheet.iter_rows(min_row=header_row, values_only=True))
    headers = [_text(value) for value in values[0]]
    return {
        header_row + offset: {headers[index]: value for index, value in enumerate(row) if headers[index]}
        for offset, row in enumerate(values[1:], 1)
        if any(value is not None and _text(value) for value in row)
    }


def _assertion_state(row: dict) -> str:
    if is_stored_machine_reference(row.get("Press/Machine #")):
        return "STORED"
    notes = _text(row.get("Notes")).casefold()
    if "cabinet" in notes:
        return "STORED"
    if "not installed" in notes or "removed from" in notes:
        return "UNKNOWN"
    if (_text(row.get("Physical Audit Verified")).casefold() == "yes"
            and _text(row.get("Audit Context")).casefold() == "installed on machine"
            and normalize_machine_reference(row.get("Press/Machine #"))):
        return "INSTALLED"
    return "UNKNOWN"


def _normalized_state(label: str) -> str:
    return {
        STATE_INSTALLED: "INSTALLED", STATE_STORED: "STORED", STATE_INACTIVE: "INACTIVE",
        STATE_CONFLICT: "CONFLICTING",
    }.get(label, "UNKNOWN")


def _wording(row: dict) -> str:
    fields = ("Audit ID", "Audit Date", "Entry Type", "Physical Audit Verified", "Audit Context", "Press/Machine #", "Notes")
    return "; ".join(f"{field}={_text(row.get(field)) or 'blank'}" for field in fields)


def build_plan(
    rows: dict[int, dict], database: dict, workbook: Path | None = None, *,
    workbook_sha256: str | None = None, source_workbook: str | None = None,
) -> dict:
    if workbook_sha256:
        digest = workbook_sha256
    elif workbook is not None:
        digest = hashlib.sha256(workbook.read_bytes()).hexdigest()
    else:
        raise ValueError("A workbook path or workbook_sha256 is required")
    workbook_name = source_workbook or (workbook.name if workbook is not None else "EOAT_Master_Tracker.xlsx")
    classified = classify_eoat_locations(rows, database)
    by_eoat: dict[str, list[tuple[int, dict]]] = {}
    for number, row in rows.items():
        if _text(row.get("Entry Type")).casefold() != "audited" or not _text(row.get("EOAT Assembly ID")):
            continue
        by_eoat.setdefault(_text(row.get("EOAT Assembly ID")), []).append((number, row))
    observations, assertions = [], []
    for record in classified["records"]:
        eoat = record["eoat_identifier"]
        source = by_eoat[eoat]
        observed_on = max(filter(None, (_date(row.get("Audit Date")) for _, row in source)), default=None)
        if observed_on is None:
            raise RuntimeError(f"Audited EOAT {eoat} has no honest observation date")
        latest = [(number, row) for number, row in source if _date(row.get("Audit Date")) == observed_on]
        source_row = min(number for number, _ in latest)
        state = _normalized_state(record["determined_physical_state"])
        observation_uuid = str(uuid5(NAMESPACE_URL, f"eoat-atlas:{digest}:{eoat}:current-state"))
        conflict_uuid = str(uuid5(NAMESPACE_URL, f"eoat-atlas:{digest}:{eoat}:conflict")) if state == "CONFLICTING" else None
        observations.append({
            "observation_uuid": observation_uuid, "eoat_identifier": eoat, "state": state,
            "machine_number": record["machine_number"] or None, "storage_location": None,
            "observed_on": observed_on.isoformat(), "source_row_number": source_row,
            "source_workbook": workbook_name, "source_worksheet": WORKSHEET,
            "original_source_wording": record["workbook_location_fields"], "confidence": record["confidence"],
            "resolution_status": "REVIEW_REQUIRED" if state == "CONFLICTING" else "CURRENT",
            "conflict_group_uuid": conflict_uuid,
        })
        for number, row in sorted(source):
            assertion_state = _assertion_state(row)
            assertions.append({
                "assertion_uuid": str(uuid5(NAMESPACE_URL, f"eoat-atlas:{digest}:{eoat}:{WORKSHEET}:{number}")),
                "observation_uuid": observation_uuid, "eoat_identifier": eoat, "state": assertion_state,
                "machine_number": normalize_machine_reference(row.get("Press/Machine #")) if assertion_state == "INSTALLED" else None,
                "observed_on": _date(row.get("Audit Date")).isoformat(), "source_row_number": number,
                "source_workbook": workbook_name, "source_worksheet": WORKSHEET,
                "original_source_wording": _wording(row), "confidence": "SOURCE_ASSERTION",
                "participates_in_conflict": state == "CONFLICTING",
            })
    return {
        "tool_version": TOOL_VERSION, "workbook": workbook_name, "workbook_sha256": digest,
        "required_schema_revision": REQUIRED_REVISION, "observations": observations, "assertions": assertions,
        "state_counts": {state: sum(row["state"] == state for row in observations) for state in ("INSTALLED", "STORED", "UNKNOWN", "INACTIVE", "CONFLICTING")},
    }


def load_database(connection) -> dict:
    with connection.cursor() as cursor:
        cursor.execute("SELECT id, business_identifier, is_active, archived_at FROM eoats ORDER BY id")
        eoats = cursor.fetchall()
        cursor.execute("SELECT e.business_identifier eoat_identifier, m.machine_number, i.removed_at IS NULL is_current FROM eoat_installations i JOIN eoats e ON e.id=i.eoat_id JOIN machines m ON m.id=i.machine_id")
        installations = cursor.fetchall()
        cursor.execute("SELECT e.business_identifier eoat_identifier, s.location_code, a.removed_from_storage_at IS NULL is_current FROM eoat_storage_assignments a JOIN eoats e ON e.id=a.eoat_id JOIN storage_locations s ON s.id=a.storage_location_id")
        storage = cursor.fetchall()
    return {"eoats": eoats, "relationships": {"installations": installations, "storage_assignments": storage}}


def apply_plan(connection, plan: dict, *, commit: bool = True) -> None:
    material_changes = 0
    with connection.cursor() as cursor:
        cursor.execute("SELECT id,business_identifier FROM eoats")
        eoats = {row["business_identifier"]: row["id"] for row in cursor.fetchall()}
        cursor.execute("SELECT id,machine_number FROM machines")
        machines = {row["machine_number"]: row["id"] for row in cursor.fetchall()}
        cursor.execute("SELECT eoat_id,source_row_number,id,source_import_batch_id FROM audit_records")
        audits = {(row["eoat_id"], row["source_row_number"]): row for row in cursor.fetchall()}
        cursor.execute("SELECT source_row_number,id,import_batch_id FROM import_rows WHERE source_sheet=%s ORDER BY id", (WORKSHEET,))
        import_rows = {row["source_row_number"]: row for row in cursor.fetchall()}
        for row in plan["observations"]:
            eoat_id = eoats[row["eoat_identifier"]]
            audit = audits.get((eoat_id, row["source_row_number"]), {})
            imported = import_rows.get(row["source_row_number"], {})
            cursor.execute(
                "INSERT IGNORE INTO eoat_location_observations (observation_uuid,eoat_id,state,machine_id,storage_location_id,observed_at,observed_on,observation_precision,source_type,source_audit_record_id,source_import_row_id,source_import_batch_id,source_workbook,source_worksheet,source_row_number,original_source_wording,confidence,resolution_status,conflict_group_uuid,is_authoritative) VALUES (%s,%s,%s,%s,NULL,NULL,%s,'DATE','MASTER_TRACKER_AUDIT',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,1)",
                (row["observation_uuid"], eoat_id, row["state"], machines.get(row["machine_number"]), row["observed_on"], audit.get("id"), imported.get("id"), audit.get("source_import_batch_id") or imported.get("import_batch_id"), row["source_workbook"], row["source_worksheet"], row["source_row_number"], row["original_source_wording"], row["confidence"], row["resolution_status"], row["conflict_group_uuid"]),
            )
            material_changes += max(0, cursor.rowcount)
        cursor.execute("SELECT id,observation_uuid FROM eoat_location_observations")
        observation_ids = {row["observation_uuid"]: row["id"] for row in cursor.fetchall()}
        for row in plan["assertions"]:
            eoat_id = eoats[row["eoat_identifier"]]
            audit = audits.get((eoat_id, row["source_row_number"]), {})
            imported = import_rows.get(row["source_row_number"], {})
            cursor.execute(
                "INSERT IGNORE INTO eoat_location_assertions (assertion_uuid,observation_id,eoat_id,state,machine_id,storage_location_id,observed_at,observed_on,observation_precision,source_type,source_audit_record_id,source_import_row_id,source_import_batch_id,source_workbook,source_worksheet,source_row_number,original_source_wording,confidence,participates_in_conflict) VALUES (%s,%s,%s,%s,%s,NULL,NULL,%s,'DATE','MASTER_TRACKER_AUDIT',%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (row["assertion_uuid"], observation_ids[row["observation_uuid"]], eoat_id, row["state"], machines.get(row["machine_number"]), row["observed_on"], audit.get("id"), imported.get("id"), audit.get("source_import_batch_id") or imported.get("import_batch_id"), row["source_workbook"], row["source_worksheet"], row["source_row_number"], row["original_source_wording"], row["confidence"], row["participates_in_conflict"]),
            )
            material_changes += max(0, cursor.rowcount)
        cursor.execute("SELECT COUNT(*) count FROM eoat_location_observations")
        if cursor.fetchone()["count"] != len(plan["observations"]):
            raise RuntimeError("Observation count mismatch; refusing partial or mixed import")
        cursor.execute("SELECT COUNT(*) count FROM eoat_location_assertions")
        if cursor.fetchone()["count"] != len(plan["assertions"]):
            raise RuntimeError("Assertion count mismatch; refusing partial or mixed import")
        import_source = f"{plan['workbook']}:{plan['workbook_sha256']}"
        if material_changes:
            cursor.execute(
                "UPDATE data_state SET current_revision=current_revision+1,data_last_modified_at=UTC_TIMESTAMP(6),"
                "last_import_at=UTC_TIMESTAMP(6),last_import_source=%s,updated_by='location-observation-import' WHERE id=1",
                (import_source,),
            )
        else:
            # A successfully deduplicated import is still useful provenance,
            # but it must never fabricate a data change or revision advance.
            cursor.execute(
                "UPDATE data_state SET last_import_at=UTC_TIMESTAMP(6),last_import_source=%s,"
                "updated_by='location-observation-import' WHERE id=1",
                (import_source,),
            )
    if commit:
        connection.commit()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--env", default="development")
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    values = read_env(args.env)
    connection = connect(values)
    try:
        identity = database_identity(connection)
        if identity["alembic_revision"] != REQUIRED_REVISION:
            raise RuntimeError(f"Schema mismatch: required {REQUIRED_REVISION}; found {identity['alembic_revision']}")
        if str(identity["database_name"]).casefold().endswith("_prod"):
            raise RuntimeError("Production database is forbidden")
        plan = build_plan(normalized_source_rows(read_rows(args.workbook)), load_database(connection), args.workbook)
        if plan["state_counts"] != load_policy()["expected_location_state_counts"]:
            raise RuntimeError(f"Unexpected location classification: {plan['state_counts']}")
        if args.apply:
            apply_plan(connection, plan)
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(json.dumps(plan, indent=2, sort_keys=True), encoding="utf-8")
        print(json.dumps({"status": "applied" if args.apply else "planned", "observations": len(plan["observations"]), "assertions": len(plan["assertions"]), "state_counts": plan["state_counts"]}, sort_keys=True))
        return 0
    finally:
        connection.close()


if __name__ == "__main__":
    raise SystemExit(main())
