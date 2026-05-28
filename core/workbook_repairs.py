from __future__ import annotations

import time
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from . import audit_field_rules as field_rules
from .audit.history import append_audit_history
from .audit_by_press import AUDIT_BY_PRESS_SHEET, audit_by_press_last_refreshed, refresh_audit_by_press_view, refresh_audit_by_press_view_action
from .audit_entries import (
    LEGACY_VACUUM_CUPS_FIELD,
    NA_VALUE,
    apply_part_present_sensor_defaults,
    audit_field_applies,
    repair_workbook_schema,
)
from .logging import log_tool_run
from .paths import resolve_project_paths
from .result import ToolResult
from .safe_files import backup_file
from .tool_fields import LEGACY_TOOL_FIELD, TOOL_FIELD
from .validation import validate_project_foundation
from .validation_findings import summarize_findings, findings_from_result
from .workbook_io import worksheet_headers
from .workbook_cache import invalidate_workbook_cache
from .workbook_locks import detect_workbook_lock
from .workbook_schema import get_expected_headers

FIX_CLEAR_STALE_HIDDEN_NA = "clear_stale_hidden_na"
FIX_REPAIR_LEGACY_HEADERS = "repair_legacy_headers"
FIX_REFRESH_GENERATED_VIEWS = "refresh_generated_views"
FIX_REAPPLY_FORMATTING = "reapply_formatting"
FIX_REBUILD_DROPDOWN_VALIDATION = "rebuild_dropdown_validation"

SAFE_FIX_IDS = {
    FIX_CLEAR_STALE_HIDDEN_NA,
    FIX_REPAIR_LEGACY_HEADERS,
    FIX_REFRESH_GENERATED_VIEWS,
    FIX_REAPPLY_FORMATTING,
    FIX_REBUILD_DROPDOWN_VALIDATION,
}


@dataclass(frozen=True)
class RepairChange:
    sheet_name: str
    row_number: int | None = None
    column_name: str = ""
    audit_id: str = ""
    machine_number: str = ""
    current_value: str = ""
    new_value: str = ""
    reason: str = ""

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class SafeFixPreview:
    fix_id: str
    title: str
    description: str
    changes: list[RepairChange] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    can_apply: bool = False

    def to_markdown(self) -> str:
        lines = [f"# {self.title}", "", self.description, "", f"**Fix ID:** {self.fix_id}"]
        if self.warnings:
            lines.extend(["", "## Warnings", *[f"- {warning}" for warning in self.warnings]])
        if not self.changes:
            lines.extend(["", "No changes are currently previewed."])
            return "\n".join(lines) + "\n"
        lines.extend(["", "## Previewed Changes"])
        for change in self.changes[:200]:
            target = f"{change.sheet_name}"
            if change.row_number:
                target += f" row {change.row_number}"
            if change.column_name:
                target += f" / {change.column_name}"
            lines.append(f"- {target}: {change.current_value or '(blank)'} -> {change.new_value or '(blank)'}")
            if change.audit_id or change.machine_number or change.reason:
                detail = " | ".join(part for part in [change.audit_id, change.machine_number, change.reason] if part)
                lines.append(f"  {detail}")
        if len(self.changes) > 200:
            lines.append(f"- ... {len(self.changes) - 200} more change(s) omitted from preview text.")
        return "\n".join(lines) + "\n"


def preview_safe_fix(project_root: str | Path, fix_id: str) -> SafeFixPreview:
    fix_id = str(fix_id or "").strip()
    if fix_id == FIX_CLEAR_STALE_HIDDEN_NA:
        return _preview_clear_stale_hidden_na(project_root)
    if fix_id == FIX_REPAIR_LEGACY_HEADERS:
        return _preview_repair_legacy_headers(project_root)
    if fix_id == FIX_REFRESH_GENERATED_VIEWS:
        return _preview_refresh_generated_views(project_root)
    if fix_id == FIX_REAPPLY_FORMATTING:
        return _generic_preview(
            fix_id,
            "Reapply Formatting",
            "Reapply established workbook formatting and inventory ranges without changing engineering values.",
            "Workbook formatting and table-like ranges will be refreshed.",
            project_root,
        )
    if fix_id == FIX_REBUILD_DROPDOWN_VALIDATION:
        return _generic_preview(
            fix_id,
            "Rebuild Dropdown Validation",
            "Rebuild known workbook dropdown and whole-number validation rules without changing cell values.",
            "Dropdown and whole-number validation rules will be rebuilt.",
            project_root,
        )
    return SafeFixPreview(
        fix_id=fix_id,
        title="Unknown Safe Fix",
        description="The requested safe fix is not registered.",
        warnings=[f"Unknown safe fix: {fix_id}"],
        can_apply=False,
    )


def preview_safe_fix_action(project_root: str | Path, fix_id: str) -> ToolResult:
    preview = preview_safe_fix(project_root, fix_id)
    return ToolResult.ok(
        "workbook_repair_preview",
        "Workbook Repair Preview",
        f"Previewed safe fix {preview.fix_id}.",
        details=preview.to_markdown().splitlines(),
        warnings=list(preview.warnings),
        metrics={"fix_id": preview.fix_id, "preview_change_count": len(preview.changes), "can_apply": preview.can_apply},
        structured_data={"safe_fix_preview": _preview_dict(preview)},
    )


def apply_safe_fix(
    project_root: str | Path,
    fix_id: str,
    *,
    confirm: bool = False,
    log_activity: bool = True,
) -> ToolResult:
    started = time.perf_counter()
    preview = preview_safe_fix(project_root, fix_id)
    if not confirm:
        return ToolResult.fail(
            "workbook_repair",
            "Workbook Repair",
            "Safe fix confirmation is required before applying workbook changes.",
            errors=["Preview the fix and rerun with confirm=True to apply it."],
            details=preview.to_markdown().splitlines(),
            metrics={"fix_id": preview.fix_id, "preview_change_count": len(preview.changes), "can_apply": preview.can_apply},
            structured_data={"safe_fix_preview": _preview_dict(preview)},
            duration_seconds=time.perf_counter() - started,
        )
    if not preview.can_apply:
        return ToolResult.ok(
            "workbook_repair",
            "Workbook Repair",
            f"No safe changes are available for {preview.fix_id}.",
            details=preview.to_markdown().splitlines(),
            warnings=list(preview.warnings),
            metrics={"fix_id": preview.fix_id, "applied_change_count": 0},
            structured_data={"safe_fix_preview": _preview_dict(preview)},
            duration_seconds=time.perf_counter() - started,
        )

    paths = resolve_project_paths(project_root)
    lock_status = detect_workbook_lock(paths.master_workbook)
    if not lock_status.can_write:
        return ToolResult.fail(
            "workbook_repair",
            "Workbook Repair",
            "Workbook repair was blocked by the workbook lock detector.",
            errors=[lock_status.message],
            warnings=[lock_status.error] if lock_status.error else [],
            metrics={"fix_id": preview.fix_id, "workbook_locked": lock_status.locked},
            structured_data={"workbook_lock": asdict(lock_status), "safe_fix_preview": _preview_dict(preview)},
            duration_seconds=time.perf_counter() - started,
        )

    if preview.fix_id == FIX_CLEAR_STALE_HIDDEN_NA:
        result = _apply_clear_stale_hidden_na(project_root, preview, started)
    elif preview.fix_id == FIX_REFRESH_GENERATED_VIEWS:
        result = refresh_audit_by_press_view_action(project_root, log_activity=False)
        result.tool_id = "workbook_repair"
        result.tool_name = "Workbook Repair"
        result.summary = "Applied safe fix: refresh generated views."
        result.metrics["fix_id"] = preview.fix_id
        result.metrics["applied_change_count"] = len(preview.changes)
        _append_workbook_repair_history(project_root, preview, result.files_modified)
    else:
        result = repair_workbook_schema(project_root, log_activity=False)
        result.tool_id = "workbook_repair"
        result.tool_name = "Workbook Repair"
        result.summary = f"Applied safe fix: {preview.title}."
        result.metrics["fix_id"] = preview.fix_id
        result.metrics["applied_change_count"] = len(preview.changes)
        _append_workbook_repair_history(project_root, preview, result.files_modified)

    validation = validate_project_foundation(project_root)
    summary_counts = summarize_findings(findings_from_result(validation))
    result.details.append(f"Validation rerun after repair: {validation.summary}")
    result.metrics["validation_after_fix_success"] = validation.success
    result.metrics["validation_after_fix_finding_count"] = summary_counts["total"]
    result.metrics["validation_after_fix_by_severity"] = summary_counts["by_severity"]
    result.structured_data["safe_fix_preview"] = _preview_dict(preview)
    result.structured_data["validation_after_fix"] = {
        "success": validation.success,
        "summary": validation.summary,
        "summary_counts": summary_counts,
    }
    result.duration_seconds = time.perf_counter() - started
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def _preview_clear_stale_hidden_na(project_root: str | Path) -> SafeFixPreview:
    paths = resolve_project_paths(project_root)
    changes: list[RepairChange] = []
    warnings: list[str] = []
    if not paths.master_workbook.exists():
        return SafeFixPreview(
            fix_id=FIX_CLEAR_STALE_HIDDEN_NA,
            title="Clear Stale Hidden Values",
            description="Set stale values in non-applicable hidden fields to N/A.",
            warnings=[f"Master workbook is missing: {paths.master_workbook}"],
            can_apply=False,
        )
    workbook = None
    try:
        workbook = load_workbook(paths.master_workbook, read_only=True, data_only=False)
        if "EOAT Inventory" not in workbook.sheetnames:
            warnings.append("EOAT Inventory sheet is missing.")
            return SafeFixPreview(
                fix_id=FIX_CLEAR_STALE_HIDDEN_NA,
                title="Clear Stale Hidden Values",
                description="Set stale values in non-applicable hidden fields to N/A.",
                warnings=warnings,
                can_apply=False,
            )
        ws = workbook["EOAT Inventory"]
        headers = worksheet_headers(ws)
        positions = {header: index for index, header in enumerate(headers)}
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_data = {header: row[index] for header, index in positions.items() if index < len(row)}
            if not _is_audit_data_row(row_data):
                continue
            row_data = apply_part_present_sensor_defaults(row_data)
            for header in get_expected_headers("EOAT Inventory"):
                if header not in positions:
                    continue
                value = row_data.get(header)
                if audit_field_applies(row_data, header) or not field_rules.is_meaningful_value(value):
                    continue
                changes.append(
                    RepairChange(
                        sheet_name="EOAT Inventory",
                        row_number=row_number,
                        column_name=header,
                        audit_id=_text(row_data.get("Audit ID")),
                        machine_number=_text(row_data.get("Press/Machine #")),
                        current_value=_text(value),
                        new_value=NA_VALUE,
                        reason=field_rules.non_applicable_reason(row_data, header),
                    )
                )
    except Exception as exc:
        warnings.append(f"Could not preview stale hidden values: {exc}")
    finally:
        if workbook is not None:
            workbook.close()
    return SafeFixPreview(
        fix_id=FIX_CLEAR_STALE_HIDDEN_NA,
        title="Clear Stale Hidden Values",
        description="Set stale values in non-applicable hidden fields to N/A. This never guesses engineering values.",
        changes=changes,
        warnings=warnings,
        can_apply=bool(changes) and not warnings,
    )


def _preview_repair_legacy_headers(project_root: str | Path) -> SafeFixPreview:
    paths = resolve_project_paths(project_root)
    changes: list[RepairChange] = []
    warnings: list[str] = []
    if not paths.master_workbook.exists():
        warnings.append(f"Master workbook is missing: {paths.master_workbook}")
    else:
        workbook = None
        try:
            workbook = load_workbook(paths.master_workbook, read_only=True, data_only=False)
            if "EOAT Inventory" not in workbook.sheetnames:
                warnings.append("EOAT Inventory sheet is missing.")
            else:
                headers = worksheet_headers(workbook["EOAT Inventory"])
                expected = get_expected_headers("EOAT Inventory")
                for header in expected:
                    if header not in headers:
                        changes.append(
                            RepairChange(
                                sheet_name="EOAT Inventory",
                                row_number=1,
                                column_name=header,
                                current_value="missing header",
                                new_value=header,
                                reason="Required header is missing.",
                            )
                        )
                if LEGACY_TOOL_FIELD in headers and TOOL_FIELD not in headers:
                    changes.append(_header_change(LEGACY_TOOL_FIELD, TOOL_FIELD))
                if LEGACY_VACUUM_CUPS_FIELD in headers:
                    changes.append(_header_change(LEGACY_VACUUM_CUPS_FIELD, "Number of Parts Picked"))
        except Exception as exc:
            warnings.append(f"Could not preview legacy header repair: {exc}")
        finally:
            if workbook is not None:
                workbook.close()
    return SafeFixPreview(
        fix_id=FIX_REPAIR_LEGACY_HEADERS,
        title="Repair Legacy Headers",
        description="Repair known legacy headers and add missing current schema headers using the established schema repair path.",
        changes=changes,
        warnings=warnings,
        can_apply=bool(changes) and not warnings,
    )


def _preview_refresh_generated_views(project_root: str | Path) -> SafeFixPreview:
    paths = resolve_project_paths(project_root)
    changes: list[RepairChange] = []
    warnings: list[str] = []
    if not paths.master_workbook.exists():
        warnings.append(f"Master workbook is missing: {paths.master_workbook}")
    else:
        workbook = None
        try:
            workbook = load_workbook(paths.master_workbook, read_only=True, data_only=False)
            if AUDIT_BY_PRESS_SHEET not in workbook.sheetnames or audit_by_press_last_refreshed(workbook) is None:
                changes.append(
                    RepairChange(
                        sheet_name=AUDIT_BY_PRESS_SHEET,
                        current_value="missing or stale generated view",
                        new_value="refreshed generated view",
                        reason="Audit by Press is a generated view from EOAT Inventory.",
                    )
                )
        except Exception as exc:
            warnings.append(f"Could not preview generated view refresh: {exc}")
        finally:
            if workbook is not None:
                workbook.close()
    return SafeFixPreview(
        fix_id=FIX_REFRESH_GENERATED_VIEWS,
        title="Refresh Generated Views",
        description="Refresh generated workbook views from source workbook data.",
        changes=changes,
        warnings=warnings,
        can_apply=bool(changes) and not warnings,
    )


def _generic_preview(fix_id: str, title: str, description: str, change_text: str, project_root: str | Path) -> SafeFixPreview:
    paths = resolve_project_paths(project_root)
    if not paths.master_workbook.exists():
        return SafeFixPreview(
            fix_id=fix_id,
            title=title,
            description=description,
            warnings=[f"Master workbook is missing: {paths.master_workbook}"],
            can_apply=False,
        )
    return SafeFixPreview(
        fix_id=fix_id,
        title=title,
        description=description,
        changes=[RepairChange(sheet_name="EOAT Inventory", current_value="current workbook", new_value="refreshed workbook structure", reason=change_text)],
        can_apply=True,
    )


def _apply_clear_stale_hidden_na(project_root: str | Path, preview: SafeFixPreview, started: float) -> ToolResult:
    paths = resolve_project_paths(project_root)
    workbook_path = paths.master_workbook
    backup = backup_file(workbook_path, workbook_path.parent / "_backups")
    workbook = None
    audit_changes: dict[str, dict[str, dict[str, Any]]] = {}
    applied = 0
    try:
        workbook = load_workbook(workbook_path)
        if "EOAT Inventory" not in workbook.sheetnames:
            raise ValueError("EOAT Inventory sheet is missing.")
        ws = workbook["EOAT Inventory"]
        headers = worksheet_headers(ws)
        header_positions = {header: index + 1 for index, header in enumerate(headers)}
        for change in preview.changes:
            if not change.row_number or change.column_name not in header_positions:
                continue
            column = header_positions[change.column_name]
            cell = ws.cell(row=change.row_number, column=column)
            before_value = cell.value
            if _text(before_value) == NA_VALUE:
                continue
            cell.value = NA_VALUE
            applied += 1
            audit_key = change.audit_id or f"row-{change.row_number}"
            audit_changes.setdefault(audit_key, {"before": {}, "after": {}})
            audit_changes[audit_key]["before"][change.column_name] = before_value
            audit_changes[audit_key]["after"][change.column_name] = NA_VALUE
        refresh_audit_by_press_view(workbook)
        workbook.save(workbook_path)
        invalidate_workbook_cache(workbook_path)
    except Exception as exc:
        return ToolResult.fail(
            "workbook_repair",
            "Workbook Repair",
            "Could not apply stale hidden value repair.",
            errors=[str(exc)],
            files_created=[str(backup)],
            duration_seconds=time.perf_counter() - started,
        )
    finally:
        if workbook is not None:
            workbook.close()

    files_modified = [str(workbook_path)]
    files_created = [str(backup)]
    for audit_id, changes in audit_changes.items():
        try:
            history_path = append_audit_history(
                project_root,
                audit_id,
                "validation_auto_fix",
                changes["before"],
                changes["after"],
                source=f"workbook_repair:{preview.fix_id}",
                files_modified=files_modified,
            )
            files_modified.append(str(history_path))
        except Exception:
            pass
    return ToolResult.ok(
        "workbook_repair",
        "Workbook Repair",
        "Applied safe fix: clear stale hidden values.",
        details=[
            f"Workbook backup: {backup}",
            f"Changed stale hidden value(s): {applied}",
            "No engineering values were guessed; non-applicable fields were set to N/A only.",
        ],
        files_created=files_created,
        files_modified=sorted(set(files_modified)),
        metrics={"fix_id": preview.fix_id, "applied_change_count": applied},
        duration_seconds=time.perf_counter() - started,
    )


def _append_workbook_repair_history(project_root: str | Path, preview: SafeFixPreview, files_modified: list[str]) -> None:
    try:
        append_audit_history(
            project_root,
            "WORKBOOK",
            "workbook_repair",
            {},
            {"fix_id": preview.fix_id, "change_count": str(len(preview.changes))},
            source=f"workbook_repair:{preview.fix_id}",
            files_modified=files_modified,
        )
    except Exception:
        pass


def _header_change(current_header: str, new_header: str) -> RepairChange:
    return RepairChange(
        sheet_name="EOAT Inventory",
        row_number=1,
        column_name=current_header,
        current_value=current_header,
        new_value=new_header,
        reason="Known legacy header repair.",
    )


def _preview_dict(preview: SafeFixPreview) -> dict[str, Any]:
    return {
        "fix_id": preview.fix_id,
        "title": preview.title,
        "description": preview.description,
        "can_apply": preview.can_apply,
        "warnings": list(preview.warnings),
        "changes": [change.to_dict() for change in preview.changes],
    }


def _is_audit_data_row(row_data: dict[str, object]) -> bool:
    values = [_text(value) for value in row_data.values()]
    if not any(values):
        return False
    if len([value for value in values if value]) == 1 and values[-1].startswith("Last Updated:"):
        return False
    return True


def _text(value: Any) -> str:
    return str(value or "").strip()


__all__ = [
    "FIX_CLEAR_STALE_HIDDEN_NA",
    "FIX_REAPPLY_FORMATTING",
    "FIX_REBUILD_DROPDOWN_VALIDATION",
    "FIX_REFRESH_GENERATED_VIEWS",
    "FIX_REPAIR_LEGACY_HEADERS",
    "SAFE_FIX_IDS",
    "RepairChange",
    "SafeFixPreview",
    "apply_safe_fix",
    "preview_safe_fix",
    "preview_safe_fix_action",
]
