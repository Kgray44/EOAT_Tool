from __future__ import annotations

import hashlib
import json
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .analysis_common import parse_score, table_from_rows, write_timestamped_report
from .annotations.service import AnnotationService
from .audit_entries import repair_legacy_audit_lookup_shift
from .logging import log_tool_run
from .open_items import list_open_items
from .paths import resolve_project_paths
from .photo_evidence import evidence_coverage_for_project
from .result import ToolResult
from .safe_files import backup_file, ensure_directory, safe_write_text
from .standards_compliance import analyze_standards_compliance
from .validation_findings import ValidationFinding
from .workbook_cache import invalidate_workbook_cache
from .workbook_io import next_empty_row, row_dicts, write_row_by_headers
from .workbook_schema import get_expected_headers

FMEA_FAILURE_MODE_LIBRARY = {
    "vacuum loss": {
        "failure_mode": "Vacuum loss",
        "effect": "Part drop, mis-pick, quality issue, or downtime.",
        "cause": "Worn cup, leaking tubing, poor seal, routing issue, or vacuum generator issue.",
        "controls": "Current controls must be verified during review.",
        "mitigation": "Review cup condition, routing, generator condition, and confirmation method before defining corrective action.",
    },
    "misalignment": {
        "failure_mode": "Misalignment",
        "effect": "Poor pickup, placement error, part damage, or process interruption.",
        "cause": "EOAT alignment drift, loose mounting, tool crash, or setup variation.",
        "controls": "Current alignment checks must be verified during review.",
        "mitigation": "Review mounting/alignment evidence and define the inspection or correction standard.",
    },
    "tubing failure": {
        "failure_mode": "Tubing failure",
        "effect": "Vacuum loss, gripper failure, intermittent part handling, or downtime.",
        "cause": "Kinked, worn, leaking, pinched, or poorly routed tubing.",
        "controls": "Visual inspection and any pneumatic checks must be verified during review.",
        "mitigation": "Review tubing routing and define corrective routing/PM actions.",
    },
    "sensor failure": {
        "failure_mode": "Sensor failure",
        "effect": "False confirmation, missed part detection, robot fault, or quality escape.",
        "cause": "Damaged sensor, poor mounting, wiring issue, or missing confirmation standard.",
        "controls": "Sensor checks and current confirmation logic must be verified during review.",
        "mitigation": "Review sensor standard, mounting, and confirmation method.",
    },
    "mechanical wear": {
        "failure_mode": "Mechanical wear",
        "effect": "Inconsistent grip, part damage, alignment issue, or repeated adjustment.",
        "cause": "Worn gripper fingers, loose hardware, missing locking hardware, or repeated impacts.",
        "controls": "Mechanical inspection controls must be verified during review.",
        "mitigation": "Review wear points, spare parts, and PM/rebuild standard.",
    },
    "quick disconnect mismatch": {
        "failure_mode": "Quick disconnect mismatch",
        "effect": "Setup delay, wrong connection, intermittent pneumatic/electrical performance, or downtime.",
        "cause": "Unstandardized quick disconnects, wear, damage, or missing labeling.",
        "controls": "Current connection standard must be verified during review.",
        "mitigation": "Review quick disconnect type, labeling, and standardization opportunity.",
    },
    "documentation failure": {
        "failure_mode": "Documentation failure",
        "effect": "Inconsistent setup, maintenance delay, incorrect spare parts, or repeat troubleshooting.",
        "cause": "Missing BOM, CAD, process binder, photo evidence, or unclear ownership.",
        "controls": "Documentation control status must be verified during review.",
        "mitigation": "Resolve documentation gaps or record them as approved follow-up items.",
    },
    "maintenance access issue": {
        "failure_mode": "Maintenance access issue",
        "effect": "Delayed PM, missed inspection, repeat failures, or avoidable downtime.",
        "cause": "Missing PM standard, unclear access points, missing spare parts, or incomplete evidence.",
        "controls": "PM readiness and access constraints must be verified during review.",
        "mitigation": "Review PM checklist readiness, access constraints, and required parts/resources.",
    },
}

FMEA_CONFIDENCE_LABELS = {"High", "Medium", "Low"}


@dataclass(frozen=True)
class FmeaSuggestion:
    suggestion_id: str
    audit_id: str
    machine: str
    failure_mode: str
    evidence: str
    suggested_severity: str
    suggested_frequency: str
    suggested_detectability: str
    suggested_mitigation: str
    source_fields_tags: str
    source_type: str
    issue_category: str = ""
    failure_effect: str = ""
    potential_cause: str = ""
    current_controls: str = ""
    review_status: str = "Suggested - user review required"
    confidence: str = "Low"
    calculated_rpn: int | str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "Suggestion ID": self.suggestion_id,
            "Audit ID": self.audit_id,
            "Press/Machine #": self.machine,
            "Failure Mode": self.failure_mode,
            "Evidence": self.evidence,
            "Suggested Severity": self.suggested_severity,
            "Suggested Frequency": self.suggested_frequency,
            "Suggested Detectability": self.suggested_detectability,
            "Suggested Mitigation": self.suggested_mitigation,
            "Source Fields/Tags": self.source_fields_tags,
            "Source Type": self.source_type,
            "Issue Category": self.issue_category,
            "Failure Effect": self.failure_effect,
            "Potential Cause": self.potential_cause,
            "Current Controls": self.current_controls,
            "Review Status": self.review_status,
            "Confidence": self.confidence,
            "Calculated RPN": self.calculated_rpn,
        }


def fmea_suggestion_decisions_path(project_root: str | Path) -> Path:
    return resolve_project_paths(project_root).project_data / "fmea_suggestion_decisions.json"


def build_fmea_suggestions(project_root: str | Path, *, include_rejected: bool = False) -> list[dict[str, Any]]:
    paths = resolve_project_paths(project_root)
    if not paths.master_workbook.exists():
        return []
    decisions = _load_decisions(project_root)
    try:
        existing_modes = {_text(row.get("Failure Mode")).casefold() for row in row_dicts(paths.master_workbook, "FMEA Draft")}
        inventory = [repair_legacy_audit_lookup_shift(row) for row in row_dicts(paths.master_workbook, "EOAT Inventory")]
        issues = row_dicts(paths.master_workbook, "Issue Log")
    except Exception:
        return []
    suggestions: list[FmeaSuggestion] = []
    suggestions.extend(_issue_suggestions(issues, existing_modes))
    suggestions.extend(_audit_field_suggestions(inventory, existing_modes))
    suggestions.extend(_annotation_suggestions(project_root, existing_modes))
    suggestions.extend(_open_item_suggestions(project_root, existing_modes))
    suggestions.extend(_validation_suggestions(project_root, existing_modes))
    suggestions.extend(_photo_evidence_gap_suggestions(project_root, existing_modes))
    suggestions.extend(_standards_failure_suggestions(project_root, existing_modes))
    rows = []
    for suggestion in _dedupe_suggestions(suggestions):
        state = decisions.get(suggestion.suggestion_id, {}).get("status", "")
        if state == "rejected" and not include_rejected:
            continue
        if state == "accepted" and not include_rejected:
            continue
        row = suggestion.to_dict()
        row["Decision"] = state or "pending"
        rows.append(row)
    return sorted(rows, key=lambda row: (str(row.get("Press/Machine #")), str(row.get("Failure Mode")), str(row.get("Source Type"))))


def accept_fmea_suggestions(project_root: str | Path, reviewed_suggestions: Iterable[dict[str, Any]], *, log_activity: bool = True) -> ToolResult:
    started = time.perf_counter()
    suggestions = [dict(item) for item in reviewed_suggestions]
    if not suggestions:
        return ToolResult.fail("fmea_suggestion_accept", "FMEA Suggestion Review", "No FMEA suggestions were selected.")
    invalid = [
        str(item.get("Suggestion ID") or item.get("Failure Mode") or "selected suggestion")
        for item in suggestions
        if not _reviewed_scores(item)
    ]
    if invalid:
        return ToolResult.fail(
            "fmea_suggestion_accept",
            "FMEA Suggestion Review",
            "Suggestions require reviewed numeric Severity, Frequency, and Detectability values before acceptance.",
            errors=invalid,
        )
    paths = resolve_project_paths(project_root)
    workbook_path = paths.master_workbook
    if not workbook_path.exists():
        return ToolResult.fail("fmea_suggestion_accept", "FMEA Suggestion Review", "Master workbook is missing.", errors=[str(workbook_path)])
    workbook = None
    try:
        backup = backup_file(workbook_path, workbook_path.parent / "_backups")
        workbook = load_workbook(workbook_path)
        if "FMEA Draft" not in workbook.sheetnames:
            raise ValueError("FMEA Draft sheet is missing.")
        ws = workbook["FMEA Draft"]
        next_ids = _next_fmea_ids(workbook_path, len(suggestions))
        rows_written: list[int] = []
        for fmea_id, suggestion in zip(next_ids, suggestions):
            sev = parse_score(suggestion.get("Suggested Severity"))
            freq = parse_score(suggestion.get("Suggested Frequency"))
            det = parse_score(suggestion.get("Suggested Detectability"))
            row_number = next_empty_row(ws)
            data = {header: "" for header in get_expected_headers("FMEA Draft")}
            data.update(
                {
                    "FMEA ID": fmea_id,
                    "Plant/Area": suggestion.get("Plant/Area", ""),
                    "Press/Machine #": suggestion.get("Press/Machine #", ""),
                    "EOAT Function": "EOAT handling / process support",
                    "Failure Mode": suggestion.get("Failure Mode", ""),
                    "Failure Effect": suggestion.get("Failure Effect", ""),
                    "Potential Cause": suggestion.get("Potential Cause", ""),
                    "Current Controls": suggestion.get("Current Controls", ""),
                    "Severity": sev,
                    "Frequency": freq,
                    "Detectability": det,
                    "RPN": int(sev or 0) * int(freq or 0) * int(det or 0),
                    "Recommended Action": suggestion.get("Suggested Mitigation", ""),
                    "Status": "Draft - Review",
                    "Notes": _accepted_notes(suggestion),
                }
            )
            write_row_by_headers(ws, row_number, data)
            rows_written.append(row_number)
        workbook.save(workbook_path)
        workbook.close()
        invalidate_workbook_cache(workbook_path)
        workbook = None
        _record_decisions(project_root, suggestions, "accepted")
    except Exception as exc:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        return ToolResult.fail("fmea_suggestion_accept", "FMEA Suggestion Review", "Could not accept FMEA suggestions.", errors=[str(exc)])
    result = ToolResult.ok(
        "fmea_suggestion_accept",
        "FMEA Suggestion Review",
        f"Accepted {len(suggestions)} reviewed FMEA suggestion(s) into FMEA Draft.",
        details=[f"Workbook backup: {backup}", f"Rows written: {', '.join(str(row) for row in rows_written)}"],
        files_created=[str(backup)],
        files_modified=[str(workbook_path)],
        metrics={"accepted_suggestions": len(suggestions)},
        duration_seconds=time.perf_counter() - started,
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def reject_fmea_suggestions(project_root: str | Path, suggestion_ids: Iterable[str], *, reason: str = "", log_activity: bool = True) -> ToolResult:
    ids = [str(item).strip() for item in suggestion_ids if str(item).strip()]
    if not ids:
        return ToolResult.fail("fmea_suggestion_reject", "FMEA Suggestion Review", "No FMEA suggestions were selected.")
    rows = [{"Suggestion ID": suggestion_id, "Reject Reason": reason} for suggestion_id in ids]
    _record_decisions(project_root, rows, "rejected")
    result = ToolResult.ok(
        "fmea_suggestion_reject",
        "FMEA Suggestion Review",
        f"Rejected {len(ids)} FMEA suggestion(s).",
        files_modified=[str(fmea_suggestion_decisions_path(project_root))],
        metrics={"rejected_suggestions": len(ids)},
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def export_fmea_suggestion_draft(project_root: str | Path, suggestions: Iterable[dict[str, Any]] | None = None, *, log_activity: bool = True) -> ToolResult:
    rows = [dict(row) for row in (suggestions if suggestions is not None else build_fmea_suggestions(project_root))]
    paths = resolve_project_paths(project_root)
    ensure_directory(paths.fmea_reports)
    markdown = "\n".join(
        [
            "# FMEA Suggestion Draft",
            "",
            "These are draft suggestions generated from existing audit evidence. They are not final engineering decisions.",
            "",
            "## Suggestions",
            *table_from_rows(
                rows,
                [
                    "Suggestion ID",
                    "Press/Machine #",
                    "Failure Mode",
                    "Confidence",
                    "Calculated RPN",
                    "Evidence",
                    "Suggested Severity",
                    "Suggested Frequency",
                    "Suggested Detectability",
                    "Suggested Mitigation",
                ],
            ),
            "",
            "## Evidence Trace",
            *table_from_rows(rows, ["Suggestion ID", "Source Type", "Source Fields/Tags", "Evidence"]),
            "",
            "## Review Required",
            "- Confirm or edit severity, frequency, and detectability before accepting any row.",
            "- Reject suggestions that do not match the physical EOAT or current engineering understanding.",
        ]
    ) + "\n"
    try:
        report = write_timestamped_report(paths.fmea_reports, "FMEA_Suggestion_Draft", markdown)
    except Exception as exc:
        return ToolResult.fail("fmea_suggestion_export", "FMEA Suggestion Review", "Could not write FMEA suggestion draft.", errors=[str(exc)])
    result = ToolResult.ok(
        "fmea_suggestion_export",
        "FMEA Suggestion Review",
        "Exported FMEA suggestion draft.",
        files_created=[str(report)],
        output_reports=[str(report)],
        metrics={"suggestion_count": len(rows)},
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def export_fmea_evidence_report(project_root: str | Path, suggestions: Iterable[dict[str, Any]] | None = None, *, log_activity: bool = True) -> ToolResult:
    return export_fmea_suggestion_draft(project_root, suggestions=suggestions, log_activity=log_activity)


def _issue_suggestions(issues: list[dict[str, Any]], existing_modes: set[str]) -> list[FmeaSuggestion]:
    suggestions: list[FmeaSuggestion] = []
    for row in issues:
        source_text = " ".join(_text(row.get(field)) for field in ["Issue Category", "Issue Description", "Suspected Cause", "Evidence/Observation", "Impact", "Notes"])
        key = _failure_key_from_text(source_text)
        suggestion = _suggestion_from_key(
            key,
            audit_id="",
            machine=_text(row.get("Press/Machine #")),
            evidence=_first_text(row, ["Issue Description", "Evidence/Observation", "Impact", "Notes"]) or f"Issue category: {_text(row.get('Issue Category'))}",
            source_fields_tags="Issue Log: Issue Category, Issue Description, Evidence/Observation",
            source_type="issue",
            issue_category=_text(row.get("Issue Category")),
            severity=_score_or_review(row.get("Severity")),
            frequency=_score_or_review(row.get("Frequency")),
            detectability=_score_or_review(row.get("Detectability")),
        )
        if suggestion.failure_mode.casefold() not in existing_modes:
            suggestions.append(suggestion)
    return suggestions


def _audit_field_suggestions(inventory: list[dict[str, Any]], existing_modes: set[str]) -> list[FmeaSuggestion]:
    suggestions: list[FmeaSuggestion] = []
    for row in inventory:
        source_text = " ".join(_text(row.get(field)) for field in ["Known Issues", "Drop/Mis-Pick History", "Tubing Condition", "Cable Management Condition", "Maintenance Frequency", "Notes"])
        if not source_text.strip():
            continue
        key = _failure_key_from_text(source_text)
        suggestion = _suggestion_from_key(
            key,
            audit_id=_text(row.get("Audit ID")),
            machine=_text(row.get("Press/Machine #")),
            evidence=_first_text(row, ["Known Issues", "Drop/Mis-Pick History", "Notes", "Tubing Condition"]) or "Audit field review needed.",
            source_fields_tags="EOAT Inventory: Known Issues, Drop/Mis-Pick History, condition fields",
            source_type="audit",
        )
        if suggestion.failure_mode.casefold() not in existing_modes:
            suggestions.append(suggestion)
    return suggestions


def _annotation_suggestions(project_root: str | Path, existing_modes: set[str]) -> list[FmeaSuggestion]:
    suggestions: list[FmeaSuggestion] = []
    try:
        service = AnnotationService(project_root)
        notes = service.search_notes(include_archived=False)
        assignments = service.list_tag_assignments(include_archived=False)
    except Exception:
        return []
    for note in notes:
        text = " ".join(_text(note.get(field)) for field in ["subject", "body_markdown", "importance", "note_type"])
        if not text.strip():
            continue
        key = _failure_key_from_text(text)
        suggestion = _suggestion_from_key(
            key,
            audit_id=_text(note.get("audit_id")),
            machine=_text(note.get("machine_id")),
            evidence=_text(note.get("subject")) or "Annotation note suggests review.",
            source_fields_tags="Note",
            source_type="note",
        )
        if suggestion.failure_mode.casefold() not in existing_modes:
            suggestions.append(suggestion)
    for assignment in assignments:
        text = " ".join(_text(assignment.get(field)) for field in ["tag_name", "comment", "target_label", "field_label"])
        key = _failure_key_from_text(text)
        suggestion = _suggestion_from_key(
            key,
            audit_id=_text(assignment.get("audit_id")),
            machine=_text(assignment.get("machine_id")),
            evidence=f"Tag: {_text(assignment.get('tag_name'))}" + (f" - {_text(assignment.get('comment'))}" if _text(assignment.get("comment")) else ""),
            source_fields_tags=f"Tag: {_text(assignment.get('tag_name'))}",
            source_type="tag",
        )
        if suggestion.failure_mode.casefold() not in existing_modes:
            suggestions.append(suggestion)
    return suggestions


def _open_item_suggestions(project_root: str | Path, existing_modes: set[str]) -> list[FmeaSuggestion]:
    suggestions: list[FmeaSuggestion] = []
    try:
        items = list_open_items(project_root, include_validation=False)
    except Exception:
        return []
    for item in items:
        text = " ".join([item.title, item.message, item.category, item.field, item.recommended_action])
        key = _failure_key_from_text(text)
        suggestion = _suggestion_from_key(
            key,
            audit_id=item.audit_id,
            machine=item.machine,
            evidence=item.message or item.title,
            source_fields_tags=f"Open Item: {item.category}",
            source_type="open_item",
        )
        if suggestion.failure_mode.casefold() not in existing_modes:
            suggestions.append(suggestion)
    return suggestions


def _validation_suggestions(project_root: str | Path, existing_modes: set[str]) -> list[FmeaSuggestion]:
    suggestions: list[FmeaSuggestion] = []
    for finding in _latest_validation_findings(project_root):
        text = " ".join([finding.message, finding.category, finding.column_name, finding.recommended_action])
        key = _failure_key_from_text(text)
        suggestion = _suggestion_from_key(
            key,
            audit_id=finding.audit_id,
            machine=finding.machine_number,
            evidence=finding.message,
            source_fields_tags=f"Validation finding: {finding.category}",
            source_type="validation",
        )
        if suggestion.failure_mode.casefold() not in existing_modes:
            suggestions.append(suggestion)
    return suggestions


def _photo_evidence_gap_suggestions(project_root: str | Path, existing_modes: set[str]) -> list[FmeaSuggestion]:
    suggestions: list[FmeaSuggestion] = []
    try:
        coverages = evidence_coverage_for_project(project_root)
    except Exception:
        return []
    for coverage in coverages:
        missing = [status.label for status in coverage.statuses if status.required and not status.present]
        if not missing:
            continue
        key = "documentation failure"
        if any("sensor" in item.casefold() for item in missing):
            key = "sensor failure"
        elif any("quick" in item.casefold() for item in missing):
            key = "quick disconnect mismatch"
        elif any("tubing" in item.casefold() for item in missing):
            key = "tubing failure"
        elif any("gripper" in item.casefold() or "mounting" in item.casefold() for item in missing):
            key = "mechanical wear"
        suggestion = _suggestion_from_key(
            key,
            audit_id=coverage.audit_id,
            machine=coverage.machine,
            evidence=f"Required evidence missing: {', '.join(missing)}.",
            source_fields_tags="Photo evidence coverage",
            source_type="photo_evidence",
        )
        if suggestion.failure_mode.casefold() not in existing_modes:
            suggestions.append(suggestion)
    return suggestions


def _standards_failure_suggestions(project_root: str | Path, existing_modes: set[str]) -> list[FmeaSuggestion]:
    summary, error = analyze_standards_compliance(project_root)
    if error or summary is None:
        return []
    suggestions: list[FmeaSuggestion] = []
    for audit in summary.audits:
        for category in [*audit.failed_standards, *audit.warnings]:
            key = _failure_key_from_text(" ".join([category.key, category.label, category.reason]))
            suggestion = _suggestion_from_key(
                key,
                audit_id=audit.audit_id,
                machine=audit.machine,
                evidence=f"{category.label}: {category.reason}",
                source_fields_tags=f"Standards compliance: {', '.join(category.related_fields)}",
                source_type="standards",
            )
            if suggestion.failure_mode.casefold() not in existing_modes:
                suggestions.append(suggestion)
    return suggestions


def _suggestion_from_key(
    key: str,
    *,
    audit_id: str,
    machine: str,
    evidence: str,
    source_fields_tags: str,
    source_type: str,
    issue_category: str = "",
    severity: str = "Review Required",
    frequency: str = "Review Required",
    detectability: str = "Review Required",
) -> FmeaSuggestion:
    details = FMEA_FAILURE_MODE_LIBRARY.get(key, FMEA_FAILURE_MODE_LIBRARY["documentation failure"])
    suggestion_id = _stable_id([audit_id, machine, details["failure_mode"], source_type, evidence, source_fields_tags])
    calculated_rpn = _calculated_rpn(severity, frequency, detectability)
    return FmeaSuggestion(
        suggestion_id=suggestion_id,
        audit_id=audit_id,
        machine=machine,
        failure_mode=details["failure_mode"],
        failure_effect=details["effect"],
        potential_cause=details["cause"],
        current_controls=details["controls"],
        evidence=evidence,
        suggested_severity=severity,
        suggested_frequency=frequency,
        suggested_detectability=detectability,
        suggested_mitigation=details["mitigation"],
        source_fields_tags=source_fields_tags,
        source_type=source_type,
        issue_category=issue_category,
        confidence=_confidence_label(evidence, severity, frequency, detectability, source_type),
        calculated_rpn=calculated_rpn,
    )


def _dedupe_suggestions(suggestions: list[FmeaSuggestion]) -> list[FmeaSuggestion]:
    grouped: dict[tuple[str, str, str], list[FmeaSuggestion]] = {}
    for suggestion in suggestions:
        key = (suggestion.audit_id.casefold(), suggestion.machine.casefold(), suggestion.failure_mode.casefold())
        grouped.setdefault(key, []).append(suggestion)
    deduped: list[FmeaSuggestion] = []
    for group in grouped.values():
        first = group[0]
        evidence = "; ".join(dict.fromkeys(item.evidence for item in group if item.evidence))[:1000]
        sources = "; ".join(dict.fromkeys(item.source_fields_tags for item in group if item.source_fields_tags))[:500]
        source_type = ", ".join(dict.fromkeys(item.source_type for item in group if item.source_type))
        calculated_rpn = _calculated_rpn(first.suggested_severity, first.suggested_frequency, first.suggested_detectability)
        deduped.append(
            FmeaSuggestion(
                **{
                    **asdict(first),
                    "suggestion_id": _stable_id([first.audit_id, first.machine, first.failure_mode, evidence, sources]),
                    "evidence": evidence,
                    "source_fields_tags": sources,
                    "source_type": source_type,
                    "confidence": _confidence_label(evidence, first.suggested_severity, first.suggested_frequency, first.suggested_detectability, source_type),
                    "calculated_rpn": calculated_rpn,
                }
            )
        )
    return deduped


def _failure_key_from_text(text: str) -> str:
    folded = text.casefold()
    if any(token in folded for token in ("vacuum", "cup", "poor seal")):
        return "vacuum loss"
    if any(token in folded for token in ("mis-pick", "mispick", "drop", "alignment", "misaligned")):
        return "misalignment"
    if any(token in folded for token in ("tubing", "pneumatic", "air line", "leak", "kink")):
        return "tubing failure"
    if any(token in folded for token in ("sensor", "part-present", "confirmation", "electrical", "cable")):
        return "sensor failure"
    if any(token in folded for token in ("gripper", "wear", "loose", "hardware", "mounting", "fastener")):
        return "mechanical wear"
    if any(token in folded for token in ("quick disconnect", "disconnect", "coupler", "mismatch")):
        return "quick disconnect mismatch"
    if any(token in folded for token in ("pm", "maintenance", "access")):
        return "maintenance access issue"
    if any(token in folded for token in ("documentation", "document", "bom", "binder", "photo", "evidence")):
        return "documentation failure"
    return "documentation failure"


def _latest_validation_findings(project_root: str | Path) -> list[ValidationFinding]:
    folder = resolve_project_paths(project_root).validation_reports
    if not folder.exists():
        return []
    for path in sorted(folder.glob("Foundation_Validation_*.json"), key=lambda item: item.stat().st_mtime, reverse=True):
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        rows = payload.get("findings", []) if isinstance(payload, dict) else []
        return [ValidationFinding.from_dict(row) for row in rows if isinstance(row, dict)]
    return []


def _reviewed_scores(row: dict[str, Any]) -> bool:
    return all(_valid_score(row.get(field)) for field in ["Suggested Severity", "Suggested Frequency", "Suggested Detectability"])


def _calculated_rpn(severity: Any, frequency: Any, detectability: Any) -> int | str:
    sev = parse_score(severity)
    freq = parse_score(frequency)
    det = parse_score(detectability)
    if all(score is not None and 1 <= int(score) <= 10 for score in [sev, freq, det]):
        return int(sev or 0) * int(freq or 0) * int(det or 0)
    return ""


def _confidence_label(evidence: Any, severity: Any, frequency: Any, detectability: Any, source_type: Any) -> str:
    source = _text(source_type).casefold()
    numeric_scores = sum(_valid_score(value) for value in [severity, frequency, detectability])
    evidence_text = _text(evidence)
    independent_sources = len([item for item in source.replace(";", ",").split(",") if item.strip()])
    strong_source = any(token in source for token in ["issue", "validation", "standards", "photo_evidence"])
    if evidence_text and numeric_scores == 3 and (strong_source or independent_sources > 1):
        return "High"
    if evidence_text and (numeric_scores >= 1 or strong_source or independent_sources > 1):
        return "Medium"
    return "Low"


def _valid_score(value: Any) -> bool:
    score = parse_score(value)
    return score is not None and 1 <= score <= 10


def _next_fmea_ids(workbook_path: str | Path, count: int) -> list[str]:
    rows = row_dicts(workbook_path, "FMEA Draft")
    prefix = f"FMEA-{datetime.now().strftime('%Y%m%d')}-"
    max_number = 0
    for row in rows:
        value = _text(row.get("FMEA ID"))
        if value.startswith(prefix):
            try:
                max_number = max(max_number, int(value.rsplit("-", 1)[1]))
            except ValueError:
                continue
    return [f"{prefix}{number:03d}" for number in range(max_number + 1, max_number + count + 1)]


def _accepted_notes(suggestion: dict[str, Any]) -> str:
    return (
        f"Accepted from FMEA suggestion {suggestion.get('Suggestion ID', '')}. "
        f"Evidence: {suggestion.get('Evidence', '')}. "
        f"Source: {suggestion.get('Source Fields/Tags', '')}. "
        "Values require engineering review before final approval."
    )


def _record_decisions(project_root: str | Path, rows: Iterable[dict[str, Any]], status: str) -> None:
    path = fmea_suggestion_decisions_path(project_root)
    ensure_directory(path.parent)
    decisions = _load_decisions(project_root)
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    for row in rows:
        suggestion_id = _text(row.get("Suggestion ID"))
        if not suggestion_id:
            continue
        decisions[suggestion_id] = {
            "status": status,
            "updated_at": now,
            "reason": _text(row.get("Reject Reason")),
        }
    safe_write_text(path, json.dumps({"decisions": decisions}, indent=2, sort_keys=True) + "\n", overwrite=True)


def _load_decisions(project_root: str | Path) -> dict[str, dict[str, Any]]:
    path = fmea_suggestion_decisions_path(project_root)
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    decisions = payload.get("decisions", {}) if isinstance(payload, dict) else {}
    return decisions if isinstance(decisions, dict) else {}


def _score_or_review(value: Any) -> str:
    score = parse_score(value)
    return str(score) if score is not None and 1 <= score <= 10 else "Review Required"


def _first_text(row: dict[str, Any], fields: list[str]) -> str:
    for field in fields:
        text = _text(row.get(field))
        if text:
            return text
    return ""


def _stable_id(parts: Iterable[Any]) -> str:
    digest = hashlib.sha256("\u241f".join(_text(part) for part in parts).encode("utf-8")).hexdigest()[:16]
    return f"fmea_sug_{digest}"


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


__all__ = [
    "FMEA_CONFIDENCE_LABELS",
    "FmeaSuggestion",
    "accept_fmea_suggestions",
    "build_fmea_suggestions",
    "export_fmea_evidence_report",
    "export_fmea_suggestion_draft",
    "fmea_suggestion_decisions_path",
    "reject_fmea_suggestions",
]
