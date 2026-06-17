from __future__ import annotations

import os
import re
import shutil
import time
from collections.abc import Mapping
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .eoat_ids import (
    EOAT_ASSEMBLY_ID_FIELD,
    build_eoat_assembly_contexts,
    infer_eoat_assembly_id_for_photo_row,
    is_valid_eoat_assembly_id,
    normalize_eoat_assembly_id,
    update_eoat_info_file,
)
from .logging import log_tool_run
from .paths import resolve_project_paths
from .result import ToolResult
from .safe_files import backup_file, ensure_directory, safe_copy_file
from .tool_fields import TOOL_FIELD
from .workbook_cache import invalidate_workbook_cache
from .workbook_io import next_empty_row, row_dicts, worksheet_headers, write_row_by_headers
from .workbook_schema import get_expected_headers

SUPPORTED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".heic", ".heif"}
LINKED_AUDIT_FIELD_HEADER = "Linked Audit Field"

PHOTO_CATEGORY_FOLDERS: tuple[str, ...] = (
    "00_Overall",
    "01_Front_View",
    "02_Side_View",
    "03_Vacuum_Cups_Grippers",
    "04_Tubing_Routing",
    "05_Sensors",
    "06_Quick_Disconnects",
    "07_Cable_Management",
    "08_Mounting_Hardware",
    "09_Wear_Damage",
    "10_Back_View",
    "11_Tool_Number",
)

PHOTO_VIEW_FOLDERS = {
    "Front View": "01_Front_View",
    "Side View": "02_Side_View",
    "Back View": "10_Back_View",
    "Tool Number": "11_Tool_Number",
    "Vacuum Cups / Grippers": "03_Vacuum_Cups_Grippers",
    "Gripper": "03_Vacuum_Cups_Grippers",
    "Tubing Routing": "04_Tubing_Routing",
    "Sensors": "05_Sensors",
    "Quick Disconnects": "06_Quick_Disconnects",
    "Cable Management": "07_Cable_Management",
    "Mounting Hardware": "08_Mounting_Hardware",
    "Wear / Damage": "09_Wear_Damage",
}

PHOTO_VIEW_FOLDER_ALIASES = {
    **PHOTO_VIEW_FOLDERS,
    "Overall": "00_Overall",
    "Overall EOAT": "00_Overall",
    "Wear/Damage": "09_Wear_Damage",
    "Back": "10_Back_View",
    "Grippers": "03_Vacuum_Cups_Grippers",
    "Vacuum Cups": "03_Vacuum_Cups_Grippers",
    "Tool Label / ID Plate": "11_Tool_Number",
    "Tool Label": "11_Tool_Number",
    "ID Plate": "11_Tool_Number",
    "Tool Connection": "00_Overall",
    "Robot Connection": "00_Overall",
    "Cylinders": "08_Mounting_Hardware",
}

LEGACY_TOP_LEVEL_PHOTO_FOLDERS: tuple[str, ...] = (
    "Overall",
    "Front_View",
    "Side_View",
    "Vacuum_Cups_Grippers",
    "Vacuum_Cups",
    "Grippers",
    "Tubing_Routing",
    "Sensors",
    "Quick_Disconnects",
    "Cable_Management",
    "Mounting_Hardware",
    "Wear_Damage",
    "Back_View",
    "Tool_Number",
    "Robot_Connection",
    "Tool_Connection",
    "EOAT_Side_Pneumatic_Circuits",
    "EOAT_Side_Pneumatics",
    "Robot_Side_Pneumatics",
    "Sensor_Mounting",
    "Tool_Label_ID_Plate",
    "Process_Binder_Reference",
    "Process_Binder_Documentation_Reference",
    "Other",
    "Incoming",
)

WINDOWS_RESERVED_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{index}" for index in range(1, 10)),
    *(f"LPT{index}" for index in range(1, 10)),
}


@dataclass(frozen=True)
class PhotoPlanItem:
    source: Path
    target: Path
    photo_id: str
    collision_avoided: bool = False
    view_type: str = ""
    description: str = ""
    related_audit_id: str = ""
    related_issue_id: str = ""
    linked_audit_field: str = ""
    eoat_assembly_id: str = ""
    tool_number: str = ""
    part_name: str = ""
    original_filename: str = ""
    stored_relative_path: str = ""


def sanitize_filename_part(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "", value.strip())
    return cleaned or "Unknown"


def sanitize_folder_part(value: str, *, max_length: int = 80) -> str:
    text = _text(value)
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", text)
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", text)
    text = re.sub(r"\.{2,}", "_", text)
    text = re.sub(r"_+", "_", text).strip(" ._")
    if not text or text in {".", ".."}:
        text = "Unknown"
    if text.upper() in WINDOWS_RESERVED_NAMES:
        text = f"{text}_"
    return text[:max_length].rstrip(" ._") or "Unknown"


def sanitize_date_filename_part(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9-]+", "", _text(value))
    return cleaned or "Unknown"


def normalize_machine_filename_part(value: str) -> str:
    text = re.sub(r"^(?:press|machine)\s*[-#:]*\s*", "", _text(value), flags=re.IGNORECASE)
    cleaned = sanitize_filename_part(text)
    return f"Machine{cleaned}"


def build_photo_filename_stem(
    plant_area: str,
    press_machine: str,
    tool_number: str,
    date_taken: str,
    view_type: str,
    eoat_assembly_id: str = "",
) -> str:
    eoat_id = normalize_eoat_assembly_id(eoat_assembly_id)
    if eoat_id:
        return (
            f"{sanitize_folder_part(eoat_id)}_"
            f"{sanitize_date_filename_part(date_taken)}_"
            f"{photo_category_filename_part(view_type)}"
        )
    return (
        f"Tool_{sanitize_folder_part(tool_number)}__"
        f"{photo_category_filename_part(view_type)}__"
        f"{sanitize_date_filename_part(date_taken)}"
    )


def photo_category_folder(view_type: str) -> str:
    folder = PHOTO_VIEW_FOLDER_ALIASES.get(_text(view_type))
    if folder:
        return folder
    candidate = sanitize_folder_part(view_type)
    if candidate in PHOTO_CATEGORY_FOLDERS:
        return candidate
    return "00_Overall"


def photo_category_filename_part(view_type: str) -> str:
    folder = photo_category_folder(view_type)
    return re.sub(r"^\d{2}_", "", folder) or sanitize_folder_part(view_type)


def tool_photo_folder_name(tool_number: str, part_name: str = "") -> str:
    clean_tool = sanitize_folder_part(tool_number)
    folder = f"Tool_{clean_tool}"
    clean_part = sanitize_folder_part(part_name, max_length=60) if _text(part_name) else ""
    if clean_part and clean_part != "Unknown":
        folder = f"{folder}__{clean_part}"
    return folder[:120].rstrip(" ._")


def tool_photo_root(project_root: str | Path, tool_number: str, part_name: str = "") -> Path:
    return resolve_project_paths(project_root).cell_photos / tool_photo_folder_name(tool_number, part_name)


def eoat_photo_root(project_root: str | Path, eoat_assembly_id: str) -> Path:
    eoat_id = normalize_eoat_assembly_id(eoat_assembly_id)
    return resolve_project_paths(project_root).cell_photos / sanitize_folder_part(eoat_id)


def ensure_tool_photo_category_folder(
    project_root: str | Path, tool_number: str, part_name: str = "", view_type: str = ""
) -> Path:
    folder = destination_folder(project_root, view_type, tool_number, part_name)
    ensure_directory(folder)
    return folder


def ensure_eoat_photo_category_folder(project_root: str | Path, eoat_assembly_id: str, view_type: str = "") -> Path:
    folder = destination_folder(project_root, view_type, eoat_assembly_id=eoat_assembly_id)
    ensure_directory(folder)
    return folder


def normalize_cell_photo_folders(project_root: str | Path, *, dry_run: bool = False) -> tuple[list[Path], list[Path]]:
    paths = resolve_project_paths(project_root)
    removed: list[Path] = []
    preserved: list[Path] = []
    if not dry_run:
        ensure_directory(paths.incoming_photos)
    for folder_name in LEGACY_TOP_LEVEL_PHOTO_FOLDERS:
        folder = paths.cell_photos / folder_name
        if not folder.exists() or not folder.is_dir():
            continue
        try:
            is_empty = not any(folder.iterdir())
        except OSError:
            preserved.append(folder)
            continue
        if is_empty:
            removed.append(folder)
            if not dry_run:
                folder.rmdir()
        else:
            preserved.append(folder)
    return removed, preserved


def incoming_photo_dir(project_root: str | Path) -> Path:
    return resolve_project_paths(project_root).incoming_photos


def list_incoming_photos(project_root: str | Path) -> list[Path]:
    folder = incoming_photo_dir(project_root)
    if not folder.exists():
        return []
    return sorted(
        path for path in folder.iterdir() if path.is_file() and path.suffix.lower() in SUPPORTED_IMAGE_EXTENSIONS
    )


def destination_folder(
    project_root: str | Path,
    view_type: str,
    tool_number: str = "",
    part_name: str = "",
    eoat_assembly_id: str = "",
) -> Path:
    folder_name = photo_category_folder(view_type)
    if normalize_eoat_assembly_id(eoat_assembly_id):
        return eoat_photo_root(project_root, eoat_assembly_id) / folder_name
    if _text(tool_number):
        return tool_photo_root(project_root, tool_number, part_name) / folder_name
    return resolve_project_paths(project_root).cell_photos / folder_name


def generate_photo_id(project_root: str | Path, taken_date: str | None = None) -> str:
    taken_date = taken_date or date.today().isoformat()
    compact = taken_date.replace("-", "")
    workbook_path = resolve_project_paths(project_root).master_workbook
    rows = row_dicts(workbook_path, "Photo Index") if workbook_path.exists() else []
    prefix = f"PHO-{compact}-"
    max_number = 0
    for row in rows:
        value = str(row.get("Photo ID") or "")
        if value.startswith(prefix):
            try:
                max_number = max(max_number, int(value.rsplit("-", 1)[1]))
            except ValueError:
                continue
    return f"{prefix}{max_number + 1:03d}"


def _next_existing_photo_sequence(
    project_root: str | Path,
    taken_date: str,
    plant_area: str,
    press_machine: str,
    view_type: str,
    tool_number: str = "",
    part_name: str = "",
    eoat_assembly_id: str = "",
) -> int:
    folder = destination_folder(project_root, view_type, tool_number, part_name, eoat_assembly_id)
    stem = build_photo_filename_stem(plant_area, press_machine, tool_number, taken_date, view_type, eoat_assembly_id)
    stem_prefix = f"{stem}_" if normalize_eoat_assembly_id(eoat_assembly_id) else f"{stem}__"
    max_number = 0
    if folder.exists():
        for item in folder.iterdir():
            if not item.is_file() or not item.stem.startswith(stem_prefix):
                continue
            try:
                max_number = max(max_number, int(item.stem.rsplit("_", 1)[1]))
            except ValueError:
                continue
    return max_number + 1


def preview_photo_intake(
    project_root: str | Path,
    photo_paths: list[str | Path],
    plant_area: str,
    press_machine: str,
    date_taken: str,
    view_type: str,
    tool_number: str = "",
    part_name: str = "",
    related_audit_id: str = "",
    per_photo_metadata: list[Mapping[str, Any]] | None = None,
    eoat_assembly_id: str = "",
) -> list[PhotoPlanItem]:
    if per_photo_metadata is None and not isinstance(tool_number, str):
        per_photo_metadata = tool_number
        tool_number = ""
    plan: list[PhotoPlanItem] = []
    next_sequences: dict[tuple[str, str, str, str], int] = {}
    photo_sequence = 1
    for item_data in _photo_items(
        photo_paths,
        view_type=view_type,
        related_audit_id=related_audit_id,
        eoat_assembly_id=eoat_assembly_id,
        tool_number=tool_number,
        part_name=part_name,
        per_photo_metadata=per_photo_metadata,
    ):
        source = Path(item_data["source"])
        item_view_type = item_data["view_type"] or view_type
        item_eoat_id = item_data["eoat_assembly_id"]
        if not item_eoat_id and item_data["related_audit_id"]:
            item_eoat_id = _eoat_assembly_id_for_audit(project_root, item_data["related_audit_id"])
        if not item_eoat_id and related_audit_id:
            item_eoat_id = _eoat_assembly_id_for_audit(project_root, related_audit_id)
        item_tool_number = item_data["tool_number"]
        item_part_name = item_data["part_name"] or part_name
        if not item_part_name and item_data["related_audit_id"]:
            item_part_name = _part_name_for_audit(project_root, item_data["related_audit_id"])
        if not item_part_name and related_audit_id:
            item_part_name = _part_name_for_audit(project_root, related_audit_id)
        if not item_view_type:
            continue
        if item_eoat_id and not is_valid_eoat_assembly_id(item_eoat_id):
            continue
        if not item_eoat_id and not item_tool_number:
            continue
        folder = destination_folder(project_root, item_view_type, item_tool_number, item_part_name, item_eoat_id)
        sequence_key = (item_view_type, item_eoat_id, item_tool_number, item_part_name)
        if sequence_key not in next_sequences:
            next_sequences[sequence_key] = _next_existing_photo_sequence(
                project_root,
                date_taken,
                plant_area,
                press_machine,
                item_view_type,
                item_tool_number,
                item_part_name,
                item_eoat_id,
            )
        sequence = next_sequences[sequence_key]
        ext = source.suffix
        if ext.lower() not in SUPPORTED_IMAGE_EXTENSIONS:
            continue
        collision_avoided = False
        stem = build_photo_filename_stem(
            plant_area, press_machine, item_tool_number, date_taken, item_view_type, item_eoat_id
        )
        while True:
            separator = "_" if item_eoat_id else "__"
            filename = f"{stem}{separator}{sequence:03d}{ext}"
            target = folder / filename
            if not target.exists() and target not in [item.target for item in plan]:
                break
            collision_avoided = True
            sequence += 1
        plan.append(
            PhotoPlanItem(
                source=source,
                target=target,
                photo_id=f"PHO-{date_taken.replace('-', '')}-{photo_sequence:03d}",
                collision_avoided=collision_avoided,
                view_type=item_view_type,
                description=item_data["description"],
                related_audit_id=item_data["related_audit_id"],
                related_issue_id=item_data["related_issue_id"],
                linked_audit_field=item_data["linked_audit_field"],
                eoat_assembly_id=item_eoat_id,
                tool_number=item_tool_number,
                part_name=item_part_name,
                original_filename=source.name,
                stored_relative_path=_relative_project_path(project_root, target),
            )
        )
        sequence += 1
        next_sequences[sequence_key] = sequence
        photo_sequence += 1
    return plan


def intake_photos(
    project_root: str | Path,
    photo_paths: list[str | Path],
    plant_area: str,
    press_machine: str,
    date_taken: str,
    view_type: str,
    related_audit_id: str = "",
    related_issue_id: str = "",
    description: str = "",
    notes: str = "",
    linked_audit_field: str = "",
    tool_number: str = "",
    part_name: str = "",
    eoat_assembly_id: str = "",
    per_photo_metadata: list[Mapping[str, Any]] | None = None,
    copy_mode: bool = True,
    log_activity: bool = True,
) -> ToolResult:
    started = time.perf_counter()
    paths = resolve_project_paths(project_root)
    workbook_path = paths.master_workbook
    if not workbook_path.exists():
        return ToolResult.fail(
            "photo_intake",
            "EOAT Photo Intake and Renaming Tool",
            "Master workbook is missing.",
            errors=[str(workbook_path)],
        )
    if not photo_paths:
        return ToolResult.fail("photo_intake", "EOAT Photo Intake and Renaming Tool", "No photos selected.")
    for field_name, value in {
        "Plant/Area": plant_area,
        "Date Taken": date_taken,
    }.items():
        if not str(value).strip():
            return ToolResult.fail(
                "photo_intake", "EOAT Photo Intake and Renaming Tool", f"Missing required field: {field_name}"
            )

    metadata = _photo_items(
        photo_paths,
        view_type=view_type,
        related_audit_id=related_audit_id,
        related_issue_id=related_issue_id,
        description=description,
        linked_audit_field=linked_audit_field,
        eoat_assembly_id=eoat_assembly_id,
        tool_number=tool_number,
        part_name=part_name,
        per_photo_metadata=per_photo_metadata,
    )
    invalid_eoat_ids = [
        str(item.get("eoat_assembly_id") or "").strip()
        for item in metadata
        if str(item.get("eoat_assembly_id") or "").strip()
        and not is_valid_eoat_assembly_id(item.get("eoat_assembly_id"))
    ]
    if invalid_eoat_ids:
        return ToolResult.fail(
            "photo_intake",
            "EOAT Photo Intake and Renaming Tool",
            "Invalid EOAT Assembly ID format.",
            errors=sorted(set(invalid_eoat_ids)),
        )
    if not all(
        str(item.get("eoat_assembly_id") or item.get("tool_number") or "").strip() for item in metadata
    ):
        return ToolResult.fail(
            "photo_intake",
            "EOAT Photo Intake and Renaming Tool",
            "Missing required field: EOAT Assembly ID or Tool #",
        )
    if any(not item["view_type"] for item in metadata):
        return ToolResult.fail(
            "photo_intake", "EOAT Photo Intake and Renaming Tool", "Missing required field: EOAT Area Shown"
        )
    selected_photo_paths = [item["source"] for item in metadata]
    plan = preview_photo_intake(
        project_root,
        selected_photo_paths,
        plant_area,
        press_machine,
        date_taken,
        view_type,
        tool_number=tool_number,
        part_name=part_name,
        related_audit_id=related_audit_id,
        per_photo_metadata=metadata,
        eoat_assembly_id=eoat_assembly_id,
    )
    if not plan:
        return ToolResult.fail(
            "photo_intake", "EOAT Photo Intake and Renaming Tool", "No supported image files selected."
        )
    missing = [str(item.source) for item in plan if not item.source.exists()]
    if missing:
        return ToolResult.fail(
            "photo_intake", "EOAT Photo Intake and Renaming Tool", "Some selected photos are missing.", errors=missing
        )

    workbook = None
    moved_or_copied: list[str] = []
    created_destination_dirs: list[Path] = []
    intake_completed = False
    try:
        for item in plan:
            if item.target.exists():
                raise FileExistsError(f"Target already exists: {item.target}")

        backup = backup_file(workbook_path, workbook_path.parent / "_backups")
        for item in plan:
            root_path = (
                eoat_photo_root(project_root, item.eoat_assembly_id)
                if item.eoat_assembly_id
                else tool_photo_root(project_root, item.tool_number, item.part_name)
            )
            target_folder = destination_folder(
                project_root, item.view_type, item.tool_number, item.part_name, item.eoat_assembly_id
            )
            root_existed = root_path.exists()
            folder_existed = target_folder.exists()
            if item.eoat_assembly_id:
                ensure_eoat_photo_category_folder(project_root, item.eoat_assembly_id, item.view_type)
            else:
                ensure_tool_photo_category_folder(project_root, item.tool_number, item.part_name, item.view_type)
            if not folder_existed:
                created_destination_dirs.append(target_folder)
            if not root_existed:
                created_destination_dirs.append(root_path)
            if copy_mode:
                _copy_photo_file(item.source, item.target)
            else:
                _move_photo_file(item.source, item.target)
            moved_or_copied.append(str(item.target))

        workbook = load_workbook(workbook_path)
        if "Photo Index" not in workbook.sheetnames:
            raise ValueError("Photo Index sheet is missing.")
        ws = workbook["Photo Index"]
        _ensure_headers(ws, get_expected_headers("Photo Index"))
        audit_rows_by_id = _audit_rows_by_id(workbook)
        rows_written: list[int] = []
        rows = row_dicts(workbook_path, "Photo Index")
        next_photo_number = 1
        prefix = f"PHO-{date_taken.replace('-', '')}-"
        imported_at = datetime.now().isoformat(timespec="seconds")
        for row in rows:
            value = str(row.get("Photo ID") or "")
            if value.startswith(prefix):
                try:
                    next_photo_number = max(next_photo_number, int(value.rsplit("-", 1)[1]) + 1)
                except ValueError:
                    continue
        for item in plan:
            photo_id = f"{prefix}{next_photo_number:03d}"
            next_photo_number += 1
            row_number = next_empty_row(ws)
            data = {header: "" for header in get_expected_headers("Photo Index")}
            audit_row = audit_rows_by_id.get(_text(item.related_audit_id).casefold(), {})
            row_eoat_assembly_id = item.eoat_assembly_id or normalize_eoat_assembly_id(
                audit_row.get(EOAT_ASSEMBLY_ID_FIELD)
            )
            row_part_name = item.part_name or _text(audit_row.get("Part Name/Description")) or _text(
                audit_row.get("Part Family")
            )
            stored_relative_path = _relative_project_path(project_root, item.target)
            stored_folder = _relative_project_path(project_root, item.target.parent)
            data.update(
                {
                    "Photo ID": photo_id,
                    "Date Taken": date_taken,
                    "Plant/Area": plant_area,
                    "Press/Machine #": press_machine,
                    EOAT_ASSEMBLY_ID_FIELD: row_eoat_assembly_id,
                    TOOL_FIELD: item.tool_number or _text(audit_row.get(TOOL_FIELD)),
                    "Part Name": row_part_name,
                    "Photo Type": item.view_type,
                    "EOAT Area Shown": item.view_type,
                    "Original Filename": item.original_filename or item.source.name,
                    "Stored Filename": item.target.name,
                    "Stored Relative Path": stored_relative_path,
                    "Imported At": imported_at,
                    "Photo Filename": item.target.name,
                    "Folder Path": stored_folder,
                    "Description": item.description,
                    "Related Audit ID": item.related_audit_id,
                    "Related Issue ID": item.related_issue_id,
                    LINKED_AUDIT_FIELD_HEADER: item.linked_audit_field,
                    "Notes": notes,
                }
            )
            write_row_by_headers(ws, row_number, data)
            rows_written.append(row_number)
        audit_update_details, audit_update_warnings, updated_audit_rows = _update_related_audit_rows(workbook, plan)
        workbook.save(workbook_path)
        workbook.close()
        invalidate_workbook_cache(workbook_path)
        metadata_warnings = _update_eoat_info_files(project_root, {item.eoat_assembly_id for item in plan})
        intake_completed = True
    except Exception as exc:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        return ToolResult.fail(
            "photo_intake",
            "EOAT Photo Intake and Renaming Tool",
            "Photo intake failed.",
            errors=[str(exc)],
            files_created=moved_or_copied,
            duration_seconds=time.perf_counter() - started,
        )
    finally:
        if not intake_completed:
            _remove_empty_directories(created_destination_dirs)

    result = ToolResult.ok(
        "photo_intake",
        "EOAT Photo Intake and Renaming Tool",
        _intake_success_summary(plan, copy_mode),
        details=[f"{item.source} -> {item.target}" for item in plan]
        + [f"Workbook backup: {backup}"]
        + audit_update_details,
        warnings=(
            ["Filename collision avoided for one or more photos."] if any(item.collision_avoided for item in plan) else []
        )
        + audit_update_warnings
        + metadata_warnings,
        files_created=[str(backup), *moved_or_copied],
        files_modified=[str(workbook_path)],
        metrics={"photo_count": len(plan), "copy_mode": copy_mode, "audit_rows_updated": updated_audit_rows},
        duration_seconds=time.perf_counter() - started,
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def repair_audit_photo_ties(project_root: str | Path, *, log_activity: bool = True) -> ToolResult:
    started = time.perf_counter()
    paths = resolve_project_paths(project_root)
    workbook_path = paths.master_workbook
    if not workbook_path.exists():
        return ToolResult.fail(
            "photo_repair_audit_ties",
            "Repair Audit Photo Ties",
            "Master workbook is missing.",
            errors=[str(workbook_path)],
        )

    workbook = None
    backup: Path | None = None
    details: list[str] = []
    warnings: list[str] = []
    files_indexed = _stored_photo_file_index(project_root)
    indexed_existing_paths: set[Path] = set()
    grouped_references: dict[str, set[str]] = {}
    photo_rows_checked = 0
    photo_rows_repaired = 0
    unresolved_rows = 0
    changed = False
    try:
        workbook = load_workbook(workbook_path)
        if "Photo Index" not in workbook.sheetnames:
            raise ValueError("Photo Index sheet is missing.")
        if "EOAT Inventory" not in workbook.sheetnames:
            raise ValueError("EOAT Inventory sheet is missing.")

        photo_ws = workbook["Photo Index"]
        added_photo_headers = _ensure_headers(photo_ws, get_expected_headers("Photo Index"))
        changed = bool(added_photo_headers)
        photo_headers = worksheet_headers(photo_ws)
        audit_rows = _audit_rows_list(workbook)
        eoat_contexts = build_eoat_assembly_contexts(audit_rows)

        for row_number in range(2, photo_ws.max_row + 1):
            row = _worksheet_row_dict(photo_ws, photo_headers, row_number)
            if not _photo_index_row_has_content(row):
                continue
            photo_rows_checked += 1
            resolved_path = _resolve_repair_photo_path(project_root, row, files_indexed)
            if resolved_path is None:
                unresolved_rows += 1
                photo_label = _text(row.get("Photo ID")) or f"row {row_number}"
                warnings.append(f"{photo_label}: stored photo file could not be found.")
                continue

            indexed_existing_paths.add(_normalized_path(resolved_path))
            stored_relative_path = _relative_project_path(project_root, resolved_path)
            folder_relative_path = _relative_project_path(project_root, resolved_path.parent)
            row_changed = False
            for header, value in {
                "Stored Relative Path": stored_relative_path,
                "Stored Filename": resolved_path.name,
                "Photo Filename": resolved_path.name,
                "Folder Path": folder_relative_path,
            }.items():
                row_changed = (
                    _set_cell_by_header_if_different(photo_ws, photo_headers, row_number, header, value) or row_changed
                )
            if row_changed:
                changed = True
                photo_rows_repaired += 1
                details.append(f"Repaired Photo Index row {row_number}: {stored_relative_path}")

            audit_ids = _related_audit_ids_for_photo_row(row, audit_rows, eoat_contexts)
            if audit_ids:
                for audit_id in audit_ids:
                    grouped_references.setdefault(audit_id, set()).add(folder_relative_path)
            else:
                warnings.append(
                    f"Photo Index row {row_number} has a stored photo but no Related Audit ID or EOAT match."
                )

        audit_details, audit_warnings, audit_rows_updated = _apply_audit_photo_references(
            workbook,
            grouped_references,
            project_root=project_root,
            repair_broken_links=True,
        )
        details.extend(audit_details)
        warnings.extend(audit_warnings)
        changed = changed or bool(audit_details)

        unindexed_files = _unindexed_stored_photo_files(files_indexed, indexed_existing_paths)
        if unindexed_files:
            warnings.append(
                f"Found {len(unindexed_files)} stored photo file(s) without Photo Index rows; "
                "audit ties can only be repaired automatically for indexed photos."
            )
            details.extend(f"Unindexed stored photo: {_relative_project_path(project_root, path)}" for path in unindexed_files[:10])

        if changed:
            backup = backup_file(workbook_path, workbook_path.parent / "_backups")
            workbook.save(workbook_path)
            invalidate_workbook_cache(workbook_path)
    except Exception as exc:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        return ToolResult.fail(
            "photo_repair_audit_ties",
            "Repair Audit Photo Ties",
            "Photo tie repair failed.",
            errors=[str(exc)],
            duration_seconds=time.perf_counter() - started,
        )
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass

    if backup is not None:
        details.append(f"Workbook backup: {backup}")
    summary = (
        f"Repaired photo ties for {audit_rows_updated} audit row(s); checked {photo_rows_checked} Photo Index row(s)."
        if changed
        else f"Photo ties already matched existing indexed photos. Checked {photo_rows_checked} Photo Index row(s)."
    )
    result = ToolResult.ok(
        "photo_repair_audit_ties",
        "Repair Audit Photo Ties",
        summary,
        details=details,
        warnings=warnings,
        files_created=[str(backup)] if backup is not None else [],
        files_modified=[str(workbook_path)] if changed else [],
        metrics={
            "photo_rows_checked": photo_rows_checked,
            "photo_rows_repaired": photo_rows_repaired,
            "audit_rows_updated": audit_rows_updated,
            "unresolved_photo_rows": unresolved_rows,
            "unindexed_stored_photos": len(unindexed_files),
        },
        duration_seconds=time.perf_counter() - started,
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def repair_photo_eoat_links(project_root: str | Path, *, log_activity: bool = True) -> ToolResult:
    started = time.perf_counter()
    paths = resolve_project_paths(project_root)
    workbook_path = paths.master_workbook
    if not workbook_path.exists():
        return ToolResult.fail(
            "repair_photo_eoat_links",
            "Repair Photo EOAT Links",
            "Master workbook is missing.",
            errors=[str(workbook_path)],
        )

    workbook = None
    repaired = 0
    skipped_ambiguous = 0
    skipped_no_match = 0
    details: list[str] = []
    warnings: list[str] = []
    try:
        workbook = load_workbook(workbook_path)
        if "Photo Index" not in workbook.sheetnames:
            raise ValueError("Photo Index sheet is missing.")
        if "EOAT Inventory" not in workbook.sheetnames:
            raise ValueError("EOAT Inventory sheet is missing.")
        photo_ws = workbook["Photo Index"]
        _ensure_headers(photo_ws, get_expected_headers("Photo Index"))
        photo_headers = worksheet_headers(photo_ws)
        eoat_col = photo_headers.index(EOAT_ASSEMBLY_ID_FIELD) + 1
        audit_rows = _audit_rows_list(workbook)
        changed = False
        for row_number in range(2, photo_ws.max_row + 1):
            row = _worksheet_row_dict(photo_ws, photo_headers, row_number)
            if not _photo_index_row_has_content(row):
                continue
            if normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD)):
                continue
            inferred, reason = infer_eoat_assembly_id_for_photo_row(row, audit_rows)
            if inferred:
                photo_ws.cell(row=row_number, column=eoat_col).value = inferred
                repaired += 1
                changed = True
                details.append(f"Photo Index row {row_number}: set {EOAT_ASSEMBLY_ID_FIELD} to {inferred}.")
            elif reason == "ambiguous":
                skipped_ambiguous += 1
                warnings.append(f"Skipped ambiguous Photo Index row {row_number}; Tool # maps to multiple EOAT IDs.")
            else:
                skipped_no_match += 1
        backup = backup_file(workbook_path, workbook_path.parent / "_backups") if changed else None
        if changed:
            workbook.save(workbook_path)
            invalidate_workbook_cache(workbook_path)
        workbook.close()
        workbook = None
    except Exception as exc:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        return ToolResult.fail(
            "repair_photo_eoat_links",
            "Repair Photo EOAT Links",
            "Photo EOAT link repair failed.",
            errors=[str(exc)],
            duration_seconds=time.perf_counter() - started,
        )

    summary = (
        f"Repaired {repaired} photo EOAT link(s)."
        if repaired
        else "No missing photo EOAT links could be repaired automatically."
    )
    if skipped_ambiguous:
        summary += f" Skipped {skipped_ambiguous} ambiguous row(s)."
    if skipped_no_match:
        summary += f" Skipped {skipped_no_match} row(s) with no matching audit/tool data."
    result = ToolResult.ok(
        "repair_photo_eoat_links",
        "Repair Photo EOAT Links",
        summary,
        details=details + ([f"Workbook backup: {backup}"] if backup is not None else []),
        warnings=warnings,
        files_created=[str(backup)] if backup is not None else [],
        files_modified=[str(workbook_path)] if repaired else [],
        metrics={
            "rows_repaired": repaired,
            "skipped_ambiguous": skipped_ambiguous,
            "skipped_no_match": skipped_no_match,
        },
        duration_seconds=time.perf_counter() - started,
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def convert_legacy_photo_tree_to_eoat_folders(
    project_root: str | Path,
    *,
    move_files: bool = False,
    log_activity: bool = True,
) -> ToolResult:
    started = time.perf_counter()
    paths = resolve_project_paths(project_root)
    workbook_path = paths.master_workbook
    if not workbook_path.exists():
        return ToolResult.fail(
            "convert_photo_tree_to_eoat",
            "Convert Photo Tree to EOAT Folders",
            "Master workbook is missing.",
            errors=[str(workbook_path)],
        )

    workbook = None
    copied_or_moved: list[str] = []
    details: list[str] = []
    warnings: list[str] = []
    converted_rows = 0
    repaired_links = 0
    skipped_ambiguous = 0
    skipped_no_match = 0
    skipped_missing_file = 0
    try:
        workbook = load_workbook(workbook_path)
        if "Photo Index" not in workbook.sheetnames:
            raise ValueError("Photo Index sheet is missing.")
        if "EOAT Inventory" not in workbook.sheetnames:
            raise ValueError("EOAT Inventory sheet is missing.")
        photo_ws = workbook["Photo Index"]
        _ensure_headers(photo_ws, get_expected_headers("Photo Index"))
        headers = worksheet_headers(photo_ws)
        positions = {header: index + 1 for index, header in enumerate(headers)}
        audit_rows = _audit_rows_list(workbook)
        backup = backup_file(workbook_path, workbook_path.parent / "_backups")
        changed = False
        next_sequences: dict[tuple[str, str], int] = {}
        planned_targets: set[Path] = set()
        updated_eoat_ids: set[str] = set()
        files_indexed = _stored_photo_file_index(project_root)
        eoat_contexts = build_eoat_assembly_contexts(audit_rows)
        grouped_references: dict[str, set[str]] = {}

        for row_number in range(2, photo_ws.max_row + 1):
            row = _worksheet_row_dict(photo_ws, headers, row_number)
            if not _photo_index_row_has_content(row):
                continue
            eoat_id = normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD))
            if not eoat_id:
                eoat_id, reason = infer_eoat_assembly_id_for_photo_row(row, audit_rows)
                if eoat_id:
                    photo_ws.cell(row=row_number, column=positions[EOAT_ASSEMBLY_ID_FIELD]).value = eoat_id
                    repaired_links += 1
                    changed = True
                elif reason == "ambiguous":
                    skipped_ambiguous += 1
                    warnings.append(f"Skipped ambiguous Photo Index row {row_number}; Tool # maps to multiple EOAT IDs.")
                    continue
                else:
                    skipped_no_match += 1
                    continue
            if not is_valid_eoat_assembly_id(eoat_id):
                warnings.append(f"Skipped Photo Index row {row_number}; invalid EOAT Assembly ID: {eoat_id}.")
                continue

            source = _resolve_repair_photo_path(project_root, row, files_indexed)
            if source is None or not source.exists():
                skipped_missing_file += 1
                warnings.append(f"Skipped Photo Index row {row_number}; could not resolve existing photo file.")
                continue
            try:
                source.relative_to(eoat_photo_root(project_root, eoat_id))
                updated_eoat_ids.add(eoat_id)
                folder_reference = _relative_project_path(project_root, source.parent)
                row_for_links = {**row, EOAT_ASSEMBLY_ID_FIELD: eoat_id}
                for audit_id in _related_audit_ids_for_photo_row(row_for_links, audit_rows, eoat_contexts):
                    grouped_references.setdefault(audit_id, set()).add(folder_reference)
                continue
            except ValueError:
                pass

            view_type = _text(row.get("EOAT Area Shown")) or _text(row.get("Photo Type")) or "Overall"
            date_taken = _text(row.get("Date Taken")) or date.today().isoformat()
            target_folder = ensure_eoat_photo_category_folder(project_root, eoat_id, view_type)
            stem = build_photo_filename_stem("", "", "", date_taken, view_type, eoat_id)
            sequence_key = (eoat_id, view_type)
            if sequence_key not in next_sequences:
                next_sequences[sequence_key] = _next_existing_photo_sequence(
                    project_root,
                    date_taken,
                    "",
                    "",
                    view_type,
                    eoat_assembly_id=eoat_id,
                )
            sequence = next_sequences[sequence_key]
            while True:
                target = target_folder / f"{stem}_{sequence:03d}{source.suffix}"
                if not target.exists() and target not in planned_targets:
                    break
                sequence += 1
            next_sequences[sequence_key] = sequence + 1
            planned_targets.add(target)
            if move_files:
                _move_photo_file(source, target)
            else:
                _copy_photo_file(source, target)
            copied_or_moved.append(str(target))
            stored_relative_path = _relative_project_path(project_root, target)
            stored_folder = _relative_project_path(project_root, target.parent)
            for header, value in {
                "Stored Relative Path": stored_relative_path,
                "Folder Path": stored_folder,
                "Stored Filename": target.name,
                "Photo Filename": target.name,
            }.items():
                if header in positions:
                    photo_ws.cell(row=row_number, column=positions[header]).value = value
            updated_eoat_ids.add(eoat_id)
            row_for_links = {**row, EOAT_ASSEMBLY_ID_FIELD: eoat_id}
            for audit_id in _related_audit_ids_for_photo_row(row_for_links, audit_rows, eoat_contexts):
                grouped_references.setdefault(audit_id, set()).add(stored_folder)
            converted_rows += 1
            changed = True
            details.append(f"Photo Index row {row_number}: {source} -> {target}")

        audit_details, audit_warnings, audit_rows_updated = _apply_audit_photo_references(
            workbook,
            grouped_references,
            project_root=project_root,
            repair_broken_links=False,
        )
        details.extend(audit_details)
        warnings.extend(audit_warnings)
        changed = changed or bool(audit_details)
        if changed:
            workbook.save(workbook_path)
            invalidate_workbook_cache(workbook_path)
        workbook.close()
        workbook = None
        warnings.extend(_update_eoat_info_files(project_root, updated_eoat_ids))
    except Exception as exc:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        return ToolResult.fail(
            "convert_photo_tree_to_eoat",
            "Convert Photo Tree to EOAT Folders",
            "Photo tree conversion failed.",
            errors=[str(exc)],
            files_created=copied_or_moved,
            duration_seconds=time.perf_counter() - started,
        )

    verb = "Moved" if move_files else "Copied"
    summary = (
        f"{verb} {converted_rows} indexed legacy photo(s) into EOAT Assembly ID folders; "
        f"repaired {repaired_links} missing EOAT link(s)."
    )
    if skipped_ambiguous or skipped_no_match or skipped_missing_file:
        summary += (
            f" Skipped {skipped_ambiguous} ambiguous, {skipped_no_match} without EOAT match, "
            f"{skipped_missing_file} missing file row(s)."
        )
    result = ToolResult.ok(
        "convert_photo_tree_to_eoat",
        "Convert Photo Tree to EOAT Folders",
        summary,
        details=[f"Workbook backup: {backup}", *details],
        warnings=warnings,
        files_created=[str(backup), *copied_or_moved],
        files_modified=[str(workbook_path)] if converted_rows or repaired_links else [],
        metrics={
            "converted_rows": converted_rows,
            "repaired_eoat_links": repaired_links,
            "audit_rows_updated": audit_rows_updated,
            "skipped_ambiguous": skipped_ambiguous,
            "skipped_no_match": skipped_no_match,
            "skipped_missing_file": skipped_missing_file,
            "move_files": move_files,
        },
        duration_seconds=time.perf_counter() - started,
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def _photo_items(
    photo_paths: list[str | Path],
    *,
    view_type: str,
    related_audit_id: str = "",
    related_issue_id: str = "",
    description: str = "",
    linked_audit_field: str = "",
    eoat_assembly_id: str = "",
    tool_number: str = "",
    part_name: str = "",
    per_photo_metadata: list[Mapping[str, Any]] | None = None,
) -> list[dict[str, str]]:
    items: list[dict[str, str]] = []
    for index, source_raw in enumerate(photo_paths):
        override = per_photo_metadata[index] if per_photo_metadata and index < len(per_photo_metadata) else {}
        include = override.get("include", True)
        if isinstance(include, str) and include.strip().casefold() in {"false", "0", "no"}:
            continue
        if include is False:
            continue
        source = Path(override.get("source") or override.get("path") or source_raw)
        items.append(
            {
                "source": str(source),
                "view_type": _text(override.get("view_type") or override.get("EOAT Area Shown") or view_type),
                "related_audit_id": _text(
                    override.get("related_audit_id") or override.get("Related Audit ID") or related_audit_id
                ),
                "related_issue_id": _text(
                    override.get("related_issue_id") or override.get("Related Issue ID") or related_issue_id
                ),
                "description": _text(override.get("description") or override.get("Description") or description),
                "linked_audit_field": _text(
                    override.get("linked_audit_field")
                    or override.get(LINKED_AUDIT_FIELD_HEADER)
                    or linked_audit_field
                ),
                "eoat_assembly_id": normalize_eoat_assembly_id(
                    override.get("eoat_assembly_id")
                    or override.get(EOAT_ASSEMBLY_ID_FIELD)
                    or eoat_assembly_id
                ),
                "tool_number": _text(override.get("tool_number") or override.get(TOOL_FIELD) or tool_number),
                "part_name": _text(
                    override.get("part_name")
                    or override.get("Part Name")
                    or override.get("Part Name/Description")
                    or part_name
                ),
            }
        )
    return items


def _ensure_headers(ws, expected_headers: list[str]) -> list[str]:
    headers = worksheet_headers(ws)
    added: list[str] = []
    for header in expected_headers:
        if header in headers:
            continue
        ws.cell(row=1, column=len(headers) + 1).value = header
        headers.append(header)
        added.append(header)
    return added


def _audit_rows_by_id(workbook) -> dict[str, dict[str, object]]:
    if "EOAT Inventory" not in workbook.sheetnames:
        return {}
    ws = workbook["EOAT Inventory"]
    headers = worksheet_headers(ws)
    positions = {header: index for index, header in enumerate(headers)}
    rows: dict[str, dict[str, object]] = {}
    if "Audit ID" not in positions:
        return rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = {header: row[index] for header, index in positions.items() if index < len(row)}
        audit_id = _text(data.get("Audit ID")).casefold()
        if audit_id:
            rows[audit_id] = data
    return rows


def _audit_rows_list(workbook) -> list[dict[str, object]]:
    if "EOAT Inventory" not in workbook.sheetnames:
        return []
    ws = workbook["EOAT Inventory"]
    headers = worksheet_headers(ws)
    rows: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = {header: row[index] for index, header in enumerate(headers) if index < len(row)}
        if any(_text(value) for value in data.values()):
            rows.append(data)
    return rows


def _part_name_for_audit(project_root: str | Path, audit_id: str) -> str:
    audit_id = _text(audit_id).casefold()
    if not audit_id:
        return ""
    workbook_path = resolve_project_paths(project_root).master_workbook
    if not workbook_path.exists():
        return ""
    try:
        rows = row_dicts(workbook_path, "EOAT Inventory")
    except Exception:
        return ""
    for row in rows:
        if _text(row.get("Audit ID")).casefold() == audit_id:
            return _text(row.get("Part Name/Description")) or _text(row.get("Part Family"))
    return ""


def _eoat_assembly_id_for_audit(project_root: str | Path, audit_id: str) -> str:
    audit_id = _text(audit_id).casefold()
    if not audit_id:
        return ""
    workbook_path = resolve_project_paths(project_root).master_workbook
    if not workbook_path.exists():
        return ""
    try:
        rows = row_dicts(workbook_path, "EOAT Inventory")
    except Exception:
        return ""
    for row in rows:
        if _text(row.get("Audit ID")).casefold() == audit_id:
            return normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD))
    return ""


def _update_eoat_info_files(project_root: str | Path, eoat_ids: set[str]) -> list[str]:
    warnings: list[str] = []
    for eoat_id in sorted({normalize_eoat_assembly_id(value) for value in eoat_ids if normalize_eoat_assembly_id(value)}):
        try:
            update_eoat_info_file(project_root, eoat_id)
        except Exception as exc:
            warnings.append(f"Could not update eoat_info.json for {eoat_id}: {exc}")
    return warnings


def _relative_project_path(project_root: str | Path, path: str | Path) -> str:
    root = Path(project_root).resolve()
    target = Path(path).resolve()
    try:
        return str(target.relative_to(root))
    except ValueError:
        return str(path)


def _intake_success_summary(plan: list[PhotoPlanItem], copy_mode: bool) -> str:
    verb = "Copied" if copy_mode else "Moved"
    folders = sorted({str(Path(item.stored_relative_path).parent) for item in plan if item.stored_relative_path})
    if any(item.eoat_assembly_id for item in plan):
        if len(folders) == 1:
            return f"{verb} and indexed {len(plan)} photo(s) into {folders[0]}."
        return f"{verb} and indexed {len(plan)} photo(s) into {len(folders)} EOAT photo folder(s)."
    if len(folders) == 1:
        return f"{verb} and indexed {len(plan)} photo(s) into {folders[0]}."
    return f"{verb} and indexed {len(plan)} photo(s) into {len(folders)} tool photo folder(s)."


def _remove_empty_directories(paths: list[Path]) -> None:
    for path in sorted({Path(item) for item in paths}, key=lambda item: len(item.parts), reverse=True):
        try:
            if path.exists() and path.is_dir() and not any(path.iterdir()):
                path.rmdir()
        except OSError:
            continue


def _copy_photo_file(source: Path, target: Path) -> Path:
    try:
        return safe_copy_file(source, target, overwrite=False)
    except FileNotFoundError as exc:
        if not _is_windows_path_length_error(exc, source, target):
            raise
        shutil.copy2(_windows_long_path(source), _windows_long_path(target))
        return target


def _move_photo_file(source: Path, target: Path) -> Path:
    try:
        shutil.move(str(source), str(target))
        return target
    except FileNotFoundError as exc:
        if not _is_windows_path_length_error(exc, source, target):
            raise
        shutil.copy2(_windows_long_path(source), _windows_long_path(target))
        Path(source).unlink()
        return target


def _is_windows_path_length_error(exc: FileNotFoundError, source: Path, target: Path) -> bool:
    return os.name == "nt" and getattr(exc, "winerror", None) == 3 and max(len(str(source)), len(str(target))) >= 240


def _windows_long_path(path: str | Path) -> str:
    text = str(Path(path).resolve())
    if os.name != "nt" or text.startswith("\\\\?\\"):
        return text
    if text.startswith("\\\\"):
        return "\\\\?\\UNC\\" + text.lstrip("\\")
    return "\\\\?\\" + text


def _worksheet_row_dict(ws, headers: list[str], row_number: int) -> dict[str, object]:
    return {header: ws.cell(row=row_number, column=index + 1).value for index, header in enumerate(headers)}


def _photo_index_row_has_content(row: Mapping[str, object]) -> bool:
    return any(
        _text(row.get(header))
        for header in (
            "Photo ID",
            "Stored Relative Path",
            "Stored Filename",
            "Photo Filename",
            "Folder Path",
            "Related Audit ID",
        )
    )


def _set_cell_by_header_if_different(
    ws,
    headers: list[str],
    row_number: int,
    header: str,
    value: str,
) -> bool:
    if header not in headers:
        return False
    cell = ws.cell(row=row_number, column=headers.index(header) + 1)
    if _text(cell.value) == value:
        return False
    cell.value = value
    return True


def _stored_photo_file_index(project_root: str | Path) -> dict[str, list[Path]]:
    paths = resolve_project_paths(project_root)
    root = paths.cell_photos
    incoming = _normalized_path(paths.incoming_photos)
    files: dict[str, list[Path]] = {}
    if not root.exists():
        return files
    for dirpath, _dirnames, filenames in os.walk(root):
        folder = Path(dirpath)
        try:
            if _path_is_relative_to(_normalized_path(folder), incoming):
                continue
        except OSError:
            continue
        for filename in filenames:
            path = folder / filename
            if path.suffix.lower() not in SUPPORTED_IMAGE_EXTENSIONS:
                continue
            files.setdefault(path.name.casefold(), []).append(path)
    return files


def _unindexed_stored_photo_files(
    files_indexed: Mapping[str, list[Path]],
    indexed_existing_paths: set[Path],
) -> list[Path]:
    unindexed: list[Path] = []
    for paths in files_indexed.values():
        for path in paths:
            normalized = _normalized_path(path)
            if normalized not in indexed_existing_paths:
                unindexed.append(path)
    return sorted(unindexed)


def _resolve_repair_photo_path(
    project_root: str | Path,
    row: Mapping[str, object],
    files_indexed: Mapping[str, list[Path]],
) -> Path | None:
    for candidate in _photo_path_candidates(project_root, row):
        if candidate.exists() and candidate.is_file():
            return candidate

    for header in ("Stored Filename", "Photo Filename", "Original Filename"):
        filename = Path(_text(row.get(header))).name
        if not filename:
            continue
        match = _pick_unique_photo_match(files_indexed.get(filename.casefold(), []), row)
        if match is not None:
            return match
    return None


def _photo_path_candidates(project_root: str | Path, row: Mapping[str, object]) -> list[Path]:
    candidates: list[Path] = []
    stored_relative_path = _text(row.get("Stored Relative Path"))
    if stored_relative_path:
        candidates.append(_path_from_reference(project_root, stored_relative_path))

    folder_path = _text(row.get("Folder Path"))
    for filename_header in ("Stored Filename", "Photo Filename"):
        filename = Path(_text(row.get(filename_header))).name
        if folder_path and filename:
            folder = _path_from_reference(project_root, folder_path)
            candidates.append(folder / filename)
    return _dedupe_paths(candidates)


def _pick_unique_photo_match(paths: list[Path], row: Mapping[str, object]) -> Path | None:
    if len(paths) == 1:
        return paths[0]
    if not paths:
        return None

    filtered = paths
    tool_number = _text(row.get(TOOL_FIELD))
    if tool_number:
        tool_prefix = f"tool_{sanitize_folder_part(tool_number).casefold()}"
        tool_filtered = [
            path
            for path in filtered
            if any(part.casefold().startswith(tool_prefix) for part in path.parts)
        ]
        if tool_filtered:
            filtered = tool_filtered

    view_type = _text(row.get("EOAT Area Shown")) or _text(row.get("Photo Type"))
    if view_type:
        folder_name = photo_category_folder(view_type).casefold()
        category_filtered = [path for path in filtered if path.parent.name.casefold() == folder_name]
        if category_filtered:
            filtered = category_filtered

    return filtered[0] if len(filtered) == 1 else None


def _dedupe_paths(paths: list[Path]) -> list[Path]:
    seen: set[str] = set()
    unique: list[Path] = []
    for path in paths:
        key = str(path).casefold()
        if key in seen:
            continue
        seen.add(key)
        unique.append(path)
    return unique


def _normalized_path(path: str | Path) -> Path:
    return Path(path).resolve(strict=False)


def _path_is_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
        return True
    except ValueError:
        return False


def _path_from_reference(project_root: str | Path, reference: str | Path) -> Path:
    text = _text(reference).strip("\"'")
    if text.casefold().startswith("file://"):
        text = text[7:]
    candidate = Path(text)
    if candidate.is_absolute():
        return candidate
    return Path(project_root) / candidate


def _photos_taken_header(headers: list[str]) -> str:
    if "Photos Taken?" in headers:
        return "Photos Taken?"
    if "Pictures Taken?" in headers:
        return "Pictures Taken?"
    return ""


def _is_local_photo_reference(project_root: str | Path, reference: str) -> bool:
    text = _text(reference)
    if not text or "://" in text and not text.casefold().startswith("file://"):
        return False
    if "cell_photos" in text.replace("\\", "/").casefold():
        return True
    candidate = _path_from_reference(project_root, text)
    paths = resolve_project_paths(project_root)
    return _path_is_relative_to(_normalized_path(candidate), _normalized_path(paths.cell_photos))


def _reference_exists(project_root: str | Path, reference: str) -> bool:
    candidate = _path_from_reference(project_root, reference)
    return candidate.exists()


def _dedupe_text_lines(lines: list[str]) -> list[str]:
    seen: set[str] = set()
    unique: list[str] = []
    for line in lines:
        text = _text(line)
        if not text:
            continue
        key = text.casefold()
        if key in seen:
            continue
        seen.add(key)
        unique.append(text)
    return unique


def _update_related_audit_rows(workbook, plan: list[PhotoPlanItem]) -> tuple[list[str], list[str], int]:
    grouped: dict[str, set[str]] = {}
    audit_rows = _audit_rows_list(workbook)
    eoat_contexts = build_eoat_assembly_contexts(audit_rows)
    for item in plan:
        reference = str(Path(item.stored_relative_path).parent)
        for audit_id in _related_audit_ids_for_plan_item(item, eoat_contexts):
            grouped.setdefault(audit_id, set()).add(reference)
    return _apply_audit_photo_references(workbook, grouped)


def _related_audit_ids_for_plan_item(item: PhotoPlanItem, eoat_contexts) -> list[str]:
    audit_ids: list[str] = []
    if _text(item.related_audit_id):
        audit_ids.append(_text(item.related_audit_id))
    eoat_id = normalize_eoat_assembly_id(item.eoat_assembly_id)
    context = eoat_contexts.get(eoat_id) if eoat_id else None
    if context is not None:
        audit_ids.extend(context.audit_ids)
    return _dedupe_text_lines(audit_ids)


def _related_audit_ids_for_photo_row(
    row: dict[str, Any],
    audit_rows: list[dict[str, object]],
    eoat_contexts,
) -> list[str]:
    audit_ids: list[str] = []
    if _text(row.get("Related Audit ID")):
        audit_ids.append(_text(row.get("Related Audit ID")))
    eoat_id = normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD))
    if not eoat_id:
        inferred, _reason = infer_eoat_assembly_id_for_photo_row(row, audit_rows)
        eoat_id = inferred
    context = eoat_contexts.get(eoat_id) if eoat_id else None
    if context is not None:
        audit_ids.extend(context.audit_ids)
    return _dedupe_text_lines(audit_ids)


def _apply_audit_photo_references(
    workbook,
    grouped: Mapping[str, set[str]],
    *,
    project_root: str | Path | None = None,
    repair_broken_links: bool = False,
) -> tuple[list[str], list[str], int]:
    if not grouped:
        return [], [], 0
    if "EOAT Inventory" not in workbook.sheetnames:
        return [], ["Related audit rows were not updated because EOAT Inventory is missing."], 0

    ws = workbook["EOAT Inventory"]
    headers = worksheet_headers(ws)
    photos_taken_header = _photos_taken_header(headers)
    required = {"Audit ID", "Photo Folder/Link"}
    missing = sorted(required - set(headers))
    if not photos_taken_header:
        missing.append("Photos Taken?")
    if missing:
        return [], [f"Related audit rows were not updated because EOAT Inventory is missing: {', '.join(missing)}"], 0

    audit_id_col = headers.index("Audit ID") + 1
    photos_taken_col = headers.index(photos_taken_header) + 1
    photo_link_col = headers.index("Photo Folder/Link") + 1
    details: list[str] = []
    warnings: list[str] = []
    updated = 0
    for audit_id, references in grouped.items():
        target_row = None
        for row_number in range(2, ws.max_row + 1):
            if _text(ws.cell(row=row_number, column=audit_id_col).value).casefold() == audit_id.casefold():
                target_row = row_number
                break
        if target_row is None:
            warnings.append(
                "Photo Index was updated, but no matching EOAT Inventory row was found "
                f"for Related Audit ID: {audit_id}"
            )
            continue
        changed = False
        if _text(ws.cell(row=target_row, column=photos_taken_col).value) != "Yes":
            ws.cell(row=target_row, column=photos_taken_col).value = "Yes"
            changed = True
        link_cell = ws.cell(row=target_row, column=photo_link_col)
        existing_lines = [_text(line) for line in _text(link_cell.value).splitlines() if _text(line)]
        kept_lines: list[str] = []
        removed_lines: list[str] = []
        for line in existing_lines:
            if repair_broken_links and project_root is not None and _is_local_photo_reference(project_root, line):
                if not _reference_exists(project_root, line):
                    removed_lines.append(line)
                    continue
            kept_lines.append(line)
        existing_folded = {line.casefold() for line in kept_lines}
        new_lines = [ref for ref in sorted(references) if ref and ref.casefold() not in existing_folded]
        repaired_lines = _dedupe_text_lines([*kept_lines, *new_lines])
        next_link_value = "\n".join(repaired_lines)
        if _text(link_cell.value) != next_link_value:
            link_cell.value = next_link_value
            changed = True
        if removed_lines:
            details.append(
                f"Removed {len(removed_lines)} broken local photo link(s) from EOAT Inventory row {target_row}."
            )
        if new_lines:
            details.append(
                f"Added {len(new_lines)} photo folder link(s) to EOAT Inventory row {target_row} for {audit_id}."
            )
        if changed:
            updated += 1
            details.append(f"Updated EOAT Inventory row {target_row} for {audit_id}: Photos Taken?=Yes.")
    return details, warnings, updated


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()
