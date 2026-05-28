from __future__ import annotations

import shutil
import time
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation

from .action_items import add_action_item
from .audit.history import append_audit_history
from .audit_constants import (
    COMPATIBILITY_SOURCE_FIELD,
    ENTRY_TYPE_AUDITED,
    ENTRY_TYPE_COMPATIBLE,
    ENTRY_TYPE_FIELD,
    SOURCE_AUDIT_ID_FIELD,
)
from .audit_by_press import refresh_audit_by_press_view
from . import audit_field_rules as field_rules
from .logging import log_tool_run
from .paths import resolve_project_paths
from .result import ToolResult
from .safe_files import backup_file
from .workbook_locks import detect_workbook_lock
from .gripper_fields import (
    CUP_COUNT_FIELD,
    GRIPPER_COUNT_FIELD,
    GRIPPER_MODEL_FIELD,
    GRIPPER_SIZE_FIELD,
    GRIPPER_TYPE_FIELD,
    GRIPPER_TYPE_VALUES,
    gripper_model_to_workbook,
)
from .tool_fields import LEGACY_TOOL_FIELD, TOOL_FIELD
from .workbook_io import find_row_by_value, next_empty_row, row_dicts, worksheet_headers, write_row_by_headers
from .workbook_schema import get_expected_headers

AUDIT_REQUIRED_FIELDS = [
    "Audit Date",
    "Auditor",
    "Plant/Area",
    "Press/Machine #",
    "Robot Type",
    "EOAT Type",
    "Status",
]

AUDIT_IMPORTANT_FIELDS = [
    "Part Family",
    "EOAT Moves",
    CUP_COUNT_FIELD,
    "Sensor Type",
    "Sensor Brand/Model",
    "Vacuum Confirmation Present?",
    "Part-Present Detection Present?",
    "Electrical Quick Disconnect Type",
    "Tubing Condition",
    "Cable Management Condition",
    "Known Issues",
    "Photos Taken?",
    "Priority",
]

CONNECTION_TYPE_FIELD = "Connection Type"
EOAT_MOVES_FIELD = "EOAT Moves"
NUMBER_OF_PARTS_PICKED_FIELD = "Number of Parts Picked"
LEGACY_VACUUM_CUPS_FIELD = "Number of Vacuum Cups"
EOAT_VACUUM_CIRCUITS_FIELD = "EOAT Vacuum Circuits"
EOAT_PRESSURE_CIRCUITS_FIELD = "EOAT Pressure Circuits"
EOAT_INTERCHANGEABLE_CIRCUITS_FIELD = "EOAT Interchangeable Circuits"
EOAT_PNEUMATIC_CIRCUIT_FIELDS = {
    EOAT_VACUUM_CIRCUITS_FIELD,
    EOAT_PRESSURE_CIRCUITS_FIELD,
    EOAT_INTERCHANGEABLE_CIRCUITS_FIELD,
}
NA_VALUE = "N/A"
UNKNOWN_NOT_CHECKED = "Unknown / Not Checked"
ELECTRICAL_WIRING_PRESENT_FIELD = field_rules.ELECTRICAL_WIRING_PRESENT_FIELD
SENSOR_TYPE_FIELD = "Sensor Type"
SENSOR_BRAND_MODEL_FIELD = "Sensor Brand/Model"
PART_PRESENT_DETECTION_FIELD = "Part-Present Detection Present?"
PART_PRESENT_SENSOR_DEFAULTS = {
    SENSOR_TYPE_FIELD: "Reed Switch",
    SENSOR_BRAND_MODEL_FIELD: "SMC",
}
VACUUM_ZONES_FIELD = "Vacuum Zones"
CONNECTION_TYPE_VALUES = ["ATI", "DoveTail", "Direct Mount", "Lever Lock"]
EOAT_MOVES_VALUES = ["Part", "Sprue", "Both"]
EOAT_TYPE_DROPDOWN_VALUES = ["Vacuum", "Mechanical / Gripper", "Hybrid", "Unknown / Needs Review", "Miscellaneous"]
CLEANROOM_DROPDOWN_VALUES = ["Cleanroom", "Non-Cleanroom", "Whiteroom", "Unknown / Not Checked"]
CLEANROOM_DEFAULT = "Whiteroom"
CUP_TYPE_DEFAULT = "Silicone"
TOOLING_COLUMN_ORDER = [
    "EOAT Type",
    EOAT_MOVES_FIELD,
    CONNECTION_TYPE_FIELD,
    NUMBER_OF_PARTS_PICKED_FIELD,
    GRIPPER_COUNT_FIELD,
    GRIPPER_TYPE_FIELD,
    GRIPPER_MODEL_FIELD,
    GRIPPER_SIZE_FIELD,
    CUP_COUNT_FIELD,
    "Cup Type/Material",
    "Cup Diameter/Size",
    "Vacuum Generator Type",
    EOAT_VACUUM_CIRCUITS_FIELD,
    EOAT_PRESSURE_CIRCUITS_FIELD,
    EOAT_INTERCHANGEABLE_CIRCUITS_FIELD,
]
VACUUM_TOOLING_FIELDS = {
    CUP_COUNT_FIELD,
    "Cup Type/Material",
    "Cup Diameter/Size",
    "Vacuum Generator Type",
}
GRIPPER_TOOLING_FIELDS = {GRIPPER_COUNT_FIELD, GRIPPER_TYPE_FIELD, GRIPPER_MODEL_FIELD, GRIPPER_SIZE_FIELD}
SENSOR_DETAIL_FIELDS = field_rules.SENSOR_DETAIL_FIELDS
ELECTRICAL_DETAIL_FIELDS = field_rules.ELECTRICAL_DETAIL_FIELDS
QUICK_DISCONNECT_DETAIL_FIELDS = field_rules.QUICK_DISCONNECT_DETAIL_FIELDS


@dataclass(frozen=True)
class AuditFieldMetadata:
    tags: frozenset[str] = field(default_factory=frozenset)
    default: str | None = None


AUDIT_FIELD_METADATA: dict[str, AuditFieldMetadata] = {
    SENSOR_TYPE_FIELD: AuditFieldMetadata(frozenset({"sensor"})),
    SENSOR_BRAND_MODEL_FIELD: AuditFieldMetadata(frozenset({"sensor"})),
    "Vacuum Confirmation Present?": AuditFieldMetadata(frozenset({"sensor"}), "Yes"),
    PART_PRESENT_DETECTION_FIELD: AuditFieldMetadata(frozenset({"sensor"}), "No"),
    ELECTRICAL_WIRING_PRESENT_FIELD: AuditFieldMetadata(frozenset({"electrical"})),
    CUP_COUNT_FIELD: AuditFieldMetadata(frozenset({"vacuum"})),
    "Vacuum Generator Type": AuditFieldMetadata(frozenset({"vacuum"}), "Venturi"),
    EOAT_INTERCHANGEABLE_CIRCUITS_FIELD: AuditFieldMetadata(frozenset({"pneumatic_circuit"}), "0"),
    "Quick Disconnects Present?": AuditFieldMetadata(frozenset({"quick_disconnect"}), "Yes"),
    "Electrical Quick Disconnect Type": AuditFieldMetadata(frozenset({"electrical"})),
    "Cable Management Condition": AuditFieldMetadata(frozenset({"wiring", "cable_management"})),
    "Spare Parts Identified?": AuditFieldMetadata(frozenset({"documentation"}), "No"),
    "Drawing/CAD Available?": AuditFieldMetadata(frozenset({"documentation"}), "No"),
    "BOM Available?": AuditFieldMetadata(frozenset({"documentation"}), "No"),
    "Process Binder Complete?": AuditFieldMetadata(frozenset({"documentation"}), "No"),
    "Photos Taken?": AuditFieldMetadata(frozenset({"photo"}), "No"),
    "Photo Folder/Link": AuditFieldMetadata(frozenset({"photo_link"})),
}

SENSOR_ELECTRICAL_TAGS = frozenset({"sensor"})
SENSOR_ELECTRICAL_FIELDS = frozenset(
    field_name
    for field_name, metadata in AUDIT_FIELD_METADATA.items()
    if metadata.tags & SENSOR_ELECTRICAL_TAGS
)
DOCUMENTATION_PHOTO_DEFAULT_FIELDS = frozenset(
    field_name
    for field_name, metadata in AUDIT_FIELD_METADATA.items()
    if metadata.default == "No" and metadata.tags & {"documentation", "photo"}
)

AUDIT_DROPDOWNS = {
    "Plant/Area": ["Plant 4", "Cleanroom"],
    "Cleanroom/Non-Cleanroom": CLEANROOM_DROPDOWN_VALUES,
    "EOAT Type": EOAT_TYPE_DROPDOWN_VALUES,
    EOAT_MOVES_FIELD: EOAT_MOVES_VALUES,
    CONNECTION_TYPE_FIELD: CONNECTION_TYPE_VALUES,
    GRIPPER_TYPE_FIELD: GRIPPER_TYPE_VALUES,
    "Robot Type": ["Wittmann R8", "Wittmann R9", "Engel Viper", "Other", "Unknown"],
    "YesNoUnknown": ["Yes", "No", UNKNOWN_NOT_CHECKED],
    "YesNoUnknownNA": ["Yes", "No", UNKNOWN_NOT_CHECKED, "Not Applicable"],
    "YesNoPartialUnknown": ["Yes", "No", "Partial", UNKNOWN_NOT_CHECKED],
    ELECTRICAL_WIRING_PRESENT_FIELD: ["Yes", "No", UNKNOWN_NOT_CHECKED],
    "Quick Disconnects Present?": ["Yes", "No", "Partial", UNKNOWN_NOT_CHECKED],
    "Tubing Condition": ["OK", "Worn", "Damaged", "Poor Routing", "Needs Follow-Up", UNKNOWN_NOT_CHECKED],
    "Cable Management Condition": ["OK", "Loose", "Damaged", "Poor Routing", "Needs Follow-Up", UNKNOWN_NOT_CHECKED],
    "Mounting Hardware Condition": ["OK", "Loose", "Missing Hardware", "Damaged", "Needs Follow-Up", UNKNOWN_NOT_CHECKED],
    "EOAT Alignment Condition": ["OK", "Slightly Off", "Misaligned", "Needs Follow-Up", UNKNOWN_NOT_CHECKED],
    "Changeover Difficulty": ["Easy", "Low", "Medium", "High", UNKNOWN_NOT_CHECKED],
    "Photos Taken?": ["Yes", "No"],
    "Status": ["Not Started", "In Progress", "Complete", "Needs Follow-Up", "Blocked"],
    "Priority": ["Low", "Medium", "High", "Critical"],
    "Pilot Candidate?": ["Yes", "No", "Maybe"],
    "Follow-Up Needed": ["Yes", "No"],
    ENTRY_TYPE_FIELD: [ENTRY_TYPE_AUDITED, ENTRY_TYPE_COMPATIBLE],
}

EOAT_TYPE_VALUES = {"Vacuum", "Mechanical gripper", "Hybrid", "Custom/other", "Unknown"}


def repair_legacy_audit_lookup_shift(row: dict[str, Any]) -> dict[str, Any]:
    """Recover short positional rows written before lookup columns expanded.

    Some older tests and ad hoc workbook edits appended compact EOAT Inventory rows
    by column position. When lookup columns are inserted ahead of EOAT condition
    fields, those compact rows can land under lookup headers. This keeps readers
    tolerant without changing the saved workbook.
    """
    legacy_lookup_shift = _text(row.get("Press Brand")) in EOAT_TYPE_VALUES
    legacy_compact_shift = _text(row.get("Cleanroom/Non-Cleanroom")) in EOAT_TYPE_VALUES and not _text(row.get("EOAT Type"))
    if _text(row.get("EOAT Type")):
        return _repair_missing_connection_type_positional_shift(
            _repair_legacy_compact_tooling_shift(
                _repair_missing_gripper_fields_positional_shift(_repair_legacy_tail_compact_shift(row))
            )
        )
    if not (legacy_lookup_shift or legacy_compact_shift):
        return row
    repaired = dict(row)
    if legacy_lookup_shift:
        fallback_map = {
            "EOAT Type": "Press Brand",
            NUMBER_OF_PARTS_PICKED_FIELD: "Press Model",
            "Cup Type/Material": "Press Tonnage",
            "Cup Diameter/Size": "Press Year",
            "Known Issues": "# of TCU's",
            "Scrap/Quality Concern?": "Screw Size",
            "Status": "Bill-to / Customer",
            "Priority": "Cycle Time (S)",
            "Pilot Candidate?": "Cavitation",
        }
    else:
        fallback_map = {
            "EOAT Type": "Cleanroom/Non-Cleanroom",
            NUMBER_OF_PARTS_PICKED_FIELD: "Vacuum Sensor",
            "Cup Type/Material": "Quick Disconnect",
            "Cup Diameter/Size": "PM Status",
            "Known Issues": "Cable Management Condition",
            "Scrap/Quality Concern?": "Estimated EOAT Weight",
            "Status": "Drawing/CAD Available?",
            "Priority": "BOM Available?",
            "Pilot Candidate?": "Process Binder Complete?",
        }
    for target, source in fallback_map.items():
        if not _text(repaired.get(target)) and _text(repaired.get(source)):
            repaired[target] = repaired.get(source)
    return repaired


def _repair_legacy_compact_tooling_shift(row: dict[str, Any]) -> dict[str, Any]:
    """Read compact rows written before EOAT Moves and Connection Type existed."""
    eoat_moves = _text(row.get(EOAT_MOVES_FIELD))
    connection_type = _text(row.get(CONNECTION_TYPE_FIELD))
    cup_type = _text(row.get("Cup Type/Material"))
    repaired = dict(row)
    if _looks_like_count(eoat_moves) and connection_type and connection_type not in CONNECTION_TYPE_VALUES:
        if not _text(repaired.get(NUMBER_OF_PARTS_PICKED_FIELD)):
            repaired[NUMBER_OF_PARTS_PICKED_FIELD] = row.get(EOAT_MOVES_FIELD)
        repaired[EOAT_MOVES_FIELD] = ""
        if not _text(repaired.get("Cup Type/Material")) or cup_type == _text(row.get("Cup Type/Material")):
            repaired["Cup Type/Material"] = row.get(CONNECTION_TYPE_FIELD)
        if cup_type and not _text(repaired.get("Cup Diameter/Size")):
            repaired["Cup Diameter/Size"] = row.get("Cup Type/Material")
        repaired[CONNECTION_TYPE_FIELD] = ""
    status_candidate = _text(repaired.get("BOM Available?"))
    priority_candidate = _text(repaired.get("Process Binder Complete?"))
    pilot_candidate = _text(repaired.get("Photos Taken?"))
    if (
        not _text(repaired.get("Status"))
        and status_candidate.lower() in {"candidate for pilot", "complete", "needs follow-up", "in progress", "blocked", "not started"}
    ):
        repaired["Status"] = repaired.get("BOM Available?")
        if priority_candidate in {"Low", "Medium", "High", "Critical"}:
            repaired["Priority"] = repaired.get("Process Binder Complete?")
        if pilot_candidate in {"Yes", "No", "Maybe"} and not _text(repaired.get("Pilot Candidate?")):
            repaired["Pilot Candidate?"] = repaired.get("Photos Taken?")
        if not _text(repaired.get("Known Issues")) or _text(repaired.get("Known Issues")) in {"Yes", "No", "Maybe"}:
            if _text(repaired.get("Mounting Hardware Condition")):
                repaired["Known Issues"] = repaired.get("Mounting Hardware Condition")
    return repaired


def _repair_missing_connection_type_positional_shift(row: dict[str, Any]) -> dict[str, Any]:
    """Read compact rows written by position before Connection Type existed."""
    connection_value = _text(row.get(CONNECTION_TYPE_FIELD))
    if _looks_like_count(connection_value) and (_text(row.get("Cup Type/Material")) or _text(row.get("Cup Diameter/Size"))):
        repaired = dict(row)
        if not _text(repaired.get(NUMBER_OF_PARTS_PICKED_FIELD)):
            repaired[NUMBER_OF_PARTS_PICKED_FIELD] = row.get(CONNECTION_TYPE_FIELD)
        repaired[CONNECTION_TYPE_FIELD] = ""
        return repaired
    shifted_after_connection = (
        _looks_like_count(connection_value)
        or (_text(row.get("Status")) in {"Low", "Medium", "High", "Critical"} and _text(row.get("Priority")) in {"Yes", "No", "Maybe"})
        or (_text(row.get("Estimated EOAT Weight")) and not _text(row.get("Known Issues")))
    )
    if not shifted_after_connection:
        return row
    headers = get_expected_headers("EOAT Inventory")
    if CONNECTION_TYPE_FIELD not in headers:
        return row
    repaired = dict(row)
    start = headers.index(CONNECTION_TYPE_FIELD)
    for index in range(len(headers) - 1, start, -1):
        repaired[headers[index]] = row.get(headers[index - 1])
    repaired[CONNECTION_TYPE_FIELD] = "" if connection_value not in CONNECTION_TYPE_VALUES else row.get(CONNECTION_TYPE_FIELD)
    return repaired


def _repair_missing_gripper_fields_positional_shift(row: dict[str, Any]) -> dict[str, Any]:
    """Read rows written by position before Gripper Model/Size existed."""
    shifted_after_gripper_fields = (
        _looks_like_count(row.get(GRIPPER_MODEL_FIELD))
        or (_text(row.get("Status")) in {"Low", "Medium", "High", "Critical"} and _text(row.get("Priority")) in {"Yes", "No", "Maybe"})
        or (_text(row.get("Estimated EOAT Weight")) and not _text(row.get("Known Issues")))
    )
    if not shifted_after_gripper_fields:
        return row
    headers = get_expected_headers("EOAT Inventory")
    if GRIPPER_MODEL_FIELD not in headers or GRIPPER_SIZE_FIELD not in headers:
        return row
    repaired = dict(row)
    start = headers.index(GRIPPER_MODEL_FIELD)
    for index in range(len(headers) - 1, start + 1, -1):
        repaired[headers[index]] = row.get(headers[index - 2])
    repaired[GRIPPER_MODEL_FIELD] = ""
    repaired[GRIPPER_SIZE_FIELD] = ""
    return repaired


def _repair_legacy_tail_compact_shift(row: dict[str, Any]) -> dict[str, Any]:
    """Recover very compact rows written before the current tail columns settled."""
    status_candidate = _text(row.get("Process Binder Complete?"))
    priority_candidate = _text(row.get("Photos Taken?"))
    pilot_candidate = _text(row.get("Photo Folder/Link"))
    if not (
        status_candidate
        and priority_candidate in {"Low", "Medium", "High", "Critical"}
        and pilot_candidate in {"Yes", "No", "Maybe"}
        and not _text(row.get("Status"))
    ):
        return row
    repaired = dict(row)
    repaired["Status"] = status_candidate
    repaired["Priority"] = priority_candidate
    repaired["Pilot Candidate?"] = pilot_candidate
    if not _text(repaired.get("Known Issues")) and _text(row.get("EOAT Alignment Condition")):
        repaired["Known Issues"] = row.get("EOAT Alignment Condition")
    return repaired


def _looks_like_count(value: Any) -> bool:
    text = _text(value)
    if not text:
        return False
    try:
        float(text)
    except ValueError:
        return False
    return True


def _text(value: Any) -> str:
    return str(value or "").strip()


def generate_audit_id(project_root: str | Path, audit_date: str | None = None) -> str:
    audit_date = audit_date or date.today().isoformat()
    compact = audit_date.replace("-", "")
    workbook_path = resolve_project_paths(project_root).master_workbook
    rows = row_dicts(workbook_path, "EOAT Inventory") if workbook_path.exists() else []
    prefix = f"AUD-{compact}-"
    max_number = 0
    for row in rows:
        value = str(row.get("Audit ID") or "")
        if value.startswith(prefix):
            try:
                max_number = max(max_number, int(value.rsplit("-", 1)[1]))
            except ValueError:
                continue
    return f"{prefix}{max_number + 1:03d}"


def normalize_audit_entry(project_root: str | Path, entry: dict[str, Any]) -> dict[str, Any]:
    normalized, _details = normalize_audit_entry_with_details(project_root, entry)
    return normalized


def normalize_audit_entry_with_details(project_root: str | Path, entry: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    headers = get_expected_headers("EOAT Inventory")
    if TOOL_FIELD not in entry and LEGACY_TOOL_FIELD in entry:
        entry = {**entry, TOOL_FIELD: entry.get(LEGACY_TOOL_FIELD, "")}
    if NUMBER_OF_PARTS_PICKED_FIELD not in entry and LEGACY_VACUUM_CUPS_FIELD in entry:
        entry = {**entry, NUMBER_OF_PARTS_PICKED_FIELD: entry.get(LEGACY_VACUUM_CUPS_FIELD, "")}
    normalized = {header: entry.get(header, "") for header in headers}
    if GRIPPER_MODEL_FIELD in normalized:
        normalized[GRIPPER_MODEL_FIELD] = gripper_model_to_workbook(normalized.get(GRIPPER_MODEL_FIELD), project_root)
    if ENTRY_TYPE_FIELD in normalized and not _text(normalized.get(ENTRY_TYPE_FIELD)):
        normalized[ENTRY_TYPE_FIELD] = ENTRY_TYPE_AUDITED
    entry_type = _text(normalized.get(ENTRY_TYPE_FIELD)).lower()
    if entry_type != ENTRY_TYPE_COMPATIBLE.lower() and not normalized.get("Audit Date"):
        normalized["Audit Date"] = date.today().isoformat()
    if not normalized.get("Audit ID"):
        normalized["Audit ID"] = generate_audit_id(project_root, str(normalized.get("Audit Date") or date.today().isoformat()))
    if not normalized.get("Cleanroom/Non-Cleanroom"):
        normalized["Cleanroom/Non-Cleanroom"] = CLEANROOM_DEFAULT
    eoat_type = normalized.get("EOAT Type")
    cleared_as_na: dict[str, str] = {}
    for header in headers:
        if not audit_field_applies(normalized, header):
            if _text(normalized.get(header)) != NA_VALUE:
                cleared_as_na[header] = field_rules.non_applicable_reason(normalized, header)
            normalized[header] = NA_VALUE
    pre_default_normalized = dict(normalized)
    for header in headers:
        if audit_field_applies(normalized, header) and not _text(normalized.get(header)):
            default = audit_field_default(header)
            if default is not None:
                normalized[header] = default
    if not _text(normalized.get("Cup Type/Material")) and cup_type_default_applies(eoat_type):
        normalized["Cup Type/Material"] = CUP_TYPE_DEFAULT
    if not tooling_field_applies(eoat_type, "Cup Type/Material") and _text(normalized.get("Cup Type/Material")) == CUP_TYPE_DEFAULT:
        normalized["Cup Type/Material"] = ""
    apply_part_present_sensor_defaults(normalized)
    if audit_field_applies(normalized, "Changeover Difficulty") and _can_apply_changeover_default(normalized.get("Changeover Difficulty")):
        connection_text = _text(normalized.get(CONNECTION_TYPE_FIELD)).casefold()
        if "ati" in connection_text:
            normalized["Changeover Difficulty"] = "Low"
        elif "dovetail" in connection_text or "dove tail" in connection_text:
            normalized["Changeover Difficulty"] = "Medium"
    for header in headers:
        if not _text(normalized.get(header)):
            if header in {SOURCE_AUDIT_ID_FIELD, COMPATIBILITY_SOURCE_FIELD, EOAT_MOVES_FIELD}:
                normalized[header] = ""
            else:
                normalized[header] = NA_VALUE
    details = {
        "fields_auto_set_to_na": cleared_as_na,
        "hybrid_warnings": field_rules.hybrid_completeness_warnings(pre_default_normalized),
        "semantic_warnings": field_rules.semantic_consistency_warnings(normalized),
    }
    return normalized, details


def validate_audit_entry(entry: dict[str, Any]) -> tuple[list[str], list[str]]:
    entry = apply_part_present_sensor_defaults(dict(entry))
    errors: list[str] = []
    warnings: list[str] = []
    requirements = field_rules.entry_type_requirements(entry)
    for field in requirements["required"]:
        if audit_field_applies(entry, field) and not _text(entry.get(field)):
            errors.append(f"Missing required field: {field}")
        elif audit_field_applies(entry, field) and is_na_value(entry.get(field)):
            warnings.append(f"Required field is marked {NA_VALUE} and needs review: {field}")
    for field in requirements["important"]:
        if audit_field_applies(entry, field) and _is_missing_audit_value(entry.get(field)):
            warnings.append(f"Missing important audit field: {field}")
    warnings.extend(field_rules.hybrid_completeness_warnings(entry))
    warnings.extend(field_rules.semantic_consistency_warnings(entry))
    gripper_type = _text(entry.get(GRIPPER_TYPE_FIELD))
    if (
        GRIPPER_TYPE_FIELD in entry
        and audit_field_applies(entry, GRIPPER_TYPE_FIELD)
        and gripper_type
        and not is_na_value(gripper_type)
        and gripper_type not in GRIPPER_TYPE_VALUES
    ):
        errors.append(f"{GRIPPER_TYPE_FIELD} must be one of: {', '.join(GRIPPER_TYPE_VALUES)}.")
    for field in {NUMBER_OF_PARTS_PICKED_FIELD, CUP_COUNT_FIELD, GRIPPER_COUNT_FIELD, *EOAT_PNEUMATIC_CIRCUIT_FIELDS}:
        if field in entry and _text(entry.get(field)) and not is_na_value(entry.get(field)) and _parse_non_negative_int(entry.get(field)) is None:
            errors.append(f"{field} must be a non-negative whole number.")
    return errors, warnings


def is_na_value(value: Any) -> bool:
    return _text(value).upper() == NA_VALUE


def audit_field_tags(field_name: str) -> frozenset[str]:
    metadata = AUDIT_FIELD_METADATA.get(field_name)
    return metadata.tags if metadata else frozenset()


def audit_field_default(field_name: str) -> str | None:
    metadata = AUDIT_FIELD_METADATA.get(field_name)
    return metadata.default if metadata else None


def _can_apply_changeover_default(value: Any) -> bool:
    text = _text(value).casefold()
    return text in {"", NA_VALUE.casefold(), UNKNOWN_NOT_CHECKED.casefold(), "not applicable"}


def part_present_sensor_value_allows_default(value: Any, default: str) -> bool:
    text = _text(value)
    return text.casefold() in {
        "",
        NA_VALUE.casefold(),
        "na",
        "not applicable",
        UNKNOWN_NOT_CHECKED.casefold(),
        "unknown",
        "not checked",
        default.casefold(),
    }


def apply_part_present_sensor_defaults(entry: dict[str, Any]) -> dict[str, Any]:
    if _text(entry.get(PART_PRESENT_DETECTION_FIELD)).casefold() != "yes":
        return entry
    for field_name, default in PART_PRESENT_SENSOR_DEFAULTS.items():
        if part_present_sensor_value_allows_default(entry.get(field_name), default):
            entry[field_name] = default
    return entry


def audit_field_has_any_tag(field_name: str, tags: set[str] | frozenset[str]) -> bool:
    return bool(audit_field_tags(field_name) & tags)


def sensor_electrical_fields_apply(entry: dict[str, Any]) -> bool:
    return _text(entry.get("Sensors Present?")).lower() != "no"


def audit_field_applies(entry: dict[str, Any], field_name: str) -> bool:
    return field_rules.field_applies(entry, field_name)


def _is_missing_audit_value(value: Any) -> bool:
    return not _text(value) or is_na_value(value)


def _parse_non_negative_int(value: Any) -> int | None:
    text = _text(value)
    if not text:
        return None
    try:
        parsed = int(text)
    except ValueError:
        return None
    if str(parsed) != text and text != f"{parsed}.0":
        return None
    return parsed if parsed >= 0 else None


def tooling_field_applies(eoat_type: Any, field: str) -> bool:
    return field_rules.field_applies({"EOAT Type": eoat_type}, field)


def cup_type_default_applies(eoat_type: Any) -> bool:
    normalized = field_rules.normalized_eoat_type(eoat_type)
    return normalized in {
        field_rules.EOAT_TYPE_BLANK,
        field_rules.EOAT_TYPE_VACUUM,
        field_rules.EOAT_TYPE_HYBRID,
        field_rules.EOAT_TYPE_UNKNOWN,
    }


def load_audit_entry(project_root: str | Path, audit_id: str) -> dict[str, Any] | None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    for row in row_dicts(workbook_path, "EOAT Inventory"):
        if str(row.get("Audit ID") or "") == str(audit_id):
            if TOOL_FIELD not in row and LEGACY_TOOL_FIELD in row:
                row = {**row, TOOL_FIELD: row.get(LEGACY_TOOL_FIELD, "")}
            if NUMBER_OF_PARTS_PICKED_FIELD not in row and LEGACY_VACUUM_CUPS_FIELD in row:
                row = {**row, NUMBER_OF_PARTS_PICKED_FIELD: row.get(LEGACY_VACUUM_CUPS_FIELD, "")}
            return {key: ("" if value is None else value) for key, value in row.items()}
    return None


def _ensure_inventory_headers(ws, required_headers: list[str]) -> list[str]:
    _migrate_legacy_tool_header(ws)
    _migrate_legacy_vacuum_cups_header(ws)
    existing = worksheet_headers(ws)
    missing = [header for header in required_headers if header not in existing]
    for header in missing:
        if header == TOOL_FIELD and "Press/Machine #" in existing:
            target_idx = existing.index("Press/Machine #") + 2
            _insert_inventory_header(ws, target_idx, header, style_from_col=target_idx - 1)
        elif header in TOOLING_COLUMN_ORDER and _tooling_insert_index(existing, header):
            target_idx = _tooling_insert_index(existing, header)
            _insert_inventory_header(ws, target_idx, header, style_from_col=max(1, target_idx - 1))
        elif header == ELECTRICAL_WIRING_PRESENT_FIELD and "Part-Present Detection Present?" in existing:
            target_idx = existing.index("Part-Present Detection Present?") + 2
            _insert_inventory_header(ws, target_idx, header, style_from_col=target_idx - 1)
        else:
            target_idx = ws.max_column + 1
            _insert_inventory_header(ws, target_idx, header, style_from_col=max(1, target_idx - 1))
        existing = worksheet_headers(ws)
    _move_tool_after_press(ws)
    _order_tooling_columns(ws)
    _style_inventory_tooling_columns(ws)
    _migrate_inventory_gripper_values(ws)
    _refresh_inventory_ranges(ws)
    _apply_inventory_validations(ws)
    return missing


def _insert_inventory_header(ws, target_idx: int, header: str, *, style_from_col: int | None = None) -> None:
    ws.insert_cols(target_idx)
    if style_from_col is not None and style_from_col >= 1 and ws.max_column > 1:
        source_col = min(style_from_col, ws.max_column)
        if source_col != target_idx:
            _copy_column_style(ws, source_col, target_idx, max_row=max(ws.max_row, 2))
    ws.cell(row=1, column=target_idx).value = header


def _migrate_electrical_wiring_presence_rows(ws) -> dict[str, int]:
    headers = worksheet_headers(ws)
    if ELECTRICAL_WIRING_PRESENT_FIELD not in headers:
        return {"rows_reviewed": 0, "set_no": 0, "set_unknown": 0, "set_yes": 0}

    positions = {header: headers.index(header) + 1 for header in headers}
    stats = {"rows_reviewed": 0, "set_no": 0, "set_unknown": 0, "set_yes": 0}
    electrical_col = positions[ELECTRICAL_WIRING_PRESENT_FIELD]
    sensor_detail_fields = {"Sensor Type", "Sensor Brand/Model"}
    electrical_evidence_fields = {"Electrical Quick Disconnect Type", "Cable Management Condition"}

    for row_number in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=row_number, column=column).value for column in range(1, len(headers) + 1)]
        if not any(_text(value) for value in row_values):
            continue
        current_value = ws.cell(row=row_number, column=electrical_col).value
        if _text(current_value) and _text(current_value).upper() != NA_VALUE:
            continue

        row_data = {header: ws.cell(row=row_number, column=column).value for header, column in positions.items()}
        stats["rows_reviewed"] += 1
        sensors_present = _text(row_data.get("Sensors Present?")).casefold()
        sensor_details_blank = not any(field_rules.is_meaningful_value(row_data.get(field)) for field in sensor_detail_fields)
        electrical_values_blank = not any(field_rules.is_meaningful_value(row_data.get(field)) for field in electrical_evidence_fields)

        if sensors_present == "no" and sensor_details_blank and electrical_values_blank:
            ws.cell(row=row_number, column=electrical_col).value = "No"
            for field in electrical_evidence_fields:
                if field in positions:
                    ws.cell(row=row_number, column=positions[field]).value = NA_VALUE
            stats["set_no"] += 1
        elif sensors_present == "yes" and not electrical_values_blank:
            ws.cell(row=row_number, column=electrical_col).value = "Yes"
            stats["set_yes"] += 1
        else:
            ws.cell(row=row_number, column=electrical_col).value = UNKNOWN_NOT_CHECKED
            stats["set_unknown"] += 1
    return stats


def _migrate_inventory_gripper_values(ws) -> dict[str, int]:
    headers = worksheet_headers(ws)
    positions = {header: headers.index(header) + 1 for header in headers}
    gripper_fields = [GRIPPER_COUNT_FIELD, GRIPPER_TYPE_FIELD, GRIPPER_MODEL_FIELD]
    if not any(field in positions for field in gripper_fields):
        return {"rows_reviewed": 0, "set_na": 0, "converted_model_presets": 0}

    stats = {"rows_reviewed": 0, "set_na": 0, "converted_model_presets": 0}
    for row_number in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=row_number, column=column).value for column in range(1, len(headers) + 1)]
        if not any(_text(value) for value in row_values):
            continue
        if len([value for value in row_values if _text(value)]) == 1 and _text(row_values[-1]).startswith("Last Updated:"):
            continue
        row_data = {header: ws.cell(row=row_number, column=column).value for header, column in positions.items()}
        stats["rows_reviewed"] += 1
        for field in gripper_fields:
            column = positions.get(field)
            if column is None:
                continue
            cell = ws.cell(row=row_number, column=column)
            if field == GRIPPER_MODEL_FIELD:
                converted = gripper_model_to_workbook(cell.value)
                if converted != _text(cell.value):
                    cell.value = converted
                    row_data[field] = converted
                    stats["converted_model_presets"] += 1
            if not _text(cell.value):
                cell.value = NA_VALUE
                row_data[field] = NA_VALUE
                stats["set_na"] += 1
    return stats


def _create_vacuum_zones_removal_backup(workbook_path: Path) -> Path:
    backup_dir = workbook_path.parent / "_backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = backup_dir / f"{workbook_path.stem}_backup_before_removing_vacuum_zones_{stamp}{workbook_path.suffix}"
    counter = 2
    while target.exists():
        target = backup_dir / f"{workbook_path.stem}_backup_before_removing_vacuum_zones_{stamp}_{counter}{workbook_path.suffix}"
        counter += 1
    shutil.copy2(workbook_path, target)
    return target


def _remove_legacy_vacuum_zones_columns(ws) -> int:
    removed = 0
    headers = worksheet_headers(ws)
    while VACUUM_ZONES_FIELD in headers:
        ws.delete_cols(headers.index(VACUUM_ZONES_FIELD) + 1)
        removed += 1
        headers = worksheet_headers(ws)
    return removed


def repair_workbook_schema(project_root: str | Path, log_activity: bool = True) -> ToolResult:
    started = time.perf_counter()
    paths = resolve_project_paths(project_root)
    workbook_path = paths.master_workbook
    if not workbook_path.exists():
        return ToolResult.fail(
            "workbook_schema_repair",
            "Workbook Schema Repair",
            "Master workbook is missing.",
            errors=[str(workbook_path)],
            duration_seconds=time.perf_counter() - started,
        )
    lock_status = detect_workbook_lock(workbook_path)
    if not lock_status.can_write:
        return ToolResult.fail(
            "workbook_schema_repair",
            "Workbook Schema Repair",
            "Workbook repair was blocked by the workbook lock detector.",
            errors=[lock_status.message],
            warnings=[lock_status.error] if lock_status.error else [],
            metrics={"workbook_locked": lock_status.locked},
            structured_data={"workbook_lock": lock_status.__dict__},
            duration_seconds=time.perf_counter() - started,
        )

    workbook = None
    vacuum_zones_backup: Path | None = None
    vacuum_zones_removed_count = 0
    try:
        backup = backup_file(workbook_path, workbook_path.parent / "_backups")
        workbook = load_workbook(workbook_path)
        if "EOAT Inventory" not in workbook.sheetnames:
            raise ValueError("EOAT Inventory sheet is missing.")
        _migrate_workbook_tool_headers(workbook)
        ws = workbook["EOAT Inventory"]
        if VACUUM_ZONES_FIELD in worksheet_headers(ws):
            vacuum_zones_backup = _create_vacuum_zones_removal_backup(workbook_path)
            vacuum_zones_removed_count = _remove_legacy_vacuum_zones_columns(ws)
        added_headers = _ensure_inventory_headers(ws, get_expected_headers("EOAT Inventory"))
        electrical_stats = _migrate_electrical_wiring_presence_rows(ws)
        refresh_audit_by_press_view(workbook)
        workbook.save(workbook_path)
        workbook.close()
    except Exception as exc:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        return ToolResult.fail(
            "workbook_schema_repair",
            "Workbook Schema Repair",
            "Could not repair workbook schema.",
            errors=[str(exc)],
            duration_seconds=time.perf_counter() - started,
        )

    details = [f"Workbook backup: {backup}"]
    files_created = [str(backup)]
    if vacuum_zones_backup is not None:
        details.append(f"Vacuum Zones removal backup: {vacuum_zones_backup}")
        details.append(f"Removed legacy Vacuum Zones column(s): {vacuum_zones_removed_count}")
        files_created.append(str(vacuum_zones_backup))
    if added_headers:
        details.append(f"Added EOAT Inventory header(s): {', '.join(added_headers)}")
    else:
        details.append("EOAT Inventory already had all expected headers.")
    details.append(
        "Electrical/Wiring Present? migration: "
        f"{electrical_stats['set_no']} row(s) set to No, "
        f"{electrical_stats['set_yes']} row(s) set to Yes, "
        f"{electrical_stats['set_unknown']} row(s) set to {UNKNOWN_NOT_CHECKED}."
    )
    result = ToolResult.ok(
        "workbook_schema_repair",
        "Workbook Schema Repair",
        "Workbook schema repair completed.",
        details=details,
        files_created=files_created,
        files_modified=[str(workbook_path)],
        metrics={
            "added_header_count": len(added_headers),
            "electrical_wiring_set_no_count": electrical_stats["set_no"],
            "electrical_wiring_set_yes_count": electrical_stats["set_yes"],
            "electrical_wiring_set_unknown_count": electrical_stats["set_unknown"],
            "vacuum_zones_columns_removed": vacuum_zones_removed_count,
        },
        duration_seconds=time.perf_counter() - started,
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def _migrate_workbook_tool_headers(workbook) -> None:
    for ws in workbook.worksheets:
        _migrate_legacy_tool_header(ws)
        _migrate_legacy_vacuum_cups_header(ws)
    if "EOAT Inventory" in workbook.sheetnames:
        _move_tool_after_press(workbook["EOAT Inventory"])
        _order_tooling_columns(workbook["EOAT Inventory"])


def _migrate_legacy_tool_header(ws) -> None:
    headers = worksheet_headers(ws)
    while LEGACY_TOOL_FIELD in headers:
        legacy_idx = headers.index(LEGACY_TOOL_FIELD) + 1
        if TOOL_FIELD in headers:
            tool_idx = headers.index(TOOL_FIELD) + 1
            for row_number in range(2, ws.max_row + 1):
                tool_cell = ws.cell(row=row_number, column=tool_idx)
                legacy_cell = ws.cell(row=row_number, column=legacy_idx)
                if tool_cell.value in (None, "") and legacy_cell.value not in (None, ""):
                    tool_cell.value = legacy_cell.value
            ws.delete_cols(legacy_idx)
        else:
            ws.cell(row=1, column=legacy_idx).value = TOOL_FIELD
        headers = worksheet_headers(ws)


def _migrate_legacy_vacuum_cups_header(ws) -> None:
    headers = worksheet_headers(ws)
    while LEGACY_VACUUM_CUPS_FIELD in headers:
        legacy_idx = headers.index(LEGACY_VACUUM_CUPS_FIELD) + 1
        if NUMBER_OF_PARTS_PICKED_FIELD in headers:
            target_idx = headers.index(NUMBER_OF_PARTS_PICKED_FIELD) + 1
            for row_number in range(2, ws.max_row + 1):
                target_cell = ws.cell(row=row_number, column=target_idx)
                legacy_cell = ws.cell(row=row_number, column=legacy_idx)
                if target_cell.value in (None, "") and legacy_cell.value not in (None, ""):
                    target_cell.value = legacy_cell.value
            ws.delete_cols(legacy_idx)
        else:
            ws.cell(row=1, column=legacy_idx).value = NUMBER_OF_PARTS_PICKED_FIELD
        headers = worksheet_headers(ws)


def _move_tool_after_press(ws) -> None:
    headers = worksheet_headers(ws)
    if "Press/Machine #" not in headers or TOOL_FIELD not in headers:
        return
    source_idx = headers.index(TOOL_FIELD) + 1
    target_idx = headers.index("Press/Machine #") + 2
    if source_idx == target_idx:
        return
    _move_column(ws, source_idx, target_idx)


def _tooling_insert_index(headers: list[str], header: str) -> int | None:
    if header not in TOOLING_COLUMN_ORDER:
        return None
    header_order_index = TOOLING_COLUMN_ORDER.index(header)
    for previous in reversed(TOOLING_COLUMN_ORDER[:header_order_index]):
        if previous in headers:
            return headers.index(previous) + 2
    for following in TOOLING_COLUMN_ORDER[header_order_index + 1 :]:
        if following in headers:
            return headers.index(following) + 1
    return None


def _order_tooling_columns(ws) -> None:
    if "EOAT Type" not in worksheet_headers(ws):
        return
    for offset, header in enumerate([header for header in TOOLING_COLUMN_ORDER if header in worksheet_headers(ws)]):
        headers = worksheet_headers(ws)
        target_idx = headers.index("EOAT Type") + 1 + offset
        source_idx = headers.index(header) + 1
        if source_idx != target_idx:
            _move_column(ws, source_idx, target_idx)


def _style_inventory_tooling_columns(ws) -> None:
    headers = worksheet_headers(ws)
    style_pairs = {
        EOAT_MOVES_FIELD: "EOAT Type",
        CONNECTION_TYPE_FIELD: "EOAT Type",
        GRIPPER_COUNT_FIELD: NUMBER_OF_PARTS_PICKED_FIELD,
        CUP_COUNT_FIELD: NUMBER_OF_PARTS_PICKED_FIELD,
        GRIPPER_TYPE_FIELD: CONNECTION_TYPE_FIELD,
        GRIPPER_MODEL_FIELD: GRIPPER_TYPE_FIELD,
        GRIPPER_SIZE_FIELD: GRIPPER_MODEL_FIELD,
        EOAT_VACUUM_CIRCUITS_FIELD: "Vacuum Generator Type",
        EOAT_PRESSURE_CIRCUITS_FIELD: EOAT_VACUUM_CIRCUITS_FIELD,
        EOAT_INTERCHANGEABLE_CIRCUITS_FIELD: EOAT_PRESSURE_CIRCUITS_FIELD,
    }
    for target_header, source_header in style_pairs.items():
        if target_header not in headers or source_header not in headers:
            continue
        target_col = headers.index(target_header) + 1
        source_col = headers.index(source_header) + 1
        _copy_column_style(ws, source_col, target_col, max_row=max(ws.max_row, 2))


def _copy_column_style(ws, source_col: int, target_col: int, *, max_row: int) -> None:
    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(target_col)
    ws.column_dimensions[target_letter].width = ws.column_dimensions[source_letter].width
    for row_number in range(1, max_row + 1):
        source = ws.cell(row=row_number, column=source_col)
        target = ws.cell(row=row_number, column=target_col)
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.fill = copy(source.fill)
        target.font = copy(source.font)
        target.border = copy(source.border)
        target.protection = copy(source.protection)


def _refresh_inventory_ranges(ws) -> None:
    headers = worksheet_headers(ws)
    if not headers:
        return
    last_column = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_column}1"
    for table in ws.tables.values():
        try:
            min_col, min_row, _max_col, max_row = range_boundaries(table.ref)
        except ValueError:
            continue
        if min_row == 1:
            table.ref = f"{get_column_letter(min_col)}{min_row}:{last_column}{max_row}"


def _move_connection_after_eoat_type(ws) -> None:
    headers = worksheet_headers(ws)
    if "EOAT Type" not in headers or CONNECTION_TYPE_FIELD not in headers:
        return
    source_idx = headers.index(CONNECTION_TYPE_FIELD) + 1
    target_idx = headers.index("EOAT Type") + 2
    if source_idx == target_idx:
        return
    _move_column(ws, source_idx, target_idx)


def _dropdown_formula(values: list[str]) -> str:
    return '"' + ",".join(value.replace('"', '""') for value in values) + '"'


def _remove_column_validations(ws, column_numbers: set[int]) -> None:
    kept = []
    for validation in ws.data_validations.dataValidation:
        ranges = getattr(validation.sqref, "ranges", [])
        if any(
            cell_range.min_col in column_numbers
            and cell_range.max_col in column_numbers
            and cell_range.min_row <= 2
            and cell_range.max_row >= 1000
            for cell_range in ranges
        ):
            continue
        kept.append(validation)
    ws.data_validations.dataValidation = kept


def _add_column_validation(ws, column_number: int, values: list[str]) -> None:
    column_letter = get_column_letter(column_number)
    validation = DataValidation(type="list", formula1=_dropdown_formula(values), allow_blank=True)
    validation.error = "Choose a value from the dropdown list."
    validation.errorTitle = "Invalid value"
    validation.prompt = "Choose a standard value, or leave blank if not known yet."
    validation.promptTitle = "Dropdown"
    ws.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}1000")


def _add_whole_number_validation(ws, column_number: int) -> None:
    column_letter = get_column_letter(column_number)
    validation = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    validation.error = "Enter a non-negative whole number, or leave blank if not known yet."
    validation.errorTitle = "Invalid whole number"
    validation.prompt = "Use a non-negative whole number."
    validation.promptTitle = "Whole number"
    ws.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}1000")


def _apply_inventory_validations(ws) -> None:
    headers = worksheet_headers(ws)
    desired = {
        "EOAT Type": [*EOAT_TYPE_DROPDOWN_VALUES, NA_VALUE],
        EOAT_MOVES_FIELD: EOAT_MOVES_VALUES,
        CONNECTION_TYPE_FIELD: [*CONNECTION_TYPE_VALUES, NA_VALUE],
        GRIPPER_TYPE_FIELD: [*GRIPPER_TYPE_VALUES, NA_VALUE],
        "Cleanroom/Non-Cleanroom": [*CLEANROOM_DROPDOWN_VALUES, NA_VALUE],
        "Electrical/Wiring Present?": ["Yes", "No", "Unknown / Not Checked", NA_VALUE],
        ENTRY_TYPE_FIELD: [ENTRY_TYPE_AUDITED, ENTRY_TYPE_COMPATIBLE],
    }
    numeric_headers = {
        NUMBER_OF_PARTS_PICKED_FIELD,
        CUP_COUNT_FIELD,
        GRIPPER_COUNT_FIELD,
        *EOAT_PNEUMATIC_CIRCUIT_FIELDS,
    }
    columns = {headers.index(header) + 1 for header in desired if header in headers}
    columns.update(headers.index(header) + 1 for header in numeric_headers if header in headers)
    if not columns:
        return
    _remove_column_validations(ws, columns)
    for header, values in desired.items():
        if header in headers:
            _add_column_validation(ws, headers.index(header) + 1, values)
    for header in numeric_headers:
        if header in headers:
            _add_whole_number_validation(ws, headers.index(header) + 1)


def _move_column(ws, source_idx: int, target_idx: int) -> None:
    width = ws.column_dimensions[get_column_letter(source_idx)].width
    cells = []
    for row_number in range(1, ws.max_row + 1):
        source = ws.cell(row=row_number, column=source_idx)
        cells.append(
            {
                "value": source.value,
                "style": copy(source._style),
                "number_format": source.number_format,
                "alignment": copy(source.alignment),
                "fill": copy(source.fill),
                "font": copy(source.font),
                "border": copy(source.border),
                "protection": copy(source.protection),
            }
        )
    ws.delete_cols(source_idx)
    if source_idx < target_idx:
        target_idx -= 1
    ws.insert_cols(target_idx)
    ws.column_dimensions[get_column_letter(target_idx)].width = width
    for row_number, snapshot in enumerate(cells, start=1):
        target = ws.cell(row=row_number, column=target_idx)
        target.value = snapshot["value"]
        target._style = snapshot["style"]
        target.number_format = snapshot["number_format"]
        target.alignment = snapshot["alignment"]
        target.fill = snapshot["fill"]
        target.font = snapshot["font"]
        target.border = snapshot["border"]
        target.protection = snapshot["protection"]


def save_audit_entry(
    project_root: str | Path,
    entry: dict[str, Any],
    allow_update: bool = False,
    create_followup_action: bool = False,
    log_activity: bool = True,
    refresh_press_view: bool = True,
    sync_linked_compatibility: bool = True,
) -> ToolResult:
    started = time.perf_counter()
    validate_started = time.perf_counter()
    timing_metrics: dict[str, float] = {}
    paths = resolve_project_paths(project_root)
    workbook_path = paths.master_workbook
    if not workbook_path.exists():
        return ToolResult.fail("eoat_audit_form", "EOAT Audit Form Tool", "Master workbook is missing.", errors=[str(workbook_path)])

    validation_entry, validation_details = normalize_audit_entry_with_details(project_root, entry)
    requirements = field_rules.entry_type_requirements(validation_entry)
    for field in requirements["required"]:
        if field != "Audit Date" and not _text(entry.get(field)):
            validation_entry[field] = ""
    errors, warnings = validate_audit_entry(validation_entry)
    timing_metrics["audit_save.validate_form_seconds"] = round(time.perf_counter() - validate_started, 3)
    if errors:
        return ToolResult.fail(
            "eoat_audit_form",
            "EOAT Audit Form Tool",
            "Audit entry failed validation.",
            errors=errors,
            warnings=warnings,
            metrics=timing_metrics,
            duration_seconds=time.perf_counter() - started,
        )
    data, normalization_details = normalize_audit_entry_with_details(project_root, entry)
    for warning in validation_details.get("hybrid_warnings", []):
        if warning not in warnings:
            warnings.append(warning)
    for warning in validation_details.get("semantic_warnings", []):
        if warning not in warnings:
            warnings.append(warning)

    workbook = None
    existing_row = None
    previous_data: dict[str, Any] | None = None
    vacuum_zones_backup: Path | None = None
    vacuum_zones_removed_count = 0
    press_view_refresh_seconds = 0.0
    write_started = time.perf_counter()
    try:
        backup = backup_file(workbook_path, workbook_path.parent / "_backups")
        workbook = load_workbook(workbook_path)
        if "EOAT Inventory" not in workbook.sheetnames:
            raise ValueError("EOAT Inventory sheet is missing.")
        _migrate_workbook_tool_headers(workbook)
        ws = workbook["EOAT Inventory"]
        if VACUUM_ZONES_FIELD in worksheet_headers(ws):
            vacuum_zones_backup = _create_vacuum_zones_removal_backup(workbook_path)
            vacuum_zones_removed_count = _remove_legacy_vacuum_zones_columns(ws)
        added_headers = _ensure_inventory_headers(ws, get_expected_headers("EOAT Inventory"))
        electrical_migration_stats = (
            _migrate_electrical_wiring_presence_rows(ws)
            if ELECTRICAL_WIRING_PRESENT_FIELD in added_headers
            else {"rows_reviewed": 0, "set_no": 0, "set_unknown": 0, "set_yes": 0}
        )
        existing_row = find_row_by_value(ws, "Audit ID", str(data["Audit ID"]))
        if existing_row:
            headers = worksheet_headers(ws)
            previous_data = {
                header: ws.cell(row=existing_row, column=column).value
                for column, header in enumerate(headers, start=1)
            }
        if existing_row and not allow_update:
            workbook.close()
            return ToolResult.fail(
                "eoat_audit_form",
                "EOAT Audit Form Tool",
                "Audit ID already exists. Re-run with update enabled to modify it.",
                errors=[str(data["Audit ID"])],
                files_created=[str(backup)],
                duration_seconds=time.perf_counter() - started,
            )
        if existing_row:
            supplied_fields = set(entry)
            if LEGACY_TOOL_FIELD in supplied_fields:
                supplied_fields.add(TOOL_FIELD)
            if LEGACY_VACUUM_CUPS_FIELD in supplied_fields:
                supplied_fields.add(NUMBER_OF_PARTS_PICKED_FIELD)
            headers = worksheet_headers(ws)
            for column, header in enumerate(headers, start=1):
                existing_value = ws.cell(row=existing_row, column=column).value
                if (
                    header in data
                    and header not in supplied_fields
                    and audit_field_applies(data, header)
                    and _text(existing_value)
                ):
                    data[header] = existing_value
                elif header in data and not audit_field_applies(data, header):
                    data[header] = NA_VALUE
                    normalization_details.setdefault("fields_auto_set_to_na", {})[header] = field_rules.non_applicable_reason(data, header)
        row_number = existing_row or next_empty_row(ws)
        write_row_by_headers(ws, row_number, data)
        if refresh_press_view:
            press_view_started = time.perf_counter()
            refresh_audit_by_press_view(workbook)
            press_view_refresh_seconds = time.perf_counter() - press_view_started
        workbook.save(workbook_path)
        workbook.close()
        timing_metrics["audit_save.write_master_seconds"] = round(time.perf_counter() - write_started, 3)
        timing_metrics["audit_save.audit_by_press_refresh_seconds"] = round(press_view_refresh_seconds, 3)
    except Exception as exc:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        return ToolResult.fail(
            "eoat_audit_form",
            "EOAT Audit Form Tool",
            "Could not save audit entry.",
            errors=[str(exc)],
            warnings=warnings,
            metrics=timing_metrics,
            duration_seconds=time.perf_counter() - started,
        )

    sync_result = None
    compatibility_started = time.perf_counter()
    if sync_linked_compatibility and _text(data.get(ENTRY_TYPE_FIELD)).lower() == ENTRY_TYPE_AUDITED.lower():
        from .audit_compatibility import sync_compatible_rows_from_source

        sync_result = sync_compatible_rows_from_source(workbook_path, str(data["Audit ID"]))
        warnings.extend(sync_result.warning_messages)
    timing_metrics["audit_save.compatibility_seconds"] = round(time.perf_counter() - compatibility_started, 3)

    details = [
        f"Audit ID: {data['Audit ID']}",
        f"Workbook row: {row_number}",
        f"Mode: {'updated existing row' if existing_row else 'added new row'}",
        f"Workbook backup: {backup}",
    ]
    files_created = [str(backup)]
    if vacuum_zones_backup is not None:
        details.append(f"Vacuum Zones removal backup: {vacuum_zones_backup}")
        details.append(f"Removed legacy Vacuum Zones column(s): {vacuum_zones_removed_count}")
        files_created.append(str(vacuum_zones_backup))
    if added_headers:
        details.append(f"Added missing EOAT Inventory headers: {', '.join(added_headers)}")
    if electrical_migration_stats["rows_reviewed"]:
        details.append(
            "Electrical/Wiring Present? migration: "
            f"{electrical_migration_stats['set_no']} row(s) set to No, "
            f"{electrical_migration_stats['set_yes']} row(s) set to Yes, "
            f"{electrical_migration_stats['set_unknown']} row(s) set to {UNKNOWN_NOT_CHECKED}."
        )
    auto_na_fields = sorted(normalization_details.get("fields_auto_set_to_na", {}))
    if auto_na_fields:
        details.append(f"Auto-set non-applicable field(s) to {NA_VALUE}: {', '.join(auto_na_fields)}")
    hybrid_warnings = normalization_details.get("hybrid_warnings", [])
    semantic_warnings = normalization_details.get("semantic_warnings", [])
    for warning in hybrid_warnings:
        if warning not in warnings:
            warnings.append(warning)
    for warning in semantic_warnings:
        if warning not in warnings:
            warnings.append(warning)
    summary = f"Saved audit entry {data['Audit ID']}."
    if sync_result is not None:
        if sync_result.backup_path:
            files_created.append(sync_result.backup_path)
        if sync_result.updated_count:
            summary = (
                f"Saved audit entry {data['Audit ID']}. "
                f"Updated {sync_result.updated_count} linked compatibility entrie(s) from this source audit."
            )
            details.append(f"Updated {sync_result.updated_count} linked compatibility entrie(s) from this source audit.")
        elif sync_result.missing_source:
            details.append("No linked compatibility entries were updated because the saved source audit could not be reloaded.")
        else:
            summary = f"Saved audit entry {data['Audit ID']}. No linked compatibility entries found."
            details.append("No linked compatibility entries found.")
        if sync_result.skipped_count:
            summary += f" {sync_result.skipped_count} compatibility entrie(s) were skipped because they were not linked compatible rows."
            details.append(f"Skipped {sync_result.skipped_count} non-compatible row(s) with this Source Audit ID.")
    files_modified = [str(workbook_path)]
    action_result = None
    if create_followup_action or str(data.get("Follow-Up Needed") or "").lower() == "yes":
        action_result = add_action_item(
            project_root,
            action_item=f"Follow up on EOAT audit {data['Audit ID']}: {data.get('Known Issues') or data.get('Notes') or 'Review audit entry.'}",
            related_cell_press=str(data.get("Press/Machine #") or ""),
            priority=str(data.get("Priority") or "Medium"),
            notes=f"Generated from audit entry {data['Audit ID']}.",
            log_activity=False,
        )
        details.append(action_result.summary)
        warnings.extend(action_result.warnings)
        if action_result.errors:
            warnings.extend(action_result.errors)
        files_modified.extend(action_result.files_modified)
        files_created.extend(action_result.files_created)

    history_started = time.perf_counter()
    try:
        history_path = append_audit_history(
            project_root,
            str(data["Audit ID"]),
            "audit_updated" if existing_row else "audit_created",
            previous_data,
            data,
            files_modified=[str(workbook_path)],
        )
        files_modified.append(str(history_path))
    except Exception as exc:
        warnings.append(f"Audit history was not updated: {exc}")
    if sync_result is not None and sync_result.updated_count:
        try:
            history_path = append_audit_history(
                project_root,
                str(data["Audit ID"]),
                "compatibility_regenerated",
                {"linked_compatibility_rows": "0"},
                {"linked_compatibility_rows": str(sync_result.updated_count)},
                source="compatibility_sync",
                files_modified=files_modified,
            )
            files_modified.append(str(history_path))
        except Exception as exc:
            warnings.append(f"Compatibility history was not updated: {exc}")
    timing_metrics["audit_save.history_seconds"] = round(time.perf_counter() - history_started, 3)

    result = ToolResult.ok(
        "eoat_audit_form",
        "EOAT Audit Form Tool",
        summary,
        details=details,
        warnings=warnings,
        files_created=sorted(set(files_created)),
        files_modified=sorted(set(files_modified)),
        metrics={
            "audit_id": data["Audit ID"],
            "row": row_number,
            "updated": bool(existing_row),
            "audit_by_press_refreshed": bool(refresh_press_view),
            "linked_compatibility_sync_requested": bool(sync_linked_compatibility),
            "compatibility_rows_synced": sync_result.updated_count if sync_result else 0,
            "compatibility_rows_skipped": sync_result.skipped_count if sync_result else 0,
            "fields_auto_set_to_na": len(auto_na_fields),
            "hybrid_warning_count": len(hybrid_warnings),
            "semantic_warning_count": len(semantic_warnings),
            "vacuum_zones_columns_removed": vacuum_zones_removed_count,
            **timing_metrics,
        },
        duration_seconds=time.perf_counter() - started,
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def save_audit_entry_with_compatibility_autorun(
    project_root: str | Path,
    entry: dict[str, Any],
    allow_update: bool = False,
    create_followup_action: bool = False,
) -> ToolResult:
    started = time.perf_counter()
    save_started = time.perf_counter()
    save_result = save_audit_entry(
        project_root,
        entry,
        allow_update=allow_update,
        create_followup_action=create_followup_action,
    )
    save_seconds = time.perf_counter() - save_started
    if not save_result.success:
        save_result.metrics["audit_save_seconds"] = round(save_seconds, 3)
        return save_result

    audit_id = str(save_result.metrics.get("audit_id") or entry.get("Audit ID") or "").strip()
    compatibility_started = time.perf_counter()
    compatibility_result = _autorun_compatibility_entry(project_root, audit_id)
    compatibility_seconds = time.perf_counter() - compatibility_started
    save_lines = _audit_save_summary_lines(save_result)
    compatibility_lines = _compatibility_summary_lines(compatibility_result)
    combined_summary = "\n".join(
        [
            "Audit Save Summary",
            "------------------",
            *save_lines,
            "",
            "Compatibility Entry Summary",
            "---------------------------",
            *compatibility_lines,
        ]
    )
    warnings = [*save_result.warnings, *compatibility_result.warnings]
    if not compatibility_result.success and compatibility_result.errors:
        warnings.extend(compatibility_result.errors)
    return ToolResult.ok(
        "audit_save_with_compatibility",
        "Save Audit Entry",
        combined_summary,
        details=[*save_result.details, *compatibility_result.details],
        warnings=warnings,
        files_created=sorted(set([*save_result.files_created, *compatibility_result.files_created])),
        files_modified=sorted(set([*save_result.files_modified, *compatibility_result.files_modified])),
        output_reports=sorted(set([*save_result.output_reports, *compatibility_result.output_reports])),
        metrics={
            **save_result.metrics,
            "compatibility_autorun_success": compatibility_result.success,
            "compatibility_created": compatibility_result.metrics.get("created", 0),
            "compatibility_conflicts": compatibility_result.metrics.get("conflicts", 0),
            "audit_save_seconds": round(save_seconds, 3),
            "compatibility_autorun_seconds": round(compatibility_seconds, 3),
        },
        duration_seconds=time.perf_counter() - started,
    )


def _autorun_compatibility_entry(project_root: str | Path, audit_id: str) -> ToolResult:
    started = time.perf_counter()
    if not audit_id:
        return ToolResult.fail(
            "compatibility_entry_autorun",
            "Compatibility Entry",
            "Compatibility update failed.",
            errors=["Saved audit ID was not available."],
            duration_seconds=time.perf_counter() - started,
        )
    try:
        from .audit_compatibility import build_compatibility_candidates, create_compatibility_entries

        candidate_result = build_compatibility_candidates(project_root, audit_id)
        if candidate_result.errors:
            return ToolResult.fail(
                "compatibility_entry_autorun",
                "Compatibility Entry",
                "Compatibility update failed.",
                errors=candidate_result.errors,
                warnings=candidate_result.warnings,
                duration_seconds=time.perf_counter() - started,
            )
        machines = [candidate.machine_no for candidate in candidate_result.candidates if candidate.can_create]
        if machines:
            result = create_compatibility_entries(project_root, audit_id, machines)
            result.tool_id = "compatibility_entry_autorun"
            result.duration_seconds = result.duration_seconds or (time.perf_counter() - started)
            return result
        conflicts = sum(1 for candidate in candidate_result.candidates if candidate.recommended_action == "Conflict / Review Needed")
        return ToolResult.ok(
            "compatibility_entry_autorun",
            "Compatibility Entry",
            f"Compatibility entries were checked for {audit_id}. No new compatibility entries were needed.",
            details=[
                f"Create-compatible candidates found: {len(machines)}",
                f"Compatibility conflicts needing review: {conflicts}",
            ],
            warnings=candidate_result.warnings,
            metrics={"created": 0, "conflicts": conflicts},
            duration_seconds=time.perf_counter() - started,
        )
    except Exception as exc:
        return ToolResult.fail(
            "compatibility_entry_autorun",
            "Compatibility Entry",
            "Compatibility update failed.",
            errors=[f"Could not run compatibility update: {exc}"],
            duration_seconds=time.perf_counter() - started,
        )


def _audit_save_summary_lines(result: ToolResult) -> list[str]:
    audit_id = result.metrics.get("audit_id", "")
    row = result.metrics.get("row", "")
    lines = [f"Saved audit entry {audit_id}." if audit_id else result.summary]
    if row:
        lines.append(f"Updated workbook row {row}.")
    if result.metrics.get("fields_auto_set_to_na", 0):
        lines.append("Blank or non-applicable unanswered fields were written as N/A.")
    if result.warnings:
        lines.append(f"Warnings: {'; '.join(result.warnings)}")
    return lines


def _compatibility_summary_lines(result: ToolResult) -> list[str]:
    if result.success:
        lines = [result.summary]
        if result.metrics.get("created"):
            lines.append(f"Created {result.metrics.get('created')} compatibility entrie(s).")
        if result.metrics.get("conflicts"):
            lines.append(f"{result.metrics.get('conflicts')} compatibility concern(s) need review.")
        if result.warnings:
            lines.append(f"Warnings: {'; '.join(result.warnings)}")
        return lines
    message = "; ".join(result.errors) if result.errors else result.summary
    return [f"Compatibility update failed: {message}", "The saved audit entry was not rolled back."]
