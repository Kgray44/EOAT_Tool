from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.audit.uninstalled import (
    UNINSTALLED_MACHINE_AND_ROBOT_CONTEXT_FIELDS,
    has_meaningful_identifier,
    normalize_identifier,
)
from core.eoat_ids import EOAT_ASSEMBLY_ID_FIELD
from core.paths import get_press_capacity_file, resolve_project_paths
from core.press_lookup import CAPACITY_FILE_NAME, part_family_from_capacity_fields, reference_data_dir
from core.tool_fields import LEGACY_TOOL_FIELD, TOOL_FIELD
from core.workbook_cache import row_dicts_cached

TOOL_IDENTIFIER_ALIASES = (
    TOOL_FIELD,
    "Tool Number",
    "Tool ID",
    "Tool",
    "Part #",
    "Part Number",
    "Part/Tool #",
    "Part Tool #",
    "NGW Part Number",
    "NGW Part #",
    "Selected Part Number",
    LEGACY_TOOL_FIELD,
    "Mold #",
    "Mold Number",
    "EOAT Number",
)

OUTPUT_FIELD_ALIASES = {
    TOOL_FIELD: TOOL_IDENTIFIER_ALIASES,
    EOAT_ASSEMBLY_ID_FIELD: (EOAT_ASSEMBLY_ID_FIELD, "EOAT ID", "Assembly ID"),
    "Part Family": ("Part Family", "Family", "Part Category"),
    "Part Name/Description": (
        "Part Name/Description",
        "Part Name",
        "Part Description",
        "Description",
        "NGW Part Description",
        "Selected Part Description",
        "Tool Description",
    ),
}

SAFE_UNINSTALLED_TOOL_LOOKUP_FIELDS = (
    TOOL_FIELD,
    EOAT_ASSEMBLY_ID_FIELD,
    "Part Family",
    "Part Name/Description",
    "EOAT Type",
    "EOAT Moves",
    "Connection Type",
    "Number of Parts Picked",
    "# of Cylinders",
    "Cylinder Type",
    "# of Cups",
    "Cup Type/Material",
    "Cup Diameter/Size",
    "Vacuum Generator Type",
    "# of Grippers",
    "Gripper Type",
    "Gripper Model",
    "Gripper Size",
    "EOAT Vacuum Circuits",
    "EOAT Pressure Circuits",
    "EOAT Interchangeable Circuits",
    "Sensors Present?",
    "Sensor Type",
    "Sensor Brand/Model",
    "Vacuum Confirmation Present?",
    "Part-Present Detection Present?",
    "Electrical/Wiring Present?",
    "Quick Disconnects Present?",
    "Pneumatic Quick Disconnect Type",
    "Electrical Quick Disconnect Type",
    "Tubing Condition",
    "Tubing Routing Notes",
    "Cable Management Condition",
    "Mounting Hardware Condition",
    "EOAT Alignment Condition",
    "Fastener/Locking Hardware Present?",
    "Known Issues",
    "Drop/Mis-Pick History",
    "Maintenance Frequency",
    "Cycle Time Concern?",
    "Scrap/Quality Concern?",
    "Changeover Difficulty",
    "Spare Parts Identified?",
    "Drawing/CAD Available?",
    "BOM Available?",
    "Process Binder Complete?",
)
_UNSAFE_LOOKUP_FIELDS = UNINSTALLED_MACHINE_AND_ROBOT_CONTEXT_FIELDS | {
    "Audit ID",
    "Audit Date",
    "Auditor",
    "Status",
    "Priority",
    "Pilot Candidate?",
    "Follow-Up Needed",
    "Photos Taken?",
    "Photo Folder/Link",
    "Notes",
    "Robot Notes",
    "Source Audit ID",
    "Compatibility Source",
}


@dataclass(frozen=True)
class ToolLookupResult:
    tool_number: str
    matched: bool
    fields: dict[str, str]
    warnings: tuple[str, ...] = ()
    source: str = ""
    match_count: int = 0
    matched_field: str = ""


@dataclass(frozen=True)
class ToolIdentifierMatch:
    tool_number: str
    description: str = ""
    source: str = ""
    matched_field: str = ""


@dataclass(frozen=True)
class ToolIdentifierSearchResult:
    query: str
    matches: tuple[ToolIdentifierMatch, ...]
    match_count: int = 0
    warnings: tuple[str, ...] = ()
    source: str = ""


@dataclass(frozen=True)
class _LookupMatch:
    row: dict[str, Any]
    source: str
    matched_field: str


def lookup_tool_details_by_identifier(project_root: str | Path, identifier: Any) -> ToolLookupResult:
    tool_text = normalize_identifier(identifier)
    if not has_meaningful_identifier(tool_text):
        return ToolLookupResult(
            tool_number=tool_text,
            matched=False,
            fields={},
            warnings=("Enter a Tool # to look up existing EOAT details.",),
        )

    warnings: list[str] = []
    searched_sources: list[str] = []
    matches: list[_LookupMatch] = []
    requested_keys = _identifier_keys(tool_text)

    inventory_rows, inventory_warning, inventory_source = _load_inventory_rows(project_root)
    if inventory_source:
        searched_sources.append(inventory_source)
    if inventory_warning:
        warnings.append(inventory_warning)
    matches.extend(_matching_rows(inventory_rows, requested_keys, inventory_source))

    for source, rows, warning in _load_reference_capacity_rows(project_root):
        if source:
            searched_sources.append(source)
        if warning:
            warnings.append(warning)
        matches.extend(_matching_rows(rows, requested_keys, source))

    if not matches:
        return ToolLookupResult(
            tool_number=tool_text,
            matched=False,
            fields={},
            warnings=tuple(
                [
                    "Tool # was not found in the existing data source. You can continue manually.",
                    *warnings,
                ]
            ),
            source="; ".join(dict.fromkeys(searched_sources)),
        )

    fields, matched_field = _fields_from_matches(tool_text, matches)
    return ToolLookupResult(
        tool_number=tool_text,
        matched=True,
        fields=fields,
        warnings=tuple(warnings),
        source="; ".join(dict.fromkeys(match.source for match in matches if match.source)),
        match_count=len(matches),
        matched_field=matched_field,
    )


def lookup_tool_details(project_root: str | Path, tool_number: Any) -> ToolLookupResult:
    return lookup_tool_details_by_identifier(project_root, tool_number)


def search_tool_identifiers(
    project_root: str | Path,
    query: Any,
    *,
    limit: int = 12,
) -> ToolIdentifierSearchResult:
    query_text = normalize_identifier(query)
    if not has_meaningful_identifier(query_text):
        return ToolIdentifierSearchResult(query=query_text, matches=(), match_count=0)

    warnings: list[str] = []
    searched_sources: list[str] = []
    matches: list[_LookupMatch] = []

    inventory_rows, inventory_warning, inventory_source = _load_inventory_rows(project_root)
    if inventory_source:
        searched_sources.append(inventory_source)
    if inventory_warning:
        warnings.append(inventory_warning)
    matches.extend(_partial_matching_rows(inventory_rows, query_text, inventory_source))

    for source, rows, warning in _load_reference_capacity_rows(project_root):
        if source:
            searched_sources.append(source)
        if warning:
            warnings.append(warning)
        matches.extend(_partial_matching_rows(rows, query_text, source))

    tool_matches = _tool_identifier_matches_from_rows(matches)
    limited = tuple(tool_matches[: max(0, limit)])
    return ToolIdentifierSearchResult(
        query=query_text,
        matches=limited,
        match_count=len(tool_matches),
        warnings=tuple(warnings),
        source="; ".join(dict.fromkeys(searched_sources)),
    )


def _load_inventory_rows(project_root: str | Path) -> tuple[list[dict[str, Any]], str, str]:
    workbook_path = resolve_project_paths(project_root).master_workbook
    source = f"{workbook_path}::EOAT Inventory"
    try:
        rows = row_dicts_cached(workbook_path, "EOAT Inventory")
    except Exception as exc:
        return [], f"Tool lookup source could not be read: {exc}", source
    return rows, "", source


def _load_reference_capacity_rows(project_root: str | Path) -> list[tuple[str, list[dict[str, Any]], str]]:
    results: list[tuple[str, list[dict[str, Any]], str]] = []
    for path in _reference_capacity_paths(project_root):
        source = str(path)
        if not path.exists():
            continue
        try:
            rows = _read_capacity_reference_rows(path)
        except Exception as exc:
            results.append((source, [], f"Reference capacity lookup source could not be read: {exc}"))
            continue
        results.append((source, rows, ""))
    return results


def _reference_capacity_paths(project_root: str | Path) -> list[Path]:
    paths = resolve_project_paths(project_root)
    data_dir = reference_data_dir(project_root)
    candidates = [
        get_press_capacity_file(project_root),
        data_dir / CAPACITY_FILE_NAME,
        paths.legacy_reference_data / CAPACITY_FILE_NAME,
        paths.project_root / "Reference_Data" / CAPACITY_FILE_NAME,
    ]
    unique: list[Path] = []
    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate.expanduser().resolve()) if candidate.exists() else str(candidate)
        if key in seen:
            continue
        seen.add(key)
        unique.append(candidate)
    return unique


def _read_capacity_reference_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        rows: list[dict[str, Any]] = []
        ordered_names = sorted(workbook.sheetnames, key=lambda name: 0 if _header_key(name) == "capacity" else 1)
        for sheet_name in ordered_names:
            rows.extend(_rows_from_identifier_header_sheet(workbook[sheet_name], path, sheet_name))
        return rows
    finally:
        workbook.close()


def _rows_from_identifier_header_sheet(ws, path: Path, sheet_name: str) -> list[dict[str, Any]]:
    max_header_row = min(int(ws.max_row or 0), 25)
    for header_index, row in enumerate(ws.iter_rows(min_row=1, max_row=max_header_row, values_only=True), start=1):
        headers = [normalize_identifier(value) for value in row]
        if not _has_identifier_header(headers):
            continue
        rows: list[dict[str, Any]] = []
        for row_number, values in enumerate(ws.iter_rows(min_row=header_index + 1, values_only=True), start=header_index + 1):
            if not any(normalize_identifier(value) for value in values):
                continue
            row_data = {
                headers[index]: value
                for index, value in enumerate(values)
                if index < len(headers) and headers[index]
            }
            row_data["_lookup_source"] = str(path)
            row_data["_lookup_sheet"] = sheet_name
            row_data["_lookup_row_number"] = row_number
            rows.append(row_data)
        return rows
    return []


def _has_identifier_header(headers: list[str]) -> bool:
    normalized_headers = {_header_key(header) for header in headers if header}
    return any(_header_key(alias) in normalized_headers for alias in TOOL_IDENTIFIER_ALIASES)


def _matching_rows(rows: list[dict[str, Any]], requested_keys: set[str], source: str) -> list[_LookupMatch]:
    matches: list[_LookupMatch] = []
    for row in rows:
        matched_field = _matched_identifier_field(row, requested_keys)
        if matched_field:
            matches.append(_LookupMatch(row=row, source=source, matched_field=matched_field))
    return matches


def _matched_identifier_field(row: dict[str, Any], requested_keys: set[str]) -> str:
    for field_name in TOOL_IDENTIFIER_ALIASES:
        value = _value_for_aliases(row, (field_name,))
        if value and _identifier_keys(value) & requested_keys:
            return field_name
    return ""


def _partial_matching_rows(rows: list[dict[str, Any]], query_text: str, source: str) -> list[_LookupMatch]:
    matches: list[_LookupMatch] = []
    query_terms = _partial_identifier_terms(query_text)
    if not query_terms:
        return matches
    for row in rows:
        matched_field = _matched_partial_identifier_field(row, query_terms)
        if matched_field:
            matches.append(_LookupMatch(row=row, source=source, matched_field=matched_field))
    return matches


def _matched_partial_identifier_field(row: dict[str, Any], query_terms: set[str]) -> str:
    for field_name in TOOL_IDENTIFIER_ALIASES:
        value = _value_for_aliases(row, (field_name,))
        if value and _identifier_contains_terms(value, query_terms):
            return field_name
    return ""


def _fields_from_matches(tool_text: str, matches: list[_LookupMatch]) -> tuple[dict[str, str], str]:
    fields: dict[str, str] = {}
    matched_field = ""
    for match in sorted(matches, key=_match_sort_key, reverse=True):
        if not matched_field:
            matched_field = match.matched_field
        for field_name in SAFE_UNINSTALLED_TOOL_LOOKUP_FIELDS:
            if field_name in _UNSAFE_LOOKUP_FIELDS or field_name in fields:
                continue
            value = _lookup_output_value(match.row, field_name)
            if has_meaningful_identifier(value):
                fields[field_name] = value
    fields[TOOL_FIELD] = tool_text
    return fields, matched_field


def _tool_identifier_matches_from_rows(matches: list[_LookupMatch]) -> list[ToolIdentifierMatch]:
    tool_matches: list[ToolIdentifierMatch] = []
    seen: set[str] = set()
    for match in sorted(matches, key=_match_sort_key, reverse=True):
        tool_number = _lookup_output_value(match.row, TOOL_FIELD) or _value_for_aliases(
            match.row, (match.matched_field,)
        )
        if not has_meaningful_identifier(tool_number):
            continue
        key = _dedupe_identifier_key(tool_number)
        if key in seen:
            continue
        seen.add(key)
        description = _lookup_output_value(match.row, "Part Name/Description") or _lookup_output_value(
            match.row, "Part Family"
        )
        tool_matches.append(
            ToolIdentifierMatch(
                tool_number=tool_number,
                description=description,
                source=match.source,
                matched_field=match.matched_field,
            )
        )
    return sorted(tool_matches, key=lambda match: _dedupe_identifier_key(match.tool_number))


def _lookup_output_value(row: dict[str, Any], field_name: str) -> str:
    value = _value_for_aliases(row, OUTPUT_FIELD_ALIASES.get(field_name, (field_name,)))
    if has_meaningful_identifier(value):
        return value
    if field_name == "Part Family":
        return normalize_identifier(part_family_from_capacity_fields(row))
    return ""


def _value_for_aliases(row: dict[str, Any], aliases: tuple[str, ...]) -> str:
    for alias in aliases:
        if alias in row:
            value = normalize_identifier(row.get(alias))
            if value:
                return value
    normalized_headers = {_header_key(header): header for header in row}
    for alias in aliases:
        header = normalized_headers.get(_header_key(alias))
        if header:
            value = normalize_identifier(row.get(header))
            if value:
                return value
    return ""


def _identifier_keys(value: Any) -> set[str]:
    text = normalize_identifier(value).casefold()
    if not text:
        return set()
    folded = " ".join(text.split())
    compact = re.sub(r"[\s\-_]+", "", folded)
    keys = {folded}
    if compact:
        keys.add(compact)
    for candidate in (folded, compact):
        if candidate and candidate.isdigit():
            keys.add(candidate.lstrip("0") or "0")
    return keys


def _partial_identifier_terms(value: Any) -> set[str]:
    text = normalize_identifier(value).casefold()
    if not text:
        return set()
    folded = " ".join(text.split())
    compact = re.sub(r"[^a-z0-9]+", "", folded)
    terms = {folded}
    if compact:
        terms.add(compact)
    return {term for term in terms if term}


def _identifier_contains_terms(value: Any, query_terms: set[str]) -> bool:
    value_terms = _partial_identifier_terms(value)
    if not value_terms:
        return False
    return any(query in candidate for query in query_terms for candidate in value_terms)


def _dedupe_identifier_key(value: Any) -> str:
    text = normalize_identifier(value).casefold()
    return re.sub(r"[^a-z0-9]+", "", text) or text


def _header_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_identifier(value).casefold())


def _match_sort_key(match: _LookupMatch) -> tuple[str, str, str]:
    row = match.row
    return (
        normalize_identifier(row.get("Audit Date")),
        normalize_identifier(row.get("Audit ID")),
        normalize_identifier(row.get("_lookup_row_number")),
    )


__all__ = [
    "SAFE_UNINSTALLED_TOOL_LOOKUP_FIELDS",
    "TOOL_IDENTIFIER_ALIASES",
    "ToolIdentifierMatch",
    "ToolIdentifierSearchResult",
    "ToolLookupResult",
    "lookup_tool_details",
    "lookup_tool_details_by_identifier",
    "search_tool_identifiers",
]
