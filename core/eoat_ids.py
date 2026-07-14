from __future__ import annotations

import json
import re
import time
from copy import copy
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .logging import log_tool_run
from .paths import get_press_capacity_file, resolve_project_paths
from .result import ToolResult
from .safe_files import backup_file, ensure_directory
from .tool_fields import TOOL_FIELD
from .workbook_cache import invalidate_workbook_cache
from .workbook_io import row_dicts, worksheet_headers

EOAT_ASSEMBLY_ID_FIELD = "EOAT Assembly ID"
EOAT_PREFIX_PLANT4 = "P4"
EOAT_PREFIX_CLEANROOM = "CL"
EOAT_ID_PREFIXES = (EOAT_PREFIX_PLANT4, EOAT_PREFIX_CLEANROOM)
EOAT_ASSEMBLY_ID_PATTERN = re.compile(r"^(P4|CL)-EOAT-(\d{4})$", re.IGNORECASE)
EOAT_ID_SEARCH_PATTERN = re.compile(r"(?<![A-Za-z0-9])(?:P4|CL)-EOAT-\d{4}(?![A-Za-z0-9])", re.IGNORECASE)
CANONICAL_AREA_CLEANROOM = "cleanroom"
CANONICAL_AREA_PLANT4 = "plant4"
CANONICAL_AREA_UNKNOWN = "unknown"
PRIMARY_AREA_FIELDS = (
    "Plant/Area",
    "Plant Area",
    "Area",
    "Plant",
    "Location",
    "Production Area",
)
CLEANROOM_FLAG_FIELDS = (
    "Cleanroom/Non-Cleanroom",
    "Cleanroom",
    "Clean Room",
    "Cleanroom Status",
    "Production Environment",
    "Environment",
)
MACHINE_FIELDS = ("Press/Machine #", "Machine #", "Machine No.", "Machine Number", "Press")


@dataclass(frozen=True)
class ParsedEOATID:
    prefix: str
    number: int

    @property
    def value(self) -> str:
        return format_eoat_id(self.prefix, self.number)


@dataclass(frozen=True)
class EOATAssignmentSummary:
    assigned_count: int = 0
    ids_created: list[str] = field(default_factory=list)
    rows_updated: list[int] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def first_assigned_id(self) -> str:
        return self.ids_created[0] if self.ids_created else ""

    @property
    def last_assigned_id(self) -> str:
        return self.ids_created[-1] if self.ids_created else ""

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class EOATAssemblyContext:
    eoat_assembly_id: str
    tools: tuple[str, ...] = ()
    machines: tuple[str, ...] = ()
    capacity_machines: tuple[str, ...] = ()
    audit_ids: tuple[str, ...] = ()
    part_names: tuple[str, ...] = ()
    photo_count: int = 0

    @property
    def tool_count(self) -> int:
        return len(self.tools)

    @property
    def is_multi_tool(self) -> bool:
        return self.tool_count > 1

    @property
    def known_machines(self) -> tuple[str, ...]:
        return tuple(_sorted_text(set(self.machines) | set(self.capacity_machines)))

    def to_dict(self) -> dict[str, Any]:
        return {
            "EOAT Assembly ID": self.eoat_assembly_id,
            "Tool Count": self.tool_count,
            "Tool #s": ", ".join(self.tools),
            "Machine #s": ", ".join(self.known_machines),
            "Known Machine #s": ", ".join(self.known_machines),
            "Audit Machine #s": ", ".join(self.machines),
            "Press Capacity Machine #s": ", ".join(self.capacity_machines),
            "Audit IDs": ", ".join(self.audit_ids),
            "Part Names": ", ".join(self.part_names),
            "Photo Count": self.photo_count,
        }


def normalize_eoat_assembly_id(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    parsed = parse_eoat_id(text)
    return parsed.value if parsed else text


def is_valid_eoat_assembly_id(value: Any) -> bool:
    return parse_eoat_id(value) is not None


def parse_eoat_id(value: Any) -> ParsedEOATID | None:
    text = "" if value is None else str(value).strip()
    match = EOAT_ASSEMBLY_ID_PATTERN.fullmatch(text)
    if not match:
        return None
    return ParsedEOATID(match.group(1).upper(), int(match.group(2)))


def find_eoat_ids(value: Any) -> list[str]:
    text = "" if value is None else str(value)
    return [normalize_eoat_assembly_id(match.group(0)) for match in EOAT_ID_SEARCH_PATTERN.finditer(text)]


def get_eoat_prefix(value: Any) -> str:
    parsed = parse_eoat_id(value)
    return parsed.prefix if parsed else ""


def get_eoat_number(value: Any) -> int | None:
    parsed = parse_eoat_id(value)
    return parsed.number if parsed else None


def format_eoat_id(prefix: str, number: int | str) -> str:
    clean_prefix = str(prefix or "").strip().upper()
    if clean_prefix not in EOAT_ID_PREFIXES:
        raise ValueError(f"Unsupported EOAT ID prefix: {prefix}")
    try:
        parsed_number = int(str(number).strip())
    except ValueError as exc:
        raise ValueError(f"Invalid EOAT ID number: {number}") from exc
    if parsed_number < 0 or parsed_number > 9999:
        raise ValueError(f"EOAT ID number must be between 0 and 9999: {number}")
    return f"{clean_prefix}-EOAT-{parsed_number:04d}"


def extract_eoat_id_number(value: Any) -> int | None:
    return get_eoat_number(value)


def normalize_area(value: Any) -> str:
    text = _area_text(value)
    if not text or text in {"n/a", "na", "none", "unknown", "unknown not checked", "unknown unchecked"}:
        return CANONICAL_AREA_UNKNOWN
    compact = re.sub(r"[^a-z0-9]+", "", text)
    if compact in {
        "plant4",
        "p4",
        "whiteroom",
        "noncleanroom",
        "noncr",
        "production",
        "prod",
    }:
        return CANONICAL_AREA_PLANT4
    if text in {"white room", "non cleanroom", "non clean room", "plant 4"}:
        return CANONICAL_AREA_PLANT4
    if "non clean" in text or "white room" in text or "whiteroom" in compact:
        return CANONICAL_AREA_PLANT4
    if compact in {"cleanroom", "cleanrm", "cleanroomarea", "cl", "cr"}:
        return CANONICAL_AREA_CLEANROOM
    if text in {"clean room", "c r", "c/r"}:
        return CANONICAL_AREA_CLEANROOM
    return CANONICAL_AREA_UNKNOWN


def canonical_area(
    row_data: dict[str, Any] | None = None,
    *,
    area: Any = None,
    cleanroom_flag: Any = None,
    machine: Any = None,
) -> str:
    row = row_data or {}
    for field_name in PRIMARY_AREA_FIELDS:
        area_value = normalize_area(row.get(field_name))
        if area_value != CANONICAL_AREA_UNKNOWN:
            return area_value
    area_value = normalize_area(area)
    if area_value != CANONICAL_AREA_UNKNOWN:
        return area_value
    for field_name in CLEANROOM_FLAG_FIELDS:
        flag_value = normalize_area(row.get(field_name))
        if flag_value != CANONICAL_AREA_UNKNOWN:
            return flag_value
    flag_value = normalize_area(cleanroom_flag)
    if flag_value != CANONICAL_AREA_UNKNOWN:
        return flag_value
    return _canonical_area_from_machine(machine)


def is_cleanroom_area(value: Any = None, **kwargs: Any) -> bool:
    if isinstance(value, dict):
        return canonical_area(value, **kwargs) == CANONICAL_AREA_CLEANROOM
    return canonical_area(area=value, **kwargs) == CANONICAL_AREA_CLEANROOM


def is_plant4_area(value: Any = None, **kwargs: Any) -> bool:
    if isinstance(value, dict):
        return canonical_area(value, **kwargs) == CANONICAL_AREA_PLANT4
    return canonical_area(area=value, **kwargs) == CANONICAL_AREA_PLANT4


def determine_eoat_prefix(
    row_data: dict[str, Any] | None = None,
    *,
    area: Any = None,
    cleanroom_flag: Any = None,
    machine: Any = None,
    default_prefix: str = EOAT_PREFIX_PLANT4,
) -> str:
    resolved_area = canonical_area(row_data, area=area, cleanroom_flag=cleanroom_flag, machine=machine)
    if resolved_area == CANONICAL_AREA_CLEANROOM:
        return EOAT_PREFIX_CLEANROOM
    if resolved_area == CANONICAL_AREA_PLANT4:
        return EOAT_PREFIX_PLANT4
    return str(default_prefix or EOAT_PREFIX_PLANT4).strip().upper()


def get_eoat_id_prefix(
    area: Any = None,
    cleanroom_flag: Any = None,
    machine: Any = None,
    row_data: dict[str, Any] | None = None,
) -> str:
    return determine_eoat_prefix(row_data, area=area, cleanroom_flag=cleanroom_flag, machine=machine)


def expected_eoat_id_for_area(row_data: dict[str, Any], current_id: Any | None = None) -> str:
    parsed = parse_eoat_id(current_id if current_id is not None else row_data.get(EOAT_ASSEMBLY_ID_FIELD))
    if parsed is None:
        return ""
    return format_eoat_id(determine_eoat_prefix(row_data), parsed.number)


def migrate_eoat_id(old_id: Any, expected_prefix: str) -> str:
    parsed = parse_eoat_id(old_id)
    if parsed is None:
        return ""
    return format_eoat_id(expected_prefix, parsed.number)


def eoat_id_prefix_matches_area(row_data: dict[str, Any], value: Any | None = None) -> bool:
    parsed = parse_eoat_id(value if value is not None else row_data.get(EOAT_ASSEMBLY_ID_FIELD))
    if parsed is None:
        return False
    return parsed.prefix == determine_eoat_prefix(row_data)


def generate_next_eoat_assembly_id(
    existing_ids: list[Any] | tuple[Any, ...] | set[Any],
    row_data: dict[str, Any] | None = None,
    *,
    prefix: str | None = None,
    area: Any = None,
    cleanroom_flag: Any = None,
    machine: Any = None,
) -> str:
    target_prefix = (
        str(prefix).strip().upper()
        if prefix
        else determine_eoat_prefix(row_data, area=area, cleanroom_flag=cleanroom_flag, machine=machine)
    )
    highest = 0
    for value in existing_ids:
        parsed = parse_eoat_id(value)
        if parsed is not None and parsed.prefix == target_prefix:
            highest = max(highest, parsed.number)
    return format_eoat_id(target_prefix, highest + 1)


def assign_missing_eoat_assembly_ids(rows: list[dict[str, Any]]) -> EOATAssignmentSummary:
    warnings: list[str] = []
    existing_ids = [normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD)) for row in rows]
    invalid_existing = sorted({value for value in existing_ids if value and not is_valid_eoat_assembly_id(value)})
    if invalid_existing:
        warnings.append(
            "Ignored invalid existing EOAT Assembly ID value(s) while finding the next number: "
            + ", ".join(invalid_existing)
        )
    next_by_prefix = {
        prefix: max(
            (parsed.number for value in existing_ids if (parsed := parse_eoat_id(value)) and parsed.prefix == prefix),
            default=0,
        )
        + 1
        for prefix in EOAT_ID_PREFIXES
    }
    ids_created: list[str] = []
    rows_updated: list[int] = []
    for index, row in enumerate(rows):
        if normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD)):
            continue
        prefix = determine_eoat_prefix(row)
        new_id = format_eoat_id(prefix, next_by_prefix[prefix])
        next_by_prefix[prefix] += 1
        row[EOAT_ASSEMBLY_ID_FIELD] = new_id
        ids_created.append(new_id)
        rows_updated.append(index)
    return EOATAssignmentSummary(
        assigned_count=len(ids_created),
        ids_created=ids_created,
        rows_updated=rows_updated,
        warnings=warnings,
    )


def build_eoat_assembly_contexts(
    audit_rows: list[dict[str, Any]],
    photo_rows: list[dict[str, Any]] | None = None,
    press_capacity_path: str | Path | None = None,
) -> dict[str, EOATAssemblyContext]:
    photo_counts: dict[str, int] = {}
    for row in photo_rows or []:
        eoat_id = normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD))
        if eoat_id:
            photo_counts[eoat_id] = photo_counts.get(eoat_id, 0) + 1

    grouped: dict[str, dict[str, set[str]]] = {}
    for row in audit_rows:
        eoat_id = normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD))
        if not eoat_id:
            continue
        bucket = grouped.setdefault(
            eoat_id,
            {
                "tools": set(),
                "machines": set(),
                "audit_ids": set(),
                "part_names": set(),
            },
        )
        _add_text(bucket["tools"], row.get(TOOL_FIELD))
        for machine in _machine_values(row):
            _add_text(bucket["machines"], machine)
        _add_text(bucket["audit_ids"], row.get("Audit ID"))
        _add_text(bucket["part_names"], row.get("Part Name/Description") or row.get("Part Family"))

    capacity_lookup = _capacity_machine_lookup_from_path(press_capacity_path) if press_capacity_path else {}
    return {
        eoat_id: EOATAssemblyContext(
            eoat_assembly_id=eoat_id,
            tools=tuple(_sorted_text(values["tools"])),
            machines=tuple(_sorted_text(values["machines"])),
            capacity_machines=tuple(_capacity_machines_for_tools_from_lookup(values["tools"], capacity_lookup)),
            audit_ids=tuple(_sorted_text(values["audit_ids"])),
            part_names=tuple(_sorted_text(values["part_names"])),
            photo_count=photo_counts.get(eoat_id, 0),
        )
        for eoat_id, values in grouped.items()
    }


def eoat_context_for_id(project_root: str | Path, eoat_assembly_id: str) -> EOATAssemblyContext | None:
    target = normalize_eoat_assembly_id(eoat_assembly_id).casefold()
    if not target:
        return None
    paths = resolve_project_paths(project_root)
    if not paths.master_workbook.exists():
        return None
    contexts = build_eoat_assembly_contexts(
        row_dicts(paths.master_workbook, "EOAT Inventory"),
        row_dicts(paths.master_workbook, "Photo Index"),
        get_press_capacity_file(project_root),
    )
    for eoat_id, context in contexts.items():
        if eoat_id.casefold() == target:
            return context
    return None


def multi_tool_eoat_rows(
    audit_rows: list[dict[str, Any]],
    photo_rows: list[dict[str, Any]] | None = None,
    press_capacity_path: str | Path | None = None,
) -> list[dict[str, Any]]:
    contexts = build_eoat_assembly_contexts(audit_rows, photo_rows, press_capacity_path)
    return [
        context.to_dict()
        for context in sorted(contexts.values(), key=lambda item: item.eoat_assembly_id.casefold())
        if context.is_multi_tool
    ]


def eoat_summary_metrics(
    audit_rows: list[dict[str, Any]],
    photo_rows: list[dict[str, Any]] | None = None,
    press_capacity_path: str | Path | None = None,
) -> dict[str, int]:
    contexts = build_eoat_assembly_contexts(audit_rows, photo_rows, press_capacity_path)
    return {
        "multi_tool_eoat_count": sum(1 for context in contexts.values() if context.is_multi_tool),
        "total_eoat_assembly_ids": len(contexts),
        "total_eoat_tool_links": sum(context.tool_count for context in contexts.values()),
        "total_eoat_known_machine_links": sum(len(context.known_machines) for context in contexts.values()),
        "eoat_linked_photo_count": sum(context.photo_count for context in contexts.values()),
    }


def assign_missing_eoat_assembly_ids_in_workbook(
    project_root: str | Path, *, log_activity: bool = True
) -> ToolResult:
    started = time.perf_counter()
    paths = resolve_project_paths(project_root)
    workbook_path = paths.master_workbook
    if not workbook_path.exists():
        return ToolResult.fail(
            "assign_missing_eoat_ids",
            "Assign Missing EOAT IDs",
            "Master workbook is missing.",
            errors=[str(workbook_path)],
        )

    workbook = None
    try:
        workbook = load_workbook(workbook_path)
        if "EOAT Inventory" not in workbook.sheetnames:
            raise ValueError("EOAT Inventory sheet is missing.")
        ws = workbook["EOAT Inventory"]
        added_column = ensure_eoat_assembly_id_column(ws)
        headers = worksheet_headers(ws)
        eoat_col = headers.index(EOAT_ASSEMBLY_ID_FIELD) + 1
        row_numbers: list[int] = []
        rows: list[dict[str, Any]] = []
        already_assigned = 0
        for row_number in range(2, ws.max_row + 1):
            row_data = {
                header: ws.cell(row=row_number, column=column).value
                for column, header in enumerate(headers, start=1)
            }
            if not _has_inventory_content(row_data):
                continue
            if normalize_eoat_assembly_id(row_data.get(EOAT_ASSEMBLY_ID_FIELD)):
                already_assigned += 1
            row_numbers.append(row_number)
            rows.append(row_data)
        summary = assign_missing_eoat_assembly_ids(rows)
        changed = added_column or bool(summary.rows_updated)
        backup = backup_file(workbook_path, workbook_path.parent / "_backups") if changed else None
        for index in summary.rows_updated:
            ws.cell(row=row_numbers[index], column=eoat_col).value = rows[index][EOAT_ASSEMBLY_ID_FIELD]
        if changed:
            workbook.save(workbook_path)
            invalidate_workbook_cache(workbook_path)
        workbook.close()
        workbook = None
    except Exception as exc:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        return ToolResult.fail(
            "assign_missing_eoat_ids",
            "Assign Missing EOAT IDs",
            "Could not assign missing EOAT Assembly IDs.",
            errors=[str(exc)],
            duration_seconds=time.perf_counter() - started,
        )

    if summary.assigned_count:
        result_summary = f"Assigned {summary.assigned_count} missing EOAT Assembly IDs."
    elif added_column:
        result_summary = "Added EOAT Assembly ID column; no missing audit rows needed IDs."
    else:
        result_summary = "All audit rows with data already have EOAT Assembly IDs."
    details = [
        f"Assigned: {summary.assigned_count}",
        f"First new ID: {summary.first_assigned_id or 'N/A'}",
        f"Last new ID: {summary.last_assigned_id or 'N/A'}",
        f"Rows already assigned: {already_assigned}",
        "Existing EOAT IDs were preserved.",
    ]
    if backup is not None:
        details.append(f"Workbook backup: {backup}")
    if added_column:
        details.append("Added EOAT Assembly ID column to EOAT Inventory.")
    result = ToolResult.ok(
        "assign_missing_eoat_ids",
        "Assign Missing EOAT IDs",
        result_summary,
        details=details,
        warnings=summary.warnings,
        files_created=[str(backup)] if backup is not None else [],
        files_modified=[str(workbook_path)] if added_column or summary.assigned_count else [],
        metrics={
            "assigned_count": summary.assigned_count,
            "already_assigned_count": already_assigned,
            "added_eoat_assembly_id_column": added_column,
            "first_assigned_id": summary.first_assigned_id,
            "last_assigned_id": summary.last_assigned_id,
        },
        structured_data=summary.to_dict(),
        duration_seconds=time.perf_counter() - started,
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def ensure_eoat_assembly_id_column(ws) -> bool:
    return _ensure_header_near(ws, EOAT_ASSEMBLY_ID_FIELD, after_header=TOOL_FIELD)


def ensure_photo_index_eoat_assembly_id_column(ws) -> bool:
    return _ensure_header_near(ws, EOAT_ASSEMBLY_ID_FIELD, after_header=TOOL_FIELD)


def update_eoat_info_file(
    project_root: str | Path,
    eoat_assembly_id: str,
    *,
    audit_rows: list[dict[str, Any]] | None = None,
) -> Path:
    eoat_id = normalize_eoat_assembly_id(eoat_assembly_id)
    if not eoat_id:
        raise ValueError("EOAT Assembly ID is required.")
    paths = resolve_project_paths(project_root)
    root = ensure_directory(paths.cell_photos / eoat_id)
    if audit_rows is None and paths.master_workbook.exists():
        audit_rows = row_dicts(paths.master_workbook, "EOAT Inventory")
    context = build_eoat_assembly_contexts(
        audit_rows or [],
        press_capacity_path=get_press_capacity_file(project_root),
    ).get(eoat_id)
    audit_machines = list(context.machines if context else ())
    capacity_machines = list(context.capacity_machines if context else ())
    known_machines = list(context.known_machines if context else ())
    payload = {
        "eoat_assembly_id": eoat_id,
        "known_tools": list(context.tools if context else ()),
        "known_machines": known_machines,
        "audit_machines": audit_machines,
        "press_capacity_machines": capacity_machines,
        "known_audit_ids": list(context.audit_ids if context else ()),
        "last_updated": datetime.now().isoformat(timespec="seconds"),
    }
    path = root / "eoat_info.json"
    path.write_text(json.dumps(payload, ensure_ascii=True, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return path


def infer_eoat_assembly_id_for_photo_row(
    photo_row: dict[str, Any], audit_rows: list[dict[str, Any]]
) -> tuple[str, str]:
    if normalize_eoat_assembly_id(photo_row.get(EOAT_ASSEMBLY_ID_FIELD)):
        return normalize_eoat_assembly_id(photo_row.get(EOAT_ASSEMBLY_ID_FIELD)), "existing"

    audit_id = _text(photo_row.get("Related Audit ID")).casefold()
    if audit_id:
        matches = [
            normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD))
            for row in audit_rows
            if _text(row.get("Audit ID")).casefold() == audit_id
            and normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD))
        ]
        unique = _unique_preserve(matches)
        if len(unique) == 1:
            return unique[0], "audit_id"
        if len(unique) > 1:
            return "", "ambiguous"

    tool = _text(photo_row.get(TOOL_FIELD)).casefold()
    machine = _first_text(photo_row, MACHINE_FIELDS).casefold()
    candidates: list[str] = []
    for row in audit_rows:
        row_tool = _text(row.get(TOOL_FIELD)).casefold()
        row_machine = _first_text(row, MACHINE_FIELDS).casefold()
        if tool and row_tool != tool:
            continue
        if machine and row_machine != machine:
            continue
        eoat_id = normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD))
        if eoat_id:
            candidates.append(eoat_id)
    unique = _unique_preserve(candidates)
    if len(unique) == 1:
        return unique[0], "tool_machine" if machine else "tool"
    if len(unique) > 1:
        return "", "ambiguous"
    return "", "no_match"


def _ensure_header_near(ws, header: str, *, after_header: str) -> bool:
    headers = worksheet_headers(ws)
    if header in headers:
        return False
    if after_header in headers:
        target_idx = headers.index(after_header) + 2
    elif "Press/Machine #" in headers:
        target_idx = headers.index("Press/Machine #") + 2
    else:
        target_idx = ws.max_column + 1
    ws.insert_cols(target_idx)
    source_col = max(1, target_idx - 1)
    if ws.max_column > 1 and source_col != target_idx:
        _copy_column_style(ws, source_col, target_idx, max_row=max(ws.max_row, 2))
    ws.cell(row=1, column=target_idx).value = header
    return True


def _copy_column_style(ws, source_col: int, target_col: int, *, max_row: int) -> None:
    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(target_col)
    ws.column_dimensions[target_letter].width = ws.column_dimensions[source_letter].width
    for row_number in range(1, max_row + 1):
        source = ws.cell(row=row_number, column=source_col)
        target = ws.cell(row=row_number, column=target_col)
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.fill = copy(source.fill)
        target.font = copy(source.font)
        target.border = copy(source.border)
        target.protection = copy(source.protection)


def _has_inventory_content(row: dict[str, Any]) -> bool:
    return any(_text(value) for value in row.values())


def _add_text(values: set[str], value: Any) -> None:
    text = _text(value)
    if text and text.casefold() not in {"n/a", "na", "none", "unknown", "unknown / not checked"}:
        values.add(text)


def _first_text(row: dict[str, Any], fields: tuple[str, ...]) -> str:
    for field_name in fields:
        text = _text(row.get(field_name))
        if text and text.casefold() not in {"n/a", "na", "none", "unknown", "unknown / not checked"}:
            return text
    return ""


def _machine_values(row: dict[str, Any]) -> list[str]:
    values: list[str] = []
    seen: set[str] = set()
    for field_name in MACHINE_FIELDS:
        text = _text(row.get(field_name))
        folded = text.casefold()
        if not text or folded in {"n/a", "na", "none", "unknown", "unknown / not checked"} or folded in seen:
            continue
        seen.add(folded)
        values.append(text)
    return values


def _capacity_machine_lookup_from_path(press_capacity_path: str | Path | None) -> dict[str, set[str]]:
    if press_capacity_path is None:
        return {}
    try:
        from .audit_compatibility import load_required_relationships
    except Exception:
        return {}
    relationships, _warnings = load_required_relationships(press_capacity_path)
    lookup: dict[str, set[str]] = {}
    for relationship in relationships:
        for key in _tool_lookup_keys(relationship.part_number):
            lookup.setdefault(key, set()).add(relationship.machine_no)
    return lookup


def _capacity_machines_for_tools_from_lookup(
    tools: tuple[str, ...] | set[str], lookup: dict[str, set[str]]
) -> list[str]:
    tool_keys = {key for tool in tools for key in _tool_lookup_keys(tool)}
    if not tool_keys:
        return []
    machines = {machine for key in tool_keys for machine in lookup.get(key, set())}
    return _sorted_text(machines)


def _tool_lookup_keys(value: Any) -> set[str]:
    text = _text(value)
    if not text:
        return set()
    keys = {text.casefold()}
    prefix = re.match(r"^[A-Za-z0-9]+", text)
    if prefix:
        keys.add(prefix.group(0).casefold())
    return keys


def _sorted_text(values: set[str]) -> list[str]:
    return sorted(values, key=lambda item: (not item.isdigit(), int(item) if item.isdigit() else item.casefold()))


def _unique_preserve(values: list[str]) -> list[str]:
    seen: set[str] = set()
    unique: list[str] = []
    for value in values:
        folded = value.casefold()
        if folded in seen:
            continue
        seen.add(folded)
        unique.append(value)
    return unique


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _area_text(value: Any) -> str:
    text = _text(value).casefold()
    text = re.sub(r"[\\/]+", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _canonical_area_from_machine(machine: Any) -> str:
    # Machine-to-area inference is intentionally disabled until a reliable site map
    # is configured; Plant/Area and Cleanroom flags are the source of truth.
    _ = machine
    return CANONICAL_AREA_UNKNOWN


__all__ = [
    "CANONICAL_AREA_CLEANROOM",
    "CANONICAL_AREA_PLANT4",
    "CANONICAL_AREA_UNKNOWN",
    "EOAT_ASSEMBLY_ID_FIELD",
    "EOAT_ASSEMBLY_ID_PATTERN",
    "EOAT_ID_PREFIXES",
    "EOAT_ID_SEARCH_PATTERN",
    "EOAT_PREFIX_CLEANROOM",
    "EOAT_PREFIX_PLANT4",
    "EOATAssignmentSummary",
    "EOATAssemblyContext",
    "ParsedEOATID",
    "assign_missing_eoat_assembly_ids",
    "assign_missing_eoat_assembly_ids_in_workbook",
    "build_eoat_assembly_contexts",
    "canonical_area",
    "determine_eoat_prefix",
    "eoat_id_prefix_matches_area",
    "eoat_context_for_id",
    "eoat_summary_metrics",
    "ensure_eoat_assembly_id_column",
    "ensure_photo_index_eoat_assembly_id_column",
    "expected_eoat_id_for_area",
    "extract_eoat_id_number",
    "find_eoat_ids",
    "format_eoat_id",
    "generate_next_eoat_assembly_id",
    "get_eoat_id_prefix",
    "get_eoat_number",
    "get_eoat_prefix",
    "infer_eoat_assembly_id_for_photo_row",
    "is_cleanroom_area",
    "is_plant4_area",
    "is_valid_eoat_assembly_id",
    "migrate_eoat_id",
    "multi_tool_eoat_rows",
    "normalize_area",
    "normalize_eoat_assembly_id",
    "parse_eoat_id",
    "update_eoat_info_file",
]
