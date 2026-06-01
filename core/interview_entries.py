from __future__ import annotations

import time
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .action_items import add_action_item
from .logging import log_tool_run
from .paths import resolve_project_paths
from .result import ToolResult
from .safe_files import backup_file
from .workbook_cache import invalidate_workbook_cache
from .workbook_io import next_empty_row, row_dicts, write_row_by_headers
from .workbook_schema import get_expected_headers

INTERVIEW_QUESTIONS = [
    "Does this EOAT drop parts?",
    "Does it ever mis-pick?",
    "What usually fails first?",
    "How often are cups or tubing replaced?",
    "Is setup/changeover difficult?",
    "What issue is most annoying?",
    "What would you improve first?",
    "Does this cause downtime, scrap, or cycle time problems?",
    "Is there existing documentation for this EOAT?",
    "Who knows this cell best?",
]


def generate_interview_id(project_root: str | Path, interview_date: str | None = None) -> str:
    interview_date = interview_date or date.today().isoformat()
    compact = interview_date.replace("-", "")
    workbook_path = resolve_project_paths(project_root).master_workbook
    rows = row_dicts(workbook_path, "Interview Notes") if workbook_path.exists() else []
    prefix = f"INT-{compact}-"
    max_number = 0
    for row in rows:
        value = str(row.get("Interview ID") or "")
        if value.startswith(prefix):
            try:
                max_number = max(max_number, int(value.rsplit("-", 1)[1]))
            except ValueError:
                continue
    return f"{prefix}{max_number + 1:03d}"


def normalize_interview_entry(project_root: str | Path, entry: dict[str, Any]) -> dict[str, Any]:
    headers = get_expected_headers("Interview Notes")
    normalized = {header: entry.get(header, "") for header in headers}
    if not normalized.get("Date"):
        normalized["Date"] = date.today().isoformat()
    if not normalized.get("Interview ID"):
        normalized["Interview ID"] = generate_interview_id(project_root, str(normalized["Date"]))
    return normalized


def validate_interview_entry(entry: dict[str, Any]) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    for field in ["Date", "Role/Department"]:
        if not str(entry.get(field) or "").strip():
            errors.append(f"Missing required field: {field}")
    if not str(entry.get("Plant/Area") or entry.get("Press/Machine #") or "").strip():
        warnings.append("Add Plant/Area or Press/Machine # when available.")
    if not str(entry.get("Notes") or entry.get("Known EOAT Issues Mentioned") or "").strip():
        errors.append("Notes or Known EOAT Issues Mentioned is required.")
    return errors, warnings


def save_interview_entry(
    project_root: str | Path,
    entry: dict[str, Any],
    create_followup_action: bool = False,
    log_activity: bool = True,
) -> ToolResult:
    started = time.perf_counter()
    paths = resolve_project_paths(project_root)
    workbook_path = paths.master_workbook
    if not workbook_path.exists():
        return ToolResult.fail(
            "interview_form",
            "Operator/Technician Interview Form Tool",
            "Master workbook is missing.",
            errors=[str(workbook_path)],
        )
    data = normalize_interview_entry(project_root, entry)
    errors, warnings = validate_interview_entry(data)
    if errors:
        return ToolResult.fail(
            "interview_form",
            "Operator/Technician Interview Form Tool",
            "Interview entry failed validation.",
            errors=errors,
            warnings=warnings,
            duration_seconds=time.perf_counter() - started,
        )

    workbook = None
    try:
        backup = backup_file(workbook_path, workbook_path.parent / "_backups")
        workbook = load_workbook(workbook_path)
        if "Interview Notes" not in workbook.sheetnames:
            raise ValueError("Interview Notes sheet is missing.")
        ws = workbook["Interview Notes"]
        row_number = next_empty_row(ws)
        write_row_by_headers(ws, row_number, data)
        workbook.save(workbook_path)
        workbook.close()
        invalidate_workbook_cache(workbook_path)
    except Exception as exc:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        return ToolResult.fail(
            "interview_form",
            "Operator/Technician Interview Form Tool",
            "Could not save interview entry.",
            errors=[str(exc)],
            warnings=warnings,
            duration_seconds=time.perf_counter() - started,
        )

    details = [f"Interview ID: {data['Interview ID']}", f"Workbook row: {row_number}", f"Workbook backup: {backup}"]
    files_modified = [str(workbook_path)]
    action_result = None
    if create_followup_action or str(data.get("Follow-Up Needed") or "").lower() == "yes":
        action_result = add_action_item(
            project_root,
            action_item=f"Follow up from interview {data['Interview ID']}: {data.get('Known EOAT Issues Mentioned') or data.get('Suggested Improvements') or 'Review interview notes.'}",
            related_cell_press=str(data.get("Press/Machine #") or ""),
            owner=str(data.get("Follow-Up Owner") or ""),
            priority="Medium",
            notes=f"Generated from interview {data['Interview ID']}.",
            log_activity=False,
        )
        details.append(action_result.summary)
        warnings.extend(action_result.warnings)
        if action_result.errors:
            warnings.extend(action_result.errors)
        files_modified.extend(action_result.files_modified)

    result = ToolResult.ok(
        "interview_form",
        "Operator/Technician Interview Form Tool",
        f"Saved interview note {data['Interview ID']}.",
        details=details,
        warnings=warnings,
        files_created=[str(backup), *(action_result.files_created if action_result else [])],
        files_modified=sorted(set(files_modified)),
        metrics={"interview_id": data["Interview ID"], "row": row_number},
        duration_seconds=time.perf_counter() - started,
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result
