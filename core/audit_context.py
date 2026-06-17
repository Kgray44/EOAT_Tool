from __future__ import annotations

from collections.abc import Mapping
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .audit_constants import (
    AUDIT_CONTEXT_BENCH,
    AUDIT_CONTEXT_COMPATIBILITY,
    AUDIT_CONTEXT_FIELD,
    AUDIT_CONTEXT_HISTORICAL,
    AUDIT_CONTEXT_INSTALLED,
    AUDIT_CONTEXT_NEEDS_REVIEW,
    AUDIT_CONTEXT_VALUES,
    COMPATIBILITY_CONFIDENCE_FIELD,
    COMPATIBILITY_SOURCE_FIELD,
    COMPATIBILITY_SOURCE_PRESS_CAPACITY,
    ENTRY_TYPE_COMPATIBLE,
    ENTRY_TYPE_FIELD,
    PHYSICAL_AUDIT_VERIFIED_FIELD,
    SOURCE_AUDIT_ID_FIELD,
)
from .result import ToolResult
from .safe_files import backup_file, ensure_directory, safe_write_text
from .workbook_cache import invalidate_workbook_cache
from .workbook_io import worksheet_headers
from .workbook_schema import get_expected_headers

MACHINE_CONTEXT_FIELDS = frozenset(
    {
        "Plant/Area",
        "Press/Machine #",
        "Robot Type",
        "Robot Model/Controller",
        "Robot Vacuum Circuits",
        "Robot Pressure Circuits",
        "Robot Interchangeable Circuits",
        "Robot Notes",
        "Cycle Time Concern?",
        "Drop/Mis-Pick History",
        "Scrap/Quality Concern?",
        "Changeover Difficulty",
    }
)

INSTALLATION_ONLY_FIELDS = frozenset(
    {
        "Plant/Area",
        "Press/Machine #",
        "Robot Type",
        "Robot Model/Controller",
        "Robot Vacuum Circuits",
        "Robot Pressure Circuits",
        "Robot Interchangeable Circuits",
        "Robot Notes",
        "EOAT Alignment Condition",
        "Drop/Mis-Pick History",
        "Cycle Time Concern?",
        "Scrap/Quality Concern?",
        "Changeover Difficulty",
    }
)

EOAT_DOCUMENTATION_FIELDS = (
    "EOAT Assembly ID",
    "Tool #",
    "Part Family",
    "Part Name/Description",
    "EOAT Type",
    "EOAT Moves",
    "Connection Type",
    "Number of Parts Picked",
    "# of Cups",
    "Cup Type/Material",
    "Cup Diameter/Size",
    "# of Grippers",
    "Gripper Type",
    "Gripper Model",
    "# of Cylinders",
    "Cylinder Type",
    "EOAT Vacuum Circuits",
    "EOAT Pressure Circuits",
    "EOAT Interchangeable Circuits",
    "Sensors Present?",
    "Sensor Type",
    "Sensor Brand/Model",
    "Vacuum Confirmation Present?",
    "Part-Present Detection Present?",
    "Electrical/Wiring Present?",
    "Quick Disconnects Present?",
    "Pneumatic Quick Disconnect Type",
    "Electrical Quick Disconnect Type",
    "Tubing Condition",
    "Tubing Routing Notes",
    "Cable Management Condition",
    "Mounting Hardware Condition",
    "Fastener/Locking Hardware Present?",
    "Estimated EOAT Weight",
    "Known Issues",
    "Maintenance Frequency",
    "Spare Parts Identified?",
    "Drawing/CAD Available?",
    "BOM Available?",
    "Process Binder Complete?",
    "Photos Taken?",
    "Photo Folder/Link",
    "Notes",
)

INSTALLATION_READINESS_FIELDS = (
    "EOAT Assembly ID",
    "Tool #",
    "Part Family",
    "Part Name/Description",
    "EOAT Type",
    "EOAT Moves",
    "Connection Type",
    "Mounting Hardware Condition",
    "Fastener/Locking Hardware Present?",
    "EOAT Vacuum Circuits",
    "EOAT Pressure Circuits",
    "EOAT Interchangeable Circuits",
    "Quick Disconnects Present?",
    "Pneumatic Quick Disconnect Type",
    "Electrical Quick Disconnect Type",
    "Tubing Condition",
    "Cable Management Condition",
    "Known Issues",
    "Drawing/CAD Available?",
    "BOM Available?",
    "Photos Taken?",
)

INSTALLED_CELL_VALIDATION_FIELDS = (
    "Press/Machine #",
    "Robot Type",
    "Robot Model/Controller",
    "EOAT Alignment Condition",
    "Robot Vacuum Circuits",
    "Robot Pressure Circuits",
    "Robot Interchangeable Circuits",
    "Drop/Mis-Pick History",
    "Cycle Time Concern?",
    "Scrap/Quality Concern?",
    "Changeover Difficulty",
)

MISSING_MACHINE_VALUES = {
    "",
    "-",
    "--",
    "n/a",
    "na",
    "none",
    "null",
    "not applicable",
    "not installed",
    "eoat not installed",
    "bench",
    "bench audit",
    "off machine",
    "off-machine",
    "uninstalled",
    "unknown",
    "unknown / not checked",
    "not checked",
}


@dataclass(frozen=True)
class AuditContextBackfillResult:
    row_number: int
    audit_id: str
    inferred_context: str
    machine: str
    reason: str
    physical_verified: str = ""
    compatibility_confidence: str = ""


def normalize_context_value(value: Any) -> str:
    text = _text(value)
    if not text:
        return ""
    folded = text.casefold()
    aliases = {
        "installed": AUDIT_CONTEXT_INSTALLED,
        "installed on press": AUDIT_CONTEXT_INSTALLED,
        "installed on machine": AUDIT_CONTEXT_INSTALLED,
        "machine audit": AUDIT_CONTEXT_INSTALLED,
        "physical audit": AUDIT_CONTEXT_INSTALLED,
        "not installed": AUDIT_CONTEXT_BENCH,
        "uninstalled": AUDIT_CONTEXT_BENCH,
        "bench": AUDIT_CONTEXT_BENCH,
        "bench audit": AUDIT_CONTEXT_BENCH,
        "off-machine audit": AUDIT_CONTEXT_BENCH,
        "off machine audit": AUDIT_CONTEXT_BENCH,
        "not installed / bench audit": AUDIT_CONTEXT_BENCH,
        "compatible": AUDIT_CONTEXT_COMPATIBILITY,
        "compatibility": AUDIT_CONTEXT_COMPATIBILITY,
        "compatibility row": AUDIT_CONTEXT_COMPATIBILITY,
        "historical": AUDIT_CONTEXT_HISTORICAL,
        "imported": AUDIT_CONTEXT_HISTORICAL,
        "historical/imported": AUDIT_CONTEXT_HISTORICAL,
        "needs review": AUDIT_CONTEXT_NEEDS_REVIEW,
        "review": AUDIT_CONTEXT_NEEDS_REVIEW,
    }
    if folded in aliases:
        return aliases[folded]
    for value in AUDIT_CONTEXT_VALUES:
        if folded == value.casefold():
            return value
    return AUDIT_CONTEXT_NEEDS_REVIEW


def infer_audit_context(row: Mapping[str, Any], *, preserve_explicit: bool = True) -> str:
    explicit = normalize_context_value(row.get(AUDIT_CONTEXT_FIELD))
    if explicit and (preserve_explicit or explicit != AUDIT_CONTEXT_NEEDS_REVIEW):
        return explicit
    if _is_compatibility_marker(row):
        return AUDIT_CONTEXT_COMPATIBILITY
    if _is_historical_marker(row):
        return AUDIT_CONTEXT_HISTORICAL
    if machine_value_is_missing(_machine_value(row)):
        return AUDIT_CONTEXT_BENCH
    return AUDIT_CONTEXT_INSTALLED


def audit_context_reason(row: Mapping[str, Any]) -> str:
    if normalize_context_value(row.get(AUDIT_CONTEXT_FIELD)):
        return "Existing Audit Context value was preserved."
    if _is_compatibility_marker(row):
        return "Entry Type or compatibility source marks this as a compatibility row."
    if _is_historical_marker(row):
        return "Historical/import marker was detected."
    if machine_value_is_missing(_machine_value(row)):
        return "Machine value is blank, N/A, Not Installed, or equivalent."
    return "Machine value is populated, so the row is treated as installed on machine."


def machine_value_is_missing(value: Any) -> bool:
    return _text(value).casefold() in MISSING_MACHINE_VALUES


def is_bench_audit_context(row: Mapping[str, Any]) -> bool:
    return infer_audit_context(row) == AUDIT_CONTEXT_BENCH


def is_installed_audit_context(row: Mapping[str, Any]) -> bool:
    return infer_audit_context(row) == AUDIT_CONTEXT_INSTALLED


def is_compatibility_audit_context(row: Mapping[str, Any]) -> bool:
    return infer_audit_context(row) == AUDIT_CONTEXT_COMPATIBILITY


def physical_audit_verified_default(context: str) -> str:
    if context == AUDIT_CONTEXT_COMPATIBILITY:
        return "No"
    if context in {AUDIT_CONTEXT_INSTALLED, AUDIT_CONTEXT_BENCH}:
        return "Yes"
    return "No"


def compatibility_confidence_default(row: Mapping[str, Any], context: str) -> str:
    if context != AUDIT_CONTEXT_COMPATIBILITY:
        return ""
    source = _text(row.get(COMPATIBILITY_SOURCE_FIELD))
    if source.casefold() == COMPATIBILITY_SOURCE_PRESS_CAPACITY.casefold():
        return "Press Capacity"
    if source:
        return source
    return "Needs review"


def backfill_audit_context(project_root: str | Path, *, log_activity: bool = True) -> ToolResult:
    from .logging import log_tool_run
    from .paths import resolve_project_paths

    started = datetime.now()
    paths = resolve_project_paths(project_root)
    workbook_path = paths.master_workbook
    if not workbook_path.exists():
        return ToolResult.fail(
            "audit_context_backfill",
            "Audit Context Backfill",
            "Master workbook is missing.",
            errors=[str(workbook_path)],
        )

    workbook = None
    changed_rows: list[AuditContextBackfillResult] = []
    added_headers: list[str] = []
    try:
        backup = backup_file(workbook_path, workbook_path.parent / "_backups")
        workbook = load_workbook(workbook_path)
        if "EOAT Inventory" not in workbook.sheetnames:
            raise ValueError("EOAT Inventory sheet is missing.")
        ws = workbook["EOAT Inventory"]
        headers = worksheet_headers(ws)
        for header in get_expected_headers("EOAT Inventory"):
            if header not in headers:
                ws.cell(row=1, column=len(headers) + 1).value = header
                headers.append(header)
                added_headers.append(header)
        positions = {header: index + 1 for index, header in enumerate(headers)}
        for row_number in range(2, ws.max_row + 1):
            row = {header: ws.cell(row=row_number, column=column).value for header, column in positions.items()}
            if not any(_text(value) for value in row.values()):
                continue
            existing_context = _text(row.get(AUDIT_CONTEXT_FIELD))
            inferred = infer_audit_context(row, preserve_explicit=False)
            physical_verified = ""
            confidence = ""
            row_changed = False
            if not existing_context or existing_context.casefold() in {"n/a", "na"}:
                ws.cell(row=row_number, column=positions[AUDIT_CONTEXT_FIELD]).value = inferred
                row_changed = True
            if PHYSICAL_AUDIT_VERIFIED_FIELD in positions and not _text(row.get(PHYSICAL_AUDIT_VERIFIED_FIELD)):
                physical_verified = physical_audit_verified_default(inferred)
                ws.cell(row=row_number, column=positions[PHYSICAL_AUDIT_VERIFIED_FIELD]).value = physical_verified
                row_changed = True
            if COMPATIBILITY_CONFIDENCE_FIELD in positions and not _text(row.get(COMPATIBILITY_CONFIDENCE_FIELD)):
                confidence = compatibility_confidence_default(row, inferred)
                if confidence:
                    ws.cell(row=row_number, column=positions[COMPATIBILITY_CONFIDENCE_FIELD]).value = confidence
                    row_changed = True
            if row_changed:
                changed_rows.append(
                    AuditContextBackfillResult(
                        row_number=row_number,
                        audit_id=_text(row.get("Audit ID")),
                        inferred_context=inferred,
                        machine=_text(row.get("Press/Machine #")),
                        reason=audit_context_reason(row),
                        physical_verified=physical_verified,
                        compatibility_confidence=confidence,
                    )
                )
        if changed_rows or added_headers:
            workbook.save(workbook_path)
            invalidate_workbook_cache(workbook_path)
        workbook.close()
        workbook = None
    except Exception as exc:
        if workbook is not None:
            workbook.close()
        return ToolResult.fail(
            "audit_context_backfill",
            "Audit Context Backfill",
            "Audit Context backfill failed.",
            errors=[str(exc)],
        )

    report_folder = ensure_directory(paths.validation_reports)
    stamp = started.strftime("%Y-%m-%d_%H%M%S")
    report = safe_write_text(
        report_folder / f"Audit_Context_Backfill_{stamp}.md",
        _backfill_report_markdown(workbook_path, backup, added_headers, changed_rows),
        overwrite=False,
    )
    result = ToolResult.ok(
        "audit_context_backfill",
        "Audit Context Backfill",
        f"Backfilled Audit Context for {len(changed_rows)} row(s).",
        details=[
            f"Workbook backup: {backup}",
            f"Migration report: {report}",
            f"Headers added: {', '.join(added_headers) if added_headers else '(none)'}",
        ],
        files_created=[str(backup), str(report)],
        files_modified=[str(workbook_path)] if changed_rows or added_headers else [],
        output_reports=[str(report)],
        metrics={
            "rows_changed": len(changed_rows),
            "headers_added": len(added_headers),
            "installed_context_count": sum(1 for row in changed_rows if row.inferred_context == AUDIT_CONTEXT_INSTALLED),
            "bench_context_count": sum(1 for row in changed_rows if row.inferred_context == AUDIT_CONTEXT_BENCH),
            "compatibility_context_count": sum(
                1 for row in changed_rows if row.inferred_context == AUDIT_CONTEXT_COMPATIBILITY
            ),
            "historical_context_count": sum(
                1 for row in changed_rows if row.inferred_context == AUDIT_CONTEXT_HISTORICAL
            ),
            "needs_review_context_count": sum(
                1 for row in changed_rows if row.inferred_context == AUDIT_CONTEXT_NEEDS_REVIEW
            ),
        },
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def _is_compatibility_marker(row: Mapping[str, Any]) -> bool:
    entry_type = _text(row.get(ENTRY_TYPE_FIELD)).casefold()
    status = _text(row.get("Status")).casefold()
    source = _text(row.get(COMPATIBILITY_SOURCE_FIELD)).casefold()
    return (
        entry_type == ENTRY_TYPE_COMPATIBLE.casefold()
        or status == ENTRY_TYPE_COMPATIBLE.casefold()
        or "compat" in source
        or source == COMPATIBILITY_SOURCE_PRESS_CAPACITY.casefold()
    )


def _is_historical_marker(row: Mapping[str, Any]) -> bool:
    haystack = " ".join(
        _text(row.get(field)).casefold()
        for field in (ENTRY_TYPE_FIELD, COMPATIBILITY_SOURCE_FIELD, SOURCE_AUDIT_ID_FIELD, "Status", "Notes")
    )
    return "historical" in haystack or "imported" in haystack or "legacy import" in haystack


def _machine_value(row: Mapping[str, Any]) -> str:
    for field in ("Press/Machine #", "Machine #", "Machine Number", "Machine No.", "Press"):
        value = _text(row.get(field))
        if value:
            return value
    return ""


def _backfill_report_markdown(
    workbook_path: Path,
    backup: Path,
    added_headers: list[str],
    changed_rows: list[AuditContextBackfillResult],
) -> str:
    lines = [
        "# Audit Context Backfill Report",
        "",
        f"- Workbook: {workbook_path}",
        f"- Backup: {backup}",
        f"- Rows changed: {len(changed_rows)}",
        f"- Headers added: {', '.join(added_headers) if added_headers else '(none)'}",
        "",
        "## Inference Rules",
        "- Compatibility rows: Entry Type or compatibility source marks the row as Compatibility row.",
        "- Bench rows: Machine value is blank, N/A, Not Installed, or equivalent.",
        "- Installed rows: Machine value is populated and no compatibility/historical marker is present.",
        "- Uncertain rows: Needs review.",
        "",
        "## Changed Rows",
    ]
    if not changed_rows:
        lines.append("No rows required changes.")
        return "\n".join(lines) + "\n"
    lines.extend(
        [
            "| Row | Audit ID | Machine # | Audit Context | Physical Audit Verified | Compatibility Confidence | Reason |",
            "| ---: | --- | --- | --- | --- | --- | --- |",
        ]
    )
    for row in changed_rows:
        lines.append(
            "| "
            + " | ".join(
                [
                    str(row.row_number),
                    row.audit_id or "N/A",
                    row.machine or "N/A",
                    row.inferred_context,
                    row.physical_verified or "",
                    row.compatibility_confidence or "",
                    row.reason,
                ]
            )
            + " |"
        )
    return "\n".join(lines) + "\n"


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


__all__ = [
    "EOAT_DOCUMENTATION_FIELDS",
    "INSTALLATION_ONLY_FIELDS",
    "INSTALLATION_READINESS_FIELDS",
    "INSTALLED_CELL_VALIDATION_FIELDS",
    "MACHINE_CONTEXT_FIELDS",
    "AuditContextBackfillResult",
    "audit_context_reason",
    "backfill_audit_context",
    "compatibility_confidence_default",
    "infer_audit_context",
    "is_bench_audit_context",
    "is_compatibility_audit_context",
    "is_installed_audit_context",
    "machine_value_is_missing",
    "normalize_context_value",
    "physical_audit_verified_default",
]
