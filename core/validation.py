from __future__ import annotations

import json
import sqlite3
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook

from .audit.relationships import is_compatibility_row, is_physical_audit_row, source_audit_for_compatibility_row
from .audit.uninstalled import UNINSTALLED_MACHINE_AND_ROBOT_CONTEXT_FIELDS, is_uninstalled_eoat_audit
from .audit_by_press import AUDIT_BY_PRESS_SHEET, audit_by_press_last_refreshed
from .audit_constants import (
    AUDIT_CONTEXT_BENCH,
    AUDIT_CONTEXT_COMPATIBILITY,
    AUDIT_CONTEXT_FIELD,
    AUDIT_CONTEXT_HISTORICAL,
    AUDIT_CONTEXT_INSTALLED,
    AUDIT_CONTEXT_NEEDS_REVIEW,
    AUDIT_CONTEXT_VALUES,
    AUTOFILLED_COMPATIBILITY_METADATA_FIELDS,
    COMPATIBILITY_CONFIDENCE_FIELD,
    COMPATIBILITY_SOURCE_FIELD,
    CYLINDER_COUNT_FIELD,
    CYLINDER_TYPE_FIELD,
    CYLINDER_TYPE_VALUES,
    ENTRY_TYPE_FIELD,
    MANUAL_COMPLETION_OVERRIDE_FIELD,
    MANUAL_COMPLETION_OVERRIDE_FIELDS,
    PHYSICAL_AUDIT_VERIFIED_FIELD,
    SOURCE_AUDIT_ID_FIELD,
)
from .audit_context import infer_audit_context
from .audit_entries import (
    AUDIT_IMPORTANT_FIELDS,
    AUDIT_REQUIRED_FIELDS,
    CLEANROOM_DROPDOWN_VALUES,
    CONNECTION_TYPE_FIELD,
    CONNECTION_TYPE_VALUES,
    CURRENT_WORKBOOK_SCHEMA_VERSION,
    EOAT_MOVES_FIELD,
    EOAT_MOVES_VALUES,
    EOAT_TYPE_DROPDOWN_VALUES,
    LEGACY_VACUUM_CUPS_FIELD,
    NA_VALUE,
    NUMBER_OF_PARTS_PICKED_FIELD,
    WORKBOOK_METADATA_SHEET,
    apply_part_present_sensor_defaults,
    audit_field_applies,
    is_na_value,
)
from .audit_field_rules import (
    ELECTRICAL_DETAIL_FIELDS,
    ELECTRICAL_WIRING_PRESENT_FIELD,
    entry_type_requirements,
    field_applies,
    hybrid_completeness_warnings,
    is_meaningful_value,
    semantic_consistency_warnings,
)
from .compatibility_health import validate_compatibility_health
from .constants import TOOLKIT_ROOT
from .eoat_ids import EOAT_ASSEMBLY_ID_FIELD, is_valid_eoat_assembly_id, normalize_eoat_assembly_id
from .git_activity import is_git_repo
from .gripper_fields import CUP_COUNT_FIELD, GRIPPER_COUNT_FIELD, GRIPPER_TYPE_FIELD, GRIPPER_TYPE_VALUES
from .logging import log_tool_run
from .paths import resolve_project_paths
from .photo_evidence import validate_photo_evidence
from .result import ToolResult
from .robot_info import ROBOT_INFO_SHEET, robot_info_workbook_path, validate_robot_info_workbook
from .safe_files import ensure_directory, safe_write_text
from .schedule import available_schedule_weeks
from .validation_findings import (
    ValidationFinding,
    ValidationSeverity,
    attach_findings,
    make_finding,
    write_validation_json_report,
)
from .workbook_io import row_dicts as workbook_row_dicts
from .workbook_schema import get_expected_headers, get_expected_sheets, get_key_inventory_headers, load_workbook_schema
from .workbook_truth import analyze_truth_from_rows

MAJOR_AUDIT_COLUMNS = {
    "Audit ID",
    "Audit Date",
    "Auditor",
    "Plant/Area",
    AUDIT_CONTEXT_FIELD,
    "Press/Machine #",
    "Tool #",
    "EOAT Type",
    CONNECTION_TYPE_FIELD,
    "Cleanroom/Non-Cleanroom",
    "Status",
    "Priority",
    "Known Issues",
    *AUDIT_REQUIRED_FIELDS,
    *AUDIT_IMPORTANT_FIELDS,
} - {EOAT_MOVES_FIELD}

AUDIT_DROPDOWN_ALLOWED_VALUES = {
    "EOAT Type": {*EOAT_TYPE_DROPDOWN_VALUES, NA_VALUE},
    EOAT_MOVES_FIELD: set(EOAT_MOVES_VALUES),
    CONNECTION_TYPE_FIELD: {*CONNECTION_TYPE_VALUES, NA_VALUE},
    CYLINDER_TYPE_FIELD: {*CYLINDER_TYPE_VALUES, NA_VALUE},
    GRIPPER_TYPE_FIELD: {*GRIPPER_TYPE_VALUES, NA_VALUE},
    "Cleanroom/Non-Cleanroom": {*CLEANROOM_DROPDOWN_VALUES, NA_VALUE},
    MANUAL_COMPLETION_OVERRIDE_FIELD: {"Yes", "No", NA_VALUE},
    AUDIT_CONTEXT_FIELD: {*AUDIT_CONTEXT_VALUES, NA_VALUE},
    PHYSICAL_AUDIT_VERIFIED_FIELD: {"Yes", "No", NA_VALUE},
    COMPATIBILITY_CONFIDENCE_FIELD: {
        "Press Capacity",
        "Manual",
        "Existing EOAT",
        "Imported",
        "Needs review",
        NA_VALUE,
    },
}
AUDIT_NUMERIC_FIELDS = {NUMBER_OF_PARTS_PICKED_FIELD, CYLINDER_COUNT_FIELD, CUP_COUNT_FIELD, GRIPPER_COUNT_FIELD}

BLANK_CELL_VALIDATION_IGNORED_FIELDS = (
    AUTOFILLED_COMPATIBILITY_METADATA_FIELDS
    | frozenset(MANUAL_COMPLETION_OVERRIDE_FIELDS)
    | frozenset({PHYSICAL_AUDIT_VERIFIED_FIELD, COMPATIBILITY_CONFIDENCE_FIELD})
)
BLANK_CELL_VALIDATION_IGNORED_FIELD_LABEL = (
    "compatibility/source metadata, physical audit verification metadata, and manual completion override metadata"
)

FIX_CLEAR_STALE_HIDDEN_NA = "clear_stale_hidden_na"
FIX_REPAIR_LEGACY_HEADERS = "repair_legacy_headers"
FIX_REFRESH_GENERATED_VIEWS = "refresh_generated_views"
FIX_NORMALIZE_DROPDOWN_CASING = "normalize_dropdown_casing"
FIX_CREATE_MISSING_REPORT_FOLDERS = "create_missing_report_folders"

SCHEMA_VERSION_UNKNOWN = "Unknown"
LEGACY_GRIPPER_SIZE_FIELD = "Gripper Size"


def validate_project_foundation(
    project_root: str | Path,
    *,
    include_open_item_cache_references: bool = True,
) -> ToolResult:
    started = time.perf_counter()
    paths = resolve_project_paths(project_root)
    details: list[str] = []
    warnings: list[str] = []
    errors: list[str] = []
    metrics: dict[str, object] = {}
    findings: list[ValidationFinding] = []

    if not paths.project_root.exists():
        findings.append(
            make_finding(
                ValidationSeverity.BLOCKER,
                "project_foundation",
                f"Missing project root: {paths.project_root}",
                expected_behavior="The selected project root should exist before validation runs.",
                recommended_action="Select an existing EOAT project root in Settings.",
                source_validator="foundation",
            )
        )
        return attach_findings(
            ToolResult.fail(
                "workbook_validator",
                "EOAT Project Foundation Validation",
                "Project root does not exist.",
                errors=[f"Missing project root: {paths.project_root}"],
                duration_seconds=time.perf_counter() - started,
            ),
            findings,
        )

    details.append(f"Project root exists: {paths.project_root}")

    for folder in paths.expected_numbered_folders():
        if folder.exists():
            details.append(f"Found folder: {folder.name}")
        else:
            warnings.append(f"Missing expected folder: {folder.name}")

    for folder_name, folder in {
        "daily reports": paths.daily_reports,
        "weekly reports": paths.weekly_reports,
        "validation reports": paths.validation_reports,
        "activity logs": paths.activity_logs,
    }.items():
        if folder.exists():
            details.append(f"Found {folder_name} folder: {folder}")
        else:
            message = f"Missing {folder_name} folder: {folder}"
            warnings.append(message)
            findings.append(
                make_finding(
                    ValidationSeverity.AUTO_FIXABLE,
                    "project_foundation",
                    message,
                    expected_behavior="Local report/log folders should exist before reports are generated.",
                    recommended_action="Preview and apply the Create Missing Report Folders safe fix.",
                    fix_available=True,
                    fix_id=FIX_CREATE_MISSING_REPORT_FOLDERS,
                    source_validator="foundation",
                )
            )

    workbook_path = paths.master_workbook
    if not workbook_path.exists():
        message = f"Missing master workbook: {workbook_path}"
        errors.append(message)
        findings.append(
            make_finding(
                ValidationSeverity.BLOCKER,
                "workbook_schema",
                message,
                expected_behavior="The master tracker workbook should exist at the project workbook path.",
                recommended_action="Restore the workbook from backup or select the correct project root.",
                source_validator="foundation",
            )
        )
        return attach_findings(
            ToolResult.fail(
                "workbook_validator",
                "EOAT Project Foundation Validation",
                "Master workbook is missing.",
                details=details,
                warnings=warnings,
                errors=errors,
                metrics=metrics,
                duration_seconds=time.perf_counter() - started,
            ),
            findings,
        )

    details.append(f"Found master workbook: {workbook_path}")
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    except Exception as exc:
        message = f"Could not open workbook with openpyxl: {exc}"
        errors.append(message)
        findings.append(
            make_finding(
                ValidationSeverity.BLOCKER,
                "workbook_lock",
                message,
                expected_behavior="Validation needs read access to the master workbook.",
                recommended_action="Close the workbook in Excel or check file permissions, then retry validation.",
                source_validator="foundation",
            )
        )
        return attach_findings(
            ToolResult.fail(
                "workbook_validator",
                "EOAT Project Foundation Validation",
                "Master workbook exists but could not be opened.",
                details=details,
                warnings=warnings,
                errors=errors,
                metrics=metrics,
                duration_seconds=time.perf_counter() - started,
            ),
            findings,
        )

    expected_sheets = get_expected_sheets()
    missing_sheets = [sheet for sheet in expected_sheets if sheet not in workbook.sheetnames]
    metrics["expected_sheet_count"] = len(expected_sheets)
    metrics["actual_sheet_count"] = len(workbook.sheetnames)
    expected_schema_version = str(load_workbook_schema().get("version") or CURRENT_WORKBOOK_SCHEMA_VERSION)
    workbook_schema_version = _workbook_schema_version(workbook)
    metrics["expected_workbook_schema_version"] = expected_schema_version
    metrics["workbook_schema_version"] = workbook_schema_version
    if workbook_schema_version == SCHEMA_VERSION_UNKNOWN:
        details.append(
            f"Workbook schema version is unknown; validating headers against template schema {expected_schema_version}."
        )
        findings.append(
            make_finding(
                ValidationSeverity.INFO,
                "workbook_schema",
                f"Workbook schema version metadata is unknown; header validation is using template schema {expected_schema_version}.",
                expected_behavior="Schema migrations should preserve local data and add explicit version metadata when the workbook supports it.",
                recommended_action="No automatic workbook rewrite is performed during validation. Run the schema repair/migration path only after previewing changes.",
                source_validator="foundation",
            )
        )
    elif workbook_schema_version != expected_schema_version:
        message = f"Workbook schema version {workbook_schema_version} differs from expected template schema {expected_schema_version}."
        warnings.append(message)
        findings.append(
            make_finding(
                ValidationSeverity.WARNING,
                "workbook_schema",
                message,
                expected_behavior="Workbook schema metadata should match the local template version after migrations.",
                recommended_action="Preview schema migration before applying any workbook rewrite.",
                source_validator="foundation",
            )
        )
    if missing_sheets:
        for sheet in missing_sheets:
            message = f"Missing expected sheet: {sheet}"
            errors.append(message)
            findings.append(
                make_finding(
                    ValidationSeverity.ERROR,
                    "workbook_schema",
                    message,
                    sheet_name=sheet,
                    expected_behavior="All expected workbook sheets should be present.",
                    recommended_action="Restore the missing sheet from the template or a trusted backup.",
                    source_validator="foundation",
                )
            )
    else:
        details.append("All expected workbook sheets are present.")
    if AUDIT_BY_PRESS_SHEET not in workbook.sheetnames:
        message = "Audit by Press view missing or stale; refresh generated view."
        warnings.append(message)
        findings.append(
            make_finding(
                ValidationSeverity.AUTO_FIXABLE,
                "generated_view",
                message,
                sheet_name=AUDIT_BY_PRESS_SHEET,
                expected_behavior="Generated Audit by Press view should be present and carry a refresh timestamp.",
                recommended_action="Preview and apply the Refresh Generated Views safe fix.",
                fix_available=True,
                fix_id="refresh_generated_views",
                source_validator="foundation",
            )
        )
    elif audit_by_press_last_refreshed(workbook) is None:
        message = "Audit by Press view missing or stale; refresh generated view."
        warnings.append(message)
        findings.append(
            make_finding(
                ValidationSeverity.AUTO_FIXABLE,
                "generated_view",
                message,
                sheet_name=AUDIT_BY_PRESS_SHEET,
                expected_behavior="Generated Audit by Press view should carry a refresh timestamp.",
                recommended_action="Preview and apply the Refresh Generated Views safe fix.",
                fix_available=True,
                fix_id="refresh_generated_views",
                source_validator="foundation",
            )
        )
    else:
        details.append("Audit by Press generated view is present.")

    if "EOAT Inventory" in workbook.sheetnames:
        ws = workbook["EOAT Inventory"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [str(value) for value in header_row if value is not None]
        required_headers = get_expected_headers("EOAT Inventory")
        missing_full_headers = [header for header in required_headers if header not in headers]
        missing_key_headers = [header for header in get_key_inventory_headers() if header not in headers]
        missing_major_headers = [
            header for header in MAJOR_AUDIT_COLUMNS if header in required_headers and header not in headers
        ]
        schema_upgrade_headers = []
        if ELECTRICAL_WIRING_PRESENT_FIELD in missing_full_headers:
            schema_upgrade_headers.append(ELECTRICAL_WIRING_PRESENT_FIELD)
        missing_detail_headers = [
            header
            for header in missing_full_headers
            if header not in missing_major_headers and header not in schema_upgrade_headers
        ]
        unexpected_headers = [
            header
            for header in headers
            if header and header not in required_headers and header not in {LEGACY_VACUUM_CUPS_FIELD, LEGACY_GRIPPER_SIZE_FIELD}
        ]
        duplicate_headers = sorted({header for header in headers if header and headers.count(header) > 1})
        legacy_vacuum_header_present = LEGACY_VACUUM_CUPS_FIELD in headers
        has_parts_picked_header = NUMBER_OF_PARTS_PICKED_FIELD in headers
        metrics["eoat_inventory_header_count"] = len(headers)
        metrics["missing_full_inventory_header_count"] = len(missing_full_headers)
        metrics["missing_key_inventory_header_count"] = len(missing_key_headers)
        metrics["missing_major_inventory_header_count"] = len(missing_major_headers)
        metrics["unexpected_inventory_header_count"] = len(unexpected_headers)
        metrics["duplicate_inventory_header_count"] = len(duplicate_headers)
        metrics["legacy_vacuum_cups_header_present"] = int(legacy_vacuum_header_present)
        if missing_major_headers:
            for header in missing_major_headers:
                message = f"Missing major EOAT Inventory header: {header}"
                errors.append(message)
                findings.append(
                    make_finding(
                        ValidationSeverity.ERROR,
                        "workbook_schema",
                        message,
                        sheet_name="EOAT Inventory",
                        row_number=1,
                        column_name=header,
                        expected_behavior="Major EOAT Inventory headers must be present for audit save/load workflows.",
                        recommended_action="Preview and apply the Repair Legacy Headers safe fix or restore the header from the workbook template.",
                        fix_available=True,
                        fix_id="repair_legacy_headers",
                        source_validator="foundation",
                    )
                )
        elif missing_key_headers:
            for header in missing_key_headers:
                message = f"Missing key EOAT Inventory header: {header}"
                warnings.append(message)
                findings.append(
                    make_finding(
                        ValidationSeverity.AUTO_FIXABLE,
                        "workbook_schema",
                        message,
                        sheet_name="EOAT Inventory",
                        row_number=1,
                        column_name=header,
                        expected_behavior="Key EOAT Inventory headers should be present.",
                        recommended_action="Preview and apply the Repair Legacy Headers safe fix.",
                        fix_available=True,
                        fix_id="repair_legacy_headers",
                        source_validator="foundation",
                    )
                )
        else:
            details.append("Key EOAT Inventory headers are present.")
        if missing_detail_headers:
            for header in missing_detail_headers:
                message = f"Missing detail EOAT Inventory header: {header}"
                warnings.append(message)
                findings.append(
                    make_finding(
                        ValidationSeverity.AUTO_FIXABLE,
                        "workbook_schema",
                        message,
                        sheet_name="EOAT Inventory",
                        row_number=1,
                        column_name=header,
                        expected_behavior="Detail EOAT Inventory headers should match the current workbook schema.",
                        recommended_action="Preview and apply the Repair Legacy Headers safe fix.",
                        fix_available=True,
                        fix_id="repair_legacy_headers",
                        source_validator="foundation",
                    )
                )
        if ELECTRICAL_WIRING_PRESENT_FIELD in schema_upgrade_headers:
            message = (
                "Workbook is missing Electrical/Wiring Present?. Run Repair Workbook Schema to upgrade old workbooks."
            )
            warnings.append(message)
            findings.append(
                make_finding(
                    ValidationSeverity.AUTO_FIXABLE,
                    "workbook_schema",
                    message,
                    sheet_name="EOAT Inventory",
                    row_number=1,
                    column_name=ELECTRICAL_WIRING_PRESENT_FIELD,
                    expected_behavior="Current workbooks include Electrical/Wiring Present? to control electrical field applicability.",
                    recommended_action="Preview and apply the Repair Legacy Headers safe fix.",
                    fix_available=True,
                    fix_id="repair_legacy_headers",
                    source_validator="foundation",
                )
            )
        if legacy_vacuum_header_present and has_parts_picked_header:
            message = (
                "EOAT Inventory has both Number of Vacuum Cups and Number of Parts Picked. "
                "Run Repair Workbook Schema to merge values into Number of Parts Picked and remove the legacy header."
            )
            warnings.append(message)
            findings.append(
                make_finding(
                    ValidationSeverity.AUTO_FIXABLE,
                    "workbook_schema",
                    message,
                    sheet_name="EOAT Inventory",
                    row_number=1,
                    column_name=LEGACY_VACUUM_CUPS_FIELD,
                    expected_behavior="Legacy vacuum-cup header should be migrated to Number of Parts Picked.",
                    recommended_action="Preview and apply the Repair Legacy Headers safe fix.",
                    fix_available=True,
                    fix_id="repair_legacy_headers",
                    source_validator="foundation",
                )
            )
        elif legacy_vacuum_header_present:
            message = (
                "EOAT Inventory still uses legacy header Number of Vacuum Cups. "
                "Run Repair Workbook Schema to rename it to Number of Parts Picked."
            )
            warnings.append(message)
            findings.append(
                make_finding(
                    ValidationSeverity.AUTO_FIXABLE,
                    "workbook_schema",
                    message,
                    sheet_name="EOAT Inventory",
                    row_number=1,
                    column_name=LEGACY_VACUUM_CUPS_FIELD,
                    expected_behavior="Legacy vacuum-cup header should be renamed to Number of Parts Picked.",
                    recommended_action="Preview and apply the Repair Legacy Headers safe fix.",
                    fix_available=True,
                    fix_id="repair_legacy_headers",
                    source_validator="foundation",
                )
            )
        if LEGACY_GRIPPER_SIZE_FIELD in headers:
            details.append("Legacy Gripper Size column is present and ignored for backward compatibility.")
        if unexpected_headers:
            for header in unexpected_headers:
                message = f"Unexpected EOAT Inventory header: {header}"
                warnings.append(message)
                findings.append(
                    make_finding(
                        ValidationSeverity.WARNING,
                        "workbook_schema",
                        message,
                        sheet_name="EOAT Inventory",
                        row_number=1,
                        column_name=header,
                        expected_behavior="Workbook headers should match the expected local schema.",
                        recommended_action="Review whether this is a local-only note column or an accidental workbook edit.",
                        source_validator="foundation",
                    )
                )
        if duplicate_headers:
            for header in duplicate_headers:
                message = f"Duplicate EOAT Inventory header: {header}"
                warnings.append(message)
                findings.append(
                    make_finding(
                        ValidationSeverity.ERROR,
                        "workbook_schema",
                        message,
                        sheet_name="EOAT Inventory",
                        row_number=1,
                        column_name=header,
                        expected_behavior="Each EOAT Inventory header should appear once.",
                        recommended_action="Review the duplicate header manually before saving more audit rows.",
                        source_validator="foundation",
                    )
                )
        if any(header in headers for header in BLANK_CELL_VALIDATION_IGNORED_FIELDS):
            details.append(
                f"Blank {BLANK_CELL_VALIDATION_IGNORED_FIELD_LABEL} cells are intentionally ignored during "
                "blank-cell validation because they are autofilled or system-managed metadata; "
                "the columns still remain part of the workbook schema."
            )
        valid_audit_ids, valid_machines = _inventory_identity_sets(ws, headers)
        inventory_warnings, inventory_metrics, inventory_findings = _validate_inventory_rows(
            ws, headers, paths.project_root
        )
        warnings.extend(inventory_warnings)
        metrics.update(inventory_metrics)
        findings.extend(inventory_findings)
        metrics["known_audit_id_count"] = len(valid_audit_ids)
        metrics["known_machine_count"] = len(valid_machines)
    else:
        message = "EOAT Inventory sheet is missing, so headers could not be checked."
        errors.append(message)
        findings.append(
            make_finding(
                ValidationSeverity.ERROR,
                "workbook_schema",
                message,
                sheet_name="EOAT Inventory",
                expected_behavior="EOAT Inventory sheet must exist for audit save/load workflows.",
                recommended_action="Restore EOAT Inventory from the template or a trusted backup.",
                source_validator="foundation",
            )
        )

    workbook.close()
    if "valid_audit_ids" in locals() and "valid_machines" in locals():
        orphan_warnings, orphan_metrics, orphan_findings = _validate_orphan_local_references(
            paths.project_root,
            valid_audit_ids,
            valid_machines,
            include_open_item_cache_references=include_open_item_cache_references,
        )
        warnings.extend(orphan_warnings)
        metrics.update(orphan_metrics)
        findings.extend(orphan_findings)
    compatibility_findings = validate_compatibility_health(workbook_path)
    findings.extend(compatibility_findings)
    metrics["compatibility_health_finding_count"] = len(compatibility_findings)

    evidence_warnings, evidence_metrics, evidence_findings = validate_photo_evidence(paths.project_root)
    warnings.extend(evidence_warnings)
    metrics.update(evidence_metrics)
    findings.extend(evidence_findings)
    if evidence_metrics.get("photo_evidence_audit_count", 0):
        details.append("Photo evidence coverage checked.")

    robot_warnings, robot_errors, robot_metrics = validate_robot_info_workbook(paths.project_root)
    warnings.extend(robot_warnings)
    errors.extend(robot_errors)
    metrics.update(robot_metrics)
    if not robot_warnings and not robot_errors:
        details.append("Robot Info workbook is present and valid.")

    weeks = available_schedule_weeks(paths.project_root)
    metrics["schedule_week_count"] = len(weeks)
    if weeks:
        details.append(f"Found schedule/task progress week files for week(s): {', '.join(str(week) for week in weeks)}")
    else:
        warnings.append("No project schedule or task progress week files found.")

    readme_path = paths.project_root / "README.md"
    if readme_path.exists():
        details.append(f"Found project README: {readme_path}")
    else:
        warnings.append(f"Missing project README: {readme_path}")

    usage_path = TOOLKIT_ROOT / "USAGE.md"
    if usage_path.exists():
        details.append(f"Found toolkit usage guide: {usage_path}")
    else:
        warnings.append(f"Missing toolkit usage guide: {usage_path}")

    git_repo, git_warning = is_git_repo(paths.project_root)
    metrics["git_repo_detected"] = git_repo
    if git_repo:
        details.append("Git repository detected.")
    elif git_warning:
        warnings.append(f"Git repository not detected or Git unavailable: {git_warning}")

    summary = "Project foundation validation completed."
    if warnings and not errors:
        summary = "Project foundation validation completed with warnings."
    if errors:
        summary = "Project foundation validation failed."
    return attach_findings(
        ToolResult(
            tool_id="workbook_validator",
            tool_name="EOAT Project Foundation Validation",
            success=not errors,
            summary=summary,
            details=details,
            warnings=warnings,
            errors=errors,
            metrics=metrics,
            duration_seconds=time.perf_counter() - started,
        ),
        findings,
    )


def _validate_inventory_rows(
    ws, headers: list[str], project_root: str | Path
) -> tuple[list[str], dict[str, int], list[ValidationFinding]]:
    warnings: list[str] = []
    findings: list[ValidationFinding] = []
    metrics = {
        "duplicate_audit_id_count": 0,
        "blank_saved_audit_cell_count": 0,
        "major_na_cell_count": 0,
        "missing_applicable_major_cell_count": 0,
        "non_applicable_na_cell_count": 0,
        "stale_hidden_value_count": 0,
        "hybrid_warning_count": 0,
        "semantic_warning_count": 0,
        "physical_audit_row_count": 0,
        "compatible_row_count": 0,
        "duplicate_physical_row_count": 0,
        "compatibility_missing_source_audit_id_count": 0,
        "compatibility_missing_source_count": 0,
        "physical_row_with_compatibility_metadata_count": 0,
        "compatibility_truth_warning_count": 0,
        "broken_photo_link_count": 0,
        "photos_yes_without_link_count": 0,
        "photo_link_while_no_count": 0,
        "indexed_photos_while_status_not_yes_count": 0,
        "robot_side_circuit_mismatch_count": 0,
        "invalid_dropdown_value_count": 0,
        "dropdown_casing_fixable_count": 0,
        "invalid_numeric_value_count": 0,
        "missing_eoat_assembly_id_count": 0,
        "invalid_eoat_assembly_id_count": 0,
        "audit_row_count": 0,
        "installed_audit_context_count": 0,
        "bench_audit_context_count": 0,
        "compatibility_audit_context_count": 0,
        "historical_audit_context_count": 0,
        "needs_review_audit_context_count": 0,
    }
    if not headers:
        return warnings, metrics, findings

    header_positions = {header: index for index, header in enumerate(headers)}
    missing_electrical_wiring_control = ELECTRICAL_WIRING_PRESENT_FIELD not in header_positions
    audit_ids: dict[str, int] = {}
    duplicate_ids: set[str] = set()
    blank_cells = 0
    major_na_examples: set[str] = set()
    non_applicable_na_examples: list[str] = []
    stale_hidden_value_examples: list[str] = []
    hybrid_warning_examples: list[str] = []
    semantic_warning_examples: list[str] = []
    invalid_dropdown_examples: list[str] = []
    dropdown_casing_examples: list[str] = []
    invalid_numeric_examples: list[str] = []
    missing_eoat_id_examples: list[str] = []
    invalid_eoat_id_examples: list[str] = []
    missing_eoat_moves_examples: list[str] = []
    duplicate_physical_examples: list[str] = []
    compatibility_warning_examples: list[str] = []
    photo_warning_examples: list[str] = []
    robot_mismatch_examples: list[str] = []
    physical_identity_rows: dict[tuple[str, str, str], int] = {}
    inventory_rows_for_truth: list[dict[str, object]] = []
    source_eoat_moves_by_audit_id: dict[str, str] = {}
    major_na_seen: set[tuple[int, str]] = set()
    indexed_photo_audit_ids, indexed_photo_machines = _indexed_photo_reference_sets(project_root)

    def add_major_na(row_number: int, field_name: str, row_data: dict[str, object]) -> None:
        if (row_number, field_name) in major_na_seen:
            return
        major_na_seen.add((row_number, field_name))
        major_na_examples.add(f"row {row_number} {field_name}")
        findings.append(
            make_finding(
                ValidationSeverity.WARNING,
                "audit_data",
                f"Applicable major EOAT Inventory cell is blank or contains {NA_VALUE}: row {row_number} {field_name}",
                sheet_name="EOAT Inventory",
                row_number=row_number,
                column_name=field_name,
                audit_id=_cell_text(row_data.get("Audit ID")),
                machine_number=_cell_text(row_data.get("Press/Machine #")),
                current_value=row_data.get(field_name),
                expected_behavior="Applicable major audit fields should contain a verified value or Unknown / Not Checked when appropriate.",
                recommended_action="Open the audit row and complete the field; do not treat N/A as complete when the field applies.",
                source_validator="inventory_rows",
            )
        )

    if EOAT_MOVES_FIELD in header_positions:
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {header: row[index] for header, index in header_positions.items() if index < len(row)}
            audit_id = _cell_text(row_data.get("Audit ID"))
            if audit_id:
                source_eoat_moves_by_audit_id[audit_id] = _cell_text(row_data.get(EOAT_MOVES_FIELD))

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {header: row[index] for header, index in header_positions.items() if index < len(row)}
        if not _is_audit_data_row(row_data):
            continue
        row_data = apply_part_present_sensor_defaults(row_data)
        audit_context = infer_audit_context(row_data)
        if audit_context == AUDIT_CONTEXT_INSTALLED:
            metrics["installed_audit_context_count"] += 1
        elif audit_context == AUDIT_CONTEXT_BENCH:
            metrics["bench_audit_context_count"] += 1
        elif audit_context == AUDIT_CONTEXT_COMPATIBILITY:
            metrics["compatibility_audit_context_count"] += 1
        elif audit_context == AUDIT_CONTEXT_HISTORICAL:
            metrics["historical_audit_context_count"] += 1
        elif audit_context == AUDIT_CONTEXT_NEEDS_REVIEW:
            metrics["needs_review_audit_context_count"] += 1
        truth_row = dict(row_data)
        truth_row["_row_index"] = row_number
        inventory_rows_for_truth.append(truth_row)
        metrics["audit_row_count"] += 1
        entry_type = _cell_text(row_data.get(ENTRY_TYPE_FIELD)).lower()
        if is_compatibility_row(row_data):
            metrics["compatible_row_count"] += 1
        else:
            metrics["physical_audit_row_count"] += 1
        audit_id = _cell_text(row_data.get("Audit ID"))
        if audit_id and not is_na_value(audit_id):
            if audit_id in audit_ids:
                duplicate_ids.add(audit_id)
                findings.append(
                    make_finding(
                        ValidationSeverity.ERROR,
                        "audit_data",
                        f"Duplicate Audit ID value: {audit_id}",
                        sheet_name="EOAT Inventory",
                        row_number=row_number,
                        column_name="Audit ID",
                        audit_id=audit_id,
                        machine_number=_cell_text(row_data.get("Press/Machine #")),
                        current_value=audit_id,
                        expected_behavior="Audit IDs must be unique across EOAT Inventory rows.",
                        recommended_action=f"Compare this row with row {audit_ids[audit_id]} and assign a unique Audit ID before saving updates.",
                        source_validator="inventory_rows",
                    )
                )
            else:
                audit_ids[audit_id] = row_number
        elif "Audit ID" in headers:
            add_major_na(row_number, "Audit ID", row_data)

        if EOAT_ASSEMBLY_ID_FIELD in header_positions:
            eoat_id = normalize_eoat_assembly_id(row_data.get(EOAT_ASSEMBLY_ID_FIELD))
            if not eoat_id:
                metrics["missing_eoat_assembly_id_count"] += 1
                missing_eoat_id_examples.append(f"row {row_number}")
                findings.append(
                    make_finding(
                        ValidationSeverity.AUTO_FIXABLE,
                        "eoat_identity",
                        f"Missing EOAT Assembly ID on EOAT Inventory row {row_number}.",
                        sheet_name="EOAT Inventory",
                        row_number=row_number,
                        column_name=EOAT_ASSEMBLY_ID_FIELD,
                        audit_id=audit_id,
                        machine_number=_cell_text(row_data.get("Press/Machine #")),
                        current_value=row_data.get(EOAT_ASSEMBLY_ID_FIELD),
                        expected_behavior="EOAT Assembly ID identifies the physical EOAT; old blank rows can be assigned safely.",
                        recommended_action="Run Assign Missing EOAT IDs.",
                        fix_available=True,
                        fix_id="assign_missing_eoat_ids",
                        source_validator="inventory_rows",
                    )
                )
            elif not is_valid_eoat_assembly_id(eoat_id):
                metrics["invalid_eoat_assembly_id_count"] += 1
                invalid_eoat_id_examples.append(f"row {row_number} {eoat_id}")
                findings.append(
                    make_finding(
                        ValidationSeverity.WARNING,
                        "eoat_identity",
                        f"Invalid EOAT Assembly ID format on EOAT Inventory row {row_number}: {eoat_id}.",
                        sheet_name="EOAT Inventory",
                        row_number=row_number,
                        column_name=EOAT_ASSEMBLY_ID_FIELD,
                        audit_id=audit_id,
                        machine_number=_cell_text(row_data.get("Press/Machine #")),
                        current_value=eoat_id,
                        expected_behavior="EOAT Assembly ID should use format P4-EOAT-0001.",
                        recommended_action="Correct the EOAT Assembly ID after confirming the physical EOAT.",
                        source_validator="inventory_rows",
                    )
                )

        if is_physical_audit_row(row_data):
            source_id = _cell_text(row_data.get(SOURCE_AUDIT_ID_FIELD))
            compatibility_source = _cell_text(row_data.get(COMPATIBILITY_SOURCE_FIELD))
            if source_id or compatibility_source:
                metrics["physical_row_with_compatibility_metadata_count"] += 1
                message = f"Physical audit row contains compatibility metadata: row {row_number}"
                compatibility_warning_examples.append(message)
                findings.append(
                    make_finding(
                        ValidationSeverity.WARNING,
                        "relationship_truth",
                        message,
                        sheet_name="EOAT Inventory",
                        row_number=row_number,
                        audit_id=audit_id,
                        machine_number=_cell_text(row_data.get("Press/Machine #")),
                        current_value=f"{SOURCE_AUDIT_ID_FIELD}={source_id}; {COMPATIBILITY_SOURCE_FIELD}={compatibility_source}",
                        expected_behavior="Physical audit rows should not carry compatibility-source metadata.",
                        recommended_action="Review whether this row should be marked Compatible or whether the metadata should be cleared manually.",
                        source_validator="inventory_relationships",
                    )
                )
            identity_key = _physical_identity_key(row_data)
            if identity_key:
                if identity_key in physical_identity_rows:
                    metrics["duplicate_physical_row_count"] += 1
                    message = f"Duplicate physical audit row identity: row {row_number} matches row {physical_identity_rows[identity_key]}"
                    duplicate_physical_examples.append(message)
                    findings.append(
                        make_finding(
                            ValidationSeverity.WARNING,
                            "relationship_truth",
                            message,
                            sheet_name="EOAT Inventory",
                            row_number=row_number,
                            audit_id=audit_id,
                            machine_number=_cell_text(row_data.get("Press/Machine #")),
                            current_value=" / ".join(identity_key),
                            expected_behavior="A physical audit identity should normally be unique for a machine, tool, and EOAT type.",
                            recommended_action="Compare the rows before deciding whether to merge, archive, or keep both as separate audits.",
                            source_validator="inventory_relationships",
                        )
                    )
                else:
                    physical_identity_rows[identity_key] = row_number
        elif is_compatibility_row(row_data):
            source_id = _cell_text(row_data.get(SOURCE_AUDIT_ID_FIELD))
            compatibility_source = _cell_text(row_data.get(COMPATIBILITY_SOURCE_FIELD))
            if not source_id:
                metrics["compatibility_missing_source_audit_id_count"] += 1
                message = f"Compatibility row is missing Source Audit ID: row {row_number}"
                compatibility_warning_examples.append(message)
                findings.append(
                    make_finding(
                        ValidationSeverity.ERROR,
                        "relationship_truth",
                        message,
                        sheet_name="EOAT Inventory",
                        row_number=row_number,
                        column_name=SOURCE_AUDIT_ID_FIELD,
                        audit_id=audit_id,
                        machine_number=_cell_text(row_data.get("Press/Machine #")),
                        current_value=row_data.get(SOURCE_AUDIT_ID_FIELD),
                        expected_behavior="Compatibility rows should link back to the physical source audit they were derived from.",
                        recommended_action="Repair compatibility metadata only after confirming the source audit.",
                        source_validator="inventory_relationships",
                    )
                )
            if not compatibility_source:
                metrics["compatibility_missing_source_count"] += 1
                message = f"Compatibility row is missing Compatibility Source: row {row_number}"
                compatibility_warning_examples.append(message)
                findings.append(
                    make_finding(
                        ValidationSeverity.ERROR,
                        "relationship_truth",
                        message,
                        sheet_name="EOAT Inventory",
                        row_number=row_number,
                        column_name=COMPATIBILITY_SOURCE_FIELD,
                        audit_id=audit_id,
                        machine_number=_cell_text(row_data.get("Press/Machine #")),
                        current_value=row_data.get(COMPATIBILITY_SOURCE_FIELD),
                        expected_behavior="Compatibility rows should record the local data source used to derive them.",
                        recommended_action="Repair compatibility metadata only after confirming the source list or source audit.",
                        source_validator="inventory_relationships",
                    )
                )
            if not source_id and (_cell_text(row_data.get("Audit Date")) or _cell_text(row_data.get("Auditor"))):
                metrics["compatibility_truth_warning_count"] += 1
                message = (
                    f"Row is marked Compatible but has physical-audit fields and no source metadata: row {row_number}"
                )
                compatibility_warning_examples.append(message)
                findings.append(
                    make_finding(
                        ValidationSeverity.WARNING,
                        "relationship_truth",
                        message,
                        sheet_name="EOAT Inventory",
                        row_number=row_number,
                        audit_id=audit_id,
                        machine_number=_cell_text(row_data.get("Press/Machine #")),
                        expected_behavior="Compatibility rows are derived rows, not physical verification rows.",
                        recommended_action="Review whether the row is a physical audit incorrectly marked compatible.",
                        source_validator="inventory_relationships",
                    )
                )

        has_indexed_photos = (
            bool(audit_id and audit_id.casefold() in indexed_photo_audit_ids)
            or bool(_cell_text(row_data.get("Press/Machine #")).casefold() in indexed_photo_machines)
        )
        for photo_finding in _photo_link_findings(
            project_root, row_number, row_data, has_indexed_photos=has_indexed_photos
        ):
            photo_warning_examples.append(photo_finding.message)
            if (
                photo_finding.column_name == "Photo Folder/Link"
                and "missing local photo evidence link" in photo_finding.message
            ):
                metrics["photos_yes_without_link_count"] += 1
            elif photo_finding.column_name == "Photos Taken?" and "marked No" in photo_finding.message:
                metrics["photo_link_while_no_count"] += 1
            elif "indexed photos exist" in photo_finding.message:
                metrics["indexed_photos_while_status_not_yes_count"] += 1
            elif "broken local photo link" in photo_finding.message:
                metrics["broken_photo_link_count"] += 1
            findings.append(photo_finding)

        requirements = entry_type_requirements(row_data)
        uninstalled_audit = is_uninstalled_eoat_audit(row_data)
        for required_field in requirements["required"]:
            if required_field in BLANK_CELL_VALIDATION_IGNORED_FIELDS:
                continue
            if uninstalled_audit and required_field in UNINSTALLED_MACHINE_AND_ROBOT_CONTEXT_FIELDS:
                continue
            if (
                required_field in header_positions
                and field_applies(row_data, required_field)
                and _is_missing_audit_value(row_data.get(required_field))
            ):
                add_major_na(row_number, required_field, row_data)

        for header in get_expected_headers("EOAT Inventory"):
            if header not in header_positions:
                continue
            if header in BLANK_CELL_VALIDATION_IGNORED_FIELDS:
                continue
            value = row_data.get(header)
            if _cell_text(value) == "":
                blank_cells += 1
            applies = audit_field_applies(row_data, header)
            if (
                missing_electrical_wiring_control
                and header in ELECTRICAL_DETAIL_FIELDS
                and not is_meaningful_value(value)
            ):
                applies = False
            if (
                header in MAJOR_AUDIT_COLUMNS
                and not (uninstalled_audit and header in UNINSTALLED_MACHINE_AND_ROBOT_CONTEXT_FIELDS)
                and (header != CUP_COUNT_FIELD or header in requirements["important"])
                and _is_missing_audit_value(value)
                and applies
            ):
                add_major_na(row_number, header, row_data)
            if not applies and is_na_value(value):
                non_applicable_na_examples.append(f"row {row_number} {header}")
            if not applies and is_meaningful_value(value):
                stale_hidden_value_examples.append(f"row {row_number} {header}")
                findings.append(
                    make_finding(
                        ValidationSeverity.AUTO_FIXABLE,
                        "audit_data",
                        f"Non-applicable EOAT Inventory cell contains a stale value: row {row_number} {header}",
                        sheet_name="EOAT Inventory",
                        row_number=row_number,
                        column_name=header,
                        audit_id=audit_id,
                        machine_number=_cell_text(row_data.get("Press/Machine #")),
                        current_value=value,
                        expected_behavior=f"Non-applicable fields should be saved as {NA_VALUE}.",
                        recommended_action="Preview and apply the Clear Stale Hidden Values safe fix.",
                        fix_available=True,
                        fix_id="clear_stale_hidden_na",
                        source_validator="inventory_rows",
                    )
                )
            if (
                header in AUDIT_DROPDOWN_ALLOWED_VALUES
                and _cell_text(value)
                and (header == EOAT_MOVES_FIELD or not is_na_value(value))
                and _cell_text(value) not in AUDIT_DROPDOWN_ALLOWED_VALUES[header]
            ):
                canonical = _canonical_dropdown_value(header, value)
                invalid_dropdown_examples.append(f"row {row_number} {header}={_cell_text(value)}")
                if canonical:
                    dropdown_casing_examples.append(f"row {row_number} {header}={_cell_text(value)} -> {canonical}")
                findings.append(
                    make_finding(
                        ValidationSeverity.AUTO_FIXABLE if canonical else ValidationSeverity.WARNING,
                        "audit_data",
                        f"Invalid EOAT Inventory dropdown value: row {row_number} {header}={_cell_text(value)}",
                        sheet_name="EOAT Inventory",
                        row_number=row_number,
                        column_name=header,
                        audit_id=audit_id,
                        machine_number=_cell_text(row_data.get("Press/Machine #")),
                        current_value=value,
                        expected_behavior=f"Value should be one of: {', '.join(sorted(AUDIT_DROPDOWN_ALLOWED_VALUES[header]))}.",
                        recommended_action=(
                            f"Preview and apply the Normalize Dropdown Casing safe fix to change this value to {canonical}."
                            if canonical
                            else "Open the audit field and choose a valid dropdown value; rebuild dropdown validation if Excel validation is missing."
                        ),
                        fix_available=bool(canonical),
                        fix_id=FIX_NORMALIZE_DROPDOWN_CASING if canonical else "",
                        source_validator="inventory_rows",
                    )
                )
            if (
                header in AUDIT_NUMERIC_FIELDS
                and applies
                and _cell_text(value)
                and not is_na_value(value)
                and not _is_non_negative_whole_number(value)
            ):
                invalid_numeric_examples.append(f"row {row_number} {header}={_cell_text(value)}")
                findings.append(
                    make_finding(
                        ValidationSeverity.ERROR,
                        "audit_data",
                        f"Invalid EOAT Inventory whole-number value: row {row_number} {header}={_cell_text(value)}",
                        sheet_name="EOAT Inventory",
                        row_number=row_number,
                        column_name=header,
                        audit_id=audit_id,
                        machine_number=_cell_text(row_data.get("Press/Machine #")),
                        current_value=value,
                        expected_behavior="Applicable count fields must be non-negative whole numbers.",
                        recommended_action="Open the audit field and correct the count; do not guess the engineering value.",
                        source_validator="inventory_rows",
                    )
                )
        if EOAT_MOVES_FIELD in header_positions and _is_missing_eoat_moves(row_data.get(EOAT_MOVES_FIELD)):
            entry_type = _cell_text(row_data.get("Entry Type")).lower()
            source_id = _cell_text(row_data.get(SOURCE_AUDIT_ID_FIELD))
            source_missing = source_id and _is_missing_eoat_moves(source_eoat_moves_by_audit_id.get(source_id))
            if entry_type == "compatible" and source_missing:
                message = f"row {row_number} Missing important audit field: {EOAT_MOVES_FIELD} inherited from source audit {source_id}"
                missing_eoat_moves_examples.append(message)
            else:
                message = f"row {row_number} Missing important audit field: {EOAT_MOVES_FIELD}"
                missing_eoat_moves_examples.append(message)
            findings.append(
                make_finding(
                    ValidationSeverity.WARNING,
                    "audit_data",
                    message,
                    sheet_name="EOAT Inventory",
                    row_number=row_number,
                    column_name=EOAT_MOVES_FIELD,
                    audit_id=audit_id,
                    machine_number=_cell_text(row_data.get("Press/Machine #")),
                    current_value=row_data.get(EOAT_MOVES_FIELD),
                    expected_behavior="EOAT Moves should be known for source audits and inherited compatibility rows.",
                    recommended_action="Open the audit row and verify whether the EOAT moves parts, sprues, or both.",
                    source_validator="inventory_rows",
                )
            )
        for warning in hybrid_completeness_warnings(row_data):
            hybrid_warning_examples.append(f"row {row_number}: {warning}")
            findings.append(
                make_finding(
                    ValidationSeverity.WARNING,
                    "audit_data",
                    f"row {row_number}: {warning}",
                    sheet_name="EOAT Inventory",
                    row_number=row_number,
                    audit_id=audit_id,
                    machine_number=_cell_text(row_data.get("Press/Machine #")),
                    expected_behavior="Hybrid audits should capture both vacuum-side and gripper/mechanical-side details.",
                    recommended_action="Open the audit row and complete the missing hybrid section.",
                    source_validator="inventory_rows",
                )
            )
        for warning in semantic_consistency_warnings(row_data):
            semantic_warning_examples.append(f"row {row_number}: {warning}")
            findings.append(
                make_finding(
                    ValidationSeverity.WARNING,
                    "audit_data",
                    f"row {row_number}: {warning}",
                    sheet_name="EOAT Inventory",
                    row_number=row_number,
                    audit_id=audit_id,
                    machine_number=_cell_text(row_data.get("Press/Machine #")),
                    expected_behavior="Audit field values should be internally consistent with field applicability rules.",
                    recommended_action="Open the audit row and resolve the conflicting field values.",
                    source_validator="inventory_rows",
                )
            )

    row_records = [dict(row) for row in inventory_rows_for_truth]
    for row in row_records:
        if not is_compatibility_row(row):
            continue
        row_number = int(row.get("_row_index") or 0) or None
        lookup = source_audit_for_compatibility_row(row_records, row)
        for warning, code in zip(lookup.warnings, lookup.warning_codes):
            if code == "missing_source_metadata":
                continue
            metrics["compatibility_truth_warning_count"] += 1
            compatibility_warning_examples.append(warning)
            findings.append(
                make_finding(
                    ValidationSeverity.WARNING,
                    "relationship_truth",
                    warning,
                    sheet_name="EOAT Inventory",
                    row_number=row_number,
                    column_name=SOURCE_AUDIT_ID_FIELD,
                    audit_id=_cell_text(row.get("Audit ID")),
                    machine_number=_cell_text(row.get("Press/Machine #")),
                    current_value=row.get(SOURCE_AUDIT_ID_FIELD),
                    expected_behavior="Compatibility rows should resolve to an existing physical source audit.",
                    recommended_action="Repair compatibility metadata only after confirming the source audit.",
                    source_validator="inventory_relationships",
                )
            )

    robot_mismatch_findings = _robot_circuit_mismatch_findings(project_root, row_records)
    for finding in robot_mismatch_findings:
        robot_mismatch_examples.append(finding.message)
    metrics["robot_side_circuit_mismatch_count"] = len(robot_mismatch_findings)
    findings.extend(robot_mismatch_findings)

    truth_summary = analyze_truth_from_rows(row_records, fields=get_expected_headers("EOAT Inventory"))
    for key, value in truth_summary.metrics.items():
        metrics[f"truth_{key}"] = value
    for state, count in truth_summary.state_counts.items():
        metrics[f"truth_state_{state}_count"] = count

    metrics["duplicate_audit_id_count"] = len(duplicate_ids)
    metrics["blank_saved_audit_cell_count"] = blank_cells
    major_na_list = sorted(major_na_examples)
    metrics["major_na_cell_count"] = len(major_na_list)
    metrics["missing_applicable_major_cell_count"] = len(major_na_list)
    metrics["non_applicable_na_cell_count"] = len(non_applicable_na_examples)
    metrics["stale_hidden_value_count"] = len(stale_hidden_value_examples)
    metrics["hybrid_warning_count"] = len(hybrid_warning_examples)
    metrics["semantic_warning_count"] = len(semantic_warning_examples)
    metrics["relationship_truth_warning_count"] = len(compatibility_warning_examples) + len(duplicate_physical_examples)
    metrics["photo_truth_warning_count"] = len(photo_warning_examples)
    metrics["invalid_dropdown_value_count"] = len(invalid_dropdown_examples)
    metrics["dropdown_casing_fixable_count"] = len(dropdown_casing_examples)
    metrics["invalid_numeric_value_count"] = len(invalid_numeric_examples)
    metrics["missing_eoat_moves_count"] = len(missing_eoat_moves_examples)
    metrics["missing_eoat_assembly_id_count"] = len(missing_eoat_id_examples)
    metrics["invalid_eoat_assembly_id_count"] = len(invalid_eoat_id_examples)

    if duplicate_ids:
        warnings.append(f"Duplicate Audit ID value(s): {', '.join(sorted(duplicate_ids))}")
    if duplicate_physical_examples:
        warnings.append(
            f"{len(duplicate_physical_examples)} duplicate physical audit row warning(s); see structured findings for row details."
        )
    if compatibility_warning_examples:
        warnings.append(
            f"{len(compatibility_warning_examples)} compatibility relationship warning(s); see structured findings for metadata details."
        )
    if photo_warning_examples:
        warnings.append(
            f"{len(photo_warning_examples)} photo link/evidence warning(s): {', '.join(photo_warning_examples[:5])}"
        )
    if robot_mismatch_examples:
        warnings.append(
            f"{len(robot_mismatch_examples)} robot-side circuit mismatch warning(s): {', '.join(robot_mismatch_examples[:5])}"
        )
    if major_na_list:
        warnings.append(
            f"{len(major_na_list)} applicable major EOAT Inventory cell(s) are blank or contain {NA_VALUE}: {', '.join(major_na_list[:10])}"
        )
    if stale_hidden_value_examples:
        warnings.append(
            f"{len(stale_hidden_value_examples)} non-applicable EOAT Inventory cell(s) contain stale values: {', '.join(stale_hidden_value_examples[:10])}"
        )
    if hybrid_warning_examples:
        warnings.append(
            f"{len(hybrid_warning_examples)} Hybrid EOAT completeness warning(s): {', '.join(hybrid_warning_examples[:5])}"
        )
    if semantic_warning_examples:
        warnings.append(
            f"{len(semantic_warning_examples)} semantic EOAT warning(s): {', '.join(semantic_warning_examples[:5])}"
        )
    if invalid_dropdown_examples:
        warnings.append(
            f"{len(invalid_dropdown_examples)} invalid EOAT Inventory dropdown value(s): {', '.join(invalid_dropdown_examples[:5])}"
        )
    if missing_eoat_id_examples:
        warnings.append(
            f"{len(missing_eoat_id_examples)} EOAT Inventory row(s) are missing EOAT Assembly ID. "
            "Run Assign Missing EOAT IDs when ready."
        )
    if invalid_eoat_id_examples:
        warnings.append(
            f"{len(invalid_eoat_id_examples)} invalid EOAT Assembly ID value(s): {', '.join(invalid_eoat_id_examples[:5])}"
        )
    if dropdown_casing_examples:
        warnings.append(
            f"{len(dropdown_casing_examples)} dropdown value(s) can be safely normalized by casing: {', '.join(dropdown_casing_examples[:5])}"
        )
    if invalid_numeric_examples:
        warnings.append(
            f"{len(invalid_numeric_examples)} invalid EOAT Inventory whole-number value(s): {', '.join(invalid_numeric_examples[:5])}"
        )
    if missing_eoat_moves_examples:
        warnings.append(
            f"{len(missing_eoat_moves_examples)} EOAT Inventory row(s) are missing EOAT Moves: {', '.join(missing_eoat_moves_examples[:5])}"
        )
    if blank_cells:
        warnings.append(
            f"{blank_cells} saved EOAT Inventory cell(s) are blank; new saves should write {NA_VALUE} "
            "for unanswered user-entered audit fields. Autofilled/system-managed compatibility metadata fields "
            "are intentionally ignored."
        )
    return warnings, metrics, findings


def _inventory_identity_sets(ws, headers: list[str]) -> tuple[set[str], set[str]]:
    header_positions = {header: index for index, header in enumerate(headers)}
    audit_ids: set[str] = set()
    machines: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {header: row[index] for header, index in header_positions.items() if index < len(row)}
        if not _is_audit_data_row(row_data):
            continue
        audit_id = _cell_text(row_data.get("Audit ID"))
        machine = _cell_text(row_data.get("Press/Machine #"))
        if audit_id and not is_na_value(audit_id):
            audit_ids.add(audit_id.casefold())
        if machine and not is_na_value(machine):
            machines.add(machine.casefold())
    return audit_ids, machines


def _workbook_schema_version(workbook) -> str:
    metadata_version = _workbook_metadata_values(workbook).get("schema_version", "")
    if metadata_version:
        return metadata_version
    return str(getattr(workbook.properties, "version", "") or "").strip() or SCHEMA_VERSION_UNKNOWN


def _workbook_metadata_values(workbook) -> dict[str, str]:
    if WORKBOOK_METADATA_SHEET not in workbook.sheetnames:
        return {}
    ws = workbook[WORKBOOK_METADATA_SHEET]
    values: dict[str, str] = {}
    for key, value in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        key_text = _cell_text(key)
        if key_text:
            values[key_text] = _cell_text(value)
    return values


def _validate_orphan_local_references(
    project_root: str | Path,
    audit_ids: set[str],
    machines: set[str],
    *,
    include_open_item_cache_references: bool = True,
) -> tuple[list[str], dict[str, int], list[ValidationFinding]]:
    warnings: list[str] = []
    findings: list[ValidationFinding] = []
    metrics = {
        "orphan_annotation_target_count": 0,
        "orphan_open_item_count": 0,
    }
    annotation_findings = _orphan_annotation_findings(project_root, audit_ids, machines)
    open_item_findings = (
        _orphan_open_item_findings(project_root, audit_ids, machines) if include_open_item_cache_references else []
    )
    metrics["orphan_annotation_target_count"] = len(annotation_findings)
    metrics["orphan_open_item_count"] = len(open_item_findings)
    findings.extend(annotation_findings)
    findings.extend(open_item_findings)
    if annotation_findings:
        warnings.append(
            f"{len(annotation_findings)} annotation note/tag target(s) reference missing audits or machines."
        )
    if open_item_findings:
        warnings.append(f"{len(open_item_findings)} cached open item(s) reference missing audits or machines.")
    return warnings, metrics, findings


def _orphan_annotation_findings(
    project_root: str | Path, audit_ids: set[str], machines: set[str]
) -> list[ValidationFinding]:
    path = resolve_project_paths(project_root).annotations_database
    if not path.exists():
        return []
    findings: list[ValidationFinding] = []
    try:
        conn = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
        conn.row_factory = sqlite3.Row
    except sqlite3.Error:
        return []
    try:
        target_rows = conn.execute(
            """
            SELECT id, target_type, target_label, audit_id, machine_id, field_key, field_label
            FROM annotation_targets
            """
        ).fetchall()
        for row in target_rows:
            target_type = _cell_text(row["target_type"]).casefold()
            audit_id = _cell_text(row["audit_id"])
            machine = _cell_text(row["machine_id"])
            missing_audit = (
                audit_id and audit_id.casefold() not in audit_ids and target_type in {"audit", "audit_field"}
            )
            missing_machine = (
                machine and machine.casefold() not in machines and target_type in {"machine", "audit", "audit_field"}
            )
            if not missing_audit and not missing_machine:
                continue
            reason = []
            if missing_audit:
                reason.append(f"audit {audit_id}")
            if missing_machine:
                reason.append(f"machine {machine}")
            findings.append(
                make_finding(
                    ValidationSeverity.WARNING,
                    "orphan_reference",
                    f"Annotation target references missing {' and '.join(reason)}.",
                    audit_id=audit_id,
                    machine_number=machine,
                    column_name=_cell_text(row["field_key"]) or _cell_text(row["field_label"]),
                    current_value=_cell_text(row["target_label"]) or _cell_text(row["id"]),
                    expected_behavior="Notes and tag assignments should point to audits or machines that still exist locally.",
                    recommended_action="Review the annotation target; do not delete notes or tags automatically.",
                    source_validator="orphan_local_references",
                )
            )
    except sqlite3.Error:
        return findings
    finally:
        conn.close()
    return findings


def _orphan_open_item_findings(
    project_root: str | Path, audit_ids: set[str], machines: set[str]
) -> list[ValidationFinding]:
    open_items_dir = resolve_project_paths(project_root).project_admin / "open_items"
    if not open_items_dir.exists():
        return []
    findings: list[ValidationFinding] = []
    seen: set[tuple[str, str, str]] = set()
    for path in sorted(open_items_dir.glob("open_item_snapshot*.json")):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if not isinstance(data, list):
            continue
        for record in data:
            if not isinstance(record, dict):
                continue
            audit_id = _cell_text(record.get("audit_id"))
            machine = _cell_text(record.get("machine"))
            item_id = _cell_text(record.get("item_id")) or _cell_text(record.get("id"))
            missing_audit = audit_id and audit_id.casefold() not in audit_ids
            missing_machine = machine and machine.casefold() not in machines
            if not missing_audit and not missing_machine:
                continue
            key = (item_id, audit_id, machine)
            if key in seen:
                continue
            seen.add(key)
            reason = []
            if missing_audit:
                reason.append(f"audit {audit_id}")
            if missing_machine:
                reason.append(f"machine {machine}")
            findings.append(
                make_finding(
                    ValidationSeverity.WARNING,
                    "orphan_reference",
                    f"Cached open item references missing {' and '.join(reason)}.",
                    audit_id=audit_id,
                    machine_number=machine,
                    current_value=_cell_text(record.get("title")) or item_id,
                    expected_behavior="Open item snapshots should point to audits or machines that still exist locally.",
                    recommended_action="Regenerate or review open items; never delete open item history automatically.",
                    source_validator="orphan_local_references",
                )
            )
    return findings


def _indexed_photo_reference_sets(project_root: str | Path) -> tuple[set[str], set[str]]:
    try:
        workbook_path = resolve_project_paths(project_root).master_workbook
        photo_rows = workbook_row_dicts(workbook_path, "Photo Index")
    except Exception:
        return set(), set()
    audit_ids: set[str] = set()
    machines_without_audit: set[str] = set()
    for row in photo_rows:
        audit_id = _cell_text(row.get("Related Audit ID")).casefold()
        machine = _cell_text(row.get("Press/Machine #")).casefold()
        if audit_id:
            audit_ids.add(audit_id)
        elif machine:
            machines_without_audit.add(machine)
    return audit_ids, machines_without_audit


def _photo_link_findings(
    project_root: str | Path,
    row_number: int,
    row_data: dict[str, object],
    *,
    has_indexed_photos: bool = False,
) -> list[ValidationFinding]:
    findings: list[ValidationFinding] = []
    audit_id = _cell_text(row_data.get("Audit ID"))
    machine = _cell_text(row_data.get("Press/Machine #"))
    photos_taken = _cell_text(row_data.get("Photos Taken?")).casefold()
    link = _cell_text(row_data.get("Photo Folder/Link"))
    if photos_taken == "yes" and not link and not has_indexed_photos:
        findings.append(
            make_finding(
                ValidationSeverity.WARNING,
                "photo_evidence",
                f"Photos Taken? is Yes but missing local photo evidence link: row {row_number}",
                sheet_name="EOAT Inventory",
                row_number=row_number,
                column_name="Photo Folder/Link",
                audit_id=audit_id,
                machine_number=machine,
                current_value=link,
                expected_behavior="Rows marked as having photos should include a local photo folder/link or indexed evidence.",
                recommended_action="Add the local photo folder/link or intake/index the evidence through the Photos page.",
                source_validator="inventory_photo_truth",
            )
        )
    if photos_taken in {"", "no"} and has_indexed_photos:
        findings.append(
            make_finding(
                ValidationSeverity.WARNING,
                "photo_evidence",
                f"Indexed photos exist but Photos Taken? is not marked Yes: row {row_number}",
                sheet_name="EOAT Inventory",
                row_number=row_number,
                column_name="Photos Taken?",
                audit_id=audit_id,
                machine_number=machine,
                current_value=row_data.get("Photos Taken?"),
                expected_behavior="Rows with indexed photo evidence should mark Photos Taken? as Yes.",
                recommended_action="Use the Photos page intake workflow or update the audit row after confirming the indexed evidence belongs to this audit.",
                source_validator="inventory_photo_truth",
            )
        )
    if photos_taken == "no" and link:
        findings.append(
            make_finding(
                ValidationSeverity.WARNING,
                "photo_evidence",
                f"Photo Folder/Link is populated while Photos Taken? is marked No: row {row_number}",
                sheet_name="EOAT Inventory",
                row_number=row_number,
                column_name="Photos Taken?",
                audit_id=audit_id,
                machine_number=machine,
                current_value=f"Photos Taken?=No; Photo Folder/Link={link}",
                expected_behavior="Photo status and evidence links should agree.",
                recommended_action="Confirm whether evidence exists before changing the workbook value.",
                source_validator="inventory_photo_truth",
            )
        )
    if link and _looks_like_local_path(link):
        missing_links = [
            candidate for candidate in _candidate_photo_links(project_root, link) if not candidate.exists()
        ]
        if missing_links and len(missing_links) == len(_split_photo_links(link)):
            findings.append(
                make_finding(
                    ValidationSeverity.WARNING,
                    "photo_evidence",
                    f"Photo Folder/Link contains a broken local photo link: row {row_number}",
                    sheet_name="EOAT Inventory",
                    row_number=row_number,
                    column_name="Photo Folder/Link",
                    audit_id=audit_id,
                    machine_number=machine,
                    current_value=link,
                    expected_behavior="Local photo links should resolve to an existing local file or folder.",
                    recommended_action="Open the Photos page and repair the folder/link; do not delete photo references automatically.",
                    source_validator="inventory_photo_truth",
                )
            )
    return findings


def _robot_circuit_mismatch_findings(
    project_root: str | Path, rows: list[dict[str, object]]
) -> list[ValidationFinding]:
    path = robot_info_workbook_path(project_root)
    if not path.exists():
        return []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    robot_rows: dict[str, dict[str, object]] = {}
    try:
        if ROBOT_INFO_SHEET not in workbook.sheetnames:
            return []
        ws = workbook[ROBOT_INFO_SHEET]
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        positions = {header: index for index, header in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {header: row[index] for header, index in positions.items() if index < len(row)}
            machine = _cell_text(row_data.get("Machine Number")).casefold()
            if machine:
                robot_rows[machine] = row_data
    finally:
        workbook.close()

    findings: list[ValidationFinding] = []
    fields = ("Robot Vacuum Circuits", "Robot Pressure Circuits", "Robot Interchangeable Circuits")
    for row in rows:
        machine = _cell_text(row.get("Press/Machine #"))
        if not machine:
            continue
        robot_row = robot_rows.get(machine.casefold())
        if not robot_row:
            continue
        for field in fields:
            inventory_value = _cell_text(row.get(field))
            robot_value = _cell_text(robot_row.get(field))
            if not inventory_value or not robot_value or is_na_value(inventory_value) or is_na_value(robot_value):
                continue
            if inventory_value == robot_value:
                continue
            findings.append(
                make_finding(
                    ValidationSeverity.WARNING,
                    "robot_truth",
                    f"Robot-side circuit mismatch for {machine}: EOAT Inventory {field}={inventory_value}, Robot_Info {field}={robot_value}.",
                    sheet_name="EOAT Inventory",
                    row_number=int(row.get("_row_index") or 0) or None,
                    column_name=field,
                    audit_id=_cell_text(row.get("Audit ID")),
                    machine_number=machine,
                    current_value=inventory_value,
                    expected_behavior="Robot-side circuit fields should agree with Robot_Info.xlsx when both sources are populated.",
                    recommended_action="Review both sources before modifying Robot_Info rows.",
                    source_validator="robot_truth",
                )
            )
    return findings


def _physical_identity_key(row_data: dict[str, object]) -> tuple[str, str, str] | None:
    machine = _cell_text(row_data.get("Press/Machine #")).casefold()
    tool = _cell_text(row_data.get("Tool #")).casefold()
    eoat_type = _cell_text(row_data.get("EOAT Type")).casefold()
    if not machine or not tool:
        return None
    return (machine, tool, eoat_type)


def _canonical_dropdown_value(header: str, value: object) -> str:
    text = _cell_text(value)
    if not text:
        return ""
    for allowed in AUDIT_DROPDOWN_ALLOWED_VALUES.get(header, set()):
        if str(allowed).casefold() == text.casefold() and str(allowed) != text:
            return str(allowed)
    return ""


def _looks_like_local_path(value: str) -> bool:
    parsed = urlparse(value)
    if parsed.scheme in {"http", "https", "mailto"}:
        return False
    return any(sep in value for sep in ("\\", "/", ":")) or value.startswith(".")


def _split_photo_links(value: str) -> list[str]:
    normalized = value.replace("\n", ";").replace("|", ";")
    return [part.strip() for part in normalized.split(";") if part.strip()]


def _candidate_photo_links(project_root: str | Path, value: str) -> list[Path]:
    candidates: list[Path] = []
    for raw in _split_photo_links(value):
        parsed = urlparse(raw)
        if parsed.scheme in {"http", "https", "mailto"}:
            continue
        text = unquote(parsed.path) if parsed.scheme == "file" else raw
        path = Path(text)
        candidates.append(path if path.is_absolute() else Path(project_root) / path)
    return candidates


def _is_audit_data_row(row_data: dict[str, object]) -> bool:
    values = [_cell_text(value) for value in row_data.values()]
    if not any(values):
        return False
    if len([value for value in values if value]) == 1 and values[-1].startswith("Last Updated:"):
        return False
    return True


def _cell_text(value: object) -> str:
    return str(value or "").strip()


def _is_missing_eoat_moves(value: object) -> bool:
    text = _cell_text(value)
    return not text or text.upper() == NA_VALUE


def _is_missing_audit_value(value: object) -> bool:
    text = _cell_text(value)
    return not text or is_na_value(text)


def _is_non_negative_whole_number(value: object) -> bool:
    text = _cell_text(value)
    if not text:
        return False
    try:
        parsed = int(text)
    except ValueError:
        return False
    return parsed >= 0 and str(parsed) == text


def write_validation_report(project_root: str | Path, result: ToolResult) -> Path:
    paths = resolve_project_paths(project_root)
    ensure_directory(paths.validation_reports)
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    report_path = paths.validation_reports / f"Foundation_Validation_{stamp}.md"
    return safe_write_text(report_path, result.to_markdown(), overwrite=False)


def run_foundation_validation(
    project_root: str | Path,
    write_report: bool = True,
    log_activity: bool = True,
) -> ToolResult:
    result = validate_project_foundation(project_root)
    if write_report and Path(project_root).exists():
        try:
            report_path = write_validation_report(project_root, result)
            result.output_reports.append(str(report_path))
            result.files_created.append(str(report_path))
            json_report_path = write_validation_json_report(project_root, result)
            result.output_reports.append(str(json_report_path))
            result.files_created.append(str(json_report_path))
        except FileExistsError:
            # Minute-level names can collide during rapid tests; retry with seconds.
            paths = resolve_project_paths(project_root)
            stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
            report_path = safe_write_text(
                paths.validation_reports / f"Foundation_Validation_{stamp}.md",
                result.to_markdown(),
                overwrite=False,
            )
            result.output_reports.append(str(report_path))
            result.files_created.append(str(report_path))
            json_report_path = write_validation_json_report(project_root, result)
            result.output_reports.append(str(json_report_path))
            result.files_created.append(str(json_report_path))
        except Exception as exc:
            result.warnings.append(f"Could not write validation report: {exc}")
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result
