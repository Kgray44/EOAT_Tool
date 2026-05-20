from __future__ import annotations

import re
from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .constants import DEFAULT_MASTER_PRESS_LIST_FILE, DEFAULT_PRESS_CAPACITY_FILE
from .paths import get_master_press_list_file, get_press_capacity_file, resolve_project_paths

MASTER_FILE_NAME = DEFAULT_MASTER_PRESS_LIST_FILE
CAPACITY_FILE_NAME = DEFAULT_PRESS_CAPACITY_FILE
MASTER_SHEET_NAME = "Machine Specifications"
CAPACITY_SHEET_NAME = "Capacity"

MASTER_SOURCE_FIELDS = [
    "U.S. Tons",
    "Press Brand",
    "Model #",
    "Year Mfg.",
    "Injection Pressure",
    "Injection Capacity",
    "Screw Diameter",
    "Controller Type",
    "Robot/Picker Brand",
    "Robot/Picker Model #",
    "Robot/Picker Serial #",
    "Robot/Picker Mfg. Date",
    "Full Servo",
    "# of TCU's",
    "EDART UNIT PRESS SIDE",
]

PRESS_MACHINE_INFO_FIELDS = [
    "Press Brand",
    "Press Model",
    "Press Tonnage",
    "Press Year",
    "Controller Type",
    "Screw Diameter",
    "Injection Pressure",
    "Injection Capacity",
]

ROBOT_INFO_FIELDS = [
    "Robot Brand",
    "Robot Model",
    "Robot Serial Number",
    "Robot Manufacturing Date",
    "Full Servo",
    "TCU Count",
    "EDART Unit Press Side",
]

MASTER_AUDIT_FIELD_MAP = {
    "U.S. Tons": "Press Tonnage",
    "Model #": "Press Model",
    "Year Mfg.": "Press Year",
    "Robot/Picker Brand": "Robot Brand",
    "Robot/Picker Model #": "Robot Model",
    "Robot/Picker Serial #": "Robot Serial Number",
    "Robot/Picker Mfg. Date": "Robot Manufacturing Date",
    "# of TCU's": "TCU Count",
    "EDART UNIT PRESS SIDE": "EDART Unit Press Side",
}

MASTER_FIELDS = list(dict.fromkeys([*PRESS_MACHINE_INFO_FIELDS, *ROBOT_INFO_FIELDS, *MASTER_SOURCE_FIELDS]))

CAPACITY_SUMMARY_FIELDS = ["Press Capacity Label", "Capacity Tonnage", "Screw Size"]

CAPACITY_PART_SOURCE_FIELDS = [
    "NGW Part Number",
    "NGW Part Description",
    "Bill-to / Customer",
    "Cycle Time (S)",
    "Cavitation",
    "Forecasted Capacity",
    "Available Capacity",
    "Hours Allocated per month",
    "Hours per week",
    "Committed Hours per Year",
]

CAPACITY_AUDIT_FIELD_MAP = {
    "NGW Part Number": "Selected Part Number",
    "NGW Part Description": "Selected Part Description",
    "Bill-to / Customer": "Customer",
    "Hours Allocated per month": "Hours Allocated per Month",
    "Hours per week": "Hours per Week",
}

CAPACITY_AUDIT_FIELDS = [
    "Selected Part Number",
    "Selected Part Description",
    "Customer",
    "Cycle Time (S)",
    "Cavitation",
    "Forecasted Capacity",
    "Available Capacity",
    "Hours Allocated per Month",
    "Hours per Week",
    "Committed Hours per Year",
]

CAPACITY_PART_FIELDS = list(dict.fromkeys([*CAPACITY_AUDIT_FIELDS, *CAPACITY_PART_SOURCE_FIELDS]))

LOOKUP_METADATA_FIELDS = [
    "Lookup Machine Number Normalized",
    "Lookup Machine Number",
    "Master Press List Matched",
    "Master Press List Rows Matched",
    "Capacity Sheet Matched",
    "Plant 4 Capacity Matched",
    "Capacity Part Rows Matched",
    "Plant 4 Capacity Rows Matched",
    "Selected Capacity Row Index or ID",
    "Lookup Warnings",
    "Lookup Errors",
    "Lookup Source Files",
]


@dataclass(frozen=True)
class CapacityPartMatch:
    row_number: int
    fields: dict[str, Any]

    def display_label(self) -> str:
        number = _display(self.fields.get("NGW Part Number") or self.fields.get("Selected Part Number"))
        description = _display(self.fields.get("NGW Part Description") or self.fields.get("Selected Part Description"))
        customer = _display(self.fields.get("Bill-to / Customer") or self.fields.get("Customer"))
        pieces = [piece for piece in [number, description, customer] if piece]
        return " - ".join(pieces) if pieces else f"Capacity row {self.row_number}"


@dataclass
class PressLookupResult:
    machine_number: int
    raw_machine_input: str = ""
    master_fields: dict[str, Any] = field(default_factory=dict)
    capacity_summary: dict[str, Any] = field(default_factory=dict)
    capacity_part_rows: list[CapacityPartMatch] = field(default_factory=list)
    master_rows_count: int = 0
    capacity_rows_count: int = 0
    selected_capacity_row_id: str = ""
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    details: list[str] = field(default_factory=list)
    master_source: Path | None = None
    capacity_source: Path | None = None

    @property
    def normalized_machine_number(self) -> int:
        return self.machine_number

    @property
    def master_match_count(self) -> int:
        return self.master_rows_count

    @property
    def capacity_match_count(self) -> int:
        return self.capacity_rows_count

    @property
    def press_fields(self) -> dict[str, Any]:
        return {field_name: self.master_fields.get(field_name, "") for field_name in PRESS_MACHINE_INFO_FIELDS}

    @property
    def robot_fields(self) -> dict[str, Any]:
        return {field_name: self.master_fields.get(field_name, "") for field_name in ROBOT_INFO_FIELDS}

    @property
    def part_rows(self) -> list[dict[str, Any]]:
        return [row.fields for row in self.capacity_part_rows]

    @property
    def selected_part_row(self) -> dict[str, Any] | None:
        if len(self.capacity_part_rows) == 1:
            return self.capacity_part_rows[0].fields
        return None

    @property
    def master_matched(self) -> bool:
        return self.master_rows_count > 0

    @property
    def capacity_matched(self) -> bool:
        return self.capacity_rows_count > 0 or bool(self.capacity_summary)

    def metadata_fields(self, part_index: int | None = None) -> dict[str, str]:
        sources = [str(path) for path in [self.master_source, self.capacity_source] if path is not None]
        selected = ""
        if part_index is not None and 0 <= part_index < len(self.capacity_part_rows):
            selected = str(self.capacity_part_rows[part_index].row_number)
        return {
            "Lookup Machine Number Normalized": str(self.machine_number),
            "Lookup Machine Number": str(self.machine_number),
            "Master Press List Matched": "Yes" if self.master_matched else "No",
            "Master Press List Rows Matched": str(self.master_rows_count),
            "Capacity Sheet Matched": "Yes" if self.capacity_matched else "No",
            "Plant 4 Capacity Matched": "Yes" if self.capacity_matched else "No",
            "Capacity Part Rows Matched": str(self.capacity_rows_count),
            "Plant 4 Capacity Rows Matched": str(self.capacity_rows_count),
            "Selected Capacity Row Index or ID": selected,
            "Lookup Warnings": "; ".join(self.warnings),
            "Lookup Errors": "; ".join(self.errors),
            "Lookup Source Files": "; ".join(sources),
        }

    def fields_for_audit(self, part_index: int | None = None) -> dict[str, Any]:
        fields: dict[str, Any] = {}
        fields.update(self.master_fields)
        fields.update(self.capacity_summary)
        if part_index is not None and 0 <= part_index < len(self.capacity_part_rows):
            fields.update(self.capacity_part_rows[part_index].fields)
        fields.update(self.metadata_fields(part_index))
        return fields

    @property
    def robot_type_suggestion(self) -> str:
        brand = _display(self.master_fields.get("Robot/Picker Brand") or self.master_fields.get("Robot Brand"))
        model = _display(self.master_fields.get("Robot/Picker Model #") or self.master_fields.get("Robot Model"))
        return " ".join(part for part in [brand, model] if part)

    @property
    def robot_model_controller_suggestion(self) -> str:
        return _display(self.master_fields.get("Robot/Picker Model #") or self.master_fields.get("Robot Model"))

    @property
    def part_options(self) -> list[dict[str, Any]]:
        options: list[dict[str, Any]] = []
        for index, row in enumerate(self.capacity_part_rows):
            fields = row.fields
            options.append(
                {
                    "index": index,
                    "row_number": row.row_number,
                    "part_number": fields.get("NGW Part Number", ""),
                    "part_description": fields.get("NGW Part Description", ""),
                    "customer": fields.get("Bill-to / Customer", ""),
                    "cycle_time": fields.get("Cycle Time (S)", ""),
                    "part_family": part_family_from_capacity_fields(fields),
                    "display_label": row.display_label(),
                }
            )
        return options

    @property
    def part_family_suggestion(self) -> str:
        if len(self.part_options) == 1:
            return str(self.part_options[0].get("part_family") or "")
        return ""


def part_family_from_capacity_fields(fields: dict[str, Any]) -> str:
    number = _display(fields.get("NGW Part Number") or fields.get("Selected Part Number"))
    description = _display(fields.get("NGW Part Description") or fields.get("Selected Part Description"))
    if number and description:
        return f"{number} - {description}"
    return number or description


MachineLookupResult = PressLookupResult


def lookup_machine(
    project_root: str | Path,
    raw_machine_value: str | int | float | None,
    reference_dir: str | Path | None = None,
) -> PressLookupResult:
    return lookup_press(project_root, raw_machine_value, reference_dir)


def normalize_machine_number(value: str | int | float | None) -> int:
    if value is None:
        raise ValueError("Machine number is required.")
    if isinstance(value, int):
        if value > 0:
            return value
        raise ValueError("Machine number must be greater than zero.")
    if isinstance(value, float) and value.is_integer():
        if value > 0:
            return int(value)
        raise ValueError("Machine number must be greater than zero.")

    text = str(value).strip()
    if not text:
        raise ValueError("Machine number is required.")

    patterns = [
        r"^(\d+)$",
        r"^(?:press|machine)\s*[-#]?\s*(\d+)$",
        r"^[pm]\s*[-#]?\s*(\d+)$",
    ]
    for pattern in patterns:
        match = re.match(pattern, text, flags=re.IGNORECASE)
        if match:
            number = int(match.group(1))
            if number > 0:
                return number
            raise ValueError("Machine number must be greater than zero.")
    raise ValueError(f"Invalid machine number: {text}. Use a value like 12, Press 12, P12, M12, or Machine 12.")


def reference_data_dir(project_root: str | Path, override: str | Path | None = None) -> Path:
    if override:
        return Path(override)
    paths = resolve_project_paths(project_root)
    preferred = paths.reference_data
    preferred_has_references = (preferred / MASTER_FILE_NAME).exists() or (preferred / CAPACITY_FILE_NAME).exists()
    if preferred.exists() and preferred_has_references:
        return preferred
    if paths.legacy_reference_data.exists():
        return paths.legacy_reference_data
    legacy_named = paths.project_root / "Reference_Data"
    if legacy_named.exists():
        return legacy_named
    return preferred


def lookup_press(
    project_root: str | Path,
    machine_input: str | int | float | None,
    reference_dir: str | Path | None = None,
) -> PressLookupResult:
    machine_number = normalize_machine_number(machine_input)
    if reference_dir:
        data_dir = Path(reference_dir)
        master_source = data_dir / MASTER_FILE_NAME
        capacity_source = data_dir / CAPACITY_FILE_NAME
    else:
        data_dir = reference_data_dir(project_root)
        master_source = get_master_press_list_file(project_root)
        capacity_source = get_press_capacity_file(project_root)
        if not master_source.exists() and data_dir != master_source.parent:
            master_source = data_dir / MASTER_FILE_NAME
        if not capacity_source.exists() and data_dir != capacity_source.parent:
            capacity_source = data_dir / CAPACITY_FILE_NAME

    result = PressLookupResult(
        machine_number=machine_number,
        raw_machine_input="" if machine_input is None else str(machine_input),
        master_source=master_source,
        capacity_source=capacity_source,
    )
    _load_master_press_list(result)
    _load_capacity_sheet(result)
    if result.master_matched and result.capacity_summary:
        master_tons = _first_number(result.master_fields.get("Press Tonnage") or result.master_fields.get("U.S. Tons"))
        capacity_tons = _first_number(result.capacity_summary.get("Capacity Tonnage"))
        if master_tons is not None and capacity_tons is not None and master_tons != capacity_tons:
            result.warnings.append(
                "Conflicting tonnage found; using Master Press List for press identity and Plant 4 Capacity for capacity data."
            )
    if not result.master_matched and result.master_source and result.master_source.exists():
        result.warnings.append(f"Machine {machine_number} was not found in the Master Press List.")
    if not result.capacity_matched and result.capacity_source and result.capacity_source.exists():
        result.warnings.append(f"Machine {machine_number} was not found in the Plant 4 Capacity sheet.")
    return result


def clear_lookup_cache() -> None:
    _load_master_rows_from_file.cache_clear()
    _load_capacity_rows_from_file.cache_clear()


def _load_master_press_list(result: PressLookupResult) -> None:
    path = result.master_source
    if path is None or not path.exists():
        result.warnings.append(f"Master Press List reference file not found: {path}")
        return

    warnings, rows = _load_master_rows(path)
    result.warnings.extend(warnings)
    if not rows:
        return

    matches = []
    for row_number, row in rows:
        try:
            number = normalize_machine_number(row.get("Machine Number"))
        except ValueError:
            continue
        if number == result.machine_number:
            matches.append((row_number, row))

    result.master_rows_count = len(matches)
    if not matches:
        return

    if len(matches) > 1:
        result.warnings.append(f"Multiple master press rows found for Machine {result.machine_number}. Review lookup details.")

    merged: dict[str, Any] = {}
    for field_name in MASTER_SOURCE_FIELDS:
        values = [_clean_value(row.get(field_name)) for _row_number, row in matches]
        nonblank = [value for value in values if value not in ("", None)]
        if not nonblank:
            continue
        first = nonblank[0]
        merged[field_name] = first
        audit_field = MASTER_AUDIT_FIELD_MAP.get(field_name)
        if audit_field:
            merged[audit_field] = first
        if any(value != first for value in nonblank[1:]):
            result.warnings.append(f"Conflicting master values for {field_name}; using the first nonblank value.")
    result.master_fields = merged
    result.details.append(f"Master Press List matched {len(matches)} row(s).")


def _load_capacity_sheet(result: PressLookupResult) -> None:
    path = result.capacity_source
    if path is None or not path.exists():
        result.warnings.append(f"Plant 4 Capacity reference file not found: {path}")
        return

    warnings, rows = _load_capacity_rows(path)
    result.warnings.extend(warnings)
    if rows:
        _parse_capacity_rows(rows, result)


def _load_master_rows(path: Path) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    stat = path.stat()
    return _load_master_rows_from_file(str(path), stat.st_mtime_ns, stat.st_size)


@lru_cache(maxsize=8)
def _load_master_rows_from_file(path_text: str, _mtime_ns: int, _size: int) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    path = Path(path_text)
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_name = _find_sheet_name(workbook.sheetnames, MASTER_SHEET_NAME, ["machine", "spec"])
        if sheet_name is None:
            return [f"Master Press List sheet not found: {MASTER_SHEET_NAME}"], []
        ws = workbook[sheet_name]
        return [], _rows_from_header_sheet(ws, "Machine Number")
    except Exception as exc:
        return [f"Master Press List could not be read: {exc}"], []
    finally:
        if workbook is not None:
            workbook.close()


def _load_capacity_rows(path: Path) -> tuple[list[str], list[tuple[int, list[Any]]]]:
    stat = path.stat()
    return _load_capacity_rows_from_file(str(path), stat.st_mtime_ns, stat.st_size)


@lru_cache(maxsize=8)
def _load_capacity_rows_from_file(path_text: str, _mtime_ns: int, _size: int) -> tuple[list[str], list[tuple[int, list[Any]]]]:
    path = Path(path_text)
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_name = _find_sheet_name(workbook.sheetnames, CAPACITY_SHEET_NAME, ["capacity"])
        if sheet_name is None:
            return [f"Plant 4 Capacity sheet not found: {CAPACITY_SHEET_NAME}"], []
        ws = workbook[sheet_name]
        return [], list(_iter_sheet_rows(ws))
    except Exception as exc:
        return [f"Plant 4 Capacity workbook could not be read: {exc}"], []
    finally:
        if workbook is not None:
            workbook.close()


def _parse_capacity_rows(rows: list[tuple[int, list[Any]]], result: PressLookupResult) -> None:
    current_summary: dict[str, Any] = {}
    header_map: dict[str, int] | None = None

    for row_number, values in rows:
        row_text = " ".join(_display(value) for value in values if _display(value))
        summary = _parse_press_summary(row_text, result.machine_number)
        if summary:
            current_summary = _merge_capacity_summary(current_summary, summary)
            result.capacity_summary = _merge_capacity_summary(result.capacity_summary, summary)

        maybe_header = _capacity_header_map(values)
        if maybe_header:
            header_map = maybe_header
            continue

        if header_map is None:
            continue

        machine_cell = _value_for(values, header_map, "Machine No.")
        machine_numbers = _machine_numbers_from_capacity_cell(machine_cell)
        if result.machine_number not in machine_numbers:
            referenced = _parse_press_summary(row_text, result.machine_number)
            if not referenced:
                continue

        fields = _capacity_fields_from_row(values, header_map)
        if current_summary:
            for field_name in CAPACITY_SUMMARY_FIELDS:
                fields.setdefault(field_name, current_summary.get(field_name, ""))
        if any(_display(fields.get(field_name)) for field_name in CAPACITY_PART_SOURCE_FIELDS):
            result.capacity_part_rows.append(CapacityPartMatch(row_number=row_number, fields=fields))
        elif summary:
            result.capacity_summary = _merge_capacity_summary(result.capacity_summary, summary)

    result.capacity_rows_count = len(result.capacity_part_rows)
    if result.capacity_part_rows and not result.capacity_summary:
        first = result.capacity_part_rows[0].fields
        result.capacity_summary = {field_name: first.get(field_name, "") for field_name in CAPACITY_SUMMARY_FIELDS if first.get(field_name, "")}
    if result.capacity_part_rows:
        result.details.append(f"Plant 4 Capacity matched {len(result.capacity_part_rows)} part row(s).")


def _find_sheet_name(sheet_names: list[str], preferred: str, words: list[str]) -> str | None:
    if preferred in sheet_names:
        return preferred
    normalized_preferred = _norm(preferred)
    for sheet_name in sheet_names:
        if _norm(sheet_name) == normalized_preferred:
            return sheet_name
    for sheet_name in sheet_names:
        normalized = _norm(sheet_name)
        if all(word in normalized for word in words):
            return sheet_name
    return None


def _rows_from_header_sheet(ws, required_header: str) -> list[tuple[int, dict[str, Any]]]:
    header_row: list[Any] | None = None
    header_index = 0
    required = _norm(required_header)
    for index, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=True), start=1):
        normalized = [_norm(value) for value in row]
        if required in normalized:
            header_row = list(row)
            header_index = index
            break
    if header_row is None:
        return []

    headers = [_display(value) for value in header_row]
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=header_index + 1, values_only=True), start=header_index + 1):
        if not any(value not in (None, "") for value in row):
            continue
        rows.append((row_number, {headers[index]: value for index, value in enumerate(row) if index < len(headers)}))
    return rows


def _iter_sheet_rows(ws):
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(value not in (None, "") for value in row):
            yield row_number, list(row)


def _capacity_header_map(values: list[Any]) -> dict[str, int] | None:
    aliases = {
        "Machine No.": ["Machine No.", "Machine No", "Machine #", "Machine Number", "Press", "Press #"],
        "NGW Part Number": ["NGW Part Number", "Part Number", "Part #", "NGW Part #"],
        "NGW Part Description": ["NGW Part Description", "Part Description", "Description"],
        "Bill-to / Customer": ["Bill-to / Customer", "Bill To / Customer", "Bill-to Customer", "Customer", "Bill To"],
        "Cycle Time (S)": ["Cycle Time (S)", "Cycle Time", "Cycle Time S", "Cycle Time Seconds"],
        "Cavitation": ["Cavitation", "Cavities", "Cavity"],
        "Forecasted Capacity": ["Forecasted Capacity", "Forecast Capacity"],
        "Available Capacity": ["Available Capacity"],
        "Hours Allocated per month": ["Hours Allocated per month", "Hours Allocated/month", "Hours Allocated"],
        "Hours per week": ["Hours per week", "Hours/week"],
        "Committed Hours per Year": ["Committed Hours per Year", "Committed Hours/year", "Committed Hours"],
    }
    normalized_values = [_norm(value) for value in values]
    machine_headers = {_norm(alias) for alias in aliases["Machine No."]}
    if not any(value in machine_headers for value in normalized_values):
        return None

    mapping: dict[str, int] = {}
    for logical, names in aliases.items():
        normalized_names = {_norm(name) for name in names}
        for index, normalized in enumerate(normalized_values):
            if normalized in normalized_names:
                mapping[logical] = index
                break
    return mapping if "Machine No." in mapping else None


def _capacity_fields_from_row(values: list[Any], header_map: dict[str, int]) -> dict[str, Any]:
    fields = {}
    for field_name in CAPACITY_PART_SOURCE_FIELDS:
        value = _clean_value(_value_for(values, header_map, field_name))
        fields[field_name] = value
        audit_field = CAPACITY_AUDIT_FIELD_MAP.get(field_name)
        if audit_field:
            fields[audit_field] = value
    return fields


def _value_for(values: list[Any], header_map: dict[str, int], field_name: str) -> Any:
    index = header_map.get(field_name)
    if index is None or index >= len(values):
        return ""
    return values[index]


def _machine_numbers_from_capacity_cell(value: Any) -> set[int]:
    text = _display(value)
    if not text:
        return set()
    return {int(match) for match in re.findall(r"\d+", text)}


def _parse_press_summary(text: str, machine_number: int) -> dict[str, str]:
    if not text:
        return {}
    match = re.search(rf"\bPress\s*#?\s*{machine_number}\b(?:\s*-\s*.*)?", text, flags=re.IGNORECASE)
    if not match:
        return {}
    label = match.group(0).strip()
    tonnage = ""
    screw = ""
    ton_match = re.search(r"(\d+(?:\.\d+)?)\s*T\b", label, flags=re.IGNORECASE)
    if ton_match:
        tonnage = f"{ton_match.group(1)}T"
    screw_match = re.search(r"(\d+(?:\.\d+)?)\s*mm\s*Screw\b", label, flags=re.IGNORECASE)
    if screw_match:
        screw = f"{screw_match.group(1)}mm Screw"
    return {
        "Press Capacity Label": label,
        "Capacity Tonnage": tonnage,
        "Screw Size": screw,
    }


def _merge_capacity_summary(existing: dict[str, Any], incoming: dict[str, Any]) -> dict[str, Any]:
    merged = dict(existing)
    incoming_has_details = bool(incoming.get("Capacity Tonnage") or incoming.get("Screw Size"))
    for field_name, value in incoming.items():
        if not value:
            continue
        if field_name == "Press Capacity Label" and merged.get(field_name) and not incoming_has_details:
            continue
        if not merged.get(field_name) or incoming_has_details:
            merged[field_name] = value
    return merged


def _first_number(value: Any) -> float | None:
    match = re.search(r"\d+(?:\.\d+)?", _display(value))
    return float(match.group(0)) if match else None


def _clean_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _display(value: Any) -> str:
    value = _clean_value(value)
    return "" if value == "" else str(value).strip()


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _display(value).lower())
