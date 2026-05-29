from __future__ import annotations

import time
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .paths import resolve_project_paths
from .result import ToolResult
from .safe_files import backup_file, ensure_directory
from .workbook_cache import invalidate_workbook_cache
from .workbook_io import next_empty_row, worksheet_headers

ROBOT_INFO_SHEET = "Robot Info"
ROBOT_NOTES_FIELD = "Robot Notes"
ROBOT_INFO_HEADERS = [
    "Plant/Area",
    "Machine Number",
    "Robot Type",
    "Robot Identifier",
    "Robot Vacuum Circuits",
    "Robot Pressure Circuits",
    "Robot Interchangeable Circuits",
    ROBOT_NOTES_FIELD,
    "Last Audit ID",
    "Last Updated",
    "Notes",
]
ROBOT_CIRCUIT_FIELDS = [
    "Robot Vacuum Circuits",
    "Robot Pressure Circuits",
    "Robot Interchangeable Circuits",
]
ROBOT_INFO_AUDIT_FIELDS = [
    *ROBOT_CIRCUIT_FIELDS,
    ROBOT_NOTES_FIELD,
]


def robot_info_workbook_path(project_root: str | Path) -> Path:
    return resolve_project_paths(project_root).robot_info_workbook


def ensure_robot_info_workbook(project_root: str | Path) -> Path:
    path = robot_info_workbook_path(project_root)
    ensure_directory(path.parent)
    if not path.exists():
        workbook = Workbook()
        ws = workbook.active
        ws.title = ROBOT_INFO_SHEET
        _write_robot_info_headers(ws)
        _style_robot_info_sheet(ws)
        workbook.save(path)
        workbook.close()
        invalidate_workbook_cache(path)
        return path

    workbook = load_workbook(path)
    try:
        ws = workbook[ROBOT_INFO_SHEET] if ROBOT_INFO_SHEET in workbook.sheetnames else workbook.create_sheet(ROBOT_INFO_SHEET)
        headers = worksheet_headers(ws)
        if not headers or not any(headers):
            _write_robot_info_headers(ws)
            headers = worksheet_headers(ws)
        for header in ROBOT_INFO_HEADERS:
            if header not in headers:
                ws.cell(row=1, column=ws.max_column + 1).value = header
                headers = worksheet_headers(ws)
        _style_robot_info_sheet(ws)
        workbook.save(path)
        invalidate_workbook_cache(path)
    finally:
        workbook.close()
    return path


def load_robot_info_for_audit_entry(project_root: str | Path, entry: dict[str, Any]) -> dict[str, object] | None:
    return load_robot_info(
        project_root,
        plant_area=str(entry.get("Plant/Area") or ""),
        machine_number=str(entry.get("Press/Machine #") or ""),
        robot_type=str(entry.get("Robot Type") or ""),
        robot_identifier=str(entry.get("Robot Identifier") or ""),
    )


def load_robot_info(
    project_root: str | Path,
    *,
    plant_area: str,
    machine_number: str,
    robot_type: str,
    robot_identifier: str = "",
) -> dict[str, object] | None:
    path = robot_info_workbook_path(project_root)
    if not path.exists():
        return None
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if ROBOT_INFO_SHEET not in workbook.sheetnames:
            return None
        ws = workbook[ROBOT_INFO_SHEET]
        headers = worksheet_headers(ws)
        key = _robot_key(plant_area, machine_number, robot_type, robot_identifier)
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            if _row_key(row_data) == key:
                return {header: row_data.get(header, "") for header in ROBOT_INFO_HEADERS}
        return None
    finally:
        workbook.close()


def upsert_robot_info_from_audit(project_root: str | Path, entry: dict[str, Any]) -> ToolResult:
    started = time.perf_counter()
    audit_id = str(entry.get("Audit ID") or "").strip()
    plant_area = str(entry.get("Plant/Area") or "").strip()
    machine_number = str(entry.get("Press/Machine #") or "").strip()
    robot_type = str(entry.get("Robot Type") or "").strip()
    robot_identifier = str(entry.get("Robot Identifier") or "").strip()
    if not machine_number:
        return ToolResult.ok(
            "robot_info_save",
            "Robot Info",
            "Robot Info was not updated because Machine Number is blank.",
            warnings=["Robot-side circuit data was skipped because Machine Number is blank."],
            duration_seconds=time.perf_counter() - started,
        )
    errors = _validate_robot_circuit_values(entry)
    if errors:
        return ToolResult.fail(
            "robot_info_save",
            "Robot Info",
            "Robot Info update failed.",
            errors=errors,
            duration_seconds=time.perf_counter() - started,
        )

    path = ensure_robot_info_workbook(project_root)
    backup = None
    workbook = None
    try:
        backup = backup_file(path, path.parent / "_backups")
        workbook = load_workbook(path)
        ws = workbook[ROBOT_INFO_SHEET]
        _ensure_headers(ws)
        headers = worksheet_headers(ws)
        existing_row = _find_robot_row(ws, headers, plant_area, machine_number, robot_type, robot_identifier)
        row_number = existing_row or next_empty_row(ws)
        existing_data = _read_robot_row(ws, row_number, headers) if existing_row else {}
        data = {
            "Plant/Area": plant_area,
            "Machine Number": machine_number,
            "Robot Type": robot_type,
            "Robot Identifier": robot_identifier,
            "Robot Vacuum Circuits": _robot_circuit_data_value(entry, existing_data, "Robot Vacuum Circuits", allow_blank=True),
            "Robot Pressure Circuits": _robot_circuit_data_value(entry, existing_data, "Robot Pressure Circuits", allow_blank=True),
            "Robot Interchangeable Circuits": _robot_circuit_data_value(entry, existing_data, "Robot Interchangeable Circuits", allow_blank=False),
            ROBOT_NOTES_FIELD: _robot_text_data_value(entry, existing_data, ROBOT_NOTES_FIELD),
            "Last Audit ID": audit_id,
            "Last Updated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "Notes": _robot_text_data_value(entry, existing_data, "Notes"),
        }
        _write_robot_row(ws, row_number, headers, data)
        _style_robot_info_sheet(ws)
        workbook.save(path)
        invalidate_workbook_cache(path)
    except Exception as exc:
        return ToolResult.fail(
            "robot_info_save",
            "Robot Info",
            "Robot Info update failed.",
            errors=[f"Could not update Robot_Info.xlsx: {exc}"],
            duration_seconds=time.perf_counter() - started,
        )
    finally:
        if workbook is not None:
            workbook.close()

    action = "Updated" if existing_row else "Created"
    details = [
        f"{action} Robot_Info.xlsx row for Machine {machine_number}.",
        f"Robot Vacuum Circuits: {data['Robot Vacuum Circuits']}",
        f"Robot Pressure Circuits: {data['Robot Pressure Circuits']}",
        f"Robot Interchangeable Circuits: {data['Robot Interchangeable Circuits']}",
    ]
    if data[ROBOT_NOTES_FIELD]:
        details.append(f"{ROBOT_NOTES_FIELD}: {data[ROBOT_NOTES_FIELD]}")
    return ToolResult.ok(
        "robot_info_save",
        "Robot Info",
        f"Updated Robot_Info.xlsx for Machine {machine_number}.",
        details=details,
        files_created=[str(backup)] if backup else [],
        files_modified=[str(path)],
        metrics={
            "machine_number": machine_number,
            "row": row_number,
            "robot_vacuum_circuits": data["Robot Vacuum Circuits"],
            "robot_pressure_circuits": data["Robot Pressure Circuits"],
            "robot_interchangeable_circuits": data["Robot Interchangeable Circuits"],
            "robot_notes": data[ROBOT_NOTES_FIELD],
        },
        duration_seconds=time.perf_counter() - started,
    )


def validate_robot_info_workbook(project_root: str | Path) -> tuple[list[str], list[str], dict[str, int]]:
    path = robot_info_workbook_path(project_root)
    warnings: list[str] = []
    errors: list[str] = []
    metrics = {
        "robot_info_missing_workbook": 0,
        "robot_info_missing_header_count": 0,
        "robot_info_duplicate_row_count": 0,
        "robot_info_invalid_circuit_count": 0,
    }
    if not path.exists():
        warnings.append(f"Missing Robot Info workbook: {path}")
        metrics["robot_info_missing_workbook"] = 1
        return warnings, errors, metrics
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if ROBOT_INFO_SHEET not in workbook.sheetnames:
            warnings.append(f"Robot Info workbook is missing sheet: {ROBOT_INFO_SHEET}")
            return warnings, errors, metrics
        ws = workbook[ROBOT_INFO_SHEET]
        headers = worksheet_headers(ws)
        missing_headers = [header for header in ROBOT_INFO_HEADERS if header not in headers]
        if missing_headers:
            metrics["robot_info_missing_header_count"] = len(missing_headers)
            warnings.extend(f"Robot Info workbook missing header: {header}" for header in missing_headers)
        header_positions = {header: index for index, header in enumerate(headers)}
        seen_keys: dict[tuple[str, str, str, str], int] = {}
        duplicate_keys: set[tuple[str, str, str, str]] = set()
        invalid_examples: list[str] = []
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_data = {header: row[index] for header, index in header_positions.items() if index < len(row)}
            if not any(str(value or "").strip() for value in row_data.values()):
                continue
            key = _row_key(row_data)
            if key in seen_keys:
                duplicate_keys.add(key)
            else:
                seen_keys[key] = row_number
            for field in ROBOT_CIRCUIT_FIELDS:
                value = row_data.get(field)
                if str(value or "").strip() and _parse_non_negative_int(value) is None:
                    invalid_examples.append(f"row {row_number} {field}={value}")
        if duplicate_keys:
            metrics["robot_info_duplicate_row_count"] = len(duplicate_keys)
            warnings.append(f"Duplicate Robot Info row key(s): {len(duplicate_keys)}")
        if invalid_examples:
            metrics["robot_info_invalid_circuit_count"] = len(invalid_examples)
            warnings.append(f"Invalid Robot Info circuit value(s): {', '.join(invalid_examples[:5])}")
    except Exception as exc:
        errors.append(f"Could not validate Robot_Info.xlsx: {exc}")
    finally:
        if workbook is not None:
            workbook.close()
    return warnings, errors, metrics


def _ensure_headers(ws) -> None:
    headers = worksheet_headers(ws)
    if not headers or not any(headers):
        _write_robot_info_headers(ws)
        headers = worksheet_headers(ws)
    for header in ROBOT_INFO_HEADERS:
        if header not in headers:
            ws.cell(row=1, column=ws.max_column + 1).value = header
            headers = worksheet_headers(ws)


def _write_robot_info_headers(ws) -> None:
    for column, header in enumerate(ROBOT_INFO_HEADERS, start=1):
        ws.cell(row=1, column=column).value = header


def _find_robot_row(ws, headers: list[str], plant_area: str, machine_number: str, robot_type: str, robot_identifier: str) -> int | None:
    key = _robot_key(plant_area, machine_number, robot_type, robot_identifier)
    for row_number in range(2, ws.max_row + 1):
        row_data = {header: ws.cell(row=row_number, column=index).value for index, header in enumerate(headers, start=1)}
        if _row_key(row_data) == key:
            return row_number
    return None


def _read_robot_row(ws, row_number: int, headers: list[str]) -> dict[str, object]:
    return {header: ws.cell(row=row_number, column=index).value for index, header in enumerate(headers, start=1)}


def _robot_key(plant_area: str, machine_number: str, robot_type: str, robot_identifier: str = "") -> tuple[str, str, str, str]:
    robot_id = str(robot_identifier or "").strip().casefold()
    return (
        str(plant_area or "").strip().casefold(),
        str(machine_number or "").strip().casefold(),
        robot_id,
        "" if robot_id else str(robot_type or "").strip().casefold(),
    )


def _row_key(row: dict[str, object]) -> tuple[str, str, str, str]:
    return _robot_key(
        str(row.get("Plant/Area") or ""),
        str(row.get("Machine Number") or ""),
        str(row.get("Robot Type") or ""),
        str(row.get("Robot Identifier") or ""),
    )


def _validate_robot_circuit_values(entry: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    for field in ROBOT_CIRCUIT_FIELDS:
        text = str(entry.get(field) or "").strip()
        if not text and field != "Robot Interchangeable Circuits":
            continue
        if _parse_non_negative_int(text or "0") is None:
            errors.append(f"{field} must be a non-negative whole number.")
    return errors


def _normalized_robot_circuit(value: Any, *, allow_blank: bool) -> int | str:
    text = str(value or "").strip()
    if not text and allow_blank:
        return ""
    parsed = _parse_non_negative_int(text or "0")
    if parsed is None:
        raise ValueError(f"Invalid circuit count: {value}")
    return parsed


def _robot_circuit_data_value(entry: dict[str, Any], existing_data: dict[str, object], field: str, *, allow_blank: bool) -> int | str:
    if field not in entry and field in existing_data:
        return existing_data.get(field, "")
    return _normalized_robot_circuit(entry.get(field), allow_blank=allow_blank)


def _robot_text_data_value(entry: dict[str, Any], existing_data: dict[str, object], field: str) -> str:
    if field not in entry and field in existing_data:
        return str(existing_data.get(field) or "").strip()
    return str(entry.get(field) or "").strip()


def _parse_non_negative_int(value: Any) -> int | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = int(text)
    except ValueError:
        return None
    if str(parsed) != text and text != f"{parsed}.0":
        return None
    return parsed if parsed >= 0 else None


def _write_robot_row(ws, row_number: int, headers: list[str], data: dict[str, object]) -> None:
    template_row = max(2, row_number - 1)
    for col, header in enumerate(headers, start=1):
        if header not in data:
            continue
        cell = ws.cell(row=row_number, column=col)
        template_cell = ws.cell(row=template_row, column=col)
        if template_cell.has_style:
            cell._style = copy(template_cell._style)
        cell.value = data[header]
        if header in ROBOT_CIRCUIT_FIELDS:
            cell.number_format = "0"


def _style_robot_info_sheet(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max(ws.max_column, len(ROBOT_INFO_HEADERS)))}1"
    widths = {
        "Plant/Area": 16,
        "Machine Number": 18,
        "Robot Type": 22,
        "Robot Identifier": 20,
        "Robot Vacuum Circuits": 22,
        "Robot Pressure Circuits": 23,
        "Robot Interchangeable Circuits": 29,
        ROBOT_NOTES_FIELD: 42,
        "Last Audit ID": 20,
        "Last Updated": 24,
        "Notes": 36,
    }
    headers = worksheet_headers(ws)
    for col, header in enumerate(headers, start=1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = widths.get(header, 18)
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2, max_row=max(ws.max_row, 2), max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
