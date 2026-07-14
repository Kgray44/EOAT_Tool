from __future__ import annotations

import json
import os
import shutil
import sqlite3
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

from .config import AtlasGlobalConfig
from .events import EventOutbox
from .pending_updates import PendingUpdateStore
from .runtime_paths import AtlasRuntimePaths, ensure_runtime_layout, get_runtime_paths


ALLOWED_FIELDS: dict[str, set[str]] = {
    "eoat": {
        "status",
        "knownissues",
        "installnotes",
        "tubingnotes",
        "connectiontype",
        "vacuuminfo",
        "pressureinfo",
        "gripperinfo",
        "sensorinfo",
        "partdescription",
    },
    "tool": {"label", "partdescription", "partfamily"},
    "machine": {"currenteoat", "robottype", "robotmodel", "controller"},
}


class ChangeValidationService:
    def __init__(self, config: AtlasGlobalConfig):
        self.config = config

    def validate_submission(self, submission: dict[str, Any]) -> tuple[bool, str]:
        required = ("entity_type", "entity_id", "field")
        missing = [key for key in required if not str(submission.get(key, "")).strip()]
        if "proposed_value" not in submission:
            missing.append("proposed_value")
        if missing:
            return False, f"Missing required fields: {', '.join(missing)}"
        entity_type = str(submission.get("entity_type") or "").strip().casefold()
        field_name = _normalize_field(str(submission.get("field") or submission.get("field_name") or ""))
        if entity_type not in ALLOWED_FIELDS:
            return False, f"Unsupported entity type: {entity_type}"
        if field_name not in ALLOWED_FIELDS[entity_type]:
            return False, f"Field is not enabled for pending updates: {submission.get('field')}"
        return True, "Validated."


@dataclass(frozen=True)
class WorkbookLock:
    path: Path
    owner_id: str
    acquired_at: str


class WorkbookLockManager:
    def __init__(self, config: AtlasGlobalConfig, *, stale_after_seconds: int = 900):
        self.config = config
        self.stale_after_seconds = stale_after_seconds

    def acquire(self, lock_dir: str | Path, *, purpose: str, attempts: int = 3, delay_seconds: float = 0.2) -> WorkbookLock:
        directory = Path(lock_dir)
        directory.mkdir(parents=True, exist_ok=True)
        lock_path = directory / "atlas_global_workbook.lock"
        owner_id = f"{self.config.app_instance_id}:{uuid4().hex}:{purpose}"
        for _ in range(max(1, attempts)):
            if self._lock_is_stale(lock_path):
                lock_path.unlink(missing_ok=True)
            try:
                with lock_path.open("x", encoding="utf-8") as stream:
                    acquired_at = datetime.now().isoformat(timespec="seconds")
                    expires_at = datetime.fromtimestamp(time.time() + self.stale_after_seconds).isoformat(timespec="seconds")
                    stream.write(
                        "\n".join(
                            (
                                f"owner_id={owner_id}",
                                f"install_id={self.config.install_id}",
                                f"app_instance_id={self.config.app_instance_id}",
                                f"computer_name={self.config.computer_name}",
                                f"windows_user={self.config.windows_user}",
                                f"process_id={os.getpid()}",
                                f"app_version={self.config.app_version}",
                                f"created_at={acquired_at}",
                                f"expires_at={expires_at}",
                                f"purpose={purpose}",
                            )
                        )
                        + "\n"
                    )
                return WorkbookLock(lock_path, owner_id, acquired_at)
            except FileExistsError:
                time.sleep(delay_seconds)
        raise TimeoutError(f"Workbook lock is held: {lock_path}")

    def release(self, lock: WorkbookLock) -> None:
        try:
            text = lock.path.read_text(encoding="utf-8")
        except OSError:
            return
        if lock.owner_id in text:
            lock.path.unlink(missing_ok=True)

    def _lock_is_stale(self, lock_path: Path) -> bool:
        if not lock_path.exists():
            return False
        try:
            age = time.time() - lock_path.stat().st_mtime
        except OSError:
            return False
        return age > self.stale_after_seconds

    def lock_metadata(self, lock_dir: str | Path) -> dict[str, str]:
        lock_path = Path(lock_dir) / "atlas_global_workbook.lock"
        try:
            text = lock_path.read_text(encoding="utf-8")
        except OSError:
            return {}
        metadata: dict[str, str] = {}
        for line in text.splitlines():
            if "=" not in line:
                continue
            key, value = line.split("=", 1)
            metadata[key.strip()] = value.strip()
        metadata["stale"] = str(self._lock_is_stale(lock_path)).lower()
        metadata["path"] = str(lock_path)
        return metadata


class WorkbookBackupManager:
    def __init__(self, config: AtlasGlobalConfig):
        self.config = config

    def create_backup(self, workbook_path: str | Path, backup_dir: str | Path) -> Path:
        if not (self.config.writes_enabled() or self.config.shadow_writes_enabled() or self.config.sandbox_writes_enabled()):
            raise PermissionError("Workbook backup is disabled until write mode is enabled or shadow mode is selected.")
        source = Path(workbook_path)
        target_dir = Path(backup_dir)
        target_dir.mkdir(parents=True, exist_ok=True)
        target = target_dir / f"{source.stem}.{datetime.now().strftime('%Y%m%d_%H%M%S')}{source.suffix}"
        shutil.copy2(source, target)
        return target


class ConflictDetectionService:
    def source_changed_since(self, workbook_path: str | Path, *, mtime_ns: int, size: int) -> bool:
        path = Path(workbook_path)
        if not path.exists():
            return True
        stat = path.stat()
        return stat.st_mtime_ns != int(mtime_ns) or stat.st_size != int(size)

    def detect_field_conflict(
        self,
        *,
        entity_type: str,
        entity_id: str,
        field: str,
        base_value: Any,
        local_value: Any,
        workbook_value: Any,
    ) -> dict[str, Any] | None:
        if _value_key(base_value) == _value_key(workbook_value):
            return None
        if _value_key(local_value) == _value_key(workbook_value):
            return None
        return {
            "conflict_id": uuid4().hex,
            "entity_type": str(entity_type or ""),
            "entity_id": str(entity_id or ""),
            "field": str(field or ""),
            "base_value": base_value,
            "local_value": local_value,
            "workbook_value": workbook_value,
            "status": "open",
            "detected_at": datetime.now().isoformat(timespec="seconds"),
        }

    def record_conflict(self, conn: sqlite3.Connection, conflict: dict[str, Any], *, update_id: str = "") -> str:
        conflict_id = str(conflict.get("conflict_id") or uuid4().hex)
        conn.execute(
            """
            INSERT OR REPLACE INTO conflicts(
                conflict_id, update_id, entity_type, entity_id, field,
                base_value, local_value, workbook_value, status, detected_at, payload_json
            )
            VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                conflict_id,
                str(update_id or conflict.get("update_id") or ""),
                str(conflict.get("entity_type") or ""),
                str(conflict.get("entity_id") or ""),
                str(conflict.get("field") or ""),
                _value_key(conflict.get("base_value")),
                _value_key(conflict.get("local_value")),
                _value_key(conflict.get("workbook_value")),
                str(conflict.get("status") or "open"),
                str(conflict.get("detected_at") or datetime.now().isoformat(timespec="seconds")),
                json.dumps(conflict, default=str, sort_keys=True),
            ),
        )
        return conflict_id


class WorkbookUpdateService:
    def __init__(self, config: AtlasGlobalConfig):
        self.config = config

    def apply_shadow_update(
        self,
        workbook_path: str | Path,
        *,
        sheet_name: str,
        row_number: int,
        field_name: str,
        proposed_value: Any,
        shadow_dir: str | Path | None = None,
        backup_dir: str | Path | None = None,
    ) -> dict[str, Any]:
        if not self.config.shadow_writes_enabled():
            return {"status": "refused", "message": "Workbook write mode is disabled; no workbook was modified."}
        source = Path(workbook_path)
        target_dir = Path(shadow_dir) if shadow_dir else source.parent / "shadow_writes"
        target_dir.mkdir(parents=True, exist_ok=True)
        target = target_dir / source.name
        shutil.copy2(source, target)
        backup_path = ""
        if backup_dir is not None:
            backup_path = str(WorkbookBackupManager(self.config).create_backup(target, backup_dir))
        workbook = load_workbook(target)
        try:
            worksheet = workbook[sheet_name]
            header = [str(cell.value or "").strip() for cell in worksheet[1]]
            try:
                column = header.index(field_name) + 1
            except ValueError:
                return {"status": "refused", "message": f"Column not found: {field_name}", "path": str(target), "backup": backup_path}
            worksheet.cell(row=row_number, column=column).value = proposed_value
            workbook.save(target)
        finally:
            workbook.close()
        return {"status": "shadow_written", "path": str(target), "backup": backup_path}

    def apply_field_update(
        self,
        workbook_path: str | Path,
        *,
        sheet_name: str,
        id_column: str,
        entity_id: str,
        field_name: str,
        proposed_value: Any,
    ) -> dict[str, Any]:
        workbook = load_workbook(workbook_path)
        try:
            worksheet = workbook[sheet_name]
            header = [str(cell.value or "").strip() for cell in worksheet[1]]
            try:
                id_index = header.index(id_column) + 1
                field_index = header.index(field_name) + 1
            except ValueError as exc:
                return {"status": "refused", "message": f"Column not found: {exc}"}
            row_number = _find_entity_row(worksheet, id_index, entity_id)
            if row_number <= 0:
                return {"status": "refused", "message": f"Entity not found: {entity_id}"}
            before_value = worksheet.cell(row=row_number, column=field_index).value
            worksheet.cell(row=row_number, column=field_index).value = proposed_value
            workbook.save(workbook_path)
        finally:
            workbook.close()
        return {
            "status": "written",
            "sheet_name": sheet_name,
            "row_number": row_number,
            "field_name": field_name,
            "before_value": before_value,
            "written_value": proposed_value,
        }

    def read_field_value(
        self,
        workbook_path: str | Path,
        *,
        sheet_name: str,
        id_column: str,
        entity_id: str,
        field_name: str,
    ) -> dict[str, Any]:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name]
            header = [str(cell.value or "").strip() for cell in worksheet[1]]
            try:
                id_index = header.index(id_column) + 1
                field_index = header.index(field_name) + 1
            except ValueError as exc:
                return {"status": "refused", "message": f"Column not found: {exc}"}
            row_number = _find_entity_row(worksheet, id_index, entity_id)
            if row_number <= 0:
                return {"status": "refused", "message": f"Entity not found: {entity_id}"}
            return {
                "status": "read",
                "sheet_name": sheet_name,
                "row_number": row_number,
                "field_name": field_name,
                "value": worksheet.cell(row=row_number, column=field_index).value,
            }
        finally:
            workbook.close()


class SyncAttemptLogger:
    def record(
        self,
        conn: sqlite3.Connection,
        *,
        status: str,
        update_id: str = "",
        event_id: str = "",
        message: str = "",
        payload: dict[str, Any] | None = None,
        attempt_id: str | None = None,
    ) -> str:
        started_at = datetime.now().isoformat(timespec="seconds")
        row_id = attempt_id or uuid4().hex
        conn.execute(
            """
            INSERT OR REPLACE INTO sync_attempts(
                attempt_id, update_id, event_id, status, started_at, completed_at, message, payload_json
            )
            VALUES(?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row_id,
                str(update_id or ""),
                str(event_id or ""),
                str(status or ""),
                started_at,
                started_at,
                str(message or ""),
                json.dumps(dict(payload or {}), default=str, sort_keys=True),
            ),
        )
        return row_id


class BackgroundSyncService:
    def __init__(self, config: AtlasGlobalConfig, *, db_path: str | Path | None = None):
        self.config = config
        self.db_path = Path(db_path) if db_path is not None else None

    def run_once(self) -> dict[str, Any]:
        if self.db_path is not None:
            try:
                with sqlite3.connect(self.db_path) as conn:
                    SyncAttemptLogger().record(
                        conn,
                        status="disabled",
                        message="Background global sync is intentionally disabled in Phase 2.",
                    )
            except sqlite3.Error:
                pass
        return {"status": "disabled", "message": "Background global sync is intentionally disabled in Phase 2."}


class SyncStatusService:
    def __init__(self, config: AtlasGlobalConfig):
        self.config = config

    def status(self) -> dict[str, Any]:
        return {
            "write_mode": self.config.write_mode,
            "writes_enabled": self.config.writes_enabled(),
            "sandbox_writes_enabled": self.config.sandbox_writes_enabled(),
            "shadow_writes_enabled": self.config.shadow_writes_enabled(),
        }


class WorkbookSyncService:
    def __init__(self, runtime: AtlasRuntimePaths | None = None, config: AtlasGlobalConfig | None = None):
        self.runtime = ensure_runtime_layout(runtime or get_runtime_paths())
        self.config = config
        if self.config is None:
            from .config import load_or_create_global_config

            self.config = load_or_create_global_config(self.runtime)

    def sync_pending_update_to_sandbox(
        self,
        update_id: str,
        *,
        workbook_path: str | Path,
        sheet_name: str,
        id_column: str,
        lock_dir: str | Path | None = None,
        backup_dir: str | Path | None = None,
    ) -> dict[str, Any]:
        store = PendingUpdateStore(self.runtime, self.config)
        update = store.get_update(update_id)
        if not update:
            raise FileNotFoundError(f"Pending update not found: {update_id}")
        workbook = Path(workbook_path)
        attempt_id = uuid4().hex
        event_writer = EventOutbox(self.runtime, self.config)
        if not self.config.sandbox_writes_enabled():
            message = "Sandbox workbook sync is disabled by config."
            self._record_attempt(attempt_id, update_id, "refused", message, {"workbook_path": str(workbook)})
            store.set_sync_status(update_id, sync_status="refused", status="pending", last_error=message, sync_attempt_id=attempt_id)
            event = self._write_sync_event(
                event_writer,
                "workbook_sync_refused",
                update,
                workbook_path=workbook,
                sync_attempt_id=attempt_id,
                write_result={"status": "refused", "message": message},
                error_type="PermissionError",
                error_message=message,
            )
            return {"status": "refused", "message": message, "event_id": event["event_id"], "sync_attempt_id": attempt_id}
        if self._is_configured_production_workbook(workbook):
            message = "Refusing sandbox sync against the configured production workbook."
            self._record_attempt(attempt_id, update_id, "refused", message, {"workbook_path": str(workbook)})
            store.set_sync_status(update_id, sync_status="refused", status="pending", last_error=message, sync_attempt_id=attempt_id)
            event = self._write_sync_event(
                event_writer,
                "workbook_sync_refused",
                update,
                workbook_path=workbook,
                sync_attempt_id=attempt_id,
                write_result={"status": "refused", "message": message},
                error_type="PermissionError",
                error_message=message,
            )
            return {"status": "refused", "message": message, "event_id": event["event_id"], "sync_attempt_id": attempt_id}

        lock_manager = WorkbookLockManager(self.config)
        lock = lock_manager.acquire(lock_dir or self.runtime.local_lock_diagnostics_dir, purpose="sandbox_sync")
        fingerprint_before = workbook_fingerprint(workbook)
        backup_path = ""
        try:
            field_name = str(update.get("field_name") or update.get("field") or "")
            expected = update.get("expected_original_value", update.get("original_value"))
            proposed = update.get("proposed_value")
            current_read = WorkbookUpdateService(self.config).read_field_value(
                workbook,
                sheet_name=sheet_name,
                id_column=id_column,
                entity_id=str(update.get("entity_id") or ""),
                field_name=field_name,
            )
            if current_read.get("status") != "read":
                raise ValueError(str(current_read.get("message") or "Could not read current workbook value."))
            current_value = current_read.get("value")
            conflict = ConflictDetectionService().detect_field_conflict(
                entity_type=str(update.get("entity_type") or ""),
                entity_id=str(update.get("entity_id") or ""),
                field=field_name,
                base_value=expected,
                local_value=proposed,
                workbook_value=current_value,
            )
            if conflict is not None:
                self._record_conflict(update_id, conflict)
                message = "Workbook value no longer matches the expected original value."
                self._record_attempt(attempt_id, update_id, "conflict", message, conflict)
                store.set_sync_status(
                    update_id,
                    sync_status="conflict",
                    status="pending",
                    conflict_status="same_field_conflict",
                    last_error=message,
                    sync_attempt_id=attempt_id,
                )
                event = self._write_sync_event(
                    event_writer,
                    "workbook_sync_conflict",
                    update,
                    workbook_path=workbook,
                    sync_attempt_id=attempt_id,
                    workbook_fingerprint_before=fingerprint_before,
                    actual_workbook_values_before_write={field_name: current_value},
                    conflict_result={"status": "conflict", "details": conflict},
                    write_result={"status": "not_written"},
                    error_type="ConflictError",
                    error_message=message,
                    lock_id=lock.owner_id,
                )
                return {"status": "conflict", "event_id": event["event_id"], "sync_attempt_id": attempt_id, "conflict": conflict}
            backup_path = str(WorkbookBackupManager(self.config).create_backup(workbook, backup_dir or self.runtime.backups_dir))
            write_result = WorkbookUpdateService(self.config).apply_field_update(
                workbook,
                sheet_name=sheet_name,
                id_column=id_column,
                entity_id=str(update.get("entity_id") or ""),
                field_name=field_name,
                proposed_value=proposed,
            )
            if write_result.get("status") != "written":
                raise ValueError(str(write_result.get("message") or "Workbook write failed."))
            fingerprint_after = workbook_fingerprint(workbook)
            self._record_attempt(attempt_id, update_id, "succeeded", "Sandbox workbook sync succeeded.", write_result)
            store.clear_after_success(update_id, sync_attempt_id=attempt_id)
            event = self._write_sync_event(
                event_writer,
                "workbook_sync_succeeded",
                update,
                workbook_path=workbook,
                sync_attempt_id=attempt_id,
                workbook_fingerprint_before=fingerprint_before,
                workbook_fingerprint_after=fingerprint_after,
                actual_workbook_values_before_write={field_name: current_value},
                values_written={field_name: proposed},
                validation_result={"status": "valid"},
                conflict_result={"status": "none"},
                write_result=write_result,
                backup_path=backup_path,
                lock_id=lock.owner_id,
            )
            return {
                "status": "succeeded",
                "event_id": event["event_id"],
                "sync_attempt_id": attempt_id,
                "backup_path": backup_path,
                "write_result": write_result,
            }
        except Exception as exc:
            message = str(exc)
            self._record_attempt(attempt_id, update_id, "failed", message, {"workbook_path": str(workbook)})
            store.set_sync_status(update_id, sync_status="failed", status="pending", last_error=message, sync_attempt_id=attempt_id)
            event = self._write_sync_event(
                event_writer,
                "workbook_sync_failed",
                update,
                workbook_path=workbook,
                sync_attempt_id=attempt_id,
                workbook_fingerprint_before=fingerprint_before,
                write_result={"status": "failed"},
                error_type=type(exc).__name__,
                error_message=message,
                backup_path=backup_path,
                lock_id=lock.owner_id,
            )
            return {"status": "failed", "event_id": event["event_id"], "sync_attempt_id": attempt_id, "message": message}
        finally:
            lock_manager.release(lock)

    def _is_configured_production_workbook(self, workbook_path: Path) -> bool:
        configured = str(self.config.source_paths().get("eoat_master_tracker") or "").strip()
        if not configured:
            return False
        try:
            return workbook_path.resolve(strict=False) == Path(configured).resolve(strict=False)
        except OSError:
            return False

    def _record_attempt(self, attempt_id: str, update_id: str, status: str, message: str, payload: dict[str, Any]) -> None:
        attempt_payload = {
            "install_id": self.config.install_id,
            "app_instance_id": self.config.app_instance_id,
            "machine_name": self.config.machine_name,
            "computer_name": self.config.computer_name,
            "windows_user": self.config.windows_user,
            "app_version": self.config.app_version,
            "release_id": self.config.release_id,
            **dict(payload or {}),
        }
        with sqlite3.connect(self.runtime.db_path) as conn:
            from .sqlite_store import initialize_schema

            conn.row_factory = sqlite3.Row
            initialize_schema(conn)
            SyncAttemptLogger().record(
                conn,
                attempt_id=attempt_id,
                status=status,
                update_id=update_id,
                message=message,
                payload=attempt_payload,
            )

    def _record_conflict(self, update_id: str, conflict: dict[str, Any]) -> None:
        with sqlite3.connect(self.runtime.db_path) as conn:
            from .sqlite_store import initialize_schema

            conn.row_factory = sqlite3.Row
            initialize_schema(conn)
            ConflictDetectionService().record_conflict(conn, conflict, update_id=update_id)

    def _write_sync_event(
        self,
        event_writer: EventOutbox,
        event_type: str,
        update: dict[str, Any],
        *,
        workbook_path: Path,
        sync_attempt_id: str,
        workbook_fingerprint_before: dict[str, Any] | None = None,
        workbook_fingerprint_after: dict[str, Any] | None = None,
        actual_workbook_values_before_write: dict[str, Any] | None = None,
        values_written: dict[str, Any] | None = None,
        validation_result: dict[str, Any] | None = None,
        conflict_result: dict[str, Any] | None = None,
        write_result: dict[str, Any] | None = None,
        error_type: str = "",
        error_message: str = "",
        backup_path: str = "",
        lock_id: str = "",
    ) -> dict[str, Any]:
        field_name = str(update.get("field_name") or update.get("field") or "")
        expected = update.get("expected_original_value", update.get("original_value"))
        proposed = update.get("proposed_value")
        return event_writer.create_event(
            event_type=event_type,
            action=str(update.get("source_action") or "sync_pending_update"),
            entity_type=str(update.get("entity_type") or ""),
            entity_id=str(update.get("entity_id") or ""),
            payload={
                "workbook_path": str(workbook_path),
                "workbook_fingerprint_before": workbook_fingerprint_before or {},
                "workbook_fingerprint_after": workbook_fingerprint_after or {},
                "lock_id": lock_id,
                "sync_attempt_id": sync_attempt_id,
                "pending_update_ids": [str(update.get("pending_update_id") or update.get("update_id") or "")],
                "field_changes": [
                    {
                        "field_name": field_name,
                        "expected_original_value": expected,
                        "proposed_value": proposed,
                    }
                ],
                "expected_original_values": {field_name: expected},
                "actual_workbook_values_before_write": actual_workbook_values_before_write or {},
                "proposed_values": {field_name: proposed},
                "values_written": values_written or {},
                "validation_result": validation_result or {"status": str(update.get("validation_status") or "not_run")},
                "conflict_result": conflict_result or {"status": "not_checked"},
                "write_result": write_result or {"status": "not_run"},
                "error_type": error_type,
                "error_message": error_message,
                "backup_path": backup_path,
                "source_view": str(update.get("source_view") or ""),
                "source_action": str(update.get("source_action") or ""),
            },
        )


def _value_key(value: Any) -> str:
    return json.dumps(value, sort_keys=True, default=str)


def workbook_fingerprint(path: str | Path) -> dict[str, Any]:
    target = Path(path)
    if not target.exists():
        return {"exists": False, "path": str(target), "size": 0, "mtime_ns": 0}
    stat = target.stat()
    return {
        "exists": True,
        "path": str(target),
        "size": int(stat.st_size),
        "mtime_ns": int(stat.st_mtime_ns),
        "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
    }


def _find_entity_row(worksheet, id_index: int, entity_id: str) -> int:
    target = str(entity_id or "").strip().casefold()
    for row_number in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row_number, column=id_index).value
        if str(value or "").strip().casefold() == target:
            return row_number
    return -1


def _normalize_field(value: str) -> str:
    return "".join(char for char in str(value or "").strip().casefold() if char.isalnum())


__all__ = [
    "ChangeValidationService",
    "ConflictDetectionService",
    "SyncAttemptLogger",
    "SyncStatusService",
    "WorkbookBackupManager",
    "WorkbookLockManager",
    "WorkbookSyncService",
    "WorkbookUpdateService",
    "workbook_fingerprint",
]
