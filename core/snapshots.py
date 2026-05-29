from __future__ import annotations

import copy
import threading
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .audit_constants import ENTRY_TYPE_AUDITED, ENTRY_TYPE_COMPATIBLE, ENTRY_TYPE_FIELD
from .paths import resolve_project_paths
from .performance import log_performance_event
from .workbook_cache import WorkbookFileSignature, workbook_file_signature
from .workbook_io import worksheet_headers

SNAPSHOT_SHEETS = (
    "EOAT Inventory",
    "Action Items",
    "Photo Index",
    "Issue Log",
    "Interview Notes",
    "Pilot Candidates",
)


@dataclass(frozen=True)
class WorkbookSnapshot:
    project_root: str
    workbook_path: str
    signature: WorkbookFileSignature
    sheets: dict[str, tuple[dict[str, Any], ...]] = field(default_factory=dict)
    audit_rows: tuple[dict[str, Any], ...] = ()
    audit_by_id: dict[str, dict[str, Any]] = field(default_factory=dict)
    audit_ids: tuple[str, ...] = ()
    press_groups: dict[str, tuple[str, ...]] = field(default_factory=dict)
    compatibility_rows: tuple[dict[str, Any], ...] = ()
    open_items_summary_inputs: dict[str, int] = field(default_factory=dict)
    dashboard_stats: dict[str, int] = field(default_factory=dict)


_LOCK = threading.RLock()
_SNAPSHOTS: dict[str, WorkbookSnapshot] = {}
_PENDING_REFRESH: dict[str, threading.Timer] = {}
_SCHEDULED_REFRESH_COUNT: dict[str, int] = {}


def get_workbook_snapshot(project_root: str | Path, *, force: bool = False) -> WorkbookSnapshot:
    paths = resolve_project_paths(project_root)
    signature = workbook_file_signature(paths.master_workbook)
    key = str(Path(paths.project_root).resolve())
    with _LOCK:
        cached = _SNAPSHOTS.get(key)
        if not force and cached and cached.signature == signature:
            return _copy_snapshot(cached)
    started = time.perf_counter()
    snapshot = _build_snapshot(paths.project_root, paths.master_workbook, signature)
    with _LOCK:
        _SNAPSHOTS[key] = snapshot
    log_performance_event(
        paths.project_root,
        "workbook_snapshot.refresh",
        time.perf_counter() - started,
        source="workbook_snapshot",
        page_tool="cache",
        details={"audit_rows": len(snapshot.audit_rows), "sheet_count": len(snapshot.sheets)},
    )
    return _copy_snapshot(snapshot)


def invalidate_workbook_snapshot(project_root: str | Path) -> None:
    key = str(Path(project_root).resolve())
    with _LOCK:
        _SNAPSHOTS.pop(key, None)


def schedule_workbook_snapshot_refresh(project_root: str | Path, *, debounce_seconds: float = 0.75) -> None:
    key = str(Path(project_root).resolve())
    with _LOCK:
        existing = _PENDING_REFRESH.get(key)
        if existing is not None:
            return
        _SCHEDULED_REFRESH_COUNT[key] = _SCHEDULED_REFRESH_COUNT.get(key, 0) + 1

        def _refresh() -> None:
            with _LOCK:
                _PENDING_REFRESH.pop(key, None)
            try:
                get_workbook_snapshot(project_root, force=True)
            except Exception as exc:
                log_performance_event(
                    project_root,
                    "workbook_snapshot.refresh",
                    0.0,
                    source="workbook_snapshot",
                    page_tool="cache",
                    success=False,
                    details={"error": str(exc)},
                    error_count=1,
                )

        timer = threading.Timer(max(0.0, debounce_seconds), _refresh)
        timer.daemon = True
        _PENDING_REFRESH[key] = timer
        timer.start()


def scheduled_snapshot_refresh_count(project_root: str | Path) -> int:
    key = str(Path(project_root).resolve())
    with _LOCK:
        return _SCHEDULED_REFRESH_COUNT.get(key, 0)


def reset_snapshot_service_for_tests() -> None:
    with _LOCK:
        for timer in _PENDING_REFRESH.values():
            timer.cancel()
        _PENDING_REFRESH.clear()
        _SNAPSHOTS.clear()
        _SCHEDULED_REFRESH_COUNT.clear()


def _build_snapshot(project_root: Path, workbook_path: Path, signature: WorkbookFileSignature) -> WorkbookSnapshot:
    if not signature.exists:
        return WorkbookSnapshot(str(project_root), str(workbook_path), signature)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheets: dict[str, tuple[dict[str, Any], ...]] = {}
        for sheet_name in SNAPSHOT_SHEETS:
            if sheet_name in workbook.sheetnames:
                sheets[sheet_name] = tuple(_sheet_rows(workbook[sheet_name]))
        audit_rows = sheets.get("EOAT Inventory", ())
        audit_by_id = {
            str(row.get("Audit ID") or ""): dict(row)
            for row in audit_rows
            if str(row.get("Audit ID") or "").strip()
        }
        compatibility_rows = tuple(
            dict(row)
            for row in audit_rows
            if str(row.get(ENTRY_TYPE_FIELD) or "").strip().casefold() == ENTRY_TYPE_COMPATIBLE.casefold()
        )
        press_groups: dict[str, list[str]] = {}
        for row in audit_rows:
            audit_id = str(row.get("Audit ID") or "").strip()
            machine = str(row.get("Press/Machine #") or "").strip() or "<unassigned>"
            if audit_id:
                press_groups.setdefault(machine, []).append(audit_id)
        dashboard_stats = {
            "audit_rows": len(audit_rows),
            "physical_audits": sum(
                1
                for row in audit_rows
                if str(row.get(ENTRY_TYPE_FIELD) or ENTRY_TYPE_AUDITED).strip().casefold() == ENTRY_TYPE_AUDITED.casefold()
            ),
            "compatibility_rows": len(compatibility_rows),
            "action_items": len(sheets.get("Action Items", ())),
            "photos": len(sheets.get("Photo Index", ())),
        }
        open_items_summary_inputs = {
            "audit_rows": len(audit_rows),
            "action_items": len(sheets.get("Action Items", ())),
            "photo_rows": len(sheets.get("Photo Index", ())),
            "issue_rows": len(sheets.get("Issue Log", ())),
        }
        return WorkbookSnapshot(
            project_root=str(project_root),
            workbook_path=str(workbook_path),
            signature=signature,
            sheets=sheets,
            audit_rows=tuple(dict(row) for row in audit_rows),
            audit_by_id=audit_by_id,
            audit_ids=tuple(sorted(audit_by_id)),
            press_groups={key: tuple(value) for key, value in press_groups.items()},
            compatibility_rows=compatibility_rows,
            open_items_summary_inputs=open_items_summary_inputs,
            dashboard_stats=dashboard_stats,
        )
    finally:
        workbook.close()


def _sheet_rows(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    if not any(headers):
        headers = worksheet_headers(ws)
    output: list[dict[str, Any]] = []
    for values in rows[1:]:
        row = {header: values[index] if index < len(values) else "" for index, header in enumerate(headers) if header}
        if any(value not in (None, "") for value in row.values()):
            output.append(row)
    return output


def _copy_snapshot(snapshot: WorkbookSnapshot) -> WorkbookSnapshot:
    return copy.deepcopy(snapshot)


__all__ = [
    "WorkbookSnapshot",
    "get_workbook_snapshot",
    "invalidate_workbook_snapshot",
    "reset_snapshot_service_for_tests",
    "schedule_workbook_snapshot_refresh",
    "scheduled_snapshot_refresh_count",
]
