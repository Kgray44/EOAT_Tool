from __future__ import annotations

import time
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from .logging import log_tool_run
from .paths import resolve_project_paths
from .result import ToolResult
from .safe_files import backup_file
from .workbook_io import next_empty_row, row_dicts, write_row_by_headers

ACTION_HEADERS = [
    "Action ID",
    "Date Added",
    "Action Item",
    "Related Cell/Press",
    "Owner",
    "Priority",
    "Due Date",
    "Status",
    "Completion Date",
    "Notes",
]


def generate_action_id(project_root: str | Path, action_date: str | None = None) -> str:
    action_date = action_date or date.today().isoformat()
    compact = action_date.replace("-", "")
    workbook_path = resolve_project_paths(project_root).master_workbook
    rows = row_dicts(workbook_path, "Action Items") if workbook_path.exists() else []
    prefix = f"ACT-{compact}-"
    max_number = 0
    for row in rows:
        value = str(row.get("Action ID") or "")
        if value.startswith(prefix):
            try:
                max_number = max(max_number, int(value.rsplit("-", 1)[1]))
            except ValueError:
                continue
    return f"{prefix}{max_number + 1:03d}"


def add_action_item(
    project_root: str | Path,
    action_item: str,
    related_cell_press: str = "",
    owner: str = "",
    priority: str = "Medium",
    due_date: str = "",
    status: str = "Open",
    notes: str = "",
    log_activity: bool = True,
) -> ToolResult:
    started = time.perf_counter()
    paths = resolve_project_paths(project_root)
    workbook_path = paths.master_workbook
    if not workbook_path.exists():
        return ToolResult.fail("action_item", "Action Item Entry", "Master workbook is missing.", errors=[str(workbook_path)])
    if not action_item.strip():
        return ToolResult.fail("action_item", "Action Item Entry", "Action item text is required.")
    action_id = generate_action_id(project_root)
    data = {
        "Action ID": action_id,
        "Date Added": date.today().isoformat(),
        "Action Item": action_item,
        "Related Cell/Press": related_cell_press,
        "Owner": owner,
        "Priority": priority,
        "Due Date": due_date,
        "Status": status,
        "Completion Date": "",
        "Notes": notes,
    }
    try:
        backup = backup_file(workbook_path, workbook_path.parent / "_backups")
        workbook = load_workbook(workbook_path)
        if "Action Items" not in workbook.sheetnames:
            raise ValueError("Action Items sheet is missing.")
        ws = workbook["Action Items"]
        row_number = next_empty_row(ws)
        write_row_by_headers(ws, row_number, data)
        workbook.save(workbook_path)
        workbook.close()
    except Exception as exc:
        return ToolResult.fail(
            "action_item",
            "Action Item Entry",
            "Could not write action item.",
            errors=[str(exc)],
            duration_seconds=time.perf_counter() - started,
        )
    result = ToolResult.ok(
        "action_item",
        "Action Item Entry",
        f"Added action item {action_id}.",
        details=[f"Row: {row_number}", f"Backup: {backup}"],
        files_created=[str(backup)],
        files_modified=[str(workbook_path)],
        metrics={"action_id": action_id, "row": row_number},
        duration_seconds=time.perf_counter() - started,
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result

