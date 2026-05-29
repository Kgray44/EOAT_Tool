from __future__ import annotations

import re
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .audit_constants import (
    CYLINDER_COUNT_FIELD,
    CYLINDER_TYPE_FIELD,
    ENTRY_TYPE_COMPATIBLE,
    ENTRY_TYPE_FIELD,
    MANUAL_COMPLETION_OVERRIDE_FIELD,
    SOURCE_AUDIT_ID_FIELD,
)
from .logging import log_tool_run
from .paths import resolve_project_paths
from .result import ToolResult
from .safe_files import backup_file
from .tool_fields import TOOL_FIELD
from .workbook_cache import invalidate_workbook_cache
from .workbook_io import worksheet_headers

AUDIT_BY_PRESS_SHEET = "Audit by Press"
AUDIT_BY_PRESS_TITLE = "EOAT Audit by Press"
UNASSIGNED_PRESS_GROUP = "Unassigned / Missing Press"
SOURCE_SHEET = "EOAT Inventory"
NA_VALUE = "N/A"
REFRESH_ACTION_NAME = "Refresh Audit by Press View"

VIEW_COLUMNS = [
    "Audit ID",
    "Audit Date",
    "Auditor",
    "Plant/Area",
    "Press/Machine #",
    TOOL_FIELD,
    "Robot Type",
    "EOAT Type",
    "EOAT Moves",
    "Connection Type",
    CYLINDER_COUNT_FIELD,
    CYLINDER_TYPE_FIELD,
    "Cleanroom/Non-Cleanroom",
    "Status",
    MANUAL_COMPLETION_OVERRIDE_FIELD,
    ENTRY_TYPE_FIELD,
    SOURCE_AUDIT_ID_FIELD,
    "Priority",
    "Known Issues",
    "Photo Folder/Link",
]

VIEW_COLUMN_LABELS = {
    "Audit Date": "Date",
    "Known Issues": "Known Issues / Observations",
    "Photo Folder/Link": "Photo Link / Photo Count",
}

_MISSING_PRESS_VALUES = {"", "N/A", "NA", "NONE", "NULL", "UNKNOWN", "UNKNOWN / NOT CHECKED", "NOT APPLICABLE"}


def refresh_audit_by_press_view_file(workbook_path: str | Path) -> None:
    """Rebuild the generated Audit by Press sheet in an existing workbook file."""
    path = Path(workbook_path)
    workbook = load_workbook(path)
    try:
        refresh_audit_by_press_view(workbook)
        workbook.save(path)
        invalidate_workbook_cache(path)
    finally:
        workbook.close()


def refresh_audit_by_press_view_action(project_root: str | Path, log_activity: bool = True) -> ToolResult:
    """Refresh the generated Audit by Press sheet as an explicit user action."""
    started = time.perf_counter()
    paths = resolve_project_paths(project_root)
    workbook_path = paths.master_workbook
    if not workbook_path.exists():
        return ToolResult.fail(
            "audit_by_press_refresh",
            REFRESH_ACTION_NAME,
            "Master workbook is missing.",
            errors=[str(workbook_path)],
            duration_seconds=time.perf_counter() - started,
        )
    try:
        backup = backup_file(workbook_path, workbook_path.parent / "_backups")
        refresh_audit_by_press_view_file(workbook_path)
    except Exception as exc:
        return ToolResult.fail(
            "audit_by_press_refresh",
            REFRESH_ACTION_NAME,
            "Could not refresh Audit by Press view.",
            errors=[str(exc)],
            duration_seconds=time.perf_counter() - started,
        )

    result = ToolResult.ok(
        "audit_by_press_refresh",
        REFRESH_ACTION_NAME,
        "Refreshed Audit by Press view.",
        details=[
            f"Generated sheet: {AUDIT_BY_PRESS_SHEET}",
            f"Source sheet: {SOURCE_SHEET}",
            f"Workbook backup: {backup}",
        ],
        files_created=[str(backup)],
        files_modified=[str(workbook_path)],
        duration_seconds=time.perf_counter() - started,
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def refresh_audit_by_press_view(workbook, refreshed_at: datetime | None = None) -> None:
    """Rebuild the generated press-grouped audit view from EOAT Inventory."""
    refreshed_at = refreshed_at or datetime.now()
    if SOURCE_SHEET not in workbook.sheetnames:
        return

    source = workbook[SOURCE_SHEET]
    source_headers = worksheet_headers(source)
    rows = _source_rows(source, source_headers)

    if AUDIT_BY_PRESS_SHEET in workbook.sheetnames:
        del workbook[AUDIT_BY_PRESS_SHEET]
    sheet = workbook.create_sheet(AUDIT_BY_PRESS_SHEET)

    view_columns = [column for column in VIEW_COLUMNS if column in source_headers]
    if not view_columns:
        view_columns = source_headers[: min(len(source_headers), 14)]

    _write_view(sheet, rows, view_columns, refreshed_at)


def audit_by_press_last_refreshed(workbook) -> datetime | None:
    if AUDIT_BY_PRESS_SHEET not in workbook.sheetnames:
        return None
    value = workbook[AUDIT_BY_PRESS_SHEET]["A2"].value
    if not isinstance(value, str) or "Last refreshed:" not in value:
        return None
    timestamp = value.split("Last refreshed:", 1)[1].strip()
    try:
        return datetime.strptime(timestamp, "%Y-%m-%d %H:%M")
    except ValueError:
        return None


def _source_rows(source, headers: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for values in source.iter_rows(min_row=2, values_only=True):
        non_empty = [value for value in values if value not in (None, "")]
        if not non_empty:
            continue
        if len(non_empty) == 1 and str(non_empty[0]).startswith("Last Updated:"):
            continue
        row = {header: values[index] for index, header in enumerate(headers) if index < len(values)}
        rows.append(row)
    return rows


def _write_view(sheet, rows: list[dict[str, Any]], view_columns: list[str], refreshed_at: datetime) -> None:
    sheet.sheet_properties.outlinePr.summaryBelow = False
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False

    max_col = len(view_columns)
    sheet.cell(row=1, column=1).value = AUDIT_BY_PRESS_TITLE
    sheet.cell(row=2, column=1).value = f"Last refreshed: {refreshed_at.strftime('%Y-%m-%d %H:%M')}"
    sheet.cell(row=2, column=max_col).value = "Generated view from EOAT Inventory"
    for column in range(1, max_col + 1):
        sheet.cell(row=3, column=column).value = VIEW_COLUMN_LABELS.get(view_columns[column - 1], view_columns[column - 1])

    current_row = 4
    grouped_rows = _group_rows(rows)
    for group in grouped_rows:
        header_row = current_row
        counts = _entry_counts(group["rows"])
        header_text = _group_header_text(group["plant"], group["press"], counts)
        sheet.cell(row=header_row, column=1).value = header_text
        sheet.cell(row=header_row, column=max_col).value = (
            f"{counts['physical']} physical / {counts['compatible']} compatible / {counts['total']} total"
        )
        current_row += 1

        detail_start = current_row
        for source_row in group["rows"]:
            for column, header in enumerate(view_columns, start=1):
                sheet.cell(row=current_row, column=column).value = source_row.get(header)
            sheet.row_dimensions[current_row].outlineLevel = 1
            sheet.row_dimensions[current_row].hidden = True
            current_row += 1

        if detail_start < current_row:
            sheet.row_dimensions[header_row].collapsed = True

    if current_row == 4:
        sheet.cell(row=current_row, column=1).value = "No audit rows found."
        current_row += 1

    sheet.auto_filter.ref = f"A3:{get_column_letter(max_col)}3"
    _apply_view_formatting(sheet, max_col, current_row - 1)


def _group_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        press = _cell_text(row.get("Press/Machine #"))
        plant = _cell_text(row.get("Plant/Area"))
        if _is_missing_press(press):
            plant = ""
            press = UNASSIGNED_PRESS_GROUP
        groups[(plant, press)].append(row)

    sorted_groups: list[dict[str, Any]] = []
    for (plant, press), group_rows in groups.items():
        group_rows.sort(key=_detail_sort_key)
        sorted_groups.append({"plant": plant, "press": press, "rows": group_rows})
    sorted_groups.sort(key=lambda group: _group_sort_key(group["plant"], group["press"]))
    return sorted_groups


def _detail_sort_key(row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        _natural_sort_key(row.get(TOOL_FIELD)),
        _natural_sort_key(row.get("Audit ID")),
        _natural_sort_key(row.get("Audit Date")),
    )


def _group_sort_key(plant: str, press: str) -> tuple[Any, ...]:
    if press == UNASSIGNED_PRESS_GROUP:
        return (1, "", (1, ""))
    return (0, _cell_text(plant).casefold(), _natural_sort_key(press))


def _natural_sort_key(value: Any) -> tuple[Any, ...]:
    text = _cell_text(value)
    if not text:
        return (1, "")
    parts = re.split(r"(\d+)", text.casefold())
    key: list[Any] = []
    for part in parts:
        if part.isdigit():
            key.append((0, int(part)))
        elif part:
            key.append((1, part))
    return (0, *key)


def _entry_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    compatible = sum(1 for row in rows if _cell_text(row.get(ENTRY_TYPE_FIELD)).casefold() == ENTRY_TYPE_COMPATIBLE.casefold())
    total = len(rows)
    return {"physical": total - compatible, "compatible": compatible, "total": total}


def _group_header_text(plant: str, press: str, counts: dict[str, int]) -> str:
    count = counts["total"]
    entry_word = "entry" if count == 1 else "entries"
    count_text = f"{counts['physical']} physical, {counts['compatible']} compatible, {count} total {entry_word}"
    if press == UNASSIGNED_PRESS_GROUP:
        return f"{UNASSIGNED_PRESS_GROUP} - {count_text}"
    press_label = _press_label(press)
    if plant and not _is_na_text(plant):
        return f"{plant} / {press_label} - {count_text}"
    return f"Press/Machine # {press} - {count_text}"


def _press_label(press: str) -> str:
    text = _cell_text(press)
    if re.search(r"\b(press|machine)\b", text, flags=re.IGNORECASE):
        return text
    return f"Press {text}"


def _apply_view_formatting(sheet, max_col: int, max_row: int) -> None:
    title_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(bold=True, color="FFFFFF", size=13)
    subtitle_font = Font(italic=True, color="666666")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F1F1F")
    group_fill = PatternFill("solid", fgColor="BDD7EE")
    group_font = Font(bold=True, color="1F1F1F")
    detail_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)

    for row in range(1, max_row + 1):
        sheet.row_dimensions[row].height = 20

    for column in range(1, max_col + 1):
        title_cell = sheet.cell(row=1, column=column)
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = Alignment(vertical="center")

        timestamp_cell = sheet.cell(row=2, column=column)
        timestamp_cell.font = subtitle_font
        timestamp_cell.alignment = Alignment(vertical="center")

        header_cell = sheet.cell(row=3, column=column)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.border = border
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[3].height = 32

    for row_number in range(4, max_row + 1):
        is_group_row = bool(sheet.cell(row=row_number, column=1).value) and sheet.row_dimensions[row_number].outlineLevel == 0
        for column in range(1, max_col + 1):
            cell = sheet.cell(row=row_number, column=column)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_group_row:
                cell.fill = group_fill
                cell.font = group_font
            else:
                cell.fill = detail_fill

    for column, width in enumerate(_column_widths(max_col), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width


def _column_widths(max_col: int) -> list[int]:
    widths = [18, 13, 14, 16, 18, 14, 16, 18, 18, 22, 14, 12, 34, 28]
    if max_col > len(widths):
        widths.extend([18] * (max_col - len(widths)))
    return widths[:max_col]


def _is_missing_press(value: Any) -> bool:
    return _cell_text(value).upper() in _MISSING_PRESS_VALUES


def _is_na_text(value: Any) -> bool:
    return _cell_text(value).upper() == NA_VALUE


def _cell_text(value: Any) -> str:
    return str(value or "").strip()
