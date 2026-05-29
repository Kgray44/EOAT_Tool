from __future__ import annotations

import csv
import json
import time
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .analysis_common import timestamp_for_report
from .logging import log_tool_run
from .paths import resolve_project_paths
from .result import ToolResult
from .safe_files import ensure_directory, safe_write_text

TOOL_ID = "data_import_wizard"
TOOL_NAME = "Data Import Wizard"


@dataclass(frozen=True)
class ImportTypeSpec:
    type_id: str
    label: str
    required_fields: tuple[str, ...]
    optional_fields: tuple[str, ...] = ()
    numeric_fields: tuple[str, ...] = ()
    aliases: dict[str, tuple[str, ...]] = field(default_factory=dict)

    @property
    def all_fields(self) -> tuple[str, ...]:
        return (*self.required_fields, *self.optional_fields)


@dataclass(frozen=True)
class ImportPreview:
    file_path: str
    import_type: str
    import_label: str
    headers: tuple[str, ...]
    row_count: int
    preview_rows: tuple[dict[str, Any], ...]
    mapping: dict[str, str]
    warnings: tuple[str, ...] = ()

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class ImportValidationIssue:
    severity: str
    field: str
    message: str
    row_number: int | None = None

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class ImportDryRun:
    file_path: str
    import_type: str
    import_label: str
    row_count: int
    mapped_rows: tuple[dict[str, Any], ...]
    mapping: dict[str, str]
    issues: tuple[ImportValidationIssue, ...]
    would_write: tuple[str, ...]
    confirmed: bool = False

    @property
    def blockers(self) -> tuple[ImportValidationIssue, ...]:
        return tuple(issue for issue in self.issues if issue.severity == "error")

    def to_dict(self) -> dict[str, Any]:
        return {
            "file_path": self.file_path,
            "import_type": self.import_type,
            "import_label": self.import_label,
            "row_count": self.row_count,
            "mapped_rows": list(self.mapped_rows),
            "mapping": self.mapping,
            "issues": [issue.to_dict() for issue in self.issues],
            "would_write": list(self.would_write),
            "confirmed": self.confirmed,
        }


IMPORT_TYPE_SPECS: dict[str, ImportTypeSpec] = {
    "press_capacity": ImportTypeSpec(
        "press_capacity",
        "Press capacity workbook",
        ("Machine No.", "NGW Part Number"),
        ("NGW Part Description", "Plant/Area", "Press Tonnage"),
        aliases={
            "Machine No.": ("Machine No.", "Machine No", "Machine #", "Machine Number", "Press", "Press #", "Press/Machine #"),
            "NGW Part Number": ("NGW Part Number", "NGW Part #", "Part Number", "Part #", "Tool #"),
            "NGW Part Description": ("NGW Part Description", "Part Description", "Description", "Part Name/Description"),
            "Plant/Area": ("Plant/Area", "Plant", "Area"),
            "Press Tonnage": ("Press Tonnage", "Tonnage", "Capacity"),
        },
    ),
    "downtime_export": ImportTypeSpec(
        "downtime_export",
        "Downtime export",
        ("Date", "Press/Machine #", "Downtime Minutes"),
        ("EOAT-Related Downtime?", "Data Source", "Notes"),
        numeric_fields=("Downtime Minutes",),
        aliases={
            "Date": ("Date", "Shift Date", "Timestamp"),
            "Press/Machine #": ("Press/Machine #", "Machine", "Machine Number", "Press", "Press #"),
            "Downtime Minutes": ("Downtime Minutes", "Downtime", "Minutes Down", "Down Minutes"),
            "EOAT-Related Downtime?": ("EOAT-Related Downtime?", "EOAT Related", "EOAT Downtime"),
            "Data Source": ("Data Source", "Source"),
            "Notes": ("Notes", "Comment", "Comments"),
        },
    ),
    "scrap_export": ImportTypeSpec(
        "scrap_export",
        "Scrap export",
        ("Date", "Press/Machine #", "Scrap Quantity"),
        ("Scrap Reason", "Data Source", "Notes"),
        numeric_fields=("Scrap Quantity",),
        aliases={
            "Date": ("Date", "Shift Date", "Timestamp"),
            "Press/Machine #": ("Press/Machine #", "Machine", "Machine Number", "Press", "Press #"),
            "Scrap Quantity": ("Scrap Quantity", "Scrap Qty", "Scrap Count", "Reject Quantity", "Rejects"),
            "Scrap Reason": ("Scrap Reason", "Reason", "Defect Reason"),
            "Data Source": ("Data Source", "Source"),
            "Notes": ("Notes", "Comment", "Comments"),
        },
    ),
    "maintenance_event_export": ImportTypeSpec(
        "maintenance_event_export",
        "Maintenance event export",
        ("Date", "Press/Machine #", "Maintenance Event Count"),
        ("Maintenance Notes", "Data Source", "Notes"),
        numeric_fields=("Maintenance Event Count",),
        aliases={
            "Date": ("Date", "Event Date", "Timestamp"),
            "Press/Machine #": ("Press/Machine #", "Machine", "Machine Number", "Press", "Press #"),
            "Maintenance Event Count": ("Maintenance Event Count", "Event Count", "Maintenance Count", "Work Order Count"),
            "Maintenance Notes": ("Maintenance Notes", "Work Order Notes", "Description"),
            "Data Source": ("Data Source", "Source"),
            "Notes": ("Notes", "Comment", "Comments"),
        },
    ),
    "cycle_time_baseline": ImportTypeSpec(
        "cycle_time_baseline",
        "Cycle-time baseline",
        ("Date", "Press/Machine #", "Cycle Time"),
        ("Part Family", "Tool #", "Data Source", "Notes"),
        numeric_fields=("Cycle Time",),
        aliases={
            "Date": ("Date", "Run Date", "Timestamp"),
            "Press/Machine #": ("Press/Machine #", "Machine", "Machine Number", "Press", "Press #"),
            "Cycle Time": ("Cycle Time", "Cycle Time Seconds", "Cycle Seconds", "Actual Cycle"),
            "Part Family": ("Part Family", "Family"),
            "Tool #": ("Tool #", "Tool", "Mold", "Mold #"),
            "Data Source": ("Data Source", "Source"),
            "Notes": ("Notes", "Comment", "Comments"),
        },
    ),
    "machine_master_list": ImportTypeSpec(
        "machine_master_list",
        "Machine master list",
        ("Press/Machine #",),
        ("Plant/Area", "Press Tonnage", "Status", "Notes"),
        aliases={
            "Press/Machine #": ("Press/Machine #", "Machine", "Machine Number", "Press", "Press #"),
            "Plant/Area": ("Plant/Area", "Plant", "Area"),
            "Press Tonnage": ("Press Tonnage", "Tonnage", "Capacity"),
            "Status": ("Status", "Active?", "State"),
            "Notes": ("Notes", "Comment", "Comments"),
        },
    ),
    "robot_list": ImportTypeSpec(
        "robot_list",
        "Robot list",
        ("Press/Machine #", "Robot Type"),
        ("Robot Model/Controller", "Plant/Area", "Notes"),
        aliases={
            "Press/Machine #": ("Press/Machine #", "Machine", "Machine Number", "Press", "Press #"),
            "Robot Type": ("Robot Type", "Robot", "Robot Make", "Robot Brand"),
            "Robot Model/Controller": ("Robot Model/Controller", "Robot Model", "Controller", "Robot Controller"),
            "Plant/Area": ("Plant/Area", "Plant", "Area"),
            "Notes": ("Notes", "Comment", "Comments"),
        },
    ),
    "pm_records": ImportTypeSpec(
        "pm_records",
        "PM records",
        ("PM Item", "Status"),
        ("Audit ID", "Press/Machine #", "Completed Date", "Notes", "Photo Evidence Link"),
        aliases={
            "PM Item": ("PM Item", "Task", "Checklist Item", "PM Task"),
            "Status": ("Status", "PM Status", "State"),
            "Audit ID": ("Audit ID", "Audit", "EOAT ID"),
            "Press/Machine #": ("Press/Machine #", "Machine", "Machine Number", "Press", "Press #"),
            "Completed Date": ("Completed Date", "Completion Date", "Date"),
            "Notes": ("Notes", "Comment", "Comments"),
            "Photo Evidence Link": ("Photo Evidence Link", "Photo Link", "Evidence Link"),
        },
    ),
}


def supported_import_types() -> list[ImportTypeSpec]:
    return list(IMPORT_TYPE_SPECS.values())


def detect_import_type(file_path: str | Path) -> str:
    headers, _rows, warnings = _read_tabular_file(file_path)
    scores = {type_id: _mapping_score(headers, spec) for type_id, spec in IMPORT_TYPE_SPECS.items()}
    detected, score = max(scores.items(), key=lambda item: item[1])
    if score <= 0:
        lowered = Path(file_path).name.casefold()
        for type_id in IMPORT_TYPE_SPECS:
            if all(part in lowered for part in type_id.split("_")[:2]):
                return type_id
        raise ValueError("Could not detect import type from headers." + (f" Warnings: {'; '.join(warnings)}" if warnings else ""))
    return detected


def preview_import_file(file_path: str | Path, *, import_type: str | None = None, max_rows: int = 10) -> ImportPreview:
    headers, rows, warnings = _read_tabular_file(file_path)
    type_id = import_type or detect_import_type(file_path)
    spec = _spec(type_id)
    mapping = suggest_column_mapping(headers, type_id)
    return ImportPreview(
        file_path=str(file_path),
        import_type=type_id,
        import_label=spec.label,
        headers=tuple(headers),
        row_count=len(rows),
        preview_rows=tuple(rows[:max_rows]),
        mapping=mapping,
        warnings=tuple(warnings),
    )


def suggest_column_mapping(headers: list[str] | tuple[str, ...], import_type: str) -> dict[str, str]:
    spec = _spec(import_type)
    normalized = {_norm(header): str(header) for header in headers}
    mapping: dict[str, str] = {}
    for field in spec.all_fields:
        aliases = spec.aliases.get(field, (field,))
        for alias in aliases:
            source = normalized.get(_norm(alias))
            if source:
                mapping[field] = source
                break
    return mapping


def validate_import_file(file_path: str | Path, *, import_type: str | None = None, column_mapping: dict[str, str] | None = None) -> tuple[ImportPreview, tuple[ImportValidationIssue, ...]]:
    preview = preview_import_file(file_path, import_type=import_type)
    type_id = preview.import_type
    spec = _spec(type_id)
    mapping = {**preview.mapping, **(column_mapping or {})}
    _headers, rows, _warnings = _read_tabular_file(file_path)
    issues: list[ImportValidationIssue] = []
    for field in spec.required_fields:
        if field not in mapping:
            issues.append(ImportValidationIssue("error", field, f"Required field is not mapped: {field}."))
    for row_number, row in enumerate(rows, start=2):
        for field in spec.required_fields:
            source = mapping.get(field, "")
            if source and _is_blank(row.get(source)):
                issues.append(ImportValidationIssue("warning", field, f"Required field is blank in source column {source}.", row_number=row_number))
        for field in spec.numeric_fields:
            source = mapping.get(field, "")
            if source and not _is_blank(row.get(source)) and _to_number(row.get(source)) is None:
                issues.append(ImportValidationIssue("warning", field, f"Numeric field could not be parsed from source column {source}.", row_number=row_number))
    return preview, tuple(issues)


def dry_run_import(
    project_root: str | Path,
    file_path: str | Path,
    *,
    import_type: str | None = None,
    column_mapping: dict[str, str] | None = None,
    max_rows: int = 50,
) -> ImportDryRun:
    preview, issues = validate_import_file(file_path, import_type=import_type, column_mapping=column_mapping)
    spec = _spec(preview.import_type)
    mapping = {**preview.mapping, **(column_mapping or {})}
    _headers, rows, _warnings = _read_tabular_file(file_path)
    mapped_rows = tuple(_mapped_row(row, spec, mapping) for row in rows[:max_rows])
    paths = resolve_project_paths(project_root)
    stamp = timestamp_for_report()
    would_write = (
        str(paths.data_imports / preview.import_type / f"{Path(file_path).stem}_{stamp}.json"),
        str(import_log_path(project_root)),
    )
    return ImportDryRun(
        file_path=str(file_path),
        import_type=preview.import_type,
        import_label=preview.import_label,
        row_count=len(rows),
        mapped_rows=mapped_rows,
        mapping=mapping,
        issues=issues,
        would_write=would_write,
        confirmed=False,
    )


def confirm_import(
    project_root: str | Path,
    file_path: str | Path,
    *,
    import_type: str | None = None,
    column_mapping: dict[str, str] | None = None,
    confirmed: bool = False,
    log_activity: bool = True,
) -> ToolResult:
    start = time.perf_counter()
    if not confirmed:
        return ToolResult.fail(TOOL_ID, TOOL_NAME, "Import was not confirmed; no files were written.")
    dry_run = dry_run_import(project_root, file_path, import_type=import_type, column_mapping=column_mapping, max_rows=100000)
    if dry_run.blockers:
        return ToolResult.fail(
            TOOL_ID,
            TOOL_NAME,
            "Import has validation blockers; no files were written.",
            errors=[issue.message for issue in dry_run.blockers],
            structured_data=dry_run.to_dict(),
        )
    paths = resolve_project_paths(project_root)
    output_dir = ensure_directory(paths.data_imports / dry_run.import_type)
    stamp = timestamp_for_report()
    output_path = output_dir / f"{Path(file_path).stem}_{stamp}.json"
    payload = {
        "import_type": dry_run.import_type,
        "import_label": dry_run.import_label,
        "source_file": str(file_path),
        "imported_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "mapping": dry_run.mapping,
        "row_count": dry_run.row_count,
        "rows": list(dry_run.mapped_rows),
        "validation_issues": [issue.to_dict() for issue in dry_run.issues],
    }
    try:
        safe_write_text(output_path, json.dumps(payload, indent=2, sort_keys=True) + "\n", overwrite=False)
        log_path = _append_import_log(project_root, payload, output_path)
    except Exception as exc:
        return ToolResult.fail(TOOL_ID, TOOL_NAME, "Could not write imported data snapshot or import log.", errors=[str(exc)])
    result = ToolResult.ok(
        TOOL_ID,
        TOOL_NAME,
        f"Imported {dry_run.row_count} row(s) into local import staging.",
        details=[
            f"Import type: {dry_run.import_label}",
            "Imported rows were staged as local JSON; master workbooks were not modified.",
            f"Snapshot: {output_path}",
            f"Import log: {log_path}",
        ],
        warnings=[issue.message for issue in dry_run.issues if issue.severity == "warning"],
        files_created=[str(output_path), str(log_path)],
        output_reports=[str(output_path), str(log_path)],
        structured_data={**dry_run.to_dict(), "confirmed": True},
        metrics={"row_count": dry_run.row_count, "warning_count": sum(issue.severity == "warning" for issue in dry_run.issues)},
        duration_seconds=time.perf_counter() - start,
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def import_log_path(project_root: str | Path) -> Path:
    return resolve_project_paths(project_root).data_imports / "import_log.jsonl"


def read_import_log(project_root: str | Path) -> list[dict[str, Any]]:
    path = import_log_path(project_root)
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return rows


def _append_import_log(project_root: str | Path, payload: dict[str, Any], output_path: Path) -> Path:
    path = import_log_path(project_root)
    ensure_directory(path.parent)
    entry = {
        "imported_at": payload["imported_at"],
        "import_type": payload["import_type"],
        "import_label": payload["import_label"],
        "source_file": payload["source_file"],
        "snapshot": str(output_path),
        "row_count": payload["row_count"],
        "warning_count": len(payload["validation_issues"]),
    }
    existing = path.read_text(encoding="utf-8") if path.exists() else ""
    safe_write_text(path, existing + json.dumps(entry, sort_keys=True) + "\n", overwrite=True)
    return path


def _mapped_row(row: dict[str, Any], spec: ImportTypeSpec, mapping: dict[str, str]) -> dict[str, Any]:
    mapped: dict[str, Any] = {}
    for field in spec.all_fields:
        source = mapping.get(field)
        value = row.get(source, "") if source else ""
        mapped[field] = _to_number(value) if field in spec.numeric_fields and not _is_blank(value) else value
    return mapped


def _read_tabular_file(file_path: str | Path) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(path)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    if suffix in {".csv", ".tsv", ".txt"}:
        return _read_delimited(path)
    raise ValueError(f"Unsupported import file type: {path.suffix}")


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        values = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()
    return _rows_from_values(values)


def _read_delimited(path: Path) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    text = path.read_text(encoding="utf-8-sig")
    sample = text[:4096]
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        delimiter = dialect.delimiter
    except csv.Error:
        pass
    reader = csv.reader(text.splitlines(), delimiter=delimiter)
    return _rows_from_values([list(row) for row in reader])


def _rows_from_values(values: list[list[Any]]) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    header_index = None
    for index, row in enumerate(values[:20]):
        non_blank = [_clean(value) for value in row if _clean(value)]
        if len(non_blank) >= 2:
            header_index = index
            break
    if header_index is None:
        return [], [], ["No header row found."]
    headers = [_clean(value) or f"Column {index + 1}" for index, value in enumerate(values[header_index])]
    rows: list[dict[str, Any]] = []
    for values_row in values[header_index + 1 :]:
        if not any(not _is_blank(value) for value in values_row):
            continue
        rows.append({headers[index]: values_row[index] if index < len(values_row) else "" for index in range(len(headers))})
    return headers, rows, warnings


def _mapping_score(headers: list[str], spec: ImportTypeSpec) -> int:
    mapping = suggest_column_mapping(headers, spec.type_id)
    return len(set(spec.required_fields).intersection(mapping)) * 10 + len(set(spec.optional_fields).intersection(mapping))


def _spec(import_type: str) -> ImportTypeSpec:
    try:
        return IMPORT_TYPE_SPECS[import_type]
    except KeyError as exc:
        raise ValueError(f"Unsupported import type: {import_type}") from exc


def _to_number(value: Any) -> float | None:
    try:
        return float(str(value).strip().replace(",", ""))
    except (TypeError, ValueError):
        return None


def _is_blank(value: Any) -> bool:
    return not _clean(value)


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _norm(value: Any) -> str:
    return "".join(char for char in _clean(value).casefold() if char.isalnum())


__all__ = [
    "IMPORT_TYPE_SPECS",
    "ImportDryRun",
    "ImportPreview",
    "ImportTypeSpec",
    "ImportValidationIssue",
    "confirm_import",
    "detect_import_type",
    "dry_run_import",
    "import_log_path",
    "preview_import_file",
    "read_import_log",
    "supported_import_types",
    "suggest_column_mapping",
    "validate_import_file",
]
