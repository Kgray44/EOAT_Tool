from __future__ import annotations

import re
import time
from collections.abc import Iterable
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .audit_constants import ENTRY_TYPE_COMPATIBLE, ENTRY_TYPE_FIELD
from .audit_context import infer_audit_context
from .eoat_ids import EOAT_ASSEMBLY_ID_FIELD, infer_eoat_assembly_id_for_photo_row, normalize_eoat_assembly_id
from .logging import log_tool_run
from .paths import resolve_project_paths
from .photo_evidence_rules import all_photo_evidence_rules, photo_evidence_aliases, photo_evidence_rule_by_key
from .result import ToolResult
from .safe_files import backup_file, ensure_directory, safe_write_text
from .tool_fields import TOOL_FIELD
from .validation_findings import ValidationFinding, ValidationSeverity, make_finding
from .workbook_cache import invalidate_workbook_cache
from .workbook_cache import row_dicts_cached as row_dicts
from .workbook_io import worksheet_headers
from .workbook_schema import get_expected_headers

PHOTO_EVIDENCE_TOOL_NAME = "EOAT Photo Evidence Coverage"
LINKED_AUDIT_FIELD_HEADER = "Linked Audit Field"

STATUS_COMPLETE = "complete"
STATUS_PARTIAL = "partial"
STATUS_MISSING = "missing"
STATUS_NOT_APPLICABLE = "not applicable"
STATUS_FOLLOW_UP = "follow-up needed"


@dataclass(frozen=True)
class PhotoEvidenceCategory:
    key: str
    label: str
    applies_when: str
    required_when: str
    recommended_when: str
    example_filename_prefix: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class EvidenceCoverageStatus:
    audit_id: str
    category: str
    label: str
    applies: bool
    required: bool
    present: bool
    photo_count: int
    status: str
    warning: str = ""

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class AuditEvidenceCoverage:
    audit_id: str
    machine: str
    statuses: tuple[EvidenceCoverageStatus, ...]
    row_data: dict[str, Any] | None = field(default=None, repr=False, compare=False)
    related_photo_count: int = 0
    inherited_photo_count: int = 0
    inherited_photo_sources: tuple[str, ...] = ()
    compatible_evidence_accepted: bool = False

    @property
    def missing_required_count(self) -> int:
        if self.compatible_evidence_accepted:
            return 0
        return sum(1 for status in self.statuses if status.required and not status.present)

    @property
    def follow_up_needed_count(self) -> int:
        return sum(1 for status in self.statuses if status.status == STATUS_FOLLOW_UP)

    @property
    def complete_count(self) -> int:
        return sum(1 for status in self.statuses if status.status == STATUS_COMPLETE)

    def to_dict(self) -> dict[str, Any]:
        return {
            "audit_id": self.audit_id,
            "machine": self.machine,
            "missing_required_count": self.missing_required_count,
            "follow_up_needed_count": self.follow_up_needed_count,
            "complete_count": self.complete_count,
            "related_photo_count": self.related_photo_count,
            "inherited_photo_count": self.inherited_photo_count,
            "inherited_photo_sources": list(self.inherited_photo_sources),
            "compatible_evidence_accepted": self.compatible_evidence_accepted,
            "statuses": [status.to_dict() for status in self.statuses],
        }


def _category_from_rule(rule) -> PhotoEvidenceCategory:
    note = rule.help_text or f"Photo evidence category: {rule.label}."
    return PhotoEvidenceCategory(
        rule.key,
        rule.label,
        note,
        note,
        note,
        rule.example_filename_prefix,
    )


PHOTO_EVIDENCE_CATEGORIES: tuple[PhotoEvidenceCategory, ...] = tuple(
    _category_from_rule(rule) for rule in all_photo_evidence_rules()
)

PHOTO_CATEGORY_ALIASES = photo_evidence_aliases()


def photo_evidence_categories() -> list[PhotoEvidenceCategory]:
    return list(PHOTO_EVIDENCE_CATEGORIES)


def audit_photo_intake_root(project_root: str | Path) -> Path:
    return resolve_project_paths(project_root).audit_root / "Photos" / "Incoming"


def audit_photo_intake_folder(project_root: str | Path, audit_id: str) -> Path:
    return audit_photo_intake_root(project_root) / _safe_folder_part(audit_id or "Unassigned_Audit")


def create_audit_photo_intake_folder(
    project_root: str | Path, audit_id: str, *, log_activity: bool = True
) -> ToolResult:
    started = time.perf_counter()
    audit_id = _text(audit_id)
    if not audit_id:
        return ToolResult.fail("photo_evidence_intake_folder", PHOTO_EVIDENCE_TOOL_NAME, "Audit ID is required.")
    folder = ensure_directory(audit_photo_intake_folder(project_root, audit_id))
    result = ToolResult.ok(
        "photo_evidence_intake_folder",
        PHOTO_EVIDENCE_TOOL_NAME,
        "Created audit photo intake folder.",
        details=[f"Audit ID: {audit_id}", f"Folder: {folder}"],
        files_created=[str(folder)],
        metrics={"audit_id": audit_id},
        duration_seconds=time.perf_counter() - started,
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def export_photo_checklist(project_root: str | Path, audit_id: str, *, log_activity: bool = True) -> ToolResult:
    started = time.perf_counter()
    audit_id = _text(audit_id)
    if not audit_id:
        return ToolResult.fail("photo_evidence_checklist", PHOTO_EVIDENCE_TOOL_NAME, "Audit ID is required.")
    folder = ensure_directory(audit_photo_intake_folder(project_root, audit_id))
    row = _find_audit_row(project_root, audit_id)
    markdown = build_photo_checklist_markdown(project_root, audit_id, row=row)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = safe_write_text(
        folder / f"Photo_Checklist_{_safe_folder_part(audit_id)}_{stamp}.md", markdown, overwrite=False
    )
    result = ToolResult.ok(
        "photo_evidence_checklist",
        PHOTO_EVIDENCE_TOOL_NAME,
        "Exported audit photo checklist.",
        details=[f"Audit ID: {audit_id}", f"Checklist: {path}", f"Intake folder: {folder}"],
        files_created=[str(path)],
        output_reports=[str(path)],
        metrics={"audit_id": audit_id},
        duration_seconds=time.perf_counter() - started,
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def build_photo_checklist_markdown(
    project_root: str | Path, audit_id: str, *, row: dict[str, Any] | None = None
) -> str:
    audit_id = _text(audit_id)
    row = row if row is not None else _find_audit_row(project_root, audit_id)
    coverage = evidence_coverage_for_audit(project_root, audit_id, row=row)
    machine = _text((row or {}).get("Press/Machine #")) or "N/A"
    audit_context = infer_audit_context(row or {}) if row else "N/A"
    eoat_assembly_id = _text((row or {}).get(EOAT_ASSEMBLY_ID_FIELD)) or "N/A"
    tool_number = _text((row or {}).get(TOOL_FIELD)) or "N/A"
    eoat_type = _text((row or {}).get("EOAT Type")) or "N/A"
    lines = [
        f"# Photo Evidence Checklist - {audit_id or 'Unassigned Audit'}",
        "",
        f"- Audit ID: {audit_id or 'N/A'}",
        f"- Audit Context: {audit_context}",
        f"- EOAT Assembly ID: {eoat_assembly_id}",
        f"- Tool #: {tool_number}",
        f"- Press/Machine #: {machine}",
        f"- EOAT Type: {eoat_type}",
        f"- Intake folder: {audit_photo_intake_folder(project_root, audit_id)}",
        "",
        "## Expected Photos",
        "| Needed | Category | Example Filename Prefix | Status | Notes |",
        "| --- | --- | --- | --- | --- |",
    ]
    status_by_key = {status.category: status for status in coverage.statuses} if coverage else {}
    for category in PHOTO_EVIDENCE_CATEGORIES:
        status = status_by_key.get(category.key)
        needed = "Required" if status and status.required else "Recommended" if status and status.applies else "N/A"
        status_text = status.status if status else "not checked"
        note = status.warning if status and status.warning else category.recommended_when
        lines.append(f"| {needed} | {category.label} | {category.example_filename_prefix} | {status_text} | {note} |")
    lines.extend(
        [
            "",
            "## Intake Notes",
            "- Keep this folder local to the project.",
            "- Do not add real photos to source control.",
            "- Rename or intake photos through the EOAT Command Center Photos page when ready.",
        ]
    )
    return "\n".join(lines) + "\n"


def evidence_coverage_for_project(project_root: str | Path) -> list[AuditEvidenceCoverage]:
    paths = resolve_project_paths(project_root)
    if not paths.master_workbook.exists():
        return []
    rows = row_dicts(paths.master_workbook, "EOAT Inventory")
    photo_rows = row_dicts(paths.master_workbook, "Photo Index")
    coverage_rows: list[AuditEvidenceCoverage] = []
    for row in rows:
        audit_id = _text(row.get("Audit ID"))
        if not audit_id:
            continue
        coverage = evidence_coverage_for_audit(
            project_root,
            audit_id,
            row=row,
            photo_rows=photo_rows,
            inventory_rows=rows,
        )
        if coverage is not None:
            coverage_rows.append(coverage)
    return coverage_rows


def evidence_coverage_for_audit(
    project_root: str | Path,
    audit_id: str,
    *,
    row: dict[str, Any] | None = None,
    photo_rows: list[dict[str, Any]] | None = None,
    inventory_rows: list[dict[str, Any]] | None = None,
) -> AuditEvidenceCoverage | None:
    audit_id = _text(audit_id)
    if not audit_id:
        return None
    row = row if row is not None else _find_audit_row(project_root, audit_id)
    if row is None:
        return None
    photo_rows = photo_rows if photo_rows is not None else _photo_rows(project_root)
    inventory_rows = inventory_rows if inventory_rows is not None else _inventory_rows(project_root)
    direct_photos = _photo_rows_for_audit(photo_rows, row)
    inherited_photos: list[dict[str, Any]] = []
    inherited_folder_count = 0
    inherited_sources: tuple[str, ...] = ()
    if _is_compatible_row(row):
        inherited_photos, inherited_folder_count, inherited_sources = _compatible_inherited_evidence(
            project_root, row, photo_rows, inventory_rows
        )
    related_photos = _dedupe_photo_rows([*direct_photos, *inherited_photos])
    compatible_evidence_accepted = _is_compatible_row(row) and bool(related_photos or inherited_folder_count)
    statuses = tuple(
        _coverage_status_for_category(row, category, related_photos) for category in PHOTO_EVIDENCE_CATEGORIES
    )
    return AuditEvidenceCoverage(
        audit_id=audit_id,
        machine=_text(row.get("Press/Machine #")),
        statuses=statuses,
        row_data=dict(row),
        related_photo_count=len(related_photos),
        inherited_photo_count=len(inherited_photos) + inherited_folder_count,
        inherited_photo_sources=inherited_sources,
        compatible_evidence_accepted=compatible_evidence_accepted,
    )


def indexed_photos_for_audit(project_root: str | Path, audit_id: str) -> list[dict[str, Any]]:
    target = _text(audit_id).casefold()
    if not target:
        return []
    audit_row = _find_audit_row(project_root, audit_id)
    photos = _photo_rows(project_root)
    if audit_row is None:
        return [dict(row) for row in photos if _text(row.get("Related Audit ID")).casefold() == target]
    return [dict(row) for row in _photo_rows_for_audit(photos, audit_row)]


def indexed_photos_for_eoat(project_root: str | Path, eoat_assembly_id: str) -> list[dict[str, Any]]:
    target = normalize_eoat_assembly_id(eoat_assembly_id).casefold()
    if not target:
        return []
    return [
        dict(row)
        for row in _photo_rows(project_root)
        if normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD)).casefold() == target
    ]


def resolve_indexed_photo_path(project_root: str | Path, photo_row: dict[str, Any]) -> Path:
    stored_relative_path = _text(photo_row.get("Stored Relative Path"))
    if stored_relative_path:
        path = Path(stored_relative_path)
        return path if path.is_absolute() else Path(project_root) / path
    folder_text = _text(photo_row.get("Folder Path"))
    filename = _text(photo_row.get("Stored Filename")) or _text(photo_row.get("Photo Filename"))
    folder = Path(folder_text) if folder_text else Path(project_root)
    if folder_text and not folder.is_absolute():
        folder = Path(project_root) / folder
    return folder / filename if filename else folder


def photo_index_path_findings(
    project_root: str | Path, photo_rows: list[dict[str, Any]] | None = None
) -> list[ValidationFinding]:
    photo_rows = photo_rows if photo_rows is not None else _photo_rows(project_root)
    inventory_by_audit_id = _inventory_rows_by_audit_id(project_root)
    findings: list[ValidationFinding] = []
    for row_number, row in enumerate(photo_rows, start=2):
        if not _has_photo_index_content(row):
            continue
        stored_relative_path = _text(row.get("Stored Relative Path"))
        folder = _text(row.get("Folder Path"))
        filename = _text(row.get("Stored Filename")) or _text(row.get("Photo Filename"))
        audit_id = _text(row.get("Related Audit ID"))
        photo_id = _text(row.get("Photo ID")) or "unidentified photo"
        machine = _text(row.get("Press/Machine #"))
        tool = _text(row.get(TOOL_FIELD))
        eoat_id = normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD))
        context = _photo_index_context(audit_id, photo_id, machine, tool)
        if not eoat_id:
            inferred, reason = infer_eoat_assembly_id_for_photo_row(row, list(inventory_by_audit_id.values()))
            if inferred:
                findings.append(
                    make_finding(
                        ValidationSeverity.AUTO_FIXABLE,
                        "photo_eoat_link",
                        f"{context}: Photo Index row is missing EOAT Assembly ID; can infer {inferred}.",
                        sheet_name="Photo Index",
                        row_number=row_number,
                        audit_id=audit_id,
                        machine_number=machine,
                        column_name=EOAT_ASSEMBLY_ID_FIELD,
                        current_value=f"Photo ID={photo_id}; Tool #={tool or 'N/A'}",
                        expected_behavior="New indexed photos should carry EOAT Assembly ID as the primary EOAT link.",
                        recommended_action="Run Repair Photo EOAT Links from the Photos page.",
                        fix_available=True,
                        fix_id="repair_photo_eoat_links",
                        source_validator="photo_evidence",
                    )
                )
            elif reason == "ambiguous":
                findings.append(
                    make_finding(
                        ValidationSeverity.WARNING,
                        "photo_eoat_link",
                        f"{context}: Photo Index row is missing EOAT Assembly ID and Tool # is ambiguous.",
                        sheet_name="Photo Index",
                        row_number=row_number,
                        audit_id=audit_id,
                        machine_number=machine,
                        column_name=EOAT_ASSEMBLY_ID_FIELD,
                        current_value=f"Photo ID={photo_id}; Tool #={tool or 'N/A'}",
                        expected_behavior="Ambiguous legacy photo links should be reviewed manually.",
                        recommended_action="Select the correct EOAT Assembly ID on the Photos page.",
                        source_validator="photo_evidence",
                    )
                )
        if not filename and not stored_relative_path:
            findings.append(
                make_finding(
                    ValidationSeverity.WARNING,
                    "photo_evidence_index",
                    f"{context}: Photo Index row is missing Photo Filename.",
                    sheet_name="Photo Index",
                    row_number=row_number,
                    audit_id=audit_id,
                    machine_number=machine,
                    column_name="Photo Filename",
                    current_value=f"Photo ID={photo_id}; Tool #={tool or 'N/A'}",
                    expected_behavior="Indexed photo rows should record the local photo filename.",
                    recommended_action="Use the Photos page to re-link the photo or intake the evidence again.",
                    source_validator="photo_evidence",
                )
            )
        if not folder and not stored_relative_path:
            findings.append(
                make_finding(
                    ValidationSeverity.WARNING,
                    "photo_evidence_index",
                    f"{context}: Photo Index row is missing Folder Path.",
                    sheet_name="Photo Index",
                    row_number=row_number,
                    audit_id=audit_id,
                    machine_number=machine,
                    column_name="Folder Path",
                    current_value=f"Photo ID={photo_id}; Tool #={tool or 'N/A'}",
                    expected_behavior="Indexed photo rows should record the local folder containing the photo.",
                    recommended_action="Use the Photos page to re-link the photo folder or intake the evidence again.",
                    source_validator="photo_evidence",
                )
            )
        if stored_relative_path or (folder and filename):
            path = resolve_indexed_photo_path(project_root, row)
            if not path.exists():
                findings.append(
                    make_finding(
                        ValidationSeverity.WARNING,
                        "photo_evidence_path",
                        f"{context}: broken photo path for {photo_id}: {path}",
                        sheet_name="Photo Index",
                        row_number=row_number,
                        audit_id=audit_id,
                        machine_number=machine,
                        column_name="Folder Path",
                        current_value=f"{path}; Tool #={tool or 'N/A'}",
                        expected_behavior="Indexed photos should point to a local file that still exists in the project photo folders.",
                        recommended_action="Use the Photos page to re-link the evidence, confirm the local folder, or intake the photo again.",
                        source_validator="photo_evidence",
                    )
                )
        if audit_id:
            audit_row = inventory_by_audit_id.get(audit_id.casefold())
            if audit_row is None:
                findings.append(
                    make_finding(
                        ValidationSeverity.WARNING,
                        "photo_evidence_relationship",
                        f"{context}: Related Audit ID does not match any EOAT Inventory row.",
                        sheet_name="Photo Index",
                        row_number=row_number,
                        audit_id=audit_id,
                        machine_number=machine,
                        column_name="Related Audit ID",
                        current_value=f"Related Audit ID={audit_id}; Tool #={tool or 'N/A'}",
                        expected_behavior="Related Audit ID values in Photo Index should reference an existing EOAT Inventory audit.",
                        recommended_action="Choose the correct audit from the Photos page dropdown or leave the relationship blank until the audit exists.",
                        source_validator="photo_evidence",
                    )
                )
                continue
            audit_tool = _text(audit_row.get(TOOL_FIELD))
            if tool and audit_tool and tool.casefold() != audit_tool.casefold():
                findings.append(
                    make_finding(
                        ValidationSeverity.WARNING,
                        "photo_evidence_relationship",
                        f"{context}: Tool # does not match EOAT Inventory Tool # {audit_tool}.",
                        sheet_name="Photo Index",
                        row_number=row_number,
                        audit_id=audit_id,
                        machine_number=machine,
                        column_name=TOOL_FIELD,
                        current_value=f"Photo Index Tool #={tool}; EOAT Inventory Tool #={audit_tool}",
                        expected_behavior="When both rows have Tool # values, indexed photos should match the related audit row.",
                        recommended_action="Confirm the correct audit relationship or update the Photo Index Tool #.",
                        source_validator="photo_evidence",
                    )
                )
    return findings


def link_photo_to_audit_field(
    project_root: str | Path, photo_id: str, audit_field: str, *, log_activity: bool = True
) -> ToolResult:
    started = time.perf_counter()
    photo_id = _text(photo_id)
    audit_field = _text(audit_field)
    if not photo_id:
        return ToolResult.fail("photo_evidence_link_field", PHOTO_EVIDENCE_TOOL_NAME, "Photo ID is required.")
    if not audit_field:
        return ToolResult.fail("photo_evidence_link_field", PHOTO_EVIDENCE_TOOL_NAME, "Audit field is required.")
    paths = resolve_project_paths(project_root)
    workbook_path = paths.master_workbook
    if not workbook_path.exists():
        return ToolResult.fail(
            "photo_evidence_link_field",
            PHOTO_EVIDENCE_TOOL_NAME,
            "Master workbook is missing.",
            errors=[str(workbook_path)],
        )

    workbook = None
    try:
        backup = backup_file(workbook_path, workbook_path.parent / "_backups")
        workbook = load_workbook(workbook_path)
        if "Photo Index" not in workbook.sheetnames:
            raise ValueError("Photo Index sheet is missing.")
        ws = workbook["Photo Index"]
        _ensure_headers(ws, get_expected_headers("Photo Index"))
        headers = worksheet_headers(ws)
        if "Photo ID" not in headers:
            raise ValueError("Photo Index sheet is missing Photo ID header.")
        photo_id_col = headers.index("Photo ID") + 1
        linked_col = headers.index(LINKED_AUDIT_FIELD_HEADER) + 1 if LINKED_AUDIT_FIELD_HEADER in headers else None
        notes_col = headers.index("Notes") + 1 if "Notes" in headers else None
        target_row = None
        for row_number in range(2, ws.max_row + 1):
            if _text(ws.cell(row=row_number, column=photo_id_col).value).casefold() == photo_id.casefold():
                target_row = row_number
                break
        if target_row is None:
            raise ValueError(f"Photo ID not found: {photo_id}")
        if linked_col is not None:
            ws.cell(row=target_row, column=linked_col).value = audit_field
        elif notes_col is not None:
            notes_cell = ws.cell(row=target_row, column=notes_col)
            existing = _text(notes_cell.value)
            link_note = f"Linked audit field: {audit_field}"
            notes_cell.value = (
                existing if link_note in existing else "\n".join(part for part in (existing, link_note) if part)
            )
        workbook.save(workbook_path)
        workbook.close()
        workbook = None
        invalidate_workbook_cache(workbook_path)
    except Exception as exc:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        return ToolResult.fail(
            "photo_evidence_link_field",
            PHOTO_EVIDENCE_TOOL_NAME,
            "Photo field link failed.",
            errors=[str(exc)],
            duration_seconds=time.perf_counter() - started,
        )

    result = ToolResult.ok(
        "photo_evidence_link_field",
        PHOTO_EVIDENCE_TOOL_NAME,
        "Linked indexed photo to audit field.",
        details=[f"Photo ID: {photo_id}", f"Audit field: {audit_field}", f"Workbook backup: {backup}"],
        files_modified=[str(workbook_path)],
        files_created=[str(backup)],
        metrics={"photo_id": photo_id, "audit_field": audit_field},
        duration_seconds=time.perf_counter() - started,
    )
    if log_activity:
        warning = log_tool_run(result, project_root)
        if warning:
            result.warnings.append(warning)
    return result


def validate_photo_evidence(project_root: str | Path) -> tuple[list[str], dict[str, int], list[ValidationFinding]]:
    coverages = evidence_coverage_for_project(project_root)
    findings = photo_evidence_findings_from_coverages(coverages)
    path_findings = photo_index_path_findings(project_root)
    findings.extend(path_findings)
    warnings = [finding.message for finding in findings]
    metrics = {
        "photo_evidence_audit_count": len(coverages),
        "photo_evidence_missing_required_count": sum(coverage.missing_required_count for coverage in coverages),
        "photo_evidence_broken_path_count": sum(
            1 for finding in path_findings if finding.category == "photo_evidence_path"
        ),
        "photo_evidence_index_finding_count": len(path_findings),
        "photo_evidence_finding_count": len(findings),
    }
    return warnings, metrics, findings


def missing_evidence_findings(project_root: str | Path) -> list[ValidationFinding]:
    return photo_evidence_findings_from_coverages(evidence_coverage_for_project(project_root))


def photo_evidence_findings_from_coverages(coverages: Iterable[AuditEvidenceCoverage]) -> list[ValidationFinding]:
    findings: list[ValidationFinding] = []
    for coverage in coverages:
        row = coverage.row_data
        if coverage.compatible_evidence_accepted:
            continue
        for status in coverage.statuses:
            if status.required and not status.present:
                findings.append(
                    make_finding(
                        ValidationSeverity.WARNING,
                        "missing_evidence",
                        f"{coverage.audit_id}: required photo evidence missing for {status.label}.",
                        sheet_name="Photo Index",
                        audit_id=coverage.audit_id,
                        machine_number=coverage.machine,
                        column_name="Related Audit ID",
                        current_value=status.status,
                        expected_behavior="Required photo evidence categories should have at least one indexed local photo before the audit is treated as complete.",
                        recommended_action="Create an audit intake folder, capture the missing evidence, and intake/index the photos locally.",
                        source_validator="photo_evidence",
                    )
                )
        findings.extend(_photo_status_consistency_findings(coverage, row))
        findings.extend(_specific_photo_evidence_findings(coverage, row))
    return findings


def pm_bom_evidence_status(project_root: str | Path, audit_id: str) -> dict[str, Any]:
    coverage = evidence_coverage_for_audit(project_root, audit_id)
    if coverage is None:
        return {"audit_id": audit_id, "missing_evidence": True, "missing_categories": []}
    missing_categories = (
        []
        if coverage.compatible_evidence_accepted
        else [status.category for status in coverage.statuses if status.required and not status.present]
    )
    return {
        "audit_id": coverage.audit_id,
        "machine": coverage.machine,
        "missing_evidence": bool(missing_categories),
        "missing_categories": missing_categories,
        "missing_required_count": coverage.missing_required_count,
    }


def _coverage_status_for_category(
    row: dict[str, Any],
    category: PhotoEvidenceCategory,
    photo_rows: list[dict[str, Any]],
) -> EvidenceCoverageStatus:
    applies = _category_applies(row, category.key)
    required = applies and _category_required(row, category.key)
    recommended = applies and _category_recommended(row, category.key)
    matching = _photo_rows_for_category(photo_rows, category.key)
    photo_count = len(matching)
    present = photo_count > 0
    if not applies:
        status = STATUS_NOT_APPLICABLE
    elif present:
        status = STATUS_COMPLETE
    elif required:
        status = STATUS_MISSING
    elif recommended:
        status = STATUS_FOLLOW_UP
    else:
        status = STATUS_NOT_APPLICABLE
    warning = f"Missing required evidence: {category.label}." if required and not present else ""
    return EvidenceCoverageStatus(
        audit_id=_text(row.get("Audit ID")),
        category=category.key,
        label=category.label,
        applies=applies,
        required=required,
        present=present,
        photo_count=photo_count,
        status=status,
        warning=warning,
    )


def _category_applies(row: dict[str, Any], key: str) -> bool:
    try:
        return photo_evidence_rule_by_key(key).applies(row)
    except KeyError:
        return False


def _category_required(row: dict[str, Any], key: str) -> bool:
    try:
        return photo_evidence_rule_by_key(key).required(row)
    except KeyError:
        return False


def _category_recommended(row: dict[str, Any], key: str) -> bool:
    try:
        return photo_evidence_rule_by_key(key).recommended(row)
    except KeyError:
        return False


def _specific_photo_evidence_findings(
    coverage: AuditEvidenceCoverage, row: dict[str, Any] | None
) -> list[ValidationFinding]:
    if row is None:
        return []
    status_by_key = {status.category: status for status in coverage.statuses}
    findings: list[ValidationFinding] = []
    if _is_complete_or_audited(row) and any(status.required and not status.present for status in coverage.statuses):
        findings.append(
            _photo_finding(
                coverage,
                "Audit is marked complete/audited but required photo evidence is missing.",
                "Status",
                row.get("Status"),
            )
        )
    if _is_pilot_candidate(row) and not _any_present(coverage):
        findings.append(
            _photo_finding(
                coverage,
                "Pilot candidate lacks before photo evidence.",
                "Pilot Candidate?",
                row.get("Pilot Candidate?"),
            )
        )
    if _has_meaningful_issue(row) and not _any_present(coverage):
        findings.append(
            _photo_finding(
                coverage, "Audit issue has no supporting photo evidence.", "Known Issues", row.get("Known Issues")
            )
        )
    if _documentation_marked_complete(row) and not (
        _text(row.get("Photo Folder/Link")) or status_by_key.get("process_binder_reference", _empty_status()).present
    ):
        findings.append(
            _photo_finding(
                coverage,
                "Documentation is marked complete but no document/photo reference is recorded.",
                "Photo Folder/Link",
                row.get("Photo Folder/Link"),
            )
        )
    sensor_status = status_by_key.get("sensors")
    if _is_yes(row.get("Sensors Present?")) and sensor_status and not sensor_status.present:
        findings.append(
            _photo_finding(
                coverage,
                "Sensors Present? is Yes but no sensor photo is indexed.",
                "Sensors Present?",
                row.get("Sensors Present?"),
            )
        )
    qd_status = status_by_key.get("quick_disconnects")
    if _is_yes(row.get("Quick Disconnects Present?")) and qd_status and not qd_status.present:
        findings.append(
            _photo_finding(
                coverage,
                "Quick Disconnects Present? is Yes but no quick disconnect photo is indexed.",
                "Quick Disconnects Present?",
                row.get("Quick Disconnects Present?"),
            )
        )
    return findings


def _photo_status_consistency_findings(
    coverage: AuditEvidenceCoverage, row: dict[str, Any] | None
) -> list[ValidationFinding]:
    if row is None:
        return []
    findings: list[ValidationFinding] = []
    photos_taken = _text(row.get("Photos Taken?")).casefold()
    link = _text(row.get("Photo Folder/Link"))
    has_indexed_photos = coverage.related_photo_count > 0
    if photos_taken == "yes" and not has_indexed_photos and not link:
        findings.append(
            _photo_finding(
                coverage,
                "Photos Taken? is Yes but no indexed photos or Photo Folder/Link are recorded.",
                "Photo Folder/Link",
                row.get("Photo Folder/Link"),
            )
        )
    if photos_taken in {"", "no"} and has_indexed_photos:
        findings.append(
            _photo_finding(
                coverage,
                "Indexed photos exist but Photos Taken? is not marked Yes.",
                "Photos Taken?",
                row.get("Photos Taken?"),
            )
        )
    return findings


def _photo_finding(
    coverage: AuditEvidenceCoverage, message: str, column_name: str, current_value: Any
) -> ValidationFinding:
    return make_finding(
        ValidationSeverity.WARNING,
        "missing_evidence",
        f"{coverage.audit_id}: {message}",
        sheet_name="EOAT Inventory",
        audit_id=coverage.audit_id,
        machine_number=coverage.machine,
        column_name=column_name,
        current_value=current_value,
        expected_behavior="Evidence-sensitive audit decisions should have local photo or document references.",
        recommended_action="Use the Photos page to create an audit intake folder, capture evidence, and intake/index the photos locally.",
        source_validator="photo_evidence",
    )


def _photo_rows_for_audit(photo_rows: list[dict[str, Any]], row: dict[str, Any]) -> list[dict[str, Any]]:
    audit_id = _text(row.get("Audit ID")).casefold()
    machine = _text(row.get("Press/Machine #")).casefold()
    tool = _text(row.get(TOOL_FIELD)).casefold()
    eoat_id = normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD)).casefold()
    matches: list[dict[str, Any]] = []
    for photo in photo_rows:
        related_audit = _text(photo.get("Related Audit ID")).casefold()
        photo_machine = _text(photo.get("Press/Machine #")).casefold()
        photo_tool = _text(photo.get(TOOL_FIELD)).casefold()
        photo_eoat_id = normalize_eoat_assembly_id(photo.get(EOAT_ASSEMBLY_ID_FIELD)).casefold()
        if audit_id and related_audit == audit_id:
            matches.append(photo)
            continue
        if related_audit:
            continue
        if eoat_id and photo_eoat_id == eoat_id:
            matches.append(photo)
            continue
        if machine and photo_machine == machine:
            if tool and photo_tool and photo_tool != tool:
                continue
            matches.append(photo)
            continue
        if tool and not photo_machine and photo_tool == tool:
            matches.append(photo)
    return matches


def _compatible_inherited_evidence(
    project_root: str | Path,
    row: dict[str, Any],
    photo_rows: list[dict[str, Any]],
    inventory_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], int, tuple[str, ...]]:
    matches: list[dict[str, Any]] = []
    sources: list[str] = []
    source_id = _text(row.get("Source Audit ID"))
    source_row = _inventory_row_by_audit_id(inventory_rows, source_id)
    if source_row is not None:
        source_photos = _photo_rows_for_audit(photo_rows, source_row)
        if source_photos:
            matches.extend(source_photos)
            sources.append(f"Source Audit ID {source_id}")

    eoat_id = normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD))
    tool = _text(row.get(TOOL_FIELD)).casefold()
    for photo in photo_rows:
        photo_eoat_id = normalize_eoat_assembly_id(photo.get(EOAT_ASSEMBLY_ID_FIELD))
        photo_tool = _text(photo.get(TOOL_FIELD)).casefold()
        if eoat_id and photo_eoat_id.casefold() == eoat_id.casefold():
            matches.append(photo)
            continue
        if tool and photo_tool == tool:
            matches.append(photo)
    if eoat_id and any(
        normalize_eoat_assembly_id(photo.get(EOAT_ASSEMBLY_ID_FIELD)).casefold() == eoat_id.casefold()
        for photo in matches
    ):
        sources.append(f"EOAT Assembly ID {eoat_id}")
    if tool and any(_text(photo.get(TOOL_FIELD)).casefold() == tool for photo in matches):
        sources.append(f"Tool # {_text(row.get(TOOL_FIELD))}")

    folder_photo_count = 0
    for folder_eoat_id in _compatible_eoat_folder_ids(row, source_row):
        count = _photo_folder_image_count(project_root, folder_eoat_id)
        if count:
            folder_photo_count += count
            sources.append(f"Cell_Photos/{folder_eoat_id}")
    return _dedupe_photo_rows(matches), folder_photo_count, tuple(_dedupe_texts(sources))


def _compatible_eoat_folder_ids(row: dict[str, Any], source_row: dict[str, Any] | None) -> tuple[str, ...]:
    ids = [
        normalize_eoat_assembly_id(row.get(EOAT_ASSEMBLY_ID_FIELD)),
        normalize_eoat_assembly_id((source_row or {}).get(EOAT_ASSEMBLY_ID_FIELD)),
    ]
    return tuple(_dedupe_texts([eoat_id for eoat_id in ids if eoat_id]))


def _photo_folder_image_count(project_root: str | Path, eoat_assembly_id: str) -> int:
    folder = resolve_project_paths(project_root).cell_photos / eoat_assembly_id
    if not folder.exists() or not folder.is_dir():
        return 0
    supported = {".jpg", ".jpeg", ".png", ".heic", ".heif"}
    return sum(1 for path in folder.rglob("*") if path.is_file() and path.suffix.lower() in supported)


def _inventory_rows(project_root: str | Path) -> list[dict[str, Any]]:
    paths = resolve_project_paths(project_root)
    if not paths.master_workbook.exists():
        return []
    return row_dicts(paths.master_workbook, "EOAT Inventory")


def _inventory_row_by_audit_id(rows: list[dict[str, Any]], audit_id: str) -> dict[str, Any] | None:
    target = _text(audit_id).casefold()
    if not target:
        return None
    return next((row for row in rows if _text(row.get("Audit ID")).casefold() == target), None)


def _dedupe_photo_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[str, str, str, str]] = set()
    unique: list[dict[str, Any]] = []
    for row in rows:
        key = (
            _text(row.get("Photo ID")).casefold(),
            _text(row.get("Stored Relative Path")).casefold(),
            _text(row.get("Folder Path")).casefold(),
            _text(row.get("Photo Filename") or row.get("Stored Filename")).casefold(),
        )
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    return unique


def _dedupe_texts(values: list[str]) -> list[str]:
    seen: set[str] = set()
    unique: list[str] = []
    for value in values:
        text = _text(value)
        key = text.casefold()
        if not text or key in seen:
            continue
        seen.add(key)
        unique.append(text)
    return unique


def _photo_rows_for_category(photo_rows: list[dict[str, Any]], category_key: str) -> list[dict[str, Any]]:
    aliases = PHOTO_CATEGORY_ALIASES.get(category_key, ())
    matches: list[dict[str, Any]] = []
    for row in photo_rows:
        haystack = _photo_search_text(row)
        if any(alias.casefold() in haystack for alias in aliases):
            matches.append(row)
    return matches


def _photo_search_text(row: dict[str, Any]) -> str:
    fields = [
        "EOAT Area Shown",
        "Photo Type",
        "Description",
        "Photo Filename",
        "Stored Filename",
        "Folder Path",
        "Stored Relative Path",
        TOOL_FIELD,
        EOAT_ASSEMBLY_ID_FIELD,
        LINKED_AUDIT_FIELD_HEADER,
        "Notes",
    ]
    return " ".join(_text(row.get(field)).casefold() for field in fields)


def linked_audit_field_for_photo(row: dict[str, Any]) -> str:
    structured = _text(row.get(LINKED_AUDIT_FIELD_HEADER))
    if structured:
        return structured
    notes = _text(row.get("Notes"))
    for line in notes.splitlines():
        if line.casefold().startswith("linked audit field:"):
            return line.split(":", 1)[1].strip()
    return ""


def _ensure_headers(ws, expected_headers: list[str]) -> list[str]:
    headers = worksheet_headers(ws)
    added: list[str] = []
    for header in expected_headers:
        if header in headers:
            continue
        ws.cell(row=1, column=len(headers) + 1).value = header
        headers.append(header)
        added.append(header)
    return added


def _photo_rows(project_root: str | Path) -> list[dict[str, Any]]:
    paths = resolve_project_paths(project_root)
    if not paths.master_workbook.exists():
        return []
    return row_dicts(paths.master_workbook, "Photo Index")


def _inventory_rows_by_audit_id(project_root: str | Path) -> dict[str, dict[str, Any]]:
    paths = resolve_project_paths(project_root)
    if not paths.master_workbook.exists():
        return {}
    rows: dict[str, dict[str, Any]] = {}
    for row in row_dicts(paths.master_workbook, "EOAT Inventory"):
        audit_id = _text(row.get("Audit ID")).casefold()
        if audit_id:
            rows[audit_id] = row
    return rows


def _find_audit_row(project_root: str | Path, audit_id: str) -> dict[str, Any] | None:
    paths = resolve_project_paths(project_root)
    if not paths.master_workbook.exists():
        return None
    target = _text(audit_id).casefold()
    for row in row_dicts(paths.master_workbook, "EOAT Inventory"):
        if _text(row.get("Audit ID")).casefold() == target:
            return row
    return None


def _has_photo_index_content(row: dict[str, Any]) -> bool:
    fields = (
        "Photo ID",
        "Date Taken",
        "Plant/Area",
        "Press/Machine #",
        TOOL_FIELD,
        EOAT_ASSEMBLY_ID_FIELD,
        "EOAT Area Shown",
        "Photo Type",
        "Original Filename",
        "Stored Filename",
        "Stored Relative Path",
        "Imported At",
        "Photo Filename",
        "Folder Path",
        "Related Audit ID",
        "Related Issue ID",
        LINKED_AUDIT_FIELD_HEADER,
        "Description",
        "Notes",
    )
    return any(_text(row.get(field)) for field in fields)


def _photo_index_context(audit_id: str, photo_id: str, machine: str, tool: str) -> str:
    parts = [audit_id or photo_id, f"Photo ID {photo_id}"]
    if machine:
        parts.append(f"Machine {machine}")
    if tool:
        parts.append(f"Tool # {tool}")
    return " | ".join(parts)


def _is_compatible_row(row: dict[str, Any]) -> bool:
    return _text(row.get(ENTRY_TYPE_FIELD)).casefold() == ENTRY_TYPE_COMPATIBLE.casefold()


def _is_complete_or_audited(row: dict[str, Any]) -> bool:
    status = _text(row.get("Status")).casefold()
    return status in {"complete", "audited"}


def _is_complete_or_pilot(row: dict[str, Any]) -> bool:
    return _is_complete_or_audited(row) or _is_pilot_candidate(row)


def _is_pilot_candidate(row: dict[str, Any]) -> bool:
    return _text(row.get("Pilot Candidate?")).casefold() in {"yes", "maybe", "candidate for pilot"}


def _is_yes(value: Any) -> bool:
    return _text(value).casefold() == "yes"


def _is_yes_or_partial(value: Any) -> bool:
    return _text(value).casefold() in {"yes", "partial"}


def _has_issue_or_damage(row: dict[str, Any]) -> bool:
    condition_fields = [
        "Tubing Condition",
        "Cable Management Condition",
        "Mounting Hardware Condition",
        "EOAT Alignment Condition",
    ]
    bad_tokens = ("worn", "damage", "damaged", "poor", "loose", "missing", "misaligned", "follow-up", "follow up")
    return _has_meaningful_issue(row) or any(
        any(token in _text(row.get(field)).casefold() for token in bad_tokens) for field in condition_fields
    )


def _has_meaningful_issue(row: dict[str, Any]) -> bool:
    issue = _text(row.get("Known Issues")).casefold()
    if not issue:
        return False
    return issue not in {
        "none",
        "no",
        "n/a",
        "na",
        "no issue observed.",
        "no issues observed",
        "unknown / not checked",
        "unknown",
    }


def _documentation_marked_complete(row: dict[str, Any]) -> bool:
    return (
        _is_yes(row.get("Process Binder Complete?"))
        or _is_yes(row.get("Drawing/CAD Available?"))
        or _is_yes(row.get("BOM Available?"))
    )


def _any_present(coverage: AuditEvidenceCoverage) -> bool:
    return any(status.present for status in coverage.statuses)


def _empty_status() -> EvidenceCoverageStatus:
    return EvidenceCoverageStatus("", "", "", False, False, False, 0, STATUS_NOT_APPLICABLE)


def _safe_folder_part(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip())
    return cleaned.strip("._-") or "Unassigned_Audit"


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


__all__ = [
    "AuditEvidenceCoverage",
    "EvidenceCoverageStatus",
    "PHOTO_EVIDENCE_CATEGORIES",
    "PhotoEvidenceCategory",
    "audit_photo_intake_folder",
    "audit_photo_intake_root",
    "build_photo_checklist_markdown",
    "create_audit_photo_intake_folder",
    "evidence_coverage_for_audit",
    "evidence_coverage_for_project",
    "export_photo_checklist",
    "indexed_photos_for_audit",
    "indexed_photos_for_eoat",
    "link_photo_to_audit_field",
    "linked_audit_field_for_photo",
    "missing_evidence_findings",
    "photo_index_path_findings",
    "photo_evidence_categories",
    "pm_bom_evidence_status",
    "resolve_indexed_photo_path",
    "validate_photo_evidence",
]
