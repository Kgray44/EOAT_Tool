from __future__ import annotations

import re
import time
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .audit_by_press import refresh_audit_by_press_view
from .audit_constants import (
    AUDIT_CONTEXT_COMPATIBILITY,
    AUDIT_CONTEXT_FIELD,
    AUDIT_CONTEXT_INSTALLED,
    COMPATIBILITY_CONFIDENCE_FIELD,
    COMPATIBILITY_SOURCE_FIELD,
    COMPATIBILITY_SOURCE_PRESS_CAPACITY,
    ENTRY_TYPE_AUDITED,
    ENTRY_TYPE_COMPATIBLE,
    ENTRY_TYPE_FIELD,
    PHYSICAL_AUDIT_VERIFIED_FIELD,
    SOURCE_AUDIT_ID_FIELD,
)
from .audit_entries import (
    ELECTRICAL_WIRING_PRESENT_FIELD,
    _ensure_inventory_headers,
    _migrate_electrical_wiring_presence_rows,
    normalize_audit_entry,
)
from .eoat_ids import EOAT_ASSEMBLY_ID_FIELD
from .logging import log_activity_event
from .paths import get_press_capacity_file, resolve_project_paths
from .press_lookup import lookup_machine
from .result import ToolResult
from .safe_files import backup_file
from .snapshots import get_workbook_snapshot
from .tool_fields import TOOL_FIELD
from .workbook_cache import invalidate_workbook_cache
from .workbook_cache import row_dicts_cached as row_dicts
from .workbook_io import next_empty_row, worksheet_headers, write_row_by_headers
from .workbook_schema import get_expected_headers

CAPACITY_SHEET_NAME = "Capacity"
RELATIONSHIP_KEY_SEPARATOR = "\u241f"
PART_NUMBER_FIELDS = ["NGW Part Number", "Selected Part Number", TOOL_FIELD, "Part Number", "Part #", "Mold #"]
PART_DESCRIPTION_FIELDS = ["NGW Part Description", "Selected Part Description", "Part Name/Description", "Part Family"]
MASTER_MACHINE_FIELDS = ["Press/Machine #", "Machine No.", "Machine Number", "Press"]
CONFLICT_FIELDS = ["NGW Part Description", "Selected Part Description", "Part Name/Description", "EOAT Type"]
OFF_MACHINE_COMPATIBILITY_UPDATE_AND_ADD = "update_and_add"
OFF_MACHINE_COMPATIBILITY_ADD_ONLY = "add_only"
OFF_MACHINE_COMPATIBILITY_LEAVE = "leave"
OFF_MACHINE_COMPATIBILITY_CHOICES = {
    OFF_MACHINE_COMPATIBILITY_UPDATE_AND_ADD,
    OFF_MACHINE_COMPATIBILITY_ADD_ONLY,
    OFF_MACHINE_COMPATIBILITY_LEAVE,
}
MISSING_MACHINE_VALUES = {
    "",
    "-",
    "--",
    "n/a",
    "na",
    "none",
    "null",
    "not applicable",
    "not installed",
    "eoat not installed",
    "bench",
    "bench audit",
    "off machine",
    "off-machine",
    "uninstalled",
    "select",
    "select machine",
    "select machine number",
    "enter machine",
    "enter machine number",
    "unknown",
    "unknown / not checked",
    "not checked",
}
MACHINE_DERIVED_AUDIT_FIELDS = (
    "Plant/Area",
    "Press/Machine #",
    "Robot Type",
    "Robot Model/Controller",
    "Cleanroom/Non-Cleanroom",
    "Part Family",
    "Part Name/Description",
)
CAPACITY_FIELD_ALIASES = {
    "Machine No.": [
        "Machine No.",
        "Machine No",
        "Machine #",
        "Machine Number",
        "Press",
        "Press #",
        "Press/Machine #",
        "Press Machine #",
    ],
    "NGW Part Number": [
        "NGW Part Number",
        "NGW Part #",
        "Tool #",
        "Tool Number",
        "Tool No.",
        "Tool",
        "Part Number",
        "Part #",
        "Mold #",
        "Mold Number",
    ],
    "NGW Part Description": [
        "NGW Part Description",
        "Part Name/Description",
        "Part Name",
        "Part Description",
        "Description",
    ],
    "Plant/Area": ["Plant/Area", "Plant Area", "Plant", "Area"],
    "Robot Type": [
        "Robot Type",
        "Robot",
        "Robot Make",
        "Robot Brand",
        "Robot Manufacturer",
        "Robot/Picker Brand",
        "Robot Picker Brand",
    ],
    "Robot Model/Controller": [
        "Robot Model/Controller",
        "Robot Model",
        "Robot Model #",
        "Robot Make/Model",
        "Robot Model Number",
        "Robot Controller",
        "Robot Controller Type",
        "Controller Type",
        "Controller",
        "Robot/Picker Model #",
        "Robot/Picker Model",
        "Robot Picker Model #",
        "Robot Picker Model",
    ],
    "Cleanroom/Non-Cleanroom": [
        "Cleanroom/Non-Cleanroom",
        "Cleanroom",
        "Clean Room",
        "Cleanroom Status",
        "Environment",
    ],
    "Part Family": ["Part Family", "Family"],
    "Part Name/Description": [
        "Part Name/Description",
        "Part Name",
        "Part Description",
        "NGW Part Description",
        "Description",
    ],
}


@dataclass(frozen=True)
class RequiredRelationship:
    machine_no: str
    part_number: str
    part_description: str = ""
    source_row: int = 0
    machine_data: dict[str, Any] = field(default_factory=dict, compare=False)

    @property
    def key(self) -> tuple[str, str]:
        return relationship_key(self.machine_no, self.part_number)


@dataclass(frozen=True)
class SourceAuditOption:
    audit_id: str
    label: str
    row: dict[str, Any]


@dataclass(frozen=True)
class CompatibilityCandidate:
    machine_no: str
    part_number: str
    part_description: str
    existing_status: str
    recommended_action: str
    existing_audit_ids: tuple[str, ...] = ()

    @property
    def can_create(self) -> bool:
        return self.recommended_action == "Create Compatible Entry"


@dataclass
class CompatibilityCandidateResult:
    source: SourceAuditOption | None
    candidates: list[CompatibilityCandidate] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class PressCapacityLookupResult:
    tool_number: str
    normalized_tool_number: str
    matches: list[RequiredRelationship] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    press_capacity_path: str = ""


@dataclass(frozen=True)
class OffMachineCompatibilityPreview:
    audit_id: str
    tool_number: str
    normalized_tool_number: str
    matches: list[RequiredRelationship] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    press_capacity_path: str = ""
    audit_row: dict[str, Any] = field(default_factory=dict, compare=False)


@dataclass
class SyncResult:
    source_audit_id: str
    updated_count: int = 0
    skipped_count: int = 0
    missing_source: bool = False
    warning_messages: list[str] = field(default_factory=list)
    backup_path: str | None = None


PROTECTED_COMPATIBILITY_SYNC_FIELDS = {
    "Audit ID",
    "Audit Date",
    "Auditor",
    "Plant/Area",
    "Press/Machine #",
    "Machine No.",
    "Machine Number",
    "Press",
    "Robot Type",
    "Robot Model/Controller",
    AUDIT_CONTEXT_FIELD,
    ENTRY_TYPE_FIELD,
    SOURCE_AUDIT_ID_FIELD,
    COMPATIBILITY_SOURCE_FIELD,
    PHYSICAL_AUDIT_VERIFIED_FIELD,
    COMPATIBILITY_CONFIDENCE_FIELD,
}


PROTECTED_COMPATIBILITY_SYNC_NAME_PARTS = (
    "created",
    "timestamp",
    "row id",
    "internal",
    "override",
)


def text_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.upper() == "N/A" else text


def display_text_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return f"{value:.0f}"
    return str(value).strip()


def normalize_tool_identifier(value: Any) -> str:
    text = display_text_value(value).strip()
    if text.casefold() in {"", "n/a", "na", "none", "null", "not applicable"}:
        return ""
    artifact = re.fullmatch(r"([0-9]+)\.0+", text)
    if artifact:
        return artifact.group(1)
    return text


def tool_identifier_key(value: Any) -> str:
    return normalize_tool_identifier(value).casefold()


def machine_identifier_is_missing(value: Any) -> bool:
    return display_text_value(value).strip().casefold() in MISSING_MACHINE_VALUES


def audit_row_has_missing_machine(row: dict[str, Any]) -> bool:
    for field_name in MASTER_MACHINE_FIELDS:
        if field_name not in row:
            continue
        value = row.get(field_name)
        if display_text_value(value).strip() and not machine_identifier_is_missing(value):
            return False
    return True


def off_machine_compatibility_lookup_needed(entry: dict[str, Any]) -> bool:
    return audit_row_has_missing_machine(entry) and bool(normalize_tool_identifier(part_number_from_row(entry)))


def normalize_entry_type(value: Any) -> str:
    text = text_value(value).lower()
    if text == ENTRY_TYPE_COMPATIBLE.lower():
        return ENTRY_TYPE_COMPATIBLE
    return ENTRY_TYPE_AUDITED


def is_unknown_entry_type(row: dict[str, Any]) -> bool:
    return not text_value(row.get(ENTRY_TYPE_FIELD))


def normalize_machine_token(value: Any) -> str:
    text = text_value(value)
    if not text:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    numeric = re.fullmatch(r"\d+(?:\.0+)?", text)
    if numeric:
        return str(int(float(text)))
    prefixed = re.fullmatch(r"(?:press|machine|p|m)\s*[-#]?\s*(\d+)", text, flags=re.IGNORECASE)
    if prefixed:
        return prefixed.group(1)
    return text


def parse_machine_tokens(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, int):
        return [str(value)] if value > 0 else []
    if isinstance(value, float):
        return [str(int(value))] if value.is_integer() and value > 0 else [str(value).strip()]
    tokens: list[str] = []
    for raw_token in str(value).split(","):
        token = normalize_machine_token(raw_token)
        if token:
            tokens.append(token)
    return sort_machine_tokens(tokens)


def sort_machine_tokens(tokens: list[str]) -> list[str]:
    unique = list(dict.fromkeys(token for token in tokens if token))

    def key(token: str) -> tuple[int, int | str]:
        return (0, int(token)) if token.isdigit() else (1, token.lower())

    return sorted(unique, key=key)


def relationship_key(machine_no: Any, part_number: Any) -> tuple[str, str]:
    return normalize_machine_token(machine_no), tool_identifier_key(part_number)


def relationship_label(machine_no: Any, part_number: Any) -> str:
    machine, part = relationship_key(machine_no, part_number)
    return f"{machine}{RELATIONSHIP_KEY_SEPARATOR}{part}"


def part_number_from_row(row: dict[str, Any]) -> str:
    for field_name in PART_NUMBER_FIELDS:
        value = text_value(row.get(field_name))
        if value:
            return value
    return ""


def part_description_from_row(row: dict[str, Any]) -> str:
    for field_name in PART_DESCRIPTION_FIELDS:
        value = text_value(row.get(field_name))
        if value:
            return value
    return ""


def machine_from_audit_row(row: dict[str, Any]) -> str:
    for field_name in MASTER_MACHINE_FIELDS:
        tokens = parse_machine_tokens(row.get(field_name))
        if tokens:
            return tokens[0]
    return ""


def machine_sort_key(value: Any) -> tuple[int, int | str, str]:
    tokens = parse_machine_tokens(value)
    first = tokens[0] if tokens else ""
    if first.isdigit():
        return (0, int(first), first)
    return (1, first.lower(), first)


def audit_row_machine_sort_key(row: dict[str, Any]) -> tuple[int, int | str, str]:
    for field_name in MASTER_MACHINE_FIELDS:
        tokens = parse_machine_tokens(row.get(field_name))
        if tokens:
            first = tokens[0]
            if first.isdigit():
                return (0, int(first), first)
            return (1, first.lower(), first)
    return (1, "", "")


def audit_option_label(row: dict[str, Any]) -> str:
    audit_id = text_value(row.get("Audit ID"))
    eoat_id = text_value(row.get(EOAT_ASSEMBLY_ID_FIELD))
    machine = machine_from_audit_row(row)
    machine_label = f"Machine {machine}" if machine else ""
    part_number = part_number_from_row(row)
    description = part_description_from_row(row)
    entry_type = normalize_entry_type(row.get(ENTRY_TYPE_FIELD))
    return " | ".join(
        piece
        for piece in [audit_id, eoat_id, machine_label, part_number, description, entry_type]
        if piece
    )


def find_capacity_file(project_root_or_path: str | Path) -> Path:
    path = Path(project_root_or_path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return path
    return get_press_capacity_file(path)


def list_audited_source_options(project_root_or_master_path: str | Path) -> list[SourceAuditOption]:
    options: list[SourceAuditOption] = []
    for row in _inventory_rows_for_options(project_root_or_master_path):
        if normalize_entry_type(row.get(ENTRY_TYPE_FIELD)) != ENTRY_TYPE_AUDITED:
            continue
        audit_id = text_value(row.get("Audit ID"))
        if not audit_id:
            continue
        options.append(SourceAuditOption(audit_id=audit_id, label=audit_option_label(row), row=row))
    return sorted(options, key=lambda option: (audit_row_machine_sort_key(option.row), option.audit_id.lower()))


def list_audit_options(project_root_or_master_path: str | Path) -> list[SourceAuditOption]:
    options: list[SourceAuditOption] = []
    for row in _inventory_rows_for_options(project_root_or_master_path):
        audit_id = text_value(row.get("Audit ID"))
        if not audit_id:
            continue
        options.append(SourceAuditOption(audit_id=audit_id, label=audit_option_label(row), row=row))
    return sorted(options, key=lambda option: (audit_row_machine_sort_key(option.row), option.audit_id.lower()))


def _inventory_rows_for_options(project_root_or_master_path: str | Path) -> list[dict[str, Any]]:
    path = Path(project_root_or_master_path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return row_dicts(_master_path(path), "EOAT Inventory")
    try:
        return [dict(row) for row in get_workbook_snapshot(path).audit_rows]
    except Exception:
        return row_dicts(_master_path(path), "EOAT Inventory")


def find_existing_audits_for_machine(
    project_root_or_master_path: str | Path, machine_number: Any
) -> list[SourceAuditOption]:
    requested = set(parse_machine_tokens(machine_number))
    if not requested:
        return []
    return [
        option
        for option in list_audit_options(project_root_or_master_path)
        if normalize_entry_type(option.row.get(ENTRY_TYPE_FIELD)) == ENTRY_TYPE_AUDITED
        and requested.intersection(_audit_row_machine_tokens(option.row))
    ]


def _audit_row_machine_tokens(row: dict[str, Any]) -> set[str]:
    tokens: set[str] = set()
    for field_name in MASTER_MACHINE_FIELDS:
        tokens.update(parse_machine_tokens(row.get(field_name)))
    return tokens


def compatible_rows_by_source_audit_id(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    by_source: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if normalize_entry_type(row.get(ENTRY_TYPE_FIELD)) != ENTRY_TYPE_COMPATIBLE:
            continue
        source_audit_id = text_value(row.get(SOURCE_AUDIT_ID_FIELD))
        if source_audit_id:
            by_source[source_audit_id].append(row)
    return dict(by_source)


def build_compatibility_candidates(
    project_root_or_master_path: str | Path,
    source_audit_id: str,
    press_capacity_path: str | Path | None = None,
) -> CompatibilityCandidateResult:
    master_path = _master_path(project_root_or_master_path)
    capacity_path = (
        Path(press_capacity_path) if press_capacity_path else find_capacity_file(project_root_or_master_path)
    )
    inventory = row_dicts(master_path, "EOAT Inventory")
    source_row = next((row for row in inventory if text_value(row.get("Audit ID")) == source_audit_id), None)
    if not source_row:
        return CompatibilityCandidateResult(source=None, errors=[f"Source audit ID not found: {source_audit_id}"])
    if normalize_entry_type(source_row.get(ENTRY_TYPE_FIELD)) != ENTRY_TYPE_AUDITED:
        return CompatibilityCandidateResult(
            source=None, errors=[f"Source audit ID is not an audited row: {source_audit_id}"]
        )

    source_option = SourceAuditOption(
        audit_id=source_audit_id,
        label=" | ".join(
            piece
            for piece in [
                source_audit_id,
                text_value(source_row.get(EOAT_ASSEMBLY_ID_FIELD)),
                machine_from_audit_row(source_row),
                part_number_from_row(source_row),
                part_description_from_row(source_row),
                text_value(source_row.get("EOAT Type")),
            ]
            if piece
        ),
        row=source_row,
    )
    part_number = part_number_from_row(source_row)
    if not part_number:
        return CompatibilityCandidateResult(
            source=source_option, errors=["Source audit row does not have an NGW Part Number / Tool #."]
        )

    required, warnings = load_required_relationships(capacity_path)
    source_part_key = tool_identifier_key(part_number)
    matching_required = [relationship for relationship in required if relationship.key[1] == source_part_key]
    if not matching_required:
        return CompatibilityCandidateResult(
            source=source_option,
            warnings=[*warnings, f"No Press Capacity relationships found for part {part_number}."],
        )

    by_key = summarize_master_relationships(inventory)
    candidates: list[CompatibilityCandidate] = []
    for relationship in matching_required:
        status = by_key.get(relationship.key, [])
        entry_types = {item["entry_type"] for item in status}
        audit_ids = tuple(
            text_value(item["row"].get("Audit ID")) for item in status if text_value(item["row"].get("Audit ID"))
        )
        conflict = relationship_has_conflict(status, relationship)
        if conflict:
            existing_status = "Conflict"
            action = "Conflict / Review Needed"
        elif ENTRY_TYPE_AUDITED in entry_types:
            existing_status = "Audited"
            action = "Already Audited"
        elif ENTRY_TYPE_COMPATIBLE in entry_types:
            compatible_source_ids = {
                text_value(item["row"].get(SOURCE_AUDIT_ID_FIELD))
                for item in status
                if item["entry_type"] == ENTRY_TYPE_COMPATIBLE
            }
            existing_status = "Compatible"
            if compatible_source_ids and compatible_source_ids <= {source_audit_id}:
                action = "Already Compatible - Linked to this source"
            else:
                action = "Already Compatible - Different Source / Review Needed"
        elif status:
            existing_status = ", ".join(sorted(entry_types))
            action = "Conflict / Review Needed"
        else:
            existing_status = ""
            action = "Create Compatible Entry"
        candidates.append(
            CompatibilityCandidate(
                machine_no=relationship.machine_no,
                part_number=relationship.part_number,
                part_description=relationship.part_description,
                existing_status=existing_status,
                recommended_action=action,
                existing_audit_ids=audit_ids,
            )
        )
    return CompatibilityCandidateResult(
        source=source_option,
        candidates=sorted(candidates, key=lambda item: _machine_sort_key(item.machine_no)),
        warnings=warnings,
    )


def sync_compatible_rows_from_source(master_audit_path: str | Path, source_audit_id: str) -> SyncResult:
    result = SyncResult(source_audit_id=str(source_audit_id))
    master_path = _master_path(master_audit_path)
    if not master_path.exists():
        result.missing_source = True
        result.warning_messages.append(f"Master workbook is missing: {master_path}")
        return result

    workbook = None
    try:
        workbook = load_workbook(master_path)
        if "EOAT Inventory" not in workbook.sheetnames:
            result.missing_source = True
            result.warning_messages.append("EOAT Inventory sheet is missing.")
            return result
        ws = workbook["EOAT Inventory"]
        added_headers = _ensure_inventory_headers(ws, get_expected_headers("EOAT Inventory"))
        if ELECTRICAL_WIRING_PRESENT_FIELD in added_headers:
            _migrate_electrical_wiring_presence_rows(ws)
        headers = worksheet_headers(ws)
        source_row_number = _find_audit_row_number(ws, headers, source_audit_id)
        if source_row_number is None:
            result.missing_source = True
            return result

        source_row = _worksheet_row_dict(ws, headers, source_row_number)
        if normalize_entry_type(source_row.get(ENTRY_TYPE_FIELD)) != ENTRY_TYPE_AUDITED:
            result.skipped_count += 1
            result.warning_messages.append(f"Source audit ID is not an audited row: {source_audit_id}")
            return result

        linked_rows: list[int] = []
        if SOURCE_AUDIT_ID_FIELD not in headers or ENTRY_TYPE_FIELD not in headers:
            return result
        source_col = headers.index(SOURCE_AUDIT_ID_FIELD) + 1
        entry_type_col = headers.index(ENTRY_TYPE_FIELD) + 1
        for row_number in range(2, ws.max_row + 1):
            if row_number == source_row_number:
                continue
            if text_value(ws.cell(row=row_number, column=source_col).value) != str(source_audit_id):
                continue
            entry_type = normalize_entry_type(ws.cell(row=row_number, column=entry_type_col).value)
            if entry_type == ENTRY_TYPE_COMPATIBLE:
                linked_rows.append(row_number)
            else:
                result.skipped_count += 1

        if not linked_rows:
            return result

        result.backup_path = str(backup_file(master_path, master_path.parent / "_backups"))
        copy_fields = [header for header in headers if _can_sync_compatibility_field(header)]
        for row_number in linked_rows:
            for header in copy_fields:
                ws.cell(row=row_number, column=headers.index(header) + 1).value = source_row.get(header)
            result.updated_count += 1
        refresh_audit_by_press_view(workbook)
        workbook.save(master_path)
        invalidate_workbook_cache(master_path)
        return result
    except Exception as exc:
        result.warning_messages.append(f"Could not update linked compatibility rows: {exc}")
        return result
    finally:
        if workbook is not None:
            workbook.close()


def create_compatibility_entries(
    project_root: str | Path,
    source_audit_id: str,
    machine_numbers: list[str],
    press_capacity_path: str | Path | None = None,
) -> ToolResult:
    started = time.perf_counter()
    paths = resolve_project_paths(project_root)
    master_path = paths.master_workbook
    if not master_path.exists():
        return ToolResult.fail(
            "compatibility_entry", "Compatibility Entry", "Master workbook is missing.", errors=[str(master_path)]
        )

    selected = set(parse_machine_tokens(",".join(str(machine) for machine in machine_numbers)))
    if not selected:
        return ToolResult.fail("compatibility_entry", "Compatibility Entry", "No compatible machines were selected.")

    candidate_result = build_compatibility_candidates(project_root, source_audit_id, press_capacity_path)
    if candidate_result.errors:
        return ToolResult.fail(
            "compatibility_entry",
            "Compatibility Entry",
            "Could not build compatibility candidates.",
            errors=candidate_result.errors,
        )
    source = candidate_result.source
    if source is None:
        return ToolResult.fail("compatibility_entry", "Compatibility Entry", "Source audit row is missing.")

    selected_candidates = [candidate for candidate in candidate_result.candidates if candidate.machine_no in selected]
    create_candidates = [candidate for candidate in selected_candidates if candidate.can_create]
    skipped_audited = sum(1 for candidate in selected_candidates if candidate.recommended_action == "Already Audited")
    skipped_compatible = sum(
        1 for candidate in selected_candidates if candidate.recommended_action.startswith("Already Compatible")
    )
    conflicts = sum(
        1 for candidate in selected_candidates if candidate.recommended_action == "Conflict / Review Needed"
    )
    if not create_candidates:
        return ToolResult.ok(
            "compatibility_entry",
            "Compatibility Entry",
            "No compatibility entries were created.",
            details=_creation_summary(0, skipped_audited, skipped_compatible, conflicts),
            warnings=candidate_result.warnings,
            metrics={
                "created": 0,
                "skipped_already_audited": skipped_audited,
                "skipped_already_compatible": skipped_compatible,
                "conflicts": conflicts,
            },
            duration_seconds=time.perf_counter() - started,
        )

    workbook = None
    try:
        backup = backup_file(master_path, master_path.parent / "_backups")
        workbook = load_workbook(master_path)
        if "EOAT Inventory" not in workbook.sheetnames:
            raise ValueError("EOAT Inventory sheet is missing.")
        ws = workbook["EOAT Inventory"]
        added_headers = _ensure_inventory_headers(ws, get_expected_headers("EOAT Inventory"))
        if ELECTRICAL_WIRING_PRESENT_FIELD in added_headers:
            _migrate_electrical_wiring_presence_rows(ws)
        headers = worksheet_headers(ws)
        existing_ids = {
            text_value(ws.cell(row=row_number, column=headers.index("Audit ID") + 1).value)
            for row_number in range(2, ws.max_row + 1)
            if "Audit ID" in headers
        }
        new_ids = _generate_audit_ids(existing_ids, len(create_candidates))
        created = 0
        for audit_id, candidate in zip(new_ids, create_candidates):
            new_row = dict(source.row)
            new_row["Audit ID"] = audit_id
            new_row["Audit Date"] = ""
            new_row["Auditor"] = ""
            new_row["Press/Machine #"] = candidate.machine_no
            new_row[TOOL_FIELD] = candidate.part_number
            if candidate.part_description:
                new_row["Part Name/Description"] = candidate.part_description
            if text_value(new_row.get("Status")).lower() == "audited":
                new_row["Status"] = "Complete"
            new_row[ENTRY_TYPE_FIELD] = ENTRY_TYPE_COMPATIBLE
            new_row[AUDIT_CONTEXT_FIELD] = AUDIT_CONTEXT_COMPATIBILITY
            new_row[SOURCE_AUDIT_ID_FIELD] = source.audit_id
            new_row[COMPATIBILITY_SOURCE_FIELD] = COMPATIBILITY_SOURCE_PRESS_CAPACITY
            new_row[PHYSICAL_AUDIT_VERIFIED_FIELD] = "No"
            new_row[COMPATIBILITY_CONFIDENCE_FIELD] = "Press Capacity"
            new_row = normalize_audit_entry(project_root, new_row)
            write_row_by_headers(ws, next_empty_row(ws), new_row)
            created += 1
        refresh_audit_by_press_view(workbook)
        workbook.save(master_path)
        workbook.close()
        invalidate_workbook_cache(master_path)
    except Exception as exc:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        return ToolResult.fail(
            "compatibility_entry",
            "Compatibility Entry",
            "Could not create compatibility entries.",
            errors=[str(exc)],
            warnings=candidate_result.warnings,
            duration_seconds=time.perf_counter() - started,
        )

    return ToolResult.ok(
        "compatibility_entry",
        "Compatibility Entry",
        f"Created {created} compatibility entries.",
        details=[
            *_creation_summary(created, skipped_audited, skipped_compatible, conflicts),
            f"Workbook backup: {backup}",
        ],
        warnings=candidate_result.warnings,
        files_created=[str(backup)],
        files_modified=[str(master_path)],
        metrics={
            "created": created,
            "skipped_already_audited": skipped_audited,
            "skipped_already_compatible": skipped_compatible,
            "conflicts": conflicts,
        },
        duration_seconds=time.perf_counter() - started,
    )


def find_press_capacity_matches_for_tool(
    project_root_or_path: str | Path,
    tool_number: Any,
    press_capacity_path: str | Path | None = None,
) -> PressCapacityLookupResult:
    capacity_path = Path(press_capacity_path) if press_capacity_path else find_capacity_file(project_root_or_path)
    normalized_tool = normalize_tool_identifier(tool_number)
    if not normalized_tool:
        return PressCapacityLookupResult(
            tool_number=display_text_value(tool_number),
            normalized_tool_number="",
            errors=["Tool # is blank; Press Capacity compatibility lookup was skipped."],
            press_capacity_path=str(capacity_path),
        )
    relationships, warnings = load_required_relationships(capacity_path)
    errors = [warning for warning in warnings if _is_capacity_load_warning(warning)]
    matches = [
        relationship for relationship in relationships if tool_identifier_key(relationship.part_number) == normalized_tool.casefold()
    ]
    return PressCapacityLookupResult(
        tool_number=display_text_value(tool_number),
        normalized_tool_number=normalized_tool,
        matches=matches,
        warnings=[warning for warning in warnings if warning not in errors],
        errors=errors,
        press_capacity_path=str(capacity_path),
    )


def _enrich_relationships_with_machine_lookup(
    project_root: str | Path,
    relationships: list[RequiredRelationship],
) -> tuple[list[RequiredRelationship], list[str]]:
    if not relationships:
        return relationships, []
    lookup_cache: dict[str, tuple[dict[str, str], list[str]]] = {}
    enriched: list[RequiredRelationship] = []
    robot_type_filled = 0
    robot_model_filled = 0
    warnings: list[str] = []
    for relationship in relationships:
        machine_data = dict(relationship.machine_data)
        if relationship.machine_no not in lookup_cache:
            lookup_cache[relationship.machine_no] = _machine_lookup_data_for_compatibility(
                project_root, relationship.machine_no
            )
        lookup_data, lookup_warnings = lookup_cache[relationship.machine_no]
        warnings.extend(lookup_warnings)
        if lookup_data.get("Robot Type") and machine_data.get("Robot Type") != lookup_data["Robot Type"]:
            machine_data["Robot Type"] = lookup_data["Robot Type"]
            robot_type_filled += 1
        if (
            lookup_data.get("Robot Model/Controller")
            and machine_data.get("Robot Model/Controller") != lookup_data["Robot Model/Controller"]
        ):
            machine_data["Robot Model/Controller"] = lookup_data["Robot Model/Controller"]
            robot_model_filled += 1
        if machine_data == relationship.machine_data:
            enriched.append(relationship)
        else:
            enriched.append(
                RequiredRelationship(
                    machine_no=relationship.machine_no,
                    part_number=relationship.part_number,
                    part_description=relationship.part_description,
                    source_row=relationship.source_row,
                    machine_data=machine_data,
                )
            )
    log_activity_event(
        project_root,
        "off_machine_compatibility_machine_metadata_enriched",
        {
            "compatible_machines_found": len(relationships),
            "robot_type_filled": robot_type_filled,
            "robot_model_controller_filled": robot_model_filled,
            "warnings": warnings[:10],
            "warning_count": len(warnings),
        },
    )
    return enriched, warnings


def _machine_lookup_data_for_compatibility(project_root: str | Path, machine_no: str) -> tuple[dict[str, str], list[str]]:
    try:
        result = lookup_machine(project_root, machine_no)
    except Exception as exc:
        return {}, [f"Machine {machine_no} robot metadata lookup failed: {exc}"]
    data: dict[str, str] = {}
    robot_type = text_value(result.robot_type_suggestion)
    robot_model = text_value(result.robot_model_controller_suggestion)
    if robot_type:
        data["Robot Type"] = robot_type
    if robot_model:
        data["Robot Model/Controller"] = robot_model
    warnings = [
        f"Machine {machine_no} robot metadata lookup warning: {warning}"
        for warning in result.warnings
        if warning and not _is_capacity_lookup_noise(warning)
    ]
    return data, warnings


def _is_capacity_lookup_noise(warning: str) -> bool:
    text = warning.casefold()
    return "capacity" in text or "plant 4" in text


def build_off_machine_compatibility_preview(
    project_root: str | Path,
    audit_id: str,
    submitted_entry: dict[str, Any] | None = None,
    press_capacity_path: str | Path | None = None,
) -> OffMachineCompatibilityPreview:
    started = time.perf_counter()
    submitted_entry = submitted_entry or {}
    clean_audit_id = text_value(audit_id) or text_value(submitted_entry.get("Audit ID"))
    audit_row = _load_inventory_row(project_root, clean_audit_id) if clean_audit_id else None
    source_row = audit_row or dict(submitted_entry)
    tool_number = part_number_from_row(source_row)
    log_activity_event(
        project_root,
        "off_machine_compatibility_lookup_started",
        {
            "audit_id": clean_audit_id,
            "tool_number": tool_number,
            "machine_missing": audit_row_has_missing_machine(source_row),
        },
    )
    if not clean_audit_id:
        errors = ["Saved audit ID was not available for off-machine compatibility lookup."]
        _log_off_machine_lookup_completed(project_root, clean_audit_id, tool_number, 0, [], errors, started)
        return OffMachineCompatibilityPreview(
            audit_id="",
            tool_number=tool_number,
            normalized_tool_number=normalize_tool_identifier(tool_number),
            errors=errors,
            audit_row=source_row,
        )
    if not off_machine_compatibility_lookup_needed(source_row):
        warnings = ["Audit is not an off-machine audit with a Tool #; compatibility lookup was skipped."]
        _log_off_machine_lookup_completed(project_root, clean_audit_id, tool_number, 0, warnings, [], started)
        return OffMachineCompatibilityPreview(
            audit_id=clean_audit_id,
            tool_number=tool_number,
            normalized_tool_number=normalize_tool_identifier(tool_number),
            warnings=warnings,
            audit_row=source_row,
        )
    lookup = find_press_capacity_matches_for_tool(project_root, tool_number, press_capacity_path)
    matches, machine_lookup_warnings = _enrich_relationships_with_machine_lookup(project_root, lookup.matches)
    errors = list(lookup.errors)
    warnings = [*lookup.warnings, *machine_lookup_warnings]
    _log_off_machine_lookup_completed(
        project_root,
        clean_audit_id,
        tool_number,
        len(matches),
        warnings,
        errors,
        started,
        press_capacity_path=lookup.press_capacity_path,
    )
    return OffMachineCompatibilityPreview(
        audit_id=clean_audit_id,
        tool_number=tool_number,
        normalized_tool_number=lookup.normalized_tool_number,
        matches=matches,
        warnings=warnings,
        errors=errors,
        press_capacity_path=lookup.press_capacity_path,
        audit_row=source_row,
    )


def apply_off_machine_compatibility_choice(
    project_root: str | Path,
    audit_id: str,
    choice: str,
    press_capacity_path: str | Path | None = None,
) -> ToolResult:
    started = time.perf_counter()
    clean_choice = str(choice or "").strip()
    if clean_choice not in OFF_MACHINE_COMPATIBILITY_CHOICES:
        return ToolResult.fail(
            "off_machine_compatibility",
            "Off-Machine Compatibility",
            "Off-machine compatibility choice was not recognized.",
            errors=[f"Unknown choice: {choice}"],
            duration_seconds=time.perf_counter() - started,
        )
    clean_audit_id = text_value(audit_id)
    log_activity_event(
        project_root,
        "off_machine_compatibility_user_choice",
        {"audit_id": clean_audit_id, "choice": clean_choice},
    )
    if clean_choice == OFF_MACHINE_COMPATIBILITY_LEAVE:
        result = ToolResult.ok(
            "off_machine_compatibility",
            "Off-Machine Compatibility",
            "Left the audit as an off-machine audit. No compatibility rows were created.",
            details=["User chose to leave the saved audit unchanged."],
            metrics={
                "choice": clean_choice,
                "compatible_machines_found": 0,
                "current_row_updated": False,
                "created": 0,
                "existing_rows_updated": 0,
                "duplicates_skipped": 0,
            },
            duration_seconds=time.perf_counter() - started,
        )
        log_activity_event(
            project_root,
            "off_machine_compatibility_applied",
            {
                "audit_id": clean_audit_id,
                "choice": clean_choice,
                "current_row_updated": False,
                "created": 0,
                "existing_rows_updated": 0,
                "duplicates_skipped": 0,
            },
        )
        return result

    preview = build_off_machine_compatibility_preview(project_root, clean_audit_id, press_capacity_path=press_capacity_path)
    if preview.errors:
        warning = "Compatibility lookup could not be completed. The saved audit was left unchanged."
        log_activity_event(
            project_root,
            "off_machine_compatibility_warning",
            {"audit_id": clean_audit_id, "errors": preview.errors, "warnings": preview.warnings},
        )
        return ToolResult.ok(
            "off_machine_compatibility",
            "Off-Machine Compatibility",
            warning,
            details=preview.errors,
            warnings=[warning, *preview.errors, *preview.warnings],
            metrics={
                "choice": clean_choice,
                "press_capacity_lookup_failed": True,
                "compatible_machines_found": 0,
                "current_row_updated": False,
                "created": 0,
                "existing_rows_updated": 0,
                "duplicates_skipped": 0,
            },
            duration_seconds=time.perf_counter() - started,
        )
    if not preview.matches:
        message = (
            f"No compatible machines were found for Tool # {preview.tool_number} "
            "in the current Press Capacity list."
        )
        return ToolResult.ok(
            "off_machine_compatibility",
            "Off-Machine Compatibility",
            message,
            details=[message],
            warnings=preview.warnings,
            metrics={
                "choice": clean_choice,
                "compatible_machines_found": 0,
                "current_row_updated": False,
                "created": 0,
                "existing_rows_updated": 0,
                "duplicates_skipped": 0,
            },
            duration_seconds=time.perf_counter() - started,
        )

    paths = resolve_project_paths(project_root)
    master_path = paths.master_workbook
    workbook = None
    try:
        workbook = load_workbook(master_path)
        if "EOAT Inventory" not in workbook.sheetnames:
            raise ValueError("EOAT Inventory sheet is missing.")
        ws = workbook["EOAT Inventory"]
        added_headers = _ensure_inventory_headers(ws, get_expected_headers("EOAT Inventory"))
        if ELECTRICAL_WIRING_PRESENT_FIELD in added_headers:
            _migrate_electrical_wiring_presence_rows(ws)
        headers = worksheet_headers(ws)
        source_row_number = _find_audit_row_number(ws, headers, clean_audit_id)
        if source_row_number is None:
            raise ValueError(f"Saved audit row was not found: {clean_audit_id}")
        source_row = _worksheet_row_dict(ws, headers, source_row_number)
        source_tool = display_text_value(source_row.get(TOOL_FIELD)) or preview.tool_number
        source_eoat_id = text_value(source_row.get(EOAT_ASSEMBLY_ID_FIELD))
        existing_ids = _existing_audit_ids(ws, headers)

        current_row_updated = False
        current_row_fields: list[str] = []
        created = 0
        existing_rows_updated = 0
        duplicates_skipped = 0
        duplicate_details: list[str] = []
        create_matches = list(preview.matches)

        if clean_choice == OFF_MACHINE_COMPATIBILITY_UPDATE_AND_ADD:
            current_match = preview.matches[0]
            duplicate_rows = _find_equivalent_compatibility_rows(
                ws,
                headers,
                source_tool,
                current_match.machine_no,
                source_eoat_id,
                exclude_rows={source_row_number},
            )
            if duplicate_rows:
                duplicates_skipped += 1
                duplicate_details.append(
                    f"Skipped current-row update for Machine {current_match.machine_no}; equivalent row already exists."
                )
                existing_rows_updated += _update_missing_machine_fields_for_rows(
                    ws, headers, duplicate_rows, current_match, source_tool
                )
            else:
                current_row_fields = _fill_missing_machine_fields(ws, headers, source_row_number, current_match, source_tool)
                current_row_updated = bool(current_row_fields)
            create_matches = list(preview.matches[1:])

        rows_to_create: list[RequiredRelationship] = []
        for match in create_matches:
            duplicate_rows = _find_equivalent_compatibility_rows(
                ws,
                headers,
                source_tool,
                match.machine_no,
                source_eoat_id,
                exclude_rows=set(),
            )
            if duplicate_rows:
                duplicates_skipped += 1
                duplicate_details.append(f"Skipped Machine {match.machine_no}; equivalent row already exists.")
                existing_rows_updated += _update_missing_machine_fields_for_rows(
                    ws, headers, duplicate_rows, match, source_tool
                )
                continue
            rows_to_create.append(match)

        backup = None
        if current_row_updated or existing_rows_updated or rows_to_create:
            backup = backup_file(master_path, master_path.parent / "_backups")
        new_ids = _generate_audit_ids(existing_ids, len(rows_to_create))
        for new_audit_id, match in zip(new_ids, rows_to_create):
            new_row = _build_off_machine_compatible_row(project_root, source_row, match, new_audit_id, clean_audit_id, source_tool)
            write_row_by_headers(ws, next_empty_row(ws), new_row)
            created += 1
        if current_row_updated or existing_rows_updated or created:
            refresh_audit_by_press_view(workbook)
            workbook.save(master_path)
            invalidate_workbook_cache(master_path)
        files_modified = [str(master_path)] if current_row_updated or existing_rows_updated or created else []
        files_created = [str(backup)] if backup else []
    except Exception as exc:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        log_activity_event(
            project_root,
            "off_machine_compatibility_error",
            {"audit_id": clean_audit_id, "choice": clean_choice, "error": str(exc)},
        )
        return ToolResult.fail(
            "off_machine_compatibility",
            "Off-Machine Compatibility",
            "Could not apply off-machine compatibility updates.",
            errors=[str(exc)],
            warnings=preview.warnings,
            duration_seconds=time.perf_counter() - started,
        )
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass

    summary = _off_machine_compatibility_summary(clean_choice, current_row_updated, created, existing_rows_updated, duplicates_skipped)
    details = [
        f"Tool #: {preview.tool_number}",
        f"Compatible machines found: {len(preview.matches)}",
        f"Current row updated: {'Yes' if current_row_updated else 'No'}",
        f"Current row fields filled: {', '.join(current_row_fields) if current_row_fields else '(none)'}",
        f"Compatibility rows created: {created}",
        f"Existing equivalent rows updated: {existing_rows_updated}",
        f"Duplicates skipped: {duplicates_skipped}",
        *duplicate_details,
    ]
    if backup:
        details.append(f"Workbook backup: {backup}")
    metrics = {
        "choice": clean_choice,
        "compatible_machines_found": len(preview.matches),
        "current_row_updated": current_row_updated,
        "current_row_fields_filled": len(current_row_fields),
        "created": created,
        "existing_rows_updated": existing_rows_updated,
        "duplicates_skipped": duplicates_skipped,
    }
    log_activity_event(
        project_root,
        "off_machine_compatibility_applied",
        {
            "audit_id": clean_audit_id,
            "choice": clean_choice,
            "current_row_updated": current_row_updated,
            "created": created,
            "existing_rows_updated": existing_rows_updated,
            "duplicates_skipped": duplicates_skipped,
            "compatible_machines_found": len(preview.matches),
        },
    )
    return ToolResult.ok(
        "off_machine_compatibility",
        "Off-Machine Compatibility",
        summary,
        details=details,
        warnings=preview.warnings,
        files_created=files_created,
        files_modified=files_modified,
        metrics=metrics,
        duration_seconds=time.perf_counter() - started,
    )


def load_required_relationships(press_capacity_path: str | Path) -> tuple[list[RequiredRelationship], list[str]]:
    path = Path(press_capacity_path)
    if not path.exists():
        return [], [f"Press Capacity reference file not found: {path}"]
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_name = _find_sheet_name(workbook.sheetnames, CAPACITY_SHEET_NAME, ["capacity"])
        if sheet_name is None:
            return [], [f"Press Capacity sheet not found: {CAPACITY_SHEET_NAME}"]
        ws = workbook[sheet_name]
        header_map: dict[str, int] | None = None
        relationships: dict[tuple[str, str], RequiredRelationship] = {}
        for row_number, values in enumerate(ws.iter_rows(values_only=True), start=1):
            values = list(values)
            maybe_header = _capacity_header_map(values)
            if maybe_header:
                header_map = maybe_header
                continue
            if header_map is None:
                continue
            part_number = display_text_value(_value_for(values, header_map, "NGW Part Number"))
            if not normalize_tool_identifier(part_number):
                continue
            description = text_value(_value_for(values, header_map, "NGW Part Description"))
            for machine_no in parse_machine_tokens(_value_for(values, header_map, "Machine No.")):
                key = relationship_key(machine_no, part_number)
                machine_data = _capacity_machine_data(values, header_map, key[0], part_number, description)
                relationships.setdefault(
                    key,
                    RequiredRelationship(
                        machine_no=key[0],
                        part_number=part_number,
                        part_description=description,
                        source_row=row_number,
                        machine_data=machine_data,
                    ),
                )
        return sorted(
            relationships.values(), key=lambda item: (_machine_sort_key(item.machine_no), item.part_number.upper())
        ), []
    except Exception as exc:
        return [], [f"Press Capacity workbook could not be read: {exc}"]
    finally:
        if workbook is not None:
            workbook.close()


def summarize_master_relationships(rows: list[dict[str, Any]]) -> dict[tuple[str, str], list[dict[str, Any]]]:
    by_key: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for index, row in enumerate(rows, start=2):
        part_number = part_number_from_row(row)
        if not part_number:
            continue
        for machine_no in parse_machine_tokens(row.get("Press/Machine #")):
            key = relationship_key(machine_no, part_number)
            by_key[key].append(
                {"row_number": index, "row": row, "entry_type": normalize_entry_type(row.get(ENTRY_TYPE_FIELD))}
            )
    return dict(by_key)


def relationship_has_conflict(
    existing_rows: list[dict[str, Any]], required: RequiredRelationship | None = None
) -> bool:
    if not existing_rows:
        return False
    if required and required.part_description:
        descriptions = {
            part_description_from_row(item["row"]) for item in existing_rows if part_description_from_row(item["row"])
        }
        if descriptions and required.part_description not in descriptions:
            return True
    for field_name in ["EOAT Type"]:
        values = {
            text_value(item["row"].get(field_name)).lower()
            for item in existing_rows
            if text_value(item["row"].get(field_name))
        }
        if len(values) > 1:
            return True
    descriptions = {
        part_description_from_row(item["row"]).lower()
        for item in existing_rows
        if part_description_from_row(item["row"])
    }
    return len(descriptions) > 1


def _can_sync_compatibility_field(header: str) -> bool:
    normalized = header.strip().lower()
    if not normalized:
        return False
    if header in PROTECTED_COMPATIBILITY_SYNC_FIELDS:
        return False
    return not any(part in normalized for part in PROTECTED_COMPATIBILITY_SYNC_NAME_PARTS)


def _find_audit_row_number(ws, headers: list[str], audit_id: str) -> int | None:
    if "Audit ID" not in headers:
        return None
    audit_id_col = headers.index("Audit ID") + 1
    for row_number in range(2, ws.max_row + 1):
        if text_value(ws.cell(row=row_number, column=audit_id_col).value) == str(audit_id):
            return row_number
    return None


def _worksheet_row_dict(ws, headers: list[str], row_number: int) -> dict[str, Any]:
    return {header: ws.cell(row=row_number, column=index).value for index, header in enumerate(headers, start=1)}


def _master_path(project_root_or_master_path: str | Path) -> Path:
    path = Path(project_root_or_master_path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return path
    return resolve_project_paths(path).master_workbook


def _generate_audit_ids(existing_ids: set[str], count: int) -> list[str]:
    compact = date.today().isoformat().replace("-", "")
    prefix = f"AUD-{compact}-"
    max_number = 0
    for value in existing_ids:
        if value.startswith(prefix):
            try:
                max_number = max(max_number, int(value.rsplit("-", 1)[1]))
            except ValueError:
                continue
    audit_ids = []
    for offset in range(1, count + 1):
        audit_id = f"{prefix}{max_number + offset:03d}"
        while audit_id in existing_ids:
            max_number += 1
            audit_id = f"{prefix}{max_number + offset:03d}"
        audit_ids.append(audit_id)
        existing_ids.add(audit_id)
    return audit_ids


def _creation_summary(created: int, skipped_audited: int, skipped_compatible: int, conflicts: int) -> list[str]:
    return [
        f"Created {created} compatibility entries.",
        f"Skipped {skipped_audited} already-audited relationships.",
        f"Skipped {skipped_compatible} already-compatible relationships.",
        f"{conflicts} conflicts need review.",
    ]


def _machine_sort_key(machine_no: str) -> tuple[int, int | str]:
    return (0, int(machine_no)) if str(machine_no).isdigit() else (1, str(machine_no).lower())


def _find_sheet_name(sheet_names: list[str], preferred: str, words: list[str]) -> str | None:
    if preferred in sheet_names:
        return preferred
    preferred_norm = _norm(preferred)
    for sheet_name in sheet_names:
        if _norm(sheet_name) == preferred_norm:
            return sheet_name
    for sheet_name in sheet_names:
        normalized = _norm(sheet_name)
        if all(word in normalized for word in words):
            return sheet_name
    return None


def _capacity_header_map(values: list[Any]) -> dict[str, int] | None:
    normalized_values = [_norm(value) for value in values]
    machine_headers = {_norm(alias) for alias in CAPACITY_FIELD_ALIASES["Machine No."]}
    if not any(value in machine_headers for value in normalized_values):
        return None
    mapping: dict[str, int] = {}
    for logical, names in CAPACITY_FIELD_ALIASES.items():
        normalized_names = {_norm(name) for name in names}
        for index, normalized in enumerate(normalized_values):
            if normalized in normalized_names:
                mapping[logical] = index
                break
    expected_headers = {_norm(header): header for header in get_expected_headers("EOAT Inventory")}
    for index, normalized in enumerate(normalized_values):
        expected_header = expected_headers.get(normalized)
        if expected_header and expected_header not in mapping:
            mapping[expected_header] = index
    required = {"Machine No.", "NGW Part Number"}
    return mapping if required.issubset(mapping) else None


def _value_for(values: list[Any], header_map: dict[str, int], field_name: str) -> Any:
    index = header_map.get(field_name)
    if index is None or index >= len(values):
        return ""
    return values[index]


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", text_value(value).lower())


def _capacity_machine_data(
    values: list[Any],
    header_map: dict[str, int],
    machine_no: str,
    part_number: str,
    description: str,
) -> dict[str, Any]:
    data: dict[str, Any] = {
        "Press/Machine #": machine_no,
        "Machine No.": machine_no,
        TOOL_FIELD: part_number,
        "NGW Part Number": part_number,
    }
    if description:
        data["NGW Part Description"] = description
        data["Part Name/Description"] = description
    for field_name in MACHINE_DERIVED_AUDIT_FIELDS:
        if field_name == "Press/Machine #":
            continue
        value = text_value(_value_for(values, header_map, field_name))
        if value:
            data[field_name] = value
    return data


def _is_capacity_load_warning(warning: str) -> bool:
    text = warning.casefold()
    return "not found" in text or "could not be read" in text or "sheet not found" in text


def _load_inventory_row(project_root: str | Path, audit_id: str) -> dict[str, Any] | None:
    if not audit_id:
        return None
    try:
        for row in row_dicts(_master_path(project_root), "EOAT Inventory"):
            if text_value(row.get("Audit ID")) == audit_id:
                return dict(row)
    except Exception:
        return None
    return None


def _log_off_machine_lookup_completed(
    project_root: str | Path,
    audit_id: str,
    tool_number: str,
    match_count: int,
    warnings: list[str],
    errors: list[str],
    started: float,
    *,
    press_capacity_path: str = "",
) -> None:
    log_activity_event(
        project_root,
        "off_machine_compatibility_lookup_completed",
        {
            "audit_id": audit_id,
            "tool_number": tool_number,
            "compatible_machines_found": match_count,
            "warnings": warnings,
            "errors": errors,
            "press_capacity_path": press_capacity_path,
            "duration_seconds": round(time.perf_counter() - started, 3),
        },
    )


def _existing_audit_ids(ws, headers: list[str]) -> set[str]:
    if "Audit ID" not in headers:
        return set()
    audit_id_col = headers.index("Audit ID") + 1
    return {
        text_value(ws.cell(row=row_number, column=audit_id_col).value)
        for row_number in range(2, ws.max_row + 1)
        if text_value(ws.cell(row=row_number, column=audit_id_col).value)
    }


def _find_equivalent_compatibility_rows(
    ws,
    headers: list[str],
    tool_number: str,
    machine_no: str,
    eoat_assembly_id: str,
    *,
    exclude_rows: set[int],
) -> list[int]:
    tool_key = tool_identifier_key(tool_number)
    machine_key = normalize_machine_token(machine_no)
    matching_rows: list[int] = []
    for row_number in range(2, ws.max_row + 1):
        if row_number in exclude_rows:
            continue
        row = _worksheet_row_dict(ws, headers, row_number)
        if tool_identifier_key(part_number_from_row(row)) != tool_key:
            continue
        if machine_key not in _audit_row_machine_tokens(row):
            continue
        if eoat_assembly_id and EOAT_ASSEMBLY_ID_FIELD in headers:
            row_eoat_id = text_value(row.get(EOAT_ASSEMBLY_ID_FIELD))
            if row_eoat_id != eoat_assembly_id:
                continue
        matching_rows.append(row_number)
    return matching_rows


def _update_missing_machine_fields_for_rows(
    ws,
    headers: list[str],
    row_numbers: list[int],
    match: RequiredRelationship,
    source_tool: str,
) -> int:
    updated = 0
    for row_number in row_numbers:
        if _fill_missing_machine_fields(ws, headers, row_number, match, source_tool):
            updated += 1
    return updated


def _fill_missing_machine_fields(
    ws,
    headers: list[str],
    row_number: int,
    match: RequiredRelationship,
    source_tool: str,
) -> list[str]:
    changed: list[str] = []
    for field_name in MACHINE_DERIVED_AUDIT_FIELDS:
        if field_name not in headers:
            continue
        value = _machine_field_value(match, field_name, source_tool)
        if not text_value(value):
            continue
        column = headers.index(field_name) + 1
        existing = ws.cell(row=row_number, column=column).value
        if _field_value_is_missing(existing):
            ws.cell(row=row_number, column=column).value = value
            changed.append(field_name)
    entry_type = ""
    if ENTRY_TYPE_FIELD in headers:
        entry_type = normalize_entry_type(ws.cell(row=row_number, column=headers.index(ENTRY_TYPE_FIELD) + 1).value)
    if AUDIT_CONTEXT_FIELD in headers:
        context_col = headers.index(AUDIT_CONTEXT_FIELD) + 1
        context_value = text_value(ws.cell(row=row_number, column=context_col).value)
        desired_context = AUDIT_CONTEXT_COMPATIBILITY if entry_type == ENTRY_TYPE_COMPATIBLE else AUDIT_CONTEXT_INSTALLED
        if _field_value_is_missing(context_value) or context_value.casefold() == "not installed / bench audit":
            ws.cell(row=row_number, column=context_col).value = desired_context
            changed.append(AUDIT_CONTEXT_FIELD)
    if PHYSICAL_AUDIT_VERIFIED_FIELD in headers:
        verified_col = headers.index(PHYSICAL_AUDIT_VERIFIED_FIELD) + 1
        desired_verified = "No" if entry_type == ENTRY_TYPE_COMPATIBLE else "Yes"
        if _field_value_is_missing(ws.cell(row=row_number, column=verified_col).value):
            ws.cell(row=row_number, column=verified_col).value = desired_verified
            changed.append(PHYSICAL_AUDIT_VERIFIED_FIELD)
    return changed


def _field_value_is_missing(value: Any) -> bool:
    text = display_text_value(value).strip()
    return not text or text.casefold() in MISSING_MACHINE_VALUES


def _machine_field_value(match: RequiredRelationship, field_name: str, source_tool: str) -> str:
    if field_name == "Press/Machine #":
        return match.machine_no
    if field_name == TOOL_FIELD:
        return source_tool
    value = text_value(match.machine_data.get(field_name))
    if value:
        return value
    if field_name == "Part Name/Description":
        return match.part_description
    return ""


def _build_off_machine_compatible_row(
    project_root: str | Path,
    source_row: dict[str, Any],
    match: RequiredRelationship,
    new_audit_id: str,
    source_audit_id: str,
    source_tool: str,
) -> dict[str, Any]:
    new_row = dict(source_row)
    new_row["Audit ID"] = new_audit_id
    new_row["Audit Date"] = ""
    new_row["Auditor"] = ""
    new_row[TOOL_FIELD] = source_tool
    for field_name in MACHINE_DERIVED_AUDIT_FIELDS:
        value = _machine_field_value(match, field_name, source_tool)
        if text_value(value):
            new_row[field_name] = value
    if text_value(new_row.get("Status")).lower() == "audited":
        new_row["Status"] = "Complete"
    new_row[ENTRY_TYPE_FIELD] = ENTRY_TYPE_COMPATIBLE
    new_row[AUDIT_CONTEXT_FIELD] = AUDIT_CONTEXT_COMPATIBILITY
    new_row[SOURCE_AUDIT_ID_FIELD] = source_audit_id
    new_row[COMPATIBILITY_SOURCE_FIELD] = COMPATIBILITY_SOURCE_PRESS_CAPACITY
    new_row[PHYSICAL_AUDIT_VERIFIED_FIELD] = "No"
    new_row[COMPATIBILITY_CONFIDENCE_FIELD] = "Press Capacity"
    return normalize_audit_entry(project_root, new_row)


def _off_machine_compatibility_summary(
    choice: str,
    current_row_updated: bool,
    created: int,
    existing_rows_updated: int,
    duplicates_skipped: int,
) -> str:
    if choice == OFF_MACHINE_COMPATIBILITY_UPDATE_AND_ADD:
        current_text = "updated the saved audit row" if current_row_updated else "left the saved audit row unchanged"
        return (
            f"Off-machine compatibility applied: {current_text}, created {created} compatibility row(s), "
            f"updated {existing_rows_updated} existing row(s), skipped {duplicates_skipped} duplicate(s)."
        )
    return (
        f"Off-machine compatibility applied: created {created} compatibility row(s), "
        f"updated {existing_rows_updated} existing row(s), skipped {duplicates_skipped} duplicate(s)."
    )
