from __future__ import annotations

from pathlib import Path
from copy import copy

from openpyxl import load_workbook

from .tool_fields import LEGACY_TOOL_FIELD, TOOL_FIELD

LEGACY_VACUUM_CUPS_FIELD = "Number of Vacuum Cups"
NUMBER_OF_PARTS_PICKED_FIELD = "Number of Parts Picked"


def workbook_sheet_names(workbook_path: str | Path) -> list[str]:
    workbook = load_workbook(Path(workbook_path), read_only=True, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def worksheet_headers(ws) -> list[str]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [str(value) if value is not None else "" for value in row]


def row_dicts(workbook_path: str | Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(Path(workbook_path), read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        ws = workbook[sheet_name]
        headers = worksheet_headers(ws)
        rows: list[dict[str, object]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            non_empty = [value for value in row if value not in (None, "")]
            if not non_empty:
                continue
            if len(non_empty) == 1 and str(non_empty[0]).startswith("Last Updated:"):
                continue
            row_data = {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            if TOOL_FIELD not in row_data and LEGACY_TOOL_FIELD in row_data:
                row_data[TOOL_FIELD] = row_data.get(LEGACY_TOOL_FIELD)
            if NUMBER_OF_PARTS_PICKED_FIELD not in row_data and LEGACY_VACUUM_CUPS_FIELD in row_data:
                row_data[NUMBER_OF_PARTS_PICKED_FIELD] = row_data.get(LEGACY_VACUUM_CUPS_FIELD)
            rows.append(row_data)
        return rows
    finally:
        workbook.close()


def find_row_by_value(ws, header_name: str, value: str) -> int | None:
    headers = worksheet_headers(ws)
    if header_name not in headers:
        return None
    column = headers.index(header_name) + 1
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=column).value or "") == str(value):
            return row
    return None


def next_empty_row(ws) -> int:
    for row in range(2, ws.max_row + 2):
        values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        if not any(value not in (None, "") for value in values):
            return row
    return ws.max_row + 1


def write_row_by_headers(ws, row_number: int, data: dict[str, object]) -> list[str]:
    headers = worksheet_headers(ws)
    written: list[str] = []
    template_row = max(2, row_number - 1)
    for col, header in enumerate(headers, start=1):
        if header not in data:
            continue
        cell = ws.cell(row=row_number, column=col)
        template_cell = ws.cell(row=template_row, column=col)
        if row_number > ws.max_row or cell.value is None:
            if template_cell.has_style:
                cell._style = copy(template_cell._style)
            if template_cell.number_format:
                cell.number_format = template_cell.number_format
            if template_cell.alignment:
                cell.alignment = copy(template_cell.alignment)
            if template_cell.fill:
                cell.fill = copy(template_cell.fill)
            if template_cell.font:
                cell.font = copy(template_cell.font)
        cell.value = data.get(header)
        written.append(header)
    return written
