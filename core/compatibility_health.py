from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .audit.relationships import is_compatibility_row, source_audit_for_compatibility_row
from .audit_compatibility import (
    _can_sync_compatibility_field,
    load_required_relationships,
    normalize_entry_type,
    part_number_from_row,
    relationship_key,
    text_value,
)
from .audit_constants import (
    COMPATIBILITY_SOURCE_FIELD,
    ENTRY_TYPE_AUDITED,
    ENTRY_TYPE_COMPATIBLE,
    ENTRY_TYPE_FIELD,
    SOURCE_AUDIT_ID_FIELD,
)
from .paths import get_press_capacity_file, resolve_project_paths
from .validation_findings import ValidationFinding, ValidationSeverity, make_finding
from .workbook_io import worksheet_headers

SOURCE_VALIDATOR = "compatibility_health"
SHEET_NAME = "EOAT Inventory"


def validate_compatibility_health(project_root_or_master_path: str | Path) -> list[ValidationFinding]:
    master_path = _master_path(project_root_or_master_path)
    if not master_path.exists():
        return []
    workbook = None
    try:
        workbook = load_workbook(master_path, read_only=True, data_only=True)
        if SHEET_NAME not in workbook.sheetnames:
            return []
        ws = workbook[SHEET_NAME]
        headers = worksheet_headers(ws)
        rows = _inventory_rows(ws, headers)
    finally:
        if workbook is not None:
            workbook.close()

    if not rows:
        return []

    row_values = [row for _row_number, row in rows]
    findings: list[ValidationFinding] = []

    for row_number, row in rows:
        audit_id = text_value(row.get("Audit ID"))
        machine = text_value(row.get("Press/Machine #"))
        entry_type = normalize_entry_type(row.get(ENTRY_TYPE_FIELD))
        source_id = text_value(row.get(SOURCE_AUDIT_ID_FIELD))
        compatibility_source = text_value(row.get(COMPATIBILITY_SOURCE_FIELD))

        if is_compatibility_row(row):
            source_lookup = source_audit_for_compatibility_row(row_values, row)
            if "missing_source_metadata" in source_lookup.warning_codes:
                findings.append(
                    make_finding(
                        ValidationSeverity.ERROR,
                        "compatibility",
                        f"Compatible row {audit_id or f'row {row_number}'} is missing Source Audit ID.",
                        sheet_name=SHEET_NAME,
                        row_number=row_number,
                        column_name=SOURCE_AUDIT_ID_FIELD,
                        audit_id=audit_id,
                        machine_number=machine,
                        expected_behavior="Compatible rows must link back to a physical source audit.",
                        recommended_action="Review the compatible row and select the source audit before relying on inherited values.",
                        source_validator=SOURCE_VALIDATOR,
                    )
                )
            elif "source_not_found" in source_lookup.warning_codes:
                findings.append(
                    make_finding(
                        ValidationSeverity.ERROR,
                        "compatibility",
                        f"Compatible row {audit_id or f'row {row_number}'} references missing source audit {source_id}.",
                        sheet_name=SHEET_NAME,
                        row_number=row_number,
                        column_name=SOURCE_AUDIT_ID_FIELD,
                        audit_id=audit_id,
                        machine_number=machine,
                        current_value=source_id,
                        expected_behavior="Source Audit ID must match an existing audited row.",
                        recommended_action="Relink the compatible row or recreate it from a valid source audit.",
                        source_validator=SOURCE_VALIDATOR,
                    )
                )
            elif "source_not_physical" in source_lookup.warning_codes:
                findings.append(
                    make_finding(
                        ValidationSeverity.ERROR,
                        "compatibility",
                        f"Compatible row {audit_id or f'row {row_number}'} references {source_id}, but that row is not a physical audit source.",
                        sheet_name=SHEET_NAME,
                        row_number=row_number,
                        column_name=SOURCE_AUDIT_ID_FIELD,
                        audit_id=audit_id,
                        machine_number=machine,
                        current_value=source_id,
                        expected_behavior="Source Audit ID must point to an audited row.",
                        recommended_action="Relink the compatible row to a physical audit source.",
                        source_validator=SOURCE_VALIDATOR,
                    )
                )
            elif source_lookup.source_audit is not None:
                _append_stale_inherited_value_findings(findings, row_number, row, source_lookup.source_audit)
            if not compatibility_source:
                findings.append(
                    make_finding(
                        ValidationSeverity.WARNING,
                        "compatibility",
                        f"Compatible row {audit_id or f'row {row_number}'} is missing Compatibility Source metadata.",
                        sheet_name=SHEET_NAME,
                        row_number=row_number,
                        column_name=COMPATIBILITY_SOURCE_FIELD,
                        audit_id=audit_id,
                        machine_number=machine,
                        expected_behavior="Compatibility Source should describe how the row was generated.",
                        recommended_action="Regenerate or review the compatible row metadata.",
                        source_validator=SOURCE_VALIDATOR,
                    )
                )
        elif source_id or compatibility_source:
            findings.append(
                make_finding(
                    ValidationSeverity.WARNING,
                    "compatibility",
                    f"Physical audit row {audit_id or f'row {row_number}'} contains system-managed compatibility metadata.",
                    sheet_name=SHEET_NAME,
                    row_number=row_number,
                    column_name=SOURCE_AUDIT_ID_FIELD if source_id else COMPATIBILITY_SOURCE_FIELD,
                    audit_id=audit_id,
                    machine_number=machine,
                    current_value=source_id or compatibility_source,
                    expected_behavior="Physical audit rows should not carry Source Audit ID or Compatibility Source values.",
                    recommended_action="Clear the compatibility metadata or convert the row back to a compatible row if appropriate.",
                    source_validator=SOURCE_VALIDATOR,
                )
            )

    findings.extend(_required_relationship_findings(project_root_or_master_path, rows))
    return findings


def _append_stale_inherited_value_findings(
    findings: list[ValidationFinding],
    row_number: int,
    compatible_row: dict[str, Any],
    source_row: dict[str, Any],
) -> None:
    audit_id = text_value(compatible_row.get("Audit ID"))
    machine = text_value(compatible_row.get("Press/Machine #"))
    mismatched_fields: list[str] = []
    for field_name, source_value in source_row.items():
        if not _can_sync_compatibility_field(field_name):
            continue
        if text_value(compatible_row.get(field_name)) != text_value(source_value):
            mismatched_fields.append(field_name)
    if not mismatched_fields:
        return
    for field_name in mismatched_fields[:5]:
        findings.append(
            make_finding(
                ValidationSeverity.WARNING,
                "compatibility",
                f"Compatible row {audit_id or f'row {row_number}'} has a stale inherited value for {field_name}.",
                sheet_name=SHEET_NAME,
                row_number=row_number,
                column_name=field_name,
                audit_id=audit_id,
                machine_number=machine,
                current_value=compatible_row.get(field_name),
                expected_behavior=f"Expected to match source audit value: {text_value(source_row.get(field_name)) or 'blank'}.",
                recommended_action="Review the linked source audit and rerun the existing compatibility sync if the inherited value should match.",
                source_validator=SOURCE_VALIDATOR,
            )
        )
    if len(mismatched_fields) > 5:
        findings.append(
            make_finding(
                ValidationSeverity.WARNING,
                "compatibility",
                f"Compatible row {audit_id or f'row {row_number}'} has {len(mismatched_fields)} stale inherited values.",
                sheet_name=SHEET_NAME,
                row_number=row_number,
                audit_id=audit_id,
                machine_number=machine,
                expected_behavior="Compatible inherited fields should match the linked source audit unless intentionally reviewed.",
                recommended_action="Review the compatible row before using it for decisions.",
                source_validator=SOURCE_VALIDATOR,
            )
        )


def _required_relationship_findings(
    project_root_or_master_path: str | Path,
    rows: list[tuple[int, dict[str, Any]]],
) -> list[ValidationFinding]:
    project_root = _project_root(project_root_or_master_path)
    if project_root is None:
        return []
    capacity_path = get_press_capacity_file(project_root)
    if not capacity_path.exists():
        return []
    required, warnings = load_required_relationships(capacity_path)
    if warnings or not required:
        return []

    required_keys = {relationship.key for relationship in required}
    covered_keys: dict[tuple[str, str], tuple[int, dict[str, Any]]] = {}
    source_part_keys = {
        text_value(row.get("Audit ID")): part_number_from_row(row)
        for _row_number, row in rows
        if normalize_entry_type(row.get(ENTRY_TYPE_FIELD)) == ENTRY_TYPE_AUDITED
    }
    findings: list[ValidationFinding] = []
    for row_number, row in rows:
        part_number = part_number_from_row(row)
        if not part_number:
            continue
        key = relationship_key(row.get("Press/Machine #"), part_number)
        if key[0] and key[1]:
            covered_keys[key] = (row_number, row)
            if normalize_entry_type(row.get(ENTRY_TYPE_FIELD)) == ENTRY_TYPE_COMPATIBLE and key not in required_keys:
                findings.append(
                    make_finding(
                        ValidationSeverity.WARNING,
                        "compatibility",
                        f"Compatible row {text_value(row.get('Audit ID')) or f'row {row_number}'} does not match a current physical/compatible capacity relationship.",
                        sheet_name=SHEET_NAME,
                        row_number=row_number,
                        column_name="Press/Machine #",
                        audit_id=text_value(row.get("Audit ID")),
                        machine_number=text_value(row.get("Press/Machine #")),
                        current_value=f"{key[0]} / {key[1]}",
                        expected_behavior="Compatible rows should correspond to a required press/part relationship when capacity data is available.",
                        recommended_action="Review whether the capacity list changed or the compatible row should be dismissed/recreated.",
                        source_validator=SOURCE_VALIDATOR,
                    )
                )

    for relationship in required:
        if relationship.key in covered_keys:
            continue
        matching_sources = [
            source_id
            for source_id, part_number in source_part_keys.items()
            if text_value(part_number).upper() == relationship.key[1]
        ]
        if not matching_sources:
            continue
        findings.append(
            make_finding(
                ValidationSeverity.WARNING,
                "compatibility",
                f"Required compatible relationship is missing for machine {relationship.machine_no} and part {relationship.part_number}.",
                sheet_name=SHEET_NAME,
                column_name="Press/Machine #",
                machine_number=relationship.machine_no,
                current_value="missing compatible row",
                expected_behavior="Every required press/part relationship should be covered by a physical audit or compatible row.",
                recommended_action=f"Use an existing source audit for this part if valid: {', '.join(sorted(matching_sources)[:3])}.",
                source_validator=SOURCE_VALIDATOR,
            )
        )
    return findings


def _inventory_rows(ws, headers: list[str]) -> list[tuple[int, dict[str, Any]]]:
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in values):
            continue
        row = {header: values[index] for index, header in enumerate(headers) if index < len(values)}
        rows.append((row_number, row))
    return rows


def _master_path(project_root_or_master_path: str | Path) -> Path:
    path = Path(project_root_or_master_path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return path
    return resolve_project_paths(path).master_workbook


def _project_root(project_root_or_master_path: str | Path) -> Path | None:
    path = Path(project_root_or_master_path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return None
    return path


__all__ = ["validate_compatibility_health"]
