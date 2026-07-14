from __future__ import annotations

import csv
import filecmp
import json
import shutil
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from .audit.relationships import is_compatibility_row
from .audit_constants import SOURCE_AUDIT_ID_FIELD
from .audit_by_press import refresh_audit_by_press_view
from .eoat_ids import (
    CANONICAL_AREA_CLEANROOM,
    CANONICAL_AREA_PLANT4,
    CANONICAL_AREA_UNKNOWN,
    EOAT_ASSEMBLY_ID_FIELD,
    EOAT_ID_SEARCH_PATTERN,
    EOAT_PREFIX_CLEANROOM,
    EOAT_PREFIX_PLANT4,
    canonical_area,
    determine_eoat_prefix,
    find_eoat_ids,
    format_eoat_id,
    get_eoat_prefix,
    is_valid_eoat_assembly_id,
    normalize_eoat_assembly_id,
    parse_eoat_id,
)
from .paths import resolve_project_paths
from .safe_files import ensure_directory, safe_write_text
from .tool_fields import TOOL_FIELD
from .workbook_cache import invalidate_workbook_cache
from .workbook_io import worksheet_headers

MIGRATION_NAME = "eoat_id_prefix_migration"
FIX_REPAIR_EOAT_ID_PREFIXES = "repair_eoat_id_prefixes"
PHOTO_EXTENSIONS = {
    ".jpg",
    ".jpeg",
    ".png",
    ".gif",
    ".bmp",
    ".tif",
    ".tiff",
    ".heic",
    ".heif",
    ".webp",
}
CACHE_TEXT_EXTENSIONS = {
    ".csv",
    ".htm",
    ".html",
    ".json",
    ".log",
    ".md",
    ".txt",
    ".xml",
    ".yaml",
    ".yml",
}


@dataclass(frozen=True)
class EOATIDMapping:
    old_id: str
    new_id: str
    source_sheet: str
    source_row: int
    audit_id: str = ""
    tool_number: str = ""
    machine_number: str = ""
    area_value: str = ""
    cleanroom_flag_value: str = ""
    reason: str = ""

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class WorkbookUpdate:
    sheet: str
    row: int
    column: str
    old_value: str
    new_value: str
    reason: str
    value_kind: str = "cell"

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class PathUpdate:
    source: str
    target: str
    kind: str
    status: str
    reason: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class ValidationIssue:
    severity: str
    source: str
    message: str
    sheet: str = ""
    row: int | None = None
    column: str = ""
    audit_id: str = ""
    source_audit_id: str = ""
    tool_number: str = ""
    machine_number: str = ""
    area_value: str = ""
    cleanroom_flag_value: str = ""
    current_eoat_id: str = ""
    expected_prefix: str = ""
    actual_prefix: str = ""
    suggested_repair: str = ""

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class EOATPrefixMigrationResult:
    timestamp: str
    dry_run: bool
    workbook_path: str
    photo_root: str = ""
    project_root: str = ""
    backup_dir: str = ""
    migration_report_md: str = ""
    migration_report_csv: str = ""
    validation_report_md: str = ""
    validation_report_json: str = ""
    mappings: list[EOATIDMapping] = field(default_factory=list)
    workbook_updates: list[WorkbookUpdate] = field(default_factory=list)
    photo_updates: list[PathUpdate] = field(default_factory=list)
    cache_updates: list[PathUpdate] = field(default_factory=list)
    backups: list[str] = field(default_factory=list)
    conflicts: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    validation_issues: list[ValidationIssue] = field(default_factory=list)

    @property
    def id_map(self) -> dict[str, str]:
        return {mapping.old_id: mapping.new_id for mapping in self.mappings}

    @property
    def success(self) -> bool:
        return not self.errors

    def to_dict(self) -> dict[str, Any]:
        return {
            "timestamp": self.timestamp,
            "dry_run": self.dry_run,
            "workbook_path": self.workbook_path,
            "photo_root": self.photo_root,
            "project_root": self.project_root,
            "backup_dir": self.backup_dir,
            "migration_report_md": self.migration_report_md,
            "migration_report_csv": self.migration_report_csv,
            "validation_report_md": self.validation_report_md,
            "validation_report_json": self.validation_report_json,
            "mappings": [item.to_dict() for item in self.mappings],
            "workbook_updates": [item.to_dict() for item in self.workbook_updates],
            "photo_updates": [item.to_dict() for item in self.photo_updates],
            "cache_updates": [item.to_dict() for item in self.cache_updates],
            "backups": list(self.backups),
            "conflicts": list(self.conflicts),
            "warnings": list(self.warnings),
            "errors": list(self.errors),
            "validation_issues": [item.to_dict() for item in self.validation_issues],
        }


def run_eoat_id_prefix_migration(
    *,
    workbook_path: str | Path,
    photo_root: str | Path | None = None,
    project_root: str | Path | None = None,
    apply: bool = False,
    report_dir: str | Path | None = None,
    backup_dir: str | Path | None = None,
    rebuild_indexes: bool = True,
    validate_only: bool = False,
    timestamp: str | None = None,
) -> EOATPrefixMigrationResult:
    started = time.perf_counter()
    stamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    workbook = Path(workbook_path)
    project = Path(project_root) if project_root else _project_root_from_workbook(workbook)
    photos = Path(photo_root) if photo_root else (resolve_project_paths(project).cell_photos if project else None)
    reports = ensure_directory(Path(report_dir) if report_dir else _default_report_dir(project, workbook))
    backups_root = Path(backup_dir) if backup_dir else _default_backup_dir(project, workbook, stamp)
    result = EOATPrefixMigrationResult(
        timestamp=stamp,
        dry_run=not apply,
        workbook_path=str(workbook),
        photo_root=str(photos or ""),
        project_root=str(project or ""),
        backup_dir=str(backups_root),
    )

    if not workbook.exists():
        result.errors.append(f"Workbook does not exist: {workbook}")
        _write_reports(result, reports)
        return result

    result.mappings = build_eoat_id_migration_map(workbook, result)
    id_map = result.id_map
    if validate_only:
        result.validation_issues = validate_eoat_id_prefixes(
            workbook,
            photo_root=photos,
            project_root=project,
            id_map=id_map,
        )
        _write_reports(result, reports)
        return result

    if not id_map:
        result.validation_issues = validate_eoat_id_prefixes(
            workbook,
            photo_root=photos,
            project_root=project,
            id_map={},
        )
        _write_reports(result, reports)
        return result

    planned_workbook_updates = apply_eoat_id_migration_to_workbook(workbook, id_map, dry_run=True)
    planned_photo_updates = apply_eoat_id_migration_to_photos(photos, id_map, dry_run=True) if photos else []
    planned_cache_updates = apply_eoat_id_migration_to_generated_files(project, id_map, dry_run=True) if project else []

    if not apply:
        result.workbook_updates = planned_workbook_updates
        result.photo_updates = planned_photo_updates
        result.cache_updates = planned_cache_updates
        result.conflicts.extend(_conflicts_from_updates(planned_photo_updates))
        result.validation_issues = validate_eoat_id_prefixes(
            workbook,
            photo_root=photos,
            project_root=project,
            id_map=id_map,
        )
        _write_reports(result, reports)
        result.warnings.append(f"Dry run completed in {time.perf_counter() - started:.2f} seconds. No files were modified.")
        return result

    backups_root.mkdir(parents=True, exist_ok=True)
    _create_backups(
        result,
        workbook,
        photos,
        planned_workbook_updates,
        planned_photo_updates,
        planned_cache_updates,
        backups_root,
        project,
    )

    result.workbook_updates = apply_eoat_id_migration_to_workbook(workbook, id_map, dry_run=False)
    if photos:
        result.photo_updates = apply_eoat_id_migration_to_photos(photos, id_map, dry_run=False)
        result.conflicts.extend(_conflicts_from_updates(result.photo_updates))
    if project:
        result.cache_updates = apply_eoat_id_migration_to_generated_files(project, id_map, dry_run=False)
    if rebuild_indexes:
        _invalidate_app_indexes(project, workbook)
    result.validation_issues = validate_eoat_id_prefixes(
        workbook,
        photo_root=photos,
        project_root=project,
        id_map=id_map,
    )
    _write_reports(result, reports)
    result.warnings.append(f"Apply completed in {time.perf_counter() - started:.2f} seconds.")
    return result


def build_eoat_id_migration_map(
    workbook_path: str | Path,
    result: EOATPrefixMigrationResult | None = None,
) -> list[EOATIDMapping]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    mappings: list[EOATIDMapping] = []
    try:
        if "EOAT Inventory" not in workbook.sheetnames:
            _append_error(result, "EOAT Inventory sheet is missing.")
            return []
        existing_locations = _eoat_id_locations(workbook)
        ws = workbook["EOAT Inventory"]
        headers = worksheet_headers(ws)
        positions = {header: index for index, header in enumerate(headers)}
        if EOAT_ASSEMBLY_ID_FIELD not in positions:
            _append_error(result, "EOAT Inventory is missing EOAT Assembly ID column.")
            return []
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_data = {header: row[index] for header, index in positions.items() if index < len(row)}
            if not _has_row_content(row_data):
                continue
            if is_compatibility_row(row_data):
                continue
            eoat_id = normalize_eoat_assembly_id(row_data.get(EOAT_ASSEMBLY_ID_FIELD))
            parsed = parse_eoat_id(eoat_id)
            if parsed is None:
                continue
            area = canonical_area(row_data)
            if area != CANONICAL_AREA_CLEANROOM or parsed.prefix != EOAT_PREFIX_PLANT4:
                continue
            new_id = format_eoat_id(EOAT_PREFIX_CLEANROOM, parsed.number)
            collisions = [
                location
                for location in existing_locations.get(new_id.casefold(), [])
                if location != ("EOAT Inventory", row_number, EOAT_ASSEMBLY_ID_FIELD)
            ]
            if collisions:
                _append_error(
                    result,
                    f"Collision blocked {eoat_id} -> {new_id}: {new_id} already exists at {_format_locations(collisions)}.",
                )
                continue
            mappings.append(
                EOATIDMapping(
                    old_id=eoat_id,
                    new_id=new_id,
                    source_sheet="EOAT Inventory",
                    source_row=row_number,
                    audit_id=_text(row_data.get("Audit ID")),
                    tool_number=_text(row_data.get(TOOL_FIELD)),
                    machine_number=_text(row_data.get("Press/Machine #")),
                    area_value=_text(row_data.get("Plant/Area") or row_data.get("Area")),
                    cleanroom_flag_value=_text(row_data.get("Cleanroom/Non-Cleanroom")),
                    reason="Cleanroom EOAT row used Plant 4 EOAT ID prefix.",
                )
            )
    finally:
        workbook.close()
    return mappings


def apply_eoat_id_migration_to_workbook(
    workbook_path: str | Path,
    id_map: dict[str, str],
    *,
    dry_run: bool = True,
) -> list[WorkbookUpdate]:
    workbook = load_workbook(workbook_path, data_only=False)
    updates: list[WorkbookUpdate] = []
    try:
        for ws in workbook.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    coordinate = _cell_reference(cell)
                    if isinstance(cell.value, str):
                        new_value = replace_mapped_eoat_ids(cell.value, id_map)
                        if new_value != cell.value:
                            updates.append(
                                WorkbookUpdate(
                                    sheet=ws.title,
                                    row=cell.row,
                                    column=coordinate,
                                    old_value=cell.value,
                                    new_value=new_value,
                                    reason="Updated migrated EOAT ID reference.",
                                )
                            )
                            if not dry_run:
                                cell.value = new_value
                    if cell.hyperlink is not None:
                        updates.extend(_replace_hyperlink_values(cell, id_map, dry_run=dry_run))
                    if cell.comment is not None:
                        new_text = replace_mapped_eoat_ids(cell.comment.text, id_map)
                        if new_text != cell.comment.text:
                            updates.append(
                                WorkbookUpdate(
                                    sheet=ws.title,
                                    row=cell.row,
                                    column=coordinate,
                                    old_value=cell.comment.text,
                                    new_value=new_text,
                                    reason="Updated migrated EOAT ID reference in cell comment.",
                                    value_kind="comment",
                                )
                            )
                            if not dry_run:
                                cell.comment = Comment(new_text, cell.comment.author)
        if not dry_run:
            try:
                refresh_audit_by_press_view(workbook)
            except Exception:
                pass
            workbook.save(workbook_path)
            invalidate_workbook_cache(workbook_path)
    finally:
        workbook.close()
    return updates


def apply_eoat_id_migration_to_photos(
    photo_root: str | Path | None,
    id_map: dict[str, str],
    *,
    dry_run: bool = True,
) -> list[PathUpdate]:
    if not photo_root:
        return []
    root = Path(photo_root)
    if not root.exists():
        return [PathUpdate(str(root), "", "photo_root", "missing", "Photo root does not exist.")]
    updates: list[PathUpdate] = []
    for folder in sorted((path for path in root.rglob("*") if path.is_dir()), key=_path_depth):
        if not folder.exists():
            continue
        target_name = replace_mapped_eoat_ids(folder.name, id_map)
        if target_name == folder.name:
            continue
        target = folder.with_name(target_name)
        updates.extend(_move_folder(folder, target, dry_run=dry_run))
    for file_path in sorted((path for path in root.rglob("*") if path.is_file()), key=_path_depth, reverse=True):
        if not file_path.exists():
            continue
        target_name = replace_mapped_eoat_ids(file_path.name, id_map)
        if target_name == file_path.name:
            continue
        target = file_path.with_name(target_name)
        updates.append(_move_file(file_path, target, dry_run=dry_run))
    return updates


def apply_eoat_id_migration_to_generated_files(
    project_root: str | Path | None,
    id_map: dict[str, str],
    *,
    dry_run: bool = True,
) -> list[PathUpdate]:
    if not project_root:
        return []
    root = Path(project_root)
    candidate_roots = [root / "00_Project_Admin" / "cache", root / ".cache"]
    updates: list[PathUpdate] = []
    for candidate_root in candidate_roots:
        if not candidate_root.exists():
            continue
        for path in sorted(candidate_root.rglob("*")):
            if not path.is_file() or path.suffix.casefold() not in CACHE_TEXT_EXTENSIONS:
                continue
            try:
                text = path.read_text(encoding="utf-8")
            except (OSError, UnicodeDecodeError):
                continue
            new_text = replace_mapped_eoat_ids(text, id_map)
            if new_text == text:
                continue
            updates.append(
                PathUpdate(
                    source=str(path),
                    target=str(path),
                    kind="generated_cache",
                    status="planned" if dry_run else "updated",
                    reason="Updated migrated EOAT ID reference in generated cache/index text file.",
                )
            )
            if not dry_run:
                path.write_text(new_text, encoding="utf-8")
    return updates


def validate_eoat_id_prefixes(
    workbook_path: str | Path,
    *,
    photo_root: str | Path | None = None,
    project_root: str | Path | None = None,
    id_map: dict[str, str] | None = None,
) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    id_map = id_map or {}
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        if "EOAT Inventory" in workbook.sheetnames:
            issues.extend(_validate_inventory_sheet(workbook["EOAT Inventory"]))
            issues.extend(_validate_compatibility_eoat_ids(workbook["EOAT Inventory"]))
        issues.extend(_validate_stale_workbook_references(workbook, id_map))
    finally:
        workbook.close()
    if photo_root:
        issues.extend(_validate_photo_paths(Path(photo_root), id_map))
    if project_root:
        issues.extend(_validate_generated_cache_references(Path(project_root), id_map))
    return issues


def replace_mapped_eoat_ids(value: Any, id_map: dict[str, str]) -> str:
    text = "" if value is None else str(value)
    if not text or not id_map:
        return text

    def replacement(match: Any) -> str:
        normalized = normalize_eoat_assembly_id(match.group(0))
        return id_map.get(normalized, match.group(0))

    return EOAT_ID_SEARCH_PATTERN.sub(replacement, text)


def write_migration_report(result: EOATPrefixMigrationResult, report_dir: str | Path) -> tuple[Path, Path]:
    report_root = ensure_directory(report_dir)
    md_path = report_root / f"eoat_id_prefix_migration_{result.timestamp}.md"
    csv_path = report_root / f"eoat_id_prefix_migration_{result.timestamp}.csv"
    safe_write_text(md_path, _migration_markdown(result), overwrite=True)
    _write_migration_csv(result, csv_path)
    result.migration_report_md = str(md_path)
    result.migration_report_csv = str(csv_path)
    return md_path, csv_path


def write_validation_report(result: EOATPrefixMigrationResult, report_dir: str | Path) -> tuple[Path, Path]:
    report_root = ensure_directory(report_dir)
    md_path = report_root / f"eoat_id_prefix_validation_{result.timestamp}.md"
    json_path = report_root / f"eoat_id_prefix_validation_{result.timestamp}.json"
    safe_write_text(md_path, _validation_markdown(result), overwrite=True)
    safe_write_text(
        json_path,
        json.dumps([issue.to_dict() for issue in result.validation_issues], indent=2, sort_keys=True) + "\n",
        overwrite=True,
    )
    result.validation_report_md = str(md_path)
    result.validation_report_json = str(json_path)
    return md_path, json_path


def _validate_inventory_sheet(ws: Any) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    headers = worksheet_headers(ws)
    positions = {header: index for index, header in enumerate(headers)}
    seen: dict[str, list[int]] = {}
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {header: row[index] for header, index in positions.items() if index < len(row)}
        if not _has_row_content(row_data):
            continue
        eoat_id = normalize_eoat_assembly_id(row_data.get(EOAT_ASSEMBLY_ID_FIELD))
        area = canonical_area(row_data)
        if not eoat_id:
            issues.append(_issue(row_data, row_number, "ERROR", "Missing EOAT Assembly ID.", area=area))
            continue
        if not is_valid_eoat_assembly_id(eoat_id):
            issues.append(_issue(row_data, row_number, "ERROR", f"Invalid EOAT Assembly ID format: {eoat_id}.", area=area))
            continue
        seen.setdefault(eoat_id.casefold(), []).append(row_number)
        if area == CANONICAL_AREA_UNKNOWN:
            issues.append(
                _issue(
                    row_data,
                    row_number,
                    "WARNING",
                    f"Area is unknown, cannot determine expected EOAT prefix for {eoat_id}.",
                    area=area,
                    suggested="Set Plant/Area or Cleanroom/Non-Cleanroom.",
                )
            )
            continue
        expected_prefix = determine_eoat_prefix(row_data)
        actual_prefix = get_eoat_prefix(eoat_id)
        if actual_prefix != expected_prefix:
            issues.append(
                _issue(
                    row_data,
                    row_number,
                    "ERROR",
                    f"EOAT Assembly ID prefix mismatch: {eoat_id} should use {expected_prefix}.",
                    area=area,
                    expected_prefix=expected_prefix,
                    actual_prefix=actual_prefix,
                    suggested="Run Repair EOAT ID Prefixes.",
                )
            )
    for normalized_id, rows in seen.items():
        if len(rows) <= 1:
            continue
        issues.append(
            ValidationIssue(
                severity="ERROR",
                source="EOAT Inventory",
                message=f"Duplicate EOAT Assembly ID {normalized_id.upper()} appears on rows {', '.join(str(row) for row in rows)}.",
                sheet="EOAT Inventory",
                row=rows[0],
                column=EOAT_ASSEMBLY_ID_FIELD,
                current_eoat_id=normalized_id.upper(),
                suggested_repair="Review duplicate physical EOAT rows before merging or renumbering.",
            )
        )
    return issues


def _validate_compatibility_eoat_ids(ws: Any) -> list[ValidationIssue]:
    headers = worksheet_headers(ws)
    positions = {header: index for index, header in enumerate(headers)}
    source_eoat_by_audit: dict[str, str] = {}
    row_data_by_number: dict[int, dict[str, Any]] = {}
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {header: row[index] for header, index in positions.items() if index < len(row)}
        row_data_by_number[row_number] = row_data
        audit_id = _text(row_data.get("Audit ID"))
        eoat_id = normalize_eoat_assembly_id(row_data.get(EOAT_ASSEMBLY_ID_FIELD))
        if audit_id and eoat_id and not is_compatibility_row(row_data):
            source_eoat_by_audit[audit_id] = eoat_id
    issues: list[ValidationIssue] = []
    for row_number, row_data in row_data_by_number.items():
        if not is_compatibility_row(row_data):
            continue
        source_id = _text(row_data.get(SOURCE_AUDIT_ID_FIELD))
        if not source_id:
            continue
        source_eoat = source_eoat_by_audit.get(source_id, "")
        row_eoat = normalize_eoat_assembly_id(row_data.get(EOAT_ASSEMBLY_ID_FIELD))
        if source_eoat and row_eoat and row_eoat != source_eoat:
            issues.append(
                _issue(
                    row_data,
                    row_number,
                    "ERROR",
                    f"Compatibility row EOAT ID {row_eoat} disagrees with source audit {source_id}: {source_eoat}.",
                    source_audit_id=source_id,
                    suggested="Refresh generated compatibility rows from the source audit.",
                )
            )
    return issues


def _validate_stale_workbook_references(workbook: Any, id_map: dict[str, str]) -> list[ValidationIssue]:
    if not id_map:
        return []
    issues: list[ValidationIssue] = []
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                values = []
                if isinstance(cell.value, str):
                    values.append(("cell", cell.value))
                hyperlink = getattr(cell, "hyperlink", None)
                if hyperlink is not None:
                    for attr in ("target", "location", "display", "tooltip"):
                        value = getattr(hyperlink, attr, None)
                        if isinstance(value, str):
                            values.append((f"hyperlink.{attr}", value))
                for value_kind, value in values:
                    stale = [eoat_id for eoat_id in find_eoat_ids(value) if eoat_id in id_map]
                    for old_id in stale:
                        issues.append(
                            ValidationIssue(
                                severity="ERROR",
                                source="workbook",
                                message=f"Workbook contains stale migrated ID {old_id}.",
                                sheet=ws.title,
                                row=cell.row,
                                column=_cell_reference(cell),
                                current_eoat_id=old_id,
                                expected_prefix=get_eoat_prefix(id_map[old_id]),
                                actual_prefix=get_eoat_prefix(old_id),
                                suggested_repair=f"Replace with {id_map[old_id]}.",
                            )
                        )
    return issues


def _validate_photo_paths(photo_root: Path, id_map: dict[str, str]) -> list[ValidationIssue]:
    if not id_map or not photo_root.exists():
        return []
    issues: list[ValidationIssue] = []
    for path in photo_root.rglob("*"):
        stale = [old_id for old_id in id_map if old_id in str(path)]
        for old_id in stale:
            issues.append(
                ValidationIssue(
                    severity="ERROR",
                    source="photo_archive",
                    message=f"Physical photo path still contains stale migrated ID {old_id}.",
                    current_eoat_id=old_id,
                    expected_prefix=get_eoat_prefix(id_map[old_id]),
                    actual_prefix=get_eoat_prefix(old_id),
                    suggested_repair=f"Rename path to use {id_map[old_id]}.",
                )
            )
    return issues


def _validate_generated_cache_references(project_root: Path, id_map: dict[str, str]) -> list[ValidationIssue]:
    if not id_map:
        return []
    issues: list[ValidationIssue] = []
    for update in apply_eoat_id_migration_to_generated_files(project_root, id_map, dry_run=True):
        issues.append(
            ValidationIssue(
                severity="ERROR",
                source="generated_cache",
                message=f"Generated cache/index file still contains stale migrated EOAT ID references: {update.source}.",
                suggested_repair="Rebuild or repair generated app indexes/caches.",
            )
        )
    return issues


def _replace_hyperlink_values(cell: Any, id_map: dict[str, str], *, dry_run: bool) -> list[WorkbookUpdate]:
    updates: list[WorkbookUpdate] = []
    hyperlink = cell.hyperlink
    for attr in ("target", "location", "display", "tooltip"):
        value = getattr(hyperlink, attr, None)
        if not isinstance(value, str):
            continue
        new_value = replace_mapped_eoat_ids(value, id_map)
        if new_value == value:
            continue
        updates.append(
            WorkbookUpdate(
                sheet=cell.parent.title,
                row=cell.row,
                column=_cell_reference(cell),
                old_value=value,
                new_value=new_value,
                reason=f"Updated migrated EOAT ID reference in hyperlink {attr}.",
                value_kind=f"hyperlink.{attr}",
            )
        )
        if not dry_run:
            try:
                setattr(hyperlink, attr, new_value)
            except Exception:
                pass
    return updates


def _move_folder(source: Path, target: Path, *, dry_run: bool) -> list[PathUpdate]:
    if source == target:
        return []
    if not target.exists():
        if not dry_run:
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(source), str(target))
        return [PathUpdate(str(source), str(target), "photo_folder", "planned" if dry_run else "renamed", "Renamed EOAT photo folder.")]
    if not target.is_dir():
        return [PathUpdate(str(source), str(target), "photo_folder", "conflict", "Target exists and is not a folder.")]
    updates = [PathUpdate(str(source), str(target), "photo_folder", "merge_planned" if dry_run else "merged", "Merged EOAT photo folder into existing target folder.")]
    for child in sorted(source.iterdir(), key=lambda path: path.name.casefold()):
        child_target = target / child.name
        if child.is_dir():
            updates.extend(_move_folder(child, child_target, dry_run=dry_run))
        else:
            updates.append(_move_file(child, child_target, dry_run=dry_run))
    if not dry_run:
        try:
            source.rmdir()
        except OSError:
            pass
    return updates


def _move_file(source: Path, target: Path, *, dry_run: bool) -> PathUpdate:
    if source == target:
        return PathUpdate(str(source), str(target), "photo_file", "unchanged", "Source and target are identical.")
    if target.exists():
        if target.is_file() and source.is_file():
            try:
                if filecmp.cmp(source, target, shallow=False):
                    return PathUpdate(
                        str(source),
                        str(target),
                        "photo_file",
                        "identical_target_exists",
                        "Target file already exists and is byte-for-byte identical; source was left in place.",
                    )
            except OSError:
                pass
        return PathUpdate(str(source), str(target), "photo_file", "conflict", "Target file already exists with different content.")
    if not dry_run:
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(source), str(target))
    return PathUpdate(str(source), str(target), "photo_file", "planned" if dry_run else "renamed", "Renamed EOAT photo file.")


def _create_backups(
    result: EOATPrefixMigrationResult,
    workbook: Path,
    photo_root: Path | None,
    workbook_updates: list[WorkbookUpdate],
    photo_updates: list[PathUpdate],
    cache_updates: list[PathUpdate],
    backup_root: Path,
    project_root: Path | None,
) -> None:
    workbook_backup = backup_root / "workbook" / workbook.name
    workbook_backup.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(workbook, workbook_backup)
    result.backups.append(str(workbook_backup))
    if photo_root and any(update.status not in {"missing", "conflict", "identical_target_exists"} for update in photo_updates):
        photo_backup = backup_root / "photos" / photo_root.name
        if photo_backup.exists():
            shutil.rmtree(photo_backup)
        shutil.copytree(photo_root, photo_backup)
        result.backups.append(str(photo_backup))
    if cache_updates and project_root:
        cache_backup_root = backup_root / "generated_cache_files"
        for update in cache_updates:
            source = Path(update.source)
            if not source.exists() or not source.is_file():
                continue
            try:
                relative = source.relative_to(project_root)
            except ValueError:
                relative = Path(source.name)
            target = cache_backup_root / relative
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)
            result.backups.append(str(target))
    if not workbook_updates:
        result.warnings.append("Workbook backup was still created before apply; no workbook cell changes were planned.")


def _invalidate_app_indexes(project_root: Path | None, workbook_path: Path) -> None:
    try:
        invalidate_workbook_cache(workbook_path)
    except Exception:
        pass
    if project_root:
        try:
            from .atlas_data_loader import invalidate_atlas_data_cache

            invalidate_atlas_data_cache(project_root)
        except Exception:
            pass


def _write_reports(result: EOATPrefixMigrationResult, report_dir: Path) -> None:
    write_migration_report(result, report_dir)
    write_validation_report(result, report_dir)


def _migration_markdown(result: EOATPrefixMigrationResult) -> str:
    lines = [
        "# EOAT ID Prefix Migration",
        "",
        f"- Timestamp: {result.timestamp}",
        f"- Mode: {'dry-run' if result.dry_run else 'apply'}",
        f"- Workbook: {result.workbook_path}",
        f"- Photo root: {result.photo_root or 'N/A'}",
        f"- Project root: {result.project_root or 'N/A'}",
        f"- Backup folder: {result.backup_dir or 'N/A'}",
        "",
        "## Summary",
        f"- IDs migrated: {len(result.mappings)}",
        f"- Workbook references updated: {len(result.workbook_updates)}",
        f"- Photo folder/file operations: {len(result.photo_updates)}",
        f"- Generated cache/index files updated: {len(result.cache_updates)}",
        f"- Conflicts: {len(result.conflicts)}",
        f"- Validation issues after scan: {len(result.validation_issues)}",
    ]
    if result.backups:
        lines.extend(["", "## Backups", *[f"- {path}" for path in result.backups]])
    if result.errors:
        lines.extend(["", "## Errors", *[f"- {error}" for error in result.errors]])
    if result.warnings:
        lines.extend(["", "## Warnings", *[f"- {warning}" for warning in result.warnings]])
    if result.conflicts:
        lines.extend(["", "## Conflicts", *[f"- {conflict}" for conflict in result.conflicts]])
    lines.extend(["", "## ID Mapping"])
    if result.mappings:
        lines.extend(["| Old ID | New ID | Sheet | Row | Audit ID | Reason |", "|---|---|---:|---:|---|---|"])
        for item in result.mappings:
            lines.append(
                f"| {item.old_id} | {item.new_id} | {item.source_sheet} | {item.source_row} | {item.audit_id} | {item.reason} |"
            )
    else:
        lines.append("No Cleanroom P4-to-CL mappings were required or safe to apply.")
    _append_updates_table(lines, "Workbook Changes", result.workbook_updates, "workbook")
    _append_path_table(lines, "Photo Changes", result.photo_updates)
    _append_path_table(lines, "Generated Cache/Index Changes", result.cache_updates)
    lines.extend(["", "## Validation Results"])
    if result.validation_issues:
        for issue in result.validation_issues[:300]:
            target = issue.source
            if issue.sheet:
                target += f" / {issue.sheet}"
            if issue.row:
                target += f" row {issue.row}"
            lines.append(f"- {issue.severity}: {target}: {issue.message}")
        if len(result.validation_issues) > 300:
            lines.append(f"- ... {len(result.validation_issues) - 300} more validation issue(s) omitted.")
    else:
        lines.append("No EOAT ID prefix validation issues were found.")
    lines.extend(
        [
            "",
            "## Numbering Strategy",
            "The app uses separate numeric sequences per EOAT prefix. Existing migrated IDs preserve their four-digit numeric suffix unless a direct collision is detected.",
        ]
    )
    return "\n".join(lines) + "\n"


def _validation_markdown(result: EOATPrefixMigrationResult) -> str:
    lines = [
        "# EOAT ID Prefix Validation",
        "",
        f"- Timestamp: {result.timestamp}",
        f"- Workbook: {result.workbook_path}",
        f"- Photo root: {result.photo_root or 'N/A'}",
        "",
        "## Issues",
    ]
    if not result.validation_issues:
        lines.append("No EOAT ID prefix validation issues were found.")
    for issue in result.validation_issues:
        target = issue.source
        if issue.sheet:
            target += f" / {issue.sheet}"
        if issue.row:
            target += f" row {issue.row}"
        lines.append(f"- {issue.severity}: {target}: {issue.message}")
        if issue.suggested_repair:
            lines.append(f"  Suggested repair: {issue.suggested_repair}")
    return "\n".join(lines) + "\n"


def _append_updates_table(lines: list[str], title: str, updates: list[WorkbookUpdate], kind: str) -> None:
    lines.extend(["", f"## {title}"])
    if not updates:
        lines.append(f"No {kind} changes.")
        return
    lines.extend(["| Sheet | Row | Column | Kind | Old Value | New Value |", "|---|---:|---|---|---|---|"])
    for update in updates[:200]:
        lines.append(
            f"| {update.sheet} | {update.row} | {update.column} | {update.value_kind} | {_table_text(update.old_value)} | {_table_text(update.new_value)} |"
        )
    if len(updates) > 200:
        lines.append(f"| ... |  |  |  | {len(updates) - 200} more change(s) omitted |  |")


def _append_path_table(lines: list[str], title: str, updates: list[PathUpdate]) -> None:
    lines.extend(["", f"## {title}"])
    if not updates:
        lines.append("No filesystem/cache changes.")
        return
    lines.extend(["| Kind | Status | Source | Target | Reason |", "|---|---|---|---|---|"])
    for update in updates[:200]:
        lines.append(
            f"| {update.kind} | {update.status} | {_table_text(update.source)} | {_table_text(update.target)} | {update.reason} |"
        )
    if len(updates) > 200:
        lines.append(f"| ... | ... | {len(updates) - 200} more change(s) omitted |  |  |")


def _write_migration_csv(result: EOATPrefixMigrationResult, path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["kind", "source", "target", "sheet", "row", "column", "old_value", "new_value", "status", "reason"],
        )
        writer.writeheader()
        for mapping in result.mappings:
            writer.writerow(
                {
                    "kind": "id_mapping",
                    "source": mapping.old_id,
                    "target": mapping.new_id,
                    "sheet": mapping.source_sheet,
                    "row": mapping.source_row,
                    "column": EOAT_ASSEMBLY_ID_FIELD,
                    "old_value": mapping.old_id,
                    "new_value": mapping.new_id,
                    "status": "mapped",
                    "reason": mapping.reason,
                }
            )
        for update in result.workbook_updates:
            writer.writerow(
                {
                    "kind": update.value_kind,
                    "source": result.workbook_path,
                    "target": result.workbook_path,
                    "sheet": update.sheet,
                    "row": update.row,
                    "column": update.column,
                    "old_value": update.old_value,
                    "new_value": update.new_value,
                    "status": "planned" if result.dry_run else "updated",
                    "reason": update.reason,
                }
            )
        for update in [*result.photo_updates, *result.cache_updates]:
            writer.writerow(
                {
                    "kind": update.kind,
                    "source": update.source,
                    "target": update.target,
                    "sheet": "",
                    "row": "",
                    "column": "",
                    "old_value": update.source,
                    "new_value": update.target,
                    "status": update.status,
                    "reason": update.reason,
                }
            )


def _eoat_id_locations(workbook: Any) -> dict[str, list[tuple[str, int, str]]]:
    locations: dict[str, list[tuple[str, int, str]]] = {}
    for ws in workbook.worksheets:
        headers = worksheet_headers(ws)
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                for eoat_id in find_eoat_ids(cell.value):
                    locations.setdefault(eoat_id.casefold(), []).append((ws.title, cell.row, _cell_reference(cell)))
    return locations


def _issue(
    row_data: dict[str, Any],
    row_number: int,
    severity: str,
    message: str,
    *,
    area: str = "",
    expected_prefix: str = "",
    actual_prefix: str = "",
    source_audit_id: str = "",
    suggested: str = "",
) -> ValidationIssue:
    return ValidationIssue(
        severity=severity,
        source="EOAT Inventory",
        message=message,
        sheet="EOAT Inventory",
        row=row_number,
        column=EOAT_ASSEMBLY_ID_FIELD,
        audit_id=_text(row_data.get("Audit ID")),
        source_audit_id=source_audit_id,
        tool_number=_text(row_data.get(TOOL_FIELD)),
        machine_number=_text(row_data.get("Press/Machine #")),
        area_value=_text(row_data.get("Plant/Area") or row_data.get("Area")),
        cleanroom_flag_value=_text(row_data.get("Cleanroom/Non-Cleanroom")),
        current_eoat_id=normalize_eoat_assembly_id(row_data.get(EOAT_ASSEMBLY_ID_FIELD)),
        expected_prefix=expected_prefix,
        actual_prefix=actual_prefix,
        suggested_repair=suggested,
    )


def _conflicts_from_updates(updates: Iterable[PathUpdate]) -> list[str]:
    return [
        f"{update.kind}: {update.source} -> {update.target}: {update.reason}"
        for update in updates
        if update.status in {"conflict", "identical_target_exists"}
    ]


def _project_root_from_workbook(workbook_path: Path) -> Path | None:
    parts = list(workbook_path.parts)
    try:
        index = parts.index("01_EOAT_Audit")
    except ValueError:
        return None
    return Path(*parts[:index])


def _default_report_dir(project_root: Path | None, workbook_path: Path) -> Path:
    if project_root:
        return project_root / "00_Project_Admin" / "Validation_Reports"
    return workbook_path.parent / "reports"


def _default_backup_dir(project_root: Path | None, workbook_path: Path, timestamp: str) -> Path:
    root = project_root if project_root else workbook_path.parent
    return root / "backups" / f"{MIGRATION_NAME}_{timestamp}"


def _append_error(result: EOATPrefixMigrationResult | None, message: str) -> None:
    if result is not None:
        result.errors.append(message)


def _format_locations(locations: list[tuple[str, int, str]]) -> str:
    return ", ".join(f"{sheet} row {row} {column}" for sheet, row, column in locations[:10])


def _has_row_content(row_data: dict[str, Any]) -> bool:
    return any(_text(value) for value in row_data.values())


def _cell_reference(cell: Any) -> str:
    return f"{get_column_letter(cell.column)}{cell.row}"


def _path_depth(path: Path) -> int:
    return len(path.parts)


def _table_text(value: Any) -> str:
    text = _text(value).replace("|", "\\|")
    text = text.replace("\n", " ")
    return text[:250] + "..." if len(text) > 250 else text


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


__all__ = [
    "EOATIDMapping",
    "EOATPrefixMigrationResult",
    "FIX_REPAIR_EOAT_ID_PREFIXES",
    "PathUpdate",
    "ValidationIssue",
    "WorkbookUpdate",
    "apply_eoat_id_migration_to_generated_files",
    "apply_eoat_id_migration_to_photos",
    "apply_eoat_id_migration_to_workbook",
    "build_eoat_id_migration_map",
    "replace_mapped_eoat_ids",
    "run_eoat_id_prefix_migration",
    "validate_eoat_id_prefixes",
    "write_migration_report",
    "write_validation_report",
]
