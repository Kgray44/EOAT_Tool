from __future__ import annotations

import sqlite3
import time
import uuid
from contextlib import contextmanager
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from core.logging import log_activity_event
from core.paths import resolve_project_paths
from core.safe_files import backup_file
from core.workbook_io import worksheet_headers

from .database import annotation_database_path, connect_annotation_database, initialize_annotation_database
from .exports import export_notes_excel, export_notes_markdown, export_tags_excel, export_tags_markdown
from .migrations import utc_now
from .models import AnnotationTarget, Note, Tag, TagAssignment
from .suggestions import suggested_annotations_for_audit
from .tag_colors import TAG_COLOR_PALETTE, excel_fill_for_color, highest_priority_tag, normalize_color_key
from .targets import display_label_for_target, normalize_target_type, target_id_for

IMPORTANCE_VALUES = ("Low", "Neutral", "Important", "Critical")
STATUS_VALUES = ("Open", "Resolved", "Archived")
NOTE_SORTS = {
    "updated": "updated_at DESC",
    "updated date": "updated_at DESC",
    "created": "created_at DESC",
    "created date": "created_at DESC",
    "subject": "lower(subject) ASC",
    "subject alphabetical": "lower(subject) ASC",
    "importance": "CASE importance WHEN 'Critical' THEN 0 WHEN 'Important' THEN 1 WHEN 'Neutral' THEN 2 WHEN 'Low' THEN 3 ELSE 4 END",
    "status": "lower(coalesce(status, '')) ASC",
    "collection": "lower(coalesce(collection, '')) ASC",
    "note_type": "lower(coalesce(note_type, '')) ASC",
    "note type": "lower(coalesce(note_type, '')) ASC",
    "follow_up_date": "coalesce(follow_up_date, '9999-99-99') ASC",
    "follow-up date": "coalesce(follow_up_date, '9999-99-99') ASC",
}
TAG_SORTS = {
    "name": "lower(tags.name) ASC",
    "tag name": "lower(tags.name) ASC",
    "color": "tags.color_key ASC",
    "updated": "tags.updated_at DESC",
    "updated date": "tags.updated_at DESC",
    "priority": "lower(tags.name) ASC",
}
ASSIGNMENT_SORTS = {
    "updated": "tag_assignments.updated_at DESC",
    "updated date": "tag_assignments.updated_at DESC",
    "tag": "lower(tags.name) ASC",
    "tag name": "lower(tags.name) ASC",
    "color": "tags.color_key ASC",
    "target type": "annotation_targets.target_type ASC",
    "target_type": "annotation_targets.target_type ASC",
}


class AnnotationService:
    def __init__(self, project_root: str | Path, db_path: str | Path | None = None, *, initialize: bool = True):
        self.project_root = Path(project_root)
        self.db_path = Path(db_path) if db_path else annotation_database_path(self.project_root)
        if initialize:
            initialize_annotation_database(self.project_root, self.db_path)

    @contextmanager
    def connection(self):
        conn = connect_annotation_database(self.db_path)
        try:
            yield conn
            conn.commit()
        finally:
            conn.close()

    def create_note(
        self,
        subject: str,
        body_markdown: str = "",
        importance: str = "Neutral",
        *,
        status: str | None = None,
        collection: str | None = None,
        note_type: str | None = None,
        follow_up_date: str | None = None,
        target_ids: Iterable[str] | None = None,
        tag_ids: Iterable[str] | None = None,
        attachment_paths: Iterable[str | Path] | None = None,
    ) -> Note:
        subject = str(subject or "").strip()
        if not subject:
            raise ValueError("Note subject is required.")
        importance = _validate_importance(importance)
        status = _validate_status(status)
        note_id = _new_id("note")
        now = utc_now()
        with self.connection() as conn:
            conn.execute(
                """
                INSERT INTO notes(id, subject, body_markdown, importance, status, collection, note_type, follow_up_date, created_at, updated_at)
                VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (note_id, subject, body_markdown or "", importance, status, _none_empty(collection), _none_empty(note_type), _none_empty(follow_up_date), now, now),
            )
            for target_id in target_ids or []:
                self._link_note_to_target(conn, note_id, str(target_id))
            for tag_id in tag_ids or []:
                self._link_note_to_tag(conn, note_id, str(tag_id))
            for file_path in attachment_paths or []:
                self._attach_file(conn, note_id=note_id, target_id=None, file_path=file_path)
        self._log("annotation_note_created", {"note_id": note_id, "subject": subject})
        return self.get_note(note_id)

    def get_note(self, note_id: str) -> Note:
        with self.connection() as conn:
            row = conn.execute("SELECT * FROM notes WHERE id = ?", (note_id,)).fetchone()
        if row is None:
            raise KeyError(note_id)
        return _note_from_row(row)

    def update_note(self, note_id: str, **updates: Any) -> Note:
        allowed = {"subject", "body_markdown", "importance", "status", "collection", "note_type", "follow_up_date"}
        fields: list[str] = []
        values: list[Any] = []
        for key, value in updates.items():
            if key not in allowed:
                continue
            if key == "importance":
                value = _validate_importance(str(value or "Neutral"))
            if key == "status":
                value = _validate_status(value)
            fields.append(f"{key} = ?")
            values.append(_none_empty(value))
        if not fields:
            return self.get_note(note_id)
        fields.append("updated_at = ?")
        values.extend([utc_now(), note_id])
        with self.connection() as conn:
            conn.execute(f"UPDATE notes SET {', '.join(fields)} WHERE id = ?", values)
        self._log("annotation_note_updated", {"note_id": note_id})
        return self.get_note(note_id)

    def archive_note(self, note_id: str) -> Note:
        now = utc_now()
        with self.connection() as conn:
            conn.execute("UPDATE notes SET archived_at = ?, updated_at = ? WHERE id = ?", (now, now, note_id))
        self._log("annotation_note_archived", {"note_id": note_id})
        return self.get_note(note_id)

    def search_notes(
        self,
        query: str = "",
        *,
        importance: str | None = None,
        status: str | None = None,
        collection: str | None = None,
        note_type: str | None = None,
        audit_id: str | None = None,
        machine_id: str | None = None,
        tag_id: str | None = None,
        tag_name: str | None = None,
        open_items_only: bool = False,
        follow_up_due_before: str | None = None,
        include_archived: bool = False,
        sort_by: str = "updated",
    ) -> list[dict[str, object]]:
        clauses: list[str] = []
        params: list[Any] = []
        if not include_archived:
            clauses.append("notes.archived_at IS NULL")
        if query:
            like = f"%{query.strip()}%"
            clauses.append(
                """
                (
                    notes.subject LIKE ?
                    OR notes.body_markdown LIKE ?
                    OR coalesce(notes.collection, '') LIKE ?
                    OR coalesce(notes.note_type, '') LIKE ?
                    OR EXISTS (
                        SELECT 1 FROM note_targets nt
                        JOIN annotation_targets at ON at.id = nt.target_id
                        WHERE nt.note_id = notes.id
                        AND (coalesce(at.audit_id, '') LIKE ? OR coalesce(at.machine_id, '') LIKE ? OR coalesce(at.target_label, '') LIKE ?)
                    )
                    OR EXISTS (
                        SELECT 1 FROM note_tags ntag
                        JOIN tags ON tags.id = ntag.tag_id
                        WHERE ntag.note_id = notes.id AND tags.name LIKE ?
                    )
                )
                """
            )
            params.extend([like, like, like, like, like, like, like, like])
        if importance and importance != "All":
            clauses.append("notes.importance = ?")
            params.append(importance)
        if status and status != "All":
            clauses.append("notes.status = ?")
            params.append(status)
        if collection and collection != "All":
            clauses.append("notes.collection = ?")
            params.append(collection)
        if note_type and note_type != "All":
            clauses.append("notes.note_type = ?")
            params.append(note_type)
        if audit_id:
            clauses.append("EXISTS (SELECT 1 FROM note_targets nt JOIN annotation_targets at ON at.id = nt.target_id WHERE nt.note_id = notes.id AND at.audit_id = ?)")
            params.append(audit_id)
        if machine_id:
            clauses.append("EXISTS (SELECT 1 FROM note_targets nt JOIN annotation_targets at ON at.id = nt.target_id WHERE nt.note_id = notes.id AND at.machine_id = ?)")
            params.append(machine_id)
        if tag_id:
            clauses.append("EXISTS (SELECT 1 FROM note_tags ntag WHERE ntag.note_id = notes.id AND ntag.tag_id = ?)")
            params.append(tag_id)
        if tag_name:
            clauses.append("EXISTS (SELECT 1 FROM note_tags ntag JOIN tags ON tags.id = ntag.tag_id WHERE ntag.note_id = notes.id AND lower(tags.name) = lower(?))")
            params.append(tag_name)
        if open_items_only:
            clauses.append("(notes.status IS NULL OR lower(notes.status) NOT IN ('resolved', 'archived'))")
        if follow_up_due_before:
            clauses.append("notes.follow_up_date IS NOT NULL AND notes.follow_up_date <= ?")
            params.append(follow_up_due_before)
        where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        order = NOTE_SORTS.get(sort_by.strip().lower(), NOTE_SORTS["updated"])
        with self.connection() as conn:
            rows = conn.execute(f"SELECT notes.* FROM notes {where} ORDER BY {order}", params).fetchall()
            return [self._enrich_note_row(conn, row) for row in rows]

    def create_tag(self, name: str, color_key: str = "yellow", *, description: str | None = None, is_default: bool = False) -> Tag:
        name = str(name or "").strip()
        if not name:
            raise ValueError("Tag name is required.")
        color_key = normalize_color_key(color_key)
        with self.connection() as conn:
            existing = conn.execute("SELECT * FROM tags WHERE lower(name) = lower(?) AND is_archived = 0", (name,)).fetchone()
            if existing:
                return _tag_from_row(existing)
            now = utc_now()
            tag_id = _new_id("tag")
            conn.execute(
                """
                INSERT INTO tags(id, name, color_key, description, is_default, is_archived, created_at, updated_at)
                VALUES(?, ?, ?, ?, ?, 0, ?, ?)
                """,
                (tag_id, name, color_key, _none_empty(description), 1 if is_default else 0, now, now),
            )
        self._log("annotation_tag_created", {"tag_id": tag_id, "name": name, "color_key": color_key})
        return self.get_tag(tag_id)

    def get_tag(self, tag_id: str) -> Tag:
        with self.connection() as conn:
            row = conn.execute("SELECT * FROM tags WHERE id = ?", (tag_id,)).fetchone()
        if row is None:
            raise KeyError(tag_id)
        return _tag_from_row(row)

    def get_tag_by_name(self, name: str) -> Tag | None:
        with self.connection() as conn:
            row = conn.execute("SELECT * FROM tags WHERE lower(name) = lower(?) AND is_archived = 0", (name,)).fetchone()
        return _tag_from_row(row) if row else None

    def update_tag(self, tag_id: str, **updates: Any) -> Tag:
        allowed = {"name", "color_key", "description"}
        fields: list[str] = []
        values: list[Any] = []
        for key, value in updates.items():
            if key not in allowed:
                continue
            if key == "color_key":
                value = normalize_color_key(str(value))
            fields.append(f"{key} = ?")
            values.append(_none_empty(value))
        if not fields:
            return self.get_tag(tag_id)
        fields.append("updated_at = ?")
        values.extend([utc_now(), tag_id])
        with self.connection() as conn:
            conn.execute(f"UPDATE tags SET {', '.join(fields)} WHERE id = ?", values)
        self._log("annotation_tag_updated", {"tag_id": tag_id})
        return self.get_tag(tag_id)

    def archive_tag(self, tag_id: str) -> Tag:
        with self.connection() as conn:
            conn.execute("UPDATE tags SET is_archived = 1, updated_at = ? WHERE id = ?", (utc_now(), tag_id))
        self._log("annotation_tag_archived", {"tag_id": tag_id})
        return self.get_tag(tag_id)

    def list_tags(self, *, include_archived: bool = False, sort_by: str = "name") -> list[Tag]:
        where = "" if include_archived else "WHERE is_archived = 0"
        order = TAG_SORTS.get(sort_by.strip().lower(), TAG_SORTS["name"]).replace("tags.", "")
        with self.connection() as conn:
            rows = conn.execute(f"SELECT * FROM tags {where} ORDER BY {order}").fetchall()
        return [_tag_from_row(row) for row in rows]

    def search_tags(
        self,
        query: str = "",
        *,
        color_key: str | None = None,
        include_archived: bool = False,
        sort_by: str = "name",
    ) -> list[Tag]:
        clauses: list[str] = []
        params: list[Any] = []
        if not include_archived:
            clauses.append("tags.is_archived = 0")
        if query:
            like = f"%{query.strip()}%"
            clauses.append("(tags.name LIKE ? OR coalesce(tags.description, '') LIKE ?)")
            params.extend([like, like])
        if color_key and color_key != "All":
            clauses.append("tags.color_key = ?")
            params.append(normalize_color_key(color_key))
        where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        order = TAG_SORTS.get(sort_by.strip().lower(), TAG_SORTS["name"])
        with self.connection() as conn:
            rows = conn.execute(f"SELECT tags.* FROM tags {where} ORDER BY {order}", params).fetchall()
        return [_tag_from_row(row) for row in rows]

    def create_or_get_target(
        self,
        target_type: str,
        *,
        target_label: str = "",
        audit_id: str = "",
        machine_id: str = "",
        field_key: str = "",
        field_label: str = "",
        sheet_name: str = "",
        header_name: str = "",
        workbook_path: str | Path | None = None,
        cached_cell_ref: str = "",
        object_ref: str = "",
    ) -> AnnotationTarget:
        normalized_type = normalize_target_type(target_type)
        identity_field_key = field_key or header_name or field_label
        if normalized_type == "audit_field":
            existing_target = self.find_audit_field_target(audit_id, identity_field_key)
            if existing_target is not None:
                with self.connection() as conn:
                    conn.execute(
                        """
                        UPDATE annotation_targets
                        SET target_label = ?, machine_id = ?, field_label = ?, sheet_name = ?, header_name = ?,
                            workbook_path = ?, cached_cell_ref = coalesce(?, cached_cell_ref), updated_at = ?
                        WHERE id = ?
                        """,
                        (
                            _none_empty(target_label or existing_target.target_label),
                            _none_empty(machine_id or existing_target.machine_id),
                            _none_empty(field_label or existing_target.field_label),
                            _none_empty(sheet_name or existing_target.sheet_name),
                            _none_empty(header_name or existing_target.header_name),
                            str(workbook_path) if workbook_path else existing_target.workbook_path,
                            _none_empty(cached_cell_ref),
                            utc_now(),
                            existing_target.id,
                        ),
                    )
                return self.get_target(existing_target.id)
        target_id = target_id_for(target_type=normalized_type, audit_id=audit_id, machine_id=machine_id, field_key=identity_field_key, object_ref=object_ref)
        label = display_label_for_target(
            target_type=normalized_type,
            target_label=target_label,
            audit_id=audit_id,
            machine_id=machine_id,
            field_label=field_label,
            field_key=field_key,
            object_ref=object_ref,
        )
        now = utc_now()
        with self.connection() as conn:
            row = conn.execute("SELECT * FROM annotation_targets WHERE id = ?", (target_id,)).fetchone()
            if row:
                conn.execute(
                    """
                    UPDATE annotation_targets
                    SET target_label = ?, audit_id = ?, machine_id = ?, field_key = ?, field_label = ?,
                        sheet_name = ?, header_name = ?, workbook_path = ?, cached_cell_ref = ?, object_ref = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        _none_empty(label),
                        _none_empty(audit_id),
                        _none_empty(machine_id),
                        _none_empty(field_key),
                        _none_empty(field_label),
                        _none_empty(sheet_name),
                        _none_empty(header_name),
                        str(workbook_path) if workbook_path else None,
                        _none_empty(cached_cell_ref),
                        _none_empty(object_ref),
                        now,
                        target_id,
                    ),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO annotation_targets(
                        id, target_type, target_label, audit_id, machine_id, field_key, field_label,
                        sheet_name, header_name, workbook_path, cached_cell_ref, object_ref, created_at, updated_at
                    )
                    VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        target_id,
                        normalized_type,
                        _none_empty(label),
                        _none_empty(audit_id),
                        _none_empty(machine_id),
                        _none_empty(field_key),
                        _none_empty(field_label),
                        _none_empty(sheet_name),
                        _none_empty(header_name),
                        str(workbook_path) if workbook_path else None,
                        _none_empty(cached_cell_ref),
                        _none_empty(object_ref),
                        now,
                        now,
                    ),
                )
        return self.get_target(target_id)

    def find_audit_field_target(self, audit_id: str, field_key: str) -> AnnotationTarget | None:
        audit_id = str(audit_id or "").strip()
        field_key = str(field_key or "").strip()
        if not audit_id or not field_key:
            return None
        with self.connection() as conn:
            row = conn.execute(
                """
                SELECT *
                FROM annotation_targets
                WHERE target_type = 'audit_field'
                AND audit_id = ?
                AND (field_key = ? OR header_name = ? OR field_label = ?)
                ORDER BY updated_at DESC
                LIMIT 1
                """,
                (audit_id, field_key, field_key, field_key),
            ).fetchone()
        return _target_from_row(row) if row else None

    def get_target(self, target_id: str) -> AnnotationTarget:
        with self.connection() as conn:
            row = conn.execute("SELECT * FROM annotation_targets WHERE id = ?", (target_id,)).fetchone()
        if row is None:
            raise KeyError(target_id)
        return _target_from_row(row)

    def assign_tag_to_target(self, tag_id: str, target_id: str, comment: str | None = None, *, sync_workbook: bool = True) -> TagAssignment:
        assignment_id: str | None = None
        now = utc_now()
        with self.connection() as conn:
            existing = conn.execute(
                "SELECT * FROM tag_assignments WHERE tag_id = ? AND target_id = ? AND archived_at IS NULL",
                (tag_id, target_id),
            ).fetchone()
            if existing:
                assignment_id = str(existing["id"])
                if comment is not None:
                    conn.execute("UPDATE tag_assignments SET comment = ?, updated_at = ? WHERE id = ?", (_none_empty(comment), now, assignment_id))
            else:
                assignment_id = _new_id("assignment")
                conn.execute(
                    """
                    INSERT INTO tag_assignments(id, tag_id, target_id, comment, created_at, updated_at)
                    VALUES(?, ?, ?, ?, ?, ?)
                    """,
                    (assignment_id, tag_id, target_id, _none_empty(comment), now, now),
                )
        self._log("annotation_tag_assigned", {"tag_id": tag_id, "target_id": target_id})
        if sync_workbook:
            self.sync_target_colors_to_workbook(target_id)
        return self.get_tag_assignment(assignment_id)

    def get_tag_assignment(self, assignment_id: str) -> TagAssignment:
        with self.connection() as conn:
            row = conn.execute("SELECT * FROM tag_assignments WHERE id = ?", (assignment_id,)).fetchone()
        if row is None:
            raise KeyError(assignment_id)
        return _assignment_from_row(row)

    def remove_tag_from_target(self, tag_id: str, target_id: str, *, sync_workbook: bool = True) -> None:
        now = utc_now()
        with self.connection() as conn:
            conn.execute(
                "UPDATE tag_assignments SET archived_at = ?, updated_at = ? WHERE tag_id = ? AND target_id = ? AND archived_at IS NULL",
                (now, now, tag_id, target_id),
            )
        self._log("annotation_tag_removed", {"tag_id": tag_id, "target_id": target_id})
        if sync_workbook:
            self.sync_target_colors_to_workbook(target_id)

    def archive_assignments(self, assignment_ids: Iterable[str], *, sync_workbook: bool = True) -> int:
        now = utc_now()
        assignment_ids = [str(value) for value in assignment_ids if value]
        if not assignment_ids:
            return 0
        with self.connection() as conn:
            targets = [
                str(row["target_id"])
                for row in conn.execute(
                    f"SELECT DISTINCT target_id FROM tag_assignments WHERE id IN ({','.join('?' for _ in assignment_ids)})",
                    assignment_ids,
                )
            ]
            conn.execute(
                f"UPDATE tag_assignments SET archived_at = ?, updated_at = ? WHERE id IN ({','.join('?' for _ in assignment_ids)}) AND archived_at IS NULL",
                [now, now, *assignment_ids],
            )
        if sync_workbook:
            for target_id in targets:
                self.sync_target_colors_to_workbook(target_id)
        return len(assignment_ids)

    def get_tags_for_target(self, target_id: str) -> list[dict[str, object]]:
        with self.connection() as conn:
            rows = conn.execute(
                """
                SELECT tags.*, tag_assignments.id AS assignment_id, tag_assignments.comment, tag_assignments.created_at AS assigned_at
                FROM tag_assignments
                JOIN tags ON tags.id = tag_assignments.tag_id
                WHERE tag_assignments.target_id = ? AND tag_assignments.archived_at IS NULL AND tags.is_archived = 0
                ORDER BY lower(tags.name)
                """,
                (target_id,),
            ).fetchall()
        return [dict(row) for row in rows]

    def get_notes_for_target(self, target_id: str) -> list[dict[str, object]]:
        with self.connection() as conn:
            rows = conn.execute(
                """
                SELECT notes.*
                FROM note_targets
                JOIN notes ON notes.id = note_targets.note_id
                WHERE note_targets.target_id = ?
                AND notes.archived_at IS NULL
                ORDER BY notes.updated_at DESC
                """,
                (target_id,),
            ).fetchall()
            return [self._enrich_note_row(conn, row) for row in rows]

    def get_targets_for_tag(self, tag_id: str) -> list[dict[str, object]]:
        with self.connection() as conn:
            rows = conn.execute(
                """
                SELECT annotation_targets.*, tag_assignments.id AS assignment_id, tag_assignments.comment, tag_assignments.updated_at AS assignment_updated_at
                FROM tag_assignments
                JOIN annotation_targets ON annotation_targets.id = tag_assignments.target_id
                WHERE tag_assignments.tag_id = ? AND tag_assignments.archived_at IS NULL
                ORDER BY annotation_targets.target_type, annotation_targets.target_label
                """,
                (tag_id,),
            ).fetchall()
        return [dict(row) for row in rows]

    def list_tag_assignments(
        self,
        query: str = "",
        *,
        tag_name: str | None = None,
        color_key: str | None = None,
        target_type: str | None = None,
        audit_id: str | None = None,
        machine_id: str | None = None,
        linked_note_id: str | None = None,
        include_archived: bool = False,
        sort_by: str = "updated",
    ) -> list[dict[str, object]]:
        clauses = ["tags.is_archived = 0"]
        params: list[Any] = []
        if not include_archived:
            clauses.append("tag_assignments.archived_at IS NULL")
        if query:
            like = f"%{query.strip()}%"
            clauses.append(
                """
                (
                    tags.name LIKE ?
                    OR coalesce(tag_assignments.comment, '') LIKE ?
                    OR coalesce(annotation_targets.target_label, '') LIKE ?
                    OR coalesce(annotation_targets.audit_id, '') LIKE ?
                    OR coalesce(annotation_targets.machine_id, '') LIKE ?
                    OR coalesce(annotation_targets.field_label, '') LIKE ?
                )
                """
            )
            params.extend([like, like, like, like, like, like])
        if tag_name and tag_name != "All":
            clauses.append("lower(tags.name) = lower(?)")
            params.append(tag_name)
        if color_key and color_key != "All":
            clauses.append("tags.color_key = ?")
            params.append(normalize_color_key(color_key))
        if target_type and target_type != "All":
            clauses.append("annotation_targets.target_type = ?")
            params.append(normalize_target_type(target_type))
        if audit_id:
            clauses.append("annotation_targets.audit_id = ?")
            params.append(audit_id)
        if machine_id:
            clauses.append("annotation_targets.machine_id = ?")
            params.append(machine_id)
        if linked_note_id:
            clauses.append(
                """
                EXISTS (
                    SELECT 1 FROM note_targets nt
                    WHERE nt.target_id = annotation_targets.id AND nt.note_id = ?
                )
                """
            )
            params.append(linked_note_id)
        where = f"WHERE {' AND '.join(clauses)}"
        order = ASSIGNMENT_SORTS.get(sort_by.strip().lower(), ASSIGNMENT_SORTS["updated"])
        with self.connection() as conn:
            rows = conn.execute(
                f"""
                SELECT
                    tag_assignments.id AS assignment_id,
                    tag_assignments.tag_id,
                    tag_assignments.target_id,
                    tag_assignments.comment,
                    tag_assignments.created_at,
                    tag_assignments.updated_at,
                    tag_assignments.archived_at,
                    tags.name AS tag_name,
                    tags.color_key,
                    tags.description,
                    annotation_targets.target_type,
                    annotation_targets.target_label,
                    annotation_targets.audit_id,
                    annotation_targets.machine_id,
                    annotation_targets.field_key,
                    annotation_targets.field_label,
                    annotation_targets.object_ref
                FROM tag_assignments
                JOIN tags ON tags.id = tag_assignments.tag_id
                JOIN annotation_targets ON annotation_targets.id = tag_assignments.target_id
                {where}
                ORDER BY {order}
                """,
                params,
            ).fetchall()
        return [dict(row) for row in rows]

    def link_note_to_target(self, note_id: str, target_id: str) -> None:
        with self.connection() as conn:
            self._link_note_to_target(conn, note_id, target_id)

    def unlink_note_from_target(self, note_id: str, target_id: str) -> None:
        with self.connection() as conn:
            conn.execute("DELETE FROM note_targets WHERE note_id = ? AND target_id = ?", (note_id, target_id))

    def link_note_to_tag(self, note_id: str, tag_id: str) -> None:
        with self.connection() as conn:
            self._link_note_to_tag(conn, note_id, tag_id)

    def unlink_note_from_tag(self, note_id: str, tag_id: str) -> None:
        with self.connection() as conn:
            conn.execute("DELETE FROM note_tags WHERE note_id = ? AND tag_id = ?", (note_id, tag_id))

    def attach_file(self, *, note_id: str | None = None, target_id: str | None = None, file_path: str | Path, display_name: str | None = None, description: str | None = None) -> str:
        with self.connection() as conn:
            return self._attach_file(conn, note_id=note_id, target_id=target_id, file_path=file_path, display_name=display_name, description=description)

    def highest_priority_color_for_target(self, target_id: str) -> str | None:
        highest = highest_priority_tag(self.get_tags_for_target(target_id))
        return str(highest["color_key"]) if highest else None

    def sync_target_colors_to_workbook(self, target_id: str) -> dict[str, object]:
        started = time.perf_counter()
        warnings: list[str] = []
        files_created: list[str] = []
        files_modified: list[str] = []
        try:
            target = self.get_target(target_id)
        except KeyError:
            return {"success": False, "warnings": [f"Annotation target not found: {target_id}"], "duration_seconds": 0.0}
        if target.target_type != "audit_field":
            return {"success": True, "warnings": ["Target is not an audit field; workbook sync skipped."], "duration_seconds": time.perf_counter() - started}
        audit_id = target.audit_id or ""
        header_name = target.header_name or target.field_key or target.field_label or ""
        sheet_name = target.sheet_name or "EOAT Inventory"
        workbook_path = Path(target.workbook_path) if target.workbook_path else resolve_project_paths(self.project_root).master_workbook
        if not audit_id:
            return {"success": False, "warnings": ["Audit field target has no Audit ID; workbook sync skipped."], "duration_seconds": time.perf_counter() - started}
        if not header_name:
            return {"success": False, "warnings": ["Audit field target has no field key/header; workbook sync skipped."], "duration_seconds": time.perf_counter() - started}
        if not workbook_path.exists():
            return {"success": False, "warnings": [f"Master workbook is missing: {workbook_path}"], "duration_seconds": time.perf_counter() - started}
        workbook = None
        try:
            backup = backup_file(workbook_path, workbook_path.parent / "_backups")
            files_created.append(str(backup))
            workbook = load_workbook(workbook_path)
            if sheet_name not in workbook.sheetnames:
                return {"success": False, "warnings": [f"Workbook sheet not found: {sheet_name}"], "files_created": files_created, "duration_seconds": time.perf_counter() - started}
            ws = workbook[sheet_name]
            headers = worksheet_headers(ws)
            audit_header = "Audit ID" if "Audit ID" in headers else "Last Audit ID" if "Last Audit ID" in headers else ""
            if not audit_header:
                return {"success": False, "warnings": ["Audit ID column is missing; workbook sync skipped."], "files_created": files_created, "duration_seconds": time.perf_counter() - started}
            column_header = _resolve_header(headers, header_name, target.field_label or "")
            if not column_header:
                return {
                    "success": False,
                    "warnings": [f"Workbook header not found for tagged field: {header_name}"],
                    "files_created": files_created,
                    "duration_seconds": time.perf_counter() - started,
                }
            audit_col = headers.index(audit_header) + 1
            target_row = None
            for row_number in range(2, ws.max_row + 1):
                if str(ws.cell(row=row_number, column=audit_col).value or "").strip() == audit_id:
                    target_row = row_number
                    break
            if target_row is None:
                return {
                    "success": False,
                    "warnings": [f"Audit ID not found for workbook tag sync: {audit_id}"],
                    "files_created": files_created,
                    "duration_seconds": time.perf_counter() - started,
                }
            target_col = headers.index(column_header) + 1
            cell = ws.cell(row=target_row, column=target_col)
            color_key = self.highest_priority_color_for_target(target_id)
            if color_key:
                cell.fill = PatternFill(fill_type="solid", fgColor=excel_fill_for_color(color_key))
            else:
                cell.fill = PatternFill(fill_type=None)
            cell_ref = cell.coordinate
            workbook.save(workbook_path)
            files_modified.append(str(workbook_path))
            with self.connection() as conn:
                conn.execute("UPDATE annotation_targets SET cached_cell_ref = ?, updated_at = ? WHERE id = ?", (cell_ref, utc_now(), target_id))
            return {
                "success": True,
                "applied_color_key": color_key,
                "cell_ref": cell_ref,
                "warnings": warnings,
                "files_created": files_created,
                "files_modified": files_modified,
                "duration_seconds": time.perf_counter() - started,
            }
        except Exception as exc:
            return {"success": False, "warnings": [f"Could not sync tag color to workbook: {exc}"], "files_created": files_created, "duration_seconds": time.perf_counter() - started}
        finally:
            if workbook is not None:
                workbook.close()

    def sync_tag_colors_to_workbook_for_audit(self, audit_id: str) -> dict[str, object]:
        target_ids = {
            str(assignment["target_id"])
            for assignment in self.list_tag_assignments(audit_id=audit_id, target_type="audit_field")
        }
        return self.sync_target_colors_to_workbook_batch(target_ids)

    def sync_target_colors_to_workbook_batch(self, target_ids: Iterable[str]) -> dict[str, object]:
        started = time.perf_counter()
        warnings: list[str] = []
        files_created: list[str] = []
        files_modified: list[str] = []
        unique_target_ids = [str(target_id) for target_id in dict.fromkeys(target_ids) if target_id]
        if not unique_target_ids:
            return {"success": True, "synced_count": 0, "warnings": [], "duration_seconds": time.perf_counter() - started}

        targets_by_workbook: dict[tuple[str, str], list[AnnotationTarget]] = {}
        for target_id in unique_target_ids:
            try:
                target = self.get_target(target_id)
            except KeyError:
                warnings.append(f"Annotation target not found: {target_id}")
                continue
            if target.target_type != "audit_field":
                warnings.append(f"Target is not an audit field; workbook sync skipped: {target_id}")
                continue
            workbook_path = Path(target.workbook_path) if target.workbook_path else resolve_project_paths(self.project_root).master_workbook
            sheet_name = target.sheet_name or "EOAT Inventory"
            targets_by_workbook.setdefault((str(workbook_path), sheet_name), []).append(target)

        synced_count = 0
        cached_updates: list[tuple[str, str]] = []
        for (workbook_path_text, sheet_name), targets in targets_by_workbook.items():
            workbook_path = Path(workbook_path_text)
            if not workbook_path.exists():
                warnings.append(f"Master workbook is missing: {workbook_path}")
                continue
            workbook = None
            changed = False
            try:
                backup = backup_file(workbook_path, workbook_path.parent / "_backups")
                files_created.append(str(backup))
                workbook = load_workbook(workbook_path)
                if sheet_name not in workbook.sheetnames:
                    warnings.append(f"Workbook sheet not found: {sheet_name}")
                    continue
                ws = workbook[sheet_name]
                headers = worksheet_headers(ws)
                audit_header = "Audit ID" if "Audit ID" in headers else "Last Audit ID" if "Last Audit ID" in headers else ""
                if not audit_header:
                    warnings.append("Audit ID column is missing; workbook sync skipped.")
                    continue
                audit_col = headers.index(audit_header) + 1
                row_by_audit = {
                    str(ws.cell(row=row_number, column=audit_col).value or "").strip(): row_number
                    for row_number in range(2, ws.max_row + 1)
                }
                for target in targets:
                    audit_id = target.audit_id or ""
                    header_name = target.header_name or target.field_key or target.field_label or ""
                    if not audit_id:
                        warnings.append("Audit field target has no Audit ID; workbook sync skipped.")
                        continue
                    if not header_name:
                        warnings.append("Audit field target has no field key/header; workbook sync skipped.")
                        continue
                    column_header = _resolve_header(headers, header_name, target.field_label or "")
                    if not column_header:
                        warnings.append(f"Workbook header not found for tagged field: {header_name}")
                        continue
                    target_row = row_by_audit.get(audit_id)
                    if target_row is None:
                        warnings.append(f"Audit ID not found for workbook tag sync: {audit_id}")
                        continue
                    target_col = headers.index(column_header) + 1
                    cell = ws.cell(row=target_row, column=target_col)
                    color_key = self.highest_priority_color_for_target(target.id)
                    if color_key:
                        cell.fill = PatternFill(fill_type="solid", fgColor=excel_fill_for_color(color_key))
                    else:
                        cell.fill = PatternFill(fill_type=None)
                    cached_updates.append((cell.coordinate, target.id))
                    synced_count += 1
                    changed = True
                if changed:
                    workbook.save(workbook_path)
                    files_modified.append(str(workbook_path))
            except Exception as exc:
                warnings.append(f"Could not sync tag colors to workbook: {exc}")
            finally:
                if workbook is not None:
                    workbook.close()

        if cached_updates:
            with self.connection() as conn:
                now = utc_now()
                conn.executemany(
                    "UPDATE annotation_targets SET cached_cell_ref = ?, updated_at = ? WHERE id = ?",
                    [(cell_ref, now, target_id) for cell_ref, target_id in cached_updates],
                )
        return {
            "success": not warnings,
            "synced_count": synced_count,
            "target_count": len(unique_target_ids),
            "warnings": warnings,
            "files_created": sorted(set(files_created)),
            "files_modified": sorted(set(files_modified)),
            "duration_seconds": time.perf_counter() - started,
        }

    def sync_all_tag_colors_to_workbook(self) -> dict[str, object]:
        with self.connection() as conn:
            target_ids = [
                str(row["id"])
                for row in conn.execute(
                    "SELECT id FROM annotation_targets WHERE target_type = 'audit_field' ORDER BY audit_id, field_key"
                ).fetchall()
            ]
        return self.sync_target_colors_to_workbook_batch(target_ids)

    def get_open_items_summary(self, *, today: date | None = None) -> dict[str, int]:
        today = today or date.today()
        due_by = (today + timedelta(days=7)).isoformat()
        with self.connection() as conn:
            open_note_clause = "(archived_at IS NULL AND (status IS NULL OR lower(status) NOT IN ('resolved', 'archived')))"
            critical = conn.execute(f"SELECT COUNT(*) AS count FROM notes WHERE {open_note_clause} AND importance = 'Critical'").fetchone()["count"]
            important = conn.execute(f"SELECT COUNT(*) AS count FROM notes WHERE {open_note_clause} AND importance = 'Important'").fetchone()["count"]
            followups = conn.execute(
                f"SELECT COUNT(*) AS count FROM notes WHERE {open_note_clause} AND follow_up_date IS NOT NULL AND follow_up_date <= ?",
                (due_by,),
            ).fetchone()["count"]

            def tag_count(name: str, target_type: str | None = None) -> int:
                params: list[Any] = [name]
                target_clause = ""
                if target_type:
                    target_clause = "AND annotation_targets.target_type = ?"
                    params.append(target_type)
                return int(
                    conn.execute(
                        f"""
                        SELECT COUNT(*) AS count
                        FROM tag_assignments
                        JOIN tags ON tags.id = tag_assignments.tag_id
                        JOIN annotation_targets ON annotation_targets.id = tag_assignments.target_id
                        WHERE tag_assignments.archived_at IS NULL
                        AND tags.is_archived = 0
                        AND lower(tags.name) = lower(?)
                        {target_clause}
                        """,
                        params,
                    ).fetchone()["count"]
                )

            return {
                "critical_notes": int(critical),
                "important_notes": int(important),
                "fields_needing_review": tag_count("Needs Review", "audit_field"),
                "data_conflicts": tag_count("Data Conflict"),
                "missing_evidence": tag_count("Missing Evidence"),
                "compatibility_concerns": tag_count("Compatibility Concern"),
                "documentation_gaps": tag_count("Documentation Gap"),
                "followups_due_soon": int(followups),
            }

    def get_suggested_annotations(self, audit_entry: dict[str, Any], *, include_ignored: bool = False) -> list[dict[str, object]]:
        suggestions: list[dict[str, object]] = []
        for suggestion in suggested_annotations_for_audit(audit_entry):
            row = dict(suggestion.__dict__)
            ignored = self.is_suggested_annotation_ignored(row)
            row["ignored"] = ignored
            row["existing_status"] = self._suggestion_existing_status(row)
            if include_ignored or not ignored:
                suggestions.append(row)
        return suggestions

    def is_suggested_annotation_ignored(self, suggestion: dict[str, object]) -> bool:
        suggestion_id = str(suggestion.get("suggestion_id") or "").strip()
        if not suggestion_id:
            return False
        with self.connection() as conn:
            row = conn.execute(
                "SELECT 1 FROM annotation_suggestion_ignores WHERE suggestion_id = ?",
                (suggestion_id,),
            ).fetchone()
        return row is not None

    def ignore_suggested_annotation(self, suggestion: dict[str, object]) -> str:
        suggestion_id = str(suggestion.get("suggestion_id") or "").strip()
        if not suggestion_id:
            raise ValueError("Suggestion ID is required.")
        ignore_id = f"ignored_{suggestion_id}"
        with self.connection() as conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO annotation_suggestion_ignores(
                    id, suggestion_id, audit_id, field_key, tag_name, reason, data_fingerprint, ignored_at
                )
                VALUES(?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    ignore_id,
                    suggestion_id,
                    _none_empty(suggestion.get("audit_id")),
                    _none_empty(suggestion.get("field_key")),
                    _none_empty(suggestion.get("tag_name")),
                    _none_empty(suggestion.get("reason")),
                    _none_empty(suggestion.get("data_fingerprint")),
                    utc_now(),
                ),
            )
        self._log("annotation_suggestion_ignored", {"suggestion_id": suggestion_id})
        return ignore_id

    def ignore_suggested_annotations(self, suggestions: Iterable[dict[str, object]]) -> int:
        count = 0
        for suggestion in suggestions:
            self.ignore_suggested_annotation(suggestion)
            count += 1
        return count

    def apply_suggested_annotation(self, suggestion: dict[str, object]) -> TagAssignment:
        tag_name = str(suggestion.get("tag_name") or "")
        tag = self.get_tag_by_name(tag_name) or self.create_tag(tag_name, "yellow")
        target = self.create_or_get_target(
            str(suggestion.get("target_type") or "project_item"),
            audit_id=str(suggestion.get("audit_id") or ""),
            machine_id=str(suggestion.get("machine_id") or ""),
            field_key=str(suggestion.get("field_key") or ""),
            field_label=str(suggestion.get("field_key") or ""),
            sheet_name="EOAT Inventory",
            header_name=str(suggestion.get("field_key") or ""),
            target_label=f"{suggestion.get('audit_id') or ''} / {suggestion.get('field_key') or ''}".strip(" /"),
        )
        return self.assign_tag_to_target(tag.id, target.id, comment=str(suggestion.get("suggested_comment") or suggestion.get("reason") or "Suggested annotation"))

    def apply_suggested_annotations(self, suggestions: Iterable[dict[str, object]]) -> list[TagAssignment]:
        return [self.apply_suggested_annotation(suggestion) for suggestion in suggestions]

    def _suggestion_existing_status(self, suggestion: dict[str, object]) -> str:
        target_type = str(suggestion.get("target_type") or "project_item")
        field_key = str(suggestion.get("field_key") or "")
        audit_id = str(suggestion.get("audit_id") or "")
        machine_id = str(suggestion.get("machine_id") or "")
        target_id = target_id_for(
            target_type=target_type,
            audit_id=audit_id,
            machine_id=machine_id,
            field_key=field_key,
            object_ref="",
        )
        tags = self.get_tags_for_target(target_id)
        notes = self.get_notes_for_target(target_id)
        tag_name = str(suggestion.get("tag_name") or "")
        parts: list[str] = []
        if any(str(tag.get("name") or "").casefold() == tag_name.casefold() for tag in tags):
            parts.append("Tag already applied")
        elif tags:
            parts.append(f"{len(tags)} existing tag(s)")
        if notes:
            parts.append(f"{len(notes)} linked note(s)")
        return "; ".join(parts) if parts else "New"

    def export_notes_markdown(self, notes: Iterable[dict[str, object]] | None = None) -> Path:
        return export_notes_markdown(self.project_root, notes if notes is not None else self.search_notes())

    def export_notes_excel(self, notes: Iterable[dict[str, object]] | None = None) -> Path:
        return export_notes_excel(self.project_root, notes if notes is not None else self.search_notes())

    def export_tags_markdown(self, assignments: Iterable[dict[str, object]] | None = None) -> Path:
        return export_tags_markdown(self.project_root, assignments if assignments is not None else self.list_tag_assignments())

    def export_tags_excel(self, assignments: Iterable[dict[str, object]] | None = None) -> Path:
        return export_tags_excel(self.project_root, assignments if assignments is not None else self.list_tag_assignments())

    def _enrich_note_row(self, conn: sqlite3.Connection, row: sqlite3.Row) -> dict[str, object]:
        note = dict(row)
        note["tags"] = [
            dict(tag_row)
            for tag_row in conn.execute(
                """
                SELECT tags.*
                FROM note_tags
                JOIN tags ON tags.id = note_tags.tag_id
                WHERE note_tags.note_id = ? AND tags.is_archived = 0
                ORDER BY lower(tags.name)
                """,
                (row["id"],),
            ).fetchall()
        ]
        note["targets"] = [
            dict(target_row)
            for target_row in conn.execute(
                """
                SELECT annotation_targets.*
                FROM note_targets
                JOIN annotation_targets ON annotation_targets.id = note_targets.target_id
                WHERE note_targets.note_id = ?
                ORDER BY annotation_targets.target_type, annotation_targets.target_label
                """,
                (row["id"],),
            ).fetchall()
        ]
        note["attachments"] = [dict(attachment_row) for attachment_row in conn.execute("SELECT * FROM attachments WHERE note_id = ? ORDER BY created_at", (row["id"],)).fetchall()]
        return note

    def _link_note_to_target(self, conn: sqlite3.Connection, note_id: str, target_id: str) -> None:
        existing = conn.execute("SELECT id FROM note_targets WHERE note_id = ? AND target_id = ?", (note_id, target_id)).fetchone()
        if existing:
            return
        conn.execute("INSERT INTO note_targets(id, note_id, target_id, created_at) VALUES(?, ?, ?, ?)", (_new_id("note_target"), note_id, target_id, utc_now()))

    def _link_note_to_tag(self, conn: sqlite3.Connection, note_id: str, tag_id: str) -> None:
        existing = conn.execute("SELECT id FROM note_tags WHERE note_id = ? AND tag_id = ?", (note_id, tag_id)).fetchone()
        if existing:
            return
        conn.execute("INSERT INTO note_tags(id, note_id, tag_id, created_at) VALUES(?, ?, ?, ?)", (_new_id("note_tag"), note_id, tag_id, utc_now()))

    def _attach_file(
        self,
        conn: sqlite3.Connection,
        *,
        note_id: str | None,
        target_id: str | None,
        file_path: str | Path,
        display_name: str | None = None,
        description: str | None = None,
    ) -> str:
        attachment_id = _new_id("attachment")
        path = Path(file_path)
        conn.execute(
            """
            INSERT INTO attachments(id, note_id, target_id, file_path, display_name, description, created_at)
            VALUES(?, ?, ?, ?, ?, ?, ?)
            """,
            (attachment_id, note_id, target_id, str(path), display_name or path.name, _none_empty(description), utc_now()),
        )
        return attachment_id

    def _log(self, event_name: str, payload: dict[str, Any]) -> None:
        log_activity_event(self.project_root, event_name, payload)


def _new_id(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex}"


def _none_empty(value: Any) -> Any:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _validate_importance(value: str) -> str:
    text = str(value or "Neutral").strip()
    if text not in IMPORTANCE_VALUES:
        raise ValueError(f"Unsupported note importance: {value}")
    return text


def _validate_status(value: Any) -> str | None:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip()
    if text not in STATUS_VALUES:
        raise ValueError(f"Unsupported note status: {value}")
    return text


def _note_from_row(row: sqlite3.Row) -> Note:
    return Note(
        id=str(row["id"]),
        subject=str(row["subject"]),
        body_markdown=str(row["body_markdown"] or ""),
        importance=str(row["importance"] or "Neutral"),
        status=row["status"],
        collection=row["collection"],
        note_type=row["note_type"],
        follow_up_date=row["follow_up_date"],
        created_at=str(row["created_at"]),
        updated_at=str(row["updated_at"]),
        archived_at=row["archived_at"],
    )


def _tag_from_row(row: sqlite3.Row) -> Tag:
    return Tag(
        id=str(row["id"]),
        name=str(row["name"]),
        color_key=str(row["color_key"]),
        description=row["description"],
        is_default=bool(row["is_default"]),
        is_archived=bool(row["is_archived"]),
        created_at=str(row["created_at"]),
        updated_at=str(row["updated_at"]),
    )


def _target_from_row(row: sqlite3.Row) -> AnnotationTarget:
    return AnnotationTarget(
        id=str(row["id"]),
        target_type=str(row["target_type"]),
        target_label=row["target_label"],
        audit_id=row["audit_id"],
        machine_id=row["machine_id"],
        field_key=row["field_key"],
        field_label=row["field_label"],
        sheet_name=row["sheet_name"],
        header_name=row["header_name"],
        workbook_path=row["workbook_path"],
        cached_cell_ref=row["cached_cell_ref"],
        object_ref=row["object_ref"],
        created_at=str(row["created_at"]),
        updated_at=str(row["updated_at"]),
    )


def _assignment_from_row(row: sqlite3.Row) -> TagAssignment:
    return TagAssignment(
        id=str(row["id"]),
        tag_id=str(row["tag_id"]),
        target_id=str(row["target_id"]),
        comment=row["comment"],
        created_at=str(row["created_at"]),
        updated_at=str(row["updated_at"]),
        archived_at=row["archived_at"],
    )


def _resolve_header(headers: list[str], preferred: str, alternate: str = "") -> str:
    for candidate in [preferred, alternate]:
        if candidate in headers:
            return candidate
    preferred_norm = _norm(preferred)
    alternate_norm = _norm(alternate)
    for header in headers:
        if _norm(header) in {preferred_norm, alternate_norm}:
            return header
    return ""


def _norm(value: str) -> str:
    return "".join(ch for ch in str(value or "").casefold() if ch.isalnum())
