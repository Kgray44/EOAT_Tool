"""
Create the Nolato EOAT Standardization project workspace.

This script is safe to run more than once:
- Existing folders are left in place.
- Existing text files are backed up before replacement.
- An existing workbook is not overwritten; a timestamped backup is made first.
"""

from __future__ import annotations

import json
import re
import argparse
from datetime import datetime
from pathlib import Path
from typing import Iterable

from core.audit_constants import (
    CYLINDER_COUNT_FIELD,
    CYLINDER_TYPE_FIELD,
    CYLINDER_TYPE_VALUES,
    IGNORED_EMPTY_FIELDS_AT_OVERRIDE_FIELD,
    MANUAL_COMPLETION_OVERRIDE_FIELD,
    MANUAL_COMPLETION_OVERRIDE_TIMESTAMP_FIELD,
    MANUAL_COMPLETION_OVERRIDE_USER_FIELD,
)
from core.gripper_fields import CUP_COUNT_FIELD, GRIPPER_COUNT_FIELD, GRIPPER_TYPE_FIELD, GRIPPER_TYPE_VALUES

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required to create the master workbook.\n"
        "Install requirements with: python -m pip install -r requirements.txt"
    ) from exc


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR / "EOAT_Standardization_Project"
WORKBOOK_PATH = (
    PROJECT_ROOT
    / "01_EOAT_Audit"
    / "EOAT_Audit_Database"
    / "EOAT_Master_Tracker.xlsx"
)
ADMIN_DIR = PROJECT_ROOT / "00_Project_Admin"
SCHEDULE_PATH = ADMIN_DIR / "project_schedule_week1.json"
PROGRESS_PATH = ADMIN_DIR / "task_progress_week1.json"
HELP_DOCUMENTS_DIR = SCRIPT_DIR / "Project_Help_Documents"


FOLDERS = [
    "00_Project_Admin/Weekly_Status_Reports",
    "00_Project_Admin/Daily_Status_Reports",
    "00_Project_Admin/Validation_Reports",
    "00_Project_Admin/Activity_Logs",
    "00_Project_Admin/Meeting_Notes",
    "00_Project_Admin/Stakeholder_Contacts",
    "00_Project_Admin/reference_data",
    "01_EOAT_Audit/EOAT_Audit_Database",
    "01_EOAT_Audit/Cell_Photos/Overall",
    "01_EOAT_Audit/Cell_Photos/Vacuum_Cups_Grippers",
    "01_EOAT_Audit/Cell_Photos/Tubing_Routing",
    "01_EOAT_Audit/Cell_Photos/Sensors",
    "01_EOAT_Audit/Cell_Photos/Quick_Disconnects",
    "01_EOAT_Audit/Cell_Photos/Mounting_Hardware",
    "01_EOAT_Audit/Cell_Photos/Cable_Management",
    "01_EOAT_Audit/Cell_Photos/Wear_Damage",
    "01_EOAT_Audit/Cell_Photos/Incoming",
    "01_EOAT_Audit/Raw_Notes",
    "02_KPI_Data/Downtime_Data",
    "02_KPI_Data/Scrap_Data",
    "02_KPI_Data/Cycle_Time_Data",
    "02_KPI_Data/Maintenance_Data",
    "03_Standards/EOAT_Design_Guideline_Draft",
    "03_Standards/PM_Checklist_Draft",
    "03_Standards/BOM_Template_Draft",
    "04_FMEA",
    "05_Pilot_Project/Candidate_Cells",
    "05_Pilot_Project/Before_After_Data",
    "05_Pilot_Project/Pilot_Implementation_Notes",
    "06_Final_Handoff/Final_Report",
    "06_Final_Handoff/Presentation",
    "06_Final_Handoff/Training_Materials",
]


def configure_project_root(project_root: Path) -> None:
    """Update module paths so setup can target a selected project root."""
    global PROJECT_ROOT, WORKBOOK_PATH, ADMIN_DIR, SCHEDULE_PATH, PROGRESS_PATH
    PROJECT_ROOT = project_root
    WORKBOOK_PATH = (
        PROJECT_ROOT
        / "01_EOAT_Audit"
        / "EOAT_Audit_Database"
        / "EOAT_Master_Tracker.xlsx"
    )
    ADMIN_DIR = PROJECT_ROOT / "00_Project_Admin"
    SCHEDULE_PATH = ADMIN_DIR / "project_schedule_week1.json"
    PROGRESS_PATH = ADMIN_DIR / "task_progress_week1.json"


YES_NO_UNKNOWN = ["Yes", "No", "Unknown / Not Checked"]
YES_NO_UNKNOWN_NA = ["Yes", "No", "Unknown / Not Checked", "Not Applicable"]
YES_NO_PARTIAL_UNKNOWN = ["Yes", "No", "Partial", "Unknown / Not Checked"]
PRIORITY_VALUES = ["Low", "Medium", "High", "Critical"]


WEEK1_DEFAULT_DAYS = {
    "1": [
        "Review EOAT project scope and deliverables",
        "Create project folder/document structure",
        "Start EOAT audit database",
        "Identify initial stakeholder questions",
        "Begin list of target robot cells",
    ],
    "2": [
        "Finalize audit template",
        "Begin target cell list",
        "Start first walkthrough/audit if approved",
        "Decide photo naming system",
        "Confirm initial audit priorities with mentor or supervisor",
    ],
    "3": [
        "Audit 2-4 accessible EOATs in detail",
        "Capture overall EOAT photos",
        "Capture vacuum cup/gripper photos",
        "Capture tubing routing photos",
        "Capture sensor and quick disconnect photos",
        "Interview operator or technician for each audited cell",
        "Clean up notes and link photos in the workbook",
    ],
    "4": [
        "Improve the audit template based on real floor use",
        "Audit another 3-5 EOATs",
        "Include at least one good example cell if possible",
        "Start Issue Log tab",
        "Group early issues into categories",
        "Flag possible pilot candidate cells",
    ],
    "5": [
        "Clean and normalize the audit database",
        "Organize photo folders",
        "Create Week 1 summary",
        "Review progress with mentor or supervisor",
        "Confirm Week 2 priorities",
        "Identify data gaps and blockers",
    ],
}


WEEK1_SCHEDULE = {
    "week": 1,
    "source": "Project_Help_Documents weekly schedule files, with fallback defaults from setup_eoat_project.py",
    "source_files": [],
    "days": WEEK1_DEFAULT_DAYS,
}


SHEETS: dict[str, list[str]] = {
    "EOAT Inventory": [
        "Audit ID",
        "Audit Date",
        "Auditor",
        "Plant/Area",
        "Press/Machine #",
        "Tool #",
        "Robot Type",
        "Robot Model/Controller",
        "Part Family",
        "Part Name/Description",
        "Cleanroom/Non-Cleanroom",
        "EOAT Type",
        "EOAT Moves",
        "Connection Type",
        "Number of Parts Picked",
        CYLINDER_COUNT_FIELD,
        CYLINDER_TYPE_FIELD,
        GRIPPER_COUNT_FIELD,
        GRIPPER_TYPE_FIELD,
        "Gripper Model",
        "Gripper Size",
        CUP_COUNT_FIELD,
        "Cup Type/Material",
        "Cup Diameter/Size",
        "Vacuum Generator Type",
        "EOAT Vacuum Circuits",
        "EOAT Pressure Circuits",
        "EOAT Interchangeable Circuits",
        "Sensors Present?",
        "Sensor Type",
        "Sensor Brand/Model",
        "Vacuum Confirmation Present?",
        "Part-Present Detection Present?",
        "Quick Disconnects Present?",
        "Pneumatic Quick Disconnect Type",
        "Electrical Quick Disconnect Type",
        "Tubing Condition",
        "Tubing Routing Notes",
        "Cable Management Condition",
        "Mounting Hardware Condition",
        "EOAT Alignment Condition",
        "Fastener/Locking Hardware Present?",
        "Estimated EOAT Weight",
        "Known Issues",
        "Drop/Mis-Pick History",
        "Maintenance Frequency",
        "Cycle Time Concern?",
        "Scrap/Quality Concern?",
        "Changeover Difficulty",
        "Spare Parts Identified?",
        "Drawing/CAD Available?",
        "BOM Available?",
        "Process Binder Complete?",
        "Photos Taken?",
        "Photo Folder/Link",
        "Status",
        "Priority",
        "Pilot Candidate?",
        "Follow-Up Needed",
        "Notes",
        MANUAL_COMPLETION_OVERRIDE_FIELD,
        MANUAL_COMPLETION_OVERRIDE_TIMESTAMP_FIELD,
        MANUAL_COMPLETION_OVERRIDE_USER_FIELD,
        IGNORED_EMPTY_FIELDS_AT_OVERRIDE_FIELD,
        "Entry Type",
        "Source Audit ID",
        "Compatibility Source",
    ],
    "Issue Log": [
        "Issue ID",
        "Date Found",
        "Plant/Area",
        "Press/Machine #",
        "Robot Type",
        "EOAT Type",
        "Issue Category",
        "Issue Description",
        "Suspected Cause",
        "Evidence/Observation",
        "Impact",
        "Severity",
        "Frequency",
        "Detectability",
        "Temporary Fix",
        "Recommended Permanent Fix",
        "Assigned To",
        "Status",
        "Follow-Up Date",
        "Notes",
    ],
    "KPI Baseline": [
        "KPI ID",
        "Date",
        "Plant/Area",
        "Press/Machine #",
        "Tool #",
        "Part Family",
        "EOAT Type",
        "Downtime Minutes",
        "EOAT-Related Downtime?",
        "Part Drops",
        "Mis-Picks",
        "Scrap Quantity",
        "Scrap Reason",
        "Cycle Time",
        "Maintenance Event Count",
        "Maintenance Notes",
        "Data Source",
        "Notes",
    ],
    "Interview Notes": [
        "Interview ID",
        "Date",
        "Person Interviewed",
        "Role/Department",
        "Shift",
        "Plant/Area",
        "Press/Machine #",
        "Main Question/Topic",
        "Notes",
        "Known EOAT Issues Mentioned",
        "Suggested Improvements",
        "Follow-Up Needed",
        "Follow-Up Owner",
    ],
    "Pilot Candidates": [
        "Candidate ID",
        "Date Added",
        "Plant/Area",
        "Press/Machine #",
        "Robot Type",
        "Tool #",
        "Part Family",
        "EOAT Type",
        "Main Problem",
        "Evidence",
        "Estimated Impact",
        "Ease of Implementation",
        "Safety/Quality Risk",
        "Required Parts/Resources",
        "Expected KPI Improvement",
        "Recommended Action",
        "Approval Status",
        "Notes",
    ],
    "FMEA Draft": [
        "FMEA ID",
        "Plant/Area",
        "Press/Machine #",
        "EOAT Function",
        "Failure Mode",
        "Failure Effect",
        "Potential Cause",
        "Current Controls",
        "Severity",
        "Frequency",
        "Detectability",
        "RPN",
        "Recommended Action",
        "Owner",
        "Target Completion Date",
        "Status",
        "Notes",
    ],
    "Action Items": [
        "Action ID",
        "Date Added",
        "Action Item",
        "Related Cell/Press",
        "Owner",
        "Priority",
        "Due Date",
        "Status",
        "Completion Date",
        "Notes",
    ],
    "Photo Index": [
        "Photo ID",
        "Date Taken",
        "Plant/Area",
        "Press/Machine #",
        "EOAT Area Shown",
        "Photo Filename",
        "Folder Path",
        "Description",
        "Related Audit ID",
        "Related Issue ID",
        "Notes",
    ],
}


VALIDATIONS: dict[str, dict[str, list[str]]] = {
    "EOAT Inventory": {
        "Plant/Area": ["Plant 4", "Cleanroom"],
        "Cleanroom/Non-Cleanroom": ["Cleanroom", "Non-Cleanroom", "Whiteroom", "Unknown / Not Checked", "N/A"],
        "EOAT Type": ["Vacuum", "Mechanical / Gripper", "Hybrid", "Unknown / Needs Review", "Miscellaneous", "N/A"],
        "EOAT Moves": ["Part", "Sprue", "Both"],
        "Entry Type": ["Audited", "Compatible"],
        "Connection Type": ["ATI", "DoveTail", "Direct Mount", "Lever Lock", "N/A"],
        CYLINDER_TYPE_FIELD: [*CYLINDER_TYPE_VALUES, "N/A"],
        GRIPPER_TYPE_FIELD: [*GRIPPER_TYPE_VALUES, "N/A"],
        MANUAL_COMPLETION_OVERRIDE_FIELD: ["Yes", "No", "N/A"],
        "Sensors Present?": YES_NO_UNKNOWN,
        "Vacuum Confirmation Present?": YES_NO_UNKNOWN_NA,
        "Part-Present Detection Present?": YES_NO_UNKNOWN_NA,
        "Quick Disconnects Present?": YES_NO_PARTIAL_UNKNOWN,
        "Fastener/Locking Hardware Present?": YES_NO_PARTIAL_UNKNOWN,
        "Cycle Time Concern?": YES_NO_UNKNOWN,
        "Scrap/Quality Concern?": YES_NO_UNKNOWN,
        "Spare Parts Identified?": YES_NO_PARTIAL_UNKNOWN,
        "Drawing/CAD Available?": YES_NO_UNKNOWN,
        "BOM Available?": YES_NO_UNKNOWN,
        "Process Binder Complete?": YES_NO_PARTIAL_UNKNOWN,
        "Photos Taken?": ["Yes", "No"],
        "Follow-Up Needed": ["Yes", "No"],
        "Tubing Condition": ["OK", "Worn", "Damaged", "Poor Routing", "Needs Follow-Up", "Unknown / Not Checked"],
        "Cable Management Condition": ["OK", "Loose", "Damaged", "Poor Routing", "Needs Follow-Up", "Unknown / Not Checked"],
        "Mounting Hardware Condition": ["OK", "Loose", "Missing Hardware", "Damaged", "Needs Follow-Up", "Unknown / Not Checked"],
        "EOAT Alignment Condition": ["OK", "Slightly Off", "Misaligned", "Needs Follow-Up", "Unknown / Not Checked"],
        "Changeover Difficulty": ["Low", "Medium", "High", "Unknown / Not Checked"],
        "Status": ["Not Started", "In Progress", "Complete", "Needs Follow-Up", "Blocked"],
        "Priority": PRIORITY_VALUES,
        "Pilot Candidate?": ["Yes", "No", "Maybe"],
    },
    "Issue Log": {
        "EOAT Type": ["Vacuum", "Mechanical gripper", "Hybrid", "Custom/other", "Unknown", "Miscellaneous"],
        "Issue Category": [
            "Vacuum loss",
            "Part drop",
            "Mis-pick",
            "Sensor issue",
            "Tubing issue",
            "Cable routing issue",
            "Mechanical wear",
            "Alignment issue",
            "Quick disconnect issue",
            "Fastener/hardware issue",
            "Documentation missing",
            "Safety concern",
            "Other",
        ],
        "Severity": ["1 - Very Low", "2 - Low", "3 - Medium", "4 - High", "5 - Critical"],
        "Frequency": ["1 - Very Low", "2 - Low", "3 - Medium", "4 - High", "5 - Critical"],
        "Detectability": ["1 - Very Low", "2 - Low", "3 - Medium", "4 - High", "5 - Critical"],
        "Status": ["Open", "Investigating", "Waiting on parts", "Waiting on approval", "In progress", "Resolved", "Closed"],
    },
    "KPI Baseline": {
        "EOAT Type": ["Vacuum", "Mechanical gripper", "Hybrid", "Custom/other", "Unknown", "Miscellaneous"],
        "EOAT-Related Downtime?": YES_NO_UNKNOWN,
    },
    "Interview Notes": {
        "Follow-Up Needed": YES_NO_UNKNOWN,
    },
    "Pilot Candidates": {
        "EOAT Type": ["Vacuum", "Mechanical gripper", "Hybrid", "Custom/other", "Unknown", "Miscellaneous"],
        "Approval Status": ["Not submitted", "Under review", "Approved", "Rejected", "Deferred"],
    },
    "FMEA Draft": {
        "Severity": ["1 - Very Low", "2 - Low", "3 - Medium", "4 - High", "5 - Critical"],
        "Frequency": ["1 - Very Low", "2 - Low", "3 - Medium", "4 - High", "5 - Critical"],
        "Detectability": ["1 - Very Low", "2 - Low", "3 - Medium", "4 - High", "5 - Critical"],
        "Status": ["Open", "In progress", "Complete", "Deferred"],
    },
    "Action Items": {
        "Priority": PRIORITY_VALUES,
        "Status": ["Open", "In progress", "Waiting", "Complete", "Deferred"],
    },
    "Photo Index": {
        "EOAT Area Shown": [
            "Overall",
            "Vacuum cups/grippers",
            "Tubing routing",
            "Sensors",
            "Quick disconnects",
            "Mounting hardware",
            "Cable management",
            "Wear/damage",
            "Other",
        ],
    },
}

WHOLE_NUMBER_VALIDATIONS: dict[str, list[str]] = {
    "EOAT Inventory": [
        "Number of Parts Picked",
        CYLINDER_COUNT_FIELD,
        CUP_COUNT_FIELD,
        GRIPPER_COUNT_FIELD,
        "EOAT Vacuum Circuits",
        "EOAT Pressure Circuits",
        "EOAT Interchangeable Circuits",
    ],
}


README_TEXT = """# EOAT Standardization Project

Local workspace for the Nolato Vermont Summer 2026 EOAT Standardization & Optimization internship project.

## Folder Guide

- `00_Project_Admin/` - project management material, daily/weekly updates, meeting notes, and stakeholder contact notes.
- `01_EOAT_Audit/` - EOAT inventory database, audit notes, and organized cell photos.
- `02_KPI_Data/` - downtime, scrap, cycle time, and maintenance data used for baselines and improvement tracking.
- `03_Standards/` - draft standard design guideline, PM checklist, and BOM template work.
- `04_FMEA/` - FMEA-lite risk assessment files and supporting notes.
- `05_Pilot_Project/` - candidate cells, before/after data, and implementation notes for the pilot optimization.
- `06_Final_Handoff/` - final report, presentation, and training materials.

## Main Workbook

The master tracker is located at:

`01_EOAT_Audit/EOAT_Audit_Database/EOAT_Master_Tracker.xlsx`

Use it as the central place for EOAT inventory, issue tracking, KPI baseline data, pilot candidates, FMEA draft items, action items, and photo indexing.
"""


USAGE_TEXT = """# EOAT Project Toolkit Usage

These tools are local-only and are intended for Windows work computers.

## Install Requirements

From this folder, run:

```powershell
python -m pip install -r requirements.txt
```

If `python-docx` is not installed, the daily report tool still creates Markdown files.

## Create or Refresh the Project Setup

```powershell
python setup_eoat_project.py
```

The setup script creates folders, `README.md`, schedule/progress JSON files for discovered weekly schedules, and the master Excel workbook. Existing workbook/text files are backed up before replacement.

Weekly schedule source files belong in `Project_Help_Documents`. Name them with the week number and schedule, for example:

```text
Week 1 Eoat Day By Day Schedule For Summarizer.pdf
Week 2 Eoat Day By Day Schedule For Summarizer.pdf
Week 3 Eoat Day By Day Schedule For Summarizer.pdf
```

When a matching weekly schedule file exists, the tool creates or uses:

```text
EOAT_Standardization_Project\\00_Project_Admin\\project_schedule_weekN.json
EOAT_Standardization_Project\\00_Project_Admin\\task_progress_weekN.json
```

## Optional: Initialize Git

Git is recommended because it gives the daily summary tool better change tracking.

```powershell
cd EOAT_Standardization_Project
git init
git add .
git commit -m "Initial EOAT project setup"
cd ..
```

## Optional: Create a File Snapshot Instead of Git

```powershell
python daily_status_summary.py --project-root EOAT_Standardization_Project --init-snapshot
```

## Run the Daily Report Interactively

Interactive mode:

```powershell
python daily_status_summary.py --project-root EOAT_Standardization_Project --week 1 --day 1 --interactive --include-git --include-snapshot
```

For later weeks, change `--week`:

```powershell
python daily_status_summary.py --project-root EOAT_Standardization_Project --week 2 --day 1 --interactive --include-git --include-snapshot
```

Command-line mode:

```powershell
python daily_status_summary.py --project-root EOAT_Standardization_Project --week 1 --day 1 --use-day1-defaults --include-git --include-snapshot
```

## Run from PowerShell

```powershell
.\\run_daily_status.ps1 -Week 1 -Day 1 -ProjectRoot "EOAT_Standardization_Project" -IncludeGit -IncludeSnapshot -Interactive
```

If your computer has more than one Python install, you can point the helper at
a specific interpreter:

```powershell
$env:PYTHON_EXE = "C:\\Path\\To\\python.exe"
.\\run_daily_status.ps1 -Week 1 -Day 1
```

## Schedule Daily at 6:00 PM with Windows Task Scheduler

1. Open Task Scheduler.
2. Choose **Create Basic Task**.
3. Name it `EOAT Daily Status Summary`.
4. Trigger: Daily at 6:00 PM.
5. Action: Start a Program.
6. Program/script: `powershell.exe`
7. Add arguments:

```powershell
-ExecutionPolicy Bypass -File "FULL_PATH_TO_THIS_FOLDER\\run_daily_status.ps1" -Week 1 -IncludeGit -IncludeSnapshot -Interactive
```

Replace `FULL_PATH_TO_THIS_FOLDER` with the folder containing these scripts.

## Codex Automation Option

If Codex Automation is available, create a daily 6:00 PM automation that runs this command from the toolkit folder:

```powershell
.\\run_daily_status.ps1 -Week 1 -IncludeGit -IncludeSnapshot -Interactive
```

Use interactive mode when you want to type the day's notes yourself.
"""


REQUIREMENTS_TEXT = """openpyxl
python-docx
pypdf
"""


GITIGNORE_TEXT = """# Local generated/cache files
__pycache__/
.activity_snapshots/
*.pyc

# Office lock files
~$*

# Timestamped safety backups from setup
*_backup_*.xlsx
*_backup_*.md
*_backup_*.txt
"""


def timestamp() -> str:
    """Return a filesystem-friendly timestamp for backup names."""
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def backup_existing(path: Path) -> Path:
    """Rename an existing file to a timestamped backup path."""
    backup_path = path.with_name(f"{path.stem}_backup_{timestamp()}{path.suffix}")
    path.rename(backup_path)
    return backup_path


def write_text_with_backup(path: Path, text: str) -> None:
    """Write a text file, backing up any existing version first."""
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        backup_existing(path)
    path.write_text(text, encoding="utf-8")


def write_json_if_missing(path: Path, data: dict) -> None:
    """Create a JSON file only when it does not already exist."""
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")


def write_schedule_if_missing_or_blank(path: Path, schedule: dict) -> None:
    """Create a schedule JSON, or refresh a blank template when real tasks are found."""
    path.parent.mkdir(parents=True, exist_ok=True)
    incoming_days = schedule.get("days", {})
    incoming_has_tasks = any(incoming_days.get(day) for day in incoming_days)
    if path.exists():
        try:
            existing = json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return
        existing_days = existing.get("days", {})
        existing_has_tasks = any(existing_days.get(day) for day in existing_days)
        if existing_has_tasks or not incoming_has_tasks:
            return
    path.write_text(json.dumps(schedule, indent=2), encoding="utf-8")


def schedule_file_matches_week(path: Path, week: int) -> bool:
    """Return True when a help document appears to belong to the requested week."""
    name = path.stem.lower()
    compact = re.sub(r"[^a-z0-9]+", "", name)
    return f"week{week}" in compact and "schedule" in compact


def schedule_help_files_for_week(week: int) -> list[Path]:
    """Find local schedule help documents for a week."""
    if not HELP_DOCUMENTS_DIR.exists():
        return []
    return [
        path
        for path in HELP_DOCUMENTS_DIR.iterdir()
        if path.is_file() and schedule_file_matches_week(path, week)
    ]


def extract_schedule_from_pdf(path: Path, week: int) -> dict[str, list[str]]:
    """Extract Dn-Tn planned task lines from a weekly schedule PDF when possible."""
    try:
        from pypdf import PdfReader
    except ImportError:
        return {}

    try:
        text = "\n".join(page.extract_text() or "" for page in PdfReader(str(path)).pages)
    except Exception:
        return {}

    days: dict[str, list[str]] = {}
    pattern = re.compile(rf"D(\d+)-T\d+\s*[-\u2013\u2014]\s*(.+)")
    for raw_line in text.splitlines():
        line = raw_line.strip()
        match = pattern.match(line)
        if not match:
            continue
        day, task_text = match.groups()
        task_text = task_text.strip().replace("\u2013", "-").replace("\u2014", "-")
        days.setdefault(day, [])
        if task_text not in days[day]:
            days[day].append(task_text)
    return days


def default_schedule_for_week(week: int) -> dict:
    """Return a generic schedule object for any project week."""
    if week == 1:
        days = WEEK1_DEFAULT_DAYS
        source = "Built-in Week 1 fallback schedule"
    else:
        days = {str(day): [] for day in range(1, 6)}
        source = "Template schedule; add a matching weekly schedule file in Project_Help_Documents"
    return {"week": week, "source": source, "source_files": [], "days": days}


def schedule_for_week(week: int) -> dict:
    """Return a schedule using Project_Help_Documents when possible."""
    schedule = default_schedule_for_week(week)
    help_files = schedule_help_files_for_week(week)
    schedule["source_files"] = [str(path.relative_to(SCRIPT_DIR)) for path in help_files]
    for help_file in help_files:
        if help_file.suffix.lower() == ".pdf":
            extracted_days = extract_schedule_from_pdf(help_file, week)
            if extracted_days:
                schedule["days"] = extracted_days
                schedule["source"] = f"Extracted from {help_file.name}"
                break
    return schedule


def discover_schedule_weeks() -> list[int]:
    """Find week numbers with schedule files, always including Week 1."""
    weeks = {1}
    if HELP_DOCUMENTS_DIR.exists():
        for path in HELP_DOCUMENTS_DIR.iterdir():
            match = re.search(r"week[\s_-]*(\d+)", path.stem, flags=re.IGNORECASE)
            if match and "schedule" in path.stem.lower():
                weeks.add(int(match.group(1)))
    return sorted(weeks)


def initial_progress_from_schedule(schedule: dict) -> dict:
    """Create initial Not started task progress rows from the schedule."""
    tasks = []
    for day_text, day_tasks in schedule.get("days", {}).items():
        day = int(day_text)
        for index, task_text in enumerate(day_tasks, start=1):
            tasks.append(
                {
                    "week": schedule.get("week", 1),
                    "day": day,
                    "task_id": f"W{schedule.get('week', 1)}D{day}T{index}",
                    "task_text": task_text,
                    "status": "Not started",
                    "completed_date": "",
                    "evidence": [],
                    "notes": "",
                }
            )
    return {"tasks": tasks}


def create_folders(dry_run: bool = False) -> list[str]:
    """Create the project folder tree."""
    actions: list[str] = []
    if not PROJECT_ROOT.exists():
        actions.append(f"Create project root: {PROJECT_ROOT}")
        if not dry_run:
            PROJECT_ROOT.mkdir(parents=True, exist_ok=True)
    for folder in FOLDERS:
        target = PROJECT_ROOT / folder
        if not target.exists():
            actions.append(f"Create folder: {target}")
            if not dry_run:
                target.mkdir(parents=True, exist_ok=True)
    return actions


def dropdown_formula(values: Iterable[str]) -> str:
    """Return an Excel list validation formula."""
    escaped_values = [value.replace('"', '""') for value in values]
    return '"' + ",".join(escaped_values) + '"'


def add_validation(ws, column_number: int, values: list[str]) -> None:
    """Add a dropdown validation to a worksheet column."""
    column_letter = get_column_letter(column_number)
    validation = DataValidation(
        type="list",
        formula1=dropdown_formula(values),
        allow_blank=True,
    )
    validation.error = "Choose a value from the dropdown list."
    validation.errorTitle = "Invalid value"
    validation.prompt = "Choose a standard value, or leave blank if not known yet."
    validation.promptTitle = "Dropdown"
    ws.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}1000")


def add_whole_number_validation(ws, column_number: int) -> None:
    """Add non-negative whole-number validation to a worksheet column."""
    column_letter = get_column_letter(column_number)
    validation = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    validation.error = "Enter a non-negative whole number, or leave blank if not known yet."
    validation.errorTitle = "Invalid whole number"
    validation.prompt = "Use a non-negative whole number."
    validation.promptTitle = "Whole number"
    ws.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}1000")


def width_for_header(header: str) -> int:
    """Choose practical column widths without letting long notes columns get huge."""
    if "Notes" in header or "Description" in header or "Observation" in header:
        return 28
    if "Recommended" in header or "Improvement" in header or "Known Issues" in header:
        return 30
    return min(max(len(header) + 3, 12), 26)


def style_sheet(ws, headers: list[str]) -> None:
    """Apply clean supervisor-friendly formatting to a worksheet."""
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F1F1F")
    note_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_number, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(column_number)].width = width_for_header(header)

    note_cell = ws.cell(row=2, column=len(headers))
    note_cell.value = f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    note_cell.fill = note_fill
    note_cell.alignment = Alignment(wrap_text=True)

    for row in ws.iter_rows(min_row=2, max_row=1000, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_sheet_validations(ws, sheet_name: str, headers: list[str]) -> None:
    """Add dropdowns to configured columns on a worksheet."""
    header_positions = {header: index for index, header in enumerate(headers, start=1)}
    for header, values in VALIDATIONS.get(sheet_name, {}).items():
        if header in header_positions:
            add_validation(ws, header_positions[header], values)
    for header in WHOLE_NUMBER_VALIDATIONS.get(sheet_name, []):
        if header in header_positions:
            add_whole_number_validation(ws, header_positions[header])


def add_fmea_formulas(ws, headers: list[str]) -> None:
    """Add simple RPN formulas for early data entry rows."""
    if ws.title != "FMEA Draft":
        return
    severity_col = get_column_letter(headers.index("Severity") + 1)
    frequency_col = get_column_letter(headers.index("Frequency") + 1)
    detectability_col = get_column_letter(headers.index("Detectability") + 1)
    rpn_col = headers.index("RPN") + 1

    for row in range(2, 101):
        ws.cell(row=row, column=rpn_col).value = (
            f'=IFERROR(VALUE(LEFT({severity_col}{row},1))*'
            f'VALUE(LEFT({frequency_col}{row},1))*'
            f'VALUE(LEFT({detectability_col}{row},1)),"")'
        )


def create_workbook(safe: bool = False, dry_run: bool = False) -> list[str]:
    """Create the master EOAT tracker workbook."""
    from core.audit_by_press import refresh_audit_by_press_view

    actions: list[str] = []
    if dry_run:
        if WORKBOOK_PATH.exists():
            actions.append(f"Workbook already exists: {WORKBOOK_PATH}")
        else:
            actions.append(f"Would create workbook: {WORKBOOK_PATH}")
        return actions
    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    if WORKBOOK_PATH.exists():
        if safe:
            actions.append(f"Workbook already exists; left unchanged: {WORKBOOK_PATH}")
            return actions
        backup = backup_existing(WORKBOOK_PATH)
        actions.append(f"Existing workbook backed up to: {backup}")

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    for sheet_name, headers in SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        style_sheet(ws, headers)
        add_sheet_validations(ws, sheet_name, headers)
        add_fmea_formulas(ws, headers)

    refresh_audit_by_press_view(wb)
    wb.save(WORKBOOK_PATH)
    actions.append(f"Created workbook: {WORKBOOK_PATH}")
    return actions


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create or verify the EOAT project workspace.")
    parser.add_argument("--project-root", default=str(PROJECT_ROOT), help="Target EOAT_Standardization_Project path.")
    parser.add_argument("--safe", action="store_true", help="Create missing items only; do not overwrite existing files.")
    parser.add_argument("--dry-run", action="store_true", help="Show what would be created without writing files.")
    return parser.parse_args()


def main() -> None:
    """Build the project setup and print a concise confirmation."""
    args = parse_args()
    configure_project_root(Path(args.project_root))

    actions = create_folders(dry_run=args.dry_run)
    actions.extend(create_workbook(safe=args.safe, dry_run=args.dry_run))
    for week in discover_schedule_weeks():
        schedule = schedule_for_week(week)
        week_schedule_path = ADMIN_DIR / f"project_schedule_week{week}.json"
        week_progress_path = ADMIN_DIR / f"task_progress_week{week}.json"
        if args.dry_run:
            if not week_schedule_path.exists():
                actions.append(f"Would create schedule: {week_schedule_path}")
            if not week_progress_path.exists():
                actions.append(f"Would create task progress: {week_progress_path}")
        else:
            write_schedule_if_missing_or_blank(week_schedule_path, schedule)
            write_json_if_missing(week_progress_path, initial_progress_from_schedule(schedule))
    if args.safe:
        for path, text in [
            (PROJECT_ROOT / "README.md", README_TEXT),
            (PROJECT_ROOT / ".gitignore", GITIGNORE_TEXT),
            (SCRIPT_DIR / "USAGE.md", USAGE_TEXT),
            (SCRIPT_DIR / "requirements.txt", REQUIREMENTS_TEXT),
        ]:
            if not path.exists():
                if args.dry_run:
                    actions.append(f"Would create file: {path}")
                else:
                    path.write_text(text, encoding="utf-8")
                    actions.append(f"Created file: {path}")
            else:
                actions.append(f"File already exists; left unchanged: {path}")
    elif not args.dry_run:
        write_text_with_backup(PROJECT_ROOT / "README.md", README_TEXT)
        if not (PROJECT_ROOT / ".gitignore").exists():
            (PROJECT_ROOT / ".gitignore").write_text(GITIGNORE_TEXT, encoding="utf-8")
        write_text_with_backup(SCRIPT_DIR / "USAGE.md", USAGE_TEXT)
        write_text_with_backup(SCRIPT_DIR / "requirements.txt", REQUIREMENTS_TEXT)

    print("EOAT project setup complete.")
    print(f"Project root: {PROJECT_ROOT}")
    print(f"Master workbook: {WORKBOOK_PATH}")
    print(f"Schedule files: {ADMIN_DIR / 'project_schedule_weekN.json'}")
    print(f"Usage guide: {SCRIPT_DIR / 'USAGE.md'}")
    if actions:
        print("Actions:")
        for action in actions:
            print(f"- {action}")


if __name__ == "__main__":
    main()
