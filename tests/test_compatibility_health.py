from __future__ import annotations

from openpyxl import load_workbook

from core.audit_compatibility import create_compatibility_entries
from core.audit_constants import ENTRY_TYPE_COMPATIBLE, ENTRY_TYPE_FIELD, SOURCE_AUDIT_ID_FIELD
from core.compatibility_health import validate_compatibility_health
from core.paths import resolve_project_paths
from tests.test_audit_compatibility import _save_audit, _write_press_capacity


def test_compatibility_health_flags_compatible_row_missing_source(fake_project):
    _save_audit(fake_project, "AUD-COMPAT-HEALTH-MISSING-SOURCE", "2", "PN-X", entry_type=ENTRY_TYPE_COMPATIBLE, source_id="")

    findings = validate_compatibility_health(fake_project)

    assert any(
        finding.severity == "ERROR"
        and finding.column_name == SOURCE_AUDIT_ID_FIELD
        and finding.audit_id == "AUD-COMPAT-HEALTH-MISSING-SOURCE"
        for finding in findings
    )


def test_compatibility_health_flags_stale_inherited_values(fake_project):
    _write_press_capacity(fake_project, [("1, 2", "PN-X", "Part X")])
    _save_audit(fake_project, "AUD-COMPAT-HEALTH-SOURCE", "1", "PN-X", description="Original")
    create_result = create_compatibility_entries(fake_project, "AUD-COMPAT-HEALTH-SOURCE", ["2"])
    assert create_result.success, create_result.errors
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    for row_number in range(2, ws.max_row + 1):
        if ws.cell(row=row_number, column=headers.index(ENTRY_TYPE_FIELD) + 1).value == ENTRY_TYPE_COMPATIBLE:
            ws.cell(row=row_number, column=headers.index("Known Issues") + 1).value = "Manual stale mismatch"
            break
    workbook.save(workbook_path)
    workbook.close()

    findings = validate_compatibility_health(fake_project)

    assert any(
        finding.column_name == "Known Issues"
        and "stale inherited value" in finding.message
        for finding in findings
    )
