from __future__ import annotations

from openpyxl import load_workbook

from core.audit_constants import (
    AUDIT_CONTEXT_BENCH,
    AUDIT_CONTEXT_COMPATIBILITY,
    AUDIT_CONTEXT_FIELD,
    AUDIT_CONTEXT_INSTALLED,
    AUDIT_CONTEXT_NEEDS_REVIEW,
    COMPATIBILITY_CONFIDENCE_FIELD,
    COMPATIBILITY_SOURCE_FIELD,
    ENTRY_TYPE_COMPATIBLE,
    ENTRY_TYPE_FIELD,
    PHYSICAL_AUDIT_VERIFIED_FIELD,
    SOURCE_AUDIT_ID_FIELD,
)
from core.audit_context import backfill_audit_context
from core.paths import resolve_project_paths
from core.workbook_io import row_dicts


def _append_inventory_row(project_root, values: dict[str, object]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    ws.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def _inventory_by_audit_id(project_root) -> dict[str, dict[str, object]]:
    workbook_path = resolve_project_paths(project_root).master_workbook
    return {str(row.get("Audit ID") or ""): row for row in row_dicts(workbook_path, "EOAT Inventory")}


def test_backfill_audit_context_infers_contexts_and_creates_backup(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-CONTEXT-BENCH",
            "Press/Machine #": "",
            "Tool #": "TOOL-BENCH",
            "Notes": "Bench record should keep notes.",
        },
    )
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-CONTEXT-INSTALLED",
            "Press/Machine #": "Press 12",
            "Tool #": "TOOL-INSTALLED",
        },
    )
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-CONTEXT-COMPAT",
            "Press/Machine #": "Press 13",
            "Tool #": "TOOL-COMPAT",
            ENTRY_TYPE_FIELD: ENTRY_TYPE_COMPATIBLE,
            SOURCE_AUDIT_ID_FIELD: "AUD-CONTEXT-INSTALLED",
            COMPATIBILITY_SOURCE_FIELD: "Press Capacity List",
        },
    )

    result = backfill_audit_context(fake_project, log_activity=False)
    rows = _inventory_by_audit_id(fake_project)

    assert result.success, result.errors
    assert result.files_created
    assert result.output_reports
    assert rows["AUD-CONTEXT-BENCH"][AUDIT_CONTEXT_FIELD] == AUDIT_CONTEXT_BENCH
    assert rows["AUD-CONTEXT-BENCH"][PHYSICAL_AUDIT_VERIFIED_FIELD] == "Yes"
    assert rows["AUD-CONTEXT-BENCH"]["Notes"] == "Bench record should keep notes."
    assert rows["AUD-CONTEXT-INSTALLED"][AUDIT_CONTEXT_FIELD] == AUDIT_CONTEXT_INSTALLED
    assert rows["AUD-CONTEXT-COMPAT"][AUDIT_CONTEXT_FIELD] == AUDIT_CONTEXT_COMPATIBILITY
    assert rows["AUD-CONTEXT-COMPAT"][PHYSICAL_AUDIT_VERIFIED_FIELD] == "No"
    assert rows["AUD-CONTEXT-COMPAT"][COMPATIBILITY_CONFIDENCE_FIELD] == "Press Capacity"


def test_backfill_preserves_user_entered_audit_context(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-CONTEXT-REVIEW",
            "Press/Machine #": "",
            "Tool #": "TOOL-REVIEW",
            AUDIT_CONTEXT_FIELD: AUDIT_CONTEXT_NEEDS_REVIEW,
        },
    )

    result = backfill_audit_context(fake_project, log_activity=False)
    row = _inventory_by_audit_id(fake_project)["AUD-CONTEXT-REVIEW"]

    assert result.success, result.errors
    assert row[AUDIT_CONTEXT_FIELD] == AUDIT_CONTEXT_NEEDS_REVIEW
