from __future__ import annotations

from datetime import date, timedelta

from openpyxl import load_workbook

from core.annotations.database import connect_annotation_database, current_schema_version, initialize_annotation_database, seed_default_tags
from core.annotations.migrations import LATEST_SCHEMA_VERSION
from core.annotations.service import AnnotationService
from core.annotations.tag_colors import DEFAULT_TAG_DEFINITIONS
from core.audit_entries import save_audit_entry
from core.paths import resolve_project_paths


def test_annotation_database_initializes_migrates_and_seeds_defaults(fake_project):
    db_path = initialize_annotation_database(fake_project)
    assert db_path == resolve_project_paths(fake_project).annotations_database
    assert db_path.exists()

    conn = connect_annotation_database(db_path)
    try:
        assert current_schema_version(conn) == LATEST_SCHEMA_VERSION
        tables = {row["name"] for row in conn.execute("SELECT name FROM sqlite_master WHERE type = 'table'").fetchall()}
        assert "annotation_suggestion_ignores" in tables
        assert "open_item_states" in tables
        before = conn.execute("SELECT COUNT(*) AS count FROM tags WHERE is_default = 1").fetchone()["count"]
        seed_default_tags(conn)
        conn.commit()
        after = conn.execute("SELECT COUNT(*) AS count FROM tags WHERE is_default = 1").fetchone()["count"]
    finally:
        conn.close()

    assert before == len(DEFAULT_TAG_DEFINITIONS)
    assert after == before


def test_note_crud_search_filter_sort_and_links(fake_project):
    service = AnnotationService(fake_project)
    tag = service.get_tag_by_name("Needs Review")
    target = service.create_or_get_target(
        "audit_field",
        audit_id="AUD-NOTE-001",
        machine_id="12",
        field_key="Sensor Type",
        field_label="Sensor Type",
        sheet_name="EOAT Inventory",
        header_name="Sensor Type",
    )
    note = service.create_note(
        "Sensor question",
        "- Verify sensor brand\n- Check spare",
        "Important",
        status="Open",
        collection="Machine 12",
        note_type="Question",
        follow_up_date="2026-05-30",
        target_ids=[target.id],
        tag_ids=[tag.id],
    )

    updated = service.update_note(note.id, subject="Sensor question updated", importance="Critical")
    assert updated.subject == "Sensor question updated"
    assert updated.importance == "Critical"

    matches = service.search_notes("sensor", importance="Critical", status="Open", tag_name="Needs Review", sort_by="subject")
    assert [item["id"] for item in matches] == [note.id]
    assert matches[0]["tags"][0]["name"] == "Needs Review"
    assert matches[0]["targets"][0]["audit_id"] == "AUD-NOTE-001"

    archived = service.archive_note(note.id)
    assert archived.archived_at is not None
    assert service.search_notes("sensor") == []


def test_tag_crud_search_filter_sort_and_multiple_tags_per_target(fake_project):
    service = AnnotationService(fake_project)
    target = service.create_or_get_target("machine", machine_id="26", target_label="Machine 26")
    review = service.get_tag_by_name("Needs Review")
    conflict = service.get_tag_by_name("Data Conflict")
    custom = service.create_tag("Custom EOAT Flag", "pink", description="Local custom tag")

    assert service.search_tags("custom")[0].name == custom.name
    service.update_tag(custom.id, color_key="teal")
    assert service.search_tags(color_key="teal")[0].id == custom.id

    service.assign_tag_to_target(review.id, target.id, sync_workbook=False)
    service.assign_tag_to_target(conflict.id, target.id, comment="Conflicting press data.", sync_workbook=False)
    service.assign_tag_to_target(custom.id, target.id, sync_workbook=False)

    tags = service.get_tags_for_target(target.id)
    assert {tag["name"] for tag in tags} == {"Needs Review", "Data Conflict", "Custom EOAT Flag"}
    assert service.highest_priority_color_for_target(target.id) == "orange"

    service.remove_tag_from_target(conflict.id, target.id, sync_workbook=False)
    assert service.highest_priority_color_for_target(target.id) == "yellow"
    service.remove_tag_from_target(review.id, target.id, sync_workbook=False)
    service.remove_tag_from_target(custom.id, target.id, sync_workbook=False)
    assert service.highest_priority_color_for_target(target.id) is None


def test_note_and_tag_exports_markdown_and_excel(fake_project):
    service = AnnotationService(fake_project)
    target = service.create_or_get_target("audit", audit_id="AUD-EXPORT-001")
    tag = service.get_tag_by_name("Documentation Gap")
    note = service.create_note("Exportable note", "Body **markdown**", "Neutral", target_ids=[target.id], tag_ids=[tag.id])
    service.assign_tag_to_target(tag.id, target.id, comment="Missing CAD.", sync_workbook=False)

    note_rows = service.search_notes("Exportable")
    tag_rows = service.list_tag_assignments(tag_name="Documentation Gap")
    note_md = service.export_notes_markdown(note_rows)
    note_xlsx = service.export_notes_excel(note_rows)
    tag_md = service.export_tags_markdown(tag_rows)
    tag_xlsx = service.export_tags_excel(tag_rows)

    assert note.subject in note_md.read_text(encoding="utf-8")
    assert note_xlsx.exists()
    assert "Documentation Gap" in tag_md.read_text(encoding="utf-8")
    assert tag_xlsx.exists()


def test_open_items_summary_counts_and_suggestions(fake_project):
    service = AnnotationService(fake_project)
    due = (date(2026, 5, 26) + timedelta(days=3)).isoformat()
    service.create_note("Critical open", "Body", "Critical", status="Open")
    service.create_note("Important open", "Body", "Important", status="Open", follow_up_date=due)
    service.create_note("Resolved critical", "Body", "Critical", status="Resolved")
    target = service.create_or_get_target("audit_field", audit_id="AUD-OPEN-001", field_key="Sensor Type", field_label="Sensor Type")
    for tag_name in ["Needs Review", "Data Conflict", "Missing Evidence", "Compatibility Concern", "Documentation Gap"]:
        service.assign_tag_to_target(service.get_tag_by_name(tag_name).id, target.id, sync_workbook=False)

    summary = service.get_open_items_summary(today=date(2026, 5, 26))

    assert summary["critical_notes"] == 1
    assert summary["important_notes"] == 1
    assert summary["followups_due_soon"] == 1
    assert summary["fields_needing_review"] == 1
    assert summary["data_conflicts"] == 1
    assert summary["missing_evidence"] == 1
    assert summary["compatibility_concerns"] == 1
    assert summary["documentation_gaps"] == 1

    suggestions = service.get_suggested_annotations(
        {
            "Audit ID": "AUD-SUGGEST-001",
            "Press/Machine #": "12",
            "EOAT Type": "Mechanical / Gripper",
            "Cup Type/Material": "Silicone",
            "Sensors Present?": "No",
            "Sensor Type": "SMC",
            "Quick Disconnects Present?": "No",
            "Pneumatic Quick Disconnect Type": "PTC",
            "Photos Taken?": "No",
            "Priority": "High",
        }
    )
    assert {item["tag_name"] for item in suggestions} == {"Data Conflict", "Missing Evidence"}


def test_workbook_cell_color_sync_priority_and_safe_clear(fake_project):
    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-COLOR-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Sensor Type": "SMC",
            "Status": "In Progress",
        },
    )
    assert result.success, result.errors
    service = AnnotationService(fake_project)
    target = service.create_or_get_target(
        "audit_field",
        audit_id="AUD-COLOR-001",
        machine_id="12",
        field_key="Sensor Type",
        field_label="Sensor Type",
        sheet_name="EOAT Inventory",
        header_name="Sensor Type",
    )
    review = service.get_tag_by_name("Needs Review")
    safety = service.get_tag_by_name("Safety Concern")
    verified = service.get_tag_by_name("Verified")

    service.assign_tag_to_target(review.id, target.id)
    assert _field_fill(fake_project, "AUD-COLOR-001", "Sensor Type") == "00FACC15"
    service.assign_tag_to_target(safety.id, target.id)
    assert service.highest_priority_color_for_target(target.id) == "red"
    assert _field_fill(fake_project, "AUD-COLOR-001", "Sensor Type") == "00EF4444"
    service.assign_tag_to_target(verified.id, target.id)
    assert _field_fill(fake_project, "AUD-COLOR-001", "Sensor Type") == "00EF4444"

    service.remove_tag_from_target(safety.id, target.id)
    assert service.highest_priority_color_for_target(target.id) == "yellow"
    assert _field_fill(fake_project, "AUD-COLOR-001", "Sensor Type") == "00FACC15"
    service.remove_tag_from_target(review.id, target.id)
    assert _field_fill(fake_project, "AUD-COLOR-001", "Sensor Type") == "0022C55E"
    service.remove_tag_from_target(verified.id, target.id)
    assert _field_fill(fake_project, "AUD-COLOR-001", "Sensor Type") in {"00000000", "000000"}


def test_missing_workbook_sync_returns_friendly_warning(tmp_path):
    service = AnnotationService(tmp_path)
    target = service.create_or_get_target("audit_field", audit_id="AUD-MISSING", field_key="Sensor Type", header_name="Sensor Type")

    result = service.sync_target_colors_to_workbook(target.id)

    assert result["success"] is False
    assert "Master workbook is missing" in result["warnings"][0]


def _field_fill(project_root, audit_id: str, header: str) -> str:
    workbook = load_workbook(resolve_project_paths(project_root).master_workbook)
    try:
        ws = workbook["EOAT Inventory"]
        headers = [cell.value for cell in ws[1]]
        row = next(row for row in range(2, ws.max_row + 1) if ws.cell(row=row, column=headers.index("Audit ID") + 1).value == audit_id)
        cell = ws.cell(row=row, column=headers.index(header) + 1)
        return cell.fill.fgColor.rgb or ""
    finally:
        workbook.close()
