from __future__ import annotations

import json
from datetime import date, timedelta

import pytest
from openpyxl import load_workbook

from core.action_items import add_action_item
from core.annotations.service import AnnotationService
from core.open_items import dismiss_open_item, list_open_items, open_items_summary, set_open_item_status
from core.paths import resolve_project_paths
from core.workbook_schema import get_expected_headers


def _append_inventory_row(project_root, values: dict[str, str]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    row = {header: "" for header in get_expected_headers("EOAT Inventory")}
    row.update(values)
    ws.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def _append_photo_index_row(project_root, values: dict[str, str]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["Photo Index"]
    headers = [cell.value for cell in ws[1]]
    ws.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def _update_inventory_cell(project_root, audit_id: str, field: str, value: str) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    audit_col = headers.index("Audit ID") + 1
    field_col = headers.index(field) + 1
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=audit_col).value or "") == audit_id:
            ws.cell(row=row, column=field_col).value = value
            break
    workbook.save(workbook_path)
    workbook.close()


def _audit_row(audit_id: str, **overrides) -> dict[str, str]:
    row = {
        "Audit ID": audit_id,
        "Audit Date": "2026-05-18",
        "Auditor": "KG",
        "Plant/Area": "Plant 4",
        "Press/Machine #": "Press 12",
        "Tool #": f"PN-{audit_id[-3:]}",
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "Status": "Complete",
        "Priority": "Medium",
        "Photos Taken?": "Yes",
        "Photo Folder/Link": "01_EOAT_Audit/Cell_Photos/synthetic",
        "Drawing/CAD Available?": "Yes",
        "BOM Available?": "Yes",
        "Process Binder Complete?": "Yes",
        "Spare Parts Identified?": "Yes",
    }
    row.update(overrides)
    return row


def test_annotation_suggestion_fingerprint_is_stable_and_ignore_tracks_data_change(fake_project):
    service = AnnotationService(fake_project)
    entry = {
        "Audit ID": "AUD-SUGGEST-STABLE-001",
        "Press/Machine #": "12",
        "EOAT Type": "Mechanical / Gripper",
        "Cup Type/Material": "Silicone",
    }

    first = service.get_suggested_annotations(entry)
    second = service.get_suggested_annotations(entry)

    assert first
    assert first[0]["suggestion_id"] == second[0]["suggestion_id"]
    assert first[0]["data_fingerprint"] == second[0]["data_fingerprint"]

    service.ignore_suggested_annotation(first[0])
    assert service.get_suggested_annotations(entry) == []

    changed = service.get_suggested_annotations({**entry, "Cup Type/Material": "Urethane"})
    assert changed
    assert changed[0]["suggestion_id"] != first[0]["suggestion_id"]


def test_apply_suggested_annotation_assigns_existing_tag_to_target(fake_project):
    service = AnnotationService(fake_project)
    suggestion = service.get_suggested_annotations(
        {
            "Audit ID": "AUD-SUGGEST-APPLY-001",
            "Press/Machine #": "12",
            "EOAT Type": "Mechanical / Gripper",
            "Cup Type/Material": "Silicone",
        }
    )[0]

    assignment = service.apply_suggested_annotation(suggestion)
    rows = service.list_tag_assignments(audit_id="AUD-SUGGEST-APPLY-001", target_type="audit_field")

    assert assignment.id
    assert rows[0]["tag_name"] == "Data Conflict"
    assert rows[0]["field_label"] == "Cup Type/Material"
    assert rows[0]["comment"] == suggestion["suggested_comment"]


def test_open_items_aggregate_notes_tags_actions_and_status_overrides(fake_project):
    service = AnnotationService(fake_project)
    target = service.create_or_get_target(
        "audit_field",
        audit_id="AUD-OPEN-AGG-001",
        machine_id="12",
        field_key="Sensor Type",
        field_label="Sensor Type",
    )
    service.create_note(
        "Important unresolved note",
        "Check sensor wiring.",
        "Important",
        status="Open",
        follow_up_date=(date.today() + timedelta(days=1)).isoformat(),
        target_ids=[target.id],
    )
    tag = service.get_tag_by_name("Needs Review")
    service.assign_tag_to_target(tag.id, target.id, comment="Confirm sensor type.", sync_workbook=False)
    action_result = add_action_item(
        fake_project, "Resolve synthetic open item.", related_cell_press="12", priority="High"
    )
    assert action_result.success

    items = list_open_items(fake_project, include_validation=False)
    sources = {item.source for item in items}

    assert {"note", "note_followup", "tag", "action_item"} <= sources
    tagged = next(item for item in items if item.source == "tag")
    assert tagged.target_payload()["audit_id"] == "AUD-OPEN-AGG-001"

    with pytest.raises(ValueError, match="cannot be manually marked resolved"):
        set_open_item_status(fake_project, tagged.id, "Resolved")
    assert tagged.id in {item.id for item in list_open_items(fake_project, include_validation=False)}

    dismiss_open_item(fake_project, tagged.id, reason="Covered by another reviewed action.")
    after = list_open_items(fake_project, include_validation=False)
    assert tagged.id not in {item.id for item in after}
    all_items = list_open_items(fake_project, include_resolved=True, include_validation=False)
    dismissed = next(item for item in all_items if item.id == tagged.id)
    assert dismissed.status == "Dismissed / Overridden"
    assert dismissed.dismissed_reason == "Covered by another reviewed action."


def test_open_items_summary_counts_validation_and_dismissal(fake_project):
    service = AnnotationService(fake_project)
    target = service.create_or_get_target(
        "audit_field", audit_id="AUD-OPEN-SUMMARY-001", field_key="Photos Taken?", field_label="Photos Taken?"
    )
    service.assign_tag_to_target(service.get_tag_by_name("Missing Evidence").id, target.id, sync_workbook=False)

    items = list_open_items(fake_project, include_validation=True)
    assert any(item.source == "validation" for item in items)
    evidence = next(item for item in items if item.category == "missing_evidence")
    summary = open_items_summary(fake_project)
    assert summary["missing_evidence_count"] >= 1

    set_open_item_status(fake_project, evidence.id, "Dismissed", reason="Synthetic test dismissal.")
    visible = list_open_items(fake_project, include_validation=False)
    assert evidence.id not in {item.id for item in visible}


def test_documentation_gap_disappears_when_source_field_is_fixed(fake_project):
    _append_inventory_row(fake_project, _audit_row("AUD-DOC-FIX-001", **{"BOM Available?": "No"}))

    before = list_open_items(fake_project, include_validation=False)
    item_id = "documentation_gap:AUD-DOC-FIX-001:BOM Available?"
    assert item_id in {item.id for item in before}
    before_count = open_items_summary(fake_project)["total_open_items"]

    _update_inventory_cell(fake_project, "AUD-DOC-FIX-001", "BOM Available?", "Yes")
    after = list_open_items(fake_project, include_validation=False)
    after_summary = open_items_summary(fake_project)

    assert item_id not in {item.id for item in after}
    assert after_summary["total_open_items"] < before_count
    assert after_summary["items_fixed_at_source_this_week"] >= 1


def test_missing_evidence_item_disappears_when_photo_index_data_is_added(fake_project):
    _append_inventory_row(fake_project, _audit_row("AUD-EVID-FIX-001"))

    before = list_open_items(fake_project, include_validation=False)
    item_id = "missing_evidence:AUD-EVID-FIX-001:overall_eoat"
    assert item_id in {item.id for item in before}
    assert next(item for item in before if item.id == item_id).target_payload()["target_type"] == "photo"

    _append_photo_index_row(
        fake_project,
        {
            "Photo ID": "PHO-OPEN-001",
            "Date Taken": "2026-05-18",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "EOAT Area Shown": "Overall EOAT",
            "Photo Filename": "synthetic_overall.jpg",
            "Folder Path": str(fake_project / "synthetic" / "photos"),
            "Related Audit ID": "AUD-EVID-FIX-001",
        },
    )
    after = list_open_items(fake_project, include_validation=False)

    assert item_id not in {item.id for item in after}


def test_validation_finding_disappears_when_workbook_value_is_corrected(fake_project):
    row = _audit_row("AUD-VALID-FIX-001")
    _append_inventory_row(fake_project, row)
    _append_inventory_row(fake_project, {**row, "Press/Machine #": "Press 13"})

    before = list_open_items(fake_project, include_validation=True)
    duplicate = next(item for item in before if item.source == "validation" and item.field == "Audit ID")

    _update_inventory_cell(fake_project, "AUD-VALID-FIX-001", "Audit ID", "AUD-VALID-FIX-002")
    after = list_open_items(fake_project, include_validation=True)

    assert duplicate.id not in {item.id for item in after}
    assert not any(item.source == "validation" and item.field == "Audit ID" for item in after)


def test_cup_count_validation_open_item_disappears_when_source_field_is_filled(fake_project):
    _append_inventory_row(fake_project, _audit_row("AUD-CUPS-VALID-FIX-001", **{"# of Cups": ""}))

    before = list_open_items(fake_project, include_validation=True)
    cup_item = next(item for item in before if item.source == "validation" and item.field == "# of Cups")

    _update_inventory_cell(fake_project, "AUD-CUPS-VALID-FIX-001", "# of Cups", "4")
    after = list_open_items(fake_project, include_validation=True)

    assert cup_item.id not in {item.id for item in after}
    assert not any(item.source == "validation" and item.field == "# of Cups" for item in after)


def test_dismiss_with_reason_hides_open_item_and_records_override(fake_project):
    _append_inventory_row(fake_project, _audit_row("AUD-DISMISS-001", **{"Process Binder Complete?": "No"}))
    item = next(
        item
        for item in list_open_items(fake_project, include_validation=False)
        if item.field == "Process Binder Complete?"
    )
    before_summary = open_items_summary(fake_project)

    dismissed = dismiss_open_item(fake_project, item.id, reason="Binder confirmed in controlled document system.")
    after_summary = open_items_summary(fake_project)

    assert dismissed is not None
    assert item.id not in {visible.id for visible in list_open_items(fake_project, include_validation=False)}
    assert after_summary["total_open_items"] == before_summary["total_open_items"] - 1
    assert after_summary["dismissed_overridden_count"] == before_summary["dismissed_overridden_count"] + 1
    assert after_summary["items_fixed_at_source_this_week"] == before_summary["items_fixed_at_source_this_week"]

    all_items = list_open_items(fake_project, include_resolved=True, include_validation=False)
    override = next(visible for visible in all_items if visible.id == item.id)
    assert override.status == "Dismissed / Overridden"
    assert override.dismissed_reason == "Binder confirmed in controlled document system."

    override_path = fake_project / "00_Project_Admin" / "open_items" / "open_item_overrides.jsonl"
    records = [json.loads(line) for line in override_path.read_text(encoding="utf-8").splitlines()]
    assert records[-1]["item_id"] == item.id
    assert records[-1]["reason"] == "Binder confirmed in controlled document system."
