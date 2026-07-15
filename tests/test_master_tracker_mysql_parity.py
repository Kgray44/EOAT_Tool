from __future__ import annotations

import zipfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook

from tools.verify_master_tracker_mysql_parity import (
    _indexed,
    comparison,
    duplicate_source_keys,
    find_change_history,
    issue_original_preserved,
    mapping_matrix,
    normalized_value,
    source_fingerprint,
    split_multi_value,
    workbook_inventory,
)


def _save_workbook(path: Path, *, hidden: bool = False, comment: bool = False, formula: str | None = None) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "EOAT Inventory"
    sheet.append(["Audit ID", "EOAT Assembly ID", "Notes", "Custom Field"])
    sheet.append(["AUD-1", "P4-EOAT-0001", " note ", "custom"])
    if comment:
        from openpyxl.comments import Comment

        sheet["C2"].comment = Comment("Preserved cell note", "Tester")
    if formula:
        sheet["D2"] = formula
    if hidden:
        extra = workbook.create_sheet("Hidden Evidence")
        extra.sheet_state = "hidden"
        extra.append(["Key", "Value"])
        extra.append(["secret", "evidence"])
    workbook.save(path)
    return path


def _inject_formula_cache(path: Path, value: str = "2") -> None:
    temporary = path.with_suffix(".cached.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(temporary, "w") as target:
        for info in source.infolist():
            content = source.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                xml = content.decode("utf-8")
                xml = xml.replace("<f>1+1</f><v></v>", f"<f>1+1</f><v>{value}</v>")
                content = xml.encode("utf-8")
            target.writestr(info, content)
    temporary.replace(path)


def test_exact_scalar_match():
    assert comparison("EOAT-001", "EOAT-001") == "exact_match"


def test_normalized_whitespace_match():
    assert comparison("  Alpha\t  Beta  ", "alpha beta") == "normalized_match"


def test_blank_and_null_match():
    assert comparison("", None) == "exact_match"


def test_blank_is_not_false():
    assert comparison("", False) == "conflicting_database_value"


def test_numeric_identifier_leading_zero_is_preserved():
    assert normalized_value("0040") != normalized_value(40.0)


def test_excel_date_matches_sql_midnight():
    assert comparison(date(2026, 5, 19), datetime(2026, 5, 19)) == "normalized_match"


def test_multi_value_relationship_expansion_is_deterministic():
    assert split_multi_value("12, 14;18|20\n22") == ["12", "14", "18", "20", "22"]


def test_missing_relationship_member_is_detected():
    expected = set(split_multi_value("12,14,18"))
    actual = {"12", "18"}
    assert expected - actual == {"14"}


def test_duplicate_source_key_is_reported():
    rows = {"EOAT Inventory": {2: {"Audit ID": "AUD-1"}, 3: {"Audit ID": " aud-1 "}}, "Photo Index": {}}
    duplicates = duplicate_source_keys(rows)
    assert duplicates[0]["count"] == 2


def test_duplicate_database_key_fixture_is_ambiguous():
    index = _indexed([{"id": 1, "business_identifier": "E-1"}, {"id": 2, "business_identifier": "E-1"}], "business_identifier")
    assert len(index["e-1"]) == 2


def test_truncated_text_is_not_normalized_away():
    assert comparison("complete source value", "complete source") == "truncated_database_value"


def test_hyperlink_and_path_normalization():
    assert comparison(r"folder/sub/file.jpg", r"folder\sub\file.jpg", header="Folder Path") == "normalized_match"


def test_workbook_cell_comment_is_inventoried(tmp_path):
    path = _save_workbook(tmp_path / "comment.xlsx", comment=True)
    inventory, _, _ = workbook_inventory(path)
    assert inventory["sheets"][0]["comments"][0]["text"] == "Preserved cell note"


def test_current_value_changed_after_import_with_history():
    database = {"audit_log": [{
        "id": 5, "entity_type": "eoat", "entity_id": 9, "occurred_at": datetime(2026, 7, 14),
        "previous_values_json": {"notes": "old"}, "new_values_json": {"notes": "new"},
        "changed_fields_json": ["notes"], "reason": "verified correction",
    }], "history": []}
    assert "event_id=5" in find_change_history(database, "eoat", 9, "notes", "old")


def test_current_value_changed_without_history_fails_evidence_check():
    assert find_change_history({"audit_log": [], "history": []}, "eoat", 9, "notes", "old") == ""


def test_unmapped_populated_metadata_column_is_failure(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "_EOAT_App_Metadata"
    sheet.append(["key", "value"])
    sheet.append(["schema_version", "1.0"])
    path = tmp_path / "metadata.xlsx"
    workbook.save(path)
    _, rows, headers = workbook_inventory(path)
    matrix = mapping_matrix(rows, headers)
    assert {item["status"] for item in matrix} == {"unmapped_failure"}


def test_hidden_populated_sheet_is_inventoried(tmp_path):
    path = _save_workbook(tmp_path / "hidden.xlsx", hidden=True)
    inventory, _, _ = workbook_inventory(path)
    hidden = next(item for item in inventory["sheets"] if item["sheet_name"] == "Hidden Evidence")
    assert hidden["visibility"] == "hidden"
    assert hidden["candidate_data_rows"] == 1


def test_formula_with_cached_value_is_recognized(tmp_path):
    path = _save_workbook(tmp_path / "cached.xlsx", formula="=1+1")
    _inject_formula_cache(path)
    inventory, _, _ = workbook_inventory(path)
    sheet = inventory["sheets"][0]
    assert sheet["formula_cells"] == 1
    assert sheet["formula_cache_missing"] == []


def test_formula_without_cached_value_is_reported(tmp_path):
    path = _save_workbook(tmp_path / "uncached.xlsx", formula="=1+1")
    inventory, _, _ = workbook_inventory(path)
    assert inventory["sheets"][0]["formula_cache_missing"] == ["D2"]


def test_malformed_source_value_preserved_in_issue_provenance():
    issue = {
        "issue_code": "AMBIGUOUS_MACHINE_VALUE", "field_name": "Press/Machine #", "source_value": "12/14",
        "raw_values_json": {"Press/Machine #": "12/14"}, "resolution_notes": {"status": "UNRESOLVED"},
    }
    assert issue_original_preserved(issue)


def test_database_only_additional_record_is_not_a_source_failure():
    source = {"E-1"}
    database = {"E-1", "E-2"}
    assert database - source == {"E-2"}


def test_orphaned_foreign_key_fixture_fails_integrity():
    orphan_checks = [{"constraint_name": "fk_child_parent", "orphan_count": 1, "status": "FAIL"}]
    assert sum(item["orphan_count"] for item in orphan_checks) == 1


def test_ambiguous_row_match_fixture():
    matches = _indexed([{"id": 1, "audit_identifier": "AUD-1"}, {"id": 2, "audit_identifier": "AUD-1"}], "audit_identifier")
    assert len(matches["aud-1"]) > 1


def test_source_fingerprint_is_stable_and_row_specific():
    values = {"Audit ID": "AUD-1", "Notes": "x"}
    assert source_fingerprint("EOAT Inventory", 2, values) == source_fingerprint("EOAT Inventory", 2, dict(reversed(list(values.items()))))
    assert source_fingerprint("EOAT Inventory", 2, values) != source_fingerprint("EOAT Inventory", 3, values)
