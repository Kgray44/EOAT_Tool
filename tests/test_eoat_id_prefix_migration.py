from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.atlas.minimalist.data import infer_search_kind
from core.atlas_exports import validate_eoat_qr_payload
from core.atlas_search import interpret_query
from core.eoat_id_migration import (
    apply_eoat_id_migration_to_photos,
    build_eoat_id_migration_map,
    run_eoat_id_prefix_migration,
    validate_eoat_id_prefixes,
)
from core.eoat_ids import EOAT_ASSEMBLY_ID_FIELD
from core.paths import resolve_project_paths
from core.photo_index import build_photo_index


def _make_project_workbook(project_root: Path) -> Path:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    ws = workbook.active
    ws.title = "EOAT Inventory"
    headers = [
        "Audit ID",
        "Plant/Area",
        "Cleanroom/Non-Cleanroom",
        "Press/Machine #",
        "Tool #",
        EOAT_ASSEMBLY_ID_FIELD,
        "Entry Type",
        "Source Audit ID",
        "Photo Folder/Link",
    ]
    ws.append(headers)
    ws.append(
        [
            "AUD-CLEAN",
            "Cleanroom",
            "Cleanroom",
            "63",
            "TOOL-CLEAN",
            "P4-EOAT-0047",
            "Audited",
            "",
            "01_EOAT_Audit/Cell_Photos/P4-EOAT-0047/P4-EOAT-0047_overall.jpg",
        ]
    )
    ws.append(["AUD-P4", "Plant 4", "Whiteroom", "12", "TOOL-P4", "P4-EOAT-0002", "Audited", "", ""])
    ws.append(["AUD-CL", "Cleanroom", "Cleanroom", "64", "TOOL-CL", "CL-EOAT-0003", "Audited", "", ""])
    ws.append(["AUD-COMPAT", "Cleanroom", "Cleanroom", "65", "TOOL-CLEAN", "P4-EOAT-0047", "Compatible", "AUD-CLEAN", ""])
    photo = workbook.create_sheet("Photo Index")
    photo.append(["Photo ID", EOAT_ASSEMBLY_ID_FIELD, "Stored Relative Path", "Photo Link"])
    photo.append(
        [
            "PHO-001",
            "P4-EOAT-0047",
            "01_EOAT_Audit/Cell_Photos/P4-EOAT-0047/P4-EOAT-0047_overall.jpg",
            "P4-EOAT-0047_overall.jpg",
        ]
    )
    generated = workbook.create_sheet("Audit by Press")
    generated.append(["Machine", EOAT_ASSEMBLY_ID_FIELD])
    generated.append(["63", "P4-EOAT-0047"])
    workbook.save(workbook_path)
    workbook.close()
    return workbook_path


def test_migration_map_only_changes_cleanroom_p4_ids(tmp_path: Path) -> None:
    workbook_path = _make_project_workbook(tmp_path)

    mappings = build_eoat_id_migration_map(workbook_path)

    assert [(item.old_id, item.new_id, item.source_row) for item in mappings] == [("P4-EOAT-0047", "CL-EOAT-0047", 2)]


def test_validation_catches_prefix_mismatches(tmp_path: Path) -> None:
    workbook_path = _make_project_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    ws.append(["AUD-BAD-P4", "Plant 4", "Whiteroom", "14", "TOOL-BAD", "CL-EOAT-0099", "Audited", "", ""])
    workbook.save(workbook_path)
    workbook.close()

    issues = validate_eoat_id_prefixes(workbook_path)
    messages = [issue.message for issue in issues]

    assert any("P4-EOAT-0047 should use CL" in message for message in messages)
    assert any("CL-EOAT-0099 should use P4" in message for message in messages)


def test_dry_run_reports_workbook_photo_and_cache_updates_without_modifying(tmp_path: Path) -> None:
    workbook_path = _make_project_workbook(tmp_path)
    photo_root = tmp_path / "01_EOAT_Audit" / "Cell_Photos"
    photo_folder = photo_root / "P4-EOAT-0047"
    photo_folder.mkdir(parents=True)
    (photo_folder / "P4-EOAT-0047_overall.jpg").write_bytes(b"photo")
    cache_dir = tmp_path / "00_Project_Admin" / "cache"
    cache_dir.mkdir(parents=True)
    cache_file = cache_dir / "atlas_index.json"
    cache_file.write_text('{"eoat":"P4-EOAT-0047"}\n', encoding="utf-8")

    result = run_eoat_id_prefix_migration(
        workbook_path=workbook_path,
        photo_root=photo_root,
        project_root=tmp_path,
        apply=False,
    )

    assert result.dry_run is True
    assert result.mappings[0].new_id == "CL-EOAT-0047"
    assert result.workbook_updates
    assert result.photo_updates
    assert result.cache_updates
    assert (photo_folder / "P4-EOAT-0047_overall.jpg").exists()
    assert "P4-EOAT-0047" in cache_file.read_text(encoding="utf-8")
    assert Path(result.migration_report_md).exists()
    assert Path(result.validation_report_json).exists()


def test_apply_migrates_workbook_photos_cache_and_creates_backups(tmp_path: Path) -> None:
    workbook_path = _make_project_workbook(tmp_path)
    photo_root = tmp_path / "01_EOAT_Audit" / "Cell_Photos"
    photo_folder = photo_root / "P4-EOAT-0047"
    photo_folder.mkdir(parents=True)
    (photo_folder / "P4-EOAT-0047_overall.jpg").write_bytes(b"photo")
    cache_dir = tmp_path / "00_Project_Admin" / "cache"
    cache_dir.mkdir(parents=True)
    cache_file = cache_dir / "atlas_index.json"
    cache_file.write_text('{"eoat":"P4-EOAT-0047"}\n', encoding="utf-8")

    result = run_eoat_id_prefix_migration(
        workbook_path=workbook_path,
        photo_root=photo_root,
        project_root=tmp_path,
        apply=True,
    )

    assert result.success, result.errors
    assert result.backups
    assert (tmp_path / "backups").exists()
    assert not (photo_root / "P4-EOAT-0047" / "P4-EOAT-0047_overall.jpg").exists()
    assert (photo_root / "CL-EOAT-0047" / "CL-EOAT-0047_overall.jpg").exists()
    assert "CL-EOAT-0047" in cache_file.read_text(encoding="utf-8")
    assert "P4-EOAT-0047" not in cache_file.read_text(encoding="utf-8")

    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        values = [
            str(cell.value)
            for ws in workbook.worksheets
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        ]
    finally:
        workbook.close()
    joined = "\n".join(values)
    assert "CL-EOAT-0047" in joined
    assert "P4-EOAT-0047" not in joined
    assert "P4-EOAT-0002" in joined
    assert Path(result.migration_report_md).exists()
    assert Path(result.migration_report_csv).exists()
    assert Path(result.validation_report_md).exists()
    assert Path(result.validation_report_json).exists()


def test_photo_file_conflict_is_reported_without_overwrite(tmp_path: Path) -> None:
    photo_root = tmp_path / "photos"
    old_folder = photo_root / "P4-EOAT-0047"
    new_folder = photo_root / "CL-EOAT-0047"
    old_folder.mkdir(parents=True)
    new_folder.mkdir(parents=True)
    old_file = old_folder / "P4-EOAT-0047_overall.jpg"
    target_file = new_folder / "P4-EOAT-0047_overall.jpg"
    old_file.write_bytes(b"old")
    target_file.write_bytes(b"new")

    updates = apply_eoat_id_migration_to_photos(
        photo_root,
        {"P4-EOAT-0047": "CL-EOAT-0047"},
        dry_run=True,
    )

    assert any(update.status == "conflict" for update in updates)
    assert old_file.read_bytes() == b"old"
    assert target_file.read_bytes() == b"new"


def test_duplicate_target_collision_blocks_mapping(tmp_path: Path) -> None:
    workbook_path = _make_project_workbook(tmp_path)
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    ws.append(["AUD-COLLIDE", "Cleanroom", "Cleanroom", "66", "TOOL-COLLIDE", "CL-EOAT-0047", "Audited", "", ""])
    workbook.save(workbook_path)
    workbook.close()

    result = run_eoat_id_prefix_migration(workbook_path=workbook_path, project_root=tmp_path, apply=False)

    assert result.mappings == []
    assert any("Collision blocked P4-EOAT-0047 -> CL-EOAT-0047" in error for error in result.errors)


def test_search_library_photo_and_export_logic_accept_cl_ids(tmp_path: Path) -> None:
    query_type, value = interpret_query("cl eoat 47")

    assert query_type == "eoat"
    assert value == "CL-EOAT-0047"
    assert infer_search_kind("CL-EOAT-0047") == "EOAT"
    assert validate_eoat_qr_payload(
        "EOAT_ATLAS_RECORD; EOAT=CL-EOAT-0047; TOOL=T-5620040010",
        eoat_id="CL-EOAT-0047",
    ) == []

    photo_root = tmp_path / "01_EOAT_Audit" / "Cell_Photos" / "CL-EOAT-0047"
    photo_root.mkdir(parents=True)
    (photo_root / "CL-EOAT-0047_overall.jpg").write_bytes(b"photo")
    photo_sets, _by_tool, warnings = build_photo_index(
        tmp_path,
        [{EOAT_ASSEMBLY_ID_FIELD: "CL-EOAT-0047", "Audit ID": "AUD-CL"}],
        [],
    )

    assert warnings == []
    assert "cleoat0047" in photo_sets
    assert photo_sets["cleoat0047"].folder_exists is True
