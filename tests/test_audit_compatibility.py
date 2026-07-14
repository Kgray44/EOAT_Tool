from __future__ import annotations

from openpyxl import Workbook, load_workbook

from core.audit_compatibility import (
    OFF_MACHINE_COMPATIBILITY_ADD_ONLY,
    OFF_MACHINE_COMPATIBILITY_LEAVE,
    OFF_MACHINE_COMPATIBILITY_UPDATE_AND_ADD,
    apply_off_machine_compatibility_choice,
    build_compatibility_candidates,
    build_off_machine_compatibility_preview,
    create_compatibility_entries,
    list_audit_options,
    list_audited_source_options,
    normalize_tool_identifier,
    parse_machine_tokens,
    sync_compatible_rows_from_source,
)
from core.audit_constants import (
    AUDIT_CONTEXT_BENCH,
    AUDIT_CONTEXT_COMPATIBILITY,
    AUDIT_CONTEXT_FIELD,
    COMPATIBILITY_CONFIDENCE_FIELD,
    COMPATIBILITY_SOURCE_FIELD,
    COMPATIBILITY_SOURCE_PRESS_CAPACITY,
    ENTRY_TYPE_COMPATIBLE,
    ENTRY_TYPE_FIELD,
    PHYSICAL_AUDIT_VERIFIED_FIELD,
    SOURCE_AUDIT_ID_FIELD,
)
from core.audit_entries import save_audit_entry, save_audit_entry_with_compatibility_autorun
from core.audit_progress import calculate_audit_progress
from core.eoat_ids import EOAT_ASSEMBLY_ID_FIELD
from core.logging import read_recent_activity
from core.paths import get_master_press_list_file, get_press_capacity_file, resolve_project_paths
from core.workbook_schema import get_expected_headers, get_expected_sheets


def _write_press_capacity(project_root, rows):
    return _write_press_capacity_with_headers(
        project_root,
        ["Machine No.", "NGW Part Number", "NGW Part Description"],
        rows,
    )


def _write_press_capacity_with_headers(project_root, headers, rows):
    path = get_press_capacity_file(project_root)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Capacity"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    workbook.save(path)
    workbook.close()
    return path


def _save_audit(
    project_root, audit_id, machine, part_number, *, description="Part X", entry_type="Audited", source_id=""
):
    entry = {
        "Audit ID": audit_id,
        "Audit Date": "2026-05-18",
        "Auditor": "KG",
        "Plant/Area": "Plant 4",
        "Press/Machine #": machine,
        "Tool #": part_number,
        "Part Name/Description": description,
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "Status": "Audited",
        "Entry Type": entry_type,
        "Source Audit ID": source_id,
    }
    if entry_type == ENTRY_TYPE_COMPATIBLE:
        entry[COMPATIBILITY_SOURCE_FIELD] = "Synthetic test link"
    result = save_audit_entry(project_root, entry)
    assert result.success, result.errors
    return result


def _inventory_rows(project_root):
    workbook = load_workbook(resolve_project_paths(project_root).master_workbook, read_only=True)
    try:
        ws = workbook["EOAT Inventory"]
        headers = [cell.value for cell in ws[1]]
        return [
            {headers[index]: value for index, value in enumerate(row)}
            for row in ws.iter_rows(min_row=2, values_only=True)
            if any(value not in (None, "") for value in row)
        ]
    finally:
        workbook.close()


def _save_off_machine_audit(project_root, audit_id, tool_number, **overrides):
    entry = {
        "Audit ID": audit_id,
        "Audit Date": "2026-05-18",
        "Auditor": "KG",
        "Plant/Area": "",
        "Press/Machine #": "",
        "Tool #": tool_number,
        EOAT_ASSEMBLY_ID_FIELD: overrides.pop(EOAT_ASSEMBLY_ID_FIELD, "P4-EOAT-9001"),
        "Robot Type": "",
        "Robot Model/Controller": "",
        "Part Family": "",
        "Part Name/Description": "",
        "EOAT Type": "Vacuum",
        "Known Issues": "Keep this EOAT note.",
        "Status": "In Progress",
    }
    entry.update(overrides)
    result = save_audit_entry(project_root, entry)
    assert result.success, result.errors
    return result


def _write_detailed_capacity(project_root, rows):
    headers = [
        "Machine No.",
        "Tool #",
        "Part Name/Description",
        "Plant/Area",
        "Robot Type",
        "Robot Model/Controller",
        "Cleanroom/Non-Cleanroom",
        "Part Family",
    ]
    return _write_press_capacity_with_headers(project_root, headers, rows)


def _write_master_press_list(project_root, rows, headers=None):
    path = get_master_press_list_file(project_root)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Machine Specifications"
    ws.append(
        headers
        or [
            "Machine Number",
            "Robot/Picker Brand",
            "Robot/Picker Model #",
        ]
    )
    for row in rows:
        ws.append(list(row))
    workbook.save(path)
    workbook.close()
    return path


def test_blank_or_missing_entry_type_counts_as_audited(fake_project):
    _write_press_capacity(fake_project, [("1", "PN-X", "Part X")])
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        ws = workbook.create_sheet(sheet_name)
        headers = [
            header
            for header in get_expected_headers(sheet_name)
            if header not in {"Entry Type", "Source Audit ID", "Compatibility Source"}
        ]
        ws.append(headers)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    row = [""] * len(headers)
    for field, value in {
        "Audit ID": "AUD-OLD-001",
        "Press/Machine #": "1",
        "Tool #": "PN-X",
        "Part Name/Description": "Part X",
        "EOAT Type": "Vacuum",
    }.items():
        row[headers.index(field)] = value
    ws.append(row)
    workbook.save(workbook_path)
    workbook.close()

    summary, error = calculate_audit_progress(fake_project)

    assert error is None
    assert summary.metrics["physically_audited_relationships"] == 1
    assert summary.entry_type_counts["Unknown treated as Audited"] == 1


def test_machine_cell_parsing_uses_exact_tokens():
    assert parse_machine_tokens("1, 2, 8, 9, 11") == ["1", "2", "8", "9", "11"]
    assert "1" in parse_machine_tokens("11, 1")
    assert parse_machine_tokens("11, 1").count("1") == 1


def test_audit_options_sort_by_machine_number_numerically(fake_project):
    for audit_id, machine in [
        ("AUD-SORT-010", "Press 10"),
        ("AUD-SORT-002", "Press 2"),
        ("AUD-SORT-WEIRD", "Bench A"),
        ("AUD-SORT-001", "Press 1"),
        ("AUD-SORT-026", "26, 70"),
    ]:
        _save_audit(fake_project, audit_id, machine, f"PN-{audit_id}", description=audit_id)
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    row = [""] * len(headers)
    for field, value in {
        "Audit ID": "AUD-SORT-BLANK",
        "Tool #": "PN-BLANK",
        "Part Name/Description": "Blank machine",
        "Entry Type": "Audited",
    }.items():
        row[headers.index(field)] = value
    ws.append(row)
    workbook.save(workbook_path)
    workbook.close()

    options = list_audit_options(fake_project)
    sorted_ids = [option.audit_id for option in options]

    assert (
        sorted_ids.index("AUD-SORT-001")
        < sorted_ids.index("AUD-SORT-002")
        < sorted_ids.index("AUD-SORT-010")
        < sorted_ids.index("AUD-SORT-026")
    )
    assert sorted_ids.index("AUD-SORT-WEIRD") > sorted_ids.index("AUD-SORT-026")
    assert sorted_ids.index("AUD-SORT-BLANK") > sorted_ids.index("AUD-SORT-026")


def test_compatibility_source_options_sort_by_machine_number_numerically(fake_project):
    for audit_id, machine in [
        ("AUD-COMPAT-SORT-011", "Press 11"),
        ("AUD-COMPAT-SORT-003", "Press 3"),
        ("AUD-COMPAT-SORT-070", "Press 70"),
    ]:
        _save_audit(fake_project, audit_id, machine, f"PN-{audit_id}", description=audit_id)

    options = list_audited_source_options(fake_project)
    sorted_ids = [option.audit_id for option in options]

    assert (
        sorted_ids.index("AUD-COMPAT-SORT-003")
        < sorted_ids.index("AUD-COMPAT-SORT-011")
        < sorted_ids.index("AUD-COMPAT-SORT-070")
    )


def test_audited_source_creates_compatibility_opportunities(fake_project):
    _write_press_capacity(fake_project, [("1, 2, 8, 9, 11", "PN-X", "Part X")])
    _save_audit(fake_project, "AUD-SOURCE-001", "1", "PN-X")

    result = build_compatibility_candidates(fake_project, "AUD-SOURCE-001")
    actions = {candidate.machine_no: candidate.recommended_action for candidate in result.candidates}

    assert actions["1"] == "Already Audited"
    assert actions["2"] == "Create Compatible Entry"
    assert actions["11"] == "Create Compatible Entry"


def test_compatibility_entry_skips_audited_and_compatible_duplicates(fake_project):
    _write_press_capacity(fake_project, [("1, 2, 3", "PN-X", "Part X")])
    _save_audit(fake_project, "AUD-SOURCE-001", "1", "PN-X")

    first = create_compatibility_entries(fake_project, "AUD-SOURCE-001", ["1", "2"])
    assert first.success, first.errors
    assert first.metrics["created"] == 1
    assert first.metrics["skipped_already_audited"] == 1

    second = create_compatibility_entries(fake_project, "AUD-SOURCE-001", ["2"])
    assert second.success, second.errors
    assert second.metrics["created"] == 0
    assert second.metrics["skipped_already_compatible"] == 1

    workbook = load_workbook(resolve_project_paths(fake_project).master_workbook, read_only=True)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    rows = [
        {headers[index]: value for index, value in enumerate(row)} for row in ws.iter_rows(min_row=2, values_only=True)
    ]
    compatible_rows = [row for row in rows if row.get(ENTRY_TYPE_FIELD) == ENTRY_TYPE_COMPATIBLE]
    assert len(compatible_rows) == 1
    assert compatible_rows[0][SOURCE_AUDIT_ID_FIELD] == "AUD-SOURCE-001"
    workbook.close()


def test_save_with_compatibility_autorun_creates_compatible_entries(fake_project):
    _write_press_capacity(fake_project, [("1, 2", "PN-X", "Part X")])

    result = save_audit_entry_with_compatibility_autorun(
        fake_project,
        {
            "Audit ID": "AUD-AUTORUN-SOURCE",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "1",
            "Tool #": "PN-X",
            "Part Name/Description": "Part X",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Status": "In Progress",
        },
    )

    assert result.success, result.errors
    assert "Audit Save Summary" in result.summary
    assert "Fit Check Entry Summary" in result.summary
    assert result.metrics["compatibility_autorun_success"] is True
    assert result.metrics["compatibility_created"] == 1
    rows = _inventory_rows(fake_project)
    compatible = next(row for row in rows if row.get("Press/Machine #") == "2")
    assert compatible[ENTRY_TYPE_FIELD] == ENTRY_TYPE_COMPATIBLE
    assert compatible[SOURCE_AUDIT_ID_FIELD] == "AUD-AUTORUN-SOURCE"


def test_off_machine_update_and_add_single_match_fills_current_row_only(fake_project):
    _write_detailed_capacity(
        fake_project,
        [("7", "PN-OFF-1", "Capacity Part", "Plant 7", "Wittmann", "R9", "Cleanroom", "Family 7")],
    )
    _save_off_machine_audit(fake_project, "AUD-OFF-SINGLE", "PN-OFF-1")

    preview = build_off_machine_compatibility_preview(fake_project, "AUD-OFF-SINGLE")
    assert not preview.errors
    assert len(preview.matches) == 1

    result = apply_off_machine_compatibility_choice(
        fake_project,
        "AUD-OFF-SINGLE",
        OFF_MACHINE_COMPATIBILITY_UPDATE_AND_ADD,
    )

    assert result.success, result.errors
    assert result.metrics["current_row_updated"] is True
    assert result.metrics["created"] == 0
    rows = [row for row in _inventory_rows(fake_project) if row.get("Tool #") == "PN-OFF-1"]
    assert len(rows) == 1
    row = rows[0]
    assert row["Audit ID"] == "AUD-OFF-SINGLE"
    assert row["Press/Machine #"] == "7"
    assert row["Plant/Area"] == "Plant 7"
    assert row["Robot Type"] == "Wittmann"
    assert row["Robot Model/Controller"] == "R9"
    assert row["Part Family"] == "Family 7"
    assert row[ENTRY_TYPE_FIELD] == "Audited"


def test_off_machine_uses_master_press_robot_fields_when_capacity_robot_fields_are_blank(fake_project):
    _write_press_capacity_with_headers(
        fake_project,
        [
            "Machine No.",
            "Tool #",
            "Part Name/Description",
            "Plant/Area",
            "Cleanroom/Non-Cleanroom",
            "Part Family",
        ],
        [("7", "PN-OFF-MASTER-ROBOT", "Capacity Part", "Plant 7", "Cleanroom", "Family 7")],
    )
    _write_master_press_list(fake_project, [("7", "Wittmann", "W833")])
    _save_off_machine_audit(fake_project, "AUD-OFF-MASTER-ROBOT", "PN-OFF-MASTER-ROBOT")

    preview = build_off_machine_compatibility_preview(fake_project, "AUD-OFF-MASTER-ROBOT")
    assert preview.matches[0].machine_data["Robot Type"] == "Wittmann W833"
    assert preview.matches[0].machine_data["Robot Model/Controller"] == "W833"

    result = apply_off_machine_compatibility_choice(
        fake_project,
        "AUD-OFF-MASTER-ROBOT",
        OFF_MACHINE_COMPATIBILITY_UPDATE_AND_ADD,
    )

    assert result.success, result.errors
    row = next(row for row in _inventory_rows(fake_project) if row["Audit ID"] == "AUD-OFF-MASTER-ROBOT")
    assert row["Press/Machine #"] == "7"
    assert row["Robot Type"] == "Wittmann W833"
    assert row["Robot Model/Controller"] == "W833"


def test_off_machine_uses_master_robot_picker_fields_for_each_compatible_machine(fake_project):
    _write_press_capacity_with_headers(
        fake_project,
        [
            "Machine No.",
            "Tool #",
            "Part Name/Description",
            "Plant/Area",
            "Cleanroom/Non-Cleanroom",
            "Part Family",
        ],
        [("1, 70", "PN-OFF-1-70", "Capacity Part", "Plant 4", "Cleanroom", "Family 170")],
    )
    _write_master_press_list(
        fake_project,
        [
            ("1", "Wittmann", "W808"),
            ("70", "Engel", "Viper"),
        ],
    )
    _save_off_machine_audit(fake_project, "AUD-OFF-1-70", "PN-OFF-1-70")

    preview = build_off_machine_compatibility_preview(fake_project, "AUD-OFF-1-70")
    preview_by_machine = {match.machine_no: match for match in preview.matches}
    assert preview_by_machine["1"].machine_data["Robot Type"] == "Wittmann W808"
    assert preview_by_machine["1"].machine_data["Robot Model/Controller"] == "W808"
    assert preview_by_machine["70"].machine_data["Robot Type"] == "Engel Viper"
    assert preview_by_machine["70"].machine_data["Robot Model/Controller"] == "Viper"

    result = apply_off_machine_compatibility_choice(
        fake_project,
        "AUD-OFF-1-70",
        OFF_MACHINE_COMPATIBILITY_UPDATE_AND_ADD,
    )

    assert result.success, result.errors
    rows = [row for row in _inventory_rows(fake_project) if row.get("Tool #") == "PN-OFF-1-70"]
    by_machine = {row["Press/Machine #"]: row for row in rows}
    assert set(by_machine) == {"1", "70"}
    assert by_machine["1"]["Audit ID"] == "AUD-OFF-1-70"
    assert by_machine["1"]["Robot Type"] == "Wittmann W808"
    assert by_machine["1"]["Robot Model/Controller"] == "W808"
    assert by_machine["70"][ENTRY_TYPE_FIELD] == ENTRY_TYPE_COMPATIBLE
    assert by_machine["70"]["Robot Type"] == "Engel Viper"
    assert by_machine["70"]["Robot Model/Controller"] == "Viper"


def test_off_machine_master_press_robot_fields_override_capacity_robot_fields(fake_project):
    _write_detailed_capacity(
        fake_project,
        [
            (
                "1",
                "PN-OFF-MASTER-WINS",
                "Capacity Part",
                "Plant 4",
                "Capacity Robot",
                "Capacity Controller",
                "Cleanroom",
                "Family 1",
            )
        ],
    )
    _write_master_press_list(fake_project, [("1", "Wittmann", "W808")])
    _save_off_machine_audit(fake_project, "AUD-OFF-MASTER-WINS", "PN-OFF-MASTER-WINS")

    result = apply_off_machine_compatibility_choice(
        fake_project,
        "AUD-OFF-MASTER-WINS",
        OFF_MACHINE_COMPATIBILITY_UPDATE_AND_ADD,
    )

    assert result.success, result.errors
    row = next(row for row in _inventory_rows(fake_project) if row["Audit ID"] == "AUD-OFF-MASTER-WINS")
    assert row["Robot Type"] == "Wittmann W808"
    assert row["Robot Model/Controller"] == "W808"


def test_off_machine_reports_when_compatible_machine_is_missing_from_master_press_list(fake_project):
    _write_press_capacity_with_headers(
        fake_project,
        [
            "Machine No.",
            "Tool #",
            "Part Name/Description",
            "Plant/Area",
        ],
        [("1, 70", "PN-OFF-MISSING-MASTER", "Capacity Part", "Plant 4")],
    )
    _write_master_press_list(fake_project, [("1", "Wittmann", "W808")])
    _save_off_machine_audit(fake_project, "AUD-OFF-MISSING-MASTER", "PN-OFF-MISSING-MASTER")

    preview = build_off_machine_compatibility_preview(fake_project, "AUD-OFF-MISSING-MASTER")
    preview_by_machine = {match.machine_no: match for match in preview.matches}

    assert preview_by_machine["1"].machine_data["Robot Type"] == "Wittmann W808"
    assert "Robot Type" not in preview_by_machine["70"].machine_data
    assert any("Machine 70 was not found in the Master Press List" in warning for warning in preview.warnings)


def test_off_machine_reads_robot_picker_aliases_from_press_capacity(fake_project):
    _write_press_capacity_with_headers(
        fake_project,
        [
            "Machine No.",
            "Tool #",
            "Part Name/Description",
            "Plant/Area",
            "Robot/Picker Brand",
            "Robot/Picker Model #",
        ],
        [("7", "PN-OFF-ROBOT-ALIASES", "Capacity Part", "Plant 7", "Wittmann", "W833")],
    )
    _save_off_machine_audit(fake_project, "AUD-OFF-ROBOT-ALIASES", "PN-OFF-ROBOT-ALIASES")

    result = apply_off_machine_compatibility_choice(
        fake_project,
        "AUD-OFF-ROBOT-ALIASES",
        OFF_MACHINE_COMPATIBILITY_UPDATE_AND_ADD,
    )

    assert result.success, result.errors
    row = next(row for row in _inventory_rows(fake_project) if row["Audit ID"] == "AUD-OFF-ROBOT-ALIASES")
    assert row["Robot Type"] == "Wittmann"
    assert row["Robot Model/Controller"] == "W833"


def test_off_machine_update_and_add_multiple_matches_creates_remaining_rows(fake_project):
    _write_detailed_capacity(
        fake_project,
        [
            ("7", "PN-OFF-2", "Capacity Part 7", "Plant 7", "Wittmann", "R9", "Cleanroom", "Family 7"),
            ("8", "PN-OFF-2", "Capacity Part 8", "Plant 8", "Engel", "Viper", "Non-Cleanroom", "Family 8"),
        ],
    )
    _save_off_machine_audit(fake_project, "AUD-OFF-MULTI", "PN-OFF-2")

    result = apply_off_machine_compatibility_choice(
        fake_project,
        "AUD-OFF-MULTI",
        OFF_MACHINE_COMPATIBILITY_UPDATE_AND_ADD,
    )

    assert result.success, result.errors
    assert result.metrics["current_row_updated"] is True
    assert result.metrics["created"] == 1
    rows = [row for row in _inventory_rows(fake_project) if row.get("Tool #") == "PN-OFF-2"]
    by_machine = {row["Press/Machine #"]: row for row in rows}
    assert set(by_machine) == {"7", "8"}
    assert by_machine["7"]["Audit ID"] == "AUD-OFF-MULTI"
    assert by_machine["8"][ENTRY_TYPE_FIELD] == ENTRY_TYPE_COMPATIBLE
    assert by_machine["8"][AUDIT_CONTEXT_FIELD] == AUDIT_CONTEXT_COMPATIBILITY
    assert by_machine["8"][PHYSICAL_AUDIT_VERIFIED_FIELD] == "No"
    assert by_machine["8"][COMPATIBILITY_CONFIDENCE_FIELD] == "Press Capacity"
    assert by_machine["8"][SOURCE_AUDIT_ID_FIELD] == "AUD-OFF-MULTI"
    assert by_machine["8"][COMPATIBILITY_SOURCE_FIELD] == COMPATIBILITY_SOURCE_PRESS_CAPACITY
    assert by_machine["8"][EOAT_ASSEMBLY_ID_FIELD] == "P4-EOAT-9001"


def test_off_machine_add_compatibility_rows_only_leaves_original_off_machine(fake_project):
    _write_detailed_capacity(
        fake_project,
        [
            ("7", "PN-OFF-3", "Capacity Part 7", "Plant 7", "Wittmann", "R9", "Cleanroom", "Family 7"),
            ("8", "PN-OFF-3", "Capacity Part 8", "Plant 8", "Engel", "Viper", "Non-Cleanroom", "Family 8"),
        ],
    )
    _save_off_machine_audit(fake_project, "AUD-OFF-ADD-ONLY", "PN-OFF-3")

    result = apply_off_machine_compatibility_choice(
        fake_project,
        "AUD-OFF-ADD-ONLY",
        OFF_MACHINE_COMPATIBILITY_ADD_ONLY,
    )

    assert result.success, result.errors
    assert result.metrics["current_row_updated"] is False
    assert result.metrics["created"] == 2
    rows = [row for row in _inventory_rows(fake_project) if row.get("Tool #") == "PN-OFF-3"]
    original = next(row for row in rows if row["Audit ID"] == "AUD-OFF-ADD-ONLY")
    assert original["Press/Machine #"] == "N/A"
    assert original[AUDIT_CONTEXT_FIELD] == AUDIT_CONTEXT_BENCH
    compatible_rows = [row for row in rows if row.get(ENTRY_TYPE_FIELD) == ENTRY_TYPE_COMPATIBLE]
    assert {row["Press/Machine #"] for row in compatible_rows} == {"7", "8"}
    assert {row[AUDIT_CONTEXT_FIELD] for row in compatible_rows} == {AUDIT_CONTEXT_COMPATIBILITY}
    assert {row[PHYSICAL_AUDIT_VERIFIED_FIELD] for row in compatible_rows} == {"No"}


def test_off_machine_leave_as_off_machine_makes_no_changes(fake_project):
    _write_detailed_capacity(
        fake_project,
        [("7", "PN-OFF-4", "Capacity Part", "Plant 7", "Wittmann", "R9", "Cleanroom", "Family 7")],
    )
    _save_off_machine_audit(fake_project, "AUD-OFF-LEAVE", "PN-OFF-4")

    result = apply_off_machine_compatibility_choice(
        fake_project,
        "AUD-OFF-LEAVE",
        OFF_MACHINE_COMPATIBILITY_LEAVE,
    )

    assert result.success, result.errors
    assert result.metrics["created"] == 0
    rows = [row for row in _inventory_rows(fake_project) if row.get("Tool #") == "PN-OFF-4"]
    assert len(rows) == 1
    assert rows[0]["Press/Machine #"] == "N/A"
    assert rows[0].get(ENTRY_TYPE_FIELD) == "Audited"


def test_off_machine_no_press_capacity_match_saves_without_compatibility_rows(fake_project):
    _write_detailed_capacity(
        fake_project,
        [("7", "DIFFERENT-TOOL", "Other Part", "Plant 7", "Wittmann", "R9", "Cleanroom", "Family 7")],
    )
    _save_off_machine_audit(fake_project, "AUD-OFF-NO-MATCH", "PN-OFF-NO-MATCH")

    result = apply_off_machine_compatibility_choice(
        fake_project,
        "AUD-OFF-NO-MATCH",
        OFF_MACHINE_COMPATIBILITY_UPDATE_AND_ADD,
    )

    assert result.success, result.errors
    assert "No compatible machines were found for Tool # PN-OFF-NO-MATCH" in result.summary
    rows = [row for row in _inventory_rows(fake_project) if row.get("Tool #") == "PN-OFF-NO-MATCH"]
    assert len(rows) == 1
    assert rows[0]["Press/Machine #"] == "N/A"
    assert rows[0]["Notes"].endswith("EOAT Not Installed.")


def test_off_machine_duplicate_prevention_updates_missing_machine_fields(fake_project):
    _write_detailed_capacity(
        fake_project,
        [("7", "PN-OFF-DUP", "Capacity Part", "Plant 7", "Wittmann", "R9", "Cleanroom", "Family 7")],
    )
    _save_off_machine_audit(fake_project, "AUD-OFF-DUP-SOURCE", "PN-OFF-DUP", **{EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-9002"})
    existing_result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-OFF-DUP-EXISTING",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "N/A",
            "Press/Machine #": "7",
            "Tool #": "PN-OFF-DUP",
            EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-9002",
            "Robot Type": "N/A",
            "Robot Model/Controller": "N/A",
            "EOAT Type": "Vacuum",
            "Status": "In Progress",
        },
    )
    assert existing_result.success, existing_result.errors

    result = apply_off_machine_compatibility_choice(
        fake_project,
        "AUD-OFF-DUP-SOURCE",
        OFF_MACHINE_COMPATIBILITY_ADD_ONLY,
    )

    assert result.success, result.errors
    assert result.metrics["created"] == 0
    assert result.metrics["duplicates_skipped"] == 1
    assert result.metrics["existing_rows_updated"] == 1
    rows = [row for row in _inventory_rows(fake_project) if row.get("Tool #") == "PN-OFF-DUP" and row.get("Press/Machine #") == "7"]
    assert len(rows) == 1
    assert rows[0]["Audit ID"] == "AUD-OFF-DUP-EXISTING"
    assert rows[0]["Robot Type"] == "Wittmann"
    assert rows[0]["Plant/Area"] == "Plant 7"


def test_off_machine_update_preserves_non_empty_eoat_audit_fields(fake_project):
    _write_detailed_capacity(
        fake_project,
        [("7", "PN-OFF-PRESERVE", "Capacity Part", "Plant 7", "Wittmann", "R9", "Cleanroom", "Family 7")],
    )
    _save_off_machine_audit(
        fake_project,
        "AUD-OFF-PRESERVE",
        "PN-OFF-PRESERVE",
        **{
            "Known Issues": "User-entered issue stays.",
            "EOAT Type": "Hybrid",
            "Part Name/Description": "User-entered part description",
        },
    )

    result = apply_off_machine_compatibility_choice(
        fake_project,
        "AUD-OFF-PRESERVE",
        OFF_MACHINE_COMPATIBILITY_UPDATE_AND_ADD,
    )

    assert result.success, result.errors
    row = next(row for row in _inventory_rows(fake_project) if row["Audit ID"] == "AUD-OFF-PRESERVE")
    assert row["Press/Machine #"] == "7"
    assert row["Robot Type"] == "Wittmann"
    assert row["Known Issues"] == "User-entered issue stays."
    assert row["EOAT Type"] == "Hybrid"
    assert row["Part Name/Description"] == "User-entered part description"


def test_off_machine_tool_number_normalization_preserves_source_tool_display(fake_project):
    long_tool = "12345678901234567890"
    assert normalize_tool_identifier(f"{long_tool}.0") == long_tool
    _write_detailed_capacity(
        fake_project,
        [("7", f"{long_tool}.0", "Long Tool Part", "Plant 7", "Wittmann", "R9", "Cleanroom", "Family 7")],
    )
    _save_off_machine_audit(fake_project, "AUD-OFF-NORMALIZE", long_tool)

    result = apply_off_machine_compatibility_choice(
        fake_project,
        "AUD-OFF-NORMALIZE",
        OFF_MACHINE_COMPATIBILITY_UPDATE_AND_ADD,
    )

    assert result.success, result.errors
    row = next(row for row in _inventory_rows(fake_project) if row["Audit ID"] == "AUD-OFF-NORMALIZE")
    assert row["Press/Machine #"] == "7"
    assert row["Tool #"] == long_tool


def test_off_machine_press_capacity_load_failure_warns_and_leaves_audit_saved(fake_project, tmp_path):
    _save_off_machine_audit(fake_project, "AUD-OFF-LOAD-FAIL", "PN-OFF-LOAD-FAIL")
    missing_capacity = tmp_path / "missing_press_capacity.xlsx"

    result = apply_off_machine_compatibility_choice(
        fake_project,
        "AUD-OFF-LOAD-FAIL",
        OFF_MACHINE_COMPATIBILITY_UPDATE_AND_ADD,
        press_capacity_path=missing_capacity,
    )

    assert result.success, result.errors
    assert result.metrics["press_capacity_lookup_failed"] is True
    assert any("Fit Check lookup could not be completed" in warning for warning in result.warnings)
    row = next(row for row in _inventory_rows(fake_project) if row["Audit ID"] == "AUD-OFF-LOAD-FAIL")
    assert row["Press/Machine #"] == "N/A"
    activity, warning = read_recent_activity(fake_project, limit=10)
    assert warning is None
    assert any(item.get("event_name") == "off_machine_compatibility_lookup_completed" for item in activity)


def test_progress_separates_covered_remaining_and_missing_reasons(fake_project):
    _write_press_capacity(fake_project, [("1, 2, 3", "PN-X", "Part X"), ("4", "PN-Y", "Part Y")])
    _save_audit(fake_project, "AUD-SOURCE-001", "1", "PN-X")
    assert create_compatibility_entries(fake_project, "AUD-SOURCE-001", ["2"]).success

    summary, error = calculate_audit_progress(fake_project)

    assert error is None
    assert summary.metrics["required_relationships"] == 4
    assert summary.metrics["physically_audited_relationships"] == 1
    assert summary.metrics["compatible_relationships"] == 1
    assert summary.metrics["total_covered_relationships"] == 2
    assert summary.metrics["remaining_relationships"] == 2
    missing_actions = {
        (row["Machine No."], row["NGW Part Number"]): row["Suggested Next Action"]
        for row in summary.missing_relationships
    }
    assert missing_actions[("3", "PN-X")] == "Use Fit Check Entry"
    assert missing_actions[("4", "PN-Y")] == "Needs Physical Audit"


def test_saving_audited_row_syncs_linked_compatible_rows_and_preserves_identity(fake_project):
    _write_press_capacity(fake_project, [("1, 2, 3", "PN-X", "Part X")])
    _save_audit(fake_project, "AUD-SOURCE-001", "1", "PN-X", description="Original part")
    create_result = create_compatibility_entries(fake_project, "AUD-SOURCE-001", ["2", "3"])
    assert create_result.success, create_result.errors

    before_rows = _inventory_rows(fake_project)
    compatible_before = [row for row in before_rows if row.get(ENTRY_TYPE_FIELD) == ENTRY_TYPE_COMPATIBLE]
    before_ids = {row["Press/Machine #"]: row["Audit ID"] for row in compatible_before}
    before_row_count = len(before_rows)

    update_result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-SOURCE-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "1",
            "Tool #": "PN-X",
            "Part Name/Description": "Updated part description",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Hybrid",
            "EOAT Moves": "Both",
            "Cup Type/Material": "Urethane",
            "Gripper Model": "Zimmer GPP",
            "Known Issues": "Updated source issue.",
            "Drop/Mis-Pick History": "Two recent drops",
            "Maintenance Frequency": "Monthly",
            "Estimated EOAT Weight": "12 lb",
            "Changeover Difficulty": "High",
            "Status": "Needs Follow-Up",
            "Priority": "High",
            "Pilot Candidate?": "Maybe",
            "Notes": "Updated shared note.",
            "Entry Type": "Audited",
        },
        allow_update=True,
        sync_linked_compatibility=True,
        refresh_press_view=True,
    )

    assert update_result.success, update_result.errors
    assert update_result.metrics["compatibility_rows_synced"] == 2
    assert "Updated 2 linked compatibility" in update_result.summary
    after_rows = _inventory_rows(fake_project)
    assert len(after_rows) == before_row_count
    compatible_after = [row for row in after_rows if row.get(ENTRY_TYPE_FIELD) == ENTRY_TYPE_COMPATIBLE]
    assert len(compatible_after) == 2
    for row in compatible_after:
        assert row["Audit ID"] == before_ids[row["Press/Machine #"]]
        assert row["Press/Machine #"] in {"2", "3"}
        assert row[ENTRY_TYPE_FIELD] == ENTRY_TYPE_COMPATIBLE
        assert row[SOURCE_AUDIT_ID_FIELD] == "AUD-SOURCE-001"
        assert row["Part Name/Description"] == "Updated part description"
        assert row["EOAT Type"] == "Hybrid"
        assert row["EOAT Moves"] == "Both"
        assert row["Cup Type/Material"] == "Urethane"
        assert row["Known Issues"] == "Updated source issue."
        assert row["Drop/Mis-Pick History"] == "Two recent drops"
        assert row["Maintenance Frequency"] == "Monthly"
        assert row["Estimated EOAT Weight"] == "12 lb"
        assert row["Changeover Difficulty"] == "High"
        assert row["Status"] == "Needs Follow-Up"
        assert row["Priority"] == "High"
        assert row["Pilot Candidate?"] == "Maybe"
        assert row["Notes"] == "Updated shared note."


def test_saving_audited_row_with_no_linked_compatibility_reports_zero(fake_project):
    _save_audit(fake_project, "AUD-NO-LINKS-001", "1", "PN-X")

    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-NO-LINKS-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "1",
            "Tool #": "PN-X",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Known Issues": "No linked rows.",
            "Status": "Audited",
            "Entry Type": "Audited",
        },
        allow_update=True,
    )

    assert result.success, result.errors
    assert result.metrics["compatibility_rows_synced"] == 0
    assert "No linked compatibility entries found" not in result.summary
    assert "Linked compatibility rows may need review." in result.warnings


def test_saving_compatible_row_does_not_recursively_sync_linked_rows(fake_project):
    _write_press_capacity(fake_project, [("1, 2, 3", "PN-X", "Part X")])
    _save_audit(fake_project, "AUD-SOURCE-001", "1", "PN-X")
    assert create_compatibility_entries(fake_project, "AUD-SOURCE-001", ["2", "3"]).success
    rows = _inventory_rows(fake_project)
    compatible_two = next(
        row for row in rows if row.get(ENTRY_TYPE_FIELD) == ENTRY_TYPE_COMPATIBLE and row["Press/Machine #"] == "2"
    )
    compatible_three_before = next(
        row for row in rows if row.get(ENTRY_TYPE_FIELD) == ENTRY_TYPE_COMPATIBLE and row["Press/Machine #"] == "3"
    )

    result = save_audit_entry(
        fake_project,
        {
            **compatible_two,
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Known Issues": "Manual compatibility note only.",
            "Entry Type": ENTRY_TYPE_COMPATIBLE,
        },
        allow_update=True,
    )

    assert result.success, result.errors
    assert result.metrics["compatibility_rows_synced"] == 0
    after_rows = _inventory_rows(fake_project)
    compatible_three_after = next(row for row in after_rows if row["Audit ID"] == compatible_three_before["Audit ID"])
    assert compatible_three_after["Known Issues"] == compatible_three_before["Known Issues"]


def test_sync_does_not_overwrite_different_source_compatible_rows(fake_project):
    _write_press_capacity(fake_project, [("1, 2", "PN-X", "Part X")])
    _save_audit(fake_project, "AUD-SOURCE-001", "1", "PN-X")
    _save_audit(fake_project, "AUD-SOURCE-OTHER", "4", "PN-X")
    _save_audit(
        fake_project,
        "AUD-COMP-DIFFERENT",
        "2",
        "PN-X",
        description="Different source child",
        entry_type=ENTRY_TYPE_COMPATIBLE,
        source_id="AUD-SOURCE-OTHER",
    )

    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-SOURCE-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "1",
            "Tool #": "PN-X",
            "Part Name/Description": "Should not touch different source",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Hybrid",
            "Known Issues": "Source changed.",
            "Status": "Audited",
            "Entry Type": "Audited",
        },
        allow_update=True,
    )

    assert result.success, result.errors
    assert result.metrics["compatibility_rows_synced"] == 0
    different_source = next(row for row in _inventory_rows(fake_project) if row["Audit ID"] == "AUD-COMP-DIFFERENT")
    assert different_source[SOURCE_AUDIT_ID_FIELD] == "AUD-SOURCE-OTHER"
    assert different_source["Part Name/Description"] == "Different source child"
    assert different_source["EOAT Type"] == "Vacuum"


def test_compatibility_candidates_distinguish_linked_and_different_sources(fake_project):
    _write_press_capacity(fake_project, [("1, 2, 3", "PN-X", "Part X")])
    _save_audit(fake_project, "AUD-SOURCE-001", "1", "PN-X")
    _save_audit(fake_project, "AUD-SOURCE-OTHER", "4", "PN-X")
    assert create_compatibility_entries(fake_project, "AUD-SOURCE-001", ["2"]).success
    _save_audit(
        fake_project, "AUD-COMP-DIFFERENT", "3", "PN-X", entry_type=ENTRY_TYPE_COMPATIBLE, source_id="AUD-SOURCE-OTHER"
    )

    result = build_compatibility_candidates(fake_project, "AUD-SOURCE-001")
    actions = {candidate.machine_no: candidate.recommended_action for candidate in result.candidates}

    assert actions["2"] == "Already Compatible - Linked to this source"
    assert actions["3"] == "Already Compatible - Different Source / Review Needed"


def test_sync_preserves_progress_counts_and_workbook_structure(fake_project):
    _write_press_capacity(fake_project, [("1, 2", "PN-X", "Part X")])
    _save_audit(fake_project, "AUD-SOURCE-001", "1", "PN-X")
    assert create_compatibility_entries(fake_project, "AUD-SOURCE-001", ["2"]).success
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    sheet_names_before = workbook.sheetnames[:]
    headers_before = [cell.value for cell in ws[1]]
    row_count_before = ws.max_row
    style_before = ws.cell(row=3, column=headers_before.index("Known Issues") + 1).style_id
    workbook.close()

    sync_result = sync_compatible_rows_from_source(workbook_path, "AUD-SOURCE-001")
    assert sync_result.updated_count == 1

    summary, error = calculate_audit_progress(fake_project)
    assert error is None
    assert summary.metrics["physically_audited_relationships"] == 1
    assert summary.metrics["compatible_relationships"] == 1
    assert summary.metrics["physical_audit_rows"] == 1
    assert summary.metrics["compatibility_rows"] == 1

    workbook = load_workbook(workbook_path)
    try:
        ws = workbook["EOAT Inventory"]
        assert workbook.sheetnames == sheet_names_before
        assert [cell.value for cell in ws[1]] == headers_before
        assert ws.max_row == row_count_before
        assert ws.cell(row=3, column=headers_before.index("Known Issues") + 1).style_id == style_before
    finally:
        workbook.close()
