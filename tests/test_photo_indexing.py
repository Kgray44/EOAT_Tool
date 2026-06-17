from __future__ import annotations

from openpyxl import load_workbook

from core.eoat_ids import EOAT_ASSEMBLY_ID_FIELD
from core.paths import resolve_project_paths
from core.photo_evidence import photo_index_path_findings
from core.photo_indexing import (
    convert_legacy_photo_tree_to_eoat_folders,
    destination_folder,
    eoat_photo_root,
    intake_photos,
    list_incoming_photos,
    normalize_cell_photo_folders,
    preview_photo_intake,
    repair_audit_photo_ties,
    repair_photo_eoat_links,
    tool_photo_folder_name,
    tool_photo_root,
)
from core.workbook_cache import invalidate_workbook_cache
from core.workbook_io import row_dicts
from core.workbook_schema import get_expected_headers


def _append_inventory_row(project_root, values: dict[str, object]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    row = {header: "" for header in get_expected_headers("EOAT Inventory")}
    row.update(values)
    ws.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()
    invalidate_workbook_cache(workbook_path)


def _delete_sheet_column(project_root, sheet_name: str, header: str) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook[sheet_name]
    headers = [cell.value for cell in ws[1]]
    if header in headers:
        ws.delete_cols(headers.index(header) + 1)
    workbook.save(workbook_path)
    workbook.close()
    invalidate_workbook_cache(workbook_path)


def test_photo_preview_and_copy_intake(fake_project):
    incoming = resolve_project_paths(fake_project).incoming_photos
    incoming.mkdir(parents=True)
    photo = incoming / "IMG_0001.jpg"
    photo.write_bytes(b"fake jpg")

    assert list_incoming_photos(fake_project) == [photo]
    plan = preview_photo_intake(
        fake_project,
        [photo],
        "Plant 4",
        "Press 12",
        "2026-05-18",
        "Tubing Routing",
        tool_number="6200171020",
        part_name="Widget Part A",
    )

    assert plan[0].target.name == "Tool_6200171020__Tubing_Routing__2026-05-18__001.jpg"
    assert plan[0].target.parent == (
        resolve_project_paths(fake_project).cell_photos
        / "Tool_6200171020__Widget_Part_A"
        / "04_Tubing_Routing"
    )

    result = intake_photos(
        fake_project,
        [photo],
        "Plant 4",
        "Press 12",
        "2026-05-18",
        "Tubing Routing",
        tool_number="6200171020",
        part_name="Widget Part A",
        copy_mode=True,
    )

    assert result.success is True
    assert photo.exists()
    assert (
        destination_folder(fake_project, "Tubing Routing", "6200171020", "Widget Part A")
        / "Tool_6200171020__Tubing_Routing__2026-05-18__001.jpg"
    ).exists()
    wb = load_workbook(resolve_project_paths(fake_project).master_workbook, read_only=True)
    ws = wb["Photo Index"]
    assert ws.max_row == 2
    wb.close()


def test_photo_preview_requires_tool_for_new_storage(fake_project):
    incoming = resolve_project_paths(fake_project).incoming_photos
    incoming.mkdir(parents=True)
    photo = incoming / "IMG_0001.jpg"
    photo.write_bytes(b"fake jpg")

    plan = preview_photo_intake(fake_project, [photo], "Plant 4", "12", "2026-05-18", "Front View")

    assert plan == []


def test_photo_filename_omits_machine_when_tool_is_off_machine(fake_project):
    incoming = resolve_project_paths(fake_project).incoming_photos
    incoming.mkdir(parents=True)
    photo = incoming / "IMG_0007.jpg"
    photo.write_bytes(b"fake jpg")

    plan = preview_photo_intake(fake_project, [photo], "Whiteroom", "", "2026-05-18", "Front View", tool_number="TOOL-12")

    assert plan[0].target.name == "Tool_TOOL-12__Front_View__2026-05-18__001.jpg"
    assert "MachineUnknown" not in plan[0].target.name


def test_photo_collision_safe_filename(fake_project):
    folder = destination_folder(fake_project, "Sensors", "6200171020")
    folder.mkdir(parents=True)
    (folder / "Tool_6200171020__Sensors__2026-05-18__001.jpg").write_bytes(b"existing")
    photo = fake_project / "photo.jpg"
    photo.write_bytes(b"new")

    plan = preview_photo_intake(
        fake_project,
        [photo],
        "Plant 4",
        "Machine 12",
        "2026-05-18",
        "Sensors",
        tool_number="6200171020",
    )

    assert plan[0].target.name == "Tool_6200171020__Sensors__2026-05-18__002.jpg"


def test_photo_move_mode_heic_and_unsupported_extension_handling(fake_project):
    incoming = resolve_project_paths(fake_project).incoming_photos
    incoming.mkdir(parents=True)
    heic = incoming / "IMG_0002.HEIC"
    heic.write_bytes(b"fake heic")
    unsupported = incoming / "notes.txt"
    unsupported.write_text("not an image", encoding="utf-8")

    assert list_incoming_photos(fake_project) == [heic]
    plan = preview_photo_intake(
        fake_project, [heic, unsupported], "Plant 4", "Press 12", "2026-05-18", "Front View", tool_number="TOOL-12"
    )

    assert len(plan) == 1
    assert plan[0].target.suffix == ".HEIC"

    result = intake_photos(
        fake_project,
        [heic],
        "Plant 4",
        "Press 12",
        "2026-05-18",
        "Front View",
        tool_number="TOOL-12",
        copy_mode=False,
    )

    assert result.success is True
    assert not heic.exists()
    assert plan[0].target.exists()


def test_default_photo_setup_creates_only_incoming_photos(tmp_path):
    import setup_eoat_project

    project_root = tmp_path / "EOAT_Standardization_Project"
    setup_eoat_project.configure_project_root(project_root)
    setup_eoat_project.create_folders()

    cell_photos = project_root / "01_EOAT_Audit" / "Cell_Photos"
    children = sorted(path.name for path in cell_photos.iterdir())

    assert children == ["Incoming_Photos"]


def test_tool_photo_category_folders_are_created_only_as_photos_arrive(fake_project):
    incoming = resolve_project_paths(fake_project).incoming_photos
    incoming.mkdir(parents=True)
    first = incoming / "IMG_0101.jpg"
    second = incoming / "IMG_0102.jpg"
    first.write_bytes(b"first")
    second.write_bytes(b"second")

    first_result = intake_photos(
        fake_project,
        [first],
        "Whiteroom",
        "",
        "2026-06-09",
        "Front View",
        tool_number="12345",
        part_name="Clean Part",
        copy_mode=False,
        log_activity=False,
    )
    tool_root = tool_photo_root(fake_project, "12345", "Clean Part")
    assert sorted(path.name for path in tool_root.iterdir() if path.is_dir()) == ["01_Front_View"]
    marker = tool_root / "do_not_touch.txt"
    marker.write_text("keep", encoding="utf-8")
    second_result = intake_photos(
        fake_project,
        [second],
        "Whiteroom",
        "",
        "2026-06-09",
        "Side View",
        tool_number="12345",
        part_name="Clean Part",
        copy_mode=False,
        log_activity=False,
    )

    assert first_result.success is True
    assert second_result.success is True
    assert sorted(path.name for path in tool_root.iterdir() if path.is_dir()) == ["01_Front_View", "02_Side_View"]
    assert marker.read_text(encoding="utf-8") == "keep"
    assert (tool_root / "01_Front_View" / "Tool_12345__Front_View__2026-06-09__001.jpg").exists()
    assert (tool_root / "02_Side_View" / "Tool_12345__Side_View__2026-06-09__001.jpg").exists()
    assert not (tool_root / "03_Vacuum_Cups_Grippers").exists()
    assert not (tool_root / "09_Wear_Damage").exists()


def test_added_photo_type_options_route_to_lazy_storage_folders(fake_project):
    incoming = resolve_project_paths(fake_project).incoming_photos
    incoming.mkdir(parents=True, exist_ok=True)
    back = incoming / "IMG_BACK.jpg"
    tool_number = incoming / "IMG_TOOL.jpg"
    gripper = incoming / "IMG_GRIP.jpg"
    back.write_bytes(b"back")
    tool_number.write_bytes(b"tool")
    gripper.write_bytes(b"grip")

    result = intake_photos(
        fake_project,
        [back, tool_number, gripper],
        "Whiteroom",
        "",
        "2026-06-09",
        "Front View",
        tool_number="12345",
        per_photo_metadata=[
            {"source": str(back), "view_type": "Back View"},
            {"source": str(tool_number), "view_type": "Tool Number"},
            {"source": str(gripper), "view_type": "Gripper"},
        ],
        copy_mode=False,
        log_activity=False,
    )
    root = tool_photo_root(fake_project, "12345")

    assert result.success is True
    assert sorted(path.name for path in root.iterdir() if path.is_dir()) == [
        "03_Vacuum_Cups_Grippers",
        "10_Back_View",
        "11_Tool_Number",
    ]
    assert (root / "10_Back_View" / "Tool_12345__Back_View__2026-06-09__001.jpg").exists()
    assert (root / "11_Tool_Number" / "Tool_12345__Tool_Number__2026-06-09__001.jpg").exists()
    assert (root / "03_Vacuum_Cups_Grippers" / "Tool_12345__Vacuum_Cups_Grippers__2026-06-09__001.jpg").exists()


def test_tool_photo_folder_name_safely_cleans_part_names():
    folder_name = tool_photo_folder_name("Tool 12/..\\Bad", " Part: A/B * Rev 2 ")

    assert folder_name == "Tool_Tool_12_Bad__Part_A_B_Rev_2"
    assert "/" not in folder_name
    assert "\\" not in folder_name
    assert ".." not in folder_name
    assert "  " not in folder_name


def test_photo_sequence_increments_and_existing_file_is_not_overwritten(fake_project):
    folder = destination_folder(fake_project, "Tubing Routing", "12345")
    folder.mkdir(parents=True)
    existing = folder / "Tool_12345__Tubing_Routing__2026-06-09__001.jpg"
    existing.write_bytes(b"existing")
    incoming = resolve_project_paths(fake_project).incoming_photos
    incoming.mkdir(parents=True, exist_ok=True)
    photo = incoming / "IMG_0110.jpg"
    photo.write_bytes(b"new")

    result = intake_photos(
        fake_project,
        [photo],
        "Whiteroom",
        "",
        "2026-06-09",
        "Tubing Routing",
        tool_number="12345",
        copy_mode=False,
        log_activity=False,
    )

    assert result.success is True
    assert existing.read_bytes() == b"existing"
    assert (folder / "Tool_12345__Tubing_Routing__2026-06-09__002.jpg").exists()


def test_failed_import_does_not_leave_empty_tool_or_category_folders(fake_project, monkeypatch):
    import core.photo_indexing as photo_indexing

    incoming = resolve_project_paths(fake_project).incoming_photos
    incoming.mkdir(parents=True, exist_ok=True)
    photo = incoming / "IMG_0120.jpg"
    photo.write_bytes(b"new")

    def fail_copy(*_args, **_kwargs):
        raise OSError("simulated copy failure")

    monkeypatch.setattr(photo_indexing, "safe_copy_file", fail_copy)

    result = intake_photos(
        fake_project,
        [photo],
        "Whiteroom",
        "",
        "2026-06-09",
        "Front View",
        tool_number="FAIL-TOOL",
        copy_mode=True,
        log_activity=False,
    )

    assert result.success is False
    assert not tool_photo_root(fake_project, "FAIL-TOOL").exists()


def test_old_empty_top_level_photo_folders_are_removed_but_non_empty_are_preserved(fake_project):
    cell_photos = resolve_project_paths(fake_project).cell_photos
    empty_legacy = cell_photos / "Front_View"
    non_empty_legacy = cell_photos / "Sensors"
    incoming = resolve_project_paths(fake_project).incoming_photos
    empty_legacy.mkdir(parents=True)
    non_empty_legacy.mkdir(parents=True)
    (non_empty_legacy / "keep.jpg").write_bytes(b"keep")

    removed, preserved = normalize_cell_photo_folders(fake_project)

    assert empty_legacy in removed
    assert not empty_legacy.exists()
    assert non_empty_legacy in preserved
    assert non_empty_legacy.exists()
    assert incoming.exists()


def test_photo_intake_updates_related_audit_row_and_structured_photo_fields(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-PHOTO-UPDATE-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "TOOL-12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Mechanical / Gripper",
            "Status": "Complete",
            "Photos Taken?": "No",
            "Photo Folder/Link": "",
        },
    )
    incoming = resolve_project_paths(fake_project).incoming_photos
    incoming.mkdir(parents=True)
    photo = incoming / "IMG_0003.jpg"
    photo.write_bytes(b"fake jpg")

    result = intake_photos(
        fake_project,
        [photo],
        "Plant 4",
        "Press 12",
        "2026-05-18",
        "Grippers",
        tool_number="TOOL-12",
        related_audit_id="AUD-PHOTO-UPDATE-001",
        linked_audit_field="Gripper Model",
        copy_mode=True,
        log_activity=False,
    )

    workbook_path = resolve_project_paths(fake_project).master_workbook
    inventory = row_dicts(workbook_path, "EOAT Inventory")
    row = next(row for row in inventory if row["Audit ID"] == "AUD-PHOTO-UPDATE-001")
    photos = row_dicts(workbook_path, "Photo Index")
    photo_row = next(row for row in photos if row["Related Audit ID"] == "AUD-PHOTO-UPDATE-001")

    assert result.success is True
    assert result.metrics["audit_rows_updated"] == 1
    assert any(path.endswith(".xlsx") for path in result.files_created)
    assert row["Photos Taken?"] == "Yes"
    assert "Tool_TOOL-12" in str(row["Photo Folder/Link"])
    assert "03_Vacuum_Cups_Grippers" in str(row["Photo Folder/Link"])
    assert photo_row["Tool #"] == "TOOL-12"
    assert photo_row["Part Name"] in {"", None}
    assert photo_row["Original Filename"] == "IMG_0003.jpg"
    assert photo_row["Stored Filename"] == photo_row["Photo Filename"]
    assert "Tool_TOOL-12" in photo_row["Stored Relative Path"]
    assert photo_row["Linked Audit Field"] == "Gripper Model"


def test_photo_intake_allows_tool_only_when_machine_is_blank(fake_project):
    incoming = resolve_project_paths(fake_project).incoming_photos
    incoming.mkdir(parents=True)
    photo = incoming / "IMG_0010.jpg"
    photo.write_bytes(b"fake jpg")

    result = intake_photos(
        fake_project,
        [photo],
        "Cleanroom",
        "",
        "2026-05-18",
        "Front View",
        tool_number="LOOSE-TOOL",
        copy_mode=True,
        log_activity=False,
    )

    rows = row_dicts(resolve_project_paths(fake_project).master_workbook, "Photo Index")
    photo_row = next(row for row in rows if row["Tool #"] == "LOOSE-TOOL")

    assert result.success is True
    assert "Tool_LOOSE-TOOL__Front_View__2026-05-18__001" in result.details[0]
    assert photo_row["Press/Machine #"] in {"", None}


def test_photo_intake_requires_tool_or_machine_context(fake_project):
    incoming = resolve_project_paths(fake_project).incoming_photos
    incoming.mkdir(parents=True)
    photo = incoming / "IMG_0011.jpg"
    photo.write_bytes(b"fake jpg")

    result = intake_photos(
        fake_project,
        [photo],
        "Cleanroom",
        "",
        "2026-05-18",
        "Front View",
        copy_mode=True,
        log_activity=False,
    )

    assert result.success is False
    assert "Tool #" in result.summary


def test_photo_intake_appends_photo_folder_link_once(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-PHOTO-APPEND-001",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "TOOL-12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Mechanical / Gripper",
            "Photos Taken?": "No",
            "Photo Folder/Link": "existing/reference",
        },
    )
    incoming = resolve_project_paths(fake_project).incoming_photos
    incoming.mkdir(parents=True)
    first = incoming / "IMG_0005.jpg"
    second = incoming / "IMG_0006.jpg"
    first.write_bytes(b"fake jpg")
    second.write_bytes(b"fake jpg")

    for photo in [first, second]:
        result = intake_photos(
            fake_project,
            [photo],
            "Plant 4",
            "Press 12",
            "2026-05-18",
            "Grippers",
            tool_number="TOOL-12",
            related_audit_id="AUD-PHOTO-APPEND-001",
            copy_mode=True,
            log_activity=False,
        )
        assert result.success is True

    row = next(
        row
        for row in row_dicts(resolve_project_paths(fake_project).master_workbook, "EOAT Inventory")
        if row["Audit ID"] == "AUD-PHOTO-APPEND-001"
    )
    folder = str(destination_folder(fake_project, "Grippers", "TOOL-12").relative_to(fake_project))
    lines = [line.strip() for line in str(row["Photo Folder/Link"]).splitlines() if line.strip()]

    assert lines[0] == "existing/reference"
    assert lines.count(folder) == 1


def test_repair_audit_photo_ties_repairs_index_paths_and_inventory_link(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-PHOTO-REPAIR-001",
            "Plant/Area": "Whiteroom",
            "Press/Machine #": "Press 42",
            "Tool #": "TOOL-REPAIR",
            "Part Name/Description": "Repair Part",
            "EOAT Type": "Vacuum",
            "Status": "Complete",
            "Photos Taken?": "No",
            "Photo Folder/Link": "01_EOAT_Audit/Cell_Photos/Overall",
        },
    )
    folder = destination_folder(fake_project, "Front View", "TOOL-REPAIR", "Repair Part")
    folder.mkdir(parents=True)
    photo = folder / "Tool_TOOL-REPAIR__Front_View__2026-06-09__001.jpg"
    photo.write_bytes(b"fake jpg")
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["Photo Index"]
    headers = [cell.value for cell in ws[1]]
    values = {
        "Photo ID": "PHO-REPAIR-001",
        "Tool #": "TOOL-REPAIR",
        "EOAT Area Shown": "Front View",
        "Stored Filename": photo.name,
        "Photo Filename": photo.name,
        "Stored Relative Path": "missing/file.jpg",
        "Folder Path": "missing/folder",
        "Related Audit ID": "AUD-PHOTO-REPAIR-001",
    }
    ws.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()
    invalidate_workbook_cache(workbook_path)

    result = repair_audit_photo_ties(fake_project, log_activity=False)

    inventory = row_dicts(workbook_path, "EOAT Inventory")
    audit_row = next(row for row in inventory if row["Audit ID"] == "AUD-PHOTO-REPAIR-001")
    photos = row_dicts(workbook_path, "Photo Index")
    photo_row = next(row for row in photos if row["Photo ID"] == "PHO-REPAIR-001")
    expected_photo_path = str(photo.resolve().relative_to(fake_project.resolve()))
    expected_folder_path = str(folder.resolve().relative_to(fake_project.resolve()))

    assert result.success is True
    assert result.metrics["photo_rows_repaired"] == 1
    assert result.metrics["audit_rows_updated"] == 1
    assert audit_row["Photos Taken?"] == "Yes"
    assert "Overall" not in str(audit_row["Photo Folder/Link"])
    assert expected_folder_path in str(audit_row["Photo Folder/Link"])
    assert photo_row["Stored Relative Path"] == expected_photo_path
    assert photo_row["Folder Path"] == expected_folder_path
    assert photo_row["Stored Filename"] == photo.name


def test_photo_intake_warns_when_related_audit_id_is_missing(fake_project):
    incoming = resolve_project_paths(fake_project).incoming_photos
    incoming.mkdir(parents=True)
    photo = incoming / "IMG_0004.jpg"
    photo.write_bytes(b"fake jpg")

    result = intake_photos(
        fake_project,
        [photo],
        "Plant 4",
        "Press 12",
        "2026-05-18",
        "Front View",
        tool_number="TOOL-404",
        related_audit_id="AUD-DOES-NOT-EXIST",
        copy_mode=True,
        log_activity=False,
    )

    assert result.success is True
    assert any("Photo Index was updated" in warning and "AUD-DOES-NOT-EXIST" in warning for warning in result.warnings)
    workbook_path = resolve_project_paths(fake_project).master_workbook
    assert any(row["Related Audit ID"] == "AUD-DOES-NOT-EXIST" for row in row_dicts(workbook_path, "Photo Index"))
    assert not any(row.get("Audit ID") == "AUD-DOES-NOT-EXIST" for row in row_dicts(workbook_path, "EOAT Inventory"))


def test_per_photo_classification_routes_to_each_shot_folder(fake_project):
    incoming = resolve_project_paths(fake_project).incoming_photos
    incoming.mkdir(parents=True)
    first = incoming / "IMG_1001.jpg"
    second = incoming / "IMG_1002.jpg"
    first.write_bytes(b"fake jpg")
    second.write_bytes(b"fake jpg")

    result = intake_photos(
        fake_project,
        [first, second],
        "Plant 4",
        "Press 12",
        "2026-05-18",
        "Front View",
        tool_number="TOOL-12",
        per_photo_metadata=[
            {"source": str(first), "view_type": "Front View", "description": "Front evidence."},
            {"source": str(second), "view_type": "Sensors", "description": "Sensor evidence."},
        ],
        copy_mode=True,
        log_activity=False,
    )
    rows = row_dicts(resolve_project_paths(fake_project).master_workbook, "Photo Index")
    filenames = {row["EOAT Area Shown"]: row["Photo Filename"] for row in rows}

    assert result.success is True
    assert (destination_folder(fake_project, "Front View", "TOOL-12") / filenames["Front View"]).exists()
    assert (destination_folder(fake_project, "Sensors", "TOOL-12") / filenames["Sensors"]).exists()
    assert "Front_View" in filenames["Front View"]
    assert "Sensors" in filenames["Sensors"]
    assert all("Tool_TOOL-12" in filename for filename in filenames.values())


def test_photo_index_without_tool_column_remains_valid_for_legacy_rows(fake_project):
    _delete_sheet_column(fake_project, "Photo Index", "Tool #")
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-LEGACY-PHOTO-001",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "TOOL-12",
            "EOAT Type": "Vacuum",
            "Photos Taken?": "Yes",
        },
    )
    folder = destination_folder(fake_project, "Overall")
    folder.mkdir(parents=True, exist_ok=True)
    photo = folder / "Plant4_Press12_EOAT_2026-05-18_Overall_001.jpg"
    photo.write_bytes(b"legacy")
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["Photo Index"]
    headers = [cell.value for cell in ws[1]]
    values = {
        "Photo ID": "PHO-LEGACY-001",
        "Date Taken": "2026-05-18",
        "Plant/Area": "Plant 4",
        "Press/Machine #": "Press 12",
        "EOAT Area Shown": "Overall",
        "Photo Filename": photo.name,
        "Folder Path": str(folder),
        "Related Audit ID": "AUD-LEGACY-PHOTO-001",
    }
    ws.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()
    invalidate_workbook_cache(workbook_path)

    findings = photo_index_path_findings(fake_project)

    assert not any("Tool #" in finding.message and "missing" in finding.message.casefold() for finding in findings)


def test_photo_index_validation_flags_relationship_and_tool_mismatch(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-TOOL-MISMATCH-001",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "TOOL-A",
            "EOAT Type": "Vacuum",
        },
    )
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-PATH-CHECK-001",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 13",
            "Tool #": "TOOL-C",
            "EOAT Type": "Vacuum",
        },
    )
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["Photo Index"]
    headers = [cell.value for cell in ws[1]]
    for values in [
        {
            "Photo ID": "PHO-MISMATCH-001",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "TOOL-B",
            "EOAT Area Shown": "Overall",
            "Photo Filename": "missing.jpg",
            "Folder Path": str(destination_folder(fake_project, "Overall")),
            "Related Audit ID": "AUD-TOOL-MISMATCH-001",
        },
        {
            "Photo ID": "PHO-NO-AUDIT-001",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 14",
            "Tool #": "TOOL-D",
            "EOAT Area Shown": "Overall",
            "Photo Filename": "missing.jpg",
            "Folder Path": str(destination_folder(fake_project, "Overall")),
            "Related Audit ID": "AUD-MISSING-RELATIONSHIP",
        },
        {
            "Photo ID": "PHO-MISSING-FOLDER-001",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 13",
            "Tool #": "TOOL-C",
            "EOAT Area Shown": "Overall",
            "Photo Filename": "",
            "Folder Path": "",
            "Related Audit ID": "AUD-PATH-CHECK-001",
        },
    ]:
        ws.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()
    invalidate_workbook_cache(workbook_path)

    findings = photo_index_path_findings(fake_project)

    assert any("Tool # does not match" in finding.message for finding in findings)
    assert any("Related Audit ID does not match" in finding.message for finding in findings)
    assert any("missing Photo Filename" in finding.message for finding in findings)
    assert any("missing Folder Path" in finding.message for finding in findings)


def test_eoat_photo_intake_uses_eoat_folder_filename_and_metadata(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-EOAT-PHOTO-001",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "26",
            EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-0007",
            "Tool #": "5116830010",
            "Part Name/Description": "Shared EOAT Part",
            "EOAT Type": "Vacuum",
        },
    )
    photo = resolve_project_paths(fake_project).incoming_photos / "IMG_EOAT.jpg"
    photo.parent.mkdir(parents=True, exist_ok=True)
    photo.write_bytes(b"fake jpg")

    result = intake_photos(
        fake_project,
        [photo],
        "Plant 4",
        "26",
        "2026-06-08",
        "Mounting Hardware",
        eoat_assembly_id="P4-EOAT-0007",
        related_audit_id="AUD-EOAT-PHOTO-001",
        copy_mode=True,
        log_activity=False,
    )
    rows = row_dicts(resolve_project_paths(fake_project).master_workbook, "Photo Index")
    photo_row = next(row for row in rows if row["Related Audit ID"] == "AUD-EOAT-PHOTO-001")
    target = eoat_photo_root(fake_project, "P4-EOAT-0007") / "08_Mounting_Hardware" / photo_row["Photo Filename"]
    metadata_path = eoat_photo_root(fake_project, "P4-EOAT-0007") / "eoat_info.json"

    assert result.success is True
    assert target.exists()
    assert photo_row[EOAT_ASSEMBLY_ID_FIELD] == "P4-EOAT-0007"
    assert photo_row["Tool #"] == "5116830010"
    assert photo_row["Photo Filename"].startswith("P4-EOAT-0007_2026-06-08_Mounting_Hardware_")
    assert metadata_path.exists()


def test_eoat_photo_intake_marks_all_audits_sharing_eoat_as_photographed(fake_project):
    for audit_id, tool in [
        ("AUD-EOAT-SHARED-001", "5116830010"),
        ("AUD-EOAT-SHARED-002", "5116830020"),
    ]:
        _append_inventory_row(
            fake_project,
            {
                "Audit ID": audit_id,
                "Plant/Area": "Plant 4",
                "Press/Machine #": "26",
                EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-0030",
                "Tool #": tool,
                "EOAT Type": "Vacuum",
                "Photos Taken?": "No",
                "Photo Folder/Link": "",
            },
        )
    photo = resolve_project_paths(fake_project).incoming_photos / "IMG_SHARED_EOAT.jpg"
    photo.parent.mkdir(parents=True, exist_ok=True)
    photo.write_bytes(b"fake jpg")

    result = intake_photos(
        fake_project,
        [photo],
        "Plant 4",
        "26",
        "2026-06-08",
        "Front View",
        eoat_assembly_id="P4-EOAT-0030",
        related_audit_id="AUD-EOAT-SHARED-001",
        copy_mode=True,
        log_activity=False,
    )
    rows = row_dicts(resolve_project_paths(fake_project).master_workbook, "EOAT Inventory")
    related = {row["Audit ID"]: row for row in rows if str(row.get("Audit ID", "")).startswith("AUD-EOAT-SHARED-")}

    assert result.success is True
    assert result.metrics["audit_rows_updated"] == 2
    assert related["AUD-EOAT-SHARED-001"]["Photos Taken?"] == "Yes"
    assert related["AUD-EOAT-SHARED-002"]["Photos Taken?"] == "Yes"
    assert "P4-EOAT-0030" in str(related["AUD-EOAT-SHARED-001"]["Photo Folder/Link"])
    assert "P4-EOAT-0030" in str(related["AUD-EOAT-SHARED-002"]["Photo Folder/Link"])


def test_repair_photo_eoat_links_repairs_clear_match_and_skips_ambiguous(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-EOAT-REPAIR-001",
            "Press/Machine #": "26",
            EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-0011",
            "Tool #": "TOOL-CLEAR",
        },
    )
    _append_inventory_row(fake_project, {EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-0012", "Tool #": "TOOL-AMB"})
    _append_inventory_row(fake_project, {EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-0013", "Tool #": "TOOL-AMB"})
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["Photo Index"]
    headers = [cell.value for cell in ws[1]]
    for values in [
        {"Photo ID": "PHO-EOAT-REPAIR-001", "Tool #": "TOOL-CLEAR", "Related Audit ID": "AUD-EOAT-REPAIR-001"},
        {"Photo ID": "PHO-EOAT-REPAIR-002", "Tool #": "TOOL-AMB"},
    ]:
        ws.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()
    invalidate_workbook_cache(workbook_path)

    result = repair_photo_eoat_links(fake_project, log_activity=False)
    rows = row_dicts(workbook_path, "Photo Index")
    repaired = next(row for row in rows if row["Photo ID"] == "PHO-EOAT-REPAIR-001")
    ambiguous = next(row for row in rows if row["Photo ID"] == "PHO-EOAT-REPAIR-002")

    assert result.success is True
    assert result.metrics["rows_repaired"] == 1
    assert result.metrics["skipped_ambiguous"] == 1
    assert repaired[EOAT_ASSEMBLY_ID_FIELD] == "P4-EOAT-0011"
    assert ambiguous.get(EOAT_ASSEMBLY_ID_FIELD) in {"", None}


def test_convert_legacy_photo_tree_copies_to_eoat_folder_without_moving_old_file(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-EOAT-CONVERT-001",
            "Press/Machine #": "31",
            EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-0020",
            "Tool #": "TOOL-CONVERT",
            "Part Name/Description": "Convert Part",
        },
    )
    legacy_folder = destination_folder(fake_project, "Front View", "TOOL-CONVERT", "Convert Part")
    legacy_folder.mkdir(parents=True)
    legacy_photo = legacy_folder / "Tool_TOOL-CONVERT__Front_View__2026-06-08__001.jpg"
    legacy_photo.write_bytes(b"legacy")
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["Photo Index"]
    headers = [cell.value for cell in ws[1]]
    values = {
        "Photo ID": "PHO-EOAT-CONVERT-001",
        "Tool #": "TOOL-CONVERT",
        "Related Audit ID": "AUD-EOAT-CONVERT-001",
        "EOAT Area Shown": "Front View",
        "Date Taken": "2026-06-08",
        "Stored Relative Path": str(legacy_photo.relative_to(fake_project)),
        "Folder Path": str(legacy_folder.relative_to(fake_project)),
        "Photo Filename": legacy_photo.name,
        "Stored Filename": legacy_photo.name,
    }
    ws.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()
    invalidate_workbook_cache(workbook_path)

    result = convert_legacy_photo_tree_to_eoat_folders(fake_project, log_activity=False)
    rows = row_dicts(workbook_path, "Photo Index")
    photo_row = next(row for row in rows if row["Photo ID"] == "PHO-EOAT-CONVERT-001")
    new_photo = fake_project / photo_row["Stored Relative Path"]

    assert result.success is True
    assert result.metrics["converted_rows"] == 1
    assert legacy_photo.exists()
    assert new_photo.exists()
    assert "Cell_Photos/P4-EOAT-0020" in photo_row["Stored Relative Path"].replace("\\", "/")
    assert photo_row[EOAT_ASSEMBLY_ID_FIELD] == "P4-EOAT-0020"
