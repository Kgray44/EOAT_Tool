from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from core.kpi_analysis import analyze_kpis, generate_kpi_dashboard_report
from core.paths import resolve_project_paths


def test_kpi_analysis_empty_baseline(fake_project):
    summary, error = analyze_kpis(fake_project)
    assert error is None
    assert summary.metrics["kpi_rows"] == 0
    assert summary.metrics["total_downtime_minutes"] == 0


def test_kpi_analysis_totals(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    wb = load_workbook(workbook_path)
    ws = wb["KPI Baseline"]
    ws.append(["KPI-1", "2026-05-18", "Plant 4", "Press 12", "", "Family", "Vacuum", 30, "Yes", 3, 2, 12, "Drop", 15.5, 2, "", "Manual", ""])
    wb.save(workbook_path)
    wb.close()

    summary, error = analyze_kpis(fake_project)
    assert error is None
    assert summary.metrics["total_downtime_minutes"] == 30
    assert summary.metrics["part_drops"] == 3
    assert summary.scrap_reasons["Drop"] == 1

    result = generate_kpi_dashboard_report(fake_project)
    assert result.success is True
    assert result.output_reports


def test_kpi_truth_labels_distinguish_measured_from_estimated(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    wb = load_workbook(workbook_path)
    ws = wb["KPI Baseline"]
    ws.append(["KPI-MEASURED", "2026-05-18", "Plant 4", "Press 21", "", "Family", "Vacuum", 42, "Yes", 2, 1, 5, "Drop", 15.5, 1, "", "MES export", "Actual measured downtime export."])
    ws.append(["KPI-EST", "2026-05-19", "Plant 4", "Press 21", "", "Family", "Vacuum", 20, "Yes", 1, 0, 3, "Drop", 15.8, 1, "", "Estimated by supervisor", "Estimated downtime from shift discussion."])
    wb.save(workbook_path)
    wb.close()

    summary, error = analyze_kpis(fake_project)

    assert error is None
    label = summary.card_truth("Downtime Minutes")
    assert label is not None
    assert label.source_breakdown["actual measured data"] == 1
    assert label.source_breakdown["estimated/subjective data"] == 1
    assert "actual measured data" in label.source_type
    assert "estimated/subjective data" in label.source_type
    assert label.date_range == "2026-05-18 to 2026-05-19"


def test_kpi_missing_data_warnings_generated(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    wb = load_workbook(workbook_path)
    ws = wb["KPI Baseline"]
    ws.append(["KPI-MISSING", "2026-05-18", "Plant 4", "Press 22", "", "Family", "Vacuum", 10, "Yes", "", 1, "", "", "", "", "", "Manual audit observation", "Audit-observed KPI row with blanks."])
    wb.save(workbook_path)
    wb.close()

    summary, error = analyze_kpis(fake_project)

    assert error is None
    drops = summary.card_truth("Part Drops")
    assert drops is not None
    assert drops.source_type == "missing data"
    assert "missing Part Drops" in drops.missing_data_warning
    assert any("missing Part Drops" in warning for warning in summary.missing_data_warnings)
    assert summary.metrics["missing_kpi_fields_total"] > 0


def test_kpi_report_includes_confidence_section(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    wb = load_workbook(workbook_path)
    ws = wb["KPI Baseline"]
    ws.append(["KPI-REPORT", "2026-05-18", "Plant 4", "Press 23", "", "Family", "Vacuum", 15, "Yes", 1, 0, 2, "Drop", 14.5, 1, "", "PLC measured export", ""])
    wb.save(workbook_path)
    wb.close()

    result = generate_kpi_dashboard_report(fake_project, log_activity=False)

    assert result.success is True
    report_text = Path(result.output_reports[0]).read_text(encoding="utf-8")
    assert "## KPI Truth And Confidence" in report_text
    assert "actual measured data" in report_text
    assert "Confidence" in report_text

