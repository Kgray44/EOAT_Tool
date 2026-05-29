from __future__ import annotations

from openpyxl import Workbook, load_workbook

from core.audit_compatibility import build_compatibility_candidates
from core.audit_constants import (
    COMPATIBILITY_SOURCE_FIELD,
    ENTRY_TYPE_COMPATIBLE,
    ENTRY_TYPE_FIELD,
    SOURCE_AUDIT_ID_FIELD,
)
from core.paths import get_press_capacity_file, resolve_project_paths
from core.press_view import build_press_view_groups, export_press_summary
from core.workbook_schema import get_expected_headers


def _append_inventory_row(workbook_path, values):
    workbook = load_workbook(workbook_path)
    try:
        ws = workbook["EOAT Inventory"]
        headers = get_expected_headers("EOAT Inventory")
        ws.append([values.get(header, "") for header in headers])
        workbook.save(workbook_path)
    finally:
        workbook.close()


def _write_press_capacity(project_root, machine_numbers, part_number="5620040010", description="Test tool"):
    path = get_press_capacity_file(project_root)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Capacity"
    ws.append(["Machine No.", "NGW Part Number", "NGW Part Description"])
    ws.append([", ".join(machine_numbers), part_number, description])
    workbook.save(path)
    workbook.close()
    return path


def test_press_view_groups_physical_and_compatible_entries(usability_fake_project):
    paths = resolve_project_paths(usability_fake_project)
    _append_inventory_row(
        paths.master_workbook,
        {
            "Audit ID": "AUD-COMPAT-101",
            "Audit Date": "2026-05-19",
            "Press/Machine #": "Press 101",
            "Tool #": "TOOL-A",
            "EOAT Type": "Vacuum",
            "Status": "Compatible",
            ENTRY_TYPE_FIELD: ENTRY_TYPE_COMPATIBLE,
            SOURCE_AUDIT_ID_FIELD: "AUD-20260518-001",
        },
    )

    groups = build_press_view_groups(usability_fake_project)
    group = next(item for item in groups if item.machine == "101")

    assert len(group.physical_audits) == 1
    assert len(group.compatible_entries) == 1
    assert group.photo_count == 1
    assert group.pilot_candidacy == "Yes"
    assert group.average_compliance_score >= 0
    assert isinstance(group.worst_compliance_category, str)


def test_press_view_counts_compatible_links_from_machine_source_audits(fake_project):
    source_audit_id = "AUD-20260519-001"
    linked_machines = ["2", "8", "9", "19", "32", "33"]
    _write_press_capacity(fake_project, ["1", *linked_machines])
    paths = resolve_project_paths(fake_project)
    _append_inventory_row(
        paths.master_workbook,
        {
            "Audit ID": source_audit_id,
            "Audit Date": "2026-05-19",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Machine 1",
            "Tool #": "5620040010",
            "Part Name/Description": "Test tool",
            "EOAT Type": "Vacuum",
            "Status": "Complete",
            ENTRY_TYPE_FIELD: "Audited",
        },
    )
    _append_inventory_row(
        paths.master_workbook,
        {
            "Audit ID": "AUD-OTHER-SOURCE",
            "Audit Date": "2026-05-19",
            "Press/Machine #": "Machine 99",
            "Tool #": "OTHER-PN",
            "EOAT Type": "Vacuum",
            "Status": "Complete",
            ENTRY_TYPE_FIELD: "Audited",
        },
    )
    _append_inventory_row(
        paths.master_workbook,
        {
            "Audit ID": "AUD-COMPAT-ASSIGNED-HERE",
            "Press/Machine #": "Machine 1",
            "Tool #": "OTHER-PN",
            "EOAT Type": "Vacuum",
            "Status": "Compatible",
            ENTRY_TYPE_FIELD: ENTRY_TYPE_COMPATIBLE,
            SOURCE_AUDIT_ID_FIELD: "AUD-OTHER-SOURCE",
            COMPATIBILITY_SOURCE_FIELD: "Synthetic test link",
        },
    )
    for machine in linked_machines:
        _append_inventory_row(
            paths.master_workbook,
            {
                "Audit ID": f"AUD-COMPAT-{machine}",
                "Press/Machine #": f"Machine {machine}",
                "Tool #": "5620040010",
                "Part Name/Description": "Test tool",
                "EOAT Type": "Vacuum",
                "Status": "Compatible",
                ENTRY_TYPE_FIELD: ENTRY_TYPE_COMPATIBLE,
                SOURCE_AUDIT_ID_FIELD: source_audit_id,
                COMPATIBILITY_SOURCE_FIELD: "Press Capacity List",
            },
        )

    groups = build_press_view_groups(fake_project)
    group = next(item for item in groups if item.machine == "1")

    assert len(group.physical_audits) == 1
    assert len(group.compatible_entries) == 1
    assert len(group.linked_compatible_entries) == 6
    assert group.compatibility_family_machine_count == 7
    assert {entry.machine for entry in group.linked_compatible_entries} == set(linked_machines)

    compatibility_tab_result = build_compatibility_candidates(fake_project, source_audit_id)
    actions = {candidate.machine_no: candidate.recommended_action for candidate in compatibility_tab_result.candidates}
    linked_in_compatibility_tab = {
        machine
        for machine, action in actions.items()
        if action == "Already Compatible - Linked to this source"
    }
    assert actions["1"] == "Already Audited"
    assert linked_in_compatibility_tab == set(linked_machines)


def test_press_view_export_is_no_overwrite_project_output(usability_fake_project):
    result = export_press_summary(usability_fake_project, "101")

    assert result.success is True
    assert result.output_reports
    assert "Press_101_Summary" in result.output_reports[0]
