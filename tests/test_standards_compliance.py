from __future__ import annotations

from openpyxl import load_workbook

from core.audit_constants import (
    AUDIT_CONTEXT_BENCH,
    AUDIT_CONTEXT_COMPATIBILITY,
    AUDIT_CONTEXT_FIELD,
    ENTRY_TYPE_COMPATIBLE,
    ENTRY_TYPE_FIELD,
)
from core.paths import resolve_project_paths
from core.standards_compliance import analyze_standards_compliance, score_audit_compliance
from core.workbook_schema import get_expected_headers


def _row(audit_id: str, eoat_type: str, **overrides) -> dict[str, str]:
    row = {header: "" for header in get_expected_headers("EOAT Inventory")}
    row.update(
        {
            "Audit ID": audit_id,
            "Audit Date": "2026-05-18",
            "Auditor": "Synthetic Auditor",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": eoat_type,
            "Number of Parts Picked": "2",
            "# of Cups": "4",
            "Cup Type/Material": "Silicone",
            "Cup Diameter/Size": "20 mm",
            "Vacuum Generator Type": "Venturi",
            "# of Grippers": "2",
            "Gripper Type": "Double Pressure",
            "Gripper Model": "Large Double Gripper",
            "Sensors Present?": "No",
            "Quick Disconnects Present?": "No",
            "Tubing Condition": "OK",
            "Cable Management Condition": "OK",
            "Mounting Hardware Condition": "OK",
            "EOAT Alignment Condition": "OK",
            "Fastener/Locking Hardware Present?": "Yes",
            "Known Issues": "No issue observed.",
            "Maintenance Frequency": "Monthly",
            "Spare Parts Identified?": "Yes",
            "Drawing/CAD Available?": "Yes",
            "BOM Available?": "Yes",
            "Process Binder Complete?": "Yes",
            "Photos Taken?": "Yes",
            "Status": "In Progress",
            "Priority": "Medium",
            "Pilot Candidate?": "No",
        }
    )
    row.update(overrides)
    return row


def _append_inventory(project_root, row: dict[str, str]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    ws.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def _category(result, key: str):
    return next(item for item in result.category_results if item.key == key)


def test_standards_compliance_scores_vacuum_mechanical_and_hybrid(fake_project):
    vacuum = score_audit_compliance(fake_project, _row("AUD-VAC-COMP", "Vacuum"))
    mechanical = score_audit_compliance(fake_project, _row("AUD-MECH-COMP", "Mechanical / Gripper"))
    hybrid = score_audit_compliance(fake_project, _row("AUD-HYB-COMP", "Hybrid"))

    assert _category(vacuum, "tooling_details_complete").status == "compliant"
    assert _category(mechanical, "tooling_details_complete").status == "compliant"
    assert _category(hybrid, "tooling_details_complete").status == "compliant"
    assert vacuum.overall_score > 0
    assert mechanical.overall_score > 0
    assert hybrid.overall_score > 0


def test_standards_compliance_treats_unknown_and_na_differently(fake_project):
    unknown = score_audit_compliance(
        fake_project, _row("AUD-UNK-COMP", "Unknown / Needs Review", **{"Sensors Present?": "Unknown / Not Checked"})
    )
    no_sensors = score_audit_compliance(fake_project, _row("AUD-NA-COMP", "Vacuum", **{"Sensors Present?": "No"}))

    assert _category(unknown, "eoat_classification_complete").status == "unknown"
    assert _category(no_sensors, "sensor_standards").status == "not applicable"
    assert _category(no_sensors, "sensor_standards").score is None


def test_standards_compliance_splits_bench_followups_from_true_failures(fake_project):
    result = score_audit_compliance(
        fake_project,
        _row(
            "AUD-BENCH-COMP",
            "Vacuum",
            **{
                AUDIT_CONTEXT_FIELD: AUDIT_CONTEXT_BENCH,
                "Press/Machine #": "",
                "Robot Type": "",
                "Robot Model/Controller": "",
                "EOAT Alignment Condition": "",
            },
        ),
    )

    installed_cell = _category(result, "installed_cell_validation_context")
    assert result.audit_context == AUDIT_CONTEXT_BENCH
    assert result.installed_cell_validation_score == "Not Installed / Pending"
    assert installed_cell.status == "not observable"
    assert result.true_fail_count == 0
    assert result.not_observable_count >= 1
    assert "audited off-machine" in result.notes_recommended_action


def test_standards_compliance_marks_compatibility_rows_pending_not_physical(fake_project):
    result = score_audit_compliance(
        fake_project,
        _row(
            "AUD-COMPAT-ROW",
            "Vacuum",
            **{
                AUDIT_CONTEXT_FIELD: AUDIT_CONTEXT_COMPATIBILITY,
                ENTRY_TYPE_FIELD: ENTRY_TYPE_COMPATIBLE,
                "Physical Audit Verified": "No",
                "Compatibility Confidence": "Press Capacity",
            },
        ),
    )

    installed_cell = _category(result, "installed_cell_validation_context")
    compatibility = _category(result, "compatibility_context")
    assert result.audit_context == AUDIT_CONTEXT_COMPATIBILITY
    assert result.installed_cell_validation_score == "Pending / Not physically verified"
    assert installed_cell.status == "follow-up required"
    assert compatibility.status == "warning"
    assert result.follow_up_count >= 1


def test_standards_compliance_rolls_up_by_press(fake_project):
    _append_inventory(
        fake_project, _row("AUD-ROLL-001", "Vacuum", **{"Press/Machine #": "Press 77", "Tubing Condition": "Damaged"})
    )
    _append_inventory(fake_project, _row("AUD-ROLL-002", "Vacuum", **{"Press/Machine #": "Press 77"}))

    summary, error = analyze_standards_compliance(fake_project)

    assert error is None
    assert summary is not None
    rollup = next(item for item in summary.press_rollups if item.machine == "Press 77")
    assert rollup.audit_count == 2
    assert rollup.open_standards_issues > 0
    assert rollup.worst_category
