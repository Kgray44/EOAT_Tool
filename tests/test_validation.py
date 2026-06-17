from __future__ import annotations

from openpyxl import Workbook, load_workbook

from core.audit_by_press import AUDIT_BY_PRESS_SHEET
from core.audit_constants import (
    AUDIT_CONTEXT_FIELD,
    AUDIT_CONTEXT_INSTALLED,
    COMPATIBILITY_SOURCE_FIELD,
    SOURCE_AUDIT_ID_FIELD,
)
from core.audit_entries import load_audit_entry
from core.audit_field_rules import ELECTRICAL_WIRING_PRESENT_FIELD
from core.audit_progress import calculate_audit_progress_from_rows
from core.constants import EXPECTED_NUMBERED_FOLDERS
from core.paths import resolve_project_paths
from core.robot_info import ROBOT_INFO_SHEET, robot_info_workbook_path, upsert_robot_info_from_audit
from core.validation import validate_project_foundation
from core.workbook_schema import get_expected_headers, get_expected_sheets


def test_validate_project_foundation_on_temp_project(tmp_path):
    for folder in EXPECTED_NUMBERED_FOLDERS:
        (tmp_path / folder).mkdir(parents=True)
    (tmp_path / "00_Project_Admin" / "Daily_Status_Reports").mkdir()
    (tmp_path / "00_Project_Admin" / "Weekly_Status_Reports").mkdir()
    (tmp_path / "00_Project_Admin" / "Activity_Logs").mkdir()
    (tmp_path / "README.md").write_text("Project", encoding="utf-8")
    (tmp_path / "00_Project_Admin" / "project_schedule_week1.json").write_text(
        '{"week": 1, "days": {}}', encoding="utf-8"
    )
    (tmp_path / "00_Project_Admin" / "task_progress_week1.json").write_text('{"tasks": []}', encoding="utf-8")
    workbook_dir = tmp_path / "01_EOAT_Audit" / "EOAT_Audit_Database"
    workbook_dir.mkdir(parents=True)

    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    for sheet_name in get_expected_sheets():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(get_expected_headers(sheet_name))
    workbook.save(workbook_dir / "EOAT_Master_Tracker.xlsx")

    result = validate_project_foundation(tmp_path)

    assert result.success is True
    assert not result.errors
    assert result.metrics["missing_key_inventory_header_count"] == 0
    assert any("Audit by Press view missing or stale" in warning for warning in result.warnings)
    assert any("Missing Robot Info workbook" in warning for warning in result.warnings)


def test_legacy_gripper_size_header_is_ignored_for_backward_compatibility(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    insert_at = headers.index("Gripper Model") + 2
    ws.insert_cols(insert_at)
    ws.cell(row=1, column=insert_at).value = "Gripper Size"
    headers = [cell.value for cell in ws[1]]
    row = {header: "" for header in headers}
    row.update(
        {
            "Audit ID": "AUD-LEGACY-GRIPPER-SIZE",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "DEMO-PN-1200",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Mechanical / Gripper",
            "# of Grippers": "2",
            "Gripper Type": "Single Pressure",
            "Gripper Model": "Zimmer GPP",
            "Gripper Size": "25 mm",
            "Status": "Complete",
        }
    )
    ws.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()

    loaded = load_audit_entry(fake_project, "AUD-LEGACY-GRIPPER-SIZE")
    result = validate_project_foundation(fake_project)

    assert loaded is not None
    assert "Gripper Size" not in loaded
    assert not any("Unexpected EOAT Inventory header: Gripper Size" in warning for warning in result.warnings)
    assert "Legacy Gripper Size column is present and ignored" in "\n".join(result.details)


def test_validate_project_foundation_missing_project(tmp_path):
    result = validate_project_foundation(tmp_path / "missing")

    assert result.success is False
    assert result.errors


def test_workbook_health_detects_robot_info_duplicate_and_invalid_rows(fake_project):
    result = upsert_robot_info_from_audit(
        fake_project,
        {
            "Audit ID": "AUD-ROBOT-HEALTH-001",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Robot Type": "Wittmann R9",
            "Robot Vacuum Circuits": "3",
            "Robot Pressure Circuits": "2",
            "Robot Interchangeable Circuits": "0",
        },
    )
    assert result.success, result.errors
    path = robot_info_workbook_path(fake_project)
    workbook = load_workbook(path)
    ws = workbook[ROBOT_INFO_SHEET]
    headers = [cell.value for cell in ws[1]]
    duplicate = [ws.cell(row=2, column=index + 1).value for index in range(len(headers))]
    duplicate[headers.index("Robot Vacuum Circuits")] = -1
    ws.append(duplicate)
    workbook.save(path)
    workbook.close()

    health = validate_project_foundation(fake_project)

    assert health.metrics["robot_info_duplicate_row_count"] == 1
    assert health.metrics["robot_info_invalid_circuit_count"] == 1
    warning_text = "\n".join(health.warnings)
    assert "Duplicate Robot Info row" in warning_text
    assert "Invalid Robot Info circuit value" in warning_text


def _append_inventory_row(workbook_path, values):
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    ws.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def _create_schema_workbook(project_root):
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(get_expected_headers(sheet_name))
    workbook.save(workbook_path)
    workbook.close()
    return workbook_path


def _create_inventory_workbook_without_electrical_wiring_header(project_root):
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        sheet = workbook.create_sheet(sheet_name)
        headers = [
            header
            for header in get_expected_headers(sheet_name)
            if sheet_name != "EOAT Inventory" or header != ELECTRICAL_WIRING_PRESENT_FIELD
        ]
        sheet.append(headers)
    workbook.save(workbook_path)
    workbook.close()
    return workbook_path


def _base_inventory_values(audit_id: str, eoat_moves: str) -> dict[str, str]:
    return {
        "Audit ID": audit_id,
        "Audit Date": "2026-05-18",
        "Auditor": "KG",
        AUDIT_CONTEXT_FIELD: AUDIT_CONTEXT_INSTALLED,
        "Plant/Area": "Plant 4",
        "Press/Machine #": "Press 12",
        "Tool #": "DEMO-PN-1200",
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "EOAT Moves": eoat_moves,
        "Connection Type": "ATI",
        "# of Cylinders": "N/A",
        "Cylinder Type": "Linear",
        "Cleanroom/Non-Cleanroom": "Whiteroom",
        "Status": "In Progress",
        "Priority": "Medium",
        "Known Issues": "No issue observed.",
    }


def _complete_inventory_values(audit_id: str, entry_type: str = "Audited") -> dict[str, str]:
    values = {header: "Filled" for header in get_expected_headers("EOAT Inventory")}
    values.update(
        {
            "Audit ID": audit_id,
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            AUDIT_CONTEXT_FIELD: AUDIT_CONTEXT_INSTALLED,
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "DEMO-PN-1200",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Hybrid",
            "EOAT Moves": "Part",
            "Connection Type": "ATI",
            "Cleanroom/Non-Cleanroom": "Whiteroom",
            "Number of Parts Picked": "2",
            "# of Cylinders": "1",
            "Cylinder Type": "Linear",
            "# of Cups": "4",
            "# of Grippers": "2",
            "Gripper Type": "Double Pressure",
            "Gripper Model": "MHZL2-16D",
            "Sensors Present?": "Yes",
            ELECTRICAL_WIRING_PRESENT_FIELD: "Yes",
            "Quick Disconnects Present?": "Yes",
            "Status": "In Progress",
            "Priority": "Medium",
            "Known Issues": "No issue observed.",
            "Photos Taken?": "No",
            "Manual Completion Override": "No",
            "Manual Completion Override Timestamp": "N/A",
            "Manual Completion Override User": "N/A",
            "Ignored Empty Fields At Override": "N/A",
            "Entry Type": entry_type,
            SOURCE_AUDIT_ID_FIELD: "",
            COMPATIBILITY_SOURCE_FIELD: "",
        }
    )
    if entry_type == "Compatible":
        values["Audit Date"] = "N/A"
        values["Auditor"] = "N/A"
    return values


def _count_blank_inventory_cells(workbook_path, fields: set[str]) -> int:
    workbook = load_workbook(workbook_path, read_only=True)
    try:
        ws = workbook["EOAT Inventory"]
        headers = [cell.value for cell in ws[1]]
        positions = {header: index for index, header in enumerate(headers) if header in fields}
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            count += sum(1 for index in positions.values() if index < len(row) and not str(row[index] or "").strip())
        return count
    finally:
        workbook.close()


def test_workbook_health_flags_missing_major_inventory_header(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        sheet = workbook.create_sheet(sheet_name)
        headers = [header for header in get_expected_headers(sheet_name) if header != "Tool #"]
        sheet.append(headers)
    workbook.save(workbook_path)
    workbook.close()

    result = validate_project_foundation(fake_project)

    assert result.success is False
    assert any("Missing major EOAT Inventory header: Tool #" in error for error in result.errors)


def test_workbook_health_tolerates_workbook_missing_eoat_moves_header(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        sheet = workbook.create_sheet(sheet_name)
        headers = [header for header in get_expected_headers(sheet_name) if header != "EOAT Moves"]
        sheet.append(headers)
    workbook.save(workbook_path)
    workbook.close()

    result = validate_project_foundation(fake_project)

    assert result.success is True
    assert not any("Missing major EOAT Inventory header: EOAT Moves" in error for error in result.errors)


def test_compatibility_metadata_headers_remain_schema_checked(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = Workbook()
    workbook.remove(workbook.active)
    metadata_fields = {SOURCE_AUDIT_ID_FIELD, COMPATIBILITY_SOURCE_FIELD}
    assert metadata_fields.issubset(set(get_expected_headers("EOAT Inventory")))
    for sheet_name in get_expected_sheets():
        sheet = workbook.create_sheet(sheet_name)
        headers = [header for header in get_expected_headers(sheet_name) if header not in metadata_fields]
        sheet.append(headers)
    workbook.save(workbook_path)
    workbook.close()

    result = validate_project_foundation(fake_project)

    assert result.metrics["missing_full_inventory_header_count"] == 2
    assert any(
        f"Missing detail EOAT Inventory header: {SOURCE_AUDIT_ID_FIELD}" in warning for warning in result.warnings
    )
    assert any(
        f"Missing detail EOAT Inventory header: {COMPATIBILITY_SOURCE_FIELD}" in warning for warning in result.warnings
    )


def test_workbook_health_ignores_blank_autofilled_metadata_cells(tmp_path):
    workbook_path = _create_schema_workbook(tmp_path)
    _append_inventory_row(workbook_path, _complete_inventory_values("AUD-META-AUDITED"))
    _append_inventory_row(workbook_path, _complete_inventory_values("AUD-META-COMPATIBLE", "Compatible"))

    result = validate_project_foundation(tmp_path)

    raw_metadata_blanks = _count_blank_inventory_cells(
        workbook_path,
        {SOURCE_AUDIT_ID_FIELD, COMPATIBILITY_SOURCE_FIELD},
    )
    warning_text = "\n".join(result.warnings)
    assert raw_metadata_blanks == 4
    assert result.metrics["blank_saved_audit_cell_count"] == 0
    assert result.metrics["major_na_cell_count"] == 0
    assert result.metrics["missing_applicable_major_cell_count"] == 0
    assert "saved EOAT Inventory cell(s) are blank" not in warning_text
    assert SOURCE_AUDIT_ID_FIELD not in warning_text
    assert COMPATIBILITY_SOURCE_FIELD not in warning_text
    assert any(
        "metadata cells are intentionally ignored during blank-cell validation"
        in detail
        for detail in result.details
    )


def test_workbook_health_blank_count_excludes_autofilled_metadata_columns(tmp_path):
    workbook_path = _create_schema_workbook(tmp_path)
    values = _complete_inventory_values("AUD-META-BLANK-COUNT")
    values["Notes"] = ""
    _append_inventory_row(workbook_path, values)

    result = validate_project_foundation(tmp_path)

    raw_blank_count = _count_blank_inventory_cells(
        workbook_path,
        {SOURCE_AUDIT_ID_FIELD, COMPATIBILITY_SOURCE_FIELD, "Notes"},
    )
    warning_text = "\n".join(result.warnings)
    assert raw_blank_count == 3
    assert result.metrics["blank_saved_audit_cell_count"] == 1
    assert any("1 saved EOAT Inventory cell(s) are blank" in warning for warning in result.warnings)
    assert SOURCE_AUDIT_ID_FIELD not in warning_text
    assert COMPATIBILITY_SOURCE_FIELD not in warning_text


def test_workbook_health_accepts_allowed_eoat_moves_values(tmp_path):
    workbook_path = _create_schema_workbook(tmp_path)
    for value in ["Part", "Sprue", "Both"]:
        _append_inventory_row(workbook_path, _base_inventory_values(f"AUD-MOVES-{value.upper()}", value))

    result = validate_project_foundation(tmp_path)

    assert result.metrics["invalid_dropdown_value_count"] == 0


def test_workbook_health_flags_invalid_eoat_moves_values(tmp_path):
    workbook_path = _create_schema_workbook(tmp_path)
    for value in ["Runner", "Parts", "Sprue Only"]:
        _append_inventory_row(workbook_path, _base_inventory_values(f"AUD-MOVES-BAD-{value.replace(' ', '-')}", value))

    result = validate_project_foundation(tmp_path)

    assert result.metrics["invalid_dropdown_value_count"] == 3
    warning_text = "\n".join(result.warnings)
    assert "EOAT Moves=Runner" in warning_text
    assert "EOAT Moves=Parts" in warning_text
    assert "EOAT Moves=Sprue Only" in warning_text


def test_workbook_health_recognizes_gripper_headers_and_validates_values(tmp_path):
    workbook_path = _create_schema_workbook(tmp_path)
    headers = get_expected_headers("EOAT Inventory")
    assert "# of Grippers" in headers
    assert "Gripper Type" in headers
    assert "Gripper Model" in headers
    assert "Vacuum Zones" not in headers

    valid = _base_inventory_values("AUD-GRIPPER-VALID", "Part")
    valid.update(
        {
            "EOAT Type": "Mechanical / Gripper",
            "# of Cups": "N/A",
            "# of Grippers": "2",
            "Gripper Type": "Single Pressure",
            "Gripper Model": "MHZL2-16D",
            "Cup Type/Material": "N/A",
            "Cup Diameter/Size": "N/A",
        }
    )
    invalid = {**valid, "Audit ID": "AUD-GRIPPER-INVALID", "# of Grippers": "two", "Gripper Type": "Parallel jaw"}
    vacuum = _base_inventory_values("AUD-GRIPPER-VACUUM-NA", "Part")
    vacuum.update({"# of Cups": "4", "# of Grippers": "N/A", "Gripper Type": "N/A", "Gripper Model": "N/A"})
    _append_inventory_row(workbook_path, valid)
    _append_inventory_row(workbook_path, invalid)
    _append_inventory_row(workbook_path, vacuum)

    result = validate_project_foundation(tmp_path)

    assert result.metrics["invalid_dropdown_value_count"] == 1
    assert result.metrics["invalid_numeric_value_count"] == 1
    warning_text = "\n".join(result.warnings)
    assert "Gripper Type=Parallel jaw" in warning_text
    assert "# of Grippers=two" in warning_text
    assert "MHZL2-16D" not in warning_text
    assert "AUD-GRIPPER-VACUUM-NA" not in warning_text


def test_workbook_health_defaults_blank_sensor_fields_when_part_present_yes(tmp_path):
    workbook_path = _create_schema_workbook(tmp_path)
    values = _complete_inventory_values("AUD-PART-PRESENT-HEALTH")
    values.update(
        {
            "Sensor Type": "",
            "Sensor Brand/Model": "",
            "Part-Present Detection Present?": "Yes",
        }
    )
    _append_inventory_row(workbook_path, values)

    result = validate_project_foundation(tmp_path)

    warning_text = "\n".join(result.warnings)
    assert "Sensor Type" not in warning_text
    assert "Sensor Brand/Model" not in warning_text


def test_workbook_health_reports_compatible_blank_eoat_moves_as_inherited(tmp_path):
    workbook_path = _create_schema_workbook(tmp_path)
    source = _base_inventory_values("AUD-MOVES-SOURCE-BLANK", "")
    source["Entry Type"] = "Audited"
    compatible = _base_inventory_values("AUD-MOVES-COMPAT-BLANK", "")
    compatible["Entry Type"] = "Compatible"
    compatible["Source Audit ID"] = "AUD-MOVES-SOURCE-BLANK"
    _append_inventory_row(workbook_path, source)
    _append_inventory_row(workbook_path, compatible)

    result = validate_project_foundation(tmp_path)

    assert result.metrics["missing_eoat_moves_count"] == 2
    warning_text = "\n".join(result.warnings)
    assert "Missing important audit field: EOAT Moves" in warning_text
    assert "inherited from source audit AUD-MOVES-SOURCE-BLANK" in warning_text


def test_workbook_health_flags_na_in_major_columns(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    _append_inventory_row(
        workbook_path,
        {
            "Audit ID": "AUD-NA-MAJOR",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "N/A",
            "Tool #": "N/A",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Connection Type": "N/A",
            "Cleanroom/Non-Cleanroom": "Whiteroom",
            "Status": "In Progress",
            "Priority": "Medium",
            "Known Issues": "N/A",
        },
    )

    result = validate_project_foundation(fake_project)

    assert result.metrics["major_na_cell_count"] >= 4
    assert any(
        "applicable major EOAT Inventory cell(s) are blank or contain N/A" in warning for warning in result.warnings
    )


def test_workbook_health_ignores_machine_fields_for_uninstalled_tool_audits(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    _append_inventory_row(
        workbook_path,
        {
            "Audit ID": "AUD-UNINSTALLED-VALID",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "N/A",
            "Press/Machine #": "N/A",
            "Tool #": "TOOL-UNINSTALLED",
            "Robot Type": "N/A",
            "Robot Model/Controller": "N/A",
            "Part Family": "Bench Tool",
            "Part Name/Description": "Bench EOAT sample",
            "EOAT Type": "Vacuum",
            "Connection Type": "ATI",
            "Cleanroom/Non-Cleanroom": "Whiteroom",
            "Status": "In Progress",
            "Priority": "Medium",
            "Known Issues": "Bench inspection.",
            "Notes": "EOAT Not Installed.",
        },
    )
    _append_inventory_row(
        workbook_path,
        {
            "Audit ID": "AUD-INSTALLED-MISSING-ROBOT",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "N/A",
            "Press/Machine #": "Press 44",
            "Tool #": "TOOL-INSTALLED",
            "Robot Type": "N/A",
            "EOAT Type": "Vacuum",
            "Connection Type": "ATI",
            "Cleanroom/Non-Cleanroom": "Whiteroom",
            "Status": "In Progress",
            "Priority": "Medium",
            "Known Issues": "Installed row should still require robot context.",
        },
    )

    result = validate_project_foundation(fake_project)
    findings = result.structured_data["validation_findings"]
    uninstalled_fields = {
        finding["column_name"] for finding in findings if finding.get("audit_id") == "AUD-UNINSTALLED-VALID"
    }
    installed_fields = {
        finding["column_name"] for finding in findings if finding.get("audit_id") == "AUD-INSTALLED-MISSING-ROBOT"
    }

    assert {"Plant/Area", "Press/Machine #", "Robot Type", "Robot Model/Controller"}.isdisjoint(uninstalled_fields)
    assert {"Plant/Area", "Robot Type"} <= installed_fields


def test_audit_progress_labels_uninstalled_rows_without_robot_missing_count(tmp_path):
    rows = [
        {
            "Audit ID": "AUD-UNINSTALLED-PROGRESS",
            "Press/Machine #": "",
            "Tool #": "TOOL-UNINSTALLED",
            "Robot Type": "",
            "EOAT Type": "Vacuum",
            "EOAT Moves": "Part",
            "# of Cups": "4",
            "Tubing Condition": "OK",
            "Cable Management Condition": "OK",
            "Known Issues": "Bench inspection.",
            "Photos Taken?": "No",
            "Status": "In Progress",
            "Priority": "Medium",
        },
        {
            "Audit ID": "AUD-INSTALLED-PROGRESS",
            "Press/Machine #": "Press 55",
            "Tool #": "TOOL-INSTALLED",
            "Robot Type": "",
            "EOAT Type": "Vacuum",
            "EOAT Moves": "Part",
            "# of Cups": "4",
            "Tubing Condition": "OK",
            "Cable Management Condition": "OK",
            "Known Issues": "Installed row.",
            "Photos Taken?": "No",
            "Status": "In Progress",
            "Priority": "Medium",
        },
    ]

    summary = calculate_audit_progress_from_rows(rows, tmp_path / "missing_capacity.xlsx")

    assert summary.missing_field_counts["Robot Type"] == 1
    assert summary.robot_type_counts["EOAT Not Installed"] == 1


def test_workbook_health_deduplicates_missing_major_row_field_pairs(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    headers = get_expected_headers("EOAT Inventory")
    values = {header: "Filled" for header in headers}
    values.update(
        {
            "Audit ID": "AUD-DEDUPE-MAJOR",
            "Audit Date": "N/A",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "DEMO-PN-1200",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Unknown / Needs Review",
            "EOAT Moves": "Part",
            "Connection Type": "ATI",
            "Cleanroom/Non-Cleanroom": "Whiteroom",
            "Number of Parts Picked": "2",
            "# of Grippers": "N/A",
            "Gripper Type": "N/A",
            "Gripper Model": "N/A",
            "Sensors Present?": "Yes",
            "Vacuum Confirmation Present?": "Yes",
            "Part-Present Detection Present?": "No",
            "Electrical/Wiring Present?": "Yes",
            "Quick Disconnects Present?": "Yes",
            "Status": "In Progress",
            "Priority": "Medium",
            "Photos Taken?": "No",
        }
    )
    _append_inventory_row(workbook_path, values)

    result = validate_project_foundation(fake_project)

    assert result.metrics["missing_applicable_major_cell_count"] == 1
    warning_text = "\n".join(result.warnings)
    assert warning_text.count("row 2 Audit Date") == 1


def test_workbook_health_allows_na_for_non_applicable_tooling_fields(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    headers = get_expected_headers("EOAT Inventory")
    values = {header: "N/A" for header in headers}
    values.update(
        {
            "Audit ID": "AUD-NONAPPLICABLE-NA",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            AUDIT_CONTEXT_FIELD: AUDIT_CONTEXT_INSTALLED,
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "DEMO-PN-1200",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Mechanical / Gripper",
            "Connection Type": "ATI",
            "Cleanroom/Non-Cleanroom": "Whiteroom",
            "Status": "In Progress",
            "Priority": "Medium",
            "Known Issues": "No issues observed.",
            "Part Family": "Housing",
            "Sensors Present?": "No",
            "Quick Disconnects Present?": "No",
            "Tubing Condition": "OK",
            "Cable Management Condition": "OK",
            "Photos Taken?": "No",
            "# of Cups": "N/A",
            "Cup Type/Material": "N/A",
            "Cup Diameter/Size": "N/A",
            "# of Grippers": "2",
            "Gripper Type": "Single Pressure",
            "Gripper Model": "Zimmer GPP",
        }
    )
    _append_inventory_row(workbook_path, values)

    result = validate_project_foundation(fake_project)

    assert result.metrics["major_na_cell_count"] == 0
    assert not any("# of Cups" in warning for warning in result.warnings)
    assert not any("Cup Diameter/Size" in warning for warning in result.warnings)
    assert not any("Cup Type/Material" in warning for warning in result.warnings)


def test_workbook_health_allows_na_for_no_sensor_wiring_fields(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    headers = get_expected_headers("EOAT Inventory")
    values = {header: "N/A" for header in headers}
    values.update(
        {
            "Audit ID": "AUD-NO-SENSOR-HEALTH",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            AUDIT_CONTEXT_FIELD: AUDIT_CONTEXT_INSTALLED,
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "DEMO-PN-1200",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "# of Cups": "4",
            "Connection Type": "ATI",
            "Cleanroom/Non-Cleanroom": "Whiteroom",
            "Sensors Present?": "No",
            "Electrical/Wiring Present?": "No",
            "Part Family": "Housing",
            "Tubing Condition": "OK",
            "Known Issues": "No sensor package.",
            "Photos Taken?": "No",
            "Status": "In Progress",
            "Priority": "Medium",
        }
    )
    _append_inventory_row(workbook_path, values)

    result = validate_project_foundation(fake_project)

    assert result.metrics["major_na_cell_count"] == 0
    warning_text = "\n".join(result.warnings)
    assert "Sensor Type" not in warning_text
    assert "Sensor Brand/Model" not in warning_text
    assert "Electrical Quick Disconnect Type" not in warning_text
    assert "Cable Management Condition" not in warning_text


def test_old_workbook_missing_electrical_wiring_header_gets_schema_warning_not_electrical_na_cascade(tmp_path):
    workbook_path = _create_inventory_workbook_without_electrical_wiring_header(tmp_path)
    _append_inventory_row(
        workbook_path,
        {
            "Audit ID": "AUD-OLD-SCHEMA-NO-WIRING",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            AUDIT_CONTEXT_FIELD: AUDIT_CONTEXT_INSTALLED,
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "DEMO-PN-1200",
            "Robot Type": "Wittmann R9",
            "Part Family": "Housing",
            "EOAT Type": "Vacuum",
            "# of Cups": "4",
            "Connection Type": "ATI",
            "Cleanroom/Non-Cleanroom": "Whiteroom",
            "Sensors Present?": "No",
            "Sensor Type": "N/A",
            "Sensor Brand/Model": "N/A",
            "Part-Present Detection Present?": "N/A",
            "Electrical Quick Disconnect Type": "N/A",
            "Cable Management Condition": "N/A",
            "Tubing Condition": "OK",
            "Known Issues": "No sensor or wiring package.",
            "Photos Taken?": "No",
            "Status": "In Progress",
            "Priority": "Medium",
        },
    )

    result = validate_project_foundation(tmp_path)

    warning_text = "\n".join(result.warnings)
    assert result.success
    assert warning_text.count("Workbook is missing Electrical/Wiring Present?") == 1
    major_warnings = [warning for warning in result.warnings if "applicable major EOAT Inventory cell" in warning]
    assert not any("Electrical Quick Disconnect Type" in warning for warning in major_warnings)
    assert not any("Cable Management Condition" in warning for warning in major_warnings)
    assert result.metrics["missing_applicable_major_cell_count"] == 0


def test_migrated_no_wiring_rows_do_not_count_electrical_na_as_major_missing(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    _append_inventory_row(
        workbook_path,
        {
            "Audit ID": "AUD-MIGRATED-NO-WIRING-HEALTH",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            AUDIT_CONTEXT_FIELD: AUDIT_CONTEXT_INSTALLED,
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "DEMO-PN-1200",
            "Robot Type": "Wittmann R9",
            "Part Family": "Housing",
            "EOAT Type": "Vacuum",
            "# of Cups": "4",
            "Connection Type": "ATI",
            "Cleanroom/Non-Cleanroom": "Whiteroom",
            "Sensors Present?": "No",
            "Electrical/Wiring Present?": "No",
            "Electrical Quick Disconnect Type": "N/A",
            "Cable Management Condition": "N/A",
            "Tubing Condition": "OK",
            "Known Issues": "No wiring package.",
            "Photos Taken?": "No",
            "Status": "In Progress",
            "Priority": "Medium",
        },
    )

    result = validate_project_foundation(fake_project)

    warning_text = "\n".join(result.warnings)
    assert result.metrics["major_na_cell_count"] == 0
    assert "Electrical Quick Disconnect Type" not in warning_text
    assert "Cable Management Condition" not in warning_text


def test_workbook_health_flags_na_for_applicable_sensor_wiring_fields(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    headers = get_expected_headers("EOAT Inventory")
    values = {header: "N/A" for header in headers}
    values.update(
        {
            "Audit ID": "AUD-SENSOR-HEALTH",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "DEMO-PN-1200",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Connection Type": "ATI",
            "Cleanroom/Non-Cleanroom": "Whiteroom",
            "Sensors Present?": "Yes",
            "Tubing Condition": "OK",
            "Known Issues": "Sensor details need review.",
            "Photos Taken?": "No",
            "Status": "In Progress",
            "Priority": "Medium",
        }
    )
    _append_inventory_row(workbook_path, values)

    result = validate_project_foundation(fake_project)

    assert result.metrics["major_na_cell_count"] >= 1
    warning_text = "\n".join(result.warnings)
    assert "Cable Management Condition" in warning_text


def test_workbook_health_summarizes_blank_saved_cells_without_per_cell_noise(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    _append_inventory_row(
        workbook_path,
        {
            "Audit ID": "AUD-BLANK-DETAILS",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "DEMO-PN-1200",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Connection Type": "ATI",
            "Cleanroom/Non-Cleanroom": "Whiteroom",
            "Status": "In Progress",
            "Priority": "Medium",
            "Known Issues": "No issue observed.",
        },
    )

    result = validate_project_foundation(fake_project)
    blank_warnings = [warning for warning in result.warnings if "saved EOAT Inventory cell(s) are blank" in warning]

    assert result.metrics["blank_saved_audit_cell_count"] > 0
    assert len(blank_warnings) == 1


def test_workbook_health_treats_audit_by_press_as_regenerable_warning(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(get_expected_headers(sheet_name))
    assert AUDIT_BY_PRESS_SHEET not in workbook.sheetnames
    workbook.save(workbook_path)
    workbook.close()

    result = validate_project_foundation(fake_project)

    assert result.success is True
    assert not any(AUDIT_BY_PRESS_SHEET in error for error in result.errors)
    assert any(
        warning == "Audit by Press view missing or stale; refresh generated view." for warning in result.warnings
    )
