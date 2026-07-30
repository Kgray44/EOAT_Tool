from __future__ import annotations

import hashlib
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from tools.migration.press_capacity_import import (
    inspect_press_capacity_workbook,
    plan_press_capacity_import,
    write_immutable_receipt,
)


def _workbook(path: Path, rows: list[tuple[object, object]]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "P4 Capacity"
    sheet.append(["Machine No.", "Press Tonnage", "NGW Part Number"])
    for machine, tons in rows:
        sheet.append([machine, tons, "part-evidence-only"])
    workbook.save(path)
    workbook.close()
    return path


def test_plan_sets_empty_capacity_and_preserves_existing_identical_value(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "capacity.xlsx", [("Press 27", 110), ("28, Machine 29", 220)])

    report = plan_press_capacity_import(
        source,
        {"27": None, "28": Decimal("220"), "29": Decimal("220")},
    )

    assert report.safe_to_execute
    assert report.source_machine_count == 3
    assert report.matched_machines == 3
    assert [(item.machine_number, item.action) for item in report.updates] == [
        ("27", "SET_PRESS_CAPACITY"),
        ("28", "UNCHANGED"),
        ("29", "UNCHANGED"),
    ]
    assert report.updates[1].source_rows == report.updates[2].source_rows == (3,)
    assert [(item.source_machine_number, item.verification_class, item.action) for item in report.mappings] == [
        ("27", "EXACT_CANONICAL_MATCH", "SET_PRESS_CAPACITY"),
        ("28", "EXACT_CANONICAL_MATCH", "UNCHANGED"),
        ("29", "EXACT_CANONICAL_MATCH", "UNCHANGED"),
    ]
    assert report.mappings[0].source_locations == (("P4 Capacity", 2),)


def test_plan_refuses_ambiguous_source_existing_conflict_and_unmatched_machine(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "capacity.xlsx", [("27", 110), ("27", 120), ("28", 220), ("99", 75)])

    report = plan_press_capacity_import(source, {"28": Decimal("200")})

    assert not report.safe_to_execute
    assert report.conflicting_source_values == {"27": ["110", "120"]}
    assert report.conflicting_existing_values == {"28": {"existing": "200", "source": "220"}}
    assert report.unmatched_machines == ["99"]


def test_invalid_capacity_is_a_fail_closed_review_item(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "capacity.xlsx", [("27", 0)])

    report = plan_press_capacity_import(source, {"27": None})

    assert not report.safe_to_execute
    assert report.invalid_rows == [{"sheet": "P4 Capacity", "row_number": 2, "issue": "INVALID_PRESS_TONNAGE"}]


def test_plan_reads_grouped_p4_press_headers_and_flags_missing_tonnage(tmp_path: Path) -> None:
    source = tmp_path / "grouped-capacity.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "P4 Capacity"
    sheet.append(["Machine No.", "NGW Part Number"])
    sheet.append(["Press 27 - 165T - 45mm Screw", None])
    sheet.append([27, "5116380010"])
    sheet.append(["Press 37", None])
    sheet.append([37, "5116400010"])
    workbook.save(source)
    workbook.close()

    report = plan_press_capacity_import(source, {"27": None, "37": None})

    assert report.source_machine_count == 1
    assert report.updates[0].machine_number == "27"
    assert report.updates[0].source_tonnage == Decimal("165")
    assert report.invalid_rows == [{"sheet": "P4 Capacity", "row_number": 4, "issue": "INVALID_PRESS_TONNAGE"}]
    assert not report.safe_to_execute


def test_plan_uses_checksums_master_press_list_only_for_missing_group_tonnage(tmp_path: Path) -> None:
    source = tmp_path / "grouped-capacity.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "P4 Capacity"
    sheet.append(["Machine No."])
    sheet.append(["Press 37"])
    workbook.save(source)
    workbook.close()
    master = tmp_path / "master-press-list.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Machine Specifications"
    sheet.append(["Machine Number", "U.S. Tons"])
    sheet.append([37, 500])
    workbook.save(master)
    workbook.close()

    report = plan_press_capacity_import(source, {"37": None}, master_press_list=master)

    assert report.safe_to_execute
    assert report.updates[0].source_tonnage == Decimal("500")
    assert report.supplementary_sources == {master.name: hashlib.sha256(master.read_bytes()).hexdigest()}


def test_receipt_is_redacted_and_immutable(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "capacity.xlsx", [("27", 110)])
    report = plan_press_capacity_import(source, {"27": None})

    receipt = write_immutable_receipt(report, tmp_path / "receipts")

    assert source.name in receipt.read_text(encoding="utf-8")
    assert str(source.parent) not in receipt.read_text(encoding="utf-8")
    with pytest.raises(FileExistsError):
        write_immutable_receipt(report, tmp_path / "receipts")


def test_inspection_and_plan_capture_workbook_structure_and_fail_closed_mappings(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "capacity.xlsx", [("27", 110), ("99", 75)])
    workbook = load_workbook(source)
    sheet = workbook.active
    sheet.row_dimensions[3].hidden = True
    sheet["C2"] = "=1+1"
    workbook.save(source)
    workbook.close()

    structure = inspect_press_capacity_workbook(source)
    report = plan_press_capacity_import(source, {"27": None})

    assert structure[0].name == "P4 Capacity"
    assert structure[0].recognized_layout
    assert structure[0].formula_cell_count == 1
    assert structure[0].hidden_row_count == 1
    assert report.workbook_sheets == structure
    assert report.mappings[-1].verification_class == "UNMAPPED"
    assert report.mappings[-1].conflict_reason == "NO_CANONICAL_MACHINE_MATCH"
    payload = report.to_dict()
    assert payload["mappings"][0]["source_tonnage"] == "110"
