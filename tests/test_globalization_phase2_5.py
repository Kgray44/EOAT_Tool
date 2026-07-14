from __future__ import annotations

import json
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from core.atlas_data_loader import _load_atlas_data_uncached
from core.globalization import workbook_import
from core.globalization.app_metadata import load_app_metadata
from core.globalization.config import load_or_create_global_config
from core.globalization.events import EventOutbox, validate_event_payload
from core.globalization.install_identity import load_or_create_install_identity
from core.globalization.pending_updates import PendingUpdateStore
from core.globalization.runtime_paths import atomic_write_json, ensure_runtime_layout, get_runtime_paths
from core.globalization.workbook_import import (
    deep_refresh_sqlite_cache,
    load_atlas_data_from_sqlite_cache,
    refresh_from_local_sqlite_cache,
)
from core.globalization.write_foundation import ChangeValidationService, WorkbookSyncService
from core.paths import resolve_project_paths
from tests.fixtures.fake_project import create_fake_eoat_project
from tests.fixtures.reference_workbooks import create_press_reference_workbooks


def test_app_metadata_and_install_identity_are_stable_and_installer_ready(tmp_path, monkeypatch) -> None:
    monkeypatch.setenv("EOAT_ATLAS_LOCALAPPDATA", str(tmp_path / "local"))
    runtime = ensure_runtime_layout(get_runtime_paths())
    metadata = load_app_metadata()
    generated = load_or_create_install_identity(runtime)
    loaded_again = load_or_create_install_identity(runtime)

    assert metadata.app_name == "EOAT Atlas"
    assert metadata.app_version == "0.9.1"
    assert metadata.release_id == "eoat-atlas-0.9.1"
    assert metadata.cache_schema_version >= 1
    assert metadata.event_schema_version == 1
    assert generated.install_id == loaded_again.install_id
    assert generated.app_instance_id == loaded_again.app_instance_id
    assert generated.generated_by == "dev_fallback"

    runtime_b = ensure_runtime_layout(get_runtime_paths(tmp_path / "installer-local"))
    atomic_write_json(
        runtime_b.install_identity_path,
        {
            "identity_schema_version": 1,
            "install_id": "installer-install-id",
            "app_instance_id": "installer-app-instance",
            "machine_name": "INSTALLER-PC",
            "windows_user": "installer-user",
            "generated_by": "installer",
        },
    )
    installer_identity = load_or_create_install_identity(runtime_b)

    assert installer_identity.install_id == "installer-install-id"
    assert installer_identity.app_instance_id == "installer-app-instance"
    assert installer_identity.generated_by == "installer"


def test_refresh_is_local_only_and_reapplies_pending_overlay(tmp_path, monkeypatch) -> None:
    monkeypatch.setenv("EOAT_ATLAS_LOCALAPPDATA", str(tmp_path / "local"))
    project = _project_with_sources(tmp_path / "project")
    source_paths = _source_paths(project)
    runtime = ensure_runtime_layout(get_runtime_paths())
    config = load_or_create_global_config(runtime)

    first = load_atlas_data_from_sqlite_cache(
        project,
        force_refresh=True,
        exclude_unaudited_tools=True,
        source_paths=source_paths,
        legacy_loader=_load_atlas_data_uncached,
    )
    eoat = next(record for record in first.eoats if record.status)
    PendingUpdateStore(runtime, config).create_update(
        entity_type="eoat",
        entity_id=eoat.eoat_id,
        field_name="status",
        expected_original_value=eoat.status,
        proposed_value="Needs Review",
        source_view="unit_test",
        source_action="local_refresh_overlay",
    )

    monkeypatch.setattr(workbook_import, "_stage_source_files", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("Refresh touched workbook staging.")))
    monkeypatch.setattr(workbook_import, "_required_sources_unavailable", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("Refresh checked workbook paths.")))

    refreshed = load_atlas_data_from_sqlite_cache(
        project,
        force_refresh=False,
        exclude_unaudited_tools=True,
        source_paths={**source_paths, "eoat_master_tracker": str(tmp_path / "missing.xlsx")},
        legacy_loader=_load_atlas_data_uncached,
    )
    effective = next(record for record in refreshed.eoats if record.eoat_id == eoat.eoat_id)

    assert refreshed.metrics["local_refresh"] is True
    assert refreshed.metrics["deep_refresh"] is False
    assert refreshed.metrics["pending_update_count"] == 1
    assert effective.status == "Needs Review"


def test_deep_refresh_uses_workbook_import_and_preserves_last_good_cache_on_failure(tmp_path, monkeypatch) -> None:
    monkeypatch.setenv("EOAT_ATLAS_LOCALAPPDATA", str(tmp_path / "local"))
    project = _project_with_sources(tmp_path / "project")
    source_paths = _source_paths(project)
    runtime = ensure_runtime_layout(get_runtime_paths())
    calls = {"stage": 0}
    original_stage = workbook_import._stage_source_files

    def counting_stage(*args, **kwargs):
        calls["stage"] += 1
        return original_stage(*args, **kwargs)

    monkeypatch.setattr(workbook_import, "_stage_source_files", counting_stage)
    first = deep_refresh_sqlite_cache(
        project,
        exclude_unaudited_tools=True,
        source_paths=source_paths,
        legacy_loader=_load_atlas_data_uncached,
        runtime=runtime,
    )

    assert calls["stage"] == 1
    assert first.metrics["deep_refresh"] is True
    assert runtime.db_path.exists()

    def failing_loader(*_args, **_kwargs):
        raise RuntimeError("simulated workbook import failure")

    with pytest.raises(RuntimeError, match="simulated workbook import failure"):
        deep_refresh_sqlite_cache(
            project,
            exclude_unaudited_tools=True,
            source_paths=source_paths,
            legacy_loader=failing_loader,
            runtime=runtime,
        )

    cached = refresh_from_local_sqlite_cache(runtime)
    failed_events = [event for event in EventOutbox(runtime, load_or_create_global_config(runtime)).list_events() if event["event_type"] == "deep_refresh_failed"]

    assert len(cached.eoats) == len(first.eoats)
    assert cached.metrics["local_refresh"] is True
    assert failed_events


def test_event_payloads_include_identity_version_and_sortable_filename(tmp_path, monkeypatch) -> None:
    monkeypatch.setenv("EOAT_ATLAS_LOCALAPPDATA", str(tmp_path / "local"))
    runtime = ensure_runtime_layout(get_runtime_paths())
    config = load_or_create_global_config(runtime)
    event = EventOutbox(runtime, config).create_event(
        event_type="unit_test_event",
        action="unit_test",
        entity_type="eoat",
        entity_id="P4-EOAT-0001",
        payload={
            "pending_update_ids": ["pending-1"],
            "field_changes": [{"field_name": "Status", "expected_original_value": "Old", "proposed_value": "Ready"}],
            "validation_result": {"status": "valid"},
            "conflict_result": {"status": "none"},
            "write_result": {"status": "not_written"},
        },
    )
    path = next(runtime.event_outbox_dir.glob(f"*_{event['event_id']}.json"))
    payload = json.loads(path.read_text(encoding="utf-8"))

    validate_event_payload(payload)
    assert path.name[:8].isdigit()
    assert payload["app_name"] == "EOAT Atlas"
    assert payload["app_version"] == config.app_version
    assert payload["release_id"] == config.release_id
    assert payload["install_id"] == config.install_id
    assert payload["app_instance_id"] == config.app_instance_id


def test_sandbox_sync_success_creates_backup_event_and_updates_pending_status(tmp_path, monkeypatch) -> None:
    monkeypatch.setenv("EOAT_ATLAS_LOCALAPPDATA", str(tmp_path / "local"))
    runtime = ensure_runtime_layout(get_runtime_paths())
    config = replace(load_or_create_global_config(runtime), write_mode="sandbox")
    workbook = _sandbox_workbook(tmp_path, status="Old")
    update = PendingUpdateStore(runtime, config).create_update(
        entity_type="eoat",
        entity_id="P4-EOAT-0001",
        field_name="Status",
        expected_original_value="Old",
        proposed_value="Ready",
        source_view="unit_test",
        source_action="sandbox_sync_success",
    )

    result = WorkbookSyncService(runtime, config).sync_pending_update_to_sandbox(
        update["pending_update_id"],
        workbook_path=workbook,
        sheet_name="EOAT Inventory",
        id_column="EOAT Assembly ID",
    )

    assert result["status"] == "succeeded"
    assert Path(result["backup_path"]).exists()
    assert _workbook_status(workbook) == "Ready"
    stored = PendingUpdateStore(runtime, config).get_update(update["pending_update_id"])
    assert stored is not None
    assert stored["sync_status"] == "applied"
    assert any(event["event_type"] == "workbook_sync_succeeded" for event in EventOutbox(runtime, config).list_events())


def test_sandbox_sync_conflict_and_failed_write_are_audited(tmp_path, monkeypatch) -> None:
    monkeypatch.setenv("EOAT_ATLAS_LOCALAPPDATA", str(tmp_path / "local"))
    runtime = ensure_runtime_layout(get_runtime_paths())
    config = replace(load_or_create_global_config(runtime), write_mode="sandbox")
    workbook = _sandbox_workbook(tmp_path, status="Changed")
    store = PendingUpdateStore(runtime, config)
    conflict_update = store.create_update(
        entity_type="eoat",
        entity_id="P4-EOAT-0001",
        field_name="Status",
        expected_original_value="Old",
        proposed_value="Ready",
        source_view="unit_test",
        source_action="sandbox_sync_conflict",
    )
    conflict = WorkbookSyncService(runtime, config).sync_pending_update_to_sandbox(
        conflict_update["pending_update_id"],
        workbook_path=workbook,
        sheet_name="EOAT Inventory",
        id_column="EOAT Assembly ID",
    )

    assert conflict["status"] == "conflict"
    assert _workbook_status(workbook) == "Changed"
    assert store.get_update(conflict_update["pending_update_id"])["sync_status"] == "conflict"

    failed_update = store.create_update(
        entity_type="eoat",
        entity_id="P4-EOAT-0001",
        field_name="Status",
        expected_original_value="Changed",
        proposed_value="Ready",
        source_view="unit_test",
        source_action="sandbox_sync_failure",
    )
    failed = WorkbookSyncService(runtime, config).sync_pending_update_to_sandbox(
        failed_update["pending_update_id"],
        workbook_path=workbook,
        sheet_name="Missing Sheet",
        id_column="EOAT Assembly ID",
    )
    event_types = {event["event_type"] for event in EventOutbox(runtime, config).list_events()}

    assert failed["status"] == "failed"
    assert "workbook_sync_conflict" in event_types
    assert "workbook_sync_failed" in event_types
    assert store.get_update(failed_update["pending_update_id"])["sync_status"] == "failed"


def test_production_writes_are_disabled_by_default(tmp_path, monkeypatch) -> None:
    monkeypatch.setenv("EOAT_ATLAS_LOCALAPPDATA", str(tmp_path / "local"))
    runtime = ensure_runtime_layout(get_runtime_paths())
    config = load_or_create_global_config(runtime)
    workbook = _sandbox_workbook(tmp_path, status="Old")
    update = PendingUpdateStore(runtime, config).create_update(
        entity_type="eoat",
        entity_id="P4-EOAT-0001",
        field_name="Status",
        expected_original_value="Old",
        proposed_value="Ready",
        source_view="unit_test",
        source_action="production_disabled",
    )

    valid, _message = ChangeValidationService(config).validate_submission(
        {"entity_type": "eoat", "entity_id": "P4-EOAT-0001", "field": "Status", "proposed_value": "Ready"}
    )
    result = WorkbookSyncService(runtime, config).sync_pending_update_to_sandbox(
        update["pending_update_id"],
        workbook_path=workbook,
        sheet_name="EOAT Inventory",
        id_column="EOAT Assembly ID",
    )

    assert valid is True
    assert result["status"] == "refused"
    assert _workbook_status(workbook) == "Old"
    assert any(event["event_type"] == "workbook_sync_refused" for event in EventOutbox(runtime, config).list_events())


def test_active_release_scope_has_no_old_app_surface_references() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    paths = [
        repo_root / "run_atlas.py",
        repo_root / "packaging" / "eoat_atlas_entry.py",
        repo_root / "EOAT_Atlas.spec",
        repo_root / "app" / "atlas",
        repo_root / "core" / "globalization",
        repo_root / "scripts" / "build_package.py",
        repo_root / "scripts" / "smoke_test_package.py",
    ]
    forbidden = ("Command Center", "command center", "classic Atlas", "legacy Atlas", "original Atlas", "atlas_window", "dashboard_ui")
    offenders: list[str] = []
    for path in paths:
        files = path.rglob("*.py") if path.is_dir() else (path,)
        for file_path in files:
            text = file_path.read_text(encoding="utf-8")
            for phrase in forbidden:
                if phrase in text:
                    offenders.append(f"{file_path.relative_to(repo_root)}: {phrase}")

    assert offenders == []


def _project_with_sources(root: Path) -> Path:
    project = create_fake_eoat_project(root)
    create_press_reference_workbooks(resolve_project_paths(project).reference_data)
    return project


def _source_paths(project: Path) -> dict[str, str]:
    paths = resolve_project_paths(project)
    return {
        "eoat_master_tracker": str(paths.master_workbook),
        "press_capacity_workbook": str(paths.reference_data / "press_capacity.xlsx"),
        "robot_workbook": str(paths.robot_info_workbook),
        "photos_root": str(paths.cell_photos),
        "output_folder": str(paths.final_handoff / "Atlas_Exports"),
        "reference_docs_folder": str(paths.standards),
    }


def _sandbox_workbook(tmp_path: Path, *, status: str) -> Path:
    path = tmp_path / f"sandbox_{status}.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "EOAT Inventory"
    sheet.append(["EOAT Assembly ID", "Status"])
    sheet.append(["P4-EOAT-0001", status])
    workbook.save(path)
    workbook.close()
    return path


def _workbook_status(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return str(workbook["EOAT Inventory"].cell(row=2, column=2).value or "")
    finally:
        workbook.close()
