from __future__ import annotations

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from core.audit_by_press import (
    AUDIT_BY_PRESS_SHEET,
    REFRESH_ACTION_NAME,
    UNASSIGNED_PRESS_GROUP,
    refresh_audit_by_press_view_action,
)
from core.audit_compatibility import create_compatibility_entries
from core.audit_constants import CYLINDER_COUNT_FIELD, CYLINDER_TYPE_FIELD, ENTRY_TYPE_COMPATIBLE, ENTRY_TYPE_FIELD
from core.audit_entries import (
    generate_audit_id,
    load_audit_entry,
    repair_workbook_schema,
    save_audit_entry,
    validate_audit_entry,
)
from core.paths import resolve_project_paths
from core.tool_fields import LEGACY_TOOL_FIELD
from core.workbook_schema import get_expected_headers, get_expected_sheets
from tests.fixtures.reference_workbooks import create_press_reference_workbooks


def test_generate_audit_id_and_add_row(fake_project):
    audit_id = generate_audit_id(fake_project, "2026-05-18")
    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": audit_id,
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "DEMO-PN-1200",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "EOAT Moves": "Both",
            "Connection Type": "ATI",
            "Gripper Model": "DESTACO-GRIP",
            "Status": "Audited",
        },
    )

    assert result.success is True
    assert audit_id.startswith("AUD-20260518-")
    assert result.files_created
    wb = load_workbook(resolve_project_paths(fake_project).master_workbook, read_only=True)
    ws = wb["EOAT Inventory"]
    assert ws.max_row == 2
    assert ws["A2"].value == audit_id
    headers = [cell.value for cell in ws[1]]
    assert headers[headers.index("Press/Machine #") + 1] == "Tool #"
    assert "Vacuum Zones" not in headers
    tooling_start = headers.index("EOAT Type")
    assert headers[tooling_start : tooling_start + 12] == [
        "EOAT Type",
        "EOAT Moves",
        "Connection Type",
        "Number of Parts Picked",
        CYLINDER_COUNT_FIELD,
        CYLINDER_TYPE_FIELD,
        "# of Grippers",
        "Gripper Type",
        "Gripper Model",
        "# of Cups",
        "Cup Type/Material",
        "Cup Diameter/Size",
    ]
    row_values = {
        headers[index]: value for index, value in enumerate(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    }
    assert row_values["Tool #"] == "DEMO-PN-1200"
    assert row_values["EOAT Moves"] == "Both"
    assert row_values["Connection Type"] == "ATI"
    assert row_values[CYLINDER_COUNT_FIELD] == "N/A"
    assert row_values[CYLINDER_TYPE_FIELD] == "N/A"
    assert row_values["Gripper Model"] == "N/A"
    assert row_values["# of Grippers"] == "N/A"
    assert row_values["Gripper Type"] == "N/A"
    assert row_values["Cleanroom/Non-Cleanroom"] == "Whiteroom"
    wb.close()


def test_number_of_parts_picked_save_load_and_header_order_are_unchanged(fake_project):
    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-PARTS-PICKED-STORAGE",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "DEMO-PN-1200",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Mechanical / Gripper",
            "EOAT Moves": "Part",
            "Connection Type": "ATI",
            "Number of Parts Picked": "3",
            "# of Grippers": "2",
            "Gripper Type": "Single Pressure",
            "Gripper Model": "Zimmer GPP",
            "Status": "Audited",
        },
    )

    assert result.success is True
    loaded = load_audit_entry(fake_project, "AUD-PARTS-PICKED-STORAGE")
    assert loaded["Number of Parts Picked"] == "3"

    wb = load_workbook(resolve_project_paths(fake_project).master_workbook, read_only=True)
    headers = [cell.value for cell in wb["EOAT Inventory"][1]]
    assert headers == list(get_expected_headers("EOAT Inventory"))
    assert headers[headers.index("Connection Type") + 1] == "Number of Parts Picked"
    assert headers[headers.index("Number of Parts Picked") + 1] == CYLINDER_COUNT_FIELD
    wb.close()


def test_save_with_current_schema_skips_heavy_schema_repair(fake_project, monkeypatch):
    repair = repair_workbook_schema(fake_project)
    assert repair.success, repair.errors

    def fail_migration(_workbook):
        raise AssertionError("current schema save should skip heavy migration")

    monkeypatch.setattr("core.audit_entries._migrate_workbook_tool_headers", fail_migration)

    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-SCHEMA-CURRENT",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Status": "In Progress",
        },
    )

    assert result.success, result.errors
    assert result.metrics["audit_save.schema_current"] is True
    assert result.metrics["audit_save.schema_repair_seconds"] == 0.0


def test_save_with_stale_schema_fails_fast_without_repair(fake_project, monkeypatch):
    import core.audit_entries as audit_entries

    calls = {"count": 0}

    def counted(_workbook):
        calls["count"] += 1
        raise AssertionError("normal save must not repair workbook schema")

    monkeypatch.setattr(audit_entries, "_migrate_workbook_tool_headers", counted)
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    ws.delete_cols(headers.index("EOAT Moves") + 1)
    workbook.save(workbook_path)
    workbook.close()

    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-SCHEMA-STALE",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Status": "In Progress",
        },
    )

    assert result.success is False
    assert "Workbook schema needs repair. Run Workbook Health > Repair Schema." in result.summary
    assert calls["count"] == 0
    assert result.metrics["audit_save.schema_current"] is False
    assert "audit_save.schema_check_seconds" in result.metrics


def test_eoat_moves_blank_stays_blank_and_warns_as_important(fake_project):
    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-MOVES-BLANK",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Status": "In Progress",
        },
    )

    assert result.success, result.errors
    assert "Missing important audit field: EOAT Moves" in "\n".join(result.warnings)
    loaded = load_audit_entry(fake_project, "AUD-MOVES-BLANK")
    assert loaded["EOAT Moves"] in (None, "")


def test_cylinder_fields_save_load_and_old_workbooks_migrate_safely(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    try:
        ws = workbook["EOAT Inventory"]
        headers = [cell.value for cell in ws[1]]
        for header in [CYLINDER_COUNT_FIELD, CYLINDER_TYPE_FIELD]:
            if header in headers:
                ws.delete_cols(headers.index(header) + 1)
                headers = [cell.value for cell in ws[1]]
        ws.append([""] * len(headers))
        row = ws.max_row
        for header, value in {
            "Audit ID": "AUD-OLD-NO-CYLINDERS",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 14",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Miscellaneous",
            "Status": "In Progress",
        }.items():
            ws.cell(row=row, column=headers.index(header) + 1).value = value
        workbook.save(workbook_path)
    finally:
        workbook.close()

    loaded_old = load_audit_entry(fake_project, "AUD-OLD-NO-CYLINDERS")
    assert loaded_old is not None
    assert CYLINDER_COUNT_FIELD not in loaded_old

    repair = repair_workbook_schema(fake_project)
    assert repair.success, repair.errors
    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-CYLINDERS-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 15",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Miscellaneous",
            CYLINDER_COUNT_FIELD: "2",
            CYLINDER_TYPE_FIELD: "Rotary",
            "Status": "In Progress",
        },
    )

    assert result.success, result.errors
    loaded = load_audit_entry(fake_project, "AUD-CYLINDERS-001")
    assert loaded[CYLINDER_COUNT_FIELD] == "2"
    assert loaded[CYLINDER_TYPE_FIELD] == "Rotary"

    default_type_result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-CYLINDERS-DEFAULT-TYPE",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 16",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Miscellaneous",
            CYLINDER_COUNT_FIELD: "3",
            CYLINDER_TYPE_FIELD: "",
            "Status": "In Progress",
        },
    )

    assert default_type_result.success, default_type_result.errors
    loaded_default = load_audit_entry(fake_project, "AUD-CYLINDERS-DEFAULT-TYPE")
    assert loaded_default[CYLINDER_COUNT_FIELD] == "3"
    assert loaded_default[CYLINDER_TYPE_FIELD] == "Linear"


def test_workbook_missing_eoat_moves_loads_and_migrates_without_overwriting_blanks(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        ws = workbook.create_sheet(sheet_name)
        headers = [header for header in get_expected_headers(sheet_name) if header != "EOAT Moves"]
        ws.append(headers)
    inventory = workbook["EOAT Inventory"]
    legacy_headers = [cell.value for cell in inventory[1]]
    for audit_id, press in [("AUD-OLD-MOVES-EDIT", "Press 12"), ("AUD-OLD-MOVES-BLANK", "Press 13")]:
        inventory.append([""] * len(legacy_headers))
        row = inventory.max_row
        inventory.cell(row=row, column=legacy_headers.index("Audit ID") + 1).value = audit_id
        inventory.cell(row=row, column=legacy_headers.index("Audit Date") + 1).value = "2026-05-18"
        inventory.cell(row=row, column=legacy_headers.index("Auditor") + 1).value = "KG"
        inventory.cell(row=row, column=legacy_headers.index("Plant/Area") + 1).value = "Plant 4"
        inventory.cell(row=row, column=legacy_headers.index("Press/Machine #") + 1).value = press
        inventory.cell(row=row, column=legacy_headers.index("Robot Type") + 1).value = "Wittmann R9"
        inventory.cell(row=row, column=legacy_headers.index("EOAT Type") + 1).value = "Vacuum"
        inventory.cell(row=row, column=legacy_headers.index("Status") + 1).value = "In Progress"
    workbook.save(workbook_path)
    workbook.close()

    loaded = load_audit_entry(fake_project, "AUD-OLD-MOVES-EDIT")
    assert loaded is not None
    assert "EOAT Moves" not in loaded

    repair = repair_workbook_schema(fake_project)
    assert repair.success, repair.errors
    loaded = load_audit_entry(fake_project, "AUD-OLD-MOVES-EDIT")
    assert loaded is not None
    result = save_audit_entry(fake_project, {**loaded, "EOAT Moves": "Sprue"}, allow_update=True)
    assert result.success, result.errors

    workbook = load_workbook(workbook_path, read_only=True)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    assert headers[headers.index("EOAT Type") + 1] == "EOAT Moves"
    rows = {
        row[headers.index("Audit ID")]: {headers[index]: value for index, value in enumerate(row)}
        for row in ws.iter_rows(min_row=2, values_only=True)
    }
    assert rows["AUD-OLD-MOVES-EDIT"]["EOAT Moves"] == "Sprue"
    assert rows["AUD-OLD-MOVES-BLANK"]["EOAT Moves"] in (None, "")
    workbook.close()


def test_compatibility_entries_copy_eoat_moves_from_source_audit(fake_project):
    reference_root = create_press_reference_workbooks(fake_project / "reference-data")
    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-MOVES-SOURCE",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "1",
            "Tool #": "DEMO-PN-0170",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "EOAT Moves": "Part",
            "Status": "Complete",
        },
    )
    assert result.success, result.errors

    result = create_compatibility_entries(
        fake_project, "AUD-MOVES-SOURCE", ["70"], reference_root / "press_capacity.xlsx"
    )

    assert result.success, result.errors
    rows = load_workbook(resolve_project_paths(fake_project).master_workbook, read_only=True)
    try:
        ws = rows["EOAT Inventory"]
        headers = [cell.value for cell in ws[1]]
        data_rows = [
            {headers[index]: value for index, value in enumerate(row)}
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]
    finally:
        rows.close()
    compatible = next(
        row
        for row in data_rows
        if row.get("Press/Machine #") == "70" and row.get(ENTRY_TYPE_FIELD) == ENTRY_TYPE_COMPATIBLE
    )
    assert compatible["EOAT Moves"] == "Part"


def test_audit_entry_update_existing(fake_project):
    entry = {
        "Audit ID": "AUD-20260518-001",
        "Audit Date": "2026-05-18",
        "Auditor": "KG",
        "Plant/Area": "Plant 4",
        "Press/Machine #": "Press 12",
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "Status": "Audited",
        "Priority": "Medium",
    }
    assert save_audit_entry(fake_project, entry).success
    entry["Priority"] = "High"
    result = save_audit_entry(fake_project, entry, allow_update=True)

    assert result.success is True
    assert result.metrics["updated"] is True


def test_blank_optional_audit_fields_save_as_na_without_replacing_defaults(fake_project):
    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-NA-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Status": "In Progress",
        },
    )

    assert result.success is True
    workbook = load_workbook(resolve_project_paths(fake_project).master_workbook, read_only=True)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    values = {
        headers[index]: value for index, value in enumerate(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    }
    assert values["Cleanroom/Non-Cleanroom"] == "Whiteroom"
    assert values["Tool #"] == "N/A"
    assert values["Connection Type"] == "N/A"
    assert values["EOAT Moves"] in (None, "")
    assert values["# of Cups"] == "N/A"
    assert values["Cup Type/Material"] == "Silicone"
    assert values["Known Issues"] == "N/A"
    assert values["Notes"] == "N/A"
    assert values["# of Grippers"] == "N/A"
    assert values["Gripper Type"] == "N/A"
    assert values["Gripper Model"] == "N/A"
    assert values["Entry Type"] == "Audited"
    assert values["Source Audit ID"] in (None, "")
    assert values["Compatibility Source"] in (None, "")
    assert all(
        value not in (None, "")
        for header, value in values.items()
        if header not in {"EOAT Moves", "Source Audit ID", "Compatibility Source"}
    )
    workbook.close()


def test_gripper_model_presets_save_actual_model_numbers(fake_project):
    for audit_id, ui_value, workbook_value in [
        ("AUD-GRIPPER-LARGE", "Large Double Gripper", "MHZL2-16D"),
        ("AUD-GRIPPER-SMALL", "Small Double Gripper", "MHZL2-10S"),
        ("AUD-GRIPPER-CUSTOM", "CUSTOM-GRIP-42", "CUSTOM-GRIP-42"),
    ]:
        result = save_audit_entry(
            fake_project,
            {
                "Audit ID": audit_id,
                "Audit Date": "2026-05-18",
                "Auditor": "KG",
                "Plant/Area": "Plant 4",
                "Press/Machine #": f"Press {audit_id}",
                "Robot Type": "Wittmann R9",
                "EOAT Type": "Mechanical / Gripper",
                "# of Grippers": "2",
                "Gripper Type": "Double Pressure",
                "Gripper Model": ui_value,
                "Status": "In Progress",
            },
        )
        assert result.success, result.errors
        loaded = load_audit_entry(fake_project, audit_id)
        assert loaded["Gripper Model"] == workbook_value


def test_gripper_count_and_pressure_type_validation(fake_project):
    bad_cups = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-CUP-COUNT-BAD",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press bad cup count",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "# of Cups": "two",
            "Status": "In Progress",
        },
    )
    assert bad_cups.success is False
    assert any("# of Cups must be a non-negative whole number" in error for error in bad_cups.errors)

    valid_cups = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-CUP-COUNT-VALID",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press valid cup count",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "# of Cups": "0",
            "Status": "In Progress",
        },
    )
    assert valid_cups.success, valid_cups.errors
    assert load_audit_entry(fake_project, "AUD-CUP-COUNT-VALID")["# of Cups"] == "0"

    bad_count = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-GRIPPER-COUNT-BAD",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press bad count",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Mechanical / Gripper",
            "# of Grippers": "two",
            "Gripper Type": "Single Pressure",
            "Status": "In Progress",
        },
    )
    assert bad_count.success is False
    assert any("# of Grippers must be a non-negative whole number" in error for error in bad_count.errors)

    bad_type = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-GRIPPER-TYPE-BAD",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press bad type",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Hybrid",
            "# of Grippers": "1",
            "Gripper Type": "Parallel jaw",
            "Status": "In Progress",
        },
    )
    assert bad_type.success is False
    assert any("Gripper Type must be one of: Single Pressure, Double Pressure" in error for error in bad_type.errors)

    bad_cylinder_count = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-CYLINDER-COUNT-BAD",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press bad cylinder count",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Miscellaneous",
            CYLINDER_COUNT_FIELD: "two",
            CYLINDER_TYPE_FIELD: "Linear",
            "Status": "In Progress",
        },
    )
    assert bad_cylinder_count.success is False
    assert any(
        f"{CYLINDER_COUNT_FIELD} must be a non-negative whole number" in error for error in bad_cylinder_count.errors
    )

    bad_cylinder_type = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-CYLINDER-TYPE-BAD",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press bad cylinder type",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Miscellaneous",
            CYLINDER_COUNT_FIELD: "1",
            CYLINDER_TYPE_FIELD: "Telescoping",
            "Status": "In Progress",
        },
    )
    assert bad_cylinder_type.success is False
    assert any(f"{CYLINDER_TYPE_FIELD} must be one of: Linear, Rotary" in error for error in bad_cylinder_type.errors)


def test_new_audit_sensor_and_documentation_defaults(fake_project):
    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-DEFAULTS-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Sensors Present?": "Yes",
            "Status": "In Progress",
        },
    )

    assert result.success, result.errors
    loaded = load_audit_entry(fake_project, "AUD-DEFAULTS-001")
    assert loaded["Vacuum Confirmation Present?"] == "Yes"
    assert loaded["Part-Present Detection Present?"] == "No"
    assert loaded["Spare Parts Identified?"] == "No"
    assert loaded["Drawing/CAD Available?"] == "No"
    assert loaded["BOM Available?"] == "No"
    assert loaded["Process Binder Complete?"] == "No"
    assert loaded["Photos Taken?"] == "No"


def test_part_present_detection_defaults_blank_sensor_fields_on_save(fake_project):
    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-PART-PRESENT-DEFAULTS",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Sensors Present?": "Yes",
            "Sensor Type": "",
            "Sensor Brand/Model": "",
            "Part-Present Detection Present?": "Yes",
            "Status": "In Progress",
        },
    )

    assert result.success, result.errors
    warning_text = "\n".join(result.warnings)
    assert "Missing important audit field: Sensor Type" not in warning_text
    assert "Missing important audit field: Sensor Brand/Model" not in warning_text
    loaded = load_audit_entry(fake_project, "AUD-PART-PRESENT-DEFAULTS")
    assert loaded["Sensor Type"] == "Reed Switch"
    assert loaded["Sensor Brand/Model"] == "SMC"


def test_part_present_detection_preserves_custom_sensor_values_on_save(fake_project):
    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-PART-PRESENT-CUSTOM",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 13",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Sensors Present?": "Yes",
            "Sensor Type": "Photoelectric",
            "Sensor Brand/Model": "Keyence PX",
            "Part-Present Detection Present?": "Yes",
            "Status": "In Progress",
        },
    )

    assert result.success, result.errors
    loaded = load_audit_entry(fake_project, "AUD-PART-PRESENT-CUSTOM")
    assert loaded["Sensor Type"] == "Photoelectric"
    assert loaded["Sensor Brand/Model"] == "Keyence PX"


def test_no_sensor_audit_saves_sensor_electrical_fields_as_na_without_named_warnings(fake_project):
    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-NO-SENSORS-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Sensors Present?": "No",
            "Tubing Condition": "OK",
            "Known Issues": "No sensor package on this EOAT.",
            "Status": "In Progress",
        },
    )

    assert result.success, result.errors
    warning_text = "\n".join(result.warnings)
    for field in [
        "Sensor Type",
        "Sensor Brand/Model",
        "Vacuum Confirmation Present?",
        "Part-Present Detection Present?",
    ]:
        assert field not in warning_text
    loaded = load_audit_entry(fake_project, "AUD-NO-SENSORS-001")
    assert loaded["Sensor Type"] == "N/A"
    assert loaded["Sensor Brand/Model"] == "N/A"
    assert loaded["Vacuum Confirmation Present?"] == "N/A"
    assert loaded["Part-Present Detection Present?"] == "N/A"
    assert loaded["Cable Management Condition"] == "N/A"
    assert loaded["Photos Taken?"] == "No"


def test_sensor_fields_are_important_when_sensors_are_applicable():
    _errors, warnings = validate_audit_entry(
        {
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Sensors Present?": "Yes",
            "Electrical/Wiring Present?": "Yes",
            "Sensor Type": "N/A",
            "Sensor Brand/Model": "",
            "Electrical Quick Disconnect Type": "",
            "Cable Management Condition": "",
            "Status": "In Progress",
            "Priority": "Medium",
            "Photos Taken?": "No",
        }
    )

    warning_text = "\n".join(warnings)
    assert "Missing important audit field: Sensor Type" in warning_text
    assert "Missing important audit field: Sensor Brand/Model" in warning_text
    assert "Missing important audit field: Electrical Quick Disconnect Type" in warning_text
    assert "Missing important audit field: Cable Management Condition" in warning_text


def test_new_audit_with_no_electrical_wiring_saves_electrical_fields_as_na(fake_project):
    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-NO-WIRING-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "DEMO-PN-1200",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Sensors Present?": "No",
            "Electrical/Wiring Present?": "No",
            "Electrical Quick Disconnect Type": "M12 stale",
            "Cable Management Condition": "Loose stale",
            "Status": "In Progress",
        },
    )

    assert result.success, result.errors
    loaded = load_audit_entry(fake_project, "AUD-NO-WIRING-001")
    assert loaded["Electrical/Wiring Present?"] == "No"
    assert loaded["Electrical Quick Disconnect Type"] == "N/A"
    assert loaded["Cable Management Condition"] == "N/A"


def test_repair_workbook_schema_adds_electrical_header_and_migrates_rows(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        ws = workbook.create_sheet(sheet_name)
        headers = [
            header
            for header in get_expected_headers(sheet_name)
            if sheet_name != "EOAT Inventory" or header != "Electrical/Wiring Present?"
        ]
        ws.append(headers)
    inventory = workbook["EOAT Inventory"]
    headers = [cell.value for cell in inventory[1]]
    quiet_no_wiring = {
        "Audit ID": "AUD-OLD-NO-WIRING",
        "Audit Date": "2026-05-18",
        "Auditor": "KG",
        "Plant/Area": "Plant 4",
        "Press/Machine #": "Press 12",
        "Tool #": "DEMO-PN-1200",
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "Sensors Present?": "No",
        "Sensor Type": "N/A",
        "Sensor Brand/Model": "Unknown / Not Checked",
        "Electrical Quick Disconnect Type": "",
        "Cable Management Condition": "N/A",
        "Status": "In Progress",
    }
    meaningful_wiring = {
        **quiet_no_wiring,
        "Audit ID": "AUD-OLD-WIRING-EVIDENCE",
        "Press/Machine #": "Press 13",
        "Electrical Quick Disconnect Type": "M12",
        "Cable Management Condition": "OK",
    }
    for values in [quiet_no_wiring, meaningful_wiring]:
        inventory.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()

    result = repair_workbook_schema(fake_project, log_activity=False)

    assert result.success, result.errors
    workbook = load_workbook(workbook_path, read_only=True)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    assert headers[headers.index("Part-Present Detection Present?") + 1] == "Electrical/Wiring Present?"
    rows = {
        row[headers.index("Audit ID")]: {headers[index]: value for index, value in enumerate(row)}
        for row in ws.iter_rows(min_row=2, values_only=True)
    }
    assert rows["AUD-OLD-NO-WIRING"]["Electrical/Wiring Present?"] == "No"
    assert rows["AUD-OLD-NO-WIRING"]["Electrical Quick Disconnect Type"] == "N/A"
    assert rows["AUD-OLD-NO-WIRING"]["Cable Management Condition"] == "N/A"
    assert rows["AUD-OLD-WIRING-EVIDENCE"]["Electrical/Wiring Present?"] == "Unknown / Not Checked"
    assert rows["AUD-OLD-WIRING-EVIDENCE"]["Electrical Quick Disconnect Type"] == "M12"
    assert rows["AUD-OLD-WIRING-EVIDENCE"]["Cable Management Condition"] == "OK"
    workbook.close()


def test_repair_workbook_schema_adds_gripper_count_and_normalizes_blank_gripper_fields(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        ws = workbook.create_sheet(sheet_name)
        headers = [
            header
            for header in get_expected_headers(sheet_name)
            if sheet_name != "EOAT Inventory" or header != "# of Grippers"
        ]
        ws.append(headers)
    inventory = workbook["EOAT Inventory"]
    headers = [cell.value for cell in inventory[1]]
    for values in [
        {
            "Audit ID": "AUD-GRIPPER-REPAIR-VACUUM",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Status": "In Progress",
        },
        {
            "Audit ID": "AUD-GRIPPER-REPAIR-PRESET",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 13",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Mechanical / Gripper",
            "Gripper Model": "Large Double Gripper",
            "Status": "In Progress",
        },
        {
            "Audit ID": "AUD-GRIPPER-REPAIR-CUSTOM",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 14",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Mechanical / Gripper",
            "Gripper Model": "CUSTOM-GRIP-42",
            "Status": "In Progress",
        },
    ]:
        inventory.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()

    result = repair_workbook_schema(fake_project, log_activity=False)

    assert result.success, result.errors
    workbook = load_workbook(workbook_path, read_only=True)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    rows = {
        row[headers.index("Audit ID")]: {headers[index]: value for index, value in enumerate(row)}
        for row in ws.iter_rows(min_row=2, values_only=True)
    }
    workbook.close()
    assert "# of Grippers" in headers
    assert rows["AUD-GRIPPER-REPAIR-VACUUM"]["# of Grippers"] == "N/A"
    assert rows["AUD-GRIPPER-REPAIR-VACUUM"]["Gripper Type"] == "N/A"
    assert rows["AUD-GRIPPER-REPAIR-VACUUM"]["Gripper Model"] == "N/A"
    assert rows["AUD-GRIPPER-REPAIR-PRESET"]["Gripper Model"] == "MHZL2-16D"
    assert rows["AUD-GRIPPER-REPAIR-PRESET"]["# of Grippers"] == "N/A"
    assert rows["AUD-GRIPPER-REPAIR-CUSTOM"]["Gripper Model"] == "CUSTOM-GRIP-42"


def test_repair_workbook_schema_adds_cup_count_before_cup_details_without_overwriting_data(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        ws = workbook.create_sheet(sheet_name)
        headers = [
            header
            for header in get_expected_headers(sheet_name)
            if sheet_name != "EOAT Inventory" or header != "# of Cups"
        ]
        ws.append(headers)
    inventory = workbook["EOAT Inventory"]
    headers = [cell.value for cell in inventory[1]]
    values = {
        "Audit ID": "AUD-CUP-COUNT-UPGRADE",
        "Audit Date": "2026-05-18",
        "Auditor": "KG",
        "Plant/Area": "Plant 4",
        "Press/Machine #": "Press 12",
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "Cup Type/Material": "Nitrile",
        "Cup Diameter/Size": "20 mm",
        "Vacuum Generator Type": "Venturi",
        "Status": "In Progress",
    }
    inventory.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()

    result = repair_workbook_schema(fake_project, log_activity=False)

    assert result.success, result.errors
    workbook = load_workbook(workbook_path, read_only=True)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    rows = {
        row[headers.index("Audit ID")]: {headers[index]: value for index, value in enumerate(row)}
        for row in ws.iter_rows(min_row=2, values_only=True)
    }
    workbook.close()
    assert headers.index("# of Cups") < headers.index("Cup Type/Material")
    assert rows["AUD-CUP-COUNT-UPGRADE"]["Cup Type/Material"] == "Nitrile"
    assert rows["AUD-CUP-COUNT-UPGRADE"]["Cup Diameter/Size"] == "20 mm"
    assert rows["AUD-CUP-COUNT-UPGRADE"]["# of Cups"] in (None, "")


def test_repair_workbook_schema_removes_vacuum_zones_with_backup_and_preserves_circuits(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        ws = workbook.create_sheet(sheet_name)
        headers = list(get_expected_headers(sheet_name))
        if sheet_name == "EOAT Inventory":
            headers.insert(headers.index("Vacuum Generator Type") + 1, "Vacuum Zones")
        ws.append(headers)
    inventory = workbook["EOAT Inventory"]
    headers = [cell.value for cell in inventory[1]]
    values = {
        "Audit ID": "AUD-VACUUM-ZONES-REMOVE",
        "Audit Date": "2026-05-18",
        "Auditor": "KG",
        "Plant/Area": "Plant 4",
        "Press/Machine #": "Press 12",
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "Vacuum Zones": "legacy value to delete",
        "EOAT Vacuum Circuits": "2",
        "EOAT Pressure Circuits": "1",
        "EOAT Interchangeable Circuits": "0",
        "Status": "In Progress",
    }
    inventory.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()

    result = repair_workbook_schema(fake_project, log_activity=False)

    assert result.success, result.errors
    assert result.metrics["vacuum_zones_columns_removed"] == 1
    backup_paths = list(
        (workbook_path.parent / "_backups").glob("EOAT_Master_Tracker_backup_before_removing_vacuum_zones_*.xlsx")
    )
    assert backup_paths
    workbook = load_workbook(workbook_path, read_only=True)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    rows = {
        row[headers.index("Audit ID")]: {headers[index]: value for index, value in enumerate(row)}
        for row in ws.iter_rows(min_row=2, values_only=True)
    }
    workbook.close()
    assert "Vacuum Zones" not in headers
    assert rows["AUD-VACUUM-ZONES-REMOVE"]["EOAT Vacuum Circuits"] == "2"
    assert rows["AUD-VACUUM-ZONES-REMOVE"]["EOAT Pressure Circuits"] == "1"
    assert rows["AUD-VACUUM-ZONES-REMOVE"]["EOAT Interchangeable Circuits"] == "0"
    assert rows["AUD-VACUUM-ZONES-REMOVE"]["Status"] == "In Progress"


def test_save_audit_entry_refuses_legacy_vacuum_zones_until_repair(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    insert_at = headers.index("Vacuum Generator Type") + 2
    ws.insert_cols(insert_at)
    ws.cell(row=1, column=insert_at).value = "Vacuum Zones"
    workbook.save(workbook_path)
    workbook.close()

    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-SAVE-REMOVES-VACUUM-ZONES",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Status": "In Progress",
        },
    )

    assert result.success is False
    assert "Workbook schema needs repair" in result.summary
    workbook = load_workbook(workbook_path, read_only=True)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    workbook.close()
    assert "Vacuum Zones" in headers


def test_cup_defaults_are_eoat_type_aware(fake_project):
    for audit_id, eoat_type, expected_cup, expected_size in [
        ("AUD-CUP-VACUUM", "Vacuum", "Silicone", "N/A"),
        ("AUD-CUP-HYBRID", "Hybrid", "Silicone", "N/A"),
        ("AUD-CUP-MECHANICAL", "Mechanical / Gripper", "N/A", "N/A"),
        ("AUD-CUP-MISC", "Miscellaneous", "N/A", "N/A"),
    ]:
        result = save_audit_entry(
            fake_project,
            {
                "Audit ID": audit_id,
                "Audit Date": "2026-05-18",
                "Auditor": "KG",
                "Plant/Area": "Plant 4",
                "Press/Machine #": f"Press {audit_id}",
                "Robot Type": "Wittmann R9",
                "EOAT Type": eoat_type,
                "Status": "In Progress",
            },
        )
        assert result.success, result.errors

    workbook = load_workbook(resolve_project_paths(fake_project).master_workbook, read_only=True)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    rows = {
        row[headers.index("Audit ID")]: {headers[index]: value for index, value in enumerate(row)}
        for row in ws.iter_rows(min_row=2, values_only=True)
    }
    assert rows["AUD-CUP-VACUUM"]["Cup Type/Material"] == "Silicone"
    assert rows["AUD-CUP-VACUUM"]["# of Cups"] == "N/A"
    assert rows["AUD-CUP-HYBRID"]["Cup Type/Material"] == "Silicone"
    assert rows["AUD-CUP-HYBRID"]["# of Cups"] == "N/A"
    assert rows["AUD-CUP-MECHANICAL"]["# of Cups"] == "N/A"
    assert rows["AUD-CUP-MECHANICAL"]["Cup Type/Material"] == "N/A"
    assert rows["AUD-CUP-MECHANICAL"]["Cup Diameter/Size"] == "N/A"
    assert rows["AUD-CUP-MECHANICAL"]["Number of Parts Picked"] == "N/A"
    assert rows["AUD-CUP-MECHANICAL"]["# of Grippers"] == "N/A"
    assert rows["AUD-CUP-MISC"]["Cup Type/Material"] == "N/A"
    workbook.close()


def test_existing_nonblank_values_are_not_overwritten_by_missing_partial_update(fake_project):
    entry = {
        "Audit ID": "AUD-PARTIAL-001",
        "Audit Date": "2026-05-18",
        "Auditor": "KG",
        "Plant/Area": "Plant 4",
        "Press/Machine #": "Press 12",
        "Tool #": "DEMO-PN-1200",
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "Connection Type": "ATI",
        "Known Issues": "Keep this observation.",
        "Status": "In Progress",
    }
    assert save_audit_entry(fake_project, entry).success
    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-PARTIAL-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Status": "Complete",
        },
        allow_update=True,
    )

    assert result.success is True
    loaded = load_audit_entry(fake_project, "AUD-PARTIAL-001")
    assert loaded["Tool #"] == "DEMO-PN-1200"
    assert loaded["Connection Type"] == "ATI"
    assert loaded["Known Issues"] == "Keep this observation."
    assert loaded["Status"] == "Complete"


def test_audit_entry_required_field_errors(fake_project):
    result = save_audit_entry(fake_project, {"Audit Date": "2026-05-18"})

    assert result.success is False
    assert any("Auditor" in error for error in result.errors)


def test_save_migrates_legacy_tool_header_and_preserves_data(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        ws = workbook.create_sheet(sheet_name)
        headers = get_expected_headers(sheet_name)
        headers = [LEGACY_TOOL_FIELD if header == "Tool #" else header for header in headers]
        ws.append(headers)
    inventory = workbook["EOAT Inventory"]
    legacy_headers = [cell.value for cell in inventory[1]]
    inventory.append([""] * len(legacy_headers))
    inventory.cell(row=2, column=legacy_headers.index("Audit ID") + 1).value = "AUD-20260518-LEG"
    inventory.cell(row=2, column=legacy_headers.index("Audit Date") + 1).value = "2026-05-18"
    inventory.cell(row=2, column=legacy_headers.index("Auditor") + 1).value = "KG"
    inventory.cell(row=2, column=legacy_headers.index("Plant/Area") + 1).value = "Plant 4"
    inventory.cell(row=2, column=legacy_headers.index("Press/Machine #") + 1).value = "Press 12"
    inventory.cell(row=2, column=legacy_headers.index(LEGACY_TOOL_FIELD) + 1).value = "PN-LEGACY"
    inventory.cell(row=2, column=legacy_headers.index("Robot Type") + 1).value = "Wittmann R9"
    inventory.cell(row=2, column=legacy_headers.index("EOAT Type") + 1).value = "Vacuum"
    inventory.cell(row=2, column=legacy_headers.index("Status") + 1).value = "Audited"
    workbook.save(workbook_path)
    workbook.close()

    loaded = load_audit_entry(fake_project, "AUD-20260518-LEG")
    assert loaded is not None
    assert loaded["Tool #"] == "PN-LEGACY"

    repair = repair_workbook_schema(fake_project)
    assert repair.success, repair.errors
    loaded = load_audit_entry(fake_project, "AUD-20260518-LEG")
    assert loaded is not None
    result = save_audit_entry(fake_project, {**loaded, "Priority": "High"}, allow_update=True)
    assert result.success is True

    workbook = load_workbook(workbook_path, read_only=True)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    values = {
        headers[index]: value for index, value in enumerate(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    }
    assert headers[headers.index("Press/Machine #") + 1] == "Tool #"
    assert LEGACY_TOOL_FIELD not in headers
    assert values["Tool #"] == "PN-LEGACY"
    assert "Setup ID" not in headers
    assert "Tool-Press Map" not in workbook.sheetnames
    for sheet in workbook.worksheets:
        assert LEGACY_TOOL_FIELD not in [cell.value for cell in sheet[1]]
    workbook.close()


def test_save_migrates_missing_connection_type_without_overwriting_existing_data(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        ws = workbook.create_sheet(sheet_name)
        headers = [header for header in get_expected_headers(sheet_name) if header != "Connection Type"]
        ws.append(headers)
    inventory = workbook["EOAT Inventory"]
    legacy_headers = [cell.value for cell in inventory[1]]
    inventory.append([""] * len(legacy_headers))
    inventory.cell(row=2, column=legacy_headers.index("Audit ID") + 1).value = "AUD-CONNECTION-MIGRATE"
    inventory.cell(row=2, column=legacy_headers.index("Audit Date") + 1).value = "2026-05-18"
    inventory.cell(row=2, column=legacy_headers.index("Auditor") + 1).value = "KG"
    inventory.cell(row=2, column=legacy_headers.index("Plant/Area") + 1).value = "Plant 4"
    inventory.cell(row=2, column=legacy_headers.index("Press/Machine #") + 1).value = "Press 12"
    inventory.cell(row=2, column=legacy_headers.index("Tool #") + 1).value = "DEMO-PN-1200"
    inventory.cell(row=2, column=legacy_headers.index("Robot Type") + 1).value = "Wittmann R9"
    inventory.cell(row=2, column=legacy_headers.index("EOAT Type") + 1).value = "Vacuum"
    inventory.cell(row=2, column=legacy_headers.index("Cup Type/Material") + 1).value = "Nitrile"
    inventory.cell(row=2, column=legacy_headers.index("Status") + 1).value = "Audited"
    workbook.save(workbook_path)
    workbook.close()

    loaded = load_audit_entry(fake_project, "AUD-CONNECTION-MIGRATE")
    assert loaded is not None
    assert "Connection Type" not in loaded
    assert loaded["Cup Type/Material"] == "Nitrile"

    repair = repair_workbook_schema(fake_project)
    assert repair.success, repair.errors
    loaded = load_audit_entry(fake_project, "AUD-CONNECTION-MIGRATE")
    assert loaded is not None
    result = save_audit_entry(fake_project, {**loaded, "Connection Type": "Lever Lock"}, allow_update=True)
    assert result.success is True

    workbook = load_workbook(workbook_path, read_only=True)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    values = {
        headers[index]: value for index, value in enumerate(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    }
    assert headers[headers.index("EOAT Type") + 2] == "Connection Type"
    assert values["Connection Type"] == "Lever Lock"
    assert values["Cup Type/Material"] == "Nitrile"
    assert "Setup ID" not in headers
    assert "Tool-Press Map" not in workbook.sheetnames
    workbook.close()


def test_save_migrates_missing_gripper_columns_without_overwriting_existing_data(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        ws = workbook.create_sheet(sheet_name)
        headers = [
            header
            for header in get_expected_headers(sheet_name)
            if header not in {"# of Grippers", "Gripper Model"}
        ]
        ws.append(headers)
    inventory = workbook["EOAT Inventory"]
    legacy_headers = [cell.value for cell in inventory[1]]
    inventory.append([""] * len(legacy_headers))
    inventory.cell(row=2, column=legacy_headers.index("Audit ID") + 1).value = "AUD-GRIPPER-MIGRATE"
    inventory.cell(row=2, column=legacy_headers.index("Audit Date") + 1).value = "2026-05-18"
    inventory.cell(row=2, column=legacy_headers.index("Auditor") + 1).value = "KG"
    inventory.cell(row=2, column=legacy_headers.index("Plant/Area") + 1).value = "Plant 4"
    inventory.cell(row=2, column=legacy_headers.index("Press/Machine #") + 1).value = "Press 12"
    inventory.cell(row=2, column=legacy_headers.index("Tool #") + 1).value = "DEMO-PN-1200"
    inventory.cell(row=2, column=legacy_headers.index("Robot Type") + 1).value = "Wittmann R9"
    inventory.cell(row=2, column=legacy_headers.index("Cleanroom/Non-Cleanroom") + 1).value = "Cleanroom"
    inventory.cell(row=2, column=legacy_headers.index("EOAT Type") + 1).value = "Mechanical / Gripper"
    inventory.cell(row=2, column=legacy_headers.index("Cup Type/Material") + 1).value = "Nitrile"
    inventory.cell(row=2, column=legacy_headers.index("Cup Diameter/Size") + 1).value = "20 mm"
    inventory.cell(row=2, column=legacy_headers.index("Status") + 1).value = "Audited"
    workbook.save(workbook_path)
    workbook.close()

    loaded = load_audit_entry(fake_project, "AUD-GRIPPER-MIGRATE")
    assert loaded is not None
    assert "# of Grippers" not in loaded
    assert "Gripper Model" not in loaded
    assert loaded["Cup Type/Material"] == "Nitrile"
    assert loaded["Cleanroom/Non-Cleanroom"] == "Cleanroom"

    repair = repair_workbook_schema(fake_project)
    assert repair.success, repair.errors
    loaded = load_audit_entry(fake_project, "AUD-GRIPPER-MIGRATE")
    assert loaded is not None
    result = save_audit_entry(
        fake_project,
        {
            **loaded,
            "# of Grippers": "2",
            "Gripper Type": "Single Pressure",
            "Gripper Model": "Zimmer GPP",
        },
        allow_update=True,
    )
    assert result.success is True

    workbook = load_workbook(workbook_path, read_only=True)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    values = {
        headers[index]: value for index, value in enumerate(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    }
    tooling_start = headers.index("EOAT Type")
    assert headers[tooling_start : tooling_start + 12] == [
        "EOAT Type",
        "EOAT Moves",
        "Connection Type",
        "Number of Parts Picked",
        CYLINDER_COUNT_FIELD,
        CYLINDER_TYPE_FIELD,
        "# of Grippers",
        "Gripper Type",
        "Gripper Model",
        "# of Cups",
        "Cup Type/Material",
        "Cup Diameter/Size",
    ]
    assert values["# of Grippers"] == "2"
    assert values["Gripper Type"] == "Single Pressure"
    assert values["Gripper Model"] == "Zimmer GPP"
    assert values["Cup Type/Material"] == "N/A"
    assert values["Cup Diameter/Size"] == "N/A"
    assert values["Cleanroom/Non-Cleanroom"] == "Cleanroom"
    assert "Setup ID" not in headers
    assert "Tool-Press Map" not in workbook.sheetnames
    workbook.close()


def test_setup_generated_master_tracker_uses_tool_header(tmp_path):
    import setup_eoat_project

    project_root = tmp_path / "EOAT_Standardization_Project"
    setup_eoat_project.configure_project_root(project_root)
    setup_eoat_project.create_workbook()

    workbook_path = project_root / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx"
    workbook = load_workbook(workbook_path, read_only=True)
    headers = [cell.value for cell in workbook["EOAT Inventory"][1]]
    assert headers[:6] == ["Audit ID", "Audit Date", "Auditor", "Plant/Area", "Press/Machine #", "Tool #"]
    assert headers[headers.index("Press/Machine #") + 1] == "Tool #"
    tooling_start = headers.index("EOAT Type")
    assert headers[tooling_start : tooling_start + 12] == [
        "EOAT Type",
        "EOAT Moves",
        "Connection Type",
        "Number of Parts Picked",
        CYLINDER_COUNT_FIELD,
        CYLINDER_TYPE_FIELD,
        "# of Grippers",
        "Gripper Type",
        "Gripper Model",
        "# of Cups",
        "Cup Type/Material",
        "Cup Diameter/Size",
    ]
    assert LEGACY_TOOL_FIELD not in headers
    assert "Setup ID" not in headers
    assert "Tool-Press Map" not in workbook.sheetnames
    workbook.close()


def test_migrated_tooling_columns_match_neighbor_formatting_and_validation(tmp_path):
    import setup_eoat_project

    project_root = tmp_path / "EOAT_Standardization_Project"
    setup_eoat_project.configure_project_root(project_root)
    setup_eoat_project.create_workbook()
    workbook_path = project_root / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx"

    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    for header in ["Connection Type", "# of Grippers", "Gripper Type", "Gripper Model"]:
        ws.delete_cols(headers.index(header) + 1)
        headers = [cell.value for cell in ws[1]]
    workbook.save(workbook_path)
    workbook.close()

    repair = repair_workbook_schema(project_root)
    assert repair.success, repair.errors
    result = save_audit_entry(
        project_root,
        {
            "Audit ID": "AUD-FORMAT-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Mechanical / Gripper",
            "Status": "In Progress",
        },
    )
    assert result.success, result.errors

    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    assert ws.auto_filter.ref == f"A1:{get_column_letter(len(headers))}1"

    def style_signature(cell):
        return (
            cell._style,
            cell.number_format,
            cell.alignment.horizontal,
            cell.alignment.vertical,
            cell.alignment.wrap_text,
            cell.fill.fill_type,
            cell.fill.fgColor.rgb,
            cell.font.bold,
            cell.border.left.style,
        )

    pairs = [
        ("EOAT Moves", "EOAT Type"),
        ("Connection Type", "EOAT Type"),
        (CYLINDER_COUNT_FIELD, "Number of Parts Picked"),
        (CYLINDER_TYPE_FIELD, CYLINDER_COUNT_FIELD),
        ("# of Grippers", "Number of Parts Picked"),
        ("# of Cups", "Number of Parts Picked"),
        ("Gripper Type", "Connection Type"),
        ("Gripper Model", "Gripper Type"),
    ]
    for target, source in pairs:
        target_col = headers.index(target) + 1
        source_col = headers.index(source) + 1
        assert (
            ws.column_dimensions[ws.cell(row=1, column=target_col).column_letter].width
            == ws.column_dimensions[ws.cell(row=1, column=source_col).column_letter].width
        )
        assert style_signature(ws.cell(row=1, column=target_col)) == style_signature(ws.cell(row=1, column=source_col))
        assert style_signature(ws.cell(row=2, column=target_col)) == style_signature(ws.cell(row=2, column=source_col))

    connection_col = headers.index("Connection Type") + 1
    connection_validations = [
        validation
        for validation in ws.data_validations.dataValidation
        for cell_range in validation.sqref.ranges
        if cell_range.min_col == connection_col
        and cell_range.max_col == connection_col
        and cell_range.min_row <= 2
        and cell_range.max_row >= 1000
    ]
    assert connection_validations
    assert "ATI" in connection_validations[0].formula1
    moves_col = headers.index("EOAT Moves") + 1
    moves_validations = [
        validation
        for validation in ws.data_validations.dataValidation
        for cell_range in validation.sqref.ranges
        if cell_range.min_col == moves_col
        and cell_range.max_col == moves_col
        and cell_range.min_row <= 2
        and cell_range.max_row >= 1000
    ]
    assert moves_validations
    assert moves_validations[0].formula1 == '"Part,Sprue,Both"'
    count_col = headers.index("# of Grippers") + 1
    count_validations = [
        validation
        for validation in ws.data_validations.dataValidation
        for cell_range in validation.sqref.ranges
        if cell_range.min_col == count_col
        and cell_range.max_col == count_col
        and cell_range.min_row <= 2
        and cell_range.max_row >= 1000
    ]
    assert count_validations
    assert count_validations[0].type == "whole"
    cup_count_col = headers.index("# of Cups") + 1
    cup_count_validations = [
        validation
        for validation in ws.data_validations.dataValidation
        for cell_range in validation.sqref.ranges
        if cell_range.min_col == cup_count_col
        and cell_range.max_col == cup_count_col
        and cell_range.min_row <= 2
        and cell_range.max_row >= 1000
    ]
    assert cup_count_validations
    assert cup_count_validations[0].type == "whole"
    cylinder_count_col = headers.index(CYLINDER_COUNT_FIELD) + 1
    cylinder_count_validations = [
        validation
        for validation in ws.data_validations.dataValidation
        for cell_range in validation.sqref.ranges
        if cell_range.min_col == cylinder_count_col
        and cell_range.max_col == cylinder_count_col
        and cell_range.min_row <= 2
        and cell_range.max_row >= 1000
    ]
    assert cylinder_count_validations
    assert cylinder_count_validations[0].type == "whole"
    cylinder_type_col = headers.index(CYLINDER_TYPE_FIELD) + 1
    cylinder_type_validations = [
        validation
        for validation in ws.data_validations.dataValidation
        for cell_range in validation.sqref.ranges
        if cell_range.min_col == cylinder_type_col
        and cell_range.max_col == cylinder_type_col
        and cell_range.min_row <= 2
        and cell_range.max_row >= 1000
    ]
    assert cylinder_type_validations
    assert "Linear" in cylinder_type_validations[0].formula1
    assert "Rotary" in cylinder_type_validations[0].formula1
    gripper_type_col = headers.index("Gripper Type") + 1
    gripper_type_validations = [
        validation
        for validation in ws.data_validations.dataValidation
        for cell_range in validation.sqref.ranges
        if cell_range.min_col == gripper_type_col
        and cell_range.max_col == gripper_type_col
        and cell_range.min_row <= 2
        and cell_range.max_row >= 1000
    ]
    assert gripper_type_validations
    assert "Single Pressure" in gripper_type_validations[0].formula1
    assert "Double Pressure" in gripper_type_validations[0].formula1
    workbook.close()


def _save_press_audit(project_root, audit_id, press, tool, plant="Plant 4", audit_date="2026-05-18"):
    result = save_audit_entry(
        project_root,
        {
            "Audit ID": audit_id,
            "Audit Date": audit_date,
            "Auditor": "KG",
            "Plant/Area": plant,
            "Press/Machine #": press,
            "Tool #": tool,
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Connection Type": "ATI",
            "Cleanroom/Non-Cleanroom": "Whiteroom",
            "Status": "In Progress",
            "Priority": "Medium",
            "Known Issues": "Review during audit.",
        },
    )
    assert result.success, result.errors
    return result


def _group_headers(ws):
    return [cell.value for cell in ws["A"] if isinstance(cell.value, str) and "total entr" in cell.value]


def test_audit_by_press_view_is_created_grouped_sorted_and_collapsible(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    _save_press_audit(fake_project, "AUD-PRESS-010", "10", "TOOL-10")
    _save_press_audit(fake_project, "AUD-PRESS-002B", "2", "TOOL-2B")
    _save_press_audit(fake_project, "AUD-PRESS-001A", "1", "TOOL-1A")
    _save_press_audit(fake_project, "AUD-PRESS-002A", "2", "TOOL-2A")
    _save_press_audit(fake_project, "AUD-PRESS-MISSING", "N/A", "TOOL-X")
    refresh = refresh_audit_by_press_view_action(fake_project, log_activity=False)
    assert refresh.success, refresh.errors

    workbook = load_workbook(workbook_path)
    inventory = workbook["EOAT Inventory"]
    inventory_headers = [cell.value for cell in inventory[1]]
    assert inventory.max_row == 6
    assert "Setup ID" not in inventory_headers
    assert "Tool-Press Map" not in workbook.sheetnames
    assert not list(inventory.merged_cells.ranges)

    assert AUDIT_BY_PRESS_SHEET in workbook.sheetnames
    view = workbook[AUDIT_BY_PRESS_SHEET]
    assert view.freeze_panes == "A4"
    assert view.auto_filter.ref == f"A3:{get_column_letter(view.max_column)}3"
    assert view["A2"].value.startswith("Last refreshed:")

    headers = _group_headers(view)
    assert headers == [
        "Plant 4 / Press 1 - 1 physical, 0 compatible, 1 total entry",
        "Plant 4 / Press 2 - 2 physical, 0 compatible, 2 total entries",
        "Plant 4 / Press 10 - 1 physical, 0 compatible, 1 total entry",
        f"{UNASSIGNED_PRESS_GROUP} - 1 physical, 0 compatible, 1 total entry",
    ]

    press_2_row = next(
        row
        for row in range(1, view.max_row + 1)
        if view.cell(row=row, column=1).value == "Plant 4 / Press 2 - 2 physical, 0 compatible, 2 total entries"
    )
    detail_rows = [press_2_row + 1, press_2_row + 2]
    assert [view.cell(row=row, column=6).value for row in detail_rows] == ["TOOL-2A", "TOOL-2B"]
    assert all(view.row_dimensions[row].outlineLevel == 1 for row in detail_rows)
    assert all(view.row_dimensions[row].hidden is True for row in detail_rows)
    assert view.row_dimensions[press_2_row].hidden is False
    assert view.row_dimensions[press_2_row].collapsed is True
    assert view.sheet_properties.outlinePr.summaryBelow is False

    assert view["A1"].fill.fgColor.rgb == "001F4E78"
    assert view["A3"].fill.fgColor.rgb == "00D9EAF7"
    assert view.cell(row=press_2_row, column=1).fill.fgColor.rgb == "00BDD7EE"
    workbook.close()


def test_audit_by_press_view_refreshes_after_new_update_and_duplicate_saves(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    _save_press_audit(fake_project, "AUD-REFRESH-001", "1", "TOOL-1")

    update = {
        "Audit ID": "AUD-REFRESH-001",
        "Audit Date": "2026-05-18",
        "Auditor": "KG",
        "Plant/Area": "Plant 4",
        "Press/Machine #": "2",
        "Tool #": "TOOL-2",
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "Connection Type": "ATI",
        "Cleanroom/Non-Cleanroom": "Whiteroom",
        "Status": "Complete",
        "Priority": "Medium",
        "Known Issues": "Updated press assignment.",
    }
    assert save_audit_entry(fake_project, update, allow_update=True).success
    _save_press_audit(fake_project, "AUD-REFRESH-002", "2", "TOOL-2B")
    refresh = refresh_audit_by_press_view_action(fake_project, log_activity=False)
    assert refresh.success, refresh.errors

    workbook = load_workbook(workbook_path)
    view = workbook[AUDIT_BY_PRESS_SHEET]
    headers = _group_headers(view)
    assert "Plant 4 / Press 1 - 1 physical, 0 compatible, 1 total entry" not in headers
    assert "Plant 4 / Press 2 - 2 physical, 0 compatible, 2 total entries" in headers
    detail_audit_ids = [
        cell.value for cell in view["A"] if isinstance(cell.value, str) and cell.value.startswith("AUD-REFRESH")
    ]
    assert detail_audit_ids == ["AUD-REFRESH-001", "AUD-REFRESH-002"]
    workbook.close()


def test_setup_generated_master_tracker_includes_audit_by_press_view(tmp_path):
    import setup_eoat_project

    project_root = tmp_path / "EOAT_Standardization_Project"
    setup_eoat_project.configure_project_root(project_root)
    setup_eoat_project.create_workbook()

    workbook_path = project_root / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx"
    workbook = load_workbook(workbook_path)
    assert AUDIT_BY_PRESS_SHEET in workbook.sheetnames
    view = workbook[AUDIT_BY_PRESS_SHEET]
    assert view["A1"].value == "EOAT Audit by Press"
    assert view["A2"].value.startswith("Last refreshed:")
    assert view["A4"].value == "No audit rows found."
    assert "Setup ID" not in [cell.value for cell in workbook["EOAT Inventory"][1]]
    assert "Tool-Press Map" not in workbook.sheetnames
    workbook.close()


def test_explicit_refresh_audit_by_press_view_action(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    _save_press_audit(fake_project, "AUD-MANUAL-001", "1", "TOOL-1")
    initial = refresh_audit_by_press_view_action(fake_project, log_activity=False)
    assert initial.success, initial.errors

    workbook = load_workbook(workbook_path)
    del workbook[AUDIT_BY_PRESS_SHEET]
    workbook.save(workbook_path)
    workbook.close()

    result = refresh_audit_by_press_view_action(fake_project, log_activity=False)

    assert result.success, result.errors
    assert result.tool_name == REFRESH_ACTION_NAME
    assert result.files_created
    workbook = load_workbook(workbook_path, read_only=True)
    assert AUDIT_BY_PRESS_SHEET in workbook.sheetnames
    workbook.close()
