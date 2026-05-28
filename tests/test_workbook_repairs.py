from __future__ import annotations

from openpyxl import Workbook, load_workbook

from core.audit.history import read_audit_history
from core.audit_entries import LEGACY_VACUUM_CUPS_FIELD, NA_VALUE
from core.paths import resolve_project_paths
from core.validation_findings import findings_from_result
from core.workbook_repairs import (
    FIX_CLEAR_STALE_HIDDEN_NA,
    FIX_REPAIR_LEGACY_HEADERS,
    apply_safe_fix,
    preview_safe_fix,
)
from core.workbook_schema import get_expected_headers, get_expected_sheets


def _append_inventory_row(workbook_path, values):
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    ws.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def test_clear_stale_hidden_values_requires_confirmation_backs_up_and_logs_history(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    _append_inventory_row(
        workbook_path,
        {
            "Audit ID": "AUD-STALE-HIDDEN-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "DEMO-PN-1200",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Sensors Present?": "No",
            "Sensor Type": "Photoeye stale value",
            "Status": "In Progress",
        },
    )

    preview = preview_safe_fix(fake_project, FIX_CLEAR_STALE_HIDDEN_NA)
    assert preview.can_apply
    assert any(change.audit_id == "AUD-STALE-HIDDEN-001" and change.column_name == "Sensor Type" for change in preview.changes)

    blocked = apply_safe_fix(fake_project, FIX_CLEAR_STALE_HIDDEN_NA)
    assert not blocked.success
    assert "confirmation" in blocked.summary.lower()

    result = apply_safe_fix(fake_project, FIX_CLEAR_STALE_HIDDEN_NA, confirm=True, log_activity=False)
    assert result.success, result.errors
    assert result.metrics["applied_change_count"] >= 1
    assert result.files_created
    assert result.metrics["validation_after_fix_finding_count"] >= 0

    workbook = load_workbook(workbook_path, read_only=True)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    rows = {
        row[headers.index("Audit ID")]: {headers[index]: value for index, value in enumerate(row)}
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[headers.index("Audit ID")]
    }
    workbook.close()
    assert rows["AUD-STALE-HIDDEN-001"]["Sensor Type"] == NA_VALUE
    assert any(record["event_type"] == "validation_auto_fix" for record in read_audit_history(fake_project))


def test_repair_legacy_headers_preview_and_apply_records_workbook_repair(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        ws = workbook.create_sheet(sheet_name)
        headers = list(get_expected_headers(sheet_name))
        if sheet_name == "EOAT Inventory":
            headers = [LEGACY_VACUUM_CUPS_FIELD if header == "Number of Parts Picked" else header for header in headers]
        ws.append(headers)
    workbook.save(workbook_path)
    workbook.close()

    preview = preview_safe_fix(fake_project, FIX_REPAIR_LEGACY_HEADERS)
    assert preview.can_apply
    assert any(change.current_value == LEGACY_VACUUM_CUPS_FIELD for change in preview.changes)

    result = apply_safe_fix(fake_project, FIX_REPAIR_LEGACY_HEADERS, confirm=True, log_activity=False)
    assert result.success, result.errors
    assert result.files_created
    assert result.metrics["validation_after_fix_success"] is True

    workbook = load_workbook(workbook_path, read_only=True)
    headers = [cell.value for cell in workbook["EOAT Inventory"][1]]
    workbook.close()
    assert LEGACY_VACUUM_CUPS_FIELD not in headers
    assert "Number of Parts Picked" in headers
    assert any(record["event_type"] == "workbook_repair" for record in read_audit_history(fake_project))


def test_validation_finding_for_stale_hidden_value_is_auto_fixable(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    _append_inventory_row(
        workbook_path,
        {
            "Audit ID": "AUD-STALE-FINDING-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 13",
            "Tool #": "DEMO-PN-1300",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Mechanical / Gripper",
            "# of Cups": "4",
            "Cup Type/Material": "stale cup",
            "Status": "In Progress",
        },
    )

    from core.validation import validate_project_foundation

    findings = findings_from_result(validate_project_foundation(fake_project))

    assert any(
        finding.audit_id == "AUD-STALE-FINDING-001"
        and finding.fix_id == FIX_CLEAR_STALE_HIDDEN_NA
        and finding.fix_available
        and finding.column_name in {"# of Cups", "Cup Type/Material"}
        for finding in findings
    )
