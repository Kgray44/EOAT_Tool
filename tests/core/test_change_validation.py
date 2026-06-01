from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from core.change_validation import (
    CHANGE_VALIDATION_ITEMS,
    build_change_validation_checklist,
    generate_change_validation_checklist,
)
from core.final_handoff import collect_handoff_sources
from core.paths import resolve_project_paths
from core.work_instruction_builder import generate_work_instructions
from core.workbook_schema import get_expected_headers


def _append_inventory_row(project_root: Path, values: dict[str, object]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    row = {header: "" for header in get_expected_headers("EOAT Inventory")}
    row.update(values)
    ws.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def test_change_validation_checklist_contains_required_items_and_missing_photo_blocker(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-CHG-001",
            "Press/Machine #": "Press 88",
            "EOAT Type": "Hybrid",
            "# of Cups": 4,
            "# of Grippers": 1,
            "Sensors Present?": "Yes",
            "Sensor Type": "Part-present sensor",
            "Quick Disconnects Present?": "Yes",
            "Pneumatic Quick Disconnect Type": "Push-to-connect",
            "Tubing Condition": "Good",
            "Cable Management Condition": "Good",
            "Mounting Hardware Condition": "Good",
            "Drop/Mis-Pick History": "Drops after changeover.",
            "Cycle Time Concern?": "Yes",
            "Scrap/Quality Concern?": "No",
            "Photos Taken?": "No",
            "Drawing/CAD Available?": "No",
            "BOM Available?": "No",
            "Process Binder Complete?": "No",
        },
    )

    checklist = build_change_validation_checklist(fake_project, audit_id="AUD-CHG-001", change_id="CHG-001")
    assert checklist is not None
    photos_item = next(item for item in checklist.items if item.item_id == "photos_captured")

    assert len(checklist.items) == len(CHANGE_VALIDATION_ITEMS)
    assert photos_item.status == "blocked - missing evidence"
    assert "photos are missing" in "\n".join(checklist.warnings)
    assert "cycle time concern" in checklist.to_markdown().casefold()


def test_generate_change_validation_writes_markdown_and_json_inside_project(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-CHG-002",
            "Press/Machine #": "Press 89",
            "EOAT Type": "Vacuum",
            "Photos Taken?": "Yes",
            "Photo Folder/Link": "01_EOAT_Audit/Cell_Photos/Press_89",
            "Drawing/CAD Available?": "Yes",
            "BOM Available?": "Yes",
            "Process Binder Complete?": "Yes",
        },
    )

    result = generate_change_validation_checklist(
        fake_project, audit_id="AUD-CHG-002", change_id="CHG-002", log_activity=False
    )

    assert result.success is True
    assert len(result.output_reports) == 2
    assert all(Path(path).is_relative_to(fake_project) for path in result.output_reports)
    json_path = next(Path(path) for path in result.output_reports if path.endswith(".json"))
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    assert payload["change_id"] == "CHG-002"
    assert len(payload["items"]) == len(CHANGE_VALIDATION_ITEMS)


def test_generated_work_instructions_and_change_validation_are_collected_for_final_handoff(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-HANDOFF-CHG",
            "Press/Machine #": "Press 90",
            "EOAT Type": "Vacuum",
            "Photos Taken?": "Yes",
            "Photo Folder/Link": "01_EOAT_Audit/Cell_Photos/Press_90",
            "Drawing/CAD Available?": "Yes",
            "BOM Available?": "Yes",
            "Process Binder Complete?": "Yes",
        },
    )
    generate_work_instructions(fake_project, audit_id="AUD-HANDOFF-CHG", log_activity=False)
    generate_change_validation_checklist(
        fake_project, audit_id="AUD-HANDOFF-CHG", change_id="CHG-HANDOFF", log_activity=False
    )

    sources = collect_handoff_sources(fake_project)

    assert any("Work_Instructions" in str(path) for path in sources["training"])
    assert any("Change_Validation" in str(path) for path in sources["change_validation"])
