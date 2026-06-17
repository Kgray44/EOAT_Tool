from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from core.audit_constants import (
    AUDIT_CONTEXT_COMPATIBILITY,
    AUDIT_CONTEXT_FIELD,
    COMPATIBILITY_CONFIDENCE_FIELD,
    COMPATIBILITY_SOURCE_FIELD,
    ENTRY_TYPE_AUDITED,
    ENTRY_TYPE_COMPATIBLE,
    ENTRY_TYPE_FIELD,
    PHYSICAL_AUDIT_VERIFIED_FIELD,
    SOURCE_AUDIT_ID_FIELD,
)
from core.compatibility_matrix import (
    COLUMN_MODE_PART_FAMILY,
    COLUMN_MODE_SOURCE_AUDIT,
    STATE_AUDITED,
    STATE_COMPATIBLE,
    STATE_CONFLICT,
    STATE_NEEDS_REVIEW,
    STATE_NOT_COMPATIBLE,
    build_compatibility_matrix,
    export_compatibility_matrix,
)
from core.paths import resolve_project_paths
from core.workbook_io import row_dicts
from core.workbook_schema import get_expected_headers


def _append_inventory_row(project_root, values: dict[str, object]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    row = {header: "" for header in get_expected_headers("EOAT Inventory")}
    row.update(
        {
            "Audit Date": "2026-05-18",
            "Auditor": "Synthetic Auditor",
            "Plant/Area": "Molding",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Part Family": "Family A",
            "Part Name/Description": "Synthetic part",
            "Status": "Audited",
            ENTRY_TYPE_FIELD: ENTRY_TYPE_AUDITED,
        }
    )
    row.update(values)
    ws.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def test_compatibility_matrix_builds_machine_rows_from_demo_data(usability_fake_project):
    summary = build_compatibility_matrix(usability_fake_project)

    assert summary.metrics["tools"] >= 1
    assert summary.metrics["machines"] >= 1
    assert summary.rows
    assert any(STATE_AUDITED in row.machine_states.values() for row in summary.rows)
    assert summary.standardization_opportunities
    assert summary.metrics["created_compatibility_rows"] == 0


def test_matrix_details_separate_physical_and_compatibility_rows(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-SOURCE-001",
            "Press/Machine #": "Press 1",
            "Tool #": "TOOL-A",
            "Part Family": "Family A",
            "Part Name/Description": "Part A",
            "EOAT Type": "Vacuum",
        },
    )
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-COMPAT-002",
            "Press/Machine #": "Press 2",
            "Tool #": "TOOL-A",
            "Part Family": "Family A",
            "Part Name/Description": "Part A",
            "EOAT Type": "Vacuum",
            ENTRY_TYPE_FIELD: ENTRY_TYPE_COMPATIBLE,
            AUDIT_CONTEXT_FIELD: AUDIT_CONTEXT_COMPATIBILITY,
            PHYSICAL_AUDIT_VERIFIED_FIELD: "No",
            COMPATIBILITY_CONFIDENCE_FIELD: "Press Capacity",
            SOURCE_AUDIT_ID_FIELD: "AUD-SOURCE-001",
            COMPATIBILITY_SOURCE_FIELD: "Press Capacity List",
        },
    )

    before_count = len(row_dicts(resolve_project_paths(fake_project).master_workbook, "EOAT Inventory"))
    summary = build_compatibility_matrix(fake_project)
    after_count = len(row_dicts(resolve_project_paths(fake_project).master_workbook, "EOAT Inventory"))
    source_cell = summary.cell("1", "TOOL-A | Part A")
    compatible_cell = summary.cell("2", "TOOL-A | Part A")

    assert before_count == after_count
    assert source_cell is not None
    assert compatible_cell is not None
    assert source_cell.compatibility_status == STATE_COMPATIBLE
    assert source_cell.physical_audit_ids == ("AUD-SOURCE-001",)
    assert source_cell.compatibility_audit_ids == ()
    assert compatible_cell.compatibility_status == STATE_COMPATIBLE
    assert compatible_cell.physical_audit_ids == ()
    assert compatible_cell.compatibility_audit_ids == ("AUD-COMPAT-002",)
    assert compatible_cell.source_audit_id == "AUD-SOURCE-001"
    assert compatible_cell.audit_context == AUDIT_CONTEXT_COMPATIBILITY
    assert compatible_cell.physical_audit_verified == "No"
    assert compatible_cell.compatibility_confidence == "Press Capacity"
    assert "Press Capacity List" in compatible_cell.compatibility_source
    assert "EOAT Type" in compatible_cell.fields_copied


def test_conflict_needs_review_and_not_compatible_states(fake_project):
    _append_inventory_row(
        fake_project,
        {"Audit ID": "AUD-CONFLICT-A", "Press/Machine #": "Press 3", "Tool #": "TOOL-B", "EOAT Type": "Vacuum"},
    )
    _append_inventory_row(
        fake_project,
        {"Audit ID": "AUD-CONFLICT-B", "Press/Machine #": "Press 3", "Tool #": "TOOL-B", "EOAT Type": "Hybrid"},
    )
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-REVIEW-004",
            "Press/Machine #": "Press 4",
            "Tool #": "TOOL-C",
            ENTRY_TYPE_FIELD: ENTRY_TYPE_COMPATIBLE,
            COMPATIBILITY_SOURCE_FIELD: "",
            SOURCE_AUDIT_ID_FIELD: "",
        },
    )
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-NOT-COMPAT-005",
            "Press/Machine #": "Press 5",
            "Tool #": "TOOL-D",
            "Status": "Not Compatible",
        },
    )

    summary = build_compatibility_matrix(fake_project)
    conflict_cell = summary.cell("3", "TOOL-B | Synthetic part")
    review_cell = summary.cell("4", "TOOL-C | Synthetic part")
    not_compatible_cell = summary.cell("5", "TOOL-D | Synthetic part")

    assert conflict_cell is not None
    assert conflict_cell.compatibility_status == STATE_CONFLICT
    assert any("Conflicting EOAT Type" in conflict for conflict in conflict_cell.conflicts)
    assert review_cell is not None
    assert review_cell.compatibility_status == STATE_NEEDS_REVIEW
    assert any("Source Audit ID" in item for item in review_cell.missing_data)
    assert not_compatible_cell is not None
    assert not_compatible_cell.compatibility_status == STATE_NOT_COMPATIBLE


def test_source_audit_and_part_family_column_modes(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-SOURCE-FAM",
            "Press/Machine #": "Press 10",
            "Tool #": "TOOL-F",
            "Part Family": "Shared Family",
        },
    )
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-COMPAT-FAM",
            "Press/Machine #": "Press 11",
            "Tool #": "TOOL-F",
            "Part Family": "Shared Family",
            ENTRY_TYPE_FIELD: ENTRY_TYPE_COMPATIBLE,
            SOURCE_AUDIT_ID_FIELD: "AUD-SOURCE-FAM",
            COMPATIBILITY_SOURCE_FIELD: "Manual Review",
        },
    )

    by_source = build_compatibility_matrix(fake_project, column_mode=COLUMN_MODE_SOURCE_AUDIT)
    by_family = build_compatibility_matrix(fake_project, column_mode=COLUMN_MODE_PART_FAMILY)

    assert any(column.source_audit_id == "AUD-SOURCE-FAM" for column in by_source.columns)
    assert by_source.cell("11", "AUD-SOURCE-FAM | TOOL-F | Synthetic part").compatibility_status == STATE_COMPATIBLE
    assert any(column.part_family == "Shared Family" for column in by_family.columns)
    assert by_family.cell("10", "Shared Family").physical_audit_ids == ("AUD-SOURCE-FAM",)


def test_export_csv_markdown_includes_cell_details_without_creating_rows(fake_project):
    _append_inventory_row(
        fake_project, {"Audit ID": "AUD-EXPORT-001", "Press/Machine #": "Press 20", "Tool #": "TOOL-X"}
    )
    before_count = len(row_dicts(resolve_project_paths(fake_project).master_workbook, "EOAT Inventory"))

    result = export_compatibility_matrix(fake_project, log_activity=False)

    after_count = len(row_dicts(resolve_project_paths(fake_project).master_workbook, "EOAT Inventory"))
    markdown = next(Path(path) for path in result.output_reports if path.endswith(".md"))
    csv = next(Path(path) for path in result.output_reports if path.endswith(".csv"))
    text = markdown.read_text(encoding="utf-8")
    csv_text = csv.read_text(encoding="utf-8")

    assert result.success is True
    assert before_count == after_count
    assert "Compatibility Matrix 2.0" in text
    assert "AUD-EXPORT-001" in text
    assert "Recommended Action" in csv_text
