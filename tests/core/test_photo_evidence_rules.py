from __future__ import annotations

import hashlib

from openpyxl import load_workbook

from core.paths import resolve_project_paths
from core.photo_evidence import (
    evidence_coverage_for_audit,
    link_photo_to_audit_field,
    linked_audit_field_for_photo,
    photo_index_path_findings,
    validate_photo_evidence,
)
from core.photo_evidence_rules import photo_evidence_rule_by_key, required_photo_evidence_rules
from core.photo_indexing import intake_photos, list_incoming_photos, preview_photo_intake
from core.search import search_project
from core.workbook_io import row_dicts
from core.workbook_schema import get_expected_headers


def _append_sheet_row(project_root, sheet_name: str, values: dict[str, str]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook[sheet_name]
    headers = [cell.value for cell in ws[1]]
    ws.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def _audit_row(audit_id: str, eoat_type: str, **overrides) -> dict[str, str]:
    row = {header: "" for header in get_expected_headers("EOAT Inventory")}
    row.update(
        {
            "Audit ID": audit_id,
            "Audit Date": "2026-05-18",
            "Auditor": "Synthetic Auditor",
            "Plant/Area": "Plant 4",
            "Press/Machine #": f"Press {audit_id[-1]}",
            "Robot Type": "Wittmann R9",
            "EOAT Type": eoat_type,
            "Status": "Complete",
            "Priority": "High",
            "Pilot Candidate?": "No",
            "Sensors Present?": "No",
            "Electrical/Wiring Present?": "No",
            "Quick Disconnects Present?": "No",
            "Known Issues": "No issue observed.",
            "Process Binder Complete?": "No",
            "Photos Taken?": "No",
        }
    )
    row.update(overrides)
    return row


def _required_keys(row: dict[str, str]) -> set[str]:
    return {rule.key for rule in required_photo_evidence_rules(row)}


def _status(coverage, category: str):
    return next(status for status in coverage.statuses if status.category == category)


def _file_hash(path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_required_shots_by_eoat_type_and_audit_details():
    vacuum = _required_keys(_audit_row("AUD-SHOT-001", "Vacuum"))
    gripper = _required_keys(_audit_row("AUD-SHOT-002", "Mechanical / Gripper"))
    hybrid = _required_keys(_audit_row("AUD-SHOT-003", "Hybrid"))

    assert "vacuum_cups" in vacuum
    assert "grippers" not in vacuum
    assert "grippers" in gripper
    assert "vacuum_cups" not in gripper
    assert {"vacuum_cups", "grippers"}.issubset(hybrid)

    detailed = _required_keys(
        _audit_row(
            "AUD-SHOT-004",
            "Hybrid",
            **{
                "# of Cylinders": "2",
                "Sensors Present?": "Yes",
                "Electrical/Wiring Present?": "Yes",
                "Quick Disconnects Present?": "Yes",
                "Robot Vacuum Circuits": "1",
                "EOAT Pressure Circuits": "1",
            },
        )
    )

    assert {
        "cylinders",
        "sensors",
        "cable_management",
        "quick_disconnects",
        "tubing_routing",
        "robot_side_pneumatics",
        "eoat_pneumatic_circuits",
    }.issubset(detailed)
    assert photo_evidence_rule_by_key("Tool Connection").key == "robot_connection"
    assert photo_evidence_rule_by_key("Grippers").linked_fields == (
        "# of Grippers",
        "Gripper Type",
        "Gripper Model",
    )


def test_broken_photo_path_creates_finding(fake_project):
    audit_id = "AUD-BROKEN-PHOTO-001"
    _append_sheet_row(fake_project, "EOAT Inventory", _audit_row(audit_id, "Vacuum"))
    _append_sheet_row(
        fake_project,
        "Photo Index",
        {
            "Photo ID": "PHO-BROKEN-001",
            "Date Taken": "2026-05-18",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 1",
            "EOAT Area Shown": "Vacuum Cups",
            "Photo Filename": "missing_vacuum_cup.jpg",
            "Folder Path": str(fake_project / "missing" / "photo-folder"),
            "Related Audit ID": audit_id,
        },
    )

    warnings, metrics, findings = validate_photo_evidence(fake_project)

    assert metrics["photo_evidence_broken_path_count"] >= 1
    assert any(
        finding.category == "photo_evidence_path" and "PHO-BROKEN-001" in finding.message for finding in findings
    )
    assert any("broken photo path" in warning.casefold() for warning in warnings)
    assert any("PHO-BROKEN-001" in finding.message for finding in photo_index_path_findings(fake_project))


def test_photo_index_search_still_returns_photos(usability_fake_project):
    results = search_project(usability_fake_project, "Press 101")

    assert any(result.result_type == "photo" for result in results)


def test_intake_preview_does_not_modify_files(usability_fake_project):
    paths = resolve_project_paths(usability_fake_project)
    workbook_hash_before = _file_hash(paths.master_workbook)
    rows_before = row_dicts(paths.master_workbook, "Photo Index")
    photos = list_incoming_photos(usability_fake_project)

    plan = preview_photo_intake(
        usability_fake_project,
        [photos[0]],
        "Molding",
        "Press 101",
        "2026-05-19",
        "Sensors",
        tool_number="TOOL-A",
    )

    assert plan
    assert _file_hash(paths.master_workbook) == workbook_hash_before
    assert row_dicts(paths.master_workbook, "Photo Index") == rows_before
    assert photos[0].exists()
    assert not plan[0].target.exists()


def test_confirmed_intake_updates_index_safely(usability_fake_project):
    photos = list_incoming_photos(usability_fake_project)

    result = intake_photos(
        usability_fake_project,
        [photos[0]],
        "Molding",
        "Press 101",
        "2026-05-19",
        "Cylinders",
        tool_number="TOOL-A",
        related_audit_id="AUD-20260518-001",
        description="Synthetic cylinder evidence.",
        notes="Linked audit field: # of Cylinders",
        copy_mode=True,
        log_activity=False,
    )

    rows = row_dicts(resolve_project_paths(usability_fake_project).master_workbook, "Photo Index")

    assert result.success is True
    assert any(path.endswith(".xlsx") for path in result.files_created)
    assert any(path.endswith(".jpg") or path.endswith(".png") for path in result.files_created)
    assert any(row["Related Audit ID"] == "AUD-20260518-001" and row["EOAT Area Shown"] == "Cylinders" for row in rows)
    assert photos[0].exists()


def test_structured_and_legacy_linked_audit_field_values_are_read(fake_project):
    _append_sheet_row(
        fake_project,
        "Photo Index",
        {
            "Photo ID": "PHO-LINK-STRUCTURED",
            "Date Taken": "2026-05-18",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 1",
            "EOAT Area Shown": "Sensors",
            "Linked Audit Field": "Sensor Type",
        },
    )
    _append_sheet_row(
        fake_project,
        "Photo Index",
        {
            "Photo ID": "PHO-LINK-LEGACY",
            "Date Taken": "2026-05-18",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 1",
            "EOAT Area Shown": "Cylinders",
            "Notes": "Linked audit field: # of Cylinders",
        },
    )

    rows = row_dicts(resolve_project_paths(fake_project).master_workbook, "Photo Index")

    assert linked_audit_field_for_photo(rows[0]) == "Sensor Type"
    assert linked_audit_field_for_photo(rows[1]) == "# of Cylinders"


def test_link_photo_to_audit_field_updates_structured_column(fake_project):
    _append_sheet_row(
        fake_project,
        "Photo Index",
        {
            "Photo ID": "PHO-LINK-WRITE",
            "Date Taken": "2026-05-18",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 1",
            "EOAT Area Shown": "Grippers",
        },
    )

    result = link_photo_to_audit_field(fake_project, "PHO-LINK-WRITE", "Gripper Model", log_activity=False)
    rows = row_dicts(resolve_project_paths(fake_project).master_workbook, "Photo Index")
    row = next(row for row in rows if row["Photo ID"] == "PHO-LINK-WRITE")

    assert result.success is True
    assert row["Linked Audit Field"] == "Gripper Model"


def test_photo_evidence_fallback_uses_machine_and_tool_when_audit_id_is_blank(fake_project):
    _append_sheet_row(
        fake_project,
        "EOAT Inventory",
        _audit_row(
            "AUD-TOOL-MATCH-001",
            "Mechanical / Gripper",
            **{"Press/Machine #": "Press 55", "Tool #": "TOOL-A"},
        ),
    )
    _append_sheet_row(
        fake_project,
        "Photo Index",
        {
            "Photo ID": "PHO-WRONG-TOOL",
            "Date Taken": "2026-05-18",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 55",
            "Tool #": "TOOL-B",
            "EOAT Area Shown": "Grippers",
        },
    )

    coverage = evidence_coverage_for_audit(fake_project, "AUD-TOOL-MATCH-001")
    assert _status(coverage, "grippers").present is False

    _append_sheet_row(
        fake_project,
        "Photo Index",
        {
            "Photo ID": "PHO-RIGHT-TOOL",
            "Date Taken": "2026-05-18",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 55",
            "Tool #": "TOOL-A",
            "EOAT Area Shown": "Grippers",
        },
    )

    coverage = evidence_coverage_for_audit(fake_project, "AUD-TOOL-MATCH-001")
    assert _status(coverage, "grippers").present is True
