from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from core.bad_actor_detector import BAD_ACTOR_SCORE_FORMULA, detect_bad_actors, export_bad_actor_report
from core.paths import resolve_project_paths
from core.workbook_io import row_dicts
from core.workbook_schema import get_expected_headers


def _append_sheet_row(project_root, sheet_name: str, values: dict[str, object]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook[sheet_name]
    headers = [cell.value for cell in ws[1]]
    row = {header: "" for header in get_expected_headers(sheet_name)}
    row.update(values)
    ws.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def _audit_row(audit_id: str, machine: str, **overrides) -> dict[str, object]:
    row = {
        "Audit ID": audit_id,
        "Audit Date": "2026-05-18",
        "Auditor": "Synthetic Auditor",
        "Plant/Area": "Molding",
        "Press/Machine #": machine,
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "Priority": "Medium",
        "Status": "Audited",
        "Follow-Up Needed": "No",
        "Drop/Mis-Pick History": "No",
        "Cycle Time Concern?": "No",
        "Scrap/Quality Concern?": "No",
        "Maintenance Frequency": "",
        "Spare Parts Identified?": "Yes",
        "BOM Available?": "Yes",
        "Drawing/CAD Available?": "Yes",
        "Process Binder Complete?": "Yes",
    }
    row.update(overrides)
    return row


def _score(summary, machine: str):
    return next(item for item in summary.rankings if item.machine == machine)


def test_bad_actor_scoring_components_and_formula(fake_project):
    _append_sheet_row(
        fake_project,
        "EOAT Inventory",
        _audit_row(
            "AUD-BA-001",
            "Press 30",
            **{
                "Priority": "High",
                "Follow-Up Needed": "Yes",
                "Drop/Mis-Pick History": "Recurring part drops.",
                "Cycle Time Concern?": "Yes",
                "Scrap/Quality Concern?": "Yes",
                "Maintenance Frequency": "Weekly",
                "BOM Available?": "No",
                "Drawing/CAD Available?": "",
            },
        ),
    )
    _append_sheet_row(
        fake_project,
        "Issue Log",
        {
            "Issue ID": "ISS-BA-001",
            "Press/Machine #": "Press 30",
            "Issue Category": "Part drop",
            "Issue Description": "Part drop and mis-pick history.",
            "Severity": 8,
            "Status": "Open",
            "Follow-Up Date": "2026-05-21",
        },
    )
    _append_sheet_row(
        fake_project,
        "KPI Baseline",
        {
            "KPI ID": "KPI-BA-001",
            "Press/Machine #": "Press 30",
            "Part Drops": 5,
            "Mis-Picks": 1,
            "Scrap Quantity": 3,
        },
    )
    _append_sheet_row(
        fake_project,
        "FMEA Draft",
        {
            "FMEA ID": "FMEA-BA-001",
            "Press/Machine #": "Press 30",
            "Failure Mode": "Vacuum loss causes drop.",
            "Severity": 9,
            "Frequency": 6,
            "Detectability": 4,
        },
    )

    summary = detect_bad_actors(fake_project)
    score = _score(summary, "30")

    assert summary.score_formula == BAD_ACTOR_SCORE_FORMULA
    assert score.issue_count == 1
    assert score.high_priority_count >= 3
    assert score.critical_priority_count >= 1
    assert score.follow_up_count == 2
    assert score.drop_history >= 3
    assert score.scrap_concern == 2
    assert score.cycle_time_concern == 1
    assert score.maintenance_frequency == 4
    assert score.documentation_gap_penalty == 4
    assert score.score > 50
    assert summary.rankings[0].machine == "30"


def test_bad_actor_missing_evidence_is_reported_not_fabricated(fake_project):
    _append_sheet_row(fake_project, "EOAT Inventory", _audit_row("AUD-BA-CLEAN", "Press 31"))

    summary = detect_bad_actors(fake_project)
    score = _score(summary, "31")

    assert score.score == 0
    assert any("No Issue Log evidence" in item for item in score.missing_evidence)
    assert any("No KPI baseline evidence" in item for item in score.missing_evidence)
    assert "collect missing sources" in score.recommended_action


def test_bad_actor_documentation_gap_penalty(fake_project):
    _append_sheet_row(
        fake_project,
        "EOAT Inventory",
        _audit_row(
            "AUD-BA-DOC",
            "Press 32",
            **{
                "Spare Parts Identified?": "",
                "BOM Available?": "No",
                "Drawing/CAD Available?": "Unknown / Not Checked",
                "Process Binder Complete?": "No",
            },
        ),
    )

    summary = detect_bad_actors(fake_project)
    score = _score(summary, "32")

    assert score.documentation_gap_penalty == 8
    assert score.score == 8


def test_bad_actor_export_includes_formula_and_does_not_modify_workbook(fake_project):
    _append_sheet_row(fake_project, "EOAT Inventory", _audit_row("AUD-BA-EXPORT", "Press 33", Priority="High"))
    workbook_path = resolve_project_paths(fake_project).master_workbook
    before_count = len(row_dicts(workbook_path, "EOAT Inventory"))

    result = export_bad_actor_report(fake_project, log_activity=False)

    after_count = len(row_dicts(workbook_path, "EOAT Inventory"))
    markdown = next(Path(path) for path in result.output_reports if path.endswith(".md"))
    text = markdown.read_text(encoding="utf-8")

    assert result.success is True
    assert before_count == after_count
    assert BAD_ACTOR_SCORE_FORMULA in text
    assert "Press 33" not in text
    assert "33" in text
