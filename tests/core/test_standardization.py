from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from core.paths import resolve_project_paths
from core.standardization import (
    analyze_standardization_opportunities,
    generate_standardization_report,
    load_part_aliases,
    normalize_part_alias,
)
from core.workbook_schema import get_expected_headers


def _append_inventory_row(project_root, values: dict[str, object]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    row = {header: "" for header in get_expected_headers("EOAT Inventory")}
    row.update(
        {
            "Audit Date": "2026-05-18",
            "Auditor": "Synthetic Auditor",
            "Plant/Area": "Molding",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Mechanical / Gripper",
            "Status": "Audited",
            "Spare Parts Identified?": "Yes",
            "BOM Available?": "Yes",
            "Drawing/CAD Available?": "Yes",
            "Process Binder Complete?": "Yes",
        }
    )
    row.update(values)
    ws.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def _component_row(analysis, category: str, component: str) -> dict[str, object]:
    return next(
        row
        for row in analysis.component_frequency_table
        if row["Category"] == category and row["Component"] == component
    )


def test_alias_normalization_and_config_alias_file(tmp_path):
    custom_alias_path = tmp_path / "part_aliases.json"
    custom_alias_path.write_text(
        json.dumps({"aliases": {"Sensor Brand/Model": {"SMC ZSE20A": "SMC ZSE20"}}}),
        encoding="utf-8",
    )

    aliases = load_part_aliases(alias_path=custom_alias_path)

    assert normalize_part_alias("Gripper Model", "Large Double Gripper", aliases) == "MHZL2-16D"
    assert normalize_part_alias("Gripper Model", "Small Double Gripper", aliases) == "MHZL2-10S"
    assert normalize_part_alias("Sensor Brand/Model", "SMC ZSE20A", aliases) == "SMC ZSE20"


def test_frequency_counts_recommendations_and_alias_cleanup(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-STD-001",
            "Press/Machine #": "Press 201",
            "Gripper Model": "Large Double Gripper",
            "Gripper Type": "Double Pressure",
            "Sensor Type": "Part-present sensor",
            "Sensor Brand/Model": "SMC",
            "Connection Type": "ATI",
        },
    )
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-STD-002",
            "Press/Machine #": "Press 202",
            "Gripper Model": "MHZL2-16D",
            "Gripper Type": "Double Pressure",
            "Sensor Type": "Part-present sensor",
            "Sensor Brand/Model": "SMC",
            "Connection Type": "ATI",
        },
    )

    analysis = analyze_standardization_opportunities(fake_project)
    gripper_model = _component_row(analysis, "gripper models", "MHZL2-16D")
    sensor_type = _component_row(analysis, "sensor types", "Part-present sensor")

    assert gripper_model["Count"] == 2
    assert "Large Double Gripper" in gripper_model["Raw Values"]
    assert sensor_type["Count"] == 2
    assert any(
        row["Recommended Part"] == "MHZL2-16D" and row["Count"] == 2 for row in analysis.recommended_standard_parts_list
    )
    assert any(
        row["Action Type"] == "Alias normalization"
        and row["Audit ID"] == "AUD-STD-001"
        and row["Recommended Value"] == "MHZL2-16D"
        for row in analysis.candidate_bom_cleanup_actions
    )


def test_unknown_model_detection(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-STD-UNKNOWN",
            "Press/Machine #": "Press 203",
            "Gripper Model": "Unknown / Not Checked",
            "Gripper Type": "N/A",
            "Sensor Brand/Model": "",
            "Connection Type": "Needs Review",
        },
    )

    analysis = analyze_standardization_opportunities(fake_project)

    assert any(
        row["Audit ID"] == "AUD-STD-UNKNOWN" and row["Field"] == "Gripper Model"
        for row in analysis.unknown_missing_part_number_table
    )
    assert any(
        row["Audit ID"] == "AUD-STD-UNKNOWN" and row["Field"] == "Sensor Brand/Model"
        for row in analysis.unknown_missing_part_number_table
    )
    assert any(
        row["Action Type"] == "Part/model lookup" and row["Audit ID"] == "AUD-STD-UNKNOWN"
        for row in analysis.candidate_bom_cleanup_actions
    )


def test_documentation_gap_cleanup_actions(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-STD-DOC",
            "Press/Machine #": "Press 204",
            "Gripper Model": "MHZL2-10S",
            "BOM Available?": "",
            "Drawing/CAD Available?": "Unknown / Not Checked",
            "Process Binder Complete?": "",
            "Spare Parts Identified?": "",
        },
    )

    analysis = analyze_standardization_opportunities(fake_project)

    assert any(
        row["Audit ID"] == "AUD-STD-DOC" and "BOM Available?" in row["Missing Fields"]
        for row in analysis.documentation_gap_table
    )
    assert any(
        row["Action Type"] == "Documentation status cleanup" and row["Audit ID"] == "AUD-STD-DOC"
        for row in analysis.candidate_bom_cleanup_actions
    )


def test_export_includes_machines_and_audits(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-STD-EXPORT-1",
            "Press/Machine #": "Press 205",
            "Gripper Model": "Small Double Gripper",
            "Gripper Type": "Single Pressure",
        },
    )
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-STD-EXPORT-2",
            "Press/Machine #": "Press 206",
            "Gripper Model": "MHZL2-10S",
            "Gripper Type": "Single Pressure",
        },
    )

    result = generate_standardization_report(fake_project)
    markdown_path = next(Path(path) for path in result.output_reports if path.endswith(".md"))
    text = markdown_path.read_text(encoding="utf-8")

    assert result.success is True
    assert "Component_Frequency_Table" in "\n".join(result.output_reports)
    assert "Recommended_Standard_Parts_List" in "\n".join(result.output_reports)
    assert "Press 205" in text
    assert "AUD-STD-EXPORT-1" in text
