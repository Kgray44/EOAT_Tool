from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from core.annotations.service import AnnotationService
from core.fmea_suggestions import (
    FMEA_CONFIDENCE_LABELS,
    accept_fmea_suggestions,
    build_fmea_suggestions,
    export_fmea_evidence_report,
    export_fmea_suggestion_draft,
    reject_fmea_suggestions,
)
from core.paths import resolve_project_paths
from core.workbook_schema import get_expected_headers


def _append_row(project_root, sheet_name: str, values: dict[str, str]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook[sheet_name]
    headers = [cell.value for cell in ws[1]]
    row = {header: "" for header in get_expected_headers(sheet_name)}
    row.update(values)
    ws.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def test_fmea_suggestion_generated_from_known_issue_and_tag(fake_project):
    _append_row(
        fake_project,
        "EOAT Inventory",
        {
            "Audit ID": "AUD-FMEA-TAG-001",
            "Press/Machine #": "Press 31",
            "EOAT Type": "Vacuum",
            "Known Issues": "Sensor intermittently fails part confirmation.",
        },
    )
    service = AnnotationService(fake_project)
    tag = service.create_tag("Sensor Failure Risk", "red")
    target = service.create_or_get_target("audit", audit_id="AUD-FMEA-TAG-001", machine_id="Press 31")
    service.assign_tag_to_target(tag.id, target.id, comment="Synthetic tag for FMEA review.", sync_workbook=False)

    suggestions = build_fmea_suggestions(fake_project)

    combined = "\n".join(str(row) for row in suggestions)
    assert "Sensor failure" in combined
    assert "Sensor Failure Risk" in combined or "Sensor intermittently fails" in combined
    assert all(row["Evidence"] for row in suggestions)
    assert all(row["Confidence"] in FMEA_CONFIDENCE_LABELS for row in suggestions)


def test_fmea_suggestion_confidence_and_rpn_from_numeric_issue_evidence(fake_project):
    _append_row(
        fake_project,
        "Issue Log",
        {
            "Issue ID": "ISS-FMEA-RPN-001",
            "Press/Machine #": "Press 39",
            "Issue Category": "Vacuum loss",
            "Issue Description": "Vacuum cups lose seal and drop parts during transfer.",
            "Severity": "8",
            "Frequency": "5",
            "Detectability": "4",
        },
    )

    suggestion = next(row for row in build_fmea_suggestions(fake_project) if row["Press/Machine #"] == "Press 39")

    assert suggestion["Confidence"] == "High"
    assert suggestion["Calculated RPN"] == 160
    assert "Vacuum cups lose seal" in suggestion["Evidence"]


def test_fmea_suggestions_require_review_before_acceptance(fake_project):
    _append_row(
        fake_project,
        "EOAT Inventory",
        {
            "Audit ID": "AUD-FMEA-REVIEW-001",
            "Press/Machine #": "Press 41",
            "EOAT Type": "Mechanical / Gripper",
            "Known Issues": "Loose gripper finger causing mechanical wear review.",
        },
    )
    suggestion = next(
        row
        for row in build_fmea_suggestions(fake_project)
        if row["Press/Machine #"] == "Press 41" and row["Failure Mode"] == "Mechanical wear"
    )

    unreviewed = accept_fmea_suggestions(fake_project, [suggestion], log_activity=False)

    assert unreviewed.success is False
    assert "reviewed numeric" in unreviewed.summary

    reviewed = dict(suggestion)
    reviewed["Suggested Severity"] = "7"
    reviewed["Suggested Frequency"] = "4"
    reviewed["Suggested Detectability"] = "5"
    reviewed["Suggested Mitigation"] = "Edited mitigation from reviewer."
    accepted = accept_fmea_suggestions(fake_project, [reviewed], log_activity=False)

    assert accepted.success is True
    rows = load_workbook(resolve_project_paths(fake_project).master_workbook, read_only=True)
    try:
        ws = rows["FMEA Draft"]
        values = [cell.value for cell in ws[ws.max_row]]
    finally:
        rows.close()
    assert "Mechanical wear" in values
    assert 140 in values
    assert "Edited mitigation from reviewer." in values


def test_fmea_suggestion_reject_and_export(fake_project):
    _append_row(
        fake_project,
        "Issue Log",
        {"Issue ID": "ISS-FMEA-001", "Press/Machine #": "Press 51", "Issue Category": "Quick disconnect issue"},
    )
    suggestion = next(row for row in build_fmea_suggestions(fake_project) if row["Press/Machine #"] == "Press 51")

    rejected = reject_fmea_suggestions(
        fake_project, [suggestion["Suggestion ID"]], reason="Synthetic test reject.", log_activity=False
    )
    exported = export_fmea_suggestion_draft(fake_project, [suggestion], log_activity=False)
    evidence_exported = export_fmea_evidence_report(fake_project, [suggestion], log_activity=False)

    assert rejected.success is True
    assert exported.success is True
    assert evidence_exported.success is True
    assert exported.output_reports
    report_text = Path(exported.output_reports[0]).read_text(encoding="utf-8")
    assert "Evidence Trace" in report_text
