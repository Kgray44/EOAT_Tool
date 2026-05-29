from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from core.paths import resolve_project_paths
from core.risk_heatmap import RISK_LEVEL_MISSING, build_risk_heatmap, export_risk_heatmap_report
from core.workbook_io import row_dicts
from core.workbook_schema import get_expected_headers


def _append_sheet_row(project_root, sheet_name: str, values: dict[str, object]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook[sheet_name]
    headers = [cell.value for cell in ws[1]]
    row = {header: "" for header in get_expected_headers(sheet_name)}
    row.update(values)
    ws.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def _audit_row(audit_id: str, machine: str, **overrides) -> dict[str, object]:
    row = {
        "Audit ID": audit_id,
        "Audit Date": "2026-05-18",
        "Auditor": "Synthetic Auditor",
        "Plant/Area": "Molding",
        "Press/Machine #": machine,
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "Priority": "Medium",
        "Status": "Audited",
        "Spare Parts Identified?": "Yes",
        "BOM Available?": "Yes",
        "Drawing/CAD Available?": "Yes",
        "Process Binder Complete?": "Yes",
    }
    row.update(overrides)
    return row


def test_risk_heatmap_uses_observed_failure_mode_evidence(fake_project):
    _append_sheet_row(
        fake_project,
        "EOAT Inventory",
        _audit_row(
            "AUD-RISK-001",
            "Press 10",
            **{
                "Priority": "High",
                "Known Issues": "Vacuum loss from tubing leak and sensor failure.",
                "Drop/Mis-Pick History": "Recurring part drops.",
                "Tubing Condition": "Leaking",
            },
        ),
    )
    _append_sheet_row(
        fake_project,
        "Issue Log",
        {
            "Issue ID": "ISS-RISK-001",
            "Press/Machine #": "Press 10",
            "Issue Category": "Alignment issue",
            "Issue Description": "EOAT alignment drift caused mis-pick.",
            "Severity": 8,
            "Status": "Open",
        },
    )

    summary = build_risk_heatmap(fake_project)
    vacuum = summary.cell("10", "vacuum loss")
    sensor = summary.cell("10", "sensor failure")
    drop = summary.cell("10", "drop/mis-pick")
    missing = summary.cell("10", "quick disconnect issue")

    assert vacuum is not None and vacuum.evidence_count >= 1 and vacuum.risk_score > 0
    assert sensor is not None and sensor.evidence_count >= 1
    assert drop is not None and drop.evidence_count >= 1
    assert missing is not None
    assert missing.risk_level == RISK_LEVEL_MISSING
    assert missing.risk_score == 0
    assert "No local evidence" in missing.missing_evidence[0]


def test_risk_heatmap_detects_pneumatic_mismatch_and_documentation_gap(fake_project):
    _append_sheet_row(
        fake_project,
        "EOAT Inventory",
        _audit_row(
            "AUD-RISK-002",
            "Press 11",
            **{
                "Known Issues": "Pneumatic circuit mismatch between robot-side and EOAT-side routing.",
                "BOM Available?": "No",
                "Drawing/CAD Available?": "",
                "Process Binder Complete?": "Unknown / Not Checked",
            },
        ),
    )

    summary = build_risk_heatmap(fake_project)
    pneumatic = summary.cell("11", "pneumatic circuit mismatch")
    documentation = summary.cell("11", "documentation gap")

    assert pneumatic is not None
    assert pneumatic.evidence_count == 1
    assert "Pneumatic circuit mismatch" in pneumatic.evidence[0].description
    assert documentation is not None
    assert documentation.evidence_count == 1
    assert "BOM Available?" in documentation.evidence[0].description


def test_risk_heatmap_uses_fmea_and_kpi_evidence(fake_project):
    _append_sheet_row(fake_project, "EOAT Inventory", _audit_row("AUD-RISK-003", "Press 12"))
    _append_sheet_row(
        fake_project,
        "FMEA Draft",
        {
            "FMEA ID": "FMEA-RISK-001",
            "Press/Machine #": "Press 12",
            "Failure Mode": "Cylinder issue causes gripper wear and misalignment.",
            "Severity": 8,
            "Frequency": 6,
            "Detectability": 4,
        },
    )
    _append_sheet_row(
        fake_project,
        "KPI Baseline",
        {
            "KPI ID": "KPI-RISK-001",
            "Press/Machine #": "Press 12",
            "Part Drops": 3,
            "Mis-Picks": 2,
            "Scrap Quantity": 5,
            "Scrap Reason": "Part drops",
        },
    )

    summary = build_risk_heatmap(fake_project)

    assert summary.cell("12", "cylinder issue").evidence_count == 1
    assert summary.cell("12", "gripper wear").evidence_count == 1
    assert summary.cell("12", "drop/mis-pick").evidence_count >= 1


def test_risk_heatmap_export_does_not_modify_workbook(fake_project):
    _append_sheet_row(fake_project, "EOAT Inventory", _audit_row("AUD-RISK-004", "Press 13", Known_Issues=""))
    workbook_path = resolve_project_paths(fake_project).master_workbook
    before_count = len(row_dicts(workbook_path, "EOAT Inventory"))

    result = export_risk_heatmap_report(fake_project, log_activity=False)

    after_count = len(row_dicts(workbook_path, "EOAT Inventory"))
    markdown = next(Path(path) for path in result.output_reports if path.endswith(".md"))
    text = markdown.read_text(encoding="utf-8")

    assert result.success is True
    assert before_count == after_count
    assert "Risk Heat Map" in text
    assert "Missing Evidence" in text
