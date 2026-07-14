from __future__ import annotations

import hashlib

from openpyxl import load_workbook

from core.paths import resolve_project_paths
from core.validation import validate_project_foundation
from core.validation_findings import findings_from_result, validation_json_payload
from core.workbook_repairs import FIX_CLEAR_STALE_HIDDEN_NA, FIX_NORMALIZE_DROPDOWN_CASING, preview_safe_fix


def _append_inventory_row(project_root, values):
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    ws.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()
    return workbook_path


def _base_row(audit_id: str, machine: str = "Press 980") -> dict[str, object]:
    return {
        "Audit ID": audit_id,
        "Audit Date": "2026-05-18",
        "Auditor": "KG",
        "Plant/Area": "Plant 4",
        "Press/Machine #": machine,
        "Tool #": f"TOOL-{audit_id}",
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "EOAT Moves": "Part",
        "Connection Type": "ATI",
        "# of Cylinders": "N/A",
        "Cylinder Type": "Linear",
        "Sensors Present?": "No",
        "Electrical/Wiring Present?": "No",
        "Quick Disconnects Present?": "No",
        "Photos Taken?": "No",
        "Status": "In Progress",
        "Priority": "Medium",
        "Entry Type": "Audited",
    }


def _hash_file(path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_truth_engine_flags_stale_hidden_values_and_preview_is_read_only(fake_project):
    workbook_path = _append_inventory_row(
        fake_project,
        {
            **_base_row("AUD-TRUTH-STALE", "Press 981"),
            "EOAT Type": "Mechanical / Gripper",
            "# of Cups": "4",
            "Cup Type/Material": "stale cup material",
        },
    )

    result = validate_project_foundation(fake_project)
    findings = findings_from_result(result)

    assert result.metrics["stale_hidden_value_count"] >= 2
    assert any(
        finding.audit_id == "AUD-TRUTH-STALE" and finding.fix_id == FIX_CLEAR_STALE_HIDDEN_NA for finding in findings
    )

    before_hash = _hash_file(workbook_path)
    preview = preview_safe_fix(fake_project, FIX_CLEAR_STALE_HIDDEN_NA)
    after_hash = _hash_file(workbook_path)

    assert preview.can_apply
    assert any(change.audit_id == "AUD-TRUTH-STALE" and change.column_name == "# of Cups" for change in preview.changes)
    assert before_hash == after_hash


def test_truth_engine_flags_compatibility_rows_without_source_metadata(fake_project):
    _append_inventory_row(
        fake_project,
        {
            **_base_row("AUD-COMPAT-MISSING-SOURCE", "Press 982"),
            "Entry Type": "Compatible",
            "Source Audit ID": "",
            "Compatibility Source": "",
        },
    )

    result = validate_project_foundation(fake_project)
    findings = findings_from_result(result)
    messages = "\n".join(finding.message for finding in findings)

    assert result.metrics["compatible_row_count"] >= 1
    assert result.metrics["compatibility_missing_source_audit_id_count"] == 1
    assert result.metrics["compatibility_missing_source_count"] == 1
    assert "Fit Check row is missing Source Audit ID" in messages
    assert "Fit Check row is missing Compatibility Source" in messages


def test_truth_engine_flags_duplicate_audit_ids(fake_project):
    duplicate = _base_row("AUD-DUPLICATE-TRUTH", "Press 983")
    _append_inventory_row(fake_project, duplicate)
    _append_inventory_row(fake_project, {**duplicate, "Tool #": "TOOL-DUPLICATE-OTHER"})

    result = validate_project_foundation(fake_project)
    findings = findings_from_result(result)

    assert result.metrics["duplicate_audit_id_count"] == 1
    assert any("Duplicate Audit ID value: AUD-DUPLICATE-TRUTH" in finding.message for finding in findings)


def test_truth_engine_flags_broken_photo_links_and_status_conflicts(fake_project):
    _append_inventory_row(
        fake_project,
        {
            **_base_row("AUD-PHOTO-BROKEN", "Press 984"),
            "Photos Taken?": "Yes",
            "Photo Folder/Link": "01_EOAT_Audit/Cell_Photos/Missing_Photo_Folder",
        },
    )
    _append_inventory_row(
        fake_project,
        {
            **_base_row("AUD-PHOTO-CONFLICT", "Press 985"),
            "Photos Taken?": "No",
            "Photo Folder/Link": "01_EOAT_Audit/Cell_Photos/Missing_While_No",
        },
    )

    result = validate_project_foundation(fake_project)
    findings = findings_from_result(result)
    messages = "\n".join(finding.message for finding in findings)

    assert result.metrics["broken_photo_link_count"] >= 2
    assert result.metrics["photos_yes_without_link_count"] == 0
    assert result.metrics["photo_link_while_no_count"] >= 1
    assert "broken local photo link" in messages
    assert "Photo Folder/Link is populated while Photos Taken? is marked No" in messages


def test_truth_engine_flags_semantic_eoat_conflicts(fake_project):
    _append_inventory_row(
        fake_project,
        {
            **_base_row("AUD-SEMANTIC-CONFLICT", "Press 986"),
            "EOAT Type": "Mechanical / Gripper",
            "# of Cups": "6",
            "Sensor Type": "Photoeye",
            "Pneumatic Quick Disconnect Type": "PTC",
            "Electrical Quick Disconnect Type": "M12",
        },
    )

    result = validate_project_foundation(fake_project)
    messages = "\n".join(finding.message for finding in findings_from_result(result))

    assert result.metrics["semantic_warning_count"] >= 1
    assert "Mechanical / Gripper EOAT has meaningful vacuum-side field values" in messages
    assert "Sensors Present? is No but sensor detail fields contain meaningful values" in messages
    assert "Quick Disconnects Present? is No but quick disconnect detail fields contain meaningful values" in messages


def test_truth_engine_json_includes_repair_suggestions_and_preview_does_not_modify_files(fake_project):
    workbook_path = _append_inventory_row(
        fake_project,
        {
            **_base_row("AUD-DROPDOWN-CASING", "Press 987"),
            "EOAT Type": "vacuum",
        },
    )

    result = validate_project_foundation(fake_project)
    payload = validation_json_payload(fake_project, result)
    before_hash = _hash_file(workbook_path)
    preview = preview_safe_fix(fake_project, FIX_NORMALIZE_DROPDOWN_CASING)
    after_hash = _hash_file(workbook_path)

    assert result.metrics["dropdown_casing_fixable_count"] >= 1
    assert any(item["fix_id"] == FIX_NORMALIZE_DROPDOWN_CASING for item in payload["repair_suggestions"])
    assert preview.can_apply
    assert any(change.audit_id == "AUD-DROPDOWN-CASING" and change.new_value == "Vacuum" for change in preview.changes)
    assert before_hash == after_hash
