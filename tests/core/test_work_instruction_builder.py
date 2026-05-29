from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from core.paths import resolve_project_paths
from core.work_instruction_builder import (
    INSTRUCTION_TYPES,
    build_work_instruction_documents,
    generate_work_instructions,
)
from core.workbook_schema import get_expected_headers


def _append_inventory_row(project_root: Path, values: dict[str, object]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    row = {header: "" for header in get_expected_headers("EOAT Inventory")}
    row.update(values)
    ws.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def test_work_instruction_documents_use_actual_audit_data_and_mark_missing_docs(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-WI-001",
            "Press/Machine #": "Press 77",
            "Robot Type": "Wittmann R9",
            "Tool #": "TOOL-WI",
            "EOAT Type": "Vacuum",
            "# of Cups": 6,
            "Cup Type/Material": "Nitrile bellows cup",
            "Cup Diameter/Size": "20 mm",
            "Vacuum Generator Type": "Venturi",
            "Sensors Present?": "Yes",
            "Sensor Type": "Vacuum switch",
            "Sensor Brand/Model": "SMC ZSE20",
            "Known Issues": "Part drops during startup.",
            "Photos Taken?": "No",
            "Drawing/CAD Available?": "No",
            "BOM Available?": "",
            "Process Binder Complete?": "No",
        },
    )

    instruction_set = build_work_instruction_documents(fake_project, audit_id="AUD-WI-001")
    combined = "\n".join(document.markdown for document in instruction_set.documents)

    assert len(instruction_set.documents) == len(INSTRUCTION_TYPES)
    assert "Press 77" in combined
    assert "Nitrile bellows cup" in combined
    assert "CAD/Drawing: Missing per audit data." in combined
    assert "BOM: Not documented." in combined
    assert "photos not documented as available" in "\n".join(instruction_set.warnings)
    assert "must be reviewed before controlled release" in combined


def test_generate_work_instructions_writes_project_root_artifacts(fake_project):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-WI-002",
            "Press/Machine #": "Press 78",
            "EOAT Type": "Mechanical / Gripper",
            "# of Grippers": 2,
            "Gripper Model": "MHZL2-10S",
            "Photos Taken?": "Yes",
            "Photo Folder/Link": "01_EOAT_Audit/Cell_Photos/Press_78",
            "Drawing/CAD Available?": "Yes",
            "BOM Available?": "Yes",
            "Process Binder Complete?": "Yes",
        },
    )

    result = generate_work_instructions(fake_project, audit_id="AUD-WI-002", log_activity=False)

    assert result.success is True
    assert result.metrics["document_count"] == len(INSTRUCTION_TYPES)
    assert len(result.output_reports) == len(INSTRUCTION_TYPES) + 1
    assert all(Path(path).is_relative_to(fake_project) for path in result.output_reports)
    assert any("Work_Instruction_Index" in Path(path).name for path in result.output_reports)


def test_work_instruction_builder_refuses_unknown_audit(fake_project):
    result = generate_work_instructions(fake_project, audit_id="AUD-DOES-NOT-EXIST", log_activity=False)

    assert result.success is False
    assert "No work instructions" in result.summary
