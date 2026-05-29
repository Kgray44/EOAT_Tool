from __future__ import annotations

import pytest
from openpyxl import load_workbook
from PySide6.QtWidgets import QComboBox, QGroupBox, QLabel, QLineEdit, QTextEdit

from app.pages.audit import AuditPage, ExistingAuditSelection, ExistingMachineAuditsDialog
from core.audit_compatibility import find_existing_audits_for_machine
from core.audit_constants import (
    COMPATIBILITY_SOURCE_FIELD,
    CYLINDER_COUNT_FIELD,
    CYLINDER_TYPE_DEFAULT,
    CYLINDER_TYPE_FIELD,
    ENTRY_TYPE_AUDITED,
    ENTRY_TYPE_COMPATIBLE,
    ENTRY_TYPE_FIELD,
    MANUAL_COMPLETION_OVERRIDE_FIELDS,
    SOURCE_AUDIT_ID_FIELD,
)
from core.paths import resolve_project_paths
from core.robot_info import upsert_robot_info_from_audit
from core.tool_fields import LEGACY_TOOL_FIELD
from core.workbook_io import row_dicts
from core.workbook_schema import get_expected_headers
from tests.fixtures.reference_workbooks import create_press_reference_workbooks
from tests.ui.helpers import click_button, wait_for_background_tasks, wait_until

pytestmark = pytest.mark.usability


REFERENCE_ONLY_FIELDS = {
    "U.S. Tons",
    "Press Brand",
    "Press Model",
    "Press Tonnage",
    "Press Year",
    "Injection Pressure",
    "Injection Capacity",
    "Screw Diameter",
    "Controller Type",
    "Robot Serial Number",
    "Robot Manufacturing Date",
    "Full Servo",
    "TCU Count",
    "EDART Unit Press Side",
    "Forecasted Capacity",
    "Available Capacity",
    "Hours Allocated per Month",
    "Hours per Week",
    "Committed Hours per Year",
    "Cycle Time (S)",
    "Cavitation",
}


def _combo_items(combo: QComboBox) -> list[str]:
    return [combo.itemText(index) for index in range(combo.count())]


def _set_field(page: AuditPage, field: str, value: str) -> None:
    widget = page.audit_fields[field]
    if isinstance(widget, QComboBox):
        if widget.findText(value) >= 0:
            widget.setCurrentText(value)
        else:
            widget.setEditText(value)
    elif isinstance(widget, QTextEdit):
        widget.setPlainText(value)
    else:
        widget.setText(value)


def _finish_machine_lookup(page: AuditPage) -> None:
    page.audit_fields["Press/Machine #"].editingFinished.emit()
    wait_until(
        lambda: page._machine_lookup_timer is None or not page._machine_lookup_timer.isActive(),
        timeout_ms=5000,
        message="machine lookup debounce timer",
    )
    wait_for_background_tasks()


def _append_inventory_row(project_root, values: dict[str, str]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    ws.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def _machine_audit_row(audit_id: str, machine: str, entry_type: str = ENTRY_TYPE_AUDITED, **overrides) -> dict[str, str]:
    row = {
        "Audit ID": audit_id,
        "Audit Date": "2026-05-18",
        "Auditor": "Synthetic Auditor",
        "Plant/Area": "Plant 4",
        "Press/Machine #": machine,
        "Tool #": f"DEMO-PN-{machine}",
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "Status": "Complete",
        ENTRY_TYPE_FIELD: entry_type,
    }
    if entry_type == ENTRY_TYPE_COMPATIBLE:
        row.update(
            {
                "Audit Date": "",
                "Auditor": "",
                SOURCE_AUDIT_ID_FIELD: "AUD-SOURCE-PHYSICAL",
                COMPATIBILITY_SOURCE_FIELD: "Synthetic compatibility test",
            }
        )
    row.update(overrides)
    return row


def test_auditor_and_plant_defaults(qapp, fake_config):
    page = AuditPage(fake_config)

    assert page.audit_fields["Auditor"].text() == "Kato Gray"
    assert page.audit_fields["Known Issues"].toPlainText() == "Unknown / Not Checked"
    assert page.audit_fields["Drop/Mis-Pick History"].toPlainText() == "Unknown / Not Checked"
    assert page.audit_fields["Maintenance Frequency"].text() == "Unknown / Not Checked"
    assert page.audit_fields["Cleanroom/Non-Cleanroom"].currentText() == "Whiteroom"
    assert page.audit_fields["Cup Type/Material"].text() == "Silicone"
    plant = page.audit_fields["Plant/Area"]
    assert isinstance(plant, QComboBox)
    assert plant.isEditable() is False
    assert _combo_items(plant) == ["Plant 4", "Cleanroom"]
    assert plant.currentText() == "Plant 4"
    assert page.audit_fields["Vacuum Confirmation Present?"].currentText() == "Yes"
    assert page.audit_fields["Part-Present Detection Present?"].currentText() == "No"
    assert page.audit_fields["Spare Parts Identified?"].currentText() == "No"
    assert page.audit_fields["Drawing/CAD Available?"].currentText() == "No"
    assert page.audit_fields["BOM Available?"].currentText() == "No"
    assert page.audit_fields["Process Binder Complete?"].currentText() == "No"
    assert page.audit_fields["Photos Taken?"].currentText() == "No"


def test_audit_page_does_not_show_reference_spreadsheet_fields(qapp, fake_config):
    page = AuditPage(fake_config)

    assert REFERENCE_ONLY_FIELDS.isdisjoint(page.audit_fields)
    workflow_metadata = {"Entry Type", "Source Audit ID", "Compatibility Source", *MANUAL_COMPLETION_OVERRIDE_FIELDS}
    assert (set(get_expected_headers("EOAT Inventory")) - workflow_metadata).issubset(page.audit_fields)
    assert {
        "Tool #",
        "Connection Type",
        "# of Cups",
        "# of Grippers",
        "Gripper Type",
        "Gripper Model",
        "Gripper Size",
        "Number of Parts Picked",
        CYLINDER_COUNT_FIELD,
        CYLINDER_TYPE_FIELD,
        "Tubing Condition",
        "Cable Management Condition",
        "BOM Available?",
        "Photos Taken?",
    }.issubset(page.audit_fields)
    assert LEGACY_TOOL_FIELD not in page.audit_fields
    label_texts = {label.text() for label in page.findChildren(QLabel)}
    assert "Tool #" in label_texts
    assert "Connection Type" in label_texts
    assert LEGACY_TOOL_FIELD not in label_texts


def test_connection_type_and_eoat_type_dropdown_options(qapp, fake_config):
    page = AuditPage(fake_config)

    eoat_moves = page.audit_fields["EOAT Moves"]
    assert isinstance(eoat_moves, QComboBox)
    assert eoat_moves.isEditable() is False
    assert _combo_items(eoat_moves) == ["", "Part", "Sprue", "Both"]
    assert eoat_moves.currentText() == ""

    connection = page.audit_fields["Connection Type"]
    assert isinstance(connection, QComboBox)
    assert connection.isEditable() is False
    assert _combo_items(connection) == ["ATI", "DoveTail", "Direct Mount", "Lever Lock"]
    assert connection.currentText() == ""

    gripper_type = page.audit_fields["Gripper Type"]
    assert isinstance(gripper_type, QComboBox)
    assert gripper_type.isEditable() is False
    assert _combo_items(gripper_type) == ["", "Single Pressure", "Double Pressure"]

    gripper_model = page.audit_fields["Gripper Model"]
    assert isinstance(gripper_model, QComboBox)
    assert gripper_model.isEditable() is True
    assert "Large Double Gripper" in _combo_items(gripper_model)
    assert "Small Double Gripper" in _combo_items(gripper_model)

    cylinder_type = page.audit_fields[CYLINDER_TYPE_FIELD]
    assert isinstance(cylinder_type, QComboBox)
    assert cylinder_type.isEditable() is False
    assert "Linear" in _combo_items(cylinder_type)
    assert "Rotary" in _combo_items(cylinder_type)
    assert cylinder_type.currentText() == ""

    eoat_type = page.audit_fields["EOAT Type"]
    assert isinstance(eoat_type, QComboBox)
    assert "Miscellaneous" in _combo_items(eoat_type)

    cleanroom = page.audit_fields["Cleanroom/Non-Cleanroom"]
    assert isinstance(cleanroom, QComboBox)
    assert "Whiteroom" in _combo_items(cleanroom)

    tooling_fields = list(page.audit_fields)
    assert tooling_fields.index("EOAT Type") < tooling_fields.index("EOAT Moves") < tooling_fields.index("Connection Type")
    assert tooling_fields.index("Number of Parts Picked") < tooling_fields.index("# of Grippers") < tooling_fields.index("Gripper Type") < tooling_fields.index("Gripper Model")
    assert tooling_fields.index("Number of Parts Picked") < tooling_fields.index(CYLINDER_COUNT_FIELD) < tooling_fields.index(CYLINDER_TYPE_FIELD)
    assert tooling_fields.index("# of Cups") < tooling_fields.index("Cup Type/Material") < tooling_fields.index("Cup Diameter/Size") < tooling_fields.index("Vacuum Generator Type")
    assert "Vacuum Zones" not in page.audit_fields
    assert tooling_fields.index("Connection Type") < tooling_fields.index("Number of Parts Picked") < tooling_fields.index("# of Cups")
    cups = page.audit_fields["# of Cups"]
    assert isinstance(cups, QLineEdit)
    cups.setText("3")
    assert cups.hasAcceptableInput()
    cups.setText("-1")
    assert not cups.hasAcceptableInput()
    cylinders = page.audit_fields[CYLINDER_COUNT_FIELD]
    assert isinstance(cylinders, QLineEdit)
    cylinders.setText("2")
    assert cylinders.hasAcceptableInput()
    cylinders.setText("-1")
    assert not cylinders.hasAcceptableInput()


@pytest.mark.parametrize(
    ("eoat_type", "vacuum_visible", "gripper_visible"),
    [
        ("Vacuum", True, False),
        ("Mechanical / Gripper", False, True),
        ("Hybrid", True, True),
        ("Unknown / Needs Review", True, False),
        ("Miscellaneous", True, True),
        ("", True, False),
    ],
)
def test_eoat_type_controls_tooling_visibility(qapp, fake_config, eoat_type, vacuum_visible, gripper_visible):
    page = AuditPage(fake_config)

    _set_field(page, "EOAT Type", eoat_type)

    assert page.audit_fields["Cup Type/Material"].isHidden() is (not vacuum_visible)
    assert page.audit_fields["Cup Diameter/Size"].isHidden() is (not vacuum_visible)
    assert page.audit_fields["# of Cups"].isHidden() is (not vacuum_visible)
    assert page.audit_fields["Number of Parts Picked"].isHidden() is False
    assert page.audit_fields["# of Grippers"].isHidden() is (not gripper_visible)
    assert page.audit_fields["Gripper Model"].isHidden() is (not gripper_visible)
    assert page.audit_fields["Gripper Size"].isHidden() is (not gripper_visible)
    assert page.audit_fields["Gripper Type"].isHidden() is (not gripper_visible)
    assert page.audit_fields[CYLINDER_COUNT_FIELD].isHidden() is False
    assert page.audit_fields[CYLINDER_TYPE_FIELD].isHidden() is False


def test_cylinder_details_group_is_always_visible(qapp, fake_config):
    page = AuditPage(fake_config)
    group_titles = {group.title(): group for group in page.findChildren(QGroupBox)}

    assert "Cylinder Details" in group_titles
    cylinder_group = group_titles["Cylinder Details"]
    for eoat_type in ["Vacuum", "Mechanical / Gripper", "Hybrid", "Miscellaneous", "Unknown / Needs Review", ""]:
        _set_field(page, "EOAT Type", eoat_type)

        assert cylinder_group.isHidden() is False
        assert page.audit_fields[CYLINDER_COUNT_FIELD].isHidden() is False
        assert page.audit_fields[CYLINDER_TYPE_FIELD].isHidden() is False


def test_cylinder_type_defaults_only_when_cylinder_count_is_used(qapp, fake_config):
    page = AuditPage(fake_config)
    count_widget = page.audit_fields[CYLINDER_COUNT_FIELD]
    type_widget = page.audit_fields[CYLINDER_TYPE_FIELD]

    assert type_widget.currentText() == ""

    count_widget.setText("2")

    assert type_widget.currentText() == CYLINDER_TYPE_DEFAULT

    count_widget.setText("")

    assert type_widget.currentText() == ""


def test_manual_cylinder_type_is_preserved_when_count_is_cleared(qapp, fake_config):
    page = AuditPage(fake_config)
    count_widget = page.audit_fields[CYLINDER_COUNT_FIELD]
    type_widget = page.audit_fields[CYLINDER_TYPE_FIELD]

    type_widget.setCurrentText("Rotary")
    count_widget.setText("2")
    count_widget.setText("")

    assert type_widget.currentText() == "Rotary"


def test_eoat_type_visibility_updates_immediately_and_preserves_hidden_values(qapp, fake_config):
    page = AuditPage(fake_config)

    _set_field(page, "EOAT Type", "Hybrid")
    _set_field(page, "Cup Type/Material", "Nitrile")
    _set_field(page, "Cup Diameter/Size", "20 mm")
    _set_field(page, "# of Cups", "4")
    _set_field(page, "Gripper Model", "Zimmer GPP")
    _set_field(page, "Gripper Size", "25 mm")
    _set_field(page, "# of Grippers", "2")
    _set_field(page, "Gripper Type", "Single Pressure")
    _set_field(page, "EOAT Type", "Mechanical / Gripper")

    assert page.audit_fields["Cup Type/Material"].isHidden() is True
    assert page.audit_fields["Cup Diameter/Size"].isHidden() is True
    assert page.audit_fields["# of Cups"].isHidden() is True
    assert page.audit_fields["Gripper Model"].isHidden() is False
    assert page.audit_fields["Gripper Size"].isHidden() is False
    assert page.audit_fields["# of Grippers"].isHidden() is False
    assert page.audit_fields["Gripper Type"].isHidden() is False
    assert page.audit_fields["Cup Type/Material"].text() == "Nitrile"
    assert page.audit_fields["Cup Diameter/Size"].text() == "20 mm"
    assert page.audit_fields["# of Cups"].text() == "4"

    _set_field(page, "EOAT Type", "Vacuum")
    assert page.audit_fields["Gripper Model"].isHidden() is True
    assert page.audit_fields["Gripper Size"].isHidden() is True
    assert page.audit_fields["# of Grippers"].isHidden() is True
    assert page.audit_fields["Gripper Type"].isHidden() is True
    assert page._field_value(page.audit_fields["Gripper Model"]) == "Zimmer GPP"
    assert page.audit_fields["Gripper Size"].text() == "25 mm"
    assert page.audit_fields["# of Grippers"].text() == "2"
    assert page._field_value(page.audit_fields["Gripper Type"]) == "Single Pressure"

    _set_field(page, "EOAT Type", "Miscellaneous")
    assert page.audit_fields["Gripper Model"].isHidden() is False
    assert page.audit_fields["Gripper Size"].isHidden() is False
    assert page.audit_fields["# of Grippers"].isHidden() is False
    assert page.audit_fields["Gripper Type"].isHidden() is False
    assert page._field_value(page.audit_fields["Gripper Model"]) == "Zimmer GPP"
    assert page.audit_fields["# of Grippers"].text() == "2"


def test_sensors_present_controls_sensor_electrical_visibility(qapp, fake_config):
    page = AuditPage(fake_config)

    _set_field(page, "Sensor Type", "Vacuum switch")
    _set_field(page, "Sensor Brand/Model", "SMC ZSE20")
    _set_field(page, "Electrical Quick Disconnect Type", "M12")
    _set_field(page, "Cable Management Condition", "OK")
    _set_field(page, "Sensors Present?", "No")

    for field in [
        "Sensor Type",
        "Sensor Brand/Model",
        "Vacuum Confirmation Present?",
        "Part-Present Detection Present?",
    ]:
        assert page.audit_fields[field].isHidden() is True
    assert page.audit_fields["Electrical Quick Disconnect Type"].isHidden() is False
    assert page.audit_fields["Cable Management Condition"].isHidden() is False

    assert page.audit_fields["Known Issues"].isHidden() is False
    _set_field(page, "Known Issues", "Unrelated value survives.")

    _set_field(page, "Sensors Present?", "Yes")

    for field in [
        "Sensor Type",
        "Sensor Brand/Model",
        "Vacuum Confirmation Present?",
        "Part-Present Detection Present?",
    ]:
        assert page.audit_fields[field].isHidden() is False
    assert page.audit_fields["Sensor Type"].text() == "Vacuum switch"
    assert page.audit_fields["Sensor Brand/Model"].text() == "SMC ZSE20"
    assert page.audit_fields["Electrical Quick Disconnect Type"].text() == "M12"
    assert page.audit_fields["Cable Management Condition"].currentText() == "OK"
    assert page.audit_fields["Known Issues"].toPlainText() == "Unrelated value survives."


def test_ui_lookup_runs_on_editing_finished_and_fills_clean_fields(qapp, fake_config, fake_project):
    create_press_reference_workbooks(fake_project / "reference-data")
    page = AuditPage(fake_config)
    page.show()

    page.audit_fields["Press/Machine #"].setText("P12")
    _finish_machine_lookup(page)

    assert page.audit_fields["Press/Machine #"].text() == "12"
    assert page.audit_fields["Robot Type"].currentText() == "Wittmann W833"
    assert page.audit_fields["Robot Model/Controller"].text() == "W833"
    assert page.audit_fields["Tool #"].text() == "DEMO-PN-1200"
    assert page.audit_fields["Part Family"].text() == "DEMO-PN-1200 - Demo housing cap"
    assert page.audit_fields["Part Name/Description"].toPlainText() == "Demo housing cap"
    assert "Robot and part info filled" in page.lookup_note_label.text()


def test_manual_lookup_button_uses_same_lookup_path(qapp, fake_config, fake_project):
    create_press_reference_workbooks(fake_project / "reference-data")
    page = AuditPage(fake_config)

    page.audit_fields["Press/Machine #"].setText("Press #12")
    click_button(page, "Lookup")

    assert page.audit_fields["Press/Machine #"].text() == "12"
    assert page.audit_fields["Robot Type"].currentText() == "Wittmann W833"


def test_multiple_capacity_rows_show_selector_without_autofilling_part(qapp, fake_config, fake_project):
    create_press_reference_workbooks(fake_project / "reference-data", multiple_capacity_rows=True)
    assert upsert_robot_info_from_audit(
        fake_project,
        {
            "Audit ID": "AUD-ROBOT-LOOKUP-012",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "12",
            "Robot Type": "Wittmann W833",
            "Robot Notes": "Lookup should load this without blocking part selection.",
        },
    ).success
    page = AuditPage(fake_config)

    page.audit_fields["Press/Machine #"].setText("Machine 12")
    _finish_machine_lookup(page)

    assert page.capacity_part_combo.isEnabled()
    assert page.capacity_part_combo.count() == 3
    assert page.capacity_matches_table.rowCount() == 2
    assert "possible parts" in page.lookup_note_label.text()
    assert "Multiple possible Tool # values found for this press" in page.lookup_note_label.text()
    assert page.audit_fields["Tool #"].text() == ""
    assert page.audit_fields["Part Family"].text() == ""
    assert page.audit_fields["Robot Notes"].toPlainText() == "Lookup should load this without blocking part selection."

    page.capacity_part_combo.setCurrentIndex(2)
    assert page.audit_fields["Tool #"].text() == "DEMO-PN-1201"
    assert page.audit_fields["Part Family"].text() == "DEMO-PN-1201 - Demo housing base"
    assert page.audit_fields["Part Name/Description"].toPlainText() == "Demo housing base"


def test_missing_machine_number_lookup_does_not_crash(qapp, fake_config):
    page = AuditPage(fake_config)

    page.audit_fields["Press/Machine #"].setText("")
    _finish_machine_lookup(page)

    assert "Invalid machine number" in page.lookup_note_label.text()


def test_machine_lookup_ignores_compatible_only_existing_rows(qapp, fake_config, fake_project):
    _append_inventory_row(fake_project, _machine_audit_row("AUD-COMPAT-044", "44", ENTRY_TYPE_COMPATIBLE))
    page = AuditPage(fake_config)
    current_audit_id = page.audit_fields["Audit ID"].text()

    loaded = page._load_or_offer_existing_audit_for_machine("44")

    assert loaded is False
    assert page._current_audit_mode == "new"
    assert page.audit_fields["Audit ID"].text() == current_audit_id
    assert page.machine_audit_match_combo.isHidden()
    assert "Machine 44 has compatible coverage entries, but no physical audit yet" in page.lookup_note_label.text()


def test_machine_lookup_single_physical_audit_opens_selection_dialog(qapp, fake_config, fake_project, monkeypatch):
    _append_inventory_row(fake_project, _machine_audit_row("AUD-PHYSICAL-045", "45", ENTRY_TYPE_AUDITED))
    page = AuditPage(fake_config)
    dialog_calls = []

    def choose(machine_number: str, matches):
        dialog_calls.append((machine_number, [match.audit_id for match in matches]))
        return ExistingAuditSelection(ExistingMachineAuditsDialog.ACTION_CANCEL)

    monkeypatch.setattr(page, "_choose_existing_machine_audit_action", choose)

    loaded = page._load_or_offer_existing_audit_for_machine("45")

    assert loaded is True
    assert dialog_calls == [("45", ["AUD-PHYSICAL-045"])]
    assert page._current_loaded_audit_id is None
    assert page._current_audit_mode == "new"
    assert "Existing audit selection canceled" in page.lookup_note_label.text()


def test_machine_lookup_clean_new_form_loads_without_unsaved_prompt(qapp, fake_config, fake_project, monkeypatch):
    _append_inventory_row(fake_project, _machine_audit_row("AUD-PHYSICAL-049", "49", ENTRY_TYPE_AUDITED))
    page = AuditPage(fake_config)
    prompt_actions = []

    def reject_unsaved_prompt(action: str) -> bool:
        prompt_actions.append(action)
        return False

    monkeypatch.setattr(page, "_confirm_unsaved_audit_changes", reject_unsaved_prompt)
    monkeypatch.setattr(
        page,
        "_choose_existing_machine_audit_action",
        lambda _machine, _matches: ExistingAuditSelection(
            ExistingMachineAuditsDialog.ACTION_CONTINUE,
            "AUD-PHYSICAL-049",
        ),
    )

    page.audit_fields["Press/Machine #"].setText("49")
    _finish_machine_lookup(page)

    assert prompt_actions == []
    assert page._current_loaded_audit_id == "AUD-PHYSICAL-049"
    assert page.audit_fields["Audit ID"].text() == "AUD-PHYSICAL-049"
    assert "Existing physical audit found for Machine 49" in page.result_panel.viewer.toPlainText()
    assert "Machine number is required" not in page.result_panel.viewer.toPlainText()
    assert page.has_unsaved_changes() is False


def test_machine_lookup_dirty_new_form_still_prompts_before_existing_load(qapp, fake_config, fake_project, monkeypatch):
    _append_inventory_row(fake_project, _machine_audit_row("AUD-PHYSICAL-051", "51", ENTRY_TYPE_AUDITED))
    page = AuditPage(fake_config)
    generated_id = page.audit_fields["Audit ID"].text()
    prompt_actions = []

    def reject_unsaved_prompt(action: str) -> bool:
        prompt_actions.append(action)
        return False

    monkeypatch.setattr(page, "_confirm_unsaved_audit_changes", reject_unsaved_prompt)
    monkeypatch.setattr(
        page,
        "_choose_existing_machine_audit_action",
        lambda _machine, _matches: ExistingAuditSelection(
            ExistingMachineAuditsDialog.ACTION_CONTINUE,
            "AUD-PHYSICAL-051",
        ),
    )

    page.audit_fields["Known Issues"].setPlainText("Meaningful unsaved issue.")
    page.audit_fields["Press/Machine #"].setText("51")
    _finish_machine_lookup(page)

    assert prompt_actions == ["load another audit"]
    assert page._current_loaded_audit_id is None
    assert page.audit_fields["Audit ID"].text() == generated_id
    assert page.audit_fields["Known Issues"].toPlainText() == "Meaningful unsaved issue."


def test_machine_lookup_uses_physical_row_when_compatible_also_exists(qapp, fake_config, fake_project, monkeypatch):
    _append_inventory_row(fake_project, _machine_audit_row("AUD-COMPAT-046", "46", ENTRY_TYPE_COMPATIBLE))
    _append_inventory_row(fake_project, _machine_audit_row("AUD-PHYSICAL-046", "46", ENTRY_TYPE_AUDITED))
    page = AuditPage(fake_config)
    monkeypatch.setattr(
        page,
        "_choose_existing_machine_audit_action",
        lambda _machine, _matches: ExistingAuditSelection(
            ExistingMachineAuditsDialog.ACTION_CONTINUE,
            "AUD-PHYSICAL-046",
        ),
    )

    loaded = page._load_or_offer_existing_audit_for_machine("46")

    assert loaded is True
    assert page._current_loaded_audit_id == "AUD-PHYSICAL-046"
    assert page.audit_fields["Audit ID"].text() == "AUD-PHYSICAL-046"
    assert "AUD-COMPAT-046" not in page.load_audit_id_combo.currentText()


def test_machine_lookup_treats_blank_entry_type_as_physical_audit(qapp, fake_config, fake_project, monkeypatch):
    _append_inventory_row(fake_project, _machine_audit_row("AUD-BLANK-TYPE-047", "47", ""))
    page = AuditPage(fake_config)
    monkeypatch.setattr(
        page,
        "_choose_existing_machine_audit_action",
        lambda _machine, _matches: ExistingAuditSelection(
            ExistingMachineAuditsDialog.ACTION_CONTINUE,
            "AUD-BLANK-TYPE-047",
        ),
    )

    loaded = page._load_or_offer_existing_audit_for_machine("47")

    assert loaded is True
    assert page._current_loaded_audit_id == "AUD-BLANK-TYPE-047"
    assert page.audit_fields["Audit ID"].text() == "AUD-BLANK-TYPE-047"


def test_machine_lookup_multiple_physical_selector_excludes_compatible_rows(qapp, fake_config, fake_project, monkeypatch):
    _append_inventory_row(fake_project, _machine_audit_row("AUD-COMPAT-048", "48", ENTRY_TYPE_COMPATIBLE))
    _append_inventory_row(fake_project, _machine_audit_row("AUD-PHYSICAL-048-A", "48", ENTRY_TYPE_AUDITED))
    _append_inventory_row(fake_project, _machine_audit_row("AUD-PHYSICAL-048-B", "48", ENTRY_TYPE_AUDITED))
    page = AuditPage(fake_config)
    dialog_calls = []

    def choose(machine_number: str, matches):
        dialog_calls.append((machine_number, [match.audit_id for match in matches]))
        return ExistingAuditSelection(ExistingMachineAuditsDialog.ACTION_CANCEL)

    monkeypatch.setattr(page, "_choose_existing_machine_audit_action", choose)

    loaded = page._load_or_offer_existing_audit_for_machine("48")

    assert loaded is True
    assert page._current_audit_mode == "new"
    assert page.machine_audit_match_combo.isHidden()
    assert dialog_calls == [("48", ["AUD-PHYSICAL-048-A", "AUD-PHYSICAL-048-B"])]


def test_existing_machine_audits_dialog_lists_required_columns_and_rows(qapp, fake_project):
    _append_inventory_row(
        fake_project,
        _machine_audit_row(
            "AUD-DIALOG-055-A",
            "55",
            ENTRY_TYPE_AUDITED,
            **{"Tool #": "TOOL-55A", "EOAT Type": "Vacuum", "Status": "In Progress", "Priority": "High"},
        ),
    )
    _append_inventory_row(
        fake_project,
        _machine_audit_row(
            "AUD-DIALOG-055-B",
            "55",
            ENTRY_TYPE_AUDITED,
            **{"Tool #": "TOOL-55B", "EOAT Type": "Mechanical / Gripper", "Status": "Complete", "Priority": "Low"},
        ),
    )
    matches = find_existing_audits_for_machine(fake_project, "55")

    dialog = ExistingMachineAuditsDialog("55", matches)
    headers = [dialog.audit_table.horizontalHeaderItem(index).text() for index in range(dialog.audit_table.columnCount())]
    label_text = "\n".join(label.text() for label in dialog.findChildren(QLabel))

    assert dialog.windowTitle() == "Existing Audits Found"
    assert "Machine 55 already has existing audit records" in label_text
    assert "Starting a new audit keeps the machine context" in label_text
    assert headers == ["Audit ID", "Audit Date", "Tool #", "EOAT Type", "Status", "Priority", "Entry Type", "Completion %"]
    assert dialog.audit_table.rowCount() == 2
    assert dialog.audit_table.item(0, 0).text() == "AUD-DIALOG-055-A"
    assert dialog.audit_table.item(1, 2).text() == "TOOL-55B"


def test_continue_selected_existing_audit_loads_selected_match(qapp, fake_config, fake_project, monkeypatch):
    _append_inventory_row(fake_project, _machine_audit_row("AUD-PHYSICAL-053-A", "53", ENTRY_TYPE_AUDITED))
    _append_inventory_row(fake_project, _machine_audit_row("AUD-PHYSICAL-053-B", "53", ENTRY_TYPE_AUDITED, **{"Tool #": "SELECTED-TOOL"}))
    page = AuditPage(fake_config)
    monkeypatch.setattr(
        page,
        "_choose_existing_machine_audit_action",
        lambda _machine, _matches: ExistingAuditSelection(
            ExistingMachineAuditsDialog.ACTION_CONTINUE,
            "AUD-PHYSICAL-053-B",
        ),
    )

    handled = page._load_or_offer_existing_audit_for_machine("53")

    assert handled is True
    assert page._current_loaded_audit_id == "AUD-PHYSICAL-053-B"
    assert page.audit_fields["Audit ID"].text() == "AUD-PHYSICAL-053-B"
    assert page.audit_fields["Tool #"].text() == "SELECTED-TOOL"


def test_start_new_audit_for_existing_machine_uses_reference_context_without_old_tooling(qapp, fake_config, fake_project, monkeypatch):
    create_press_reference_workbooks(fake_project / "reference-data")
    _append_inventory_row(
        fake_project,
        _machine_audit_row(
            "AUD-OLD-012",
            "12",
            ENTRY_TYPE_AUDITED,
            **{
                "Tool #": "OLD-TOOL",
                "EOAT Type": "Mechanical / Gripper",
                "Number of Parts Picked": "8",
                "# of Cups": "6",
                "Cup Type/Material": "Nitrile",
                "Cup Diameter/Size": "30 mm",
                "Vacuum Generator Type": "Old generator",
                "# of Grippers": "4",
                "Gripper Type": "Double Pressure",
                "Gripper Model": "Old gripper",
                "Sensor Type": "Old sensor",
                "Sensor Brand/Model": "Old sensor brand",
                "Known Issues": "Old audit issue should not copy.",
                "Drop/Mis-Pick History": "Old drop history should not copy.",
                "Photos Taken?": "Yes",
                "Photo Folder/Link": "old-folder",
                "Notes": "Old notes should not copy.",
                "Final Notes": "Old final notes should not copy.",
            },
        ),
    )
    page = AuditPage(fake_config)
    monkeypatch.setattr(
        page,
        "_choose_existing_machine_audit_action",
        lambda _machine, _matches: ExistingAuditSelection(ExistingMachineAuditsDialog.ACTION_START_NEW),
    )

    page.audit_fields["Press/Machine #"].setText("12")
    _finish_machine_lookup(page)

    assert page._current_audit_mode == "new"
    assert page._editing_audit_id is None
    assert page._current_loaded_audit_id is None
    assert page.audit_fields["Audit ID"].text() != "AUD-OLD-012"
    assert page.audit_fields["Press/Machine #"].text() == "12"
    assert page.audit_fields["Robot Type"].currentText() == "Wittmann W833"
    assert page.audit_fields["Robot Model/Controller"].text() == "W833"
    assert page.audit_fields["Tool #"].text() == "DEMO-PN-1200"
    assert page.audit_fields["EOAT Type"].currentText() != "Mechanical / Gripper"
    assert page.audit_fields["# of Cups"].text() != "6"
    assert page.audit_fields["Cup Type/Material"].text() != "Nitrile"
    assert page.audit_fields["Known Issues"].toPlainText() != "Old audit issue should not copy."
    assert page.audit_fields["Drop/Mis-Pick History"].toPlainText() != "Old drop history should not copy."
    assert page.audit_fields["Photo Folder/Link"].text() == ""
    assert page.audit_fields["Notes"].toPlainText() == ""
    assert page.has_unsaved_changes() is False


def test_start_new_existing_machine_prompt_is_one_shot_and_keeps_capacity_selector(
    qapp,
    fake_config,
    fake_project,
    monkeypatch,
):
    create_press_reference_workbooks(fake_project / "reference-data", multiple_capacity_rows=True)
    _append_inventory_row(fake_project, _machine_audit_row("AUD-PHYSICAL-012", "12", ENTRY_TYPE_AUDITED))
    page = AuditPage(fake_config)
    dialog_calls = []

    def choose(machine_number: str, matches):
        dialog_calls.append((machine_number, [match.audit_id for match in matches]))
        return ExistingAuditSelection(ExistingMachineAuditsDialog.ACTION_START_NEW)

    monkeypatch.setattr(page, "_choose_existing_machine_audit_action", choose)

    page.audit_fields["Press/Machine #"].setText("12")
    _finish_machine_lookup(page)
    first_audit_id = page.audit_fields["Audit ID"].text()

    assert dialog_calls == [("12", ["AUD-PHYSICAL-012"])]
    assert page._current_audit_mode == "new"
    assert page._editing_audit_id is None
    assert page._current_loaded_audit_id is None
    assert first_audit_id != "AUD-PHYSICAL-012"
    assert page.audit_fields["Press/Machine #"].text() == "12"
    assert page.capacity_part_combo.isEnabled()
    assert _combo_items(page.capacity_part_combo) == [
        "Select current running part...",
        "DEMO-PN-1200 - Demo housing cap - Demo Customer A",
        "DEMO-PN-1201 - Demo housing base - Demo Customer A",
    ]

    _finish_machine_lookup(page)

    assert dialog_calls == [("12", ["AUD-PHYSICAL-012"])]
    assert page._current_audit_mode == "new"
    assert page._editing_audit_id is None
    assert page._current_loaded_audit_id is None
    assert page.audit_fields["Audit ID"].text() == first_audit_id
    assert page.capacity_part_combo.isEnabled()
    assert page.capacity_part_combo.count() == 3

    page.capacity_part_combo.setCurrentIndex(2)

    assert page.audit_fields["Tool #"].text() == "DEMO-PN-1201"
    assert page.audit_fields["Part Name/Description"].toPlainText() == "Demo housing base"


def test_continue_existing_machine_prompt_is_one_shot_during_loaded_audit_hydration(
    qapp,
    fake_config,
    fake_project,
    monkeypatch,
):
    _append_inventory_row(fake_project, _machine_audit_row("AUD-PHYSICAL-060-A", "60", ENTRY_TYPE_AUDITED))
    _append_inventory_row(fake_project, _machine_audit_row("AUD-PHYSICAL-060-B", "60", ENTRY_TYPE_AUDITED))
    page = AuditPage(fake_config)
    dialog_calls = []

    def choose(machine_number: str, matches):
        dialog_calls.append((machine_number, [match.audit_id for match in matches]))
        return ExistingAuditSelection(
            ExistingMachineAuditsDialog.ACTION_CONTINUE,
            "AUD-PHYSICAL-060-B",
        )

    monkeypatch.setattr(page, "_choose_existing_machine_audit_action", choose)

    page.audit_fields["Press/Machine #"].setText("60")
    _finish_machine_lookup(page)
    _finish_machine_lookup(page)

    assert dialog_calls == [("60", ["AUD-PHYSICAL-060-A", "AUD-PHYSICAL-060-B"])]
    assert page._current_loaded_audit_id == "AUD-PHYSICAL-060-B"
    assert page.audit_fields["Audit ID"].text() == "AUD-PHYSICAL-060-B"
    assert page._current_audit_mode == "edit"


def test_programmatic_existing_audit_load_does_not_open_same_machine_dialog(
    qapp,
    fake_config,
    fake_project,
    monkeypatch,
):
    _append_inventory_row(fake_project, _machine_audit_row("AUD-PHYSICAL-061-A", "61", ENTRY_TYPE_AUDITED))
    _append_inventory_row(fake_project, _machine_audit_row("AUD-PHYSICAL-061-B", "61", ENTRY_TYPE_AUDITED))
    page = AuditPage(fake_config)
    dialog_calls = []

    def choose(machine_number: str, matches):
        dialog_calls.append((machine_number, [match.audit_id for match in matches]))
        return ExistingAuditSelection(ExistingMachineAuditsDialog.ACTION_CANCEL)

    monkeypatch.setattr(page, "_choose_existing_machine_audit_action", choose)

    assert page.load_existing_audit("AUD-PHYSICAL-061-A", confirm_unsaved=False) is True
    _finish_machine_lookup(page)

    assert dialog_calls == []
    assert page._current_loaded_audit_id == "AUD-PHYSICAL-061-A"
    assert page.audit_fields["Audit ID"].text() == "AUD-PHYSICAL-061-A"


def test_cancel_existing_machine_prompt_is_one_shot_and_leaves_form_stable(
    qapp,
    fake_config,
    fake_project,
    monkeypatch,
):
    _append_inventory_row(fake_project, _machine_audit_row("AUD-PHYSICAL-062", "62", ENTRY_TYPE_AUDITED))
    page = AuditPage(fake_config)
    page.audit_fields["Known Issues"].setPlainText("Keep this note.")
    before_prompt = page._current_audit_form_values()
    dialog_calls = []

    def choose(machine_number: str, matches):
        dialog_calls.append((machine_number, [match.audit_id for match in matches]))
        return ExistingAuditSelection(ExistingMachineAuditsDialog.ACTION_CANCEL)

    monkeypatch.setattr(page, "_choose_existing_machine_audit_action", choose)

    page.audit_fields["Press/Machine #"].setText("62")
    _finish_machine_lookup(page)
    after_cancel = page._current_audit_form_values()
    _finish_machine_lookup(page)

    assert dialog_calls == [("62", ["AUD-PHYSICAL-062"])]
    assert after_cancel["Press/Machine #"] == "62"
    assert after_cancel["Known Issues"] == before_prompt["Known Issues"]
    assert page._current_audit_form_values() == after_cancel
    assert page._current_loaded_audit_id is None
    assert page._current_audit_mode == "new"


def test_cancel_existing_audit_selection_leaves_form_values_unchanged(qapp, fake_config, fake_project, monkeypatch):
    _append_inventory_row(fake_project, _machine_audit_row("AUD-PHYSICAL-054", "54", ENTRY_TYPE_AUDITED))
    page = AuditPage(fake_config)
    original_id = page.audit_fields["Audit ID"].text()
    page.audit_fields["Known Issues"].setPlainText("Keep my unsaved issue.")
    page.audit_fields["Tool #"].setText("MANUAL-TOOL")
    before = page._current_audit_form_values()
    monkeypatch.setattr(
        page,
        "_choose_existing_machine_audit_action",
        lambda _machine, _matches: ExistingAuditSelection(ExistingMachineAuditsDialog.ACTION_CANCEL),
    )

    handled = page._load_or_offer_existing_audit_for_machine("54")

    assert handled is True
    assert page._current_loaded_audit_id is None
    assert page.audit_fields["Audit ID"].text() == original_id
    assert page._current_audit_form_values() == before


def test_saving_start_new_same_machine_creates_unique_row_without_overwriting_unfinished_audit(
    qapp,
    fake_config,
    fake_project,
    monkeypatch,
):
    create_press_reference_workbooks(fake_project / "reference-data")
    _append_inventory_row(
        fake_project,
        _machine_audit_row(
            "AUD-UNFINISHED-012",
            "12",
            ENTRY_TYPE_AUDITED,
            **{
                "Tool #": "OLD-UNFINISHED-TOOL",
                "EOAT Type": "Vacuum",
                "Status": "In Progress",
                "Known Issues": "Unfinished audit issue remains.",
                "Notes": "Unfinished audit note remains.",
            },
        ),
    )
    page = AuditPage(fake_config)
    monkeypatch.setattr(
        page,
        "_choose_existing_machine_audit_action",
        lambda _machine, _matches: ExistingAuditSelection(ExistingMachineAuditsDialog.ACTION_START_NEW),
    )

    page.audit_fields["Press/Machine #"].setText("12")
    _finish_machine_lookup(page)
    new_audit_id = page.audit_fields["Audit ID"].text()
    _set_field(page, "EOAT Type", "Vacuum")
    _set_field(page, "Status", "In Progress")
    _set_field(page, "Notes", "Brand-new audit for swapped EOAT.")

    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()

    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
    old_row = next(row for row in rows if row["Audit ID"] == "AUD-UNFINISHED-012")
    new_row = next(row for row in rows if row["Audit ID"] == new_audit_id)
    machine_12_ids = {row["Audit ID"] for row in rows if row["Press/Machine #"] == "12"}

    assert new_audit_id != "AUD-UNFINISHED-012"
    assert {"AUD-UNFINISHED-012", new_audit_id}.issubset(machine_12_ids)
    assert old_row["Known Issues"] == "Unfinished audit issue remains."
    assert old_row["Notes"] == "Unfinished audit note remains."
    assert old_row["Tool #"] == "OLD-UNFINISHED-TOOL"
    assert new_row["Press/Machine #"] == "12"
    assert new_row["Tool #"] == "DEMO-PN-1200"
    assert new_row["Notes"] == "Brand-new audit for swapped EOAT."


def test_save_writes_clean_audit_fields_not_reference_values(qapp, fake_config, fake_project, frozen_project_date):
    create_press_reference_workbooks(fake_project / "reference-data")
    page = AuditPage(fake_config)
    page.show()

    page.audit_fields["Press/Machine #"].setText("P12")
    _finish_machine_lookup(page)
    _set_field(page, "EOAT Type", "Vacuum")
    _set_field(page, "Sensors Present?", "Yes")
    _set_field(page, "Vacuum Confirmation Present?", "Yes")
    _set_field(page, "Quick Disconnects Present?", "Unknown / Not Checked")
    _set_field(page, "Tubing Condition", "OK")
    _set_field(page, "Known Issues", "No issue during test.")
    _set_field(page, "Notes", "Clean form save test.")

    audit_id = page.audit_fields["Audit ID"].text()
    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()

    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
    row = next(row for row in rows if row["Audit ID"] == audit_id)
    assert row["Auditor"] == "Kato Gray"
    assert row["Plant/Area"] == "Plant 4"
    assert row["Press/Machine #"] == "12"
    assert row["Tool #"] == "DEMO-PN-1200"
    assert row["Robot Type"] == "Wittmann W833"
    assert row["Robot Model/Controller"] == "W833"
    assert row["Part Family"] == "DEMO-PN-1200 - Demo housing cap"
    assert row["Part Name/Description"] == "Demo housing cap"
    assert row["Sensors Present?"] == "Yes"
    assert row["Vacuum Confirmation Present?"] == "Yes"
    assert row["Quick Disconnects Present?"] == "Unknown / Not Checked"
    assert row["Tubing Condition"] == "OK"
    assert row["Notes"] == "Clean form save test."
    for field in REFERENCE_ONLY_FIELDS:
        assert field not in row or row[field] in ("", None)


def test_lookup_does_not_overwrite_manual_robot_type(qapp, fake_config, fake_project):
    create_press_reference_workbooks(fake_project / "reference-data")
    page = AuditPage(fake_config)

    page.audit_fields["Robot Type"].setEditText("Manual robot")
    page.audit_fields["Press/Machine #"].setText("12")
    _finish_machine_lookup(page)

    assert page.audit_fields["Robot Type"].currentText() == "Manual robot"
    assert "different Robot Type suggestion" in page.lookup_note_label.text()


def test_lookup_does_not_overwrite_manual_tool_number(qapp, fake_config, fake_project):
    create_press_reference_workbooks(fake_project / "reference-data")
    page = AuditPage(fake_config)

    page.audit_fields["Tool #"].setText("MANUAL-TOOL")
    page.audit_fields["Press/Machine #"].setText("12")
    _finish_machine_lookup(page)

    assert page.audit_fields["Tool #"].text() == "MANUAL-TOOL"
    assert "different Tool # suggestion" in page.lookup_note_label.text()

