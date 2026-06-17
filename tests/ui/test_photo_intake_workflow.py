from __future__ import annotations

import json
import os
from datetime import datetime

import pytest
from openpyxl import load_workbook
from PySide6.QtCore import Qt
from PySide6.QtTest import QTest
from PySide6.QtWidgets import QLabel, QTableWidgetItem, QWidget

from app.pages.photos import PhotosPage
from core.audit_field_links import build_audit_field_link, serialize_audit_field_link
from core.paths import resolve_project_paths
from core.workbook_cache import invalidate_workbook_cache
from core.workbook_io import row_dicts
from core.workbook_schema import get_expected_headers
from tests.ui.helpers import click_button, process_events, wait_for_background_tasks

pytestmark = pytest.mark.usability


def _combo_data_values(combo) -> set[str]:
    return {str(combo.itemData(index) or "") for index in range(combo.count())}


def _combo_text_values(combo) -> set[str]:
    return {combo.itemText(index) for index in range(combo.count())}


class _LinkHost(QWidget):
    def __init__(self):
        super().__init__()
        self.opened_link = ""

    def navigate_to_audit_field_link(self, link_text: str) -> bool:
        self.opened_link = link_text
        return True


def _sample_link() -> str:
    link = build_audit_field_link(
        {
            "Audit ID": "AUD-20260518-001",
            "Press/Machine #": "Press 101",
            "Tool #": "TOOL-A",
        },
        "Tubing Routing Notes",
        "Tubing Routing Notes",
        created_at="2026-06-02T12:00:00+00:00",
    )
    return serialize_audit_field_link(link)


def _append_inventory_row(project_root, values: dict[str, object]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    row = {header: "" for header in get_expected_headers("EOAT Inventory")}
    row.update(values)
    ws.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()
    invalidate_workbook_cache(workbook_path)


def test_photo_intake_previews_copies_and_indexes_selected_images(qapp, fake_config, fake_project, frozen_project_date):
    page = PhotosPage(fake_config)
    page.show()

    click_button(page, "Refresh Incoming Photos")
    assert page.incoming_list.count() >= 2
    page.incoming_list.item(0).setSelected(True)
    page.incoming_list.item(1).setSelected(True)
    page.plant_combo.setCurrentText("Whiteroom")
    page._select_tool_number("TOOL-A", machine="Press 101")
    page.date_edit.setText("2026-05-19")
    page.view_combo.setCurrentText("Vacuum Cups / Grippers")
    page.description_edit.setPlainText("Fake intake usability photos.")
    page.notes_edit.setPlainText("No real internship image files used.")

    click_button(page, "Preview Intake")
    preview_text = page.result_panel.viewer.toPlainText()
    assert "Tool_TOOL-A__Vacuum_Cups_Grippers__2026-05-19__001" in preview_text

    click_button(page, "Save Photos to EOAT Folder")
    wait_for_background_tasks()

    target_dir = (
        fake_project
        / "01_EOAT_Audit"
        / "Cell_Photos"
        / "Tool_TOOL-A__Vacuum_EOAT_family_A_sample"
        / "03_Vacuum_Cups_Grippers"
    )
    imported = list(target_dir.glob("Tool_TOOL-A__Vacuum_Cups_Grippers__2026-05-19__*"))
    assert len(imported) == 2
    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "Photo Index")
    assert sum(1 for row in rows if row["Related Audit ID"] == "AUD-20260518-001") >= 2

    activity_log = fake_project / "00_Project_Admin" / "Activity_Logs" / "activity_log.jsonl"
    entries = [json.loads(line) for line in activity_log.read_text(encoding="utf-8").splitlines() if line.strip()]
    assert any(entry["tool_id"] == "photo_intake" and entry["success"] for entry in entries)


def test_photos_page_related_audit_and_issue_fields_are_dropdowns(qapp, usability_fake_config):
    from PySide6.QtWidgets import QComboBox

    page = PhotosPage(usability_fake_config)
    page.show()

    assert isinstance(page.audit_id_edit, QComboBox)
    assert isinstance(page.issue_id_edit, QComboBox)
    assert not any(label.text() == "Audit Lookup" for label in page.findChildren(QLabel))
    assert {"AUD-20260518-001", "AUD-20260518-002", "AUD-20260518-003"}.issubset(
        _combo_data_values(page.audit_id_edit)
    )
    assert {"ISS-001", "ISS-002", "ISS-003"}.issubset(_combo_data_values(page.issue_id_edit))

    page.audit_id_edit.setCurrentIndex(page.audit_id_edit.findData("AUD-20260518-002"))
    page.issue_id_edit.setCurrentIndex(page.issue_id_edit.findData("ISS-002"))

    assert page.audit_id_edit.text() == "AUD-20260518-002"
    assert page.issue_id_edit.text() == "ISS-002"
    metadata = page.metadata()
    assert metadata["related_audit_id"] == "AUD-20260518-002"
    assert metadata["related_issue_id"] == "ISS-002"


def test_photos_page_eoat_area_options_use_expected_photo_types(qapp, usability_fake_config):
    page = PhotosPage(usability_fake_config)
    page.show()

    options = _combo_text_values(page.view_combo)

    assert {"Front View", "Side View", "Back View", "Tool Number", "Gripper"} <= options
    assert "Overall" not in options
    assert "Overall EOAT" not in options


def test_photos_page_receives_pending_audit_field_link(qapp, usability_fake_config):
    page = PhotosPage(usability_fake_config)
    page.show()
    link_text = _sample_link()

    page.apply_pending_audit_field_link(link_text)

    assert page.audit_field_link_edit.text() == link_text
    assert page.audit_id_edit.text() == "AUD-20260518-001"
    assert page._current_tool_number() == "TOOL-A"
    assert "Field: Tubing Routing Notes" in page.link_display_label.text()
    assert page.go_to_link_button.isEnabled()


def test_photos_page_go_to_link_uses_navigation_host(qapp, usability_fake_config):
    host = _LinkHost()
    page = PhotosPage(usability_fake_config, parent=host)
    page.show()
    link_text = _sample_link()
    page.apply_pending_audit_field_link(link_text)

    click_button(page, "Go to Link")

    assert host.opened_link == link_text


def test_photos_page_go_to_link_uses_selected_indexed_photo_row(qapp, usability_fake_config):
    host = _LinkHost()
    page = PhotosPage(usability_fake_config, parent=host)
    page.show()
    link_text = _sample_link()
    page.audit_field_link_edit.clear()
    page.indexed_photos_table.setRowCount(1)
    page.indexed_photos_table.setItem(0, 0, QTableWidgetItem("PHO-LINK-001"))
    page.indexed_photos_table.setItem(0, 4, QTableWidgetItem(link_text))
    page.indexed_photos_table.setCurrentCell(0, 4)
    page._update_go_to_link_button()

    click_button(page, "Go to Link")

    assert host.opened_link == link_text


def test_photos_page_invalid_and_legacy_links_are_unavailable(qapp, usability_fake_config):
    page = PhotosPage(usability_fake_config)
    page.show()
    page.audit_field_link_edit.setText("Machine 12 - Tubing Routing Notes")

    assert not page.go_to_link_button.isEnabled()

    page.go_to_audit_field_link()

    assert "unavailable" in page.result_panel.viewer.toPlainText()


def test_photo_intake_empty_incoming_folder_has_helpful_empty_state(qapp, fake_config, fake_project):
    incoming = fake_project / "01_EOAT_Audit" / "Cell_Photos" / "Incoming_Photos"
    for path in incoming.iterdir():
        path.unlink()

    page = PhotosPage(fake_config)
    page.show()
    click_button(page, "Refresh Incoming Photos")

    assert page.incoming_list.count() == 0
    assert "No incoming photos found" in page.empty_hint.text()
    assert "Incoming_Photos" in page.result_panel.viewer.toPlainText()


def test_photos_page_spacebar_opens_keyboard_preview(qapp, usability_fake_config):
    page = PhotosPage(usability_fake_config)
    page.show()
    click_button(page, "Refresh Incoming Photos")
    page.incoming_list.setCurrentRow(0)
    page.incoming_list.item(0).setSelected(True)

    QTest.keyClick(page.incoming_list, Qt.Key.Key_Space)
    process_events(100)

    dialog = page._photo_preview_dialog
    assert dialog is not None
    assert dialog.isVisible()
    assert dialog.current_path == page._incoming_paths[0]

    QTest.keyClick(dialog, Qt.Key.Key_Right)
    process_events(50)

    assert dialog.current_path == page._incoming_paths[1]

    QTest.keyClick(dialog, Qt.Key.Key_Escape)
    process_events(50)

    assert page._photo_preview_dialog is None or not page._photo_preview_dialog.isVisible()


def test_photos_page_preview_selected_photos_opens_contact_sheet(qapp, usability_fake_config):
    page = PhotosPage(usability_fake_config)
    page.show()
    click_button(page, "Refresh Incoming Photos")
    page.incoming_list.item(0).setSelected(True)
    page.incoming_list.item(1).setSelected(True)

    click_button(page, "Preview Selected Photos")
    process_events(100)

    dialog = page._contact_sheet_dialog
    assert dialog is not None
    assert dialog.isVisible()
    assert dialog.paths == page.selected_photos()


def test_photos_page_autofills_date_from_selected_photo_mtime(qapp, usability_fake_config):
    incoming = resolve_project_paths(usability_fake_config.project_root).incoming_photos
    photo = sorted(incoming.iterdir())[0]
    stamp = datetime(2026, 2, 3, 8, 30).timestamp()
    photo.touch()
    os.utime(photo, (stamp, stamp))

    page = PhotosPage(usability_fake_config)
    page.show()
    click_button(page, "Refresh Incoming Photos")

    assert page.date_edit.text() == "2026-02-03"


def test_photos_page_audit_lookup_autofill_and_next_missing_shot(
    qapp, usability_fake_config, usability_fake_project
):
    page = PhotosPage(usability_fake_config)
    page.show()

    target_index = page.audit_id_edit.findData("AUD-20260518-002")
    assert target_index > 0

    page.audit_id_edit.setCurrentIndex(target_index)
    page.refresh_audit_context()

    assert page.audit_id_edit.text() == "AUD-20260518-002"
    assert page._press_machine == "Press 102"
    assert page._current_tool_number() == "TOOL-B"
    assert page.tool_combo.currentText() == "TOOL-B"
    assert " | " not in page.tool_combo.currentText()
    assert "Tool #: TOOL-B" in page.audit_context_label.text()
    assert "Gripper Size" not in page.audit_context_label.text()

    click_button(page, "Use Next Missing Shot Type")

    assert page.view_combo.currentText() in {"Front View", "Tool Connection", "Grippers", "Tool Label / ID Plate"}


def test_tool_dropdown_shows_machine_context_from_completed_physical_audits(
    qapp, usability_fake_config, usability_fake_project
):
    _append_inventory_row(
        usability_fake_project,
        {
            "Audit ID": "AUD-COMPAT-PHOTO-001",
            "Press/Machine #": "Press 999",
            "Tool #": "TOOL-COMPAT",
            "Status": "Complete",
            "Entry Type": "Compatible",
        },
    )
    _append_inventory_row(
        usability_fake_project,
        {
            "Audit ID": "AUD-INPROGRESS-PHOTO-001",
            "Press/Machine #": "Press 888",
            "Tool #": "TOOL-INPROGRESS",
            "Status": "In Progress",
        },
    )
    _append_inventory_row(
        usability_fake_project,
        {
            "Audit ID": "AUD-AUDITED-INPROGRESS-PHOTO-001",
            "Press/Machine #": "Press 777",
            "Tool #": "TOOL-AUDITED-INPROGRESS",
            "Status": "In Progress",
            "Entry Type": "Audited",
        },
    )
    _append_inventory_row(
        usability_fake_project,
        {
            "Audit ID": "AUD-NOMACHINE-PHOTO-001",
            "Tool #": "TOOL-NOMACHINE",
            "Status": "Complete",
        },
    )
    _append_inventory_row(
        usability_fake_project,
        {
            "Audit ID": "AUD-MACHINE-NA-PHOTO-001",
            "Press/Machine #": "N/A",
            "Tool #": "TOOL-MACHINE-NA",
            "Status": "In Progress",
            "Entry Type": "Audited",
        },
    )
    _append_inventory_row(
        usability_fake_project,
        {
            "Audit ID": "AUD-NA-TOOL-PHOTO-001",
            "Press/Machine #": "Press 555",
            "Tool #": "N/A",
            "Status": "Complete",
        },
    )

    page = PhotosPage(usability_fake_config)
    page.show()
    option_texts = [page.tool_combo.itemText(index) for index in range(page.tool_combo.count())]

    assert "TOOL-B (Press 102)" in option_texts
    assert "TOOL-AUDITED-INPROGRESS (Press 777)" in option_texts
    assert "TOOL-COMPAT (Press 999)" not in option_texts
    assert "TOOL-INPROGRESS (Press 888)" not in option_texts
    assert "TOOL-NOMACHINE" in option_texts
    assert "TOOL-MACHINE-NA" in option_texts
    assert "TOOL-MACHINE-NA (N/A)" not in option_texts
    assert "N/A (Press 555)" not in option_texts

    page.tool_combo.setCurrentIndex(option_texts.index("TOOL-B (Press 102)"))

    assert page.tool_combo.currentText() == "TOOL-B"
    assert page._current_tool_number() == "TOOL-B"
    assert "Machine: Press 102" in page.audit_context_label.text()


def test_photos_page_manual_tool_entry_clears_machine_context(qapp, usability_fake_config):
    page = PhotosPage(usability_fake_config)
    page.show()
    page._select_tool_number("TOOL-B", machine="Press 102")
    assert page._press_machine == "Press 102"

    page.tool_combo.setEditText("OFF-MACHINE-TOOL")

    assert page._current_tool_number() == "OFF-MACHINE-TOOL"
    assert page._press_machine == ""


def test_photos_page_batch_review_supports_per_photo_shot_types(
    qapp, usability_fake_config, usability_fake_project, frozen_project_date
):
    page = PhotosPage(usability_fake_config)
    page.show()

    click_button(page, "Refresh Incoming Photos")
    page.incoming_list.item(0).setSelected(True)
    page.incoming_list.item(1).setSelected(True)
    page.plant_combo.setCurrentText("Whiteroom")
    page._select_tool_number("TOOL-A", machine="Press 101")
    page.date_edit.setText("2026-05-19")

    click_button(page, "Build Batch Review")
    first_view = page.batch_table.cellWidget(0, page.BATCH_VIEW_COL)
    second_view = page.batch_table.cellWidget(1, page.BATCH_VIEW_COL)
    first_view.setCurrentText("Front View")
    second_view.setCurrentText("Sensors")

    click_button(page, "Preview Intake")
    preview_text = page.result_panel.viewer.toPlainText()

    assert "Front_View" in preview_text
    assert "Sensors" in preview_text

    click_button(page, "Save Photos to EOAT Folder")
    wait_for_background_tasks()

    rows = row_dicts(
        usability_fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx",
        "Photo Index",
    )
    related = [row for row in rows if row["Related Audit ID"] == "AUD-20260518-001"]
    assert any(row["EOAT Area Shown"] == "Front View" for row in related)
    assert any(row["EOAT Area Shown"] == "Sensors" for row in related)


def test_photos_page_batch_review_related_ids_are_dropdowns(
    qapp, usability_fake_config, usability_fake_project, frozen_project_date
):
    from PySide6.QtWidgets import QComboBox

    page = PhotosPage(usability_fake_config)
    page.show()

    click_button(page, "Refresh Incoming Photos")
    page.incoming_list.item(0).setSelected(True)
    page.plant_combo.setCurrentText("Whiteroom")
    page._select_tool_number("TOOL-A", machine="Press 101")
    page.date_edit.setText("2026-05-19")
    page.issue_id_edit.setText("ISS-001")

    click_button(page, "Build Batch Review")
    audit_widget = page.batch_table.cellWidget(0, page.BATCH_AUDIT_COL)
    issue_widget = page.batch_table.cellWidget(0, page.BATCH_ISSUE_COL)

    assert isinstance(audit_widget, QComboBox)
    assert isinstance(issue_widget, QComboBox)
    assert "AUD-20260518-002" in _combo_data_values(audit_widget)
    assert "ISS-002" in _combo_data_values(issue_widget)

    audit_widget.setCurrentIndex(audit_widget.findData("AUD-20260518-002"))
    issue_widget.setCurrentIndex(issue_widget.findData("ISS-002"))

    _photos, metadata = page.selected_batch_metadata()
    assert metadata[0]["related_audit_id"] == "AUD-20260518-002"
    assert metadata[0]["related_issue_id"] == "ISS-002"


def test_photos_page_repair_audit_photo_ties_button_runs(qapp, usability_fake_config):
    page = PhotosPage(usability_fake_config)
    page.show()

    click_button(page, "Repair Audit Photo Ties")
    wait_for_background_tasks()

    result_text = page.result_panel.viewer.toPlainText()
    assert "Repair Audit Photo Ties" in result_text
    assert "Photo Index row" in result_text or "Photo ties" in result_text
