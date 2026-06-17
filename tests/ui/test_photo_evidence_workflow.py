from __future__ import annotations

import pytest
from openpyxl import load_workbook

from app.pages.photos import PhotosPage
from core.eoat_ids import EOAT_ASSEMBLY_ID_FIELD
from core.paths import resolve_project_paths
from core.photo_evidence import audit_photo_intake_folder
from core.photo_indexing import eoat_photo_root
from core.workbook_schema import get_expected_headers
from tests.ui.helpers import click_button

pytestmark = pytest.mark.usability


def _append_inventory_row(project_root, values: dict[str, str]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    row = {header: "" for header in get_expected_headers("EOAT Inventory")}
    row.update(values)
    ws.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def test_photos_page_evidence_folder_checklist_and_open_actions(
    qapp, fake_config, fake_project, captured_open_requests
):
    _append_inventory_row(
        fake_project,
        {
            "Audit ID": "AUD-UI-EVID-001",
            "Audit Date": "2026-05-18",
            "Auditor": "Synthetic Auditor",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-0101",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Status": "Complete",
            "Sensors Present?": "No",
            "Quick Disconnects Present?": "No",
        },
    )
    page = PhotosPage(fake_config)
    page.show()
    page._set_eoat_combo_value("P4-EOAT-0101")
    page.audit_id_edit.setText("AUD-UI-EVID-001")

    click_button(page, "Refresh Coverage")
    assert page.evidence_table.rowCount() > 0
    assert "required missing" in page.result_panel.viewer.toPlainText()

    click_button(page, "Create EOAT Photo Folder")
    eoat_folder = eoat_photo_root(fake_project, "P4-EOAT-0101")
    assert eoat_folder.exists()
    folder = audit_photo_intake_folder(fake_project, "AUD-UI-EVID-001")

    click_button(page, "Export Photo Checklist")
    assert list(folder.glob("Photo_Checklist_AUD-UI-EVID-001_*.md"))

    click_button(page, "Open EOAT Folder")
    assert captured_open_requests[-1] == eoat_folder
