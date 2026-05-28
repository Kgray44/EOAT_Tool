from __future__ import annotations

import json
import re
from collections import Counter

import pytest
from openpyxl import load_workbook
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QComboBox, QGroupBox, QPushButton, QTextEdit

from app.pages.audit import AUDIT_SECTIONS, AUDIT_SECTION_GROUPS, AuditPage
from core.audit_constants import (
    IGNORED_EMPTY_FIELDS_AT_OVERRIDE_FIELD,
    MANUAL_COMPLETION_OVERRIDE_FIELD,
    MANUAL_COMPLETION_OVERRIDE_TIMESTAMP_FIELD,
    MANUAL_COMPLETION_OVERRIDE_USER_FIELD,
)
from core.audit_entries import save_audit_entry
from core.robot_info import load_robot_info_for_audit_entry, robot_info_workbook_path, upsert_robot_info_from_audit
from core.workbook_io import row_dicts, workbook_sheet_names
from tests.fixtures.reference_workbooks import create_press_reference_workbooks
from tests.ui.helpers import click_button, wait_for_background_tasks, wait_until


pytestmark = pytest.mark.usability


def _set_field(page: AuditPage, field: str, value: str) -> None:
    widget = page.audit_fields[field]
    if isinstance(widget, QComboBox):
        if widget.findText(value) >= 0:
            widget.setCurrentText(value)
        else:
            widget.setEditText(value)
    elif isinstance(widget, QTextEdit):
        widget.setPlainText(value)
    else:
        widget.setText(value)


def _field_value(page: AuditPage, field: str) -> str:
    return page._field_value(page.audit_fields[field])


def _finish_machine_lookup(page: AuditPage) -> None:
    page.audit_fields["Press/Machine #"].editingFinished.emit()
    wait_until(
        lambda: page._machine_lookup_timer is None or not page._machine_lookup_timer.isActive(),
        timeout_ms=5000,
        message="machine lookup debounce timer",
    )
    wait_for_background_tasks()


def _seed_audit(project_root, audit_id: str, **overrides):
    entry = {
        "Audit ID": audit_id,
        "Audit Date": "2026-05-18",
        "Auditor": "Original Auditor",
        "Plant/Area": "Plant 4",
        "Press/Machine #": "Press 12",
        "Tool #": "DEMO-PN-1200",
        "Robot Type": "Wittmann R9",
        "Robot Model/Controller": "W833",
        "Part Family": "DEMO-PN-1200 - Demo housing cap",
        "Part Name/Description": "Demo housing cap",
        "EOAT Type": "Vacuum",
        "Connection Type": "Direct Mount",
        "Cleanroom/Non-Cleanroom": "Cleanroom",
        "Cup Type/Material": "Silicone",
        "Gripper Model": "Zimmer GPP",
        "Gripper Size": "25 mm",
        "Tubing Condition": "OK",
        "Cable Management Condition": "OK",
        "Known Issues": "Original known issue.",
        "Photos Taken?": "Yes",
        "Photo Folder/Link": "01_EOAT_Audit/Cell_Photos/Overall/original",
        "Status": "In Progress",
        "Priority": "Medium",
        "Follow-Up Needed": "No",
        "Notes": "Original notes.",
    }
    entry.update(overrides)
    result = save_audit_entry(project_root, entry)
    assert result.success, result.errors
    return entry


def _empty_only_count_message(page: AuditPage) -> int:
    match = re.search(r"Empty Only is showing (\d+) blank or N/A field", page.result_panel.viewer.toPlainText())
    assert match
    return int(match.group(1))


def _visible_empty_only_fields(page: AuditPage) -> list[str]:
    return [field for field, widget in page.audit_fields.items() if not widget.isHidden() and not _field_value(page, field).strip()]


def test_save_complete_and_optional_missing_audit_entries(qapp, fake_config, fake_project, frozen_project_date):
    page = AuditPage(fake_config)
    page.show()

    complete_values = {
        "Auditor": "Usability Tester",
        "Plant/Area": "Plant 4",
        "Press/Machine #": "Press 104",
        "Robot Type": "Wittmann R9",
        "Robot Model/Controller": "Wittmann synthetic",
        "Part Family": "Part family D",
        "EOAT Type": "Vacuum",
        "EOAT Moves": "Part",
        "Connection Type": "ATI",
        "Tubing Condition": "OK",
        "Cable Management Condition": "OK",
        "Known Issues": "No issue found during fake test.",
        "Photos Taken?": "Yes",
        "Status": "Complete",
        "Priority": "Medium",
        "Follow-Up Needed": "No",
        "Notes": "Complete audit entry from usability test.",
    }
    for field, value in complete_values.items():
        _set_field(page, field, value)
    complete_id = page.audit_fields["Audit ID"].text()
    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()

    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
    assert any(row["Audit ID"] == complete_id and row["Press/Machine #"] == "Press 104" for row in rows)
    saved_row = next(row for row in rows if row["Audit ID"] == complete_id)
    assert saved_row["Connection Type"] == "ATI"
    assert saved_row["EOAT Moves"] == "Part"
    assert saved_row["Cup Type/Material"] == "Silicone"
    assert saved_row["Cleanroom/Non-Cleanroom"] == "Whiteroom"
    assert saved_row["# of Grippers"] == "N/A"
    assert saved_row["Gripper Type"] == "N/A"
    assert saved_row["Gripper Model"] == "N/A"
    assert saved_row["Notes"] == "Complete audit entry from usability test."
    assert "Saved audit entry" in page.result_panel.viewer.toPlainText()

    page.clear_audit_form()
    for field, value in {
        "Auditor": "Usability Tester",
        "Plant/Area": "Plant 4",
        "Press/Machine #": "Press 105",
        "Robot Type": "Engel Viper",
        "EOAT Type": "Mechanical / Gripper",
    }.items():
        _set_field(page, field, value)
    optional_missing_id = page.audit_fields["Audit ID"].text()
    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()
    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
    assert any(row["Audit ID"] == optional_missing_id and row["Press/Machine #"] == "Press 105" for row in rows)
    optional_row = next(row for row in rows if row["Audit ID"] == optional_missing_id)
    assert optional_row["Cleanroom/Non-Cleanroom"] == "Whiteroom"
    assert optional_row["Cup Type/Material"] == "N/A"
    assert optional_row["Cup Diameter/Size"] == "N/A"
    assert optional_row["Number of Parts Picked"] == "N/A"
    assert optional_row["# of Cups"] == "N/A"
    assert optional_row["# of Grippers"] == "N/A"
    assert optional_row["Gripper Type"] == "N/A"
    assert optional_row["Connection Type"] == "N/A"
    assert optional_row["Known Issues"] == "Unknown / Not Checked"
    assert optional_row["Drop/Mis-Pick History"] == "Unknown / Not Checked"
    assert optional_row["Maintenance Frequency"] == "Unknown / Not Checked"

    activity_log = fake_project / "00_Project_Admin" / "Activity_Logs" / "activity_log.jsonl"
    entries = [json.loads(line) for line in activity_log.read_text(encoding="utf-8").splitlines() if line.strip()]
    assert any(entry.get("tool_id") == "eoat_audit_form" and entry.get("success") for entry in entries)


def test_invalid_audit_entry_shows_friendly_error_without_workbook_row(qapp, fake_config, fake_project, frozen_project_date):
    page = AuditPage(fake_config)
    page.show()
    page.clear_audit_form()
    bad_id = page.audit_fields["Audit ID"].text()
    _set_field(page, "Auditor", "")
    _set_field(page, "Plant/Area", "Plant 4")
    _set_field(page, "Press/Machine #", "Press BAD")
    _set_field(page, "Robot Type", "")
    _set_field(page, "EOAT Type", "")

    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()
    text = page.result_panel.viewer.toPlainText()
    assert "Audit entry failed validation" in text
    assert "Missing required field" in text
    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
    assert all(row["Audit ID"] != bad_id for row in rows)


def test_loading_existing_audit_preserves_saved_reliability_values(qapp, fake_config, fake_project, frozen_project_date):
    original = _seed_audit(
        fake_project,
        "AUD-RELIABILITY-VALUES-001",
        **{
            "Known Issues": "Vacuum cup wear.",
            "Drop/Mis-Pick History": "Two drops last week.",
            "Maintenance Frequency": "Weekly",
        },
    )
    page = AuditPage(fake_config)
    page.show()

    page.load_audit_id_edit.setText(original["Audit ID"])
    click_button(page, "Load Existing Audit ID")

    assert page.audit_fields["Known Issues"].toPlainText() == "Vacuum cup wear."
    assert page.audit_fields["Drop/Mis-Pick History"].toPlainText() == "Two drops last week."
    assert page.audit_fields["Maintenance Frequency"].text() == "Weekly"


def test_audit_id_dropdown_includes_all_audits_and_selection_loads_row(qapp, fake_config, fake_project, frozen_project_date):
    first = _seed_audit(fake_project, "AUD-DROPDOWN-001", **{"Press/Machine #": "Press 26", "Tool #": "PN-26", "Part Name/Description": "Latch", "Notes": "First row."})
    second = _seed_audit(fake_project, "AUD-DROPDOWN-002", **{"Press/Machine #": "Press 2", "Tool #": "PN-02", "Part Name/Description": "Clip", "Notes": "Second row."})
    page = AuditPage(fake_config)
    page.show()
    page._load_lazy_audit_indexes()

    audit_ids = [page.load_audit_id_combo.itemData(index) for index in range(page.load_audit_id_combo.count())]
    labels = [page.load_audit_id_combo.itemText(index) for index in range(page.load_audit_id_combo.count())]
    assert first["Audit ID"] in audit_ids
    assert second["Audit ID"] in audit_ids
    assert any("AUD-DROPDOWN-001 | Machine 26 | PN-26 | Latch | Audited" == label for label in labels)

    index = page.load_audit_id_combo.findData(second["Audit ID"])
    assert index >= 0
    page.load_audit_id_combo.setCurrentIndex(index)
    page._on_audit_selector_activated(index)

    assert page.audit_fields["Audit ID"].text() == second["Audit ID"]
    assert page.audit_fields["Press/Machine #"].text() == "Press 2"
    assert page.audit_fields["Notes"].toPlainText() == "Second row."


def test_audit_id_dropdown_refreshes_after_save(qapp, fake_config, fake_project, frozen_project_date):
    page = AuditPage(fake_config)
    page.show()

    _set_field(page, "Press/Machine #", "Press 8")
    _set_field(page, "Robot Type", "Wittmann R9")
    _set_field(page, "EOAT Type", "Vacuum")
    audit_id = page.audit_fields["Audit ID"].text()
    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()

    assert page.load_audit_id_combo.findData(audit_id) >= 0


def test_audit_and_compatibility_dropdowns_sort_by_machine_number(qapp, fake_config, fake_project, frozen_project_date):
    _seed_audit(fake_project, "AUD-UI-SORT-010", **{"Press/Machine #": "Press 10", "Tool #": "PN-10"})
    _seed_audit(fake_project, "AUD-UI-SORT-002", **{"Press/Machine #": "Press 2", "Tool #": "PN-02"})
    _seed_audit(fake_project, "AUD-UI-SORT-026", **{"Press/Machine #": "26, 70", "Tool #": "PN-26"})
    _seed_audit(fake_project, "AUD-UI-SORT-WEIRD", **{"Press/Machine #": "Bench A", "Tool #": "PN-WEIRD"})
    page = AuditPage(fake_config)
    page.show()
    page._load_lazy_audit_indexes()

    audit_ids = [page.load_audit_id_combo.itemData(index) for index in range(page.load_audit_id_combo.count())]
    source_ids = [page.compatibility_source_combo.itemData(index) for index in range(page.compatibility_source_combo.count())]

    assert audit_ids.index("AUD-UI-SORT-002") < audit_ids.index("AUD-UI-SORT-010") < audit_ids.index("AUD-UI-SORT-026") < audit_ids.index("AUD-UI-SORT-WEIRD")
    assert source_ids.index("AUD-UI-SORT-002") < source_ids.index("AUD-UI-SORT-010") < source_ids.index("AUD-UI-SORT-026") < source_ids.index("AUD-UI-SORT-WEIRD")


def test_new_audit_defaults_quick_disconnect_fields(qapp, fake_config, fake_project, frozen_project_date):
    page = AuditPage(fake_config)
    page.show()

    assert page.audit_fields["Quick Disconnects Present?"].currentText() == "Yes"
    assert page.audit_fields["Pneumatic Quick Disconnect Type"].text() == "PTC"


def test_part_present_detection_yes_autofills_sensor_fields(qapp, fake_config):
    page = AuditPage(fake_config)
    page.show()

    _set_field(page, "Part-Present Detection Present?", "Yes")

    assert page.audit_fields["Sensor Type"].text() == "Reed Switch"
    assert page.audit_fields["Sensor Brand/Model"].text() == "SMC"


def test_part_present_detection_autofill_preserves_custom_sensor_values(qapp, fake_config):
    page = AuditPage(fake_config)
    page.show()

    _set_field(page, "Sensor Type", "Photoelectric")
    _set_field(page, "Sensor Brand/Model", "Keyence PX")
    _set_field(page, "Part-Present Detection Present?", "Yes")

    assert page.audit_fields["Sensor Type"].text() == "Photoelectric"
    assert page.audit_fields["Sensor Brand/Model"].text() == "Keyence PX"


def test_existing_audit_load_preserves_saved_quick_disconnect_fields(qapp, fake_config, fake_project, frozen_project_date):
    original = _seed_audit(
        fake_project,
        "AUD-QD-SAVED-001",
        **{
            "Quick Disconnects Present?": "No",
            "Pneumatic Quick Disconnect Type": "Barb",
        },
    )
    page = AuditPage(fake_config)

    page.load_audit_id_edit.setText(original["Audit ID"])
    click_button(page, "Load Existing Audit ID")

    assert page.audit_fields["Quick Disconnects Present?"].currentText() == "No"
    assert page.audit_fields["Pneumatic Quick Disconnect Type"].text() == ""
    assert page.audit_fields["Pneumatic Quick Disconnect Type"].isHidden() is True


def test_connection_type_smart_defaults_changeover_difficulty_when_unset(qapp, fake_config, fake_project, frozen_project_date):
    page = AuditPage(fake_config)
    page.show()

    _set_field(page, "Connection Type", "ATI")
    assert page.audit_fields["Changeover Difficulty"].currentText() == "Low"

    second_page = AuditPage(fake_config)
    second_page.show()
    _set_field(second_page, "Connection Type", "DoveTail")
    assert second_page.audit_fields["Changeover Difficulty"].currentText() == "Medium"


def test_user_modified_changeover_difficulty_is_not_overwritten(qapp, fake_config, fake_project, frozen_project_date):
    page = AuditPage(fake_config)
    page.show()

    _set_field(page, "Changeover Difficulty", "High")
    _set_field(page, "Connection Type", "ATI")
    _set_field(page, "Connection Type", "DoveTail")

    assert page.audit_fields["Changeover Difficulty"].currentText() == "High"


def test_existing_saved_changeover_difficulty_is_not_overwritten_on_load(qapp, fake_config, fake_project, frozen_project_date):
    original = _seed_audit(
        fake_project,
        "AUD-CHANGEOVER-SAVED-001",
        **{
            "Connection Type": "ATI",
            "Changeover Difficulty": "High",
        },
    )
    page = AuditPage(fake_config)

    page.load_audit_id_edit.setText(original["Audit ID"])
    click_button(page, "Load Existing Audit ID")

    assert page.audit_fields["Connection Type"].currentText() == "ATI"
    assert page.audit_fields["Changeover Difficulty"].currentText() == "High"


def test_generate_new_audit_id_does_not_load_first_audit_and_save_adds_row(qapp, fake_config, fake_project, frozen_project_date):
    first = _seed_audit(fake_project, "AUD-20260519-001", **{"Press/Machine #": "Press 1", "Notes": "First row must survive."})
    page = AuditPage(fake_config)
    page.show()

    click_button(page, "Generate New Audit ID")
    generated_id = page.audit_fields["Audit ID"].text()

    assert generated_id != first["Audit ID"]
    assert page.audit_fields["Press/Machine #"].text() == ""
    assert "Generated new audit ID" in page.result_panel.viewer.toPlainText()
    assert "Loaded audit entry" not in page.result_panel.viewer.toPlainText()

    _set_field(page, "Press/Machine #", "Press 200")
    _set_field(page, "Robot Type", "Manual robot")
    _set_field(page, "EOAT Type", "Vacuum")
    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()

    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
    assert next(row for row in rows if row["Audit ID"] == first["Audit ID"])["Notes"] == "First row must survive."
    assert any(row["Audit ID"] == generated_id and row["Press/Machine #"] == "Press 200" for row in rows)


def test_typing_machine_with_one_existing_audit_loads_that_audit(qapp, fake_config, fake_project, frozen_project_date):
    original = _seed_audit(fake_project, "AUD-MACHINE-026-001", **{"Press/Machine #": "26", "Tool #": "PN-26"})
    page = AuditPage(fake_config)

    page.audit_fields["Press/Machine #"].setText("26")
    _finish_machine_lookup(page)

    assert page.audit_fields["Audit ID"].text() == original["Audit ID"]
    assert page.audit_fields["Tool #"].text() == "PN-26"
    assert "Existing physical audit found for Machine 26" in page.result_panel.viewer.toPlainText()


def test_machine_lookup_uses_exact_normalized_machine_match(qapp, fake_config, fake_project, frozen_project_date):
    _seed_audit(fake_project, "AUD-MACHINE-011-001", **{"Press/Machine #": "11", "Tool #": "PN-11"})
    page = AuditPage(fake_config)

    page.audit_fields["Press/Machine #"].setText("1")
    _finish_machine_lookup(page)

    assert page.audit_fields["Audit ID"].text() != "AUD-MACHINE-011-001"
    assert page.audit_fields["Tool #"].text() == ""


def test_machine_lookup_multiple_existing_audits_requires_selection(qapp, fake_config, fake_project, frozen_project_date):
    first = _seed_audit(fake_project, "AUD-MACHINE-050-001", **{"Press/Machine #": "50", "Tool #": "PN-50A"})
    second = _seed_audit(fake_project, "AUD-MACHINE-050-002", **{"Press/Machine #": "50", "Tool #": "PN-50B"})
    page = AuditPage(fake_config)
    page.show()
    starting_id = page.audit_fields["Audit ID"].text()

    page.audit_fields["Press/Machine #"].setText("50")
    _finish_machine_lookup(page)

    assert page.audit_fields["Audit ID"].text() == starting_id
    assert page.machine_audit_match_combo.isHidden() is False
    audit_ids = [page.machine_audit_match_combo.itemData(index) for index in range(page.machine_audit_match_combo.count())]
    assert first["Audit ID"] in audit_ids
    assert second["Audit ID"] in audit_ids
    assert "Multiple existing physical audits found" in page.result_panel.viewer.toPlainText()


def test_empty_only_count_ignores_fields_filled_by_new_defaults(qapp, fake_config, fake_project, frozen_project_date):
    original = _seed_audit(
        fake_project,
        "AUD-EMPTY-DEFAULTS-001",
        **{
            "Connection Type": "ATI",
            "Changeover Difficulty": "Easy",
            "Quick Disconnects Present?": "Yes",
            "Pneumatic Quick Disconnect Type": "PTC",
            "Known Issues": "",
        },
    )
    page = AuditPage(fake_config)

    page.load_audit_id_edit.setText(original["Audit ID"])
    click_button(page, "Load Existing Audit ID")

    visible_empty_fields = _visible_empty_only_fields(page)
    assert "Known Issues" in visible_empty_fields
    assert "Changeover Difficulty" not in visible_empty_fields
    assert "Quick Disconnects Present?" not in visible_empty_fields
    assert "Pneumatic Quick Disconnect Type" not in visible_empty_fields
    assert _empty_only_count_message(page) == len(visible_empty_fields)


def test_audit_selector_refresh_does_not_load_first_audit(qapp, fake_config, fake_project, frozen_project_date):
    first = _seed_audit(fake_project, "AUD-SELECTOR-FIRST-001", **{"Press/Machine #": "1", "Tool #": "PN-FIRST"})
    page = AuditPage(fake_config)
    page.show()
    generated_id = page.audit_fields["Audit ID"].text()

    page.refresh_audit_selector()

    assert page.audit_fields["Audit ID"].text() == generated_id
    assert page.audit_fields["Tool #"].text() == ""
    assert page.load_audit_id_combo.findData(first["Audit ID"]) >= 0


def test_duplicate_audit_creates_new_unsaved_copy_and_saves_without_overwriting_original(qapp, fake_config, fake_project, frozen_project_date):
    original = _seed_audit(fake_project, "AUD-DUP-SOURCE-001")
    page = AuditPage(fake_config)
    page.show()

    page.load_audit_id_edit.setText(original["Audit ID"])
    click_button(page, "Load Existing Audit ID")
    click_button(page, "Duplicate Audit")

    duplicate_id = page.audit_fields["Audit ID"].text()
    assert duplicate_id != original["Audit ID"]
    assert page.audit_fields["Audit Date"].text() == "2026-05-19"
    assert page.audit_fields["Press/Machine #"].text() == original["Press/Machine #"]
    assert page.audit_fields["Tool #"].text() == original["Tool #"]
    assert page.audit_fields["Connection Type"].currentText() == original["Connection Type"]
    assert page.audit_fields["Cleanroom/Non-Cleanroom"].currentText() == original["Cleanroom/Non-Cleanroom"]
    assert page.audit_fields["Cup Type/Material"].text() == original["Cup Type/Material"]
    assert _field_value(page, "Gripper Model") == ""
    assert page.audit_fields["Gripper Size"].text() == ""
    assert page.audit_fields["Photo Folder/Link"].text() == ""
    assert "new unsaved" in page.lookup_note_label.text()

    _set_field(page, "Notes", "Duplicated audit notes.")
    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()

    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
    original_row = next(row for row in rows if row["Audit ID"] == original["Audit ID"])
    duplicate_row = next(row for row in rows if row["Audit ID"] == duplicate_id)
    assert original_row["Notes"] == original["Notes"]
    assert original_row["Photo Folder/Link"] == original["Photo Folder/Link"]
    assert duplicate_row["Notes"] == "Duplicated audit notes."
    assert duplicate_row["Connection Type"] == original["Connection Type"]
    assert duplicate_row["Cleanroom/Non-Cleanroom"] == original["Cleanroom/Non-Cleanroom"]
    assert duplicate_row["Cup Type/Material"] == original["Cup Type/Material"]
    assert duplicate_row["# of Grippers"] == "N/A"
    assert duplicate_row["Gripper Type"] == "N/A"
    assert duplicate_row["Gripper Model"] == "N/A"
    assert duplicate_row["Gripper Size"] == "N/A"
    assert duplicate_row["Photo Folder/Link"] == "N/A"


def test_load_and_duplicate_audit_with_na_values(qapp, fake_config, fake_project, frozen_project_date):
    original = _seed_audit(
        fake_project,
        "AUD-DUP-NA-001",
        **{
            "Connection Type": "",
            "Known Issues": "",
            "Gripper Model": "",
            "Gripper Size": "",
        },
    )
    page = AuditPage(fake_config)
    page.show()

    page.load_audit_id_edit.setText(original["Audit ID"])
    click_button(page, "Load Existing Audit ID")

    assert page.audit_fields["Connection Type"].currentText() == ""
    assert page.audit_fields["Known Issues"].toPlainText() == ""
    assert _field_value(page, "Gripper Model") == ""

    click_button(page, "Duplicate Audit")
    assert page.audit_fields["Connection Type"].currentText() == ""
    assert page.audit_fields["Known Issues"].toPlainText() == ""
    assert _field_value(page, "Gripper Model") == ""
    duplicate_id = page.audit_fields["Audit ID"].text()
    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()

    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
    duplicate_row = next(row for row in rows if row["Audit ID"] == duplicate_id)
    assert duplicate_row["Connection Type"] == "N/A"
    assert duplicate_row["Known Issues"] == "N/A"
    assert duplicate_row["# of Grippers"] == "N/A"
    assert duplicate_row["Gripper Type"] == "N/A"
    assert duplicate_row["Gripper Model"] == "N/A"
    assert duplicate_row["Gripper Size"] == "N/A"


def test_loaded_na_field_displays_blank_and_can_be_filled(qapp, fake_config, fake_project, frozen_project_date):
    original = _seed_audit(fake_project, "AUD-LOAD-NA-001", **{"Known Issues": "", "Gripper Model": "", "EOAT Moves": "Both"})
    page = AuditPage(fake_config)
    page.show()

    page.load_audit_id_edit.setText(original["Audit ID"])
    click_button(page, "Load Existing Audit ID")

    assert page.audit_fields["EOAT Moves"].currentText() == "Both"
    assert page.audit_fields["Known Issues"].toPlainText() == ""
    assert _field_value(page, "Gripper Model") == ""

    _set_field(page, "Known Issues", "Added after review.")
    _set_field(page, "Gripper Model", "Zimmer GPP")
    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()

    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
    updated_row = next(row for row in rows if row["Audit ID"] == original["Audit ID"])
    assert updated_row["Known Issues"] == "Added after review."
    assert updated_row["Gripper Model"] == "N/A"


def test_gripper_model_workbook_values_load_as_friendly_labels(qapp, fake_config, fake_project, frozen_project_date):
    for audit_id, workbook_value, expected_ui_value in [
        ("AUD-GRIPPER-LOAD-LARGE", "MHZL2-16D", "Large Double Gripper"),
        ("AUD-GRIPPER-LOAD-SMALL", "MHZL2-10S", "Small Double Gripper"),
        ("AUD-GRIPPER-LOAD-CUSTOM", "CUSTOM-GRIP-42", "CUSTOM-GRIP-42"),
        ("AUD-GRIPPER-LOAD-NA", "N/A", ""),
    ]:
        result = save_audit_entry(
            fake_project,
            {
                "Audit ID": audit_id,
                "Audit Date": "2026-05-18",
                "Auditor": "KG",
                "Plant/Area": "Plant 4",
                "Press/Machine #": f"Press {audit_id}",
                "Robot Type": "Wittmann R9",
                "EOAT Type": "Mechanical / Gripper",
                "# of Grippers": "2",
                "Gripper Type": "Double Pressure",
                "Gripper Model": workbook_value,
                "Status": "In Progress",
            },
        )
        assert result.success, result.errors

        page = AuditPage(fake_config)
        page.show()
        page.load_audit_id_edit.setText(audit_id)
        click_button(page, "Load Existing Audit ID")

        assert _field_value(page, "Gripper Model") == expected_ui_value


def test_vacuum_only_audit_loads_stale_gripper_values_as_blank(qapp, fake_config, fake_project, frozen_project_date):
    original = _seed_audit(fake_project, "AUD-VACUUM-STALE-GRIPPER-001")
    workbook_path = fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    row_number = next(row for row in range(2, ws.max_row + 1) if ws.cell(row=row, column=headers.index("Audit ID") + 1).value == original["Audit ID"])
    ws.cell(row=row_number, column=headers.index("# of Grippers") + 1).value = "2"
    ws.cell(row=row_number, column=headers.index("Gripper Type") + 1).value = "Double Pressure"
    ws.cell(row=row_number, column=headers.index("Gripper Model") + 1).value = "MHZL2-16D"
    workbook.save(workbook_path)
    workbook.close()

    page = AuditPage(fake_config)
    page.show()
    page.load_existing_audit(original["Audit ID"])

    assert _field_value(page, "# of Grippers") == ""
    assert _field_value(page, "Gripper Type") == ""
    assert _field_value(page, "Gripper Model") == ""


def test_load_existing_audit_defaults_to_empty_only_and_preserves_edits_when_switching_views(qapp, fake_config, fake_project, frozen_project_date):
    original = _seed_audit(
        fake_project,
        "AUD-EMPTY-ONLY-001",
        **{
            "Known Issues": "",
            "Sensor Type": "Vacuum switch",
            "Sensor Brand/Model": "SMC ZSE20",
            "Sensors Present?": "Yes",
            "Vacuum Confirmation Present?": "Yes",
            "Part-Present Detection Present?": "No",
            "Electrical Quick Disconnect Type": "M12",
            "Cable Management Condition": "OK",
        },
    )
    page = AuditPage(fake_config)
    page.show()

    page.load_audit_id_edit.setText(original["Audit ID"])
    click_button(page, "Load Existing Audit ID")

    assert page.audit_view_mode_combo.currentText() == "Empty Only"
    assert page.audit_fields["Known Issues"].isHidden() is False
    assert page.audit_fields["Known Issues"].toPlainText() == ""
    assert page.audit_fields["Press/Machine #"].isHidden() is True
    assert page.audit_fields["Sensor Type"].isHidden() is True

    _set_field(page, "Known Issues", "Filled while in empty-only mode.")
    page.audit_view_mode_combo.setCurrentText("Full Audit")

    assert page.audit_fields["Press/Machine #"].isHidden() is False
    assert page.audit_fields["Sensor Type"].isHidden() is False
    assert page.audit_fields["Known Issues"].toPlainText() == "Filled while in empty-only mode."

    page.audit_view_mode_combo.setCurrentText("Empty Only")

    assert page.audit_fields["Press/Machine #"].isHidden() is True
    assert page.audit_fields["Known Issues"].isHidden() is False
    assert page.audit_fields["Known Issues"].toPlainText() == "Filled while in empty-only mode."


def test_empty_only_respects_no_sensor_not_applicable_fields(qapp, fake_config, fake_project, frozen_project_date):
    original = _seed_audit(
        fake_project,
        "AUD-EMPTY-NO-SENSORS-001",
        **{
            "Sensors Present?": "No",
            "Sensor Type": "",
            "Sensor Brand/Model": "",
            "Vacuum Confirmation Present?": "",
            "Part-Present Detection Present?": "",
            "Electrical Quick Disconnect Type": "",
            "Cable Management Condition": "",
            "Known Issues": "",
        },
    )
    page = AuditPage(fake_config)
    page.show()

    page.load_audit_id_edit.setText(original["Audit ID"])
    click_button(page, "Load Existing Audit ID")

    assert page.audit_view_mode_combo.currentText() == "Empty Only"
    assert page.audit_fields["Known Issues"].isHidden() is False
    assert page.audit_fields["Sensor Type"].isHidden() is True
    assert page.audit_fields["Sensor Brand/Model"].isHidden() is True
    assert page.audit_fields["Electrical Quick Disconnect Type"].isHidden() is False
    assert page.audit_fields["Cable Management Condition"].isHidden() is False

    page.audit_view_mode_combo.setCurrentText("Full Audit")

    assert page.audit_fields["Known Issues"].isHidden() is False
    assert page.audit_fields["Sensor Type"].isHidden() is True
    assert page.audit_fields["Cable Management Condition"].isHidden() is False


def test_empty_only_info_count_matches_visible_fields_and_ignores_hidden_irrelevant_fields(qapp, fake_config, fake_project, frozen_project_date):
    original = _seed_audit(
        fake_project,
        "AUD-EMPTY-COUNT-001",
        **{
            "EOAT Type": "Mechanical / Gripper",
            "Sensors Present?": "No",
            "Known Issues": "",
            "Drop/Mis-Pick History": "",
            "Maintenance Frequency": "",
            "Sensor Type": "",
            "Sensor Brand/Model": "",
            "Electrical Quick Disconnect Type": "",
            "Cup Type/Material": "",
            "Cup Diameter/Size": "",
            "Number of Parts Picked": "",
        },
    )
    page = AuditPage(fake_config)
    page.show()

    page.load_audit_id_edit.setText(original["Audit ID"])
    click_button(page, "Load Existing Audit ID")

    visible_empty_fields = _visible_empty_only_fields(page)
    assert _empty_only_count_message(page) == len(visible_empty_fields)
    assert "Known Issues" in visible_empty_fields
    assert "Drop/Mis-Pick History" in visible_empty_fields
    assert "Maintenance Frequency" in visible_empty_fields
    assert "Sensor Type" not in visible_empty_fields
    assert "Electrical Quick Disconnect Type" in visible_empty_fields
    assert "Cup Type/Material" not in visible_empty_fields


def test_empty_only_treats_whitespace_and_na_as_empty(qapp, fake_config, fake_project, frozen_project_date):
    original = _seed_audit(fake_project, "AUD-EMPTY-WHITESPACE-001", **{"Known Issues": "Saved value."})
    workbook_path = fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    row_number = next(row for row in range(2, ws.max_row + 1) if ws.cell(row=row, column=headers.index("Audit ID") + 1).value == original["Audit ID"])
    ws.cell(row=row_number, column=headers.index("Known Issues") + 1).value = "   "
    ws.cell(row=row_number, column=headers.index("Drop/Mis-Pick History") + 1).value = "N/A"
    workbook.save(workbook_path)
    workbook.close()

    page = AuditPage(fake_config)
    page.show()
    page.load_audit_id_edit.setText(original["Audit ID"])
    click_button(page, "Load Existing Audit ID")

    visible_empty_fields = _visible_empty_only_fields(page)
    assert "Known Issues" in visible_empty_fields
    assert "Drop/Mis-Pick History" in visible_empty_fields
    assert _empty_only_count_message(page) == len(visible_empty_fields)


def test_loaded_blank_workbook_values_display_blank_and_save_na(qapp, fake_config, fake_project, frozen_project_date):
    original = _seed_audit(fake_project, "AUD-LOAD-BLANK-001")
    workbook_path = fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    row_number = next(row for row in range(2, ws.max_row + 1) if ws.cell(row=row, column=headers.index("Audit ID") + 1).value == original["Audit ID"])
    ws.cell(row=row_number, column=headers.index("Known Issues") + 1).value = None
    ws.cell(row=row_number, column=headers.index("Gripper Model") + 1).value = ""
    workbook.save(workbook_path)
    workbook.close()

    page = AuditPage(fake_config)
    page.show()
    page.load_audit_id_edit.setText(original["Audit ID"])
    click_button(page, "Load Existing Audit ID")

    assert page.audit_fields["EOAT Moves"].currentText() == ""
    assert page.audit_fields["Known Issues"].toPlainText() == ""
    assert _field_value(page, "Gripper Model") == ""

    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()

    rows = row_dicts(workbook_path, "EOAT Inventory")
    updated_row = next(row for row in rows if row["Audit ID"] == original["Audit ID"])
    assert updated_row["EOAT Moves"] in (None, "")
    assert updated_row["Known Issues"] == "N/A"
    assert updated_row["Gripper Model"] == "N/A"


def test_hidden_tooling_values_are_saved_when_switching_eoat_type(qapp, fake_config, fake_project, frozen_project_date):
    page = AuditPage(fake_config)
    page.show()

    _set_field(page, "Auditor", "Usability Tester")
    _set_field(page, "Plant/Area", "Plant 4")
    _set_field(page, "Press/Machine #", "Press 106")
    _set_field(page, "Robot Type", "Wittmann R9")
    _set_field(page, "EOAT Type", "Hybrid")
    _set_field(page, "Connection Type", "ATI")
    _set_field(page, "Cup Type/Material", "Nitrile")
    _set_field(page, "Cup Diameter/Size", "20 mm")
    _set_field(page, "# of Cups", "4")
    _set_field(page, "# of Grippers", "2")
    _set_field(page, "Gripper Type", "Double Pressure")
    _set_field(page, "Gripper Model", "Zimmer GPP")
    _set_field(page, "Gripper Size", "25 mm")
    _set_field(page, "EOAT Type", "Mechanical / Gripper")

    audit_id = page.audit_fields["Audit ID"].text()
    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()

    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
    row = next(row for row in rows if row["Audit ID"] == audit_id)
    assert row["EOAT Type"] == "Mechanical / Gripper"
    assert row["# of Cups"] == "N/A"
    assert row["Cup Type/Material"] == "N/A"
    assert row["Cup Diameter/Size"] == "N/A"
    assert row["# of Grippers"] == "2"
    assert row["Gripper Type"] == "Double Pressure"
    assert row["Gripper Model"] == "Zimmer GPP"
    assert row["Gripper Size"] == "25 mm"


def test_miscellaneous_audit_saves_and_loads_gripper_fields(qapp, fake_config, fake_project, frozen_project_date):
    page = AuditPage(fake_config)
    page.show()

    _set_field(page, "Press/Machine #", "Press 107")
    _set_field(page, "Robot Type", "Wittmann R9")
    _set_field(page, "EOAT Type", "Miscellaneous")
    _set_field(page, "Connection Type", "ATI")
    _set_field(page, "# of Grippers", "1")
    _set_field(page, "Gripper Type", "Single Pressure")
    _set_field(page, "Gripper Model", "Zimmer GPP")
    _set_field(page, "Gripper Size", "25 mm")

    audit_id = page.audit_fields["Audit ID"].text()
    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()

    loaded = AuditPage(fake_config)
    loaded.show()
    assert loaded.load_existing_audit(audit_id) is True

    assert loaded.audit_fields["EOAT Type"].currentText() == "Miscellaneous"
    assert loaded.audit_fields["# of Grippers"].text() == "1"
    assert loaded.audit_fields["Gripper Type"].currentText() == "Single Pressure"
    assert loaded.audit_fields["Gripper Model"].currentText() == "Zimmer GPP"
    loaded.audit_view_mode_combo.setCurrentText("Full Audit")
    assert loaded.audit_fields["Gripper Model"].isHidden() is False


def test_duplicate_audit_can_change_press_and_autofill_tool_from_new_part_number(qapp, fake_config, fake_project, frozen_project_date):
    create_press_reference_workbooks(fake_project / "reference-data")
    original = _seed_audit(fake_project, "AUD-DUP-AUTOFILL-001")
    page = AuditPage(fake_config)

    page.load_audit_id_edit.setText(original["Audit ID"])
    click_button(page, "Load Existing Audit ID")
    click_button(page, "Duplicate Audit")
    duplicate_id = page.audit_fields["Audit ID"].text()

    page.audit_fields["Press/Machine #"].setText("Press 70")
    _finish_machine_lookup(page)
    assert page.audit_fields["Press/Machine #"].text() == "70"
    assert page.audit_fields["Tool #"].text() == "DEMO-PN-0170"

    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()

    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
    duplicate_row = next(row for row in rows if row["Audit ID"] == duplicate_id)
    original_row = next(row for row in rows if row["Audit ID"] == original["Audit ID"])
    assert duplicate_row["Press/Machine #"] == "70"
    assert duplicate_row["Tool #"] == "DEMO-PN-0170"
    assert original_row["Press/Machine #"] == original["Press/Machine #"]
    assert original_row["Tool #"] == original["Tool #"]


def test_manual_audit_id_change_saves_new_row_and_existing_id_collision_is_blocked(qapp, fake_config, fake_project, frozen_project_date):
    original = _seed_audit(fake_project, "AUD-MANUAL-SOURCE-001", Notes="Keep original.")
    other = _seed_audit(fake_project, "AUD-MANUAL-EXISTING-002")
    page = AuditPage(fake_config)

    page.load_audit_id_edit.setText(original["Audit ID"])
    click_button(page, "Load Existing Audit ID")
    _set_field(page, "Audit ID", "AUD-MANUAL-NEW")
    _set_field(page, "Press/Machine #", "Press 200")
    _set_field(page, "Tool #", "PN-2000")
    _set_field(page, "Notes", "Manual save-as row.")
    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()

    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
    assert next(row for row in rows if row["Audit ID"] == original["Audit ID"])["Notes"] == "Keep original."
    manual_row = next(row for row in rows if row["Audit ID"] == "AUD-MANUAL-NEW")
    assert manual_row["Press/Machine #"] == "Press 200"
    assert manual_row["Tool #"] == "PN-2000"

    page.load_audit_id_edit.setText(original["Audit ID"])
    click_button(page, "Load Existing Audit ID")
    _set_field(page, "Audit ID", other["Audit ID"])
    _set_field(page, "Notes", "Should not overwrite existing row.")
    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()

    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
    other_row = next(row for row in rows if row["Audit ID"] == other["Audit ID"])
    assert other_row["Notes"] == other["Notes"]
    assert "Audit ID already exists" in page.result_panel.viewer.toPlainText()
    headers = set(rows[0].keys())
    assert "Setup ID" not in headers
    assert "Tool-Press Map" not in workbook_sheet_names(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx")


def test_pneumatic_circuits_tab_groups_fields_defaults_and_tag_buttons(qapp, fake_config):
    page = AuditPage(fake_config)
    page.show()

    tab_titles = [page.audit_section_tabs.tabText(index) for index in range(page.audit_section_tabs.count())]
    assert "Pneumatic Circuits" in tab_titles
    group_titles = {group.title() for group in page.findChildren(QGroupBox)}
    assert {"EOAT Side", "Robot Side"}.issubset(group_titles)
    for field in [
        "EOAT Vacuum Circuits",
        "EOAT Pressure Circuits",
        "EOAT Interchangeable Circuits",
        "Robot Vacuum Circuits",
        "Robot Pressure Circuits",
        "Robot Interchangeable Circuits",
    ]:
        assert field in page.audit_fields
        assert field in page._field_tag_buttons
    assert page.audit_fields["EOAT Interchangeable Circuits"].text() == "0"
    assert page.audit_fields["Robot Interchangeable Circuits"].text() == "0"


def test_audit_form_uses_grouped_panels_without_losing_fields(qapp, fake_config):
    page = AuditPage(fake_config)
    page.show()

    tab_titles = [page.audit_section_tabs.tabText(index) for index in range(page.audit_section_tabs.count())]
    assert tab_titles == list(AUDIT_SECTIONS)

    expected_fields = [field for fields in AUDIT_SECTIONS.values() for field in fields]
    grouped_fields = [field for groups in AUDIT_SECTION_GROUPS.values() for _group_title, fields in groups for field in fields]
    assert Counter(expected_fields) == Counter(grouped_fields)
    assert set(page.audit_fields) == set(expected_fields)

    group_titles = {group.title() for group in page.findChildren(QGroupBox)}
    expected_group_titles = {group_title for groups in AUDIT_SECTION_GROUPS.values() for group_title, _fields in groups}
    assert expected_group_titles.issubset(group_titles)

    for field in expected_fields:
        assert field in page._field_tag_buttons
        assert field in page._audit_field_rows
        assert field in page._audit_field_group_keys


def test_audit_output_panel_has_readable_splitter_and_controls(qapp, fake_config):
    page = AuditPage(fake_config)
    page.show()

    assert page.audit_output_splitter.orientation() == Qt.Orientation.Vertical
    assert page.result_panel.minimumHeight() >= 180
    assert page.result_panel.viewer.minimumHeight() >= 160

    buttons = {button.text() for button in page.findChildren(QPushButton)}
    assert {"Expand Output", "Collapse Output", "Copy Output", "Clear Output"}.issubset(buttons)


def test_save_audit_workflow_creates_robot_info_and_summary(qapp, fake_config, fake_project, frozen_project_date):
    page = AuditPage(fake_config)
    page.show()
    audit_id = page.audit_fields["Audit ID"].text()
    values = {
        "Press/Machine #": "Press 212",
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "EOAT Vacuum Circuits": "2",
        "EOAT Pressure Circuits": "1",
        "EOAT Interchangeable Circuits": "0",
        "Robot Vacuum Circuits": "3",
        "Robot Pressure Circuits": "2",
        "Robot Interchangeable Circuits": "0",
    }
    for field, value in values.items():
        _set_field(page, field, value)

    result = page._save_audit_workflow(
        {field: page._field_value(widget) for field, widget in page.audit_fields.items()},
        allow_update=False,
        create_followup_action=False,
    )

    assert result.success, result.errors
    assert "Robot Info Summary" in result.summary
    assert "Updated Robot_Info.xlsx for Machine Press 212" in result.summary
    assert robot_info_workbook_path(fake_project).exists()
    robot_info = load_robot_info_for_audit_entry(
        fake_project,
        {
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 212",
            "Robot Type": "Wittmann R9",
        },
    )
    assert robot_info is not None
    assert robot_info["Robot Vacuum Circuits"] == 3
    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
    saved_row = next(row for row in rows if row["Audit ID"] == audit_id)
    assert saved_row["EOAT Vacuum Circuits"] == "2"
    assert "Robot Vacuum Circuits" not in saved_row


def test_save_audit_workflow_skips_robot_info_when_circuit_values_are_default(qapp, fake_config, fake_project, monkeypatch, frozen_project_date):
    page = AuditPage(fake_config)
    page.show()
    _set_field(page, "Press/Machine #", "Press 213")
    _set_field(page, "Robot Type", "Wittmann R9")
    _set_field(page, "EOAT Type", "Vacuum")

    def fail_upsert(*_args, **_kwargs):
        raise AssertionError("Robot_Info.xlsx should not be opened for write when only default robot circuit values exist.")

    monkeypatch.setattr("app.pages.audit_save_workflow.upsert_robot_info_from_audit", fail_upsert)
    result = page._save_audit_workflow(
        {field: page._field_value(widget) for field, widget in page.audit_fields.items()},
        allow_update=False,
        create_followup_action=False,
    )

    assert result.success, result.errors
    assert result.metrics["robot_info_save_skipped"] is True
    assert "Robot Info skipped" in result.summary
    assert not robot_info_workbook_path(fake_project).exists()


def _patch_manual_override_confirmation(monkeypatch, button_text: str) -> None:
    clicked = {}

    def fake_exec(box):
        clicked["button"] = next(button for button in box.buttons() if button.text() == button_text)
        return 0

    def fake_clicked_button(_box):
        return clicked["button"]

    monkeypatch.setattr("app.pages.audit.QMessageBox.exec", fake_exec)
    monkeypatch.setattr("app.pages.audit.QMessageBox.clickedButton", fake_clicked_button)


def test_manual_completion_override_cancel_does_not_mark_complete(qapp, fake_config, fake_project, monkeypatch):
    audit_id = "AUD-MANUAL-OVERRIDE-CANCEL"
    _seed_audit(
        fake_project,
        audit_id,
        Status="In Progress",
        **{"Sensor Brand/Model": "", "Photos Taken?": "Yes", "Photo Folder/Link": ""},
    )
    page = AuditPage(fake_config)
    page.show()
    assert page.load_existing_audit(audit_id) is True
    _patch_manual_override_confirmation(monkeypatch, "Cancel")

    click_button(page, "Manual Override: Mark Audit Complete")

    assert page.audit_coach_panel.summary.manual_completion_override is False
    row = next(row for row in row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory") if row["Audit ID"] == audit_id)
    assert row.get(MANUAL_COMPLETION_OVERRIDE_FIELD) != "Yes"


def test_manual_completion_override_saves_persists_and_stays_audit_specific(qapp, fake_config, fake_project, monkeypatch):
    audit_id = "AUD-MANUAL-OVERRIDE-YES"
    other_id = "AUD-MANUAL-OVERRIDE-OTHER"
    _seed_audit(
        fake_project,
        audit_id,
        Status="In Progress",
        **{"Sensor Brand/Model": "", "Photos Taken?": "Yes", "Photo Folder/Link": ""},
    )
    _seed_audit(fake_project, other_id, Status="In Progress")
    page = AuditPage(fake_config)
    page.show()
    assert page.load_existing_audit(audit_id) is True
    _patch_manual_override_confirmation(monkeypatch, "Mark Complete")

    click_button(page, "Manual Override: Mark Audit Complete")
    wait_for_background_tasks()

    assert page.audit_coach_panel.summary.percent_complete == 100
    assert page.audit_coach_panel.summary.manual_completion_override is True
    assert "Manual completion override applied" in page.manual_override_status_label.text()
    assert page.has_unsaved_changes() is False
    rows = row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
    row = next(row for row in rows if row["Audit ID"] == audit_id)
    other = next(row for row in rows if row["Audit ID"] == other_id)
    assert row[MANUAL_COMPLETION_OVERRIDE_FIELD] == "Yes"
    assert row[MANUAL_COMPLETION_OVERRIDE_TIMESTAMP_FIELD]
    assert row[MANUAL_COMPLETION_OVERRIDE_USER_FIELD] == "Original Auditor"
    assert "Photo Folder/Link" in row[IGNORED_EMPTY_FIELDS_AT_OVERRIDE_FIELD]
    assert other.get(MANUAL_COMPLETION_OVERRIDE_FIELD) != "Yes"

    loaded = AuditPage(fake_config)
    loaded.show()
    assert loaded.load_existing_audit(audit_id) is True
    assert loaded.audit_coach_panel.summary.percent_complete == 100
    assert loaded.audit_coach_panel.summary.manual_completion_override is True

    _set_field(loaded, "Notes", "Edited after override.")
    click_button(loaded, "Save Audit Entry")
    wait_for_background_tasks()
    reloaded_row = next(
        row
        for row in row_dicts(fake_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx", "EOAT Inventory")
        if row["Audit ID"] == audit_id
    )
    assert reloaded_row[MANUAL_COMPLETION_OVERRIDE_FIELD] == "Yes"
    assert reloaded_row[MANUAL_COMPLETION_OVERRIDE_TIMESTAMP_FIELD] == row[MANUAL_COMPLETION_OVERRIDE_TIMESTAMP_FIELD]
    assert reloaded_row[IGNORED_EMPTY_FIELDS_AT_OVERRIDE_FIELD] == row[IGNORED_EMPTY_FIELDS_AT_OVERRIDE_FIELD]
    assert loaded.has_unsaved_changes() is False


def test_save_audit_workflow_skips_heavy_derived_systems(qapp, fake_config, monkeypatch, frozen_project_date):
    page = AuditPage(fake_config)
    page.show()
    _set_field(page, "Press/Machine #", "Press 214")
    _set_field(page, "Robot Type", "Wittmann R9")
    _set_field(page, "EOAT Type", "Vacuum")

    def fail_heavy(*_args, **_kwargs):
        raise AssertionError("Normal audit save should not run heavy derived aggregation or scans.")

    monkeypatch.setattr("core.audit_entries.refresh_audit_by_press_view", fail_heavy)
    monkeypatch.setattr("core.open_items.list_open_items", fail_heavy)
    monkeypatch.setattr("core.standards_compliance.analyze_standards_compliance", fail_heavy)
    monkeypatch.setattr("core.fmea_suggestions.build_fmea_suggestions", fail_heavy)
    monkeypatch.setattr("core.photo_evidence.evidence_coverage_for_project", fail_heavy)
    monkeypatch.setattr("core.search.search_project", fail_heavy)
    monkeypatch.setattr("core.final_handoff_readiness.build_final_handoff_readiness", fail_heavy)

    result = page._save_audit_workflow(
        {field: page._field_value(widget) for field, widget in page.audit_fields.items()},
        allow_update=False,
        create_followup_action=False,
    )

    assert result.success, result.errors
    assert result.metrics["audit_by_press_refreshed"] is False
    assert result.metrics["linked_compatibility_sync_requested"] is False
    assert result.metrics["compatibility_autorun_skipped"] is True


def test_load_existing_audit_restores_eoat_and_robot_pneumatic_fields(qapp, fake_config, fake_project):
    entry = _seed_audit(
        fake_project,
        "AUD-PNEUMATIC-LOAD-001",
        **{
            "Press/Machine #": "Press 88",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "EOAT Vacuum Circuits": "2",
            "EOAT Pressure Circuits": "1",
            "EOAT Interchangeable Circuits": "0",
        },
    )
    assert upsert_robot_info_from_audit(
        fake_project,
        {
            **entry,
            "Robot Vacuum Circuits": "4",
            "Robot Pressure Circuits": "3",
            "Robot Interchangeable Circuits": "0",
        },
    ).success
    page = AuditPage(fake_config)
    page.show()

    page.load_existing_audit("AUD-PNEUMATIC-LOAD-001")

    assert page.audit_fields["EOAT Vacuum Circuits"].text() == "2"
    assert page.audit_fields["EOAT Pressure Circuits"].text() == "1"
    assert page.audit_fields["EOAT Interchangeable Circuits"].text() == "0"
    assert page.audit_fields["Robot Vacuum Circuits"].text() == "4"
    assert page.audit_fields["Robot Pressure Circuits"].text() == "3"
    assert page.audit_fields["Robot Interchangeable Circuits"].text() == "0"


def test_pneumatic_target_navigation_selects_pneumatic_tab(qapp, fake_config):
    page = AuditPage(fake_config)
    page.show()

    page.focus_annotation_target(
        {
            "target_type": "audit_field",
            "audit_id": "AUD-NAV-PNEUMATIC-001",
            "field_label": "Robot Pressure Circuits",
            "field_key": "Robot Pressure Circuits",
        }
    )

    current_title = page.audit_section_tabs.tabText(page.audit_section_tabs.currentIndex())
    assert current_title == "Pneumatic Circuits"
    assert "Target field: Robot Pressure Circuits" in page.result_panel.viewer.toPlainText()

