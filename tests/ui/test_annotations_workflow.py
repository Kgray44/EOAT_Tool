from __future__ import annotations

import pytest
from openpyxl import load_workbook
from PySide6.QtWidgets import QDialog, QMessageBox, QPushButton

from app.dashboard_ui import DashboardWindow
from app.pages.audit import AuditPage, audit_section_for_field
from app.pages.notes import NotesPage
from app.pages.tags import TagsPage
from app.widgets import annotation_target_navigator
from app.widgets.field_tag_button import FieldNoteDialog, FieldTagButton, FieldTagDialog
from app.widgets.annotation_target_navigator import AnnotationTargetPickerDialog
from app.widgets.note_editor import NoteEditor
from core.annotations.service import AnnotationService
from core.audit_entries import save_audit_entry
from core.paths import resolve_project_paths
from tests.ui.helpers import click_button, wait_for_background_tasks


pytestmark = pytest.mark.usability


def test_notes_page_opens_and_creates_note(qapp, fake_config):
    page = NotesPage(fake_config)
    page.show()

    page.editor.subject_edit.setText("Audit observation")
    page.editor.body_edit.setPlainText("- Check sensor bracket")
    page.editor.importance_combo.setCurrentText("Important")
    click_button(page, "Save Note")

    assert page.table.rowCount() == 1
    assert page.table.item(0, 0).text() == "Audit observation"


def test_tags_page_opens_and_creates_custom_tag(qapp, fake_config):
    page = TagsPage(fake_config)
    page.show()

    names = [page.tag_table.item(row, 0).text() for row in range(page.tag_table.rowCount())]
    assert "Info" in names

    click_button(page, "+ New Tag")
    page.name_edit.setText("Fixture Trial")
    page.color_combo.setCurrentText("Pink")
    click_button(page, "Save Tag")

    names = [page.tag_table.item(row, 0).text() for row in range(page.tag_table.rowCount())]
    assert "Fixture Trial" in names


def test_audit_page_has_tiny_field_tag_buttons_and_indicators(qapp, fake_config):
    page = AuditPage(fake_config)
    page.show()

    buttons = page.findChildren(FieldTagButton)
    assert buttons
    assert all(button.text() == "" for button in buttons)

    service = AnnotationService(fake_config.project_root)
    target = page._create_or_get_field_tag_target("Sensor Type")
    tag = service.get_tag_by_name("Needs Review")
    service.assign_tag_to_target(tag.id, target.id, sync_workbook=False)
    page._refresh_field_tag_indicators()

    assert "Needs Review" in page._field_tag_buttons["Sensor Type"].toolTip()


def test_field_tag_dialog_persists_assignment_and_tags_page_returns_target(qapp, fake_config, fake_project):
    save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-FIELD-TAG-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "EOAT Vacuum Circuits": "2",
            "Status": "In Progress",
        },
    )
    page = AuditPage(fake_config)
    page.load_existing_audit("AUD-FIELD-TAG-001")
    target = page._create_or_get_field_tag_target("EOAT Vacuum Circuits")
    dialog = FieldTagDialog(page.annotation_service, target, field_label="EOAT Vacuum Circuits", current_value="2")
    dialog.show_add_tag_controls()
    dialog.tag_combo.setCurrentIndex(dialog.tag_combo.findText("Needs Review"))
    dialog.comment_edit.setPlainText("Check circuit split.")

    dialog.add_tag()
    service = AnnotationService(fake_config.project_root)
    assert service.list_tag_assignments(audit_id="AUD-FIELD-TAG-001", target_type="audit_field") == []
    assert dialog.primary_button.text() == "Save"
    assert dialog.commit_changes()
    page._refresh_field_tag_indicators()

    assignments = service.list_tag_assignments(audit_id="AUD-FIELD-TAG-001", target_type="audit_field")
    assert len(assignments) == 1
    assert assignments[0]["tag_name"] == "Needs Review"
    assert assignments[0]["field_label"] == "EOAT Vacuum Circuits"
    assert assignments[0]["comment"] == "Check circuit split."

    tags_page = TagsPage(fake_config)
    assert tags_page.assignment_table.rowCount() == 1
    row_values = [tags_page.assignment_table.item(0, col).text() for col in range(tags_page.assignment_table.columnCount())]
    assert "Needs Review" in row_values
    assert "audit_field" in row_values
    assert "AUD-FIELD-TAG-001" in row_values
    assert "EOAT Vacuum Circuits" in row_values
    assert _field_fill(fake_project, "AUD-FIELD-TAG-001", "EOAT Vacuum Circuits") == "00FACC15"


def test_field_note_dialog_stages_linked_note_and_notes_page_returns_after_commit(qapp, fake_config, fake_project, monkeypatch):
    service = AnnotationService(fake_config.project_root)
    target = service.create_or_get_target(
        "audit_field",
        audit_id="AUD-FIELD-NOTE-001",
        machine_id="12",
        field_key="EOAT Vacuum Circuits",
        field_label="EOAT Vacuum Circuits",
        sheet_name="EOAT Inventory",
        header_name="EOAT Vacuum Circuits",
    )

    def save_note_editor(note_dialog):
        note_dialog.subject_edit.setText("Review vacuum circuit split")
        note_dialog.body_edit.setPlainText("Confirm whether the zones are independently valved.")
        note_dialog.importance_combo.setCurrentText("Important")
        note_dialog.save_note()
        return QDialog.DialogCode.Accepted

    monkeypatch.setattr(FieldNoteDialog, "exec", save_note_editor)
    dialog = FieldTagDialog(service, target, field_label="EOAT Vacuum Circuits", current_value="2")
    dialog.create_note()

    assert service.search_notes("vacuum circuit") == []
    assert dialog.primary_button.text() == "Done"
    assert dialog.commit_changes()

    notes = service.search_notes("vacuum circuit")
    assert len(notes) == 1
    assert notes[0]["subject"] == "Review vacuum circuit split"
    assert notes[0]["targets"][0]["id"] == target.id

    notes_page = NotesPage(fake_config)
    assert notes_page.table.rowCount() == 1
    assert notes_page.table.item(0, 0).text() == "Review vacuum circuit split"
    assert "1 target" in notes_page.table.item(0, 6).text()


def test_field_tag_popup_uses_read_only_tag_color_preview(qapp, fake_config):
    service = AnnotationService(fake_config.project_root)
    target = service.create_or_get_target("audit_field", audit_id="AUD-COLOR-PREVIEW", field_key="EOAT Vacuum Circuits", field_label="EOAT Vacuum Circuits")
    dialog = FieldTagDialog(service, target, field_label="EOAT Vacuum Circuits", current_value="")

    assert not hasattr(dialog, "color_combo")
    assert dialog.add_tag_panel.isHidden()
    assert dialog.tag_combo.findText("Info") >= 0
    index = dialog.tag_combo.findText("Needs Review")
    dialog.tag_combo.setCurrentIndex(index)

    assert dialog.color_preview_label.text() == "Yellow"


def test_field_annotation_icon_is_calm_and_active_for_tags_or_notes(qapp, fake_config):
    page = AuditPage(fake_config)
    page.show()
    button = page._field_tag_buttons["EOAT Vacuum Circuits"]

    assert button.property("annotation_icon") == "flag"
    assert button.property("annotation_active") is False
    assert "warning" not in button.toolTip().lower()
    assert "QToolButton:hover" in button.styleSheet()
    assert "QToolButton:pressed" in button.styleSheet()

    target = page._create_or_get_field_tag_target("EOAT Vacuum Circuits")
    service = AnnotationService(fake_config.project_root)
    service.create_note("Note for EOAT Vacuum Circuits", "Body", "Neutral", target_ids=[target.id])
    page._refresh_field_tag_indicators()

    assert button.property("annotation_active") is True
    assert "Note for EOAT Vacuum Circuits" in button.toolTip()


def test_field_tag_popup_existing_rows_stage_delete_and_cancel_truthfully(qapp, fake_config, monkeypatch):
    service = AnnotationService(fake_config.project_root)
    target = service.create_or_get_target("audit_field", audit_id="AUD-TAG-ROW", field_key="EOAT Vacuum Circuits", field_label="EOAT Vacuum Circuits")
    tag = service.get_tag_by_name("Documentation Gap")
    service.assign_tag_to_target(tag.id, target.id, comment="Ask Jake", sync_workbook=False)

    dialog = FieldTagDialog(service, target, field_label="EOAT Vacuum Circuits", current_value="2")
    assert dialog.existing_list.item(0).text() == "Documentation Gap | Ask Jake"

    monkeypatch.setattr(dialog, "_confirm_remove_tag", lambda: True)
    dialog.existing_list.setCurrentRow(0)
    dialog.remove_selected_tag()
    assert len(service.list_tag_assignments(audit_id="AUD-TAG-ROW", target_type="audit_field")) == 1
    assert dialog.commit_changes()
    assert service.list_tag_assignments(audit_id="AUD-TAG-ROW", target_type="audit_field") == []
    assert service.get_tag(tag.id).name == "Documentation Gap"

    dialog = FieldTagDialog(service, target, field_label="EOAT Vacuum Circuits", current_value="2")
    dialog.show_add_tag_controls()
    dialog.tag_combo.setCurrentIndex(dialog.tag_combo.findText("Needs Review"))
    dialog.comment_edit.setPlainText("Staged only")
    dialog.add_tag()
    dialog.reject()
    assert service.list_tag_assignments(audit_id="AUD-TAG-ROW", target_type="audit_field") == []


def test_field_tag_popup_edit_tag_updates_assignment_only_after_commit(qapp, fake_config):
    service = AnnotationService(fake_config.project_root)
    target = service.create_or_get_target("audit_field", audit_id="AUD-EDIT-TAG", field_key="Gripper Model", field_label="Gripper Model")
    tag = service.get_tag_by_name("Documentation Gap")
    service.assign_tag_to_target(tag.id, target.id, comment="Ask Jake", sync_workbook=False)

    dialog = FieldTagDialog(service, target, field_label="Gripper Model", current_value="")
    assert dialog.edit_tag_button.text() == "Edit Tag"
    dialog.existing_list.setCurrentRow(0)
    dialog.show_edit_tag_controls()
    assert dialog.color_preview_label.text() == "Orange"
    dialog.comment_edit.setPlainText("Ask Jake about gripper model number.")
    dialog.save_tag_editor()

    assert dialog.existing_list.item(0).text() == "Documentation Gap | Ask Jake about gripper model number."
    assert service.list_tag_assignments(audit_id="AUD-EDIT-TAG", target_type="audit_field")[0]["comment"] == "Ask Jake"
    assert dialog.commit_changes()

    assignment = service.list_tag_assignments(audit_id="AUD-EDIT-TAG", target_type="audit_field")[0]
    assert assignment["tag_name"] == "Documentation Gap"
    assert assignment["comment"] == "Ask Jake about gripper model number."
    assert service.get_tag(tag.id).name == "Documentation Gap"
    assert service.get_tag(tag.id).color_key == "orange"


def test_field_tag_popup_edit_tag_can_change_assignment_tag_and_cancel(qapp, fake_config):
    service = AnnotationService(fake_config.project_root)
    target = service.create_or_get_target("audit_field", audit_id="AUD-EDIT-CHANGE", field_key="Gripper Model", field_label="Gripper Model")
    original = service.get_tag_by_name("Documentation Gap")
    replacement = service.get_tag_by_name("Maintenance Concern")
    service.assign_tag_to_target(original.id, target.id, comment="Ask Jake", sync_workbook=False)

    dialog = FieldTagDialog(service, target, field_label="Gripper Model", current_value="")
    dialog.existing_list.setCurrentRow(0)
    dialog.show_edit_tag_controls()
    dialog.tag_combo.setCurrentIndex(dialog.tag_combo.findText("Maintenance Concern"))
    dialog.comment_edit.setPlainText("Ask Jake about pneumatic routing")
    dialog.save_tag_editor()
    dialog.reject()

    assignment = service.list_tag_assignments(audit_id="AUD-EDIT-CHANGE", target_type="audit_field")[0]
    assert assignment["tag_name"] == "Documentation Gap"
    assert assignment["comment"] == "Ask Jake"

    dialog = FieldTagDialog(service, target, field_label="Gripper Model", current_value="")
    dialog.existing_list.setCurrentRow(0)
    dialog.show_edit_tag_controls()
    dialog.tag_combo.setCurrentIndex(dialog.tag_combo.findText("Maintenance Concern"))
    dialog.comment_edit.setPlainText("Ask Jake about pneumatic routing")
    dialog.save_tag_editor()
    assert dialog.commit_changes()

    assignments = service.list_tag_assignments(audit_id="AUD-EDIT-CHANGE", target_type="audit_field")
    assert len(assignments) == 1
    assert assignments[0]["tag_id"] == replacement.id
    assert assignments[0]["tag_name"] == "Maintenance Concern"
    assert assignments[0]["comment"] == "Ask Jake about pneumatic routing"
    assert service.get_tag(original.id).name == "Documentation Gap"


def test_field_tag_row_blank_comment_uses_fallback(qapp, fake_config):
    service = AnnotationService(fake_config.project_root)
    target = service.create_or_get_target("audit_field", audit_id="AUD-NO-COMMENT", field_key="EOAT Vacuum Circuits", field_label="EOAT Vacuum Circuits")
    tag = service.get_tag_by_name("Needs Review")
    service.assign_tag_to_target(tag.id, target.id, sync_workbook=False)

    dialog = FieldTagDialog(service, target, field_label="EOAT Vacuum Circuits", current_value="")

    assert dialog.existing_list.item(0).text() == "Needs Review | No comment"


def test_field_note_cancel_discards_staged_note(qapp, fake_config, monkeypatch):
    service = AnnotationService(fake_config.project_root)
    target = service.create_or_get_target("audit_field", audit_id="AUD-NOTE-CANCEL", field_key="EOAT Vacuum Circuits", field_label="EOAT Vacuum Circuits")

    def save_note_editor(note_dialog):
        note_dialog.subject_edit.setText("Temporary field note")
        note_dialog.body_edit.setPlainText("Discard me.")
        note_dialog.save_note()
        return QDialog.DialogCode.Accepted

    monkeypatch.setattr(FieldNoteDialog, "exec", save_note_editor)
    dialog = FieldTagDialog(service, target, field_label="EOAT Vacuum Circuits", current_value="")
    dialog.create_note()
    dialog.reject()

    assert service.search_notes("Temporary field note") == []


def test_notes_add_field_options_include_expanded_annotation_links():
    labels = {label for label, _key in NoteEditor.OPTIONAL_FIELDS}

    for label in [
        "Status",
        "Collection / Folder",
        "Note Type",
        "Follow-Up Date",
        "Linked Audit ID",
        "Linked Machine",
        "Linked EOAT / Tool",
        "Linked Audit Field",
        "Linked Compatibility Entry",
        "Linked Photo / Attachment",
        "Linked Workbook Health Warning",
        "Linked Pilot Candidate",
        "Related Tags",
        "Created By",
        "Due / Review Date",
        "Priority Reason",
        "Source / Evidence",
        "Assigned To / Owner",
        "Resolution Notes",
        "Related Report",
        "Related PM Checklist Item",
        "Related FMEA Item",
        "Related Issue",
        "Related Standard / Guideline",
        "Related Spare Part / BOM Item",
    ]:
        assert label in labels


def test_navigation_to_notes_and_tags_refreshes_after_external_annotation_change(qapp, fake_config):
    window = DashboardWindow(fake_config)
    window._show_page("notes")
    notes_page = window.pages["notes"]
    assert notes_page.table.rowCount() == 0
    window._show_page("tags")
    tags_page = window.pages["tags"]
    assert tags_page.assignment_table.rowCount() == 0

    service = AnnotationService(fake_config.project_root)
    target = service.create_or_get_target("audit_field", audit_id="AUD-NAV-001", field_key="EOAT Vacuum Circuits", field_label="EOAT Vacuum Circuits")
    tag = service.get_tag_by_name("Needs Review")
    service.assign_tag_to_target(tag.id, target.id, sync_workbook=False)
    service.create_note("Navigation refresh note", "Body", "Neutral", target_ids=[target.id])

    window._show_page("notes")
    assert notes_page.table.rowCount() == 1
    window._show_page("tags")
    assert tags_page.assignment_table.rowCount() == 1


def test_tags_page_go_to_target_loads_audit_field(qapp, fake_config, fake_project):
    save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-GO-TARGET-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "EOAT Vacuum Circuits": "2",
            "Status": "In Progress",
        },
    )
    service = AnnotationService(fake_config.project_root)
    target = service.create_or_get_target("audit_field", audit_id="AUD-GO-TARGET-001", machine_id="12", field_key="EOAT Vacuum Circuits", field_label="EOAT Vacuum Circuits")
    tag = service.get_tag_by_name("Documentation Gap")
    service.assign_tag_to_target(tag.id, target.id, comment="What do i put here for this type? took picture.", sync_workbook=False)

    window = DashboardWindow(fake_config)
    window._show_page("tags")
    tags_page = window.pages["tags"]
    tags_page.assignment_table.selectRow(0)
    tags_page.go_to_target()

    audit_page = window.pages["audit"]
    assert audit_page.audit_fields["Audit ID"].text() == "AUD-GO-TARGET-001"
    assert "Target field: EOAT Vacuum Circuits" in audit_page.result_panel.viewer.toPlainText()
    assert tags_page.assignment_table.wordWrap() is True
    assert tags_page.assignment_table.columnWidth(7) <= 300
    assert tags_page.assignment_table.item(0, 7).toolTip() == "What do i put here for this type? took picture."


def test_go_to_target_switches_to_field_section_and_preserves_tag_color(qapp, fake_config, fake_project):
    save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-GRIPPER-TARGET-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Mechanical / Gripper",
            "Gripper Model": "GM-1",
            "Status": "In Progress",
        },
    )
    service = AnnotationService(fake_config.project_root)
    target = service.create_or_get_target("audit_field", audit_id="AUD-GRIPPER-TARGET-001", machine_id="12", field_key="Gripper Model", field_label="Gripper Model")
    tag = service.get_tag_by_name("Documentation Gap")
    service.assign_tag_to_target(tag.id, target.id, comment="Ask Jake", sync_workbook=False)

    window = DashboardWindow(fake_config)
    window._show_page("tags")
    tags_page = window.pages["tags"]
    tags_page.assignment_table.selectRow(0)
    tags_page.go_to_target()

    audit_page = window.pages["audit"]
    assert audit_section_for_field("Gripper Model") == "EOAT Type and Tooling"
    assert audit_page.audit_section_tabs.tabText(audit_page.audit_section_tabs.currentIndex()) == "EOAT Type and Tooling"
    field_style = audit_page.audit_fields["Gripper Model"].styleSheet()
    assert "#f97316" in field_style
    assert "#eff6ff" not in field_style
    assert "#2563eb" in audit_page._audit_field_rows["Gripper Model"].styleSheet()
    audit_page._clear_navigation_highlight()
    assert "#f97316" in audit_page.audit_fields["Gripper Model"].styleSheet()


def test_unknown_field_target_loads_audit_with_friendly_message(qapp, fake_config, fake_project):
    save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-UNKNOWN-FIELD-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Status": "In Progress",
        },
    )
    window = DashboardWindow(fake_config)
    navigator = annotation_target_navigator.AnnotationTargetNavigator(window)

    assert navigator.open_target({"target_type": "audit_field", "audit_id": "AUD-UNKNOWN-FIELD-001", "field_key": "Mystery Field"})

    audit_page = window.pages["audit"]
    assert audit_page.audit_fields["Audit ID"].text() == "AUD-UNKNOWN-FIELD-001"
    assert "Could not determine section" in audit_page.result_panel.viewer.toPlainText()


def test_notes_page_go_to_target_loads_linked_audit_field(qapp, fake_config, fake_project):
    save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-NOTE-GO-001",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "EOAT Vacuum Circuits": "2",
            "Status": "In Progress",
        },
    )
    service = AnnotationService(fake_config.project_root)
    target = service.create_or_get_target("audit_field", audit_id="AUD-NOTE-GO-001", machine_id="12", field_key="EOAT Vacuum Circuits", field_label="EOAT Vacuum Circuits")
    note = service.create_note("Linked navigation note", "Body", "Important", target_ids=[target.id])

    window = DashboardWindow(fake_config)
    window._show_page("notes")
    notes_page = window.pages["notes"]
    notes_page.select_note(note.id)
    notes_page.go_to_target()

    audit_page = window.pages["audit"]
    assert audit_page.audit_fields["Audit ID"].text() == "AUD-NOTE-GO-001"
    assert "Target field: EOAT Vacuum Circuits" in audit_page.result_panel.viewer.toPlainText()


def test_multiple_tag_targets_use_picker_before_navigation(qapp, fake_config, monkeypatch):
    service = AnnotationService(fake_config.project_root)
    tag = service.get_tag_by_name("Needs Review")
    first = service.create_or_get_target("audit_field", audit_id="AUD-PICK-001", field_key="EOAT Vacuum Circuits", field_label="EOAT Vacuum Circuits")
    second = service.create_or_get_target("audit_field", audit_id="AUD-PICK-002", field_key="Sensor Type", field_label="Sensor Type")
    service.assign_tag_to_target(tag.id, first.id, comment="First", sync_workbook=False)
    service.assign_tag_to_target(tag.id, second.id, comment="Second", sync_workbook=False)
    opened = {}

    def fake_exec(self):
        opened["rows"] = self.table.rowCount()
        self.selected_target = self.targets[1]
        return QDialog.DialogCode.Accepted

    def fake_open_target(self, target):
        opened["audit_id"] = target["audit_id"]
        return True

    monkeypatch.setattr(AnnotationTargetPickerDialog, "exec", fake_exec)
    monkeypatch.setattr(annotation_target_navigator.AnnotationTargetNavigator, "open_target", fake_open_target)

    navigator = annotation_target_navigator.AnnotationTargetNavigator(None)
    assert navigator.open_targets(service.get_targets_for_tag(tag.id), title="Select Target for Tag") is True
    assert opened == {"rows": 2, "audit_id": "AUD-PICK-002"}


def test_tags_page_tag_level_go_to_target_uses_multiple_target_picker(qapp, fake_config, monkeypatch):
    service = AnnotationService(fake_config.project_root)
    tag = service.get_tag_by_name("Needs Review")
    first = service.create_or_get_target("audit_field", audit_id="AUD-PAGE-PICK-001", field_key="EOAT Vacuum Circuits", field_label="EOAT Vacuum Circuits")
    second = service.create_or_get_target("audit_field", audit_id="AUD-PAGE-PICK-002", field_key="Sensor Type", field_label="Sensor Type")
    service.assign_tag_to_target(tag.id, first.id, comment="First", sync_workbook=False)
    service.assign_tag_to_target(tag.id, second.id, comment="Second", sync_workbook=False)
    captured = {}

    def fake_open_targets(self, targets, *, title):
        captured["title"] = title
        captured["count"] = len(targets)
        return True

    monkeypatch.setattr(annotation_target_navigator.AnnotationTargetNavigator, "open_targets", fake_open_targets)
    page = TagsPage(fake_config)
    page.select_tag_or_assignment(tag_id=tag.id)
    page.go_to_target()

    assert captured == {"title": "Select Target for Tag", "count": 2}


def _field_fill(project_root, audit_id: str, header: str) -> str:
    workbook = load_workbook(resolve_project_paths(project_root).master_workbook)
    try:
        ws = workbook["EOAT Inventory"]
        headers = [cell.value for cell in ws[1]]
        row = next(row for row in range(2, ws.max_row + 1) if ws.cell(row=row, column=headers.index("Audit ID") + 1).value == audit_id)
        return ws.cell(row=row, column=headers.index(header) + 1).fill.fgColor.rgb or ""
    finally:
        workbook.close()


def test_clear_form_confirmation_can_be_suppressed_and_clears_summary(qapp, fake_config, monkeypatch):
    page = AuditPage(fake_config)
    page.show()
    page.result_panel.show_text("Summary to clear")
    page.audit_fields["Press/Machine #"].setText("12")

    def accept_and_check(box):
        box.checkBox().setChecked(True)
        return QMessageBox.StandardButton.Ok

    monkeypatch.setattr(QMessageBox, "exec", accept_and_check)
    page.clear_audit_form(confirm=True)

    assert page._suppress_clear_confirm_this_session is True
    assert page.audit_fields["Press/Machine #"].text() == ""
    assert page.result_panel.viewer.toPlainText() == ""


def test_save_audit_preserves_form_and_combined_summary(qapp, fake_config, fake_project, frozen_project_date):
    page = AuditPage(fake_config)
    page.show()
    audit_id = page.audit_fields["Audit ID"].text()
    page.audit_fields["Press/Machine #"].setText("12")
    page.audit_fields["Robot Type"].setCurrentText("Wittmann R9")
    page.audit_fields["EOAT Type"].setCurrentText("Vacuum")

    click_button(page, "Save Audit Entry")
    wait_for_background_tasks()

    text = page.result_panel.viewer.toPlainText()
    assert "Audit Save Summary" in text
    assert "Robot Info Summary" in text
    assert "Compatibility Entry Summary" in text
    assert f"Saved audit entry {audit_id}" in text
    assert page.audit_fields["Press/Machine #"].text() == "12"
    assert page.audit_fields["Audit ID"].text() == audit_id
    assert page.has_unsaved_changes() is False


def test_save_audit_workflow_records_timing_and_defers_annotation_sync(qapp, fake_config, monkeypatch, frozen_project_date):
    page = AuditPage(fake_config)
    page.show()
    audit_id = page.audit_fields["Audit ID"].text()
    page.audit_fields["Press/Machine #"].setText("12")
    page.audit_fields["Robot Type"].setCurrentText("Wittmann R9")
    page.audit_fields["EOAT Type"].setCurrentText("Vacuum")
    calls = {"targeted": 0, "full": 0}

    def targeted_sync(self, sync_audit_id):
        calls["targeted"] += 1
        raise AssertionError("Normal audit save should not sync workbook annotation colors.")

    def full_sync(self):
        calls["full"] += 1
        raise AssertionError("Normal audit save should not run full workbook annotation sync.")

    monkeypatch.setattr(AnnotationService, "sync_tag_colors_to_workbook_for_audit", targeted_sync)
    monkeypatch.setattr(AnnotationService, "sync_all_tag_colors_to_workbook", full_sync)

    result = page._save_audit_workflow(
        {field: page._field_value(widget) for field, widget in page.audit_fields.items()},
        allow_update=False,
        create_followup_action=False,
    )

    assert result.success, result.errors
    assert result.metrics["audit_id"] == audit_id
    assert calls == {"targeted": 0, "full": 0}
    assert result.metrics["annotation_color_sync_deferred"] is True
    assert "audit_save_timing" in result.metrics
    assert "audit_save" in result.metrics["audit_save_timing"]
    assert "robot_info_save" in result.metrics["audit_save_timing"]
    assert "compatibility_autorun" in result.metrics["audit_save_timing"]
    assert "annotation_color_sync" in result.metrics["audit_save_timing"]


def test_updated_audit_defaults(qapp, fake_config):
    page = AuditPage(fake_config)
    page.show()

    assert page.audit_fields["Vacuum Generator Type"].text() == "Venturi"
    assert page.audit_fields["Quick Disconnects Present?"].currentText() == "Yes"
    assert page.audit_fields["Vacuum Confirmation Present?"].currentText() == "Yes"
    assert page.audit_fields["Part-Present Detection Present?"].currentText() == "No"
    assert page.audit_fields["Drawing/CAD Available?"].currentText() == "No"
    assert page.audit_fields["BOM Available?"].currentText() == "No"
    assert page.audit_fields["Process Binder Complete?"].currentText() == "No"
    assert page.audit_fields["Photos Taken?"].currentText() == "No"
    assert page.audit_fields["EOAT Interchangeable Circuits"].text() == "0"
    assert page.audit_fields["Robot Interchangeable Circuits"].text() == "0"

    page.audit_fields["Connection Type"].setCurrentText("ATI")
    assert page.audit_fields["Changeover Difficulty"].currentText() == "Low"
