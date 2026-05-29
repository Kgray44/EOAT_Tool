from __future__ import annotations

import pytest

from app.pages.audit_progress import AuditProgressPage
from app.pages.bom_spares import BomSparesPage
from app.pages.fmea import FmeaPage
from app.pages.issue_analysis import IssueAnalysisPage
from app.pages.kpi_dashboard import KpiDashboardPage
from app.pages.pilot_candidates import PilotCandidatesPage
from app.pages.pm_checklists import PmChecklistsPage
from app.pages.standards_docs import StandardsDocsPage
from tests.fixtures.fake_project import create_fake_eoat_project
from tests.ui.helpers import click_button, table_text, wait_for_background_tasks

pytestmark = pytest.mark.usability


def test_audit_progress_metrics_and_report(qapp, fake_config, fake_project):
    page = AuditProgressPage(fake_config)
    page.show()

    assert int(page.cards["Physical Audit Rows"].value_label.text()) > 0
    assert int(page.cards["Issues Logged"].value_label.text()) > 0
    click_button(page, "Generate Progress Report")
    wait_for_background_tasks()
    assert list((fake_project / "01_EOAT_Audit" / "Audit_Progress_Reports").glob("Audit_Progress_*.md"))


def test_issue_analysis_reports_fake_categories_and_missing_risk(qapp, fake_config, fake_project):
    page = IssueAnalysisPage(fake_config)
    page.show()

    click_button(page, "Run Issue Analysis")
    wait_for_background_tasks()

    combined = "\n".join([table_text(page.category_table), table_text(page.missing_table), table_text(page.fmea_table)])
    assert "Vacuum loss" in combined
    assert "Sensor failure" in combined
    assert "Tubing wear" in combined
    assert "ISS-002" in combined
    assert list((fake_project / "01_EOAT_Audit" / "Issue_Analysis_Reports").glob("Issue_Analysis_*.md"))


def test_fmea_lite_refresh_run_and_disabled_planned_button(qapp, fake_config, fake_project):
    page = FmeaPage(fake_config)
    page.show()

    click_button(page, "Calculate RPN / Refresh")
    click_button(page, "Suggest FMEA Entries")
    assert int(page.cards["Existing Rows"].value_label.text()) >= 1
    assert int(page.cards["Top RPN"].value_label.text()) >= 200
    assert click_button(page, "Edit Before Accepting").isEnabled()
    assert page.suggest_table.horizontalHeaderItem(0).text() == "Accept"
    click_button(page, "Run FMEA Analysis")
    wait_for_background_tasks()
    assert list((fake_project / "04_FMEA" / "FMEA_Reports").glob("FMEA_Lite_Report_*.md"))


def test_pilot_ranking_includes_candidate_and_empty_state(qapp, fake_config, fake_project, tmp_path):
    page = PilotCandidatesPage(fake_config)
    page.show()

    click_button(page, "Run Pilot Ranking")
    wait_for_background_tasks()
    assert "PILOT-001" in table_text(page.table)
    click_button(page, "Generate Evidence Packet")
    wait_for_background_tasks()
    assert list((fake_project / "05_Pilot_Project" / "Candidate_Cells").glob("Pilot_Candidate_Ranking_*.md"))
    assert list((fake_project / "05_Pilot_Project" / "Candidate_Cells").glob("Pilot_Evidence_Packet_*.md"))

    empty_project = create_fake_eoat_project(tmp_path / "empty_candidate_case")
    from openpyxl import load_workbook

    workbook_path = empty_project / "01_EOAT_Audit" / "EOAT_Audit_Database" / "EOAT_Master_Tracker.xlsx"
    wb = load_workbook(workbook_path)
    ws = wb["Pilot Candidates"]
    ws.delete_rows(2, ws.max_row)
    inventory = wb["EOAT Inventory"]
    headers = [cell.value for cell in inventory[1]]
    pilot_col = headers.index("Pilot Candidate?") + 1
    status_col = headers.index("Status") + 1
    for row in range(2, inventory.max_row + 1):
        inventory.cell(row=row, column=pilot_col).value = "No"
        inventory.cell(row=row, column=status_col).value = "Audited"
    wb.save(workbook_path)
    wb.close()
    empty_config = type(fake_config)(**{**fake_config.to_dict(), "project_root": str(empty_project)})
    empty_page = PilotCandidatesPage(empty_config)
    empty_page.show()
    assert empty_page.cards["Candidates Evaluated"].value_label.text() == "0"


def test_kpi_dashboard_cards_and_report(qapp, fake_config, fake_project):
    page = KpiDashboardPage(fake_config)
    page.show()

    click_button(page, "Run KPI Analysis")
    wait_for_background_tasks()
    assert float(page.cards["Downtime Minutes"].value_label.text()) >= 59
    assert float(page.cards["Part Drops"].value_label.text()) >= 10
    assert float(page.cards["Mis-Picks"].value_label.text()) >= 10
    assert "Press 101" in table_text(page.by_press_table)
    assert list((fake_project / "02_KPI_Data" / "Dashboard_Exports").glob("KPI_Dashboard_Report_*.md"))


def test_standards_documentation_gap_scan(qapp, fake_config, fake_project):
    page = StandardsDocsPage(fake_config)
    page.show()

    click_button(page, "Run Documentation Gap Scan")
    wait_for_background_tasks()
    assert int(page.cards["EOATs Scanned"].value_label.text()) > 0
    assert int(page.cards["Total Gaps"].value_label.text()) > 0
    assert int(page.cards["Avg Compliance"].value_label.text()) >= 0
    assert "Press 102" in table_text(page.top_table)
    assert list((fake_project / "03_Standards" / "Documentation_Gap_Reports").glob("Documentation_Gap_Report_*.md"))


def test_pm_checklist_generic_specific_and_invalid_friendly_fallback(qapp, fake_config, fake_project):
    page = PmChecklistsPage(fake_config)
    page.show()

    click_button(page, "Generate Generic Templates")
    wait_for_background_tasks()
    assert list((fake_project / "03_Standards" / "PM_Checklist_Draft" / "Generated_Checklists").glob("PM_Checklist_Generic_*.md"))

    page.audit_id_edit.setText("AUD-20260518-001")
    click_button(page, "Generate PM Checklist")
    wait_for_background_tasks()
    assert "PM Checklist - Press 101" in page.preview.toPlainText()

    page.audit_id_edit.setText("AUD-DOES-NOT-EXIST")
    click_button(page, "Generate PM Checklist")
    wait_for_background_tasks()
    assert "No matching EOAT rows found" in page.result_panel.viewer.toPlainText()


def test_bom_spare_parts_analysis_populates_common_and_missing_tables(qapp, fake_config, fake_project):
    page = BomSparesPage(fake_config)
    page.show()

    click_button(page, "Run BOM/Spare Parts Analysis")
    wait_for_background_tasks()
    assert int(page.cards["EOATs Scanned"].value_label.text()) > 0
    assert "Nitrile bellows cup" in table_text(page.common_table)
    assert "Press 102" in table_text(page.missing_table)
    assert list((fake_project / "03_Standards" / "BOM_Template_Draft" / "BOM_Standardization_Reports").glob("BOM_Standardization_Report_*.md"))
