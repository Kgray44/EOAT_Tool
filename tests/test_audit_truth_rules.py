from __future__ import annotations

from openpyxl import load_workbook

from app.pages.audit import get_empty_only_visible_fields
from core.audit_by_press import AUDIT_BY_PRESS_SHEET
from core.audit_compatibility import create_compatibility_entries
from core.audit_constants import COMPATIBILITY_SOURCE_FIELD, ENTRY_TYPE_COMPATIBLE, ENTRY_TYPE_FIELD, SOURCE_AUDIT_ID_FIELD
from core.audit_entries import load_audit_entry, save_audit_entry, validate_audit_entry
from core.audit_field_rules import (
    field_applies,
    field_group,
    hybrid_completeness_warnings,
    semantic_consistency_warnings,
)
from core.paths import resolve_project_paths
from core.validation import validate_project_foundation
from tests.fixtures.reference_workbooks import create_press_reference_workbooks


def test_shared_rules_match_eoat_type_visibility_and_normalization(fake_project):
    vacuum = {"EOAT Type": "Vacuum"}
    assert field_applies(vacuum, "# of Cups")
    assert field_applies(vacuum, "Cup Type/Material")
    assert not field_applies(vacuum, "Gripper Model")
    assert not field_applies(vacuum, "# of Grippers")

    mechanical = {"EOAT Type": "Mechanical / Gripper"}
    assert field_applies(mechanical, "Gripper Model")
    assert field_applies(mechanical, "# of Grippers")
    assert not field_applies(mechanical, "# of Cups")
    assert not field_applies(mechanical, "Cup Type/Material")

    hybrid = {"EOAT Type": "Hybrid"}
    assert field_applies(hybrid, "# of Cups")
    assert field_applies(hybrid, "Cup Type/Material")
    assert field_applies(hybrid, "Gripper Model")
    assert field_applies(hybrid, "# of Grippers")

    miscellaneous = {"EOAT Type": "Miscellaneous"}
    assert field_applies(miscellaneous, "Gripper Model")
    assert field_applies(miscellaneous, "# of Grippers")

    unknown = {"EOAT Type": "Unknown / Needs Review"}
    assert field_applies(unknown, "# of Cups")
    assert field_applies(unknown, "Cup Type/Material")
    assert not field_applies(unknown, "Gripper Model")
    assert not field_applies(unknown, "# of Grippers")

    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-TRUTH-VACUUM",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Tool #": "DEMO-PN-1200",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "# of Grippers": "2",
            "Gripper Type": "Single Pressure",
            "Gripper Model": "STALE-GRIPPER",
            "Gripper Size": "25 mm",
            "Status": "In Progress",
        },
    )

    assert result.success, result.errors
    assert result.metrics["fields_auto_set_to_na"] >= 2
    loaded = load_audit_entry(fake_project, "AUD-TRUTH-VACUUM")
    assert loaded["# of Grippers"] == "N/A"
    assert loaded["Gripper Type"] == "N/A"
    assert loaded["Gripper Model"] == "N/A"
    assert loaded["Gripper Size"] == "N/A"
    health = validate_project_foundation(fake_project)
    assert "Gripper Model" not in "\n".join(health.warnings)


def test_hybrid_and_semantic_warnings_are_non_blocking(fake_project):
    entry = {
        "Audit ID": "AUD-TRUTH-HYBRID",
        "Audit Date": "2026-05-18",
        "Auditor": "KG",
        "Plant/Area": "Plant 4",
        "Press/Machine #": "Press 14",
        "Tool #": "DEMO-PN-1400",
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Hybrid",
        "Status": "In Progress",
    }
    assert len(hybrid_completeness_warnings(entry)) == 2
    assert "vacuum-side" not in " ".join(hybrid_completeness_warnings({**entry, "# of Cups": "4"}))
    result = save_audit_entry(fake_project, entry)
    assert result.success, result.errors
    assert result.metrics["hybrid_warning_count"] == 2

    warnings = semantic_consistency_warnings(
        {
            "EOAT Type": "Mechanical / Gripper",
            "Gripper Type": "Vacuum",
            "# of Cups": "4",
            "Cup Type/Material": "Silicone",
            "Gripper Model": "Rubber",
            "Sensors Present?": "No",
            "Sensor Type": "Photoeye",
            "Quick Disconnects Present?": "No",
            "Pneumatic Quick Disconnect Type": "PTC",
        }
    )
    assert len(warnings) >= 5


def test_sensor_electrical_and_quick_disconnect_rules_are_separate():
    no_sensors = {"EOAT Type": "Vacuum", "Sensors Present?": "No", "Electrical/Wiring Present?": "Unknown / Not Checked"}
    assert not field_applies(no_sensors, "Sensor Type")
    assert not field_applies(no_sensors, "Sensor Brand/Model")
    assert not field_applies(no_sensors, "Part-Present Detection Present?")
    assert field_applies(no_sensors, "Cable Management Condition")

    no_wiring = {**no_sensors, "Electrical/Wiring Present?": "No"}
    assert not field_applies(no_wiring, "Cable Management Condition")
    assert not field_applies(no_wiring, "Electrical Quick Disconnect Type")

    old_schema_no_wiring_control = {"EOAT Type": "Vacuum", "Sensors Present?": "No", "Cable Management Condition": "N/A"}
    assert not field_applies(old_schema_no_wiring_control, "Cable Management Condition")
    old_schema_with_evidence = {"EOAT Type": "Vacuum", "Sensors Present?": "No", "Cable Management Condition": "OK"}
    assert field_applies(old_schema_with_evidence, "Cable Management Condition")

    no_qd = {"Quick Disconnects Present?": "No"}
    assert not field_applies(no_qd, "Pneumatic Quick Disconnect Type")
    assert not field_applies(no_qd, "Electrical Quick Disconnect Type")


def test_compatible_rows_validate_differently_from_physical_audits(fake_project):
    _errors, warnings = validate_audit_entry(
        {
            ENTRY_TYPE_FIELD: ENTRY_TYPE_COMPATIBLE,
            "Press/Machine #": "70",
            "Tool #": "DEMO-PN-0170",
            SOURCE_AUDIT_ID_FIELD: "AUD-SOURCE",
            COMPATIBILITY_SOURCE_FIELD: "Synthetic compatibility map",
            "Audit Date": "",
            "Auditor": "",
        }
    )
    assert not _errors
    assert not any("Audit Date" in warning or "Auditor" in warning for warning in warnings)

    errors, _warnings = validate_audit_entry(
        {
            ENTRY_TYPE_FIELD: ENTRY_TYPE_COMPATIBLE,
            "Press/Machine #": "70",
            "Tool #": "DEMO-PN-0170",
        }
    )
    assert not errors


def test_physical_audit_blank_tool_number_is_high_priority_warning(fake_project):
    result = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-MISSING-TOOL",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Status": "In Progress",
        },
    )

    assert result.success, result.errors
    assert "Missing important audit field: Tool #" in "\n".join(result.warnings)


def test_cup_count_is_important_for_vacuum_and_hybrid_but_not_gripper():
    _errors, vacuum_warnings = validate_audit_entry({"EOAT Type": "Vacuum", "# of Cups": ""})
    _errors, hybrid_warnings = validate_audit_entry({"EOAT Type": "Hybrid", "# of Cups": ""})
    _errors, gripper_warnings = validate_audit_entry({"EOAT Type": "Mechanical / Gripper", "# of Cups": "N/A"})

    assert "Missing important audit field: # of Cups" in "\n".join(vacuum_warnings)
    assert "Missing important audit field: # of Cups" in "\n".join(hybrid_warnings)
    assert "Missing important audit field: # of Cups" not in "\n".join(gripper_warnings)


def test_connection_type_is_mounting_connection_not_qd_detail():
    entry = {"Quick Disconnects Present?": "No"}

    assert field_group("Connection Type") == "tool_mounting_connection"
    assert field_applies(entry, "Connection Type")
    assert not field_applies(entry, "Pneumatic Quick Disconnect Type")
    assert not field_applies(entry, "Electrical Quick Disconnect Type")


def test_empty_only_filters_to_applicable_empty_fields():
    row = {
        "EOAT Type": "Mechanical / Gripper",
        "Gripper Model": "N/A",
        "Cup Type/Material": "N/A",
        "Sensors Present?": "No",
        "Electrical/Wiring Present?": "Unknown / Not Checked",
        "Sensor Type": "N/A",
        "Cable Management Condition": "N/A",
    }
    fields = ["Gripper Model", "Cup Type/Material", "Sensor Type", "Cable Management Condition"]

    visible = get_empty_only_visible_fields(row, {field: object() for field in fields}, lambda field: field_applies(row, field))

    assert [field.name for field in visible] == ["Gripper Model", "Cable Management Condition"]


def test_audit_by_press_counts_physical_and_compatible_rows(fake_project):
    reference_root = create_press_reference_workbooks(fake_project / "reference-data")
    source = save_audit_entry(
        fake_project,
        {
            "Audit ID": "AUD-PRESS-SOURCE",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "1",
            "Tool #": "DEMO-PN-0170",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "EOAT Moves": "Part",
            "Status": "Complete",
        },
    )
    assert source.success, source.errors
    created = create_compatibility_entries(fake_project, "AUD-PRESS-SOURCE", ["70"], reference_root / "press_capacity.xlsx")
    assert created.success, created.errors

    workbook = load_workbook(resolve_project_paths(fake_project).master_workbook, read_only=True)
    try:
        view = workbook[AUDIT_BY_PRESS_SHEET]
        headers = [cell.value for cell in view[3]]
        assert ENTRY_TYPE_FIELD in headers
        assert SOURCE_AUDIT_ID_FIELD in headers
        group_headers = [
            row[0]
            for row in view.iter_rows(min_col=1, max_col=1, values_only=True)
            if isinstance(row[0], str) and "total entr" in row[0]
        ]
        assert any("1 physical, 0 compatible, 1 total entry" in value for value in group_headers)
        assert any("0 physical, 1 compatible, 1 total entry" in value for value in group_headers)
    finally:
        workbook.close()

    rows = load_workbook(resolve_project_paths(fake_project).master_workbook, read_only=True)
    try:
        ws = rows["EOAT Inventory"]
        headers = [cell.value for cell in ws[1]]
        data_rows = [
            {headers[index]: value for index, value in enumerate(row)}
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]
        compatible = next(row for row in data_rows if row.get(ENTRY_TYPE_FIELD) == ENTRY_TYPE_COMPATIBLE)
        assert compatible["Audit ID"]
        assert compatible["Press/Machine #"] == "70"
        assert compatible["Tool #"] == "DEMO-PN-0170"
        assert compatible[ENTRY_TYPE_FIELD] == ENTRY_TYPE_COMPATIBLE
        assert compatible[SOURCE_AUDIT_ID_FIELD] == "AUD-PRESS-SOURCE"
        assert compatible[COMPATIBILITY_SOURCE_FIELD]
        assert compatible["Audit Date"] == "N/A"
        assert compatible["Auditor"] == "N/A"
        assert compatible["Gripper Model"] == "N/A"
    finally:
        rows.close()
