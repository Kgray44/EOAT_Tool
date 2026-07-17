from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from pydantic import ValidationError

import server.eoat_api.database.models  # noqa: F401
from core.domain.models import EOAT
from server.eoat_api.database.base import Base
from tools.migration.excel_to_mysql import analyze_workbook


def test_domain_model_enforces_identifiers_counts_and_timezone():
    with pytest.raises(ValidationError):
        EOAT(business_identifier="", number_of_grippers=-1)
    with pytest.raises(ValidationError):
        EOAT(business_identifier="P4-EOAT-TEST", created_at=datetime(2026, 1, 1))


def test_sqlalchemy_foundation_has_concrete_tables_and_constraints():
    assert len(Base.metadata.tables) == 54
    assert "eoat_installations" in Base.metadata.tables
    table = Base.metadata.tables["eoat_installations"]
    names = {constraint.name for constraint in table.constraints}
    assert "uq_active_installation_eoat" in names
    assert "uq_active_installation_machine" in names


def test_dry_run_preserves_source_and_refuses_to_infer_parts_or_installations(tmp_path: Path):
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    inventory = workbook.active
    inventory.title = "EOAT Inventory"
    inventory.append([
        "Audit ID", "Audit Date", "Plant/Area", "Press/Machine #", "Tool #", "EOAT Assembly ID",
        "Part Name/Description", "Entry Type", "EOAT Type", "Connection Type", "Cleanroom/Non-Cleanroom",
    ])
    inventory.append(["AUD-1", "2026-01-01", "Plant 4", "1", "T-1", "P4-EOAT-0001", "Part A", "Audited", "Vacuum", "ATI", "Whiteroom"])
    photos = workbook.create_sheet("Photo Index")
    photos.append(["Photo ID", "Stored Relative Path", "Stored Filename", "Folder Path", "Photo Filename"])
    metadata = workbook.create_sheet("_EOAT_App_Metadata")
    metadata.append(["key", "value"])
    metadata.append(["schema_version", "test"])
    workbook.save(path)
    before = path.read_bytes()
    report = analyze_workbook(path)
    assert path.read_bytes() == before
    assert report.source_unchanged is True
    assert report.staged_counts["eoats"] == 1
    assert report.staged_counts["parts"] == 0
    assert report.staged_counts["installation_records"] == 0
    assert any(issue.code == "PART_IDENTIFIER_AMBIGUITY" for issue in report.issues)
