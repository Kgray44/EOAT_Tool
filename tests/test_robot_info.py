from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.audit_entries import load_audit_entry, repair_workbook_schema, save_audit_entry
from core.paths import resolve_project_paths
from core.robot_info import (
    ROBOT_INFO_HEADERS,
    ROBOT_INFO_SHEET,
    ROBOT_NOTES_FIELD,
    ensure_robot_info_workbook,
    load_robot_info_for_audit_entry,
    robot_info_workbook_path,
    upsert_robot_info_from_audit,
    validate_robot_info_workbook,
)
from core.workbook_schema import get_expected_headers


def _required_audit_entry(**overrides) -> dict[str, str]:
    entry = {
        "Audit ID": "AUD-PNEUMATIC-001",
        "Audit Date": "2026-05-27",
        "Auditor": "KG",
        "Plant/Area": "Plant 4",
        "Press/Machine #": "Press 12",
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "Status": "In Progress",
    }
    entry.update(overrides)
    return entry


def test_save_audit_writes_eoat_pneumatic_fields_without_robot_columns(fake_project):
    result = save_audit_entry(
        fake_project,
        _required_audit_entry(
            **{
                "Number of Parts Picked": "2",
                "EOAT Vacuum Circuits": "2",
                "EOAT Pressure Circuits": "1",
                "EOAT Interchangeable Circuits": "0",
                "Robot Vacuum Circuits": "3",
                ROBOT_NOTES_FIELD: "Robot-side note should stay out of EOAT Inventory.",
            }
        ),
    )

    assert result.success, result.errors
    saved = load_audit_entry(fake_project, "AUD-PNEUMATIC-001")
    assert saved["Number of Parts Picked"] == "2"
    assert saved["EOAT Vacuum Circuits"] == "2"
    assert saved["EOAT Pressure Circuits"] == "1"
    assert saved["EOAT Interchangeable Circuits"] == "0"
    assert "Robot Vacuum Circuits" not in saved
    assert ROBOT_NOTES_FIELD not in saved


def test_save_audit_rejects_negative_eoat_circuit_count(fake_project):
    result = save_audit_entry(
        fake_project,
        _required_audit_entry(
            **{
                "Audit ID": "AUD-PNEUMATIC-BAD-001",
                "EOAT Vacuum Circuits": "-1",
            }
        ),
    )

    assert result.success is False
    assert any("EOAT Vacuum Circuits must be a non-negative whole number" in error for error in result.errors)


def test_legacy_vacuum_cups_header_migrates_to_parts_picked(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    parts_col = headers.index("Number of Parts Picked") + 1
    ws.cell(row=1, column=parts_col).value = "Number of Vacuum Cups"
    ws.column_dimensions[ws.cell(row=1, column=parts_col).column_letter].width = 23.5
    ws.cell(row=1, column=parts_col).fill = PatternFill(fill_type="solid", fgColor="FFC000")
    ws.cell(row=1, column=parts_col).font = Font(name="Calibri", bold=True, italic=True, color="FF111111")
    ws.cell(row=1, column=parts_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=1, column=parts_col).border = Border(bottom=Side(style="thick", color="FF000000"))
    ws.append(
        [
            "AUD-LEGACY-CUPS-001" if header == "Audit ID" else 6 if header == "Number of Parts Picked" else ""
            for header in headers
        ]
    )
    workbook.save(workbook_path)
    workbook.close()

    result = repair_workbook_schema(fake_project, log_activity=False)

    assert result.success, result.errors
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    migrated_headers = [cell.value for cell in ws[1]]
    rows = {
        row[migrated_headers.index("Audit ID")]: {migrated_headers[index]: value for index, value in enumerate(row)}
        for row in ws.iter_rows(min_row=2, values_only=True)
    }
    workbook.close()
    assert "Number of Vacuum Cups" not in migrated_headers
    assert "Number of Parts Picked" in migrated_headers
    assert "# of Cups" in migrated_headers
    assert rows["AUD-LEGACY-CUPS-001"]["Number of Parts Picked"] == 6
    assert rows["AUD-LEGACY-CUPS-001"]["# of Cups"] in (None, "")
    migrated_col = migrated_headers.index("Number of Parts Picked") + 1
    migrated_cell = ws.cell(row=1, column=migrated_col)
    assert ws.column_dimensions[migrated_cell.column_letter].width == 23.5
    assert migrated_cell.fill.fgColor.rgb == "00FFC000"
    assert migrated_cell.font.bold is True
    assert migrated_cell.font.italic is True
    assert migrated_cell.font.color.rgb == "FF111111"
    assert migrated_cell.alignment.horizontal == "center"
    assert migrated_cell.alignment.vertical == "center"
    assert migrated_cell.alignment.wrap_text is True
    assert migrated_cell.border.bottom.style == "thick"


def test_expected_workbook_schema_uses_parts_picked_not_vacuum_cups():
    headers = get_expected_headers("EOAT Inventory")

    assert "Number of Parts Picked" in headers
    assert "# of Cups" in headers
    assert "Number of Vacuum Cups" not in headers


def test_robot_info_workbook_is_created_and_upserted_by_machine_identity(fake_project):
    first = upsert_robot_info_from_audit(
        fake_project,
        _required_audit_entry(
            **{
                "Audit ID": "AUD-ROBOT-001",
                "Robot Vacuum Circuits": "3",
                "Robot Pressure Circuits": "2",
                "Robot Interchangeable Circuits": "0",
            }
        ),
    )
    second = upsert_robot_info_from_audit(
        fake_project,
        _required_audit_entry(
            **{
                "Audit ID": "AUD-ROBOT-002",
                "Robot Vacuum Circuits": "4",
                "Robot Pressure Circuits": "1",
                "Robot Interchangeable Circuits": "0",
            }
        ),
    )

    assert first.success, first.errors
    assert second.success, second.errors
    path = robot_info_workbook_path(fake_project)
    assert path.exists()
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook[ROBOT_INFO_SHEET]
    headers = [cell.value for cell in ws[1]]
    rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]
    workbook.close()
    assert headers[: len(ROBOT_INFO_HEADERS)] == ROBOT_INFO_HEADERS
    assert len(rows) == 1
    assert rows[0]["Machine Number"] == "Press 12"
    assert rows[0]["Robot Vacuum Circuits"] == 4
    assert rows[0]["Robot Pressure Circuits"] == 1
    assert rows[0]["Last Audit ID"] == "AUD-ROBOT-002"
    assert ROBOT_NOTES_FIELD in headers


def test_robot_info_schema_repair_adds_robot_notes_without_destroying_data(fake_project):
    path = robot_info_workbook_path(fake_project)
    workbook = load_workbook(path) if path.exists() else None
    if workbook is None:
        from openpyxl import Workbook

        workbook = Workbook()
    ws = workbook.active
    ws.title = ROBOT_INFO_SHEET
    old_headers = [header for header in ROBOT_INFO_HEADERS if header != ROBOT_NOTES_FIELD]
    ws.append(old_headers)
    ws.append(
        [
            "Plant 4" if header == "Plant/Area" else
            "Press 12" if header == "Machine Number" else
            "Wittmann R9" if header == "Robot Type" else
            3 if header == "Robot Vacuum Circuits" else
            2 if header == "Robot Pressure Circuits" else
            0 if header == "Robot Interchangeable Circuits" else
            "AUD-OLD" if header == "Last Audit ID" else
            ""
            for header in old_headers
        ]
    )
    workbook.save(path)
    workbook.close()

    ensure_robot_info_workbook(fake_project)

    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook[ROBOT_INFO_SHEET]
    headers = [cell.value for cell in ws[1]]
    row = dict(zip(headers, next(ws.iter_rows(min_row=2, values_only=True))))
    workbook.close()

    assert ROBOT_NOTES_FIELD in headers
    assert row["Machine Number"] == "Press 12"
    assert row["Robot Vacuum Circuits"] == 3
    assert row[ROBOT_NOTES_FIELD] in ("", None)


def test_load_robot_info_for_audit_entry_returns_matching_robot_row(fake_project):
    entry = _required_audit_entry(
        **{
            "Audit ID": "AUD-ROBOT-LOAD-001",
            "Robot Vacuum Circuits": "5",
            "Robot Pressure Circuits": "2",
            "Robot Interchangeable Circuits": "0",
            ROBOT_NOTES_FIELD: "Shared robot-side manifold note.",
        }
    )
    assert upsert_robot_info_from_audit(fake_project, entry).success

    loaded = load_robot_info_for_audit_entry(fake_project, entry)

    assert loaded is not None
    assert loaded["Robot Vacuum Circuits"] == 5
    assert loaded["Robot Pressure Circuits"] == 2
    assert loaded["Robot Interchangeable Circuits"] == 0
    assert loaded[ROBOT_NOTES_FIELD] == "Shared robot-side manifold note."


def test_robot_notes_can_be_saved_and_intentionally_cleared(fake_project):
    entry = _required_audit_entry(
        **{
            "Audit ID": "AUD-ROBOT-NOTES-001",
            "Robot Vacuum Circuits": "5",
            "Robot Pressure Circuits": "2",
            "Robot Interchangeable Circuits": "0",
            ROBOT_NOTES_FIELD: "Check wrist air labels.",
        }
    )
    assert upsert_robot_info_from_audit(fake_project, entry).success
    loaded = load_robot_info_for_audit_entry(fake_project, entry)
    assert loaded is not None
    assert loaded[ROBOT_NOTES_FIELD] == "Check wrist air labels."

    cleared = {**entry, "Audit ID": "AUD-ROBOT-NOTES-002", ROBOT_NOTES_FIELD: ""}
    assert upsert_robot_info_from_audit(fake_project, cleared).success

    loaded = load_robot_info_for_audit_entry(fake_project, entry)
    assert loaded is not None
    assert loaded[ROBOT_NOTES_FIELD] in ("", None)


def test_robot_info_health_detects_duplicate_rows_and_invalid_counts(fake_project):
    path = robot_info_workbook_path(fake_project)
    assert upsert_robot_info_from_audit(
        fake_project,
        _required_audit_entry(
            **{
                "Robot Vacuum Circuits": "3",
                "Robot Pressure Circuits": "2",
                "Robot Interchangeable Circuits": "0",
            }
        ),
    ).success
    workbook = load_workbook(path)
    ws = workbook[ROBOT_INFO_SHEET]
    headers = [cell.value for cell in ws[1]]
    first_row = [ws.cell(row=2, column=index + 1).value for index in range(len(headers))]
    duplicate = list(first_row)
    duplicate[headers.index("Robot Pressure Circuits")] = -1
    ws.append(duplicate)
    workbook.save(path)
    workbook.close()

    warnings, errors, metrics = validate_robot_info_workbook(fake_project)

    assert errors == []
    assert metrics["robot_info_duplicate_row_count"] == 1
    assert metrics["robot_info_invalid_circuit_count"] == 1
    assert any("Duplicate Robot Info row" in warning for warning in warnings)
    assert any("Invalid Robot Info circuit value" in warning for warning in warnings)
