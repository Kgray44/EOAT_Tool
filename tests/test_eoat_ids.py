from __future__ import annotations

import json

import pytest
from openpyxl import Workbook, load_workbook

from core.audit_constants import ENTRY_TYPE_COMPATIBLE, ENTRY_TYPE_FIELD, SOURCE_AUDIT_ID_FIELD
from core.audit_entries import save_audit_entry
from core.eoat_ids import (
    CANONICAL_AREA_CLEANROOM,
    CANONICAL_AREA_PLANT4,
    EOAT_ASSEMBLY_ID_FIELD,
    assign_missing_eoat_assembly_ids,
    assign_missing_eoat_assembly_ids_in_workbook,
    build_eoat_assembly_contexts,
    canonical_area,
    determine_eoat_prefix,
    format_eoat_id,
    generate_next_eoat_assembly_id,
    is_valid_eoat_assembly_id,
    normalize_area,
    parse_eoat_id,
    update_eoat_info_file,
)
from core.paths import get_press_capacity_file, resolve_project_paths
from core.workbook_cache import invalidate_workbook_cache
from core.workbook_io import row_dicts
from core.workbook_schema import get_expected_headers


def test_eoat_id_generation_uses_next_highest_number() -> None:
    assert generate_next_eoat_assembly_id(["P4-EOAT-0001", "P4-EOAT-0005", "bad"]) == "P4-EOAT-0006"


def test_eoat_id_generation_uses_area_prefix_sequences() -> None:
    existing = ["P4-EOAT-0005", "CL-EOAT-0002", "bad"]

    assert generate_next_eoat_assembly_id(existing, {"Plant/Area": "Cleanroom"}) == "CL-EOAT-0003"
    assert generate_next_eoat_assembly_id(existing, {"Plant/Area": "Plant 4"}) == "P4-EOAT-0006"


@pytest.mark.parametrize("value", ["Cleanroom", "Clean Room", "CL", "C/R", "CR", "cleanroom", "clean room"])
def test_cleanroom_area_aliases(value: str) -> None:
    assert normalize_area(value) == CANONICAL_AREA_CLEANROOM
    assert determine_eoat_prefix({"Plant/Area": value}) == "CL"


@pytest.mark.parametrize(
    "value",
    ["Plant 4", "P4", "Whiteroom", "White Room", "Non-Cleanroom", "Non Cleanroom", "Production"],
)
def test_plant4_area_aliases(value: str) -> None:
    assert normalize_area(value) == CANONICAL_AREA_PLANT4
    assert determine_eoat_prefix({"Plant/Area": value}) == "P4"


def test_area_field_preferred_over_cleanroom_flag() -> None:
    row = {"Plant/Area": "Cleanroom", "Cleanroom/Non-Cleanroom": "Whiteroom"}

    assert canonical_area(row) == CANONICAL_AREA_CLEANROOM
    assert determine_eoat_prefix(row) == "CL"


def test_eoat_id_regex_accepts_p4_and_cl_and_rejects_malformed() -> None:
    assert parse_eoat_id("P4-EOAT-0001").value == "P4-EOAT-0001"
    assert parse_eoat_id("cl-eoat-0047").value == "CL-EOAT-0047"
    assert format_eoat_id("CL", 47) == "CL-EOAT-0047"
    assert is_valid_eoat_assembly_id("CL-EOAT-0047")
    assert not is_valid_eoat_assembly_id("P5-EOAT-0047")
    assert not is_valid_eoat_assembly_id("CL-EOAT-47")


def test_assign_missing_eoat_ids_only_fills_blanks() -> None:
    rows = [
        {EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-0002", "Tool #": "TOOL-A"},
        {EOAT_ASSEMBLY_ID_FIELD: "", "Tool #": "TOOL-B"},
        {EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-0002", "Tool #": "TOOL-C"},
        {"Tool #": "TOOL-D"},
    ]

    summary = assign_missing_eoat_assembly_ids(rows)

    assert summary.assigned_count == 2
    assert summary.ids_created == ["P4-EOAT-0003", "P4-EOAT-0004"]
    assert rows[0][EOAT_ASSEMBLY_ID_FIELD] == "P4-EOAT-0002"
    assert rows[2][EOAT_ASSEMBLY_ID_FIELD] == "P4-EOAT-0002"


def test_assign_missing_eoat_ids_uses_row_area_prefix() -> None:
    rows = [
        {EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-0002", "Plant/Area": "Plant 4"},
        {EOAT_ASSEMBLY_ID_FIELD: "CL-EOAT-0004", "Plant/Area": "Cleanroom"},
        {EOAT_ASSEMBLY_ID_FIELD: "", "Plant/Area": "Cleanroom"},
        {EOAT_ASSEMBLY_ID_FIELD: "", "Plant/Area": "Plant 4"},
    ]

    summary = assign_missing_eoat_assembly_ids(rows)

    assert summary.ids_created == ["CL-EOAT-0005", "P4-EOAT-0003"]
    assert rows[2][EOAT_ASSEMBLY_ID_FIELD] == "CL-EOAT-0005"
    assert rows[3][EOAT_ASSEMBLY_ID_FIELD] == "P4-EOAT-0003"


def test_eoat_context_collects_machine_number_aliases() -> None:
    contexts = build_eoat_assembly_contexts(
        [
            {
                EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-0007",
                "Tool #": "TOOL-A",
                "Machine #": "12",
                "Audit ID": "AUD-001",
            },
            {
                EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-0007",
                "Tool #": "TOOL-B",
                "Press/Machine #": "N/A",
                "Machine No.": "13",
                "Audit ID": "AUD-002",
            },
        ]
    )

    assert contexts["P4-EOAT-0007"].machines == ("12", "13")


def test_eoat_info_file_includes_press_capacity_machines(fake_project) -> None:
    capacity_path = get_press_capacity_file(fake_project)
    capacity_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Capacity"
    ws.append(["Machine No.", "NGW Part Number", "NGW Part Description"])
    ws.append(["12, 13, 18", "TOOL-CAPACITY", "Capacity tool"])
    workbook.save(capacity_path)
    workbook.close()

    path = update_eoat_info_file(
        fake_project,
        "P4-EOAT-0007",
        audit_rows=[
            {
                EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-0007",
                "Tool #": "TOOL-CAPACITY",
                "Press/Machine #": "12",
                "Audit ID": "AUD-001",
            }
        ],
    )
    payload = json.loads(path.read_text(encoding="utf-8"))

    assert payload["audit_machines"] == ["12"]
    assert payload["press_capacity_machines"] == ["12", "13", "18"]
    assert payload["known_machines"] == ["12", "13", "18"]


def test_assign_missing_eoat_ids_in_workbook_adds_column_safely(fake_project) -> None:
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    if EOAT_ASSEMBLY_ID_FIELD in headers:
        ws.delete_cols(headers.index(EOAT_ASSEMBLY_ID_FIELD) + 1)
        headers.remove(EOAT_ASSEMBLY_ID_FIELD)
    row = {header: "" for header in get_expected_headers("EOAT Inventory")}
    row.update({"Audit ID": "AUD-EOAT-BACKFILL-001", "Tool #": "TOOL-BACKFILL"})
    ws.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()
    invalidate_workbook_cache(workbook_path)

    result = assign_missing_eoat_assembly_ids_in_workbook(fake_project, log_activity=False)
    rows = row_dicts(workbook_path, "EOAT Inventory")
    saved = next(row for row in rows if row["Audit ID"] == "AUD-EOAT-BACKFILL-001")

    assert result.success is True
    assert result.metrics["assigned_count"] >= 1
    assert saved[EOAT_ASSEMBLY_ID_FIELD].startswith("P4-EOAT-")


def test_audit_save_generates_missing_eoat_id_and_preserves_existing(fake_project) -> None:
    generated = save_audit_entry(
        fake_project,
        {
            "Audit Date": "2026-06-08",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "26",
            "Tool #": "TOOL-GENERATED",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Status": "Audited",
        },
    )
    preserved = save_audit_entry(
        fake_project,
        {
            "Audit Date": "2026-06-08",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "31",
            EOAT_ASSEMBLY_ID_FIELD: "P4-EOAT-0099",
            "Tool #": "TOOL-PRESERVED",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Status": "Audited",
        },
    )
    rows = row_dicts(resolve_project_paths(fake_project).master_workbook, "EOAT Inventory")
    generated_row = next(row for row in rows if row["Tool #"] == "TOOL-GENERATED")
    preserved_row = next(row for row in rows if row["Tool #"] == "TOOL-PRESERVED")

    assert generated.success is True
    assert preserved.success is True
    assert generated_row[EOAT_ASSEMBLY_ID_FIELD].startswith("P4-EOAT-")
    assert preserved_row[EOAT_ASSEMBLY_ID_FIELD] == "P4-EOAT-0099"


def test_cleanroom_audit_save_generates_cl_id(fake_project) -> None:
    generated = save_audit_entry(
        fake_project,
        {
            "Audit Date": "2026-06-08",
            "Auditor": "KG",
            "Plant/Area": "Cleanroom",
            "Press/Machine #": "63",
            "Tool #": "TOOL-CLEANROOM",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Status": "Audited",
        },
    )
    rows = row_dicts(resolve_project_paths(fake_project).master_workbook, "EOAT Inventory")
    generated_row = next(row for row in rows if row["Tool #"] == "TOOL-CLEANROOM")

    assert generated.success is True, generated.errors
    assert generated_row[EOAT_ASSEMBLY_ID_FIELD].startswith("CL-EOAT-")


def test_compatible_entry_inherits_source_eoat_id(fake_project) -> None:
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    source = {header: "" for header in headers}
    source.update(
        {
            "Audit ID": "AUD-SOURCE-CL",
            "Plant/Area": "Cleanroom",
            "Press/Machine #": "63",
            "Tool #": "TOOL-SOURCE",
            EOAT_ASSEMBLY_ID_FIELD: "CL-EOAT-0047",
            "Entry Type": "Audited",
        }
    )
    ws.append([source.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()
    invalidate_workbook_cache(workbook_path)

    from core.audit_entries import normalize_audit_entry

    normalized = normalize_audit_entry(
        fake_project,
        {
            ENTRY_TYPE_FIELD: ENTRY_TYPE_COMPATIBLE,
            SOURCE_AUDIT_ID_FIELD: "AUD-SOURCE-CL",
            "Plant/Area": "Cleanroom",
            "Press/Machine #": "64",
            "Tool #": "TOOL-SOURCE",
        },
    )

    assert normalized[EOAT_ASSEMBLY_ID_FIELD] == "CL-EOAT-0047"
