from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from core.paths import resolve_project_paths
from core.photo_evidence import (
    audit_photo_intake_folder,
    create_audit_photo_intake_folder,
    evidence_coverage_for_audit,
    export_photo_checklist,
    validate_photo_evidence,
)
from core.workbook_schema import get_expected_headers


def _append_sheet_row(project_root, sheet_name: str, values: dict[str, str]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook[sheet_name]
    headers = [cell.value for cell in ws[1]]
    ws.append([values.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def _audit_row(audit_id: str, eoat_type: str, **overrides) -> dict[str, str]:
    row = {header: "" for header in get_expected_headers("EOAT Inventory")}
    row.update(
        {
            "Audit ID": audit_id,
            "Audit Date": "2026-05-18",
            "Auditor": "Synthetic Auditor",
            "Plant/Area": "Plant 4",
            "Press/Machine #": f"Press {audit_id[-1]}",
            "Robot Type": "Wittmann R9",
            "EOAT Type": eoat_type,
            "Status": "Complete",
            "Priority": "High",
            "Pilot Candidate?": "No",
            "Sensors Present?": "No",
            "Quick Disconnects Present?": "No",
            "Known Issues": "No issue observed.",
            "Process Binder Complete?": "No",
            "Photos Taken?": "No",
        }
    )
    row.update(overrides)
    return row


def _status(coverage, category: str):
    return next(status for status in coverage.statuses if status.category == category)


def test_photo_evidence_category_applicability_tracks_eoat_type(fake_project):
    for row in [
        _audit_row("AUD-VAC-001", "Vacuum"),
        _audit_row("AUD-MECH-002", "Mechanical / Gripper"),
        _audit_row("AUD-HYB-003", "Hybrid"),
    ]:
        _append_sheet_row(fake_project, "EOAT Inventory", row)

    vacuum = evidence_coverage_for_audit(fake_project, "AUD-VAC-001")
    mechanical = evidence_coverage_for_audit(fake_project, "AUD-MECH-002")
    hybrid = evidence_coverage_for_audit(fake_project, "AUD-HYB-003")

    assert _status(vacuum, "vacuum_cups").applies is True
    assert _status(vacuum, "grippers").applies is False
    assert _status(mechanical, "grippers").applies is True
    assert _status(mechanical, "vacuum_cups").applies is False
    assert _status(hybrid, "grippers").applies is True
    assert _status(hybrid, "vacuum_cups").applies is True


def test_missing_required_photo_evidence_creates_structured_findings(fake_project):
    _append_sheet_row(
        fake_project,
        "EOAT Inventory",
        _audit_row(
            "AUD-EVID-004",
            "Mechanical / Gripper",
            **{
                "Sensors Present?": "Yes",
                "Quick Disconnects Present?": "Yes",
                "Process Binder Complete?": "Yes",
                "Known Issues": "Loose gripper jaw needs review.",
                "Pilot Candidate?": "Yes",
            },
        ),
    )

    warnings, metrics, findings = validate_photo_evidence(fake_project)

    assert metrics["photo_evidence_finding_count"] >= 1
    assert any(finding.category == "missing_evidence" for finding in findings)
    assert any("sensor photo" in finding.message.casefold() for finding in findings)
    assert any("quick disconnect" in warning.casefold() for warning in warnings)


def test_photo_evidence_uses_photo_index_without_requiring_photo_files(fake_project):
    _append_sheet_row(
        fake_project,
        "EOAT Inventory",
        _audit_row("AUD-PHOTO-005", "Vacuum", **{"Sensors Present?": "Yes"}),
    )
    _append_sheet_row(
        fake_project,
        "Photo Index",
        {
            "Photo ID": "PHO-20260518-001",
            "Date Taken": "2026-05-18",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 5",
            "EOAT Area Shown": "Sensors",
            "Photo Filename": "synthetic_sensor_photo.jpg",
            "Folder Path": str(fake_project / "synthetic" / "not-real"),
            "Related Audit ID": "AUD-PHOTO-005",
        },
    )

    coverage = evidence_coverage_for_audit(fake_project, "AUD-PHOTO-005")

    assert _status(coverage, "sensors").present is True
    assert _status(coverage, "sensors").photo_count == 1


def test_compatible_row_inherits_source_photo_evidence(fake_project):
    _append_sheet_row(
        fake_project,
        "EOAT Inventory",
        _audit_row(
            "AUD-PHOTO-SOURCE-007",
            "Mechanical / Gripper",
            **{
                "Tool #": "TOOL-INHERIT-PHOTO",
                "EOAT Assembly ID": "P4-EOAT-9907",
                "Photos Taken?": "Yes",
            },
        ),
    )
    _append_sheet_row(
        fake_project,
        "EOAT Inventory",
        _audit_row(
            "AUD-PHOTO-COMPAT-007",
            "Mechanical / Gripper",
            **{
                "Entry Type": "Compatible",
                "Source Audit ID": "AUD-PHOTO-SOURCE-007",
                "Compatibility Source": "Press Capacity",
                "Press/Machine #": "Press 70",
                "Tool #": "TOOL-INHERIT-PHOTO",
                "EOAT Assembly ID": "P4-EOAT-9907",
                "Photos Taken?": "No",
                "Sensors Present?": "Yes",
                "Quick Disconnects Present?": "Yes",
                "Known Issues": "Review inherited setup.",
            },
        ),
    )
    _append_sheet_row(
        fake_project,
        "Photo Index",
        {
            "Photo ID": "PHO-INHERIT-001",
            "Date Taken": "2026-05-18",
            "EOAT Assembly ID": "P4-EOAT-9907",
            "Tool #": "TOOL-INHERIT-PHOTO",
            "EOAT Area Shown": "Overall EOAT",
            "Related Audit ID": "AUD-PHOTO-SOURCE-007",
        },
    )

    warnings, _metrics, findings = validate_photo_evidence(fake_project)
    coverage = evidence_coverage_for_audit(fake_project, "AUD-PHOTO-COMPAT-007")

    assert coverage.compatible_evidence_accepted is True
    assert coverage.missing_required_count == 0
    assert not any(finding.audit_id == "AUD-PHOTO-COMPAT-007" for finding in findings)
    assert not any("AUD-PHOTO-COMPAT-007" in warning for warning in warnings)


def test_photo_evidence_links_bench_audit_by_eoat_without_machine(fake_project):
    _append_sheet_row(
        fake_project,
        "EOAT Inventory",
        _audit_row(
            "AUD-BENCH-PHOTO-006",
            "Vacuum",
            **{
                "Press/Machine #": "",
                "Tool #": "TOOL-BENCH-PHOTO",
                "EOAT Assembly ID": "P4-EOAT-9901",
                "Photos Taken?": "Yes",
            },
        ),
    )
    _append_sheet_row(
        fake_project,
        "Photo Index",
        {
            "Photo ID": "PHO-BENCH-001",
            "Date Taken": "2026-05-18",
            "EOAT Assembly ID": "P4-EOAT-9901",
            "Tool #": "TOOL-BENCH-PHOTO",
            "EOAT Area Shown": "Overall EOAT",
            "Photo Filename": "bench_overall.jpg",
            "Folder Path": str(fake_project / "synthetic" / "bench"),
        },
    )

    coverage = evidence_coverage_for_audit(fake_project, "AUD-BENCH-PHOTO-006")
    checklist = export_photo_checklist(fake_project, "AUD-BENCH-PHOTO-006", log_activity=False)
    text = Path(next(iter(checklist.output_reports))).read_text(encoding="utf-8")

    assert coverage.related_photo_count == 1
    assert coverage.machine == ""
    assert "- EOAT Assembly ID: P4-EOAT-9901" in text
    assert "- Tool #: TOOL-BENCH-PHOTO" in text


def test_indexed_photos_with_photos_taken_no_create_actionable_status_finding(fake_project):
    _append_sheet_row(
        fake_project,
        "EOAT Inventory",
        _audit_row("AUD-PHOTO-STATUS-006", "Vacuum", **{"Photos Taken?": "No"}),
    )
    _append_sheet_row(
        fake_project,
        "Photo Index",
        {
            "Photo ID": "PHO-STATUS-001",
            "Date Taken": "2026-05-18",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 6",
            "EOAT Area Shown": "Overall EOAT",
            "Related Audit ID": "AUD-PHOTO-STATUS-006",
        },
    )

    warnings, _metrics, findings = validate_photo_evidence(fake_project)

    assert any("Photos Taken? is not marked Yes" in warning for warning in warnings)
    assert any(finding.audit_id == "AUD-PHOTO-STATUS-006" for finding in findings)


def test_audit_photo_intake_folder_and_checklist_export(fake_project):
    result = create_audit_photo_intake_folder(fake_project, "AUD-CHECK-006", log_activity=False)

    assert result.success is True
    folder = audit_photo_intake_folder(fake_project, "AUD-CHECK-006")
    assert folder.exists()

    checklist = export_photo_checklist(fake_project, "AUD-CHECK-006", log_activity=False)

    assert checklist.success is True
    path = next(iter(checklist.output_reports))
    text = Path(path).read_text(encoding="utf-8")
    assert "Front View" in text
    assert "Quick Disconnects" in text
    assert "Do not add real photos to source control" in text
