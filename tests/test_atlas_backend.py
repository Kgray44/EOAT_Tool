from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from core.atlas_data_loader import invalidate_atlas_data_cache, load_atlas_data
from core.atlas_exports import (
    build_eoat_qr_payload,
    build_install_packet,
    decode_qr_payload_from_image,
    export_eoat_qr_label,
    export_install_packet,
    qr_payload_warning,
    validate_eoat_qr_payload,
)
from core.atlas_models import DocumentationStatus, EOATRecord
from core.atlas_recommendations import recommend_for_query
from core.atlas_record_details import (
    RecordDetailData,
    RecordField,
    RecordPhoto,
    RecordPhotoGroup,
    RecordSection,
    build_record_detail_data,
)
from core.atlas_search import search_atlas
from core.atlas_utils import row_value
from core.compatibility_engine import compatibility_matrix_rows, machine_to_eoats, tool_to_eoats
from core.documentation_score import calculate_documentation_status
from core.paths import resolve_project_paths
from core.reporting.pdf_footer import LEGAL_FOOTER_TEXT
from core.reporting.pdf_image_utils import PdfImageResult, prepare_image_for_pdf
from core.reporting.pdf_record_report import export_record_pdf
from core.reporting.record_report_options import ReportOptions
from tests.fixtures.fake_images import create_fake_images
from tests.fixtures.fake_project import create_fake_eoat_project
from tests.fixtures.reference_workbooks import create_press_reference_workbooks


def test_atlas_header_alias_lookup() -> None:
    row = {"Tool Number": "12345", "Machine Number": "Press 12", "EOAT ID": "P4-EOAT-0001"}

    assert row_value(row, ("Tool #", "Tool Number")) == "12345"
    assert row_value(row, ("Machine #", "Machine Number")) == "Press 12"
    assert row_value(row, ("EOAT Assembly ID", "EOAT ID")) == "P4-EOAT-0001"


def test_atlas_load_builds_fast_lookup_indexes(tmp_path: Path) -> None:
    root = create_fake_eoat_project(tmp_path)
    create_press_reference_workbooks(resolve_project_paths(root).reference_data, multiple_capacity_rows=True)
    invalidate_atlas_data_cache(root)

    bundle = load_atlas_data(root, force_refresh=True)

    assert len(bundle.eoats) == 3
    assert "toola" in bundle.indexes.eoats_by_tool
    assert bundle.indexes.eoats_by_machine["101"] == ("AUD-20260518-001",)
    assert bundle.metrics["workbook_load_ms"] >= 0
    assert bundle.metrics["cache_build_ms"] >= 0


def test_atlas_search_and_recommendation_use_exact_tool_match(tmp_path: Path) -> None:
    root = create_fake_eoat_project(tmp_path)
    create_press_reference_workbooks(resolve_project_paths(root).reference_data)
    bundle = load_atlas_data(root, force_refresh=True)

    matches = search_atlas(bundle, "Tool TOOL-A")
    recommendation = recommend_for_query(bundle, "Tool TOOL-A")

    assert matches[0].result_type in {"eoat", "tool"}
    assert recommendation.best is not None
    assert recommendation.best.eoat_id == "AUD-20260518-001"
    assert [candidate.eoat_id for candidate in recommendation.candidates] == ["AUD-20260518-001"]
    assert recommendation.install_checklist


def test_atlas_compatibility_engine_answers_tool_and_machine(tmp_path: Path) -> None:
    root = create_fake_eoat_project(tmp_path)
    create_press_reference_workbooks(resolve_project_paths(root).reference_data)
    bundle = load_atlas_data(root, force_refresh=True)

    assert tool_to_eoats(bundle, "TOOL-A")[0].eoat_id == "AUD-20260518-001"
    assert machine_to_eoats(bundle, "101")[0].eoat_id == "AUD-20260518-001"
    rows = compatibility_matrix_rows(bundle)
    assert any(row["EOAT"] == "AUD-20260518-001" and row["Machine"] == "101" for row in rows)


def test_atlas_excludes_unaudited_capacity_only_tools_by_default(tmp_path: Path) -> None:
    root = create_fake_eoat_project(tmp_path)
    create_press_reference_workbooks(resolve_project_paths(root).reference_data)

    bundle = load_atlas_data(root, force_refresh=True)

    assert bundle.metrics["exclude_unaudited_tools"] is True
    assert bundle.metrics["unaudited_press_capacity_relationships_excluded"] == 3
    assert "DEMO-PN-1200" not in {tool.tool for tool in bundle.tools}
    assert "12" not in {machine.machine for machine in bundle.machines}
    assert not bundle.press_capacity_rows


def test_atlas_can_include_unaudited_capacity_tools_when_setting_is_off(tmp_path: Path) -> None:
    root = create_fake_eoat_project(tmp_path)
    create_press_reference_workbooks(resolve_project_paths(root).reference_data)

    bundle = load_atlas_data(root, force_refresh=True, exclude_unaudited_tools=False)

    assert bundle.metrics["exclude_unaudited_tools"] is False
    assert "DEMO-PN-1200" in {tool.tool for tool in bundle.tools}
    assert "12" in {machine.machine for machine in bundle.machines}
    capacity_row = next(row for row in bundle.press_capacity_rows if row.get("NGW Part Number") == "DEMO-PN-1200")
    assert capacity_row["Bill-to / Customer"] == "Demo Customer A"
    assert str(capacity_row["Cycle Time (S)"]) == "18.5"


def test_record_pdf_includes_press_capacity_appendix_fields(tmp_path: Path) -> None:
    from pypdf import PdfReader

    root = create_fake_eoat_project(tmp_path)
    create_press_reference_workbooks(resolve_project_paths(root).reference_data)
    bundle = load_atlas_data(root, force_refresh=True, exclude_unaudited_tools=False)
    detail_data = build_record_detail_data(bundle, "tool", "DEMO-PN-1200")

    pdf_path = export_record_pdf(detail_data, tmp_path / "Tool_Report_DEMO-PN-1200.pdf", project_root=root)
    text = "\n".join(page.extract_text() or "" for page in PdfReader(str(pdf_path)).pages)

    assert "Workbook Data Appendix" in text
    assert "Press Capacity Rows" in text
    assert "Demo Customer A" in text
    assert "Cycle Time (S)" in text


def test_record_pdf_legal_footer_for_eoat_tool_and_machine_reports(tmp_path: Path) -> None:
    from pypdf import PdfReader

    for record_type, record_id in (("eoat", "P4-EOAT-0002"), ("tool", "6200360010"), ("machine", "36")):
        detail_data = _simple_record_detail(record_type, record_id)
        pdf_path = export_record_pdf(
            detail_data,
            tmp_path / f"{record_type}_{record_id}.pdf",
            project_root=tmp_path,
            options=ReportOptions(include_photos=False),
        )
        text = "\n".join(page.extract_text() or "" for page in PdfReader(str(pdf_path)).pages)

        assert LEGAL_FOOTER_TEXT in text
        assert f"EOAT Atlas - {record_id}" in text
        assert "Page 1" in text


def test_record_pdf_legal_footer_appears_on_every_page(tmp_path: Path) -> None:
    from pypdf import PdfReader

    sections = tuple(
        RecordSection(
            f"Long Section {index + 1}",
            tuple(RecordField(f"Field {index + 1}.{field + 1}", "Reference information " * 5) for field in range(4)),
        )
        for index in range(36)
    )
    detail_data = _simple_record_detail("eoat", "P4-EOAT-MULTI", report_sections=sections)
    pdf_path = export_record_pdf(
        detail_data,
        tmp_path / "EOAT_Report_Multipage_Footer.pdf",
        project_root=tmp_path,
        options=ReportOptions(include_photos=False),
    )
    pages = PdfReader(str(pdf_path)).pages

    assert len(pages) > 1
    assert all(LEGAL_FOOTER_TEXT in (page.extract_text() or "") for page in pages)


def test_record_pdf_resolves_photo_candidates_and_lists_missing_images(tmp_path: Path) -> None:
    from pypdf import PdfReader

    image_path = create_fake_images(tmp_path / "photos")[1]
    missing_path = tmp_path / "photos" / "missing_front.jpg"
    detail_data = RecordDetailData(
        record_type="eoat",
        record_id="P4-EOAT-0005",
        title="P4-EOAT-0005",
        subtitle="Vacuum EOAT",
        condition="In Service",
        plant_area="Plant 4",
        hero_fields=(RecordField("Type", "Vacuum"),),
        detail_sections=(),
        documentation_fields=(),
        photo_groups=(
            RecordPhotoGroup(
                "Overall / Front View",
                (
                    RecordPhoto(
                        path=str(tmp_path / "photos" / "stale_front.png"),
                        filename=image_path.name,
                        category="Front View",
                        photo_id="front",
                        path_candidates=(str(tmp_path / "photos" / "stale_front.png"), str(image_path)),
                    ),
                    RecordPhoto(
                        path=str(missing_path),
                        filename=missing_path.name,
                        category="Side View",
                        photo_id="missing",
                        path_candidates=(str(missing_path),),
                    ),
                ),
            ),
        ),
        history_fields=(),
        summary_fields=(),
        report_sections=(),
    )

    pdf_path = export_record_pdf(
        detail_data,
        tmp_path / "EOAT_Report_Image_Handling.pdf",
        project_root=tmp_path,
        options=ReportOptions(include_photo_appendix=True, include_missing_photo_status=True),
    )
    text = "\n".join(page.extract_text() or "" for page in PdfReader(str(pdf_path)).pages)

    assert pdf_path.exists()
    assert "Front View" in text
    assert "Missing photo status" in text
    assert "missing_front.jpg" in text


def test_record_pdf_skips_invalid_heic_without_crashing(tmp_path: Path) -> None:
    from pypdf import PdfReader

    heic_path = tmp_path / "photos" / "P4-EOAT-0002_2026-06-10_Front_View_001.HEIC"
    heic_path.parent.mkdir(parents=True, exist_ok=True)
    heic_path.write_bytes(b"not a heic image")
    detail_data = _photo_report_detail(
        tmp_path,
        (
            RecordPhoto(
                path=str(heic_path),
                filename=heic_path.name,
                category="Front View",
                photo_id="heic-front",
                path_candidates=(str(heic_path),),
            ),
        ),
    )

    pdf_path = export_record_pdf(
        detail_data,
        tmp_path / "EOAT_Report_Invalid_HEIC.pdf",
        project_root=tmp_path,
        options=ReportOptions(include_photo_appendix=True, include_missing_photo_status=True),
    )
    text = "\n".join(page.extract_text() or "" for page in PdfReader(str(pdf_path)).pages)

    assert pdf_path.exists()
    assert "Image unavailable" in text
    assert "Photos skipped or unavailable" in text
    assert heic_path.name in text


def test_record_pdf_contact_sheet_keeps_valid_image_when_one_photo_is_bad(tmp_path: Path) -> None:
    from pypdf import PdfReader

    valid_path = create_fake_images(tmp_path / "photos")[0]
    bad_path = tmp_path / "photos" / "bad_side.HEIC"
    bad_path.write_bytes(b"bad image data")
    detail_data = _photo_report_detail(
        tmp_path,
        (
            RecordPhoto(path=str(valid_path), filename=valid_path.name, category="Front View", photo_id="front", path_candidates=(str(valid_path),)),
            RecordPhoto(path=str(bad_path), filename=bad_path.name, category="Side View", photo_id="bad", path_candidates=(str(bad_path),)),
        ),
    )

    pdf_path = export_record_pdf(
        detail_data,
        tmp_path / "EOAT_Report_Mixed_Photos.pdf",
        project_root=tmp_path,
        options=ReportOptions(include_photo_appendix=True, include_missing_photo_status=True),
    )
    text = "\n".join(page.extract_text() or "" for page in PdfReader(str(pdf_path)).pages)

    assert pdf_path.exists()
    assert "Front View" in text
    assert "Side View" in text
    assert "Image unavailable" in text
    assert "Photos skipped or unavailable" in text


def test_pdf_image_conversion_cache_reuses_prepared_image(tmp_path: Path) -> None:
    from PIL import Image

    image_path = tmp_path / "photos" / "valid.png"
    image_path.parent.mkdir(parents=True, exist_ok=True)
    Image.new("RGB", (80, 60), (24, 141, 255)).save(image_path)

    first = prepare_image_for_pdf(image_path, tmp_path, max_size=(320, 240))
    second = prepare_image_for_pdf(image_path, tmp_path, max_size=(320, 240))

    assert first.ok
    assert second.ok
    assert first.pdf_safe_path == second.pdf_safe_path
    assert second.cache_hit
    assert Path(second.pdf_safe_path or "").exists()


def test_reportlab_image_receives_converted_path_not_raw_heic(tmp_path: Path, monkeypatch) -> None:
    from core.reporting import pdf_record_report

    raw_heic = tmp_path / "photos" / "front.HEIC"
    safe_jpg = tmp_path / "00_Project_Admin" / "cache" / "pdf_images" / "front.jpg"
    raw_heic.parent.mkdir(parents=True, exist_ok=True)
    safe_jpg.parent.mkdir(parents=True, exist_ok=True)
    raw_heic.write_bytes(b"fake heic payload")
    safe_jpg.write_bytes(b"fake prepared jpeg")
    captured_paths: list[str] = []
    detail_data = _photo_report_detail(
        tmp_path,
        (
            RecordPhoto(path=str(raw_heic), filename=raw_heic.name, category="Front View", photo_id="front", path_candidates=(str(raw_heic),)),
        ),
    )

    def fake_prepare(path, project_root, max_size=(1200, 900), prefer_format="JPEG"):
        return PdfImageResult(ok=True, pdf_safe_path=str(safe_jpg), original_path=str(path), converted=True)

    def fake_image(path, *_args, **_kwargs):
        captured_paths.append(str(path))
        return pdf_record_report.Paragraph("prepared image", pdf_record_report._styles()["body"])

    monkeypatch.setattr(pdf_record_report, "prepare_image_for_pdf", fake_prepare)
    monkeypatch.setattr(pdf_record_report, "Image", fake_image)

    export_record_pdf(
        detail_data,
        tmp_path / "EOAT_Report_No_Raw_HEIC.pdf",
        project_root=tmp_path,
        options=ReportOptions(include_photo_appendix=True),
    )

    assert str(safe_jpg) in captured_paths
    assert str(raw_heic) not in captured_paths


def test_atlas_indexes_photo_folder_by_eoat_id(tmp_path: Path) -> None:
    root = create_fake_eoat_project(tmp_path, with_photos=False)
    paths = resolve_project_paths(root)
    _set_first_inventory_eoat_id(paths.master_workbook, "P4-EOAT-0001")
    photo_folder = paths.cell_photos / "P4-EOAT-0001" / "00_Overall"
    photo_folder.mkdir(parents=True)
    (photo_folder / "P4-EOAT-0001_overall_001.jpg").write_bytes(b"not a real image but enough for indexing")

    bundle = load_atlas_data(root, force_refresh=True)
    eoat = next(record for record in bundle.eoats if record.eoat_id == "P4-EOAT-0001")

    assert eoat.photo_count == 1
    assert bundle.indexes.photos_by_eoat["p4eoat0001"]


def test_atlas_handles_missing_optional_sources(tmp_path: Path) -> None:
    root = create_fake_eoat_project(tmp_path, with_photos=False)
    bundle = load_atlas_data(root, force_refresh=True)

    assert len(bundle.eoats) == 3
    assert any(not source.available for source in bundle.source_statuses if source.label in {"Press Capacity", "Robot Info"})
    assert bundle.warnings


def test_atlas_registers_root_standardization_document(tmp_path: Path) -> None:
    root = create_fake_eoat_project(tmp_path, with_photos=False)
    root_standard = root / "EOAT Standardization Guide.md"
    root_standard.write_text("# EOAT Standardization Guide\n\nSynthetic design guidelines.\n", encoding="utf-8")
    invalidate_atlas_data_cache(root)

    bundle = load_atlas_data(root, force_refresh=True)
    paths = resolve_project_paths(root)

    assert (paths.standards / root_standard.name).exists()
    assert any(standard.title == "Eoat Standardization Guide" for standard in bundle.standards)
    assert any(standard.category == "eoat standards" for standard in bundle.standards)


def test_atlas_cache_detects_replaced_standard_documents(tmp_path: Path) -> None:
    root = create_fake_eoat_project(tmp_path, with_photos=False)
    paths = resolve_project_paths(root)
    paths.standards.mkdir(parents=True, exist_ok=True)
    old_standard = paths.standards / "EOAT_Standardization_Work_Instruction_Spaced_Annotated.pdf"
    revised_standard = paths.standards / "EOAT_Standardization_Work_Instruction_Revised.pdf"
    old_standard.write_bytes(b"old standard document")
    invalidate_atlas_data_cache(root)

    bundle = load_atlas_data(root, force_refresh=True)
    assert any(standard.title == "Eoat Standardization Work Instruction Spaced Annotated" for standard in bundle.standards)

    old_standard.unlink()
    revised_standard.write_bytes(b"revised standard document with updated content")
    bundle = load_atlas_data(root)
    titles = {standard.title for standard in bundle.standards}

    assert "Eoat Standardization Work Instruction Revised" in titles
    assert "Eoat Standardization Work Instruction Spaced Annotated" not in titles


def test_eoat_qr_payload_modes_store_offline_record(tmp_path: Path) -> None:
    root = create_fake_eoat_project(tmp_path)
    create_press_reference_workbooks(resolve_project_paths(root).reference_data)
    bundle = load_atlas_data(root, force_refresh=True)
    eoat = bundle.eoats[0]

    compact = build_eoat_qr_payload(eoat, mode="compact")
    deep_link = build_eoat_qr_payload(eoat, mode="deep_link")
    json_payload = build_eoat_qr_payload(eoat, mode="json")
    full = build_eoat_qr_payload(eoat, mode="full")

    assert compact.startswith("EOAT_ATLAS_RECORD")
    assert f"EOAT={eoat.eoat_id}" in compact
    assert "TOOL=T-" in compact
    assert "MACHINES=M-" in compact
    assert not compact.strip().isdigit()
    assert not compact.casefold().startswith(("tel:", "call:"))
    assert validate_eoat_qr_payload(compact, mode="compact", eoat_id=eoat.eoat_id) == []
    assert deep_link.startswith(f"eoat-atlas://record/eoat/{eoat.eoat_id}")
    assert "tool=T-" in deep_link
    assert '"app":"EOAT Atlas"' in json_payload
    assert '"record_type":"eoat"' in json_payload
    assert f'"eoat_id":"{eoat.eoat_id}"' in json_payload
    assert '"tools":' in json_payload
    assert '"machines":' in json_payload
    assert "KNOWN_ISSUES:" in full
    assert len(full) > len(compact)
    assert qr_payload_warning(full, mode="full")
    try:
        import qrcode  # noqa: F401
    except ImportError:
        return
    label_path = export_eoat_qr_label(bundle, eoat, payload_mode="compact")
    assert label_path.exists()
    assert "Atlas_Exports" in str(label_path)
    assert "QR_Labels" in str(label_path)
    assert decode_qr_payload_from_image(label_path).payload == compact


def test_eoat_qr_payload_validation_rejects_phone_like_values() -> None:
    assert validate_eoat_qr_payload("5620040010", mode="compact", eoat_id="P4-EOAT-0001")
    assert validate_eoat_qr_payload("tel:5620040010", mode="compact", eoat_id="P4-EOAT-0001")
    assert validate_eoat_qr_payload("call: 5620040010", mode="compact", eoat_id="P4-EOAT-0001")
    assert validate_eoat_qr_payload(
        "EOAT_ATLAS_RECORD; EOAT=P4-EOAT-0001; TOOL=5620040010",
        mode="compact",
        eoat_id="P4-EOAT-0001",
    )


def test_compact_eoat_qr_payload_decodes_not_phone_like_for_reported_tools(tmp_path: Path) -> None:
    for eoat_id, tool in [("P4-EOAT-0001", "5620040010"), ("P4-EOAT-0002", "5116950010")]:
        eoat = EOATRecord(
            eoat_id=eoat_id,
            display_id=eoat_id,
            tools=(tool,),
            machines=("1", "2", "8", "9", "19", "32", "33"),
            eoat_type="Mechanical/Gripper",
            documentation=DocumentationStatus(score=83, status_label="Good"),
        )
        bundle = load_atlas_data(create_fake_eoat_project(tmp_path / eoat_id), force_refresh=True)
        payload = build_eoat_qr_payload(eoat, mode="compact")
        label_path = export_eoat_qr_label(bundle, eoat, payload_mode="compact")
        decoded = decode_qr_payload_from_image(label_path).payload

        assert decoded == payload
        assert decoded.startswith("EOAT_ATLAS_RECORD")
        assert eoat_id in decoded
        assert f"T-{tool}" in decoded
        assert decoded != tool
        assert not decoded[:1].isdigit()
        assert not decoded.casefold().startswith(("call:", "tel:"))


def test_install_packet_exports_timestamped_markdown(tmp_path: Path) -> None:
    root = create_fake_eoat_project(tmp_path)
    create_press_reference_workbooks(resolve_project_paths(root).reference_data)
    bundle = load_atlas_data(root, force_refresh=True)
    eoat = bundle.eoats[0]

    packet = build_install_packet(bundle, eoat=eoat, context="Unit Test")
    path = export_install_packet(bundle, packet)

    text = path.read_text(encoding="utf-8")
    assert path.exists()
    assert "Install_Packets" in str(path)
    assert path.name.startswith("Atlas_Install_Packet_")
    assert eoat.eoat_id in text
    assert "## Fit Check" in text


def test_documentation_score_flags_missing_critical_fields() -> None:
    status = calculate_documentation_status({"Tool #": "123", "EOAT Type": "Vacuum"}, photo_count=0)

    assert status.score < 75
    assert "EOAT Assembly ID" in status.critical_missing_fields
    assert status.status_label in {"Critical gaps", "Missing important info"}


def _photo_report_detail(tmp_path: Path, photos: tuple[RecordPhoto, ...]) -> RecordDetailData:
    return RecordDetailData(
        record_type="eoat",
        record_id="P4-EOAT-0002",
        title="P4-EOAT-0002",
        subtitle="Vacuum EOAT",
        condition="In Service",
        plant_area="Plant 4",
        hero_fields=(RecordField("Type", "Vacuum"),),
        detail_sections=(),
        documentation_fields=(),
        photo_groups=(RecordPhotoGroup("Overall / Front View", photos),),
        history_fields=(),
        summary_fields=(),
        report_sections=(),
    )


def _simple_record_detail(record_type: str, record_id: str, *, report_sections: tuple[RecordSection, ...] = ()) -> RecordDetailData:
    return RecordDetailData(
        record_type=record_type,
        record_id=record_id,
        title=record_id,
        subtitle=f"{record_type.title()} report",
        condition="In Service",
        plant_area="Plant 4",
        hero_fields=(RecordField("Record Type", record_type), RecordField("Record ID", record_id)),
        detail_sections=(),
        documentation_fields=(RecordField("Documentation Status", "Indexed"),),
        photo_groups=(),
        history_fields=(),
        summary_fields=(),
        report_sections=report_sections,
    )


def _set_first_inventory_eoat_id(workbook_path: Path, eoat_id: str) -> None:
    workbook = load_workbook(workbook_path)
    try:
        ws = workbook["EOAT Inventory"]
        headers = [cell.value for cell in ws[1]]
        column = headers.index("EOAT Assembly ID") + 1
        ws.cell(row=2, column=column).value = eoat_id
        workbook.save(workbook_path)
    finally:
        workbook.close()
