from __future__ import annotations

import json
import re
import subprocess
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from core.morning_context import detect_generic_morning_report
from core.morning_planner import build_morning_plan_details_markdown, build_morning_plan_markdown, generate_morning_plan
from core.paths import resolve_project_paths

FORBIDDEN_MORNING_HEADINGS = [
    "Confidence / Data Quality",
    "Source Availability",
    "Recent Activity",
    "Workbook State",
    "Recent Modified Files",
    "App State Signals",
    "Planner Reasoning Summary",
    "Plan Sources",
    "What Changed Since Yesterday",
    "Yesterday -> Today Continuity",
]


def _section(markdown: str, heading: str, next_heading: str | None = None) -> str:
    chunk = markdown.split(f"## {heading}", 1)[1]
    if next_heading:
        chunk = chunk.split(f"## {next_heading}", 1)[0]
    return chunk.strip()


def _word_count(markdown: str) -> int:
    return len(re.findall(r"\b[\w#/.-]+\b", markdown))


def _list_item_count(markdown: str) -> int:
    return sum(1 for line in markdown.splitlines() if re.match(r"^(?:- |\d+\. )", line.strip()))


def _major_sections(markdown: str) -> list[str]:
    return [line[3:].strip() for line in markdown.splitlines() if line.startswith("## ")]


def _assert_todo_contract(markdown: str) -> None:
    assert _word_count(markdown) <= 250
    assert _list_item_count(markdown) <= 12
    assert _major_sections(markdown) == [
        "Today's Mission",
        "Do First",
        "Main TODO",
        "Ask Today",
        "If Blocked",
        "Done When",
    ]
    for heading in FORBIDDEN_MORNING_HEADINGS:
        assert f"## {heading}" not in markdown
    assert not re.search(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}", markdown)
    assert not re.search(r"(?:[A-Z]:\\|\\\\|/[^ \n]+/)", markdown)
    assert "source availability" not in markdown.lower()
    assert "confidence" not in markdown.lower()
    assert "recent activity" not in markdown.lower()
    assert "workbook rows" not in markdown.lower()
    ask_section = _section(markdown, "Ask Today", "If Blocked")
    assert sum(1 for line in ask_section.splitlines() if line.startswith("- ")) <= 4
    list_texts = [
        re.sub(r"^(?:- \[ \] |- |\d+\. )", "", line).strip().lower()
        for line in markdown.splitlines()
        if re.match(r"^(?:- |\d+\. )", line.strip())
    ]
    assert len(list_texts) == len(set(list_texts))


def _morning_plan_quality_fixture(fake_project):
    admin = fake_project / "00_Project_Admin"
    (admin / "project_schedule_week1.json").write_text(
        json.dumps(
            {
                "days": {
                    "1": ["Confirm kickoff expectations"],
                    "2": [
                        "Begin target cell list",
                        "Start first walkthrough/audit if approved",
                        "Decide photo naming system",
                        "Run workbook validation after Day 2 changes",
                    ],
                    "3": ["Review documentation gaps", "Stretch report"],
                    "4": ["Compare BOM common components"],
                }
            }
        ),
        encoding="utf-8",
    )
    (admin / "task_progress_week1.json").write_text(
        json.dumps(
            {
                "tasks": [
                    {
                        "task_id": "W1D1T1",
                        "day": "1",
                        "task_text": "Confirm kickoff expectations",
                        "status": "Complete",
                    },
                    {"task_id": "W1D2T1", "day": "2", "task_text": "Begin target cell list", "status": "Not started"},
                    {
                        "task_id": "W1D2T2",
                        "day": "2",
                        "task_text": "Start first walkthrough/audit if approved",
                        "status": "Not started",
                    },
                    {
                        "task_id": "W1D2T3",
                        "day": "2",
                        "task_text": "Decide photo naming system",
                        "status": "Not started",
                    },
                    {
                        "task_id": "W1D2T4",
                        "day": "2",
                        "task_text": "Run workbook validation after Day 2 changes",
                        "status": "Not started",
                    },
                    {
                        "task_id": "W1D3T1",
                        "day": "3",
                        "task_text": "Review documentation gaps",
                        "status": "Not started",
                    },
                    {"task_id": "W1D3T2", "day": "3", "task_text": "Stretch report", "status": "Not started"},
                    {
                        "task_id": "W1D4T1",
                        "day": "4",
                        "task_text": "Compare BOM common components",
                        "status": "Not started",
                    },
                ]
            }
        ),
        encoding="utf-8",
    )


def test_morning_planner_handles_missing_schedule(fake_project):
    markdown, warnings, metrics = build_morning_plan_markdown(fake_project, week=2, day=1)

    assert "Week 2 Day 1 Morning Plan" in markdown
    assert warnings
    assert metrics["primary_focus_items"] >= 0


def test_morning_planner_prioritizes_carryover_blocked_and_actions(fake_project):
    admin = fake_project / "00_Project_Admin"
    (admin / "project_schedule_week1.json").write_text(
        json.dumps({"days": {"1": ["Audit Press 12"], "2": ["Review data"]}}), encoding="utf-8"
    )
    (admin / "task_progress_week1.json").write_text(
        json.dumps(
            {
                "tasks": [
                    {"id": "T1", "day": "1", "task": "Audit Press 12", "status": "In progress"},
                    {"id": "T2", "day": "2", "task": "Review data", "status": "Blocked"},
                    {"id": "T3", "day": "3", "task": "Stretch report", "status": "Not started"},
                    {"id": "T4", "day": "1", "task": "Already done", "status": "Complete"},
                ]
            }
        ),
        encoding="utf-8",
    )
    paths = resolve_project_paths(fake_project)
    wb = load_workbook(paths.master_workbook)
    ws = wb["Action Items"]
    ws.append(
        ["ACT-1", "2026-05-18", "Ask maintenance for spare cup info", "Press 12", "KG", "High", "", "Open", "", ""]
    )
    wb.save(paths.master_workbook)
    wb.close()

    result = generate_morning_plan(fake_project, week=1, day=2)

    assert result.success is True
    assert result.metrics["carryover_tasks"] == 1
    assert result.metrics["blocked_tasks"] == 1
    assert result.output_reports


def test_morning_plan_filename_uses_resolved_project_day(fake_project):
    admin = fake_project / "00_Project_Admin"
    (admin / "project_schedule_week1.json").write_text(
        json.dumps({"days": {"2": ["Begin target cell list"]}}), encoding="utf-8"
    )
    (admin / "task_progress_week1.json").write_text(
        json.dumps(
            {
                "tasks": [
                    {"task_id": "W1D2T1", "day": "2", "task_text": "Begin target cell list", "status": "Not started"}
                ]
            }
        ),
        encoding="utf-8",
    )

    result = generate_morning_plan(
        fake_project, project_start_date="2026-05-18", current_date=date(2026, 5, 19), manual_override=False
    )

    assert result.success is True
    assert "Week1_Day2" in result.output_reports[0]


def test_morning_plan_content_is_specific_and_has_sources(fake_project):
    admin = fake_project / "00_Project_Admin"
    (admin / "project_schedule_week1.json").write_text(
        json.dumps({"days": {"2": ["Begin target cell list", "Start first walkthrough/audit if approved"]}}),
        encoding="utf-8",
    )
    (admin / "task_progress_week1.json").write_text(
        json.dumps(
            {
                "tasks": [
                    {"task_id": "W1D2T1", "day": "2", "task_text": "Begin target cell list", "status": "Not started"},
                    {
                        "task_id": "W1D2T2",
                        "day": "2",
                        "task_text": "Start first walkthrough/audit if approved",
                        "status": "Not started",
                    },
                ]
            }
        ),
        encoding="utf-8",
    )

    markdown, warnings, metrics = build_morning_plan_markdown(
        fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
    )

    _assert_todo_contract(markdown)
    assert "## Today's Mission" in markdown
    assert "## Do First" in markdown
    assert "## Main TODO" in markdown
    assert "## Ask Today" in markdown
    assert "## If Blocked" in markdown
    assert "## Done When" in markdown
    assert "If ahead, start: " not in markdown
    assert "choose the next highest-value task" not in markdown
    assert metrics["resolved_day"] == 2


def test_morning_plan_task_selection_uses_today_and_carryover(fake_project):
    admin = fake_project / "00_Project_Admin"
    (admin / "project_schedule_week1.json").write_text(
        json.dumps(
            {
                "days": {
                    "1": ["Resolve mentor blocker"],
                    "2": ["Begin target cell list", "Already done"],
                    "3": ["Stretch task"],
                }
            }
        ),
        encoding="utf-8",
    )
    (admin / "task_progress_week1.json").write_text(
        json.dumps(
            {
                "tasks": [
                    {"task_id": "W1D1T1", "day": "1", "task_text": "Resolve mentor blocker", "status": "Blocked"},
                    {"task_id": "W1D2T1", "day": "2", "task_text": "Begin target cell list", "status": "Not started"},
                    {"task_id": "W1D2T2", "day": "2", "task_text": "Already done", "status": "Complete"},
                    {"task_id": "W1D3T1", "day": "3", "task_text": "Stretch task", "status": "Not started"},
                ]
            }
        ),
        encoding="utf-8",
    )

    markdown, warnings, metrics = build_morning_plan_markdown(
        fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
    )
    main = _section(markdown, "Main TODO", "Ask Today")

    assert "Already done" not in markdown
    assert "Resolve mentor blocker" in markdown
    assert "Resolve or document the blocker" in main


def test_morning_planner_cli_still_accepts_week_and_day(fake_project):
    admin = fake_project / "00_Project_Admin"
    (admin / "project_schedule_week1.json").write_text(json.dumps({"days": {"2": ["Review data"]}}), encoding="utf-8")
    (admin / "task_progress_week1.json").write_text(
        json.dumps({"tasks": [{"task_id": "W1D2T1", "day": "2", "task_text": "Review data", "status": "Not started"}]}),
        encoding="utf-8",
    )

    completed = subprocess.run(
        [
            sys.executable,
            str(Path.cwd() / "tools" / "morning_project_planner.py"),
            "--project-root",
            str(fake_project),
            "--week",
            "1",
            "--day",
            "2",
        ],
        text=True,
        capture_output=True,
        check=False,
    )

    assert completed.returncode == 0
    assert "Generated Week 1 Day 2 morning plan." in completed.stdout


def test_morning_plan_primary_focus_is_mission_style(fake_project):
    _morning_plan_quality_fixture(fake_project)

    markdown, _warnings, metrics = build_morning_plan_markdown(
        fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
    )
    primary = _section(markdown, "Today's Mission", "Do First")

    assert metrics["resolved_day"] == 2
    assert "EOAT discovery" in primary
    assert "making visible progress on:" not in primary
    assert "Begin target cell list, Start first walkthrough/audit if approved" not in primary


def test_morning_plan_recommended_next_actions_are_multiple_and_practical(fake_project):
    _morning_plan_quality_fixture(fake_project)

    markdown, _warnings, metrics = build_morning_plan_markdown(
        fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
    )
    do_first = _section(markdown, "Do First", "Main TODO")
    main = _section(markdown, "Main TODO", "Ask Today")
    actions = [line for line in (do_first + "\n" + main).splitlines() if line.startswith(("1. ", "- [ ] "))]

    assert 4 <= len(actions) <= 6
    assert actions[0].startswith("1. ")
    assert any("target-cell list" in line for line in actions)
    assert any("photo naming" in line for line in actions)
    assert "continue project work" not in (do_first + main).lower()


def test_morning_plan_if_blocked_section_is_useful_and_conditional(fake_project):
    _morning_plan_quality_fixture(fake_project)
    admin = fake_project / "00_Project_Admin"
    progress = json.loads((admin / "task_progress_week1.json").read_text(encoding="utf-8"))
    progress["tasks"].append(
        {
            "task_id": "W1D1BLOCK",
            "day": "1",
            "task_text": "Start first walkthrough/audit if approved",
            "status": "Blocked",
        }
    )
    (admin / "task_progress_week1.json").write_text(json.dumps(progress), encoding="utf-8")

    markdown, _warnings, metrics = build_morning_plan_markdown(
        fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
    )
    section = _section(markdown, "If Blocked", "Done When")
    bullets = [line for line in section.splitlines() if line.startswith("- ")]

    assert metrics["if_blocked_items"] >= 2
    assert 2 <= len(bullets) <= 3
    assert any("blocked" in line.lower() for line in bullets)
    assert "floor access is approved" not in section.lower()


def test_morning_plan_definition_of_done_is_measurable(fake_project):
    _morning_plan_quality_fixture(fake_project)

    markdown, _warnings, metrics = build_morning_plan_markdown(
        fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
    )
    section = _section(markdown, "Done When")
    bullets = [line for line in section.splitlines() if line.startswith("- ")]

    assert metrics["definition_of_done_items"] >= 2
    assert 2 <= len(bullets) <= 3
    assert any("target-cell list" in line.lower() for line in bullets)


def test_morning_plan_has_no_blank_bullets(fake_project):
    _morning_plan_quality_fixture(fake_project)

    markdown, _warnings, _metrics = build_morning_plan_markdown(
        fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
    )

    assert "- If ahead, start:" not in markdown
    assert "\n- \n" not in markdown
    assert not any(line.rstrip() == "-" for line in markdown.splitlines())


def test_morning_plan_omits_optional_stretch_from_main_checklist(fake_project):
    _morning_plan_quality_fixture(fake_project)

    markdown, _warnings, _metrics = build_morning_plan_markdown(
        fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
    )
    _assert_todo_contract(markdown)
    assert "## Optional Stretch" not in markdown
    assert "Stretch report" not in markdown

    admin = fake_project / "00_Project_Admin"
    progress = json.loads((admin / "task_progress_week1.json").read_text(encoding="utf-8"))
    progress["tasks"] = [
        task for task in progress["tasks"] if str(task.get("day")) != "3" and str(task.get("day")) != "4"
    ]
    (admin / "task_progress_week1.json").write_text(json.dumps(progress), encoding="utf-8")

    markdown, _warnings, _metrics = build_morning_plan_markdown(
        fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
    )
    assert "No stretch tasks suggested yet." not in markdown


def test_morning_plan_latest_tool_metadata_uses_activity_log_truthfully(fake_project):
    _morning_plan_quality_fixture(fake_project)
    activity_path = fake_project / "00_Project_Admin" / "Activity_Logs" / "activity_log.jsonl"
    activity_path.parent.mkdir(parents=True, exist_ok=True)
    activity_path.write_text(
        json.dumps(
            {
                "timestamp": "2026-05-19T08:12:33+00:00",
                "tool_name": "EOAT Audit Progress Dashboard Tool",
                "success": True,
                "summary": "Synthetic activity entry.",
            }
        )
        + "\n",
        encoding="utf-8",
    )

    markdown = build_morning_plan_details_markdown(fake_project, 1, 2, date(2026, 5, 19))
    activity = _section(markdown, "Recent Activity", "Workbook State")
    assert "EOAT Audit Progress Dashboard Tool" in activity
    assert "Synthetic activity entry." in activity

    activity_path.write_text(json.dumps({"tool_name": "Workbook Validation"}) + "\n", encoding="utf-8")
    markdown = build_morning_plan_details_markdown(fake_project, 1, 2, date(2026, 5, 19))
    activity = _section(markdown, "Recent Activity", "Workbook State")
    assert "Workbook Validation" in activity
    assert "Workbook Validation - success" not in activity
    assert "Workbook Validation - failed" not in activity


def test_morning_plan_is_concise_while_details_report_has_source_quality(fake_project):
    _morning_plan_quality_fixture(fake_project)
    activity_path = fake_project / "00_Project_Admin" / "Activity_Logs" / "activity_log.jsonl"
    activity_path.write_text(
        json.dumps(
            {
                "timestamp": "2026-05-19T12:22:16+00:00",
                "tool_name": "EOAT Audit Progress Dashboard Tool",
                "success": True,
                "files_created": ["01_EOAT_Audit/Audit_Progress_Reports/Audit_Progress_2026-05-19_0822.md"],
                "summary": "Generated audit progress report.",
            }
        )
        + "\n",
        encoding="utf-8",
    )
    (fake_project / "01_EOAT_Audit" / "recent_state_marker.md").write_text("recent app/project state", encoding="utf-8")

    markdown, _warnings, metrics = build_morning_plan_markdown(
        fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
    )

    details = build_morning_plan_details_markdown(fake_project, 1, 2, date(2026, 5, 19))

    assert metrics["resolved_day"] == 2
    _assert_todo_contract(markdown)
    assert "## Main TODO" in markdown
    assert "## Source Availability" not in markdown
    assert "## Confidence / Data Quality" not in markdown
    assert "## Recently Modified Files" not in markdown
    assert "Generated audit progress report" not in markdown
    assert "## Source Availability" in details
    assert "activity_log.jsonl: found" in details
    assert "## Confidence / Data Quality" in details
    assert "recent_state_marker.md" in details


def test_morning_plan_detail_level_debug_returns_diagnostics(fake_project):
    _morning_plan_quality_fixture(fake_project)

    markdown, _warnings, metrics = build_morning_plan_markdown(
        fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
        detail_level="debug",
    )

    assert metrics["resolved_day"] == 2
    assert "## Source Availability" in markdown
    assert "## Confidence / Data Quality" in markdown


def test_morning_plan_reports_blocked_tasks_and_truthful_action_item_source(fake_project):
    admin = fake_project / "00_Project_Admin"
    (admin / "project_schedule_week1.json").write_text(json.dumps({"days": {"2": ["Review data"]}}), encoding="utf-8")
    (admin / "task_progress_week1.json").write_text(
        json.dumps({"tasks": [{"task_id": "W1D2T1", "day": "2", "task_text": "Review data", "status": "Blocked"}]}),
        encoding="utf-8",
    )

    markdown, _warnings, metrics = build_morning_plan_markdown(
        fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
    )

    assert metrics["blocked_tasks"] == 1
    assert "Clear the blocker on Review data" in markdown
    assert "No open action items found" not in markdown
    assert "Resolve or document the blocker on Review data" in markdown


def test_morning_plan_missing_action_source_does_not_claim_none(minimal_fake_project):
    admin = minimal_fake_project / "00_Project_Admin"
    admin.mkdir(parents=True, exist_ok=True)
    (admin / "project_schedule_week1.json").write_text(json.dumps({"days": {"2": ["Review data"]}}), encoding="utf-8")

    markdown, _warnings, _metrics = build_morning_plan_markdown(
        minimal_fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
    )

    assert "No open action items found" not in markdown
    assert "## Source Availability" not in markdown
    _assert_todo_contract(markdown)


def test_morning_plan_ranked_actions_are_not_schedule_echoes(fake_project):
    _morning_plan_quality_fixture(fake_project)

    markdown, _warnings, _metrics = build_morning_plan_markdown(
        fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
    )
    do_first = _section(markdown, "Do First", "Main TODO")
    main = _section(markdown, "Main TODO", "Ask Today")
    section = do_first + "\n" + main

    assert "1. Open the EOAT Audit page" in section
    assert "Ask mentor/supervisor" in section
    assert "1. Begin target cell list" not in section
    assert "Start first walkthrough/audit if approved." not in section


def test_morning_plan_changes_when_activity_and_workbook_state_change(fake_project):
    _morning_plan_quality_fixture(fake_project)

    before, _warnings, _metrics = build_morning_plan_markdown(
        fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
    )

    paths = resolve_project_paths(fake_project)
    wb = load_workbook(paths.master_workbook)
    ws = wb["Action Items"]
    ws.append(
        [
            "ACT-2",
            "2026-05-19",
            "Confirm Plant 4 first audit cell",
            "Plant 4",
            "KG",
            "High",
            "",
            "Open",
            "",
            "Manual override",
        ]
    )
    wb.save(paths.master_workbook)
    wb.close()
    activity_path = fake_project / "00_Project_Admin" / "Activity_Logs" / "activity_log.jsonl"
    activity_path.write_text(
        json.dumps(
            {
                "timestamp": "2026-05-19T14:00:00+00:00",
                "tool_name": "Workbook Validation",
                "success": True,
                "summary": "Validated workbook after action update.",
            }
        )
        + "\n",
        encoding="utf-8",
    )

    after, _warnings, _metrics = build_morning_plan_markdown(
        fake_project,
        project_start_date="2026-05-18",
        current_date=date(2026, 5, 19),
        manual_override=False,
    )

    assert before != after
    assert "Confirm Plant 4 first audit cell" in after
    assert "Validated workbook after action update" not in after
    assert "Confirm Plant 4 first audit cell" in after


def test_generic_output_detector_flags_static_reports():
    static_report = """# Week 1 Day 2 Morning Plan

## Scheduled Tasks
- Begin target cell list
- Start first walkthrough/audit if approved
- Decide photo naming system

## Recommended Next Actions
- Begin target cell list
- Start first walkthrough/audit if approved
- Decide photo naming system

## Carryover / Blockers
- No carryover or blocked tasks found.

## Open Action Items
- No open action items found.
"""

    issues = detect_generic_morning_report(
        static_report,
        ["Begin target cell list", "Start first walkthrough/audit if approved", "Decide photo naming system"],
    )

    assert "missing blocked/no-floor-access fallback" in issues
    assert "recommended actions are not ranked" in issues
