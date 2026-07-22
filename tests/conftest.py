from __future__ import annotations

import gc
import os
from pathlib import Path

import pytest
from openpyxl import Workbook

from core.audit_entries import CURRENT_WORKBOOK_SCHEMA_VERSION, WORKBOOK_METADATA_SHEET
from core.config import UserConfig
from core.constants import EXPECTED_NUMBERED_FOLDERS
from core.workbook_schema import get_expected_headers, get_expected_sheets
from tests.fixtures.fake_config import create_fake_config
from tests.fixtures.fake_project import create_fake_eoat_project, create_minimal_fake_project
from tests.qt_lifecycle_diagnostics import (
    assert_post_test_invariants as assert_qt_post_test_invariants,
)
from tests.qt_lifecycle_diagnostics import (
    drain_deferred_deletes,
    settle_pending_widget_updates,
)
from tests.qt_lifecycle_diagnostics import (
    snapshot as qt_lifecycle_snapshot,
)

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
os.environ.setdefault("EOAT_DISABLE_GLOBAL_TYPE_SEARCH", "1")
if os.environ.get("QT_QPA_PLATFORM", "").strip().casefold() == "offscreen":
    system_font_directory = Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts"
    if system_font_directory.is_dir():
        # The Qt offscreen plugin on this Windows image does not enumerate
        # installed fonts unless its font directory is explicit.  Tests and
        # native evidence use installed system fonts only; no font is bundled.
        os.environ.setdefault("QT_QPA_FONTDIR", str(system_font_directory))


def _load_local_development_database_environment() -> None:
    """Make the approved local database available to in-process API integration tests."""
    environment_file = Path(os.environ.get("LOCALAPPDATA", "")) / "EOAT Atlas Development" / "database.env"
    if not environment_file.is_file():
        return
    for raw_line in environment_file.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        if key.startswith("EOAT_DB_"):
            os.environ.setdefault(key, value)


_load_local_development_database_environment()


@pytest.fixture(autouse=True)
def isolated_eoat_atlas_runtime(monkeypatch, tmp_path):
    monkeypatch.setenv("EOAT_ATLAS_LOCALAPPDATA", str(tmp_path / "localappdata"))


@pytest.fixture(scope="session")
def qapp():
    pytest.importorskip("PySide6")
    from PySide6.QtGui import QFont, QGuiApplication
    from PySide6.QtWidgets import QApplication

    app = QApplication.instance() or QApplication([])
    if QGuiApplication.platformName().casefold() == "offscreen":
        app.setFont(QFont("Segoe UI", 9))
    yield app
    app.processEvents()


@pytest.fixture(autouse=True)
def cleanup_qt_widgets(request):
    qt_lifecycle_snapshot(request.node.nodeid, "before_test")
    yield
    try:
        from PySide6.QtWidgets import QApplication
    except ImportError:
        return
    app = QApplication.instance()
    if app is None:
        return
    qt_lifecycle_snapshot(request.node.nodeid, "before_cleanup")
    settle_pending_widget_updates(app)
    from shiboken6 import isValid

    for widget in app.topLevelWidgets():
        if not isValid(widget):
            continue
        try:
            # Popup/view helper widgets can be reported as top-level even
            # while they remain QObject children of a test window.  Closing
            # one is correct; scheduling it for deletion in addition to its
            # owning window creates a competing destruction path.
            is_owned_popup = widget.parentWidget() is not None or widget.parent() is not None
            widget.close()
            if not is_owned_popup:
                widget.deleteLater()
        except RuntimeError:
            # A parent can destroy a sibling while QApplication is enumerating
            # top-level widgets.  Never invoke another Qt method on that stale
            # wrapper.
            continue
    drain_deferred_deletes(app)
    qt_lifecycle_snapshot(request.node.nodeid, "after_cleanup")
    assert_qt_post_test_invariants(app)


def pytest_sessionfinish(session, exitstatus):
    try:
        from PySide6.QtWidgets import QApplication
    except ImportError:
        return
    app = QApplication.instance()
    if app is not None:
        from shiboken6 import isValid

        for widget in app.topLevelWidgets():
            if not isValid(widget):
                continue
            try:
                is_owned_popup = widget.parentWidget() is not None or widget.parent() is not None
                widget.close()
                if not is_owned_popup:
                    widget.deleteLater()
            except RuntimeError:
                continue
        drain_deferred_deletes(app)
    try:
        from core.performance import flush_performance_log_queue

        flush_performance_log_queue(timeout=2.0)
    except Exception:
        return
    try:
        gc.freeze()
    except Exception:
        return


@pytest.fixture
def fake_project(tmp_path) -> Path:
    for folder in EXPECTED_NUMBERED_FOLDERS:
        (tmp_path / folder).mkdir(parents=True, exist_ok=True)
    (tmp_path / "00_Project_Admin" / "Daily_Status_Reports").mkdir(exist_ok=True)
    (tmp_path / "00_Project_Admin" / "Weekly_Status_Reports").mkdir(exist_ok=True)
    (tmp_path / "00_Project_Admin" / "Activity_Logs").mkdir(exist_ok=True)
    workbook_dir = tmp_path / "01_EOAT_Audit" / "EOAT_Audit_Database"
    workbook_dir.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in get_expected_sheets():
        ws = workbook.create_sheet(sheet_name)
        ws.append(get_expected_headers(sheet_name))
    metadata = workbook.create_sheet(WORKBOOK_METADATA_SHEET)
    metadata.sheet_state = "hidden"
    metadata.append(["key", "value"])
    metadata.append(["schema_version", CURRENT_WORKBOOK_SCHEMA_VERSION])
    metadata.append(["app_name", "EOAT Atlas"])
    workbook.save(workbook_dir / "EOAT_Master_Tracker.xlsx")
    workbook.close()
    return tmp_path


@pytest.fixture
def usability_fake_project(tmp_path) -> Path:
    return create_fake_eoat_project(tmp_path)


@pytest.fixture
def minimal_fake_project(tmp_path) -> Path:
    return create_minimal_fake_project(tmp_path)


@pytest.fixture
def fake_config(fake_project: Path) -> UserConfig:
    return create_fake_config(fake_project)


@pytest.fixture
def usability_fake_config(usability_fake_project: Path) -> UserConfig:
    return create_fake_config(usability_fake_project)
