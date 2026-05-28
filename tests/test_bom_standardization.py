from __future__ import annotations

from openpyxl import load_workbook

from core.bom_standardization import analyze_bom_standardization, generate_bom_standardization_report
from core.paths import resolve_project_paths


def test_bom_standardization_empty_inventory(fake_project):
    data, warnings, details = analyze_bom_standardization(fake_project)

    assert warnings == []
    assert details
    assert data["rows"] == []
    assert "Start by auditing" in data["opportunities"][0]


def test_bom_standardization_common_parts_and_missing_data(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    wb = load_workbook(workbook_path)
    ws = wb["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    for row in [
        {
            "Audit ID": "AUD-1",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Number of Parts Picked": 4,
            "# of Cups": 4,
            "Cup Type/Material": "Silicone",
            "Cup Diameter/Size": "20mm",
        },
        {
            "Audit ID": "AUD-2",
            "Audit Date": "2026-05-18",
            "Auditor": "KG",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 13",
            "Robot Type": "Wittmann R9",
            "EOAT Type": "Vacuum",
            "Number of Parts Picked": 6,
            "# of Cups": 4,
            "Cup Type/Material": "Silicone",
            "Cup Diameter/Size": "20mm",
        },
    ]:
        ws.append([row.get(header, "") for header in headers])
    wb.save(workbook_path)
    wb.close()

    data, warnings, _details = analyze_bom_standardization(fake_project)
    assert warnings == []
    assert data["counts"]["vacuum cup counts"]["4"] == 2
    assert data["counts"]["vacuum cup materials"]["Silicone"] == 2
    assert data["counts"]["vacuum cup sizes"]["20mm"] == 2
    assert data["missing_rows"]

    result = generate_bom_standardization_report(fake_project)
    assert result.success is True
    assert result.output_reports
    assert any(path.endswith(".csv") for path in result.files_created)
