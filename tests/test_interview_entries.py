from __future__ import annotations

from openpyxl import load_workbook

from core.interview_entries import generate_interview_id, save_interview_entry
from core.paths import resolve_project_paths


def test_generate_interview_id_and_add_row(fake_project):
    interview_id = generate_interview_id(fake_project, "2026-05-18")
    result = save_interview_entry(
        fake_project,
        {
            "Interview ID": interview_id,
            "Date": "2026-05-18",
            "Role/Department": "Technician",
            "Plant/Area": "Plant 4",
            "Press/Machine #": "Press 12",
            "Notes": "Cups wear out first.",
        },
    )

    assert result.success is True
    wb = load_workbook(resolve_project_paths(fake_project).master_workbook, read_only=True)
    ws = wb["Interview Notes"]
    assert ws.max_row == 2
    assert ws["A2"].value == interview_id
    wb.close()


def test_interview_validation_requires_notes(fake_project):
    result = save_interview_entry(fake_project, {"Date": "2026-05-18", "Role/Department": "Operator"})

    assert result.success is False
    assert any("Notes" in error for error in result.errors)

