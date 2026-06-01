from __future__ import annotations

import json

from openpyxl import load_workbook

from core.paths import resolve_project_paths
from core.validation import run_foundation_validation, validate_project_foundation
from core.validation_findings import ValidationFinding, ValidationSeverity, findings_from_result


def test_validation_finding_model_stable_id_and_severity():
    finding = ValidationFinding(
        finding_id="",
        severity="error",
        category="audit_data",
        sheet_name="EOAT Inventory",
        row_number=2,
        column_name="Audit ID",
        audit_id="AUD-1",
        message="Duplicate Audit ID value: AUD-1",
    )
    same = ValidationFinding.from_dict(finding.to_dict())

    assert finding.severity == ValidationSeverity.ERROR.value
    assert finding.finding_id == same.finding_id


def test_validation_emits_structured_findings_json_and_preserves_markdown(fake_project):
    workbook_path = resolve_project_paths(fake_project).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook["EOAT Inventory"]
    headers = [cell.value for cell in ws[1]]
    row = [""] * len(headers)
    for field, value in {
        "Audit ID": "AUD-DUPLICATE-001",
        "Audit Date": "2026-05-18",
        "Auditor": "KG",
        "Plant/Area": "Plant 4",
        "Press/Machine #": "Press 998",
        "Tool #": "DEMO-PN-998",
        "Robot Type": "Wittmann R9",
        "EOAT Type": "Vacuum",
        "Status": "In Progress",
    }.items():
        row[headers.index(field)] = value
    ws.append(row)
    duplicate = list(row)
    duplicate[headers.index("Press/Machine #")] = "Press 999"
    ws.append(duplicate)
    workbook.save(workbook_path)
    workbook.close()

    result = run_foundation_validation(fake_project, write_report=True, log_activity=False)
    findings = findings_from_result(result)

    assert any(
        finding.severity == ValidationSeverity.ERROR.value and finding.column_name == "Audit ID" for finding in findings
    )
    assert "finding_id" not in result.to_markdown()
    json_reports = sorted(
        (fake_project / "00_Project_Admin" / "Validation_Reports").glob("Foundation_Validation_*.json")
    )
    md_reports = sorted((fake_project / "00_Project_Admin" / "Validation_Reports").glob("Foundation_Validation_*.md"))
    assert json_reports
    assert md_reports
    payload = json.loads(json_reports[-1].read_text(encoding="utf-8"))
    assert payload["summary_counts"]["total"] == len(findings)
    assert any(item["column_name"] == "Audit ID" for item in payload["findings"])


def test_validation_structures_compatibility_missing_source_without_legacy_warning_noise(tmp_path):
    from tests.test_validation import _append_inventory_row, _complete_inventory_values, _create_schema_workbook

    workbook_path = _create_schema_workbook(tmp_path)
    _append_inventory_row(workbook_path, _complete_inventory_values("AUD-COMPAT-BLANK-SOURCE", "Compatible"))

    result = validate_project_foundation(tmp_path)
    findings = findings_from_result(result)

    assert any(finding.column_name == "Source Audit ID" and finding.severity == "ERROR" for finding in findings)
    assert "Source Audit ID" not in "\n".join(result.warnings)
