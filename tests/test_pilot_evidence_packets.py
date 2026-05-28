from __future__ import annotations

from openpyxl import load_workbook

from core.paths import resolve_project_paths
from core.pilot_evidence_packets import build_pilot_evidence_packet, generate_pilot_evidence_packet
from core.workbook_schema import get_expected_headers


def _append_row(project_root, sheet_name: str, values: dict[str, str]) -> None:
    workbook_path = resolve_project_paths(project_root).master_workbook
    workbook = load_workbook(workbook_path)
    ws = workbook[sheet_name]
    headers = [cell.value for cell in ws[1]]
    row = {header: "" for header in get_expected_headers(sheet_name)}
    row.update(values)
    ws.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    workbook.close()


def test_pilot_evidence_packet_includes_expected_sections(fake_project):
    _append_row(
        fake_project,
        "EOAT Inventory",
        {
            "Audit ID": "AUD-PILOT-PACKET-001",
            "Press/Machine #": "Press 61",
            "EOAT Type": "Vacuum",
            "Known Issues": "Vacuum drop during startup.",
            "Pilot Candidate?": "Yes",
            "Status": "Complete",
            "Tubing Condition": "Worn",
            "BOM Available?": "No",
        },
    )
    _append_row(
        fake_project,
        "Pilot Candidates",
        {
            "Candidate ID": "AUD-PILOT-PACKET-001",
            "Press/Machine #": "Press 61",
            "EOAT Type": "Vacuum",
            "Main Problem": "Vacuum drop during startup.",
            "Expected KPI Improvement": "Reliability review only; quantify later.",
            "Ease of Implementation": "Medium",
        },
    )

    packet, error = build_pilot_evidence_packet(fake_project, candidate_id="AUD-PILOT-PACKET-001")

    assert error is None
    assert packet is not None
    markdown = packet.to_markdown()
    assert "Failure Modes For Review" in markdown
    assert "Standards Gaps" in markdown
    assert "Photo / Evidence Coverage" in markdown
    assert "does not claim pilot success" in markdown


def test_generate_pilot_evidence_packet_writes_timestamped_markdown(fake_project):
    _append_row(
        fake_project,
        "EOAT Inventory",
        {
            "Audit ID": "AUD-PILOT-PACKET-002",
            "Press/Machine #": "Press 62",
            "EOAT Type": "Mechanical / Gripper",
            "Known Issues": "Loose gripper jaw.",
            "Pilot Candidate?": "Maybe",
        },
    )

    result = generate_pilot_evidence_packet(fake_project, audit_id="AUD-PILOT-PACKET-002", log_activity=False)

    assert result.success is True
    assert result.output_reports
    assert "Pilot_Evidence_Packet" in result.output_reports[0]
